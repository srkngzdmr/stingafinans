# -*- coding: utf-8 -*-
# ╔══════════════════════════════════════════════════════════════╗
# ║       STINGA PRO FINANCE v17.0 - ULTRA EDITION              ║
# ║  Geliştiren: AI ile birlikte - Gemini 2.5 Flash Destekli    ║
# ╚══════════════════════════════════════════════════════════════╝

import streamlit as st
import pandas as pd
import google.generativeai as genai
from datetime import datetime, timedelta

# ─── İSTANBUL SAAT DİLİMİ ──────────────────────────────────────
def now_ist():
    """UTC+3 Istanbul saati döndürür (pytz gerektirmez, Streamlit Cloud uyumlu)."""
    import time as _time
    # os.environ TZ'ye güvenmek yerine direkt UTC offset kullan
    from datetime import timezone as _tz
    return datetime.now(_tz.utc).replace(tzinfo=None) + timedelta(hours=3)
import os
import json
import re
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from PIL import Image
import hashlib
# ─── RAPOR KÜTÜPHANELERİ ──────────────────────────────────────
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
try:
    from reportlab.platypus import Image as RLImage
except ImportError:
    RLImage = None
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment as XLAlign, Border, Side
from openpyxl.utils import get_column_letter
try:
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    XLImage = None
import time
import random
import base64
from io import BytesIO

# ─── SAYFA YAPISI ─────────────────────────────────────────────
st.set_page_config(
    page_title="Stinga Pro Finance v17.0 ⚡",
    layout="wide",
    page_icon="⚡",
    initial_sidebar_state="expanded"
)

# ─── ULTRA DARK THEME CSS ─────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;1,400&family=JetBrains+Mono:wght@400;500;700&display=swap');

/* ════════════════════════════════════════════════════════
   STINGA PRO — LIGHT THEME (21st.dev inspired)
   Sidebar  : #08090c (siyah)
   Content  : #f8f9fb (açık)
   Green    : #11855B  Navy: #2F3C6E
════════════════════════════════════════════════════════ */

:root {
    --sg:        #11855B;
    --sg-hi:     #17a870;
    --sg-lo:     #0c6344;
    --sn:        #2F3C6E;
    --sn-hi:     #3d4e8a;

    /* SIDEBAR — açık tema */
    --sb:        #ffffff;
    --sb-s:      #f4f7f5;
    --sb-bdr:    rgba(17,133,91,0.16);

    /* CONTENT — açık */
    --bg:        #f4f6f8;
    --bg-2:      #eef0f3;
    --bg-card:   #ffffff;
    --bg-hover:  #f0f4f1;

    /* TEXT */
    --t1:  #0f1923;
    --t2:  #3d5260;
    --t3:  #7a96a4;

    /* ACCENTS */
    --amber: #d97706;
    --red:   #dc2626;
    --cyan:  #0891b2;

    --shadow: 0 1px 4px rgba(15,25,35,0.08), 0 4px 16px rgba(15,25,35,0.06);
    --shadow-lg: 0 4px 12px rgba(15,25,35,0.1), 0 16px 48px rgba(15,25,35,0.08);
    --bdr:   rgba(17,133,91,0.15);
    --bdr-m: rgba(17,133,91,0.3);
    --glow:  0 4px 20px rgba(17,133,91,0.15);
}

* { box-sizing: border-box; }

html, body, [class*="css"] {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    background: var(--bg) !important;
    color: var(--t1) !important;
}

.stApp {
    background:
        radial-gradient(ellipse 70% 50% at 0% 0%,   rgba(17,133,91,0.05) 0%, transparent 55%),
        radial-gradient(ellipse 60% 45% at 100% 100%, rgba(47,60,110,0.06) 0%, transparent 55%),
        var(--bg) !important;
}

/* ── SIDEBAR ─────────────────────────── */
[data-testid="stSidebar"] {
    background: var(--sb) !important;
    border-right: 1px solid rgba(17,133,91,0.14) !important;
    box-shadow: 4px 0 16px rgba(0,0,0,0.06) !important;
}
[data-testid="stSidebar"] > div:first-child { padding: 0 !important; }
[data-testid="stSidebar"] * { color: inherit !important; }

[data-testid="stSidebar"] .stRadio > div { gap: 0 !important; }
[data-testid="stSidebar"] .stRadio label {
    display: flex !important; align-items: center !important;
    padding: 11px 16px !important; margin: 1px 8px !important;
    border-radius: 9px !important; border: 1px solid transparent !important;
    background: transparent !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 0.87rem !important; font-weight: 600 !important;
    color: #5a7a6a !important; cursor: pointer !important;
    transition: all 0.16s ease !important; letter-spacing: 0.01em !important;
    position: relative !important;
}
[data-testid="stSidebar"] .stRadio label::before {
    content: ''; position: absolute; left: 0; top: 18%; bottom: 18%;
    width: 3px; border-radius: 0 3px 3px 0;
    background: var(--sg); opacity: 0; transition: opacity 0.15s;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(17,133,91,0.07) !important;
    color: #11855B !important;
    border-color: rgba(17,133,91,0.15) !important;
    transform: translateX(4px) !important;
}
[data-testid="stSidebar"] .stRadio label:hover::before { opacity: 0.6 !important; }
[data-testid="stSidebar"] .stRadio label[data-checked="true"] {
    background: linear-gradient(90deg,rgba(17,133,91,0.1),rgba(17,133,91,0.03)) !important;
    color: var(--sg) !important;
    border-color: rgba(17,133,91,0.22) !important; font-weight: 700 !important;
}
[data-testid="stSidebar"] .stRadio label[data-checked="true"]::before { opacity: 1 !important; }
[data-testid="stSidebar"] .stRadio input[type="radio"] { display: none !important; }
[data-testid="stSidebar"] .stRadio > div > label > div:first-child { display: none !important; }

/* ── LOGO ANIMATION ──────────────────── */
.slogo-frame {
    display: inline-block; border-radius: 50%; position: relative;
    animation: sBreath 4s ease-in-out infinite;
}
@keyframes sBreath {
    0%,100% { transform: scale(0.97); filter: drop-shadow(0 2px 8px rgba(17,133,91,0.15)); }
    50%     { transform: scale(1.06); filter: drop-shadow(0 6px 22px rgba(17,133,91,0.5)); }
}
.slogo-frame::after {
    content: ''; position: absolute; inset: -5px; border-radius: 50%;
    border: 1.5px solid rgba(17,133,91,0.4);
    animation: sRing 4s ease-in-out infinite; pointer-events: none;
}
@keyframes sRing {
    0%,100% { transform: scale(1); opacity: 0.6; }
    50%     { transform: scale(1.18); opacity: 0; }
}
.slogo-core {
    background: transparent; border-radius: 50%; padding: 4px;
    display: flex; align-items: center; justify-content: center;
    width: 192px; height: 192px;
}
.slogo-core img { width: 176px; height: 176px; object-fit: contain; border-radius: 50%; }

/* ── USER CARD ───────────────────────── */
.suser-card {
    margin: 8px 10px 6px; padding: 14px 15px;
    background: var(--sb-s);
    border: 1px solid rgba(17,133,91,0.2); border-radius: 14px;
    transition: all 0.22s ease; position: relative; overflow: hidden;
}
.suser-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, var(--sg-lo), var(--sg), var(--sg-lo)); opacity: 0.8;
}
.suser-card:hover { border-color: rgba(17,133,91,0.4); box-shadow: 0 4px 16px rgba(17,133,91,0.15); transform: translateY(-1px); }
.suser-top { display: flex; align-items: center; gap: 11px; }
.suser-ava {
    width: 44px; height: 44px; border-radius: 11px;
    background: linear-gradient(135deg,rgba(17,133,91,0.2),rgba(47,60,110,0.3));
    border: 1px solid rgba(17,133,91,0.3);
    display: flex; align-items: center; justify-content: center;
    font-size: 1.4rem; transition: transform 0.25s ease;
}
.suser-card:hover .suser-ava { transform: scale(1.08) rotate(-4deg); }
.suser-card:hover .suser-ava { transform: scale(1.08) rotate(-4deg); }
.suser-name { font-size: 0.93rem; font-weight: 700; color: #0f1923 !important; }
.suser-role { font-family: 'JetBrains Mono',monospace; font-size: 0.59rem; margin-top: 3px; letter-spacing: 0.1em; }
.sxp-row { display:flex; justify-content:space-between; font-family:'JetBrains Mono',monospace; font-size:0.58rem; color:#7a96a4 !important; margin:10px 0 4px; }
.sxp-track { height: 3px; background: rgba(0,0,0,0.08); border-radius:99px; position: relative; overflow: visible; }
.sxp-fill { height: 100%; border-radius:99px; background: linear-gradient(90deg,var(--sg-lo),var(--sg)); position: relative; }
.sxp-fill::after { content:''; position:absolute; right:-1px; top:50%; transform:translateY(-50%); width:7px; height:7px; background:var(--sg); border-radius:50%; box-shadow:0 0 6px rgba(17,133,91,0.5); }
.smeta { display:flex; justify-content:space-between; margin-top:9px; font-family:'JetBrains Mono',monospace; font-size:0.58rem; color:#7a96a4 !important; }
.ssep { height:1px; background:linear-gradient(90deg,transparent,rgba(17,133,91,0.2),transparent); margin:5px 12px; }

.snotif {
    margin:5px 10px; padding:9px 13px;
    background:rgba(220,38,38,0.06); border:1px solid rgba(220,38,38,0.18);
    border-radius:10px; display:flex; align-items:center; gap:8px;
    animation:snotifP 2.5s ease-in-out infinite;
}
.snotif:hover { background:rgba(220,38,38,0.1); border-color:rgba(220,38,38,0.3); }
@keyframes snotifP { 0%,100%{box-shadow:none;}50%{box-shadow:0 0 0 4px rgba(220,38,38,0.05);} }
.snotif-lbl { font-size:0.72rem; font-weight:600; color:#dc2626 !important; flex:1; }
.snotif-num { background:var(--red); color:#fff !important; font-size:0.58rem; font-weight:700; padding:2px 6px; border-radius:99px; font-family:'JetBrains Mono',monospace; }
.snav-hdr { font-family:'JetBrains Mono',monospace; font-size:0.51rem; color:#a0b8ae !important; letter-spacing:0.2em; padding:6px 18px 3px; text-transform:uppercase; }

.slimit { margin:5px 10px 10px; padding:13px 15px; background:var(--sb-s); border:1px solid rgba(17,133,91,0.12); border-radius:14px; transition:all 0.22s ease; }
.slimit:hover { border-color:rgba(17,133,91,0.26); }
.slimit-lbl { font-family:'JetBrains Mono',monospace; font-size:0.51rem; color:#a0b8ae !important; letter-spacing:0.16em; text-transform:uppercase; margin-bottom:6px; }
.slimit-row { display:flex; justify-content:space-between; align-items:baseline; margin-bottom:6px; }
.slimit-pct { font-size:1.3rem; font-weight:800; font-family:'Plus Jakarta Sans',sans-serif; color:#0f1923; }
.slimit-val { font-family:'JetBrains Mono',monospace; font-size:0.54rem; color:#7a96a4 !important; }
.slimit-bar { height:4px; background:rgba(0,0,0,0.07); border-radius:99px; overflow:hidden; }

/* ════════════════════════════════════════
   MAIN CONTENT — LIGHT
════════════════════════════════════════ */

.ultra-card {
    background: var(--bg-card); border: 1px solid rgba(0,0,0,0.07);
    border-radius: 16px; padding: 24px; margin: 10px 0;
    position: relative; overflow: hidden;
    transition: all 0.22s ease; box-shadow: var(--shadow);
}
.ultra-card::before {
    content: ''; position: absolute; top:0; left:0; right:0; height:2px;
    background: linear-gradient(90deg,var(--sg),var(--sn-hi),var(--sg));
    opacity:0; transition:opacity 0.22s;
}
.ultra-card:hover { border-color: var(--bdr-m); box-shadow: var(--glow); transform: translateY(-2px); }
.ultra-card:hover::before { opacity: 1; }

.metric-card {
    background: var(--bg-card); border: 1px solid rgba(0,0,0,0.07);
    border-radius: 16px; padding: 22px; text-align: center;
    position: relative; overflow: hidden;
    transition: all 0.22s ease; box-shadow: var(--shadow);
}
.metric-card:hover { transform: translateY(-4px); border-color: var(--bdr-m); box-shadow: var(--glow); }
.metric-value { font-size: 2rem; font-weight: 800; line-height:1.1; margin:6px 0 4px; color:var(--t1); }
.metric-label { font-size:0.67rem; font-weight:700; color:var(--t3); text-transform:uppercase; letter-spacing:0.13em; }

.risk-badge { display:inline-block; padding:4px 12px; border-radius:20px; font-family:'JetBrains Mono',monospace; font-size:0.72rem; font-weight:700; }
.risk-low  { background:rgba(17,133,91,0.1);  color:var(--sg);   border:1px solid rgba(17,133,91,0.25); }
.risk-mid  { background:rgba(217,119,6,0.1);  color:var(--amber);border:1px solid rgba(217,119,6,0.25); }
.risk-high { background:rgba(220,38,38,0.1);  color:var(--red);  border:1px solid rgba(220,38,38,0.25); }

.stButton > button {
    background: linear-gradient(135deg,var(--sg-lo) 0%,var(--sg) 100%) !important;
    color:#fff !important; border:none !important; border-radius:10px !important;
    font-family:'Plus Jakarta Sans',sans-serif !important; font-weight:700 !important;
    font-size:0.88rem !important; letter-spacing:0.02em !important;
    transition:all 0.2s ease !important; padding:0.58rem 1.3rem !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg,var(--sg),var(--sg-hi)) !important;
    transform:translateY(-2px) !important; box-shadow:0 8px 20px rgba(17,133,91,0.35) !important;
}

.page-header { display:flex; align-items:center; gap:16px; margin-bottom:28px; padding-bottom:16px; border-bottom:1px solid rgba(0,0,0,0.08); }
.page-title { font-size:2.4rem; font-weight:800; letter-spacing:-0.02em; line-height:1; color:var(--t1); }

[data-testid="stDataFrame"] { border:1px solid rgba(0,0,0,0.08) !important; border-radius:14px !important; overflow:hidden !important; }

.stTextInput > div > div > input,
.stSelectbox > div > div > div,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea {
    background: #fff !important; border:1px solid rgba(0,0,0,0.12) !important;
    border-radius:10px !important; color:var(--t1) !important;
    font-family:'Plus Jakarta Sans',sans-serif !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color:var(--sg) !important; box-shadow:0 0 0 3px rgba(17,133,91,0.12) !important;
}

.stTabs [data-baseweb="tab-list"] { background:var(--bg-2) !important; border-radius:12px !important; padding:4px !important; border:1px solid rgba(0,0,0,0.07) !important; }
.stTabs [data-baseweb="tab"] { background:transparent !important; border-radius:9px !important; color:var(--t3) !important; font-family:'Plus Jakarta Sans',sans-serif !important; font-weight:600 !important; transition:all 0.18s !important; }
.stTabs [aria-selected="true"] { background:#fff !important; color:var(--sg) !important; box-shadow:var(--shadow) !important; }

.streamlit-expanderHeader { background:#fff !important; border:1px solid rgba(0,0,0,0.08) !important; border-radius:10px !important; color:var(--t1) !important; }

.stSuccess { background:rgba(17,133,91,0.08) !important; border-left:3px solid var(--sg) !important; color:var(--t1) !important; }
.stError   { background:rgba(220,38,38,0.07) !important; border-left:3px solid var(--red) !important; }
.stWarning { background:rgba(217,119,6,0.07) !important; border-left:3px solid var(--amber) !important; }
.stInfo    { background:rgba(8,145,178,0.07) !important; border-left:3px solid var(--cyan) !important; }

.stProgress > div > div { background:linear-gradient(90deg,var(--sg-lo),var(--sg)) !important; }
hr { border-color:rgba(0,0,0,0.07) !important; }

::-webkit-scrollbar { width:5px; }
::-webkit-scrollbar-track { background:var(--bg); }
::-webkit-scrollbar-thumb { background:rgba(17,133,91,0.2); border-radius:3px; }
::-webkit-scrollbar-thumb:hover { background:rgba(17,133,91,0.4); }

header[data-testid="stHeader"] { background:var(--bg) !important; border-bottom:1px solid rgba(0,0,0,0.07) !important; box-shadow:0 1px 4px rgba(0,0,0,0.05) !important; }
[data-testid="stDecoration"] { display:none !important; }
.block-container { background:transparent !important; padding-top:3rem !important; }
section[data-testid="stMain"] > div { background:transparent !important; }

[data-testid="stForm"] { background:#fff !important; border:1px solid rgba(0,0,0,0.09) !important; border-radius:20px !important; padding:28px !important; box-shadow:var(--shadow-lg) !important; }
.login-container { max-width:440px; margin:0 auto; padding:48px 40px; background:#fff; border:1px solid rgba(0,0,0,0.08); border-radius:24px; box-shadow:var(--shadow-lg); }
.logo-circle { display:inline-flex; align-items:center; justify-content:center; background:#fff; border-radius:50%; padding:10px; box-shadow:0 4px 20px rgba(17,133,91,0.25); }
.sidebar-logo-wrap { display:inline-flex; align-items:center; justify-content:center; background:#fff; border-radius:50%; padding:6px; }

/* LOGIN PAGE — animated gradient mesh background */
.login-bg {
    position: fixed; inset: 0; z-index: -1; overflow: hidden;
    background: linear-gradient(135deg, #f0f7f4 0%, #e8eef7 50%, #f4f0f7 100%);
}
.login-bg::before {
    content: ''; position: absolute; inset: -50%;
    background:
        radial-gradient(ellipse 80% 60% at 20% 20%, rgba(17,133,91,0.12) 0%, transparent 60%),
        radial-gradient(ellipse 70% 50% at 80% 80%, rgba(47,60,110,0.1) 0%, transparent 60%),
        radial-gradient(ellipse 60% 70% at 60% 10%, rgba(17,133,91,0.06) 0%, transparent 50%);
    animation: meshMove 12s ease-in-out infinite alternate;
}
@keyframes meshMove {
    0%   { transform: translate(0%, 0%) rotate(0deg); }
    33%  { transform: translate(3%, 2%) rotate(2deg); }
    66%  { transform: translate(-2%, 3%) rotate(-1deg); }
    100% { transform: translate(1%, -2%) rotate(1deg); }
}
.login-bg::after {
    content: ''; position: absolute; inset: 0;
    background: url("data:image/svg+xml,%3Csvg width='60' height='60' viewBox='0 0 60 60' xmlns='http://www.w3.org/2000/svg'%3E%3Cg fill='none' fill-rule='evenodd'%3E%3Cg fill='%2311855B' fill-opacity='0.03'%3E%3Ccircle cx='30' cy='30' r='1'/%3E%3C/g%3E%3C/g%3E%3C/svg%3E");
}

/* LOGIN LOGO — büyük ve animasyonlu */
.login-logo-wrap {
    display: inline-block; position: relative;
    animation: loginBreath 3.5s ease-in-out infinite;
}
@keyframes loginBreath {
    0%,100% { transform: scale(0.97); filter: drop-shadow(0 6px 20px rgba(17,133,91,0.2)); }
    50%     { transform: scale(1.04); filter: drop-shadow(0 10px 36px rgba(17,133,91,0.45)); }
}
.login-logo-ring {
    position: absolute; inset: -10px; border-radius: 50%;
    border: 2px solid rgba(17,133,91,0.45);
    animation: loginRing 3.5s ease-in-out infinite;
    pointer-events: none;
}
.login-logo-ring2 {
    position: absolute; inset: -20px; border-radius: 50%;
    border: 1.5px solid rgba(47,60,110,0.25);
    animation: loginRing 3.5s ease-in-out infinite 0.5s;
    pointer-events: none;
}
@keyframes loginRing {
    0%,100% { transform: scale(1); opacity: 0.7; }
    50%     { transform: scale(1.22); opacity: 0; }
}

/* FLOATING PARTICLES on login */
.particle {
    position: absolute; border-radius: 50%;
    background: radial-gradient(circle, rgba(17,133,91,0.15), transparent);
    animation: floatP linear infinite;
    pointer-events: none;
}
@keyframes floatP {
    0%   { transform: translateY(100vh) scale(0); opacity:0; }
    10%  { opacity: 1; }
    90%  { opacity: 1; }
    100% { transform: translateY(-10vh) scale(1); opacity:0; }
}

@keyframes pulse-glow { 0%,100%{box-shadow:0 0 10px rgba(17,133,91,0.15);}50%{box-shadow:0 0 25px rgba(17,133,91,0.4);} }
@keyframes float { 0%,100%{transform:translateY(0);}50%{transform:translateY(-10px);} }
.floating{animation:float 4s ease-in-out infinite;}
.pulsing{animation:pulse-glow 2s ease-in-out infinite;}
.glow-text{text-shadow:0 0 30px rgba(17,133,91,0.5);}

.feed-item{display:flex;align-items:flex-start;gap:12px;padding:12px 0;border-bottom:1px solid rgba(0,0,0,0.06);}
.feed-dot{width:8px;height:8px;border-radius:50%;margin-top:5px;flex-shrink:0;}

.status-pill{display:inline-block;padding:3px 10px;border-radius:20px;font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:1px;}
.pill-approved{background:rgba(17,133,91,0.1);color:var(--sg);border:1px solid rgba(17,133,91,0.2);}
.pill-pending{background:rgba(217,119,6,0.1);color:var(--amber);border:1px solid rgba(217,119,6,0.2);}
.pill-rejected{background:rgba(220,38,38,0.1);color:var(--red);border:1px solid rgba(220,38,38,0.2);}

.nav-item{padding:10px 16px;border-radius:10px;margin:2px 0;cursor:pointer;transition:all 0.18s;font-weight:500;}
.nav-item:hover{background:var(--bg-hover);}
.nav-item.active{background:rgba(17,133,91,0.1);color:var(--sg);border-left:3px solid var(--sg);}

.ai-bubble{background:linear-gradient(135deg,rgba(17,133,91,0.06),rgba(47,60,110,0.06));border:1px solid rgba(17,133,91,0.14);border-radius:14px;padding:18px 20px;margin:12px 0;position:relative;}
.ai-bubble::before{content:'⚡ AI';position:absolute;top:-10px;left:16px;background:linear-gradient(135deg,var(--sg-lo),var(--sn));color:#fff;font-size:0.62rem;font-weight:700;padding:2px 9px;border-radius:99px;letter-spacing:1.5px;}

.voice-note{background:rgba(17,133,91,0.07);border:1px solid rgba(17,133,91,0.18);border-radius:12px;padding:12px 16px;font-family:'JetBrains Mono',monospace;font-size:0.85rem;color:var(--sg);}
.anomaly-alert{background:linear-gradient(135deg,rgba(220,38,38,0.06),rgba(217,119,6,0.06));border:1px solid rgba(220,38,38,0.15);border-radius:12px;padding:16px;margin:8px 0;}
.budget-track{background:rgba(0,0,0,0.06);border-radius:8px;height:8px;overflow:hidden;margin:6px 0;}
.budget-fill{height:100%;border-radius:8px;transition:width 0.8s ease;}

.lb-item{display:flex;align-items:center;gap:12px;padding:13px 16px;border-radius:12px;margin:4px 0;background:#fff;border:1px solid rgba(0,0,0,0.07);transition:all 0.18s ease;box-shadow:var(--shadow);}
.lb-item:hover{border-color:var(--bdr-m);transform:translateX(3px);}
.lb-rank{font-family:'Plus Jakarta Sans';font-weight:900;font-size:1.3rem;color:var(--t3);width:30px;}
.lb-rank.gold{color:#d97706;text-shadow:0 0 12px rgba(217,119,6,0.4);}
.lb-rank.silver{color:#64748b;}
.lb-rank.bronze{color:#92400e;}

.heat-cell{display:inline-block;width:14px;height:14px;border-radius:3px;margin:1px;}
.info-tooltip{background:var(--bg-hover);border:1px solid rgba(0,0,0,0.08);border-radius:8px;padding:8px 12px;font-size:0.8rem;color:var(--t2);}
.empty-state{text-align:center;padding:60px 20px;color:var(--t3);}
.empty-icon{font-size:3rem;margin-bottom:12px;opacity:0.35;}
.upload-zone{border:2px dashed rgba(0,0,0,0.1);border-radius:16px;padding:40px;text-align:center;transition:all 0.3s;cursor:pointer;}
.upload-zone:hover{border-color:var(--sg);background:rgba(17,133,91,0.03);}
.spark-container{display:flex;align-items:flex-end;gap:2px;height:30px;}
.spark-bar{background:linear-gradient(180deg,var(--sg),var(--sg-lo));border-radius:2px;min-width:4px;opacity:0.6;}

@keyframes sdotBlink{0%,100%{opacity:1;}50%{opacity:0.2;}}
.sactive-dot{width:5px;height:5px;background:var(--sg);border-radius:50%;box-shadow:0 0 6px var(--sg);animation:sdotBlink 2s ease-in-out infinite;display:inline-block;margin-left:auto;}

/* ── ONAY ANİMASYONU (tam ekran overlay) ────────────────── */
#sg-approve-overlay {
    position:fixed; inset:0; z-index:99999;
    background:#000;
    display:flex; align-items:center; justify-content:center;
    flex-direction:column; gap:20px;
    animation: sgOverlayOut 0.4s ease 1.8s forwards;
    pointer-events:all;
}
@keyframes sgOverlayOut {
    to { opacity:0; pointer-events:none; }
}
.sg-approve-icon {
    font-size: 7rem;
    animation: sgIconPop 0.5s cubic-bezier(0.34,1.56,0.64,1) 0.2s both;
    filter: drop-shadow(0 0 40px rgba(17,233,150,0.9));
}
@keyframes sgIconPop {
    from { transform:scale(0) rotate(-30deg); opacity:0; }
    to   { transform:scale(1) rotate(0deg);   opacity:1; }
}
.sg-approve-text {
    font-family:'Plus Jakarta Sans',sans-serif;
    font-size: clamp(2rem, 6vw, 4.5rem);
    font-weight:900; letter-spacing:0.06em;
    background:linear-gradient(135deg,#00e896,#17a870,#2d4a8a);
    -webkit-background-clip:text; -webkit-text-fill-color:transparent;
    filter:drop-shadow(0 0 30px rgba(0,232,150,0.5));
    animation: sgTextRise 0.6s cubic-bezier(0.16,1,0.3,1) 0.4s both;
}
.sg-approve-sub {
    font-family:'JetBrains Mono',monospace;
    font-size:0.85rem; letter-spacing:0.2em;
    color:rgba(0,200,140,0.6);
    animation: sgTextRise 0.6s cubic-bezier(0.16,1,0.3,1) 0.6s both;
}
@keyframes sgTextRise {
    from { opacity:0; transform:translateY(20px); }
    to   { opacity:1; transform:translateY(0); }
}

/* ── TOAST BİLDİRİM ──────────────────────────────────────── */
#sg-toast-container {
    position:fixed; bottom:28px; right:28px;
    z-index:9999; display:flex; flex-direction:column; gap:10px;
    pointer-events:none;
}
.sg-toast {
    background:#fff; border-radius:14px;
    padding:14px 18px; min-width:280px; max-width:360px;
    box-shadow:0 8px 32px rgba(0,0,0,0.12), 0 2px 8px rgba(0,0,0,0.08);
    border-left:4px solid var(--sg);
    display:flex; align-items:flex-start; gap:12px;
    animation: sgToastIn 0.4s cubic-bezier(0.16,1,0.3,1) both,
               sgToastOut 0.35s ease 4.5s forwards;
    pointer-events:all; cursor:pointer;
}
.sg-toast.warning { border-left-color: var(--amber); }
.sg-toast.error   { border-left-color: var(--red); }
@keyframes sgToastIn  { from{transform:translateX(110%);opacity:0;} to{transform:translateX(0);opacity:1;} }
@keyframes sgToastOut { to{transform:translateX(110%);opacity:0;} }
.sg-toast-icon { font-size:1.4rem; flex-shrink:0; margin-top:1px; }
.sg-toast-body { flex:1; }
.sg-toast-title { font-weight:700; font-size:0.82rem; color:#0f1923; margin-bottom:2px; }
.sg-toast-msg   { font-size:0.74rem; color:#5a7a6a; line-height:1.4; }

/* ── AI ROBOT (fare takipçi) ─────────────────────────────── */
#sg-robot-wrap {
    position:fixed; bottom:28px; left:28px;
    z-index:9998; pointer-events:none;
    transition: transform 0.18s ease;
}
#sg-robot-svg {
    width:72px; height:72px;
    filter: drop-shadow(0 4px 16px rgba(17,133,91,0.35));
    animation: sgRobotFloat 3s ease-in-out infinite;
    cursor: pointer; pointer-events:all;
}
@keyframes sgRobotFloat {
    0%,100%{transform:translateY(0);} 50%{transform:translateY(-8px);}
}
#sg-robot-bubble {
    position:absolute; bottom:80px; left:0;
    background:#fff; border:1px solid rgba(17,133,91,0.2);
    border-radius:14px 14px 14px 4px;
    padding:10px 14px; min-width:180px; max-width:240px;
    font-family:'Plus Jakarta Sans',sans-serif;
    font-size:0.75rem; color:#0f1923; line-height:1.5;
    box-shadow:0 4px 20px rgba(0,0,0,0.1);
    opacity:0; transform:translateY(6px) scale(0.95);
    transition:all 0.25s ease;
    pointer-events:none;
}
#sg-robot-wrap:hover #sg-robot-bubble,
#sg-robot-bubble.visible {
    opacity:1; transform:translateY(0) scale(1);
}
#sg-robot-bubble::before {
    content:'⚡ STINGA AI';
    display:block; font-size:0.58rem; font-weight:700;
    color:var(--sg); letter-spacing:1.5px; margin-bottom:4px;
}

</style>
""", unsafe_allow_html=True)

# ─── KULLANICI SİSTEMİ ────────────────────────────────────────
def hash_password(pwd):
    return hashlib.sha256(pwd.encode()).hexdigest()

# ── Kullanıcı Sistemi ─────────────────────────────────────────
# role: "admin"   → tüm fişleri görür + onaylayabilir (Zeynep, Serkan)
# role: "user"    → sadece kendi fişlerini görür (Okan İlhan, Şenol Faik Özyaman)
# role: "admin"   → tüm fişleri görür (Zeynep Özyaman, Serkan Güzdemir)
#
# Şifre → WhatsApp bot sifre alanıyla eşleşiyor:
#   Zeynep Özyaman: 789 | Serkan Güzdemir: 123 | Okan İlhan: 321 | Şenol Faik Özyaman: 456
USERS = {
    "zeynep": {
        "name": "Zeynep Özyaman",
        "password": hash_password("789"),
        "role": "admin",
        "avatar": "👑",
        "title": "Yönetim Kurulu Başkanı",
        "department": "Yönetim Kurulu",
        "monthly_limit": 50000,
        "xp": 0
    },
    "senol": {
        "name": "Şenol Faik Özyaman",
        "password": hash_password("456"),
        "role": "user",
        "avatar": "🏢",
        "title": "Genel Müdür",
        "department": "Genel Müdürlük",
        "monthly_limit": 30000,
        "xp": 0
    },
    "serkan": {
        "name": "Serkan Güzdemir",
        "password": hash_password("123"),
        "role": "admin",
        "avatar": "📊",
        "title": "İşletme Müdürü",
        "department": "İşletme Müdürlüğü",
        "monthly_limit": 25000,
        "xp": 0
    },
    "okan": {
        "name": "Okan İlhan",
        "password": hash_password("321"),
        "role": "user",
        "avatar": "🔧",
        "title": "Saha Personeli",
        "department": "Saha",
        "monthly_limit": 5000,
        "xp": 0
    },
}

# ─── SESSION STATE ────────────────────────────────────────────
defaults = {
    'authenticated': False,
    'user_info': None,
    'splash_done': False,
    'current_key_idx': 0,
    'chat_history': [],
    'dark_mode': True,
    'last_ai_insight': None,
    'tour_done': False,
    'voice_input': None,
    'realtime_alerts': [],
    'selected_page': '🏠 Dashboard'
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─── API ──────────────────────────────────────────────────────
API_KEYS = [
    st.secrets.get("GEMINI_API_KEY_1", ""),
    st.secrets.get("GEMINI_API_KEY_2", "")
]
API_KEYS = [k for k in API_KEYS if k]  # boş olanları çıkar

def configure_ai():
    try:
        k = API_KEYS[st.session_state.current_key_idx]
        genai.configure(api_key=k)
        return genai.GenerativeModel('models/gemini-2.5-flash')
    except Exception as e:
        st.error(f"AI Hatası: {e}")
        return None

# ─── RAILWAY API BAĞLANTISI ──────────────────────────────────
# Bot Railway'de çalışıyor. Dashboard oradan okur/yazar.
import requests as _req

RAILWAY_URL = st.secrets.get("BOT_API_URL", "https://stingafinans-production.up.railway.app")
_API_TIMEOUT = 12

def init_db():
    """API hazır mı kontrol et, session_state'i hazırla."""
    pass  # Veri Railway'den geliyor, lokal init gerekmiyor


@st.cache_data(ttl=15)
def load_data():
    """Railway API'den tüm veriyi çek. 15 sn cache."""
    try:
        r = _req.get(f"{RAILWAY_URL}/all-data", timeout=_API_TIMEOUT)
        r.raise_for_status()
        data = r.json()
        # ── Expense verilerindeki HTML kalıntılarını temizle ──
        _html_fields = ("AI_Audit", "Notlar", "AI_Anomali_Aciklama", "IlgincDetay", "Firma")
        for exp in data.get("expenses", []):
            for _fld in _html_fields:
                val = exp.get(_fld)
                if val and isinstance(val, str) and ('<' in val or '&nbsp' in val):
                    exp[_fld] = _re.sub(r'<[^>]+>', '', val)
                    exp[_fld] = _re.sub(r'&nbsp;?', ' ', exp[_fld])
                    exp[_fld] = _re.sub(r'&[a-z]+;', ' ', exp[_fld])
                    exp[_fld] = _re.sub(r'\s+', ' ', exp[_fld]).strip()
        return data
    except Exception as e:
        st.error(f"⚠️ Railway bağlantı hatası: {e}")
        return {
            "expenses": [], "wallets": {}, "budgets": {}, "ledger": [],
            "notifications": [], "xp": {}, "fis_sayaci": {}, "badges": [],
            "ai_insights": [], "anomaly_log": []
        }


def save_data(d):
    """Lokal kayıt yok — Railway API üzerinden yazılıyor. Bu fonksiyon artık kullanılmıyor."""
    pass


def _api_post(endpoint: str, payload: dict) -> dict:
    """Railway API'ye POST atar, sonucu döndürür."""
    try:
        r = _req.post(f"{RAILWAY_URL}{endpoint}", json=payload, timeout=_API_TIMEOUT)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        st.error(f"API Hatası ({endpoint}): {e}")
        return {"ok": False, "error": str(e)}


def api_approve(fis_id: str, action: str, approver: str) -> bool:
    """Fişi onayla veya reddet."""
    res = _api_post("/approve", {"ID": fis_id, "action": action, "approver": approver})
    if res.get("ok"):
        st.cache_data.clear()
        return True
    return False


def api_transfer(hedef: str, miktar: float, aciklama: str, gonderen: str) -> bool:
    """Harcırah transferi."""
    res = _api_post("/transfer", {"hedef": hedef, "miktar": miktar,
                                   "aciklama": aciklama, "gonderen": gonderen})
    if res.get("ok"):
        st.cache_data.clear()
        return True
    return False


def api_add_expense(expense: dict) -> bool:
    """Dashboard'dan fiş ekle."""
    res = _api_post("/add-expense", expense)
    if res.get("ok"):
        st.cache_data.clear()
        return True
    return False


def add_notify(target, message, notif_type="info", d=None):
    """Eski uyumluluk shim — artık API üzerinden çalışıyor."""
    pass


def add_xp(user_name, amount, reason="", d=None):
    """XP günceller — bildirim GÖNDERİLMEZ."""
    pass  # Railway API üzerinden yönetiliyor, local bildirim yok


def compute_wallet(user_nm: str, data_store: dict) -> float:
    """
    Kişi bakiyesini hesapla:
      + ledger'dan gelen avans transferleri
      - onaylı harcırah/nakit/kisisel harcamalar
    Negatif olabilir (şirkete borç).
    """
    def _nm(a, b):
        """İsim eşleştirme: tam eşit, ya da biri diğerini içeriyor (büyük/küçük harf yok sayılır)."""
        a = str(a).lower().strip()
        b = str(b).lower().strip()
        if not a or not b:
            return False
        return a == b or a in b or b in a

    ledger    = data_store.get("ledger", [])
    expenses  = data_store.get("expenses", [])

    # Yöneticinin verdiği avans (ledger transferleri)
    avans = sum(
        float(e.get("miktar", e.get("Miktar", 0)))
        for e in ledger
        if _nm(e.get("hedef", e.get("Hedef", "")), user_nm)
    )

    # Onaylı harcırah/nakit/cep harcamaları
    harcama = sum(
        float(e.get("Tutar", 0))
        for e in expenses
        if _nm(e.get("Kullanıcı", ""), user_nm)
        and str(e.get("Durum", "")).strip() in ("Onaylandı", "Onaylandi", "onaylandi", "onaylandı")
        and str(e.get("Odeme_Turu", "")).lower().strip() in (
            "harcirah", "nakit", "kisisel",
            "harcırahtan düş", "harcirahtan dus",
            "harcırahtan düş (nakit / kişisel kart)"
        )
    )

    return avans - harcama  # Negatif = şirkete borç


def get_user_wallet_balance(user_nm: str, data_store: dict) -> float:
    """
    Kullanıcının güncel harcırah bakiyesini döndürür.
    1. wallets dict'inde ara (API'den direkt gelen değer) — hem isim hem username dener
    2. Ledger tabanlı compute_wallet ile hesapla
    Birden fazla varyasyonla eşleştirme yapar (Serkan, serkan, Serkan Güzdemir vb.)
    """
    def _nm(a, b):
        a = str(a).lower().strip()
        b = str(b).lower().strip()
        if not a or not b:
            return False
        return a == b or a in b or b in a

    # Aranacak isim varyasyonları: verilen isim + USERS'dan eşleşen display name ve username
    search_names = {str(user_nm).lower().strip()}
    for _ukey, _uinfo in USERS.items():
        if _nm(_uinfo.get("name", ""), user_nm) or _nm(_ukey, user_nm):
            search_names.add(_uinfo.get("name", "").lower().strip())
            search_names.add(_ukey.lower().strip())

    # wallets dict görmezden gelinir — herkes sıfır bakiye ile başlar
    # Sadece gerçek ledger avans transferleri bakiyeyi değiştirir

    # Ledger tabanlı hesapla (tüm varyasyonlarla)
    best = compute_wallet(user_nm, data_store)
    if best == 0:
        for alt_name in search_names:
            if alt_name != str(user_nm).lower().strip():
                alt_val = compute_wallet(alt_name, data_store)
                if alt_val != 0:
                    return alt_val
    return best


# ─── YARDIMCI ─────────────────────────────────────────────────
def extract_json(text):
    try:
        text = text.replace("```json", "").replace("```", "").strip()
        m = re.search(r'\{.*\}', text, re.DOTALL)
        return json.loads(m.group()) if m else json.loads(text)
    except:
        return None

def tr_fix(text):
    if not isinstance(text, str): return str(text)
    chars = {"İ":"I","ı":"i","Ş":"S","ş":"s","Ğ":"G","ğ":"g","Ç":"C","ç":"c","Ö":"O","ö":"o","Ü":"U","ü":"u"}
    for t, e in chars.items(): text = text.replace(t, e)
    return text

def get_risk_html(score):
    if score < 30:
        return f'<span class="risk-badge risk-low">✓ {score}% DÜŞÜK</span>'
    elif score < 70:
        return f'<span class="risk-badge risk-mid">⚠ {score}% ORTA</span>'
    else:
        return f'<span class="risk-badge risk-high">⛔ {score}% KRİTİK</span>'

def get_status_html(status):
    mapping = {
        "Onaylandı":    ("pill-approved", "✓ ONAYLANDI"),
        "Onay Bekliyor":("pill-pending",  "⏳ BEKLİYOR"),
        "Reddedildi":   ("pill-rejected", "✗ REDDEDİLDİ")
    }
    cls, label = mapping.get(status, ("pill-pending", status))
    return f'<span class="status-pill {cls}">{label}</span>'

def img_to_b64(img_path):
    if os.path.exists(img_path):
        with open(img_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return ""

# ─── TÜRKÇE FONT KAYDI ────────────────────────────────────────
# Önce sistemde arar, yoksa otomatik indirir (Railway uyumlu)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import urllib.request

_FONT_CACHE = os.path.join(os.path.expanduser("~"), ".stinga_fonts")
os.makedirs(_FONT_CACHE, exist_ok=True)

_DEJAVU_URLS = {
    "DejaVuSans.ttf":            "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf",
    "DejaVuSans-Bold.ttf":       "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans-Bold.ttf",
    "DejaVuSans-Oblique.ttf":    "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans-Oblique.ttf",
    "DejaVuSans-BoldOblique.ttf":"https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans-BoldOblique.ttf",
}

def _find_or_dl(filename):
    # Önce sabit yollar + repo'nun yanındaki fonts/ klasörü
    _script_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else "."
    search = [
        os.path.join(_script_dir, "fonts"),          # repo'daki fonts/ klasörü
        "/usr/share/fonts/truetype/dejavu",
        "/usr/share/fonts/dejavu",
        _FONT_CACHE,
        "fonts",
        ".",
    ]
    for d in search:
        p = os.path.join(d, filename)
        if os.path.exists(p): return p
    dest = os.path.join(_FONT_CACHE, filename)
    if not os.path.exists(dest):
        try: urllib.request.urlretrieve(_DEJAVU_URLS[filename], dest)
        except: return None
    return dest if os.path.exists(dest) else None

def _reg_font(name, filename):
    p = _find_or_dl(filename)
    if p:
        try: pdfmetrics.registerFont(TTFont(name, p)); return True
        except: pass
    return False

_fonts_ok = (
    _reg_font("DJ",    "DejaVuSans.ttf") and
    _reg_font("DJ-B",  "DejaVuSans-Bold.ttf") and
    _reg_font("DJ-I",  "DejaVuSans-Oblique.ttf") and
    _reg_font("DJ-BI", "DejaVuSans-BoldOblique.ttf")
)
_FN  = "DJ"   if _fonts_ok else "Helvetica"
_FNB = "DJ-B" if _fonts_ok else "Helvetica-Bold"
_FNI = "DJ-I" if _fonts_ok else "Helvetica-Oblique"

_TR_MAP = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSOoCc")
def _tr(t):
    """Font varsa Türkçe, yoksa ASCII fallback."""
    if _fonts_ok: return str(t)
    return str(t).translate(_TR_MAP)

# ─── RAPOR RENK PALETİ ────────────────────────────────────────
_R_DARK  = rl_colors.HexColor("#1B3A5C")
_R_GREEN = rl_colors.HexColor("#1D7A5F")
_R_LIGHT = rl_colors.HexColor("#EEF2F7")
_R_LINE  = rl_colors.HexColor("#D1D5DB")
_R_RED   = rl_colors.HexColor("#DC2626")
_R_ORN   = rl_colors.HexColor("#D97706")
_R_GRN   = rl_colors.HexColor("#059669")

_KDV_ORANLARI = {"yakıt":0.20,"yemek":0.10,"konaklama":0.20,"ulaşım":0.20,"kırtasiye":0.20}

def _kdv_hesapla(tutar, kdv_field, kategori):
    try:
        kdv = float(kdv_field) if kdv_field and float(kdv_field) > 0 else None
    except (TypeError, ValueError):
        kdv = None
    if kdv and kdv < tutar:
        return round(tutar - kdv, 2), round(kdv, 2)
    k = str(kategori).lower()
    oran = next((v for key,v in _KDV_ORANLARI.items() if key in k), 0.20)
    net = round(tutar / (1 + oran), 2)
    return net, round(tutar - net, 2)

def _rl_styles():
    def s(name, **kw): return ParagraphStyle(name, **kw)
    return {
        "title":   s("T",  fontName=_FNB, fontSize=16, textColor=rl_colors.white, alignment=TA_CENTER, leading=20),
        "sub":     s("ST", fontName=_FN,  fontSize=9,  textColor=rl_colors.HexColor("#B0C4D8"), alignment=TA_CENTER, leading=13),
        "sec":     s("SC", fontName=_FNB, fontSize=10, textColor=_R_DARK, spaceBefore=12, spaceAfter=4),
        "n":       s("N",  fontName=_FN,  fontSize=8.5,textColor=rl_colors.HexColor("#374151"), leading=12),
        "sm":      s("SM", fontName=_FN,  fontSize=7.5,textColor=rl_colors.HexColor("#6B7280"), leading=11),
        "ft":      s("FT", fontName=_FNI, fontSize=7,  textColor=rl_colors.HexColor("#6B7280"), alignment=TA_CENTER),
        "r":       s("R",  fontName=_FN,  fontSize=7.5,textColor=rl_colors.HexColor("#374151"), alignment=TA_RIGHT),
        "b":       s("B",  fontName=_FNB, fontSize=8.5,textColor=_R_DARK),
        "gr":      s("GR", fontName=_FNB, fontSize=9,  textColor=_R_GRN),
    }

def _tbl_s(hbg=None):
    if hbg is None: hbg = _R_DARK
    ts = TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), hbg),
        ("TEXTCOLOR",     (0,0),(-1,0), rl_colors.white),
        ("FONTNAME",      (0,0),(-1,0), _FNB),
        ("FONTSIZE",      (0,0),(-1,0), 8),
        ("TOPPADDING",    (0,0),(-1,0), 6), ("BOTTOMPADDING",(0,0),(-1,0),6),
        ("FONTNAME",      (0,1),(-1,-1), _FN),
        ("FONTSIZE",      (0,1),(-1,-1), 7.5),
        ("TOPPADDING",    (0,1),(-1,-1), 4), ("BOTTOMPADDING",(0,1),(-1,-1),4),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [rl_colors.white, _R_LIGHT]),
        ("GRID",          (0,0),(-1,-1), 0.4, _R_LINE),
        ("LINEBELOW",     (0,0),(-1,0), 1, _R_GREEN),
        ("LEFTPADDING",   (0,0),(-1,-1), 6), ("RIGHTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ])
    return ts

def _ph(txt, st): return Paragraph(txt, st)

def export_pdf_muhasebe(df_raw, title="Mali Rapor", donem="Tüm Zamanlar", logo_path=None):
    """Profesyonel muhasebe PDF raporu — KDV ayrıntılı, kategori kırılımlı."""
    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=1.8*cm, rightMargin=1.8*cm,
            topMargin=1.5*cm, bottomMargin=2*cm)
        W = A4[0] - 3.6*cm
        ST = _rl_styles()
        story = []
        df = df_raw.copy()

        if df.empty:
            story.append(_ph("Bu döneme ait veri bulunmamaktadır.", ST["n"]))
            doc.build(story); return buf.getvalue()

        # KDV hesapları
        df["_t"] = pd.to_numeric(df.get("Tutar", 0), errors="coerce").fillna(0)
        df["_n"] = 0.0; df["_k"] = 0.0
        for i, row in df.iterrows():
            kv = df.get("KDV", pd.Series([None]*len(df))).iloc[i] if "KDV" in df.columns else None
            kt = df.get("Kategori", pd.Series([""]*len(df))).iloc[i] if "Kategori" in df.columns else ""
            n, k = _kdv_hesapla(row["_t"], kv, kt)
            df.at[i,"_n"] = n; df.at[i,"_k"] = k

        def col(*names):
            for n in names:
                if n in df.columns: return n
            return None
        c_tar = col("Tarih","Fis_Tarihi"); c_per = col("Kullanıcı","Personel")
        c_fir = col("Firma"); c_kat = col("Kategori"); c_dur = col("Durum")
        c_ode = col("Odeme_Turu","OdemeTipi")
        def v(row, c): return str(row[c]) if c and c in row.index else "-"

        tot_b = df["_t"].sum(); tot_k = df["_k"].sum(); tot_n = df["_n"].sum()
        onay_df = df[df[c_dur].str.contains("Onay",case=False,na=False)] if c_dur else df
        onay_b = onay_df["_t"].sum(); onay_k = onay_df["_k"].sum(); onay_n = onay_df["_n"].sum()
        fmt = lambda x: f"₺{x:,.2f}"

        # ── BAŞLIK ──
        logo_cell = Spacer(3*cm, 3*cm)
        if logo_path and RLImage:
            try:
                from PIL import Image as _PI
                import numpy as _np, io as _bio
                _img = _PI.open(logo_path).convert('RGBA')
                _d = _np.array(_img)
                _d[(_d[:,:,0]<50)&(_d[:,:,1]<50)&(_d[:,:,2]<50), 3] = 0
                _img = _PI.fromarray(_d, 'RGBA')
                _w, _h = _img.size; _s = min(_w,_h)
                _img = _img.crop(((_w-_s)//2,(_h-_s)//2,(_w+_s)//2,(_h+_s)//2)).resize((400,400), _PI.LANCZOS)
                _lb = _bio.BytesIO(); _img.save(_lb, format='PNG'); _lb.seek(0)
                logo_cell = RLImage(_lb, width=4*cm, height=4*cm)
            except:
                pass
        _logo_w = 5.2*cm
        _title_w = W - _logo_w
        hdr_items = [
            Spacer(1,0.05*cm),
            _ph("STİNGA ENERJİ A.Ş.", ST["title"]),
            _ph("GİDER VE KDV RAPORU", ST["title"]),
            _ph(f"Dönem: {donem}  |  {now_ist().strftime('%d.%m.%Y %H:%M')}  |  Stinga Pro v15.0", ST["sub"]),
        ]
        hdr = Table([[logo_cell, hdr_items]], colWidths=[_logo_w, _title_w])
        hdr.setStyle(TableStyle([
            ("BACKGROUND",  (0,0),(0,0), rl_colors.white),
            ("ALIGN",       (0,0),(0,0), "CENTER"),
            ("VALIGN",      (0,0),(0,0), "MIDDLE"),
            ("TOPPADDING",  (0,0),(0,0), 3), ("BOTTOMPADDING",(0,0),(0,0), 3),
            ("LEFTPADDING", (0,0),(0,0), 6), ("RIGHTPADDING", (0,0),(0,0), 6),
            ("BACKGROUND",  (1,0),(1,0), _R_DARK),
            ("ALIGN",       (1,0),(1,0), "CENTER"),
            ("VALIGN",      (1,0),(1,0), "MIDDLE"),
            ("TOPPADDING",  (1,0),(1,0), 16), ("BOTTOMPADDING",(1,0),(1,0), 16),
            ("LEFTPADDING", (1,0),(1,0), 16), ("RIGHTPADDING", (1,0),(1,0), 20),
        ]))
        story.append(hdr); story.append(Spacer(1,0.5*cm))

        # ── ÖZET KARTLAR ──
        def card(lbl, br, net, kdv, bg):
            return [
                _ph(f"<b>{lbl}</b>", ParagraphStyle("cl",fontName=_FNB,fontSize=7.5,
                    textColor=rl_colors.HexColor("#6B7280"),alignment=TA_CENTER)),
                _ph(f"<b>{fmt(br)}</b>", ParagraphStyle("cv",fontName=_FNB,fontSize=11,
                    textColor=_R_DARK,alignment=TA_CENTER,spaceBefore=2)),
                _ph(f"KDV Hariç: {fmt(net)}", ParagraphStyle("cs",fontName=_FN,fontSize=7,
                    textColor=rl_colors.HexColor("#6B7280"),alignment=TA_CENTER)),
                _ph(f"KDV: {fmt(kdv)}", ParagraphStyle("ck",fontName=_FN,fontSize=7,
                    textColor=_R_GRN,alignment=TA_CENTER)),
            ]
        onay_sayi = len(onay_df)
        cards = Table([[
            card("TOPLAM GİDER (BRÜT)", tot_b, tot_n, tot_k, _R_LIGHT),
            card("ONAYLANAN GİDER", onay_b, onay_n, onay_k, rl_colors.HexColor("#ECFDF5")),
            [_ph("<b>KDV İADE HAKKI</b>", ParagraphStyle("cl",fontName=_FNB,fontSize=7.5,
                 textColor=rl_colors.HexColor("#6B7280"),alignment=TA_CENTER)),
             _ph(f"<b>{fmt(onay_k)}</b>", ParagraphStyle("cv",fontName=_FNB,fontSize=14,
                 textColor=_R_GRN,alignment=TA_CENTER,spaceBefore=2)),
             _ph(f"{onay_sayi} onaylı fiş", ParagraphStyle("cs",fontName=_FN,fontSize=7,
                 textColor=rl_colors.HexColor("#6B7280"),alignment=TA_CENTER)),
             _ph(f"Oran: %{(onay_k/onay_b*100) if onay_b else 0:.1f}",
                 ParagraphStyle("ck",fontName=_FN,fontSize=7,textColor=_R_GRN,alignment=TA_CENTER))],
        ]], colWidths=[W/3]*3)
        cards.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,0),_R_LIGHT),
            ("BACKGROUND",(1,0),(1,0),rl_colors.HexColor("#ECFDF5")),
            ("BACKGROUND",(2,0),(2,0),rl_colors.HexColor("#EFF6FF")),
            ("BOX",(0,0),(0,0),0.5,_R_LINE),("BOX",(1,0),(1,0),0.5,_R_LINE),("BOX",(2,0),(2,0),0.5,_R_LINE),
            ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
        ]))
        story.append(cards); story.append(Spacer(1,0.6*cm))

        # ── KATEGORİ ANALİZİ ──
        story.append(_ph("KATEGORİ BAZLI GİDER ANALİZİ", ST["sec"]))
        story.append(HRFlowable(width=W, thickness=1, color=_R_GREEN, spaceAfter=6))
        if c_kat:
            grp = df.groupby(c_kat).agg(Fis=("_t","count"),Brut=("_t","sum"),Net=("_n","sum"),KDV=("_k","sum")).reset_index().sort_values("Brut",ascending=False)
            def ph(t): return _ph(f"<b>{t}</b>", ParagraphStyle("h",fontName=_FNB,fontSize=8,textColor=rl_colors.white))
            def pr(t,al=TA_LEFT,bold=False,col=None):
                c2 = col or rl_colors.HexColor("#374151")
                fn = _FNB if bold else _FN
                return _ph(t, ParagraphStyle("d",fontName=fn,fontSize=7.5,textColor=c2,alignment=al))
            kat_data = [[ph("KATEGORİ"),ph("FİŞ"),ph("BRÜT"),ph("KDV HARİÇ"),ph("KDV"),ph("ORAN")]]
            for _,r in grp.iterrows():
                oran = r["KDV"]/r["Brut"]*100 if r["Brut"] else 0
                kat_data.append([pr(str(r[c_kat]).title()),pr(str(int(r["Fis"])),TA_CENTER),
                    pr(fmt(r["Brut"]),TA_RIGHT),pr(fmt(r["Net"]),TA_RIGHT),
                    pr(fmt(r["KDV"]),TA_RIGHT,True,_R_GRN),pr(f"%{oran:.0f}",TA_CENTER)])
            last = len(kat_data)
            kat_data.append([pr("<b>TOPLAM</b>",TA_LEFT,True),pr(f"<b>{len(df)}</b>",TA_CENTER,True),
                pr(f"<b>{fmt(tot_b)}</b>",TA_RIGHT,True),pr(f"<b>{fmt(tot_n)}</b>",TA_RIGHT,True),
                pr(f"<b>{fmt(tot_k)}</b>",TA_RIGHT,True,_R_GRN),
                pr(f"<b>%{(tot_k/tot_b*100) if tot_b else 0:.1f}</b>",TA_CENTER,True)])
            cw = [W*.22,W*.10,W*.17,W*.17,W*.17,W*.17]
            kt = Table(kat_data, colWidths=cw)
            ts2 = _tbl_s()
            ts2.add("BACKGROUND",(0,last),(-1,last),rl_colors.HexColor("#DBEAFE"))
            ts2.add("LINEABOVE",(0,last),(-1,last),1,_R_DARK)
            kt.setStyle(ts2)
            story.append(kt)
        story.append(Spacer(1,0.6*cm))

        # ── PERSONEL DAĞILIMI ──
        if c_per:
            story.append(_ph("PERSONEL BAZLI GİDER DAĞILIMI", ST["sec"]))
            story.append(HRFlowable(width=W,thickness=1,color=_R_GREEN,spaceAfter=6))
            pgrp = df.groupby(c_per).agg(Fis=("_t","count"),Brut=("_t","sum"),Net=("_n","sum"),KDV=("_k","sum")).reset_index().sort_values("Brut",ascending=False)
            def phh(t): return _ph(f"<b>{t}</b>",ParagraphStyle("ph",fontName=_FNB,fontSize=8,textColor=rl_colors.white))
            def prd(t,al=TA_LEFT,bold=False,col=None):
                c2=col or rl_colors.HexColor("#374151"); fn=_FNB if bold else _FN
                return _ph(t,ParagraphStyle("pd",fontName=fn,fontSize=7.5,textColor=c2,alignment=al))
            pd_data=[[phh("PERSONEL"),phh("FİŞ"),phh("TOPLAM BRÜT"),phh("KDV HARİÇ"),phh("KDV TUTARI")]]
            for _,r in pgrp.iterrows():
                pd_data.append([prd(str(r[c_per])),prd(str(int(r["Fis"])),TA_CENTER),
                    prd(fmt(r["Brut"]),TA_RIGHT),prd(fmt(r["Net"]),TA_RIGHT),
                    prd(fmt(r["KDV"]),TA_RIGHT,True,_R_GRN)])
            pt=Table(pd_data,colWidths=[W*.30,W*.10,W*.20,W*.20,W*.20])
            pt.setStyle(_tbl_s()); story.append(pt); story.append(Spacer(1,0.6*cm))

        # ── DETAY FİŞ TABLOSU ──
        story.append(_ph("GİDER FİŞ DETAY LİSTESİ", ST["sec"]))
        story.append(HRFlowable(width=W,thickness=1,color=_R_GREEN,spaceAfter=6))
        def dh(t): return _ph(f"<b>{t}</b>",ParagraphStyle("dh",fontName=_FNB,fontSize=7.5,textColor=rl_colors.white,alignment=TA_CENTER))
        det=[[dh("TARİH"),dh("PERSONEL"),dh("FİRMA"),dh("KATEGORİ"),dh("BRÜT"),dh("KDV HARİÇ"),dh("KDV"),dh("ÖDEME"),dh("DURUM")]]
        for _,row in df.iterrows():
            dur=v(row,c_dur)
            dc=_R_GRN if ("Onay" in dur and "Bekl" not in dur) else (_R_RED if "Red" in dur else _R_ORN)
            def dd(t,al=TA_LEFT,c=None,bold=False):
                return _ph(t,ParagraphStyle("dd",fontName=_FNB if bold else _FN,fontSize=7.5,
                    textColor=c or rl_colors.HexColor("#374151"),alignment=al))
            det.append([dd(v(row,c_tar)),dd(v(row,c_per)),dd(v(row,c_fir)),dd(v(row,c_kat).title()),
                dd(fmt(row["_t"]),TA_RIGHT),dd(fmt(row["_n"]),TA_RIGHT),
                dd(fmt(row["_k"]),TA_RIGHT,_R_GRN,True),
                dd(v(row,c_ode).replace("_"," ").title() if c_ode else "-"),
                dd(dur,TA_CENTER,dc,True)])
        cw3=[W*.09,W*.13,W*.16,W*.10,W*.10,W*.10,W*.09,W*.10,W*.13]
        dt=Table(det,colWidths=cw3,repeatRows=1); dt.setStyle(_tbl_s()); story.append(dt)
        story.append(Spacer(1,0.8*cm))

        # ── KDV İADE KUTUSU ──
        story.append(KeepTogether([
            HRFlowable(width=W,thickness=1,color=_R_DARK,spaceBefore=4,spaceAfter=8),
            _ph("KDV İADE ÖZETİ", ST["sec"]),
        ]))
        def kd(l,v2,bold=False,vcol=None):
            fn=_FNB if bold else _FN
            sz=12 if bold else 9
            return [
                _ph(l,ParagraphStyle("kl",fontName=fn,fontSize=sz,textColor=_R_DARK)),
                _ph(v2,ParagraphStyle("kv",fontName=fn,fontSize=sz,textColor=vcol or _R_DARK,alignment=TA_RIGHT))
            ]
        kdv_tbl=Table([
            kd("Onaylanan Toplam Brüt Tutar", fmt(onay_b)),
            kd("Onaylanan Toplam KDV Hariç Tutar", fmt(onay_n)),
            kd("İade Talep Edilebilecek KDV", f"<b>{fmt(onay_k)}</b>", bold=True, vcol=_R_GRN),
        ], colWidths=[W*.6,W*.4])
        kdv_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,1),_R_LIGHT),
            ("BACKGROUND",(0,2),(-1,2),rl_colors.HexColor("#ECFDF5")),
            ("LINEBELOW",(0,1),(-1,1),1,_R_DARK),
            ("BOX",(0,0),(-1,-1),1,_R_GREEN),
            ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
            ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(kdv_tbl)
        story.append(Spacer(1,0.5*cm))
        story.append(HRFlowable(width=W,thickness=0.5,color=_R_LINE))
        story.append(Spacer(1,0.2*cm))
        story.append(_ph(_tr("Bu rapor Stinga Pro v15.0 Mali Yonetim Sistemi tarafindan otomatik uretilmistir. KDV hesaplamalari fis verileri ve standart KDV oranlari baz alinarak yapilmistir. Resmi muhasebe islemleri icin yetkili mali musavirinize danisiniz."), ST["ft"]))

        doc.build(story)
        return buf.getvalue()
    except Exception as e:
        return b"%PDF-1.4\n"


def export_excel_muhasebe(df_raw, donem="Tüm Zamanlar", logo_path=None):
    """Profesyonel muhasebe Excel raporu — 4 sayfalı, KDV kırılımlı."""
    try:
        df = df_raw.copy()
        df["_t"] = pd.to_numeric(df.get("Tutar",0), errors="coerce").fillna(0)
        df["_n"] = 0.0; df["_k"] = 0.0
        for i, row in df.iterrows():
            kv = df["KDV"].iloc[i] if "KDV" in df.columns else None
            kt = df["Kategori"].iloc[i] if "Kategori" in df.columns else ""
            n, k = _kdv_hesapla(row["_t"], kv, kt)
            df.at[i,"_n"] = n; df.at[i,"_k"] = k

        def col(*names):
            for n in names:
                if n in df.columns: return n
            return None
        c_tar=col("Tarih","Fis_Tarihi"); c_per=col("Kullanıcı","Personel")
        c_fir=col("Firma"); c_kat=col("Kategori"); c_dur=col("Durum")
        c_ode=col("Odeme_Turu","OdemeTipi")
        def v(row,c): return str(row[c]) if c and c in row.index else "-"

        wb = Workbook()
        MONEY  = '#,##0.00 ₺'
        BDR    = Border(left=Side(style="thin",color="D1D5DB"),right=Side(style="thin",color="D1D5DB"),
                        top=Side(style="thin",color="D1D5DB"),bottom=Side(style="thin",color="D1D5DB"))
        C = XLAlign(horizontal="center",vertical="center",wrap_text=True)
        R = XLAlign(horizontal="right",vertical="center")
        L = XLAlign(horizontal="left",vertical="center")

        def hc(ws,r,c,t,bg="1B3A5C"):
            cell=ws.cell(row=r,column=c,value=t)
            cell.font=XLFont(name="Calibri",bold=True,size=10,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor=bg); cell.alignment=C; cell.border=BDR
        def dc(ws,r,c,v2,fmt=None,bold=False,bg=None,al=None,col=None):
            cell=ws.cell(row=r,column=c,value=v2)
            cell.font=XLFont(name="Calibri",size=10,bold=bold,color=col or "000000")
            cell.alignment=al or L; cell.border=BDR
            if fmt: cell.number_format=fmt
            if bg: cell.fill=PatternFill("solid",fgColor=bg)

        tot_b=df["_t"].sum(); tot_n=df["_n"].sum(); tot_k=df["_k"].sum()
        onay_df=df[df[c_dur].str.contains("Onay",case=False,na=False)] if c_dur else df

        # ── SAYFA 1: ÖZET ──
        ws1=wb.active; ws1.title="📊 Özet"; ws1.sheet_view.showGridLines=False
        ws1.column_dimensions["A"].width=32; ws1.column_dimensions["B"].width=18
        ws1.column_dimensions["C"].width=22
        if logo_path and XLImage:
            try:
                xi=XLImage(logo_path); xi.width=80; xi.height=80; ws1.add_image(xi,"A1")
            except: pass
        ws1.row_dimensions[1].height=70
        ws1.merge_cells("B1:C1"); ws1["B1"].value="STİNGA ENERJİ A.Ş. — GİDER VE KDV RAPORU"
        ws1["B1"].font=XLFont(name="Calibri",bold=True,size=15,color="1B3A5C"); ws1["B1"].alignment=C
        ws1.merge_cells("A2:C2"); ws1["A2"].value=f"Dönem: {donem}  |  {now_ist().strftime('%d.%m.%Y %H:%M')}  |  Stinga Pro v15.0"
        ws1["A2"].font=XLFont(name="Calibri",size=10,italic=True,color="6B7280"); ws1["A2"].alignment=C
        r=4
        for lbl,b,n,k,clr in [
            ("TOPLAM GİDER",tot_b,tot_n,tot_k,"1B3A5C"),
            ("ONAYLANAN GİDER",onay_df["_t"].sum(),onay_df["_n"].sum(),onay_df["_k"].sum(),"1D7A5F"),
        ]:
            ws1.merge_cells(f"A{r}:A{r+2}"); c2=ws1[f"A{r}"]
            c2.value=lbl; c2.font=XLFont(name="Calibri",bold=True,size=11,color="FFFFFF")
            c2.fill=PatternFill("solid",fgColor=clr); c2.alignment=C; c2.border=BDR
            for lbl2,val,rr in [("Toplam Brüt",b,r),("KDV Hariç",n,r+1),("İndirilecek KDV",k,r+2)]:
                ws1[f"B{rr}"].value=lbl2; ws1[f"B{rr}"].font=XLFont(name="Calibri",size=10); ws1[f"B{rr}"].border=BDR; ws1[f"B{rr}"].alignment=L
                ws1[f"C{rr}"].value=val; ws1[f"C{rr}"].number_format=MONEY
                ws1[f"C{rr}"].font=XLFont(name="Calibri",bold=True,size=11,color=clr)
                ws1[f"C{rr}"].alignment=R; ws1[f"C{rr}"].border=BDR
            r+=4

        # ── SAYFA 2: KATEGORİ ──
        ws2=wb.create_sheet("📂 Kategori Analizi"); ws2.sheet_view.showGridLines=False
        for i,w in enumerate([28,10,18,18,18,12],1): ws2.column_dimensions[get_column_letter(i)].width=w
        ws2.merge_cells("A1:F1"); ws2["A1"].value="KATEGORİ BAZLI GİDER ANALİZİ"
        ws2["A1"].font=XLFont(name="Calibri",bold=True,size=13,color="FFFFFF")
        ws2["A1"].fill=PatternFill("solid",fgColor="1B3A5C"); ws2["A1"].alignment=C; ws2.row_dimensions[1].height=28
        for ci,h in enumerate(["KATEGORİ","FİŞ SAYISI","TOPLAM BRÜT","KDV HARİÇ","KDV TUTARI","KDV ORANI"],1):
            hc(ws2,2,ci,h,"1D7A5F")
        ws2.row_dimensions[2].height=22
        if c_kat:
            grp=df.groupby(c_kat).agg(Fis=("_t","count"),Brut=("_t","sum"),Net=("_n","sum"),KDV=("_k","sum")).reset_index().sort_values("Brut",ascending=False)
            ri2=3
            for _,row in grp.iterrows():
                bg="EEF2F7" if ri2%2==0 else None
                oran=row["KDV"]/row["Brut"] if row["Brut"] else 0
                dc(ws2,ri2,1,str(row[c_kat]).title(),bg=bg)
                dc(ws2,ri2,2,int(row["Fis"]),bg=bg,al=C)
                dc(ws2,ri2,3,row["Brut"],MONEY,bg=bg,al=R)
                dc(ws2,ri2,4,row["Net"],MONEY,bg=bg,al=R)
                dc(ws2,ri2,5,row["KDV"],MONEY,True,bg="ECFDF5",al=R,col="059669")
                dc(ws2,ri2,6,oran,"0%",bg=bg,al=C)
                ri2+=1
            tr=ri2+1
            ws2[f"A{tr}"].value="GENEL TOPLAM"; ws2[f"A{tr}"].font=XLFont(name="Calibri",bold=True,size=11,color="1B3A5C")
            ws2[f"A{tr}"].fill=PatternFill("solid",fgColor="DBEAFE"); ws2[f"A{tr}"].alignment=L; ws2[f"A{tr}"].border=BDR
            ws2[f"B{tr}"].value=f"=SUM(B3:B{ri2-1})"; ws2[f"B{tr}"].font=XLFont(name="Calibri",bold=True,size=11)
            ws2[f"B{tr}"].fill=PatternFill("solid",fgColor="DBEAFE"); ws2[f"B{tr}"].alignment=C; ws2[f"B{tr}"].border=BDR
            for cl in ["C","D","E"]:
                ws2[f"{cl}{tr}"].value=f"=SUM({cl}3:{cl}{ri2-1})"
                ws2[f"{cl}{tr}"].number_format=MONEY
                ws2[f"{cl}{tr}"].font=XLFont(name="Calibri",bold=True,size=11,color="1B3A5C")
                ws2[f"{cl}{tr}"].fill=PatternFill("solid",fgColor="DBEAFE"); ws2[f"{cl}{tr}"].alignment=R; ws2[f"{cl}{tr}"].border=BDR
            ws2[f"F{tr}"].value=f"=E{tr}/C{tr}"; ws2[f"F{tr}"].number_format="0%"
            ws2[f"F{tr}"].font=XLFont(name="Calibri",bold=True,size=11)
            ws2[f"F{tr}"].fill=PatternFill("solid",fgColor="DBEAFE"); ws2[f"F{tr}"].alignment=C; ws2[f"F{tr}"].border=BDR

        # ── SAYFA 3: DETAY ──
        ws3=wb.create_sheet("📋 Gider Detayı"); ws3.sheet_view.showGridLines=False; ws3.freeze_panes="A3"
        for i,w in enumerate([13,20,28,14,15,15,15,16,15],1): ws3.column_dimensions[get_column_letter(i)].width=w
        ws3.merge_cells("A1:I1"); ws3["A1"].value=f"GİDER FİŞ DETAY LİSTESİ — {donem}"
        ws3["A1"].font=XLFont(name="Calibri",bold=True,size=13,color="FFFFFF")
        ws3["A1"].fill=PatternFill("solid",fgColor="1B3A5C"); ws3["A1"].alignment=C; ws3.row_dimensions[1].height=28
        for ci,h in enumerate(["TARİH","PERSONEL","FİRMA","KATEGORİ","BRÜT TUTAR","KDV HARİÇ","KDV TUTARI","ÖDEME","DURUM"],1):
            hc(ws3,2,ci,h,"1D7A5F")
        ws3.row_dimensions[2].height=22
        ri3=3
        for _,row in df.iterrows():
            bg="EEF2F7" if ri3%2==0 else None
            dur=v(row,c_dur)
            dc3="059669" if ("Onay" in dur and "Bekl" not in dur) else ("DC2626" if "Red" in dur else "D97706")
            dc(ws3,ri3,1,v(row,c_tar),bg=bg); dc(ws3,ri3,2,v(row,c_per),bg=bg)
            dc(ws3,ri3,3,v(row,c_fir),bg=bg); dc(ws3,ri3,4,v(row,c_kat).title() if c_kat else "-",bg=bg)
            dc(ws3,ri3,5,row["_t"],MONEY,bg=bg,al=R); dc(ws3,ri3,6,row["_n"],MONEY,bg=bg,al=R)
            dc(ws3,ri3,7,row["_k"],MONEY,True,bg="ECFDF5",al=R,col="059669")
            ode=v(row,c_ode).replace("_"," ").title() if c_ode else "-"
            dc(ws3,ri3,8,ode,bg=bg); dc(ws3,ri3,9,dur,bold=True,bg=bg,al=C,col=dc3)
            ri3+=1
        tr3=ri3+1
        for cl,ci in [("E",5),("F",6),("G",7)]:
            ws3[f"{cl}{tr3}"].value=f"=SUM({cl}3:{cl}{ri3-1})"
            ws3[f"{cl}{tr3}"].number_format=MONEY
            ws3[f"{cl}{tr3}"].font=XLFont(name="Calibri",bold=True,size=11,color="1B3A5C")
            ws3[f"{cl}{tr3}"].fill=PatternFill("solid",fgColor="DBEAFE"); ws3[f"{cl}{tr3}"].alignment=R; ws3[f"{cl}{tr3}"].border=BDR

        # ── SAYFA 4: KDV İADE ──
        ws4=wb.create_sheet("🧾 KDV İade Tablosu"); ws4.sheet_view.showGridLines=False
        for i,w in enumerate([30,12,20,20,12],1): ws4.column_dimensions[get_column_letter(i)].width=w
        ws4.merge_cells("A1:E1"); ws4["A1"].value="KDV İADE BAŞVURU TABLOSU — ONAYLANAN GİDERLER"
        ws4["A1"].font=XLFont(name="Calibri",bold=True,size=13,color="FFFFFF")
        ws4["A1"].fill=PatternFill("solid",fgColor="1D7A5F"); ws4["A1"].alignment=C; ws4.row_dimensions[1].height=30
        ws4.merge_cells("A2:E2"); ws4["A2"].value=f"STİNGA ENERJİ A.Ş.  |  Dönem: {donem}  |  Tarih: {now_ist().strftime('%d.%m.%Y')}"
        ws4["A2"].font=XLFont(name="Calibri",size=10,italic=True,color="6B7280"); ws4["A2"].alignment=C
        for ci,h in enumerate(["KATEGORİ","FİŞ SAYISI","MATRAH (KDV HARİÇ)","HESAPLANAN KDV","KDV ORANI"],1):
            hc(ws4,3,ci,h,"1D7A5F")
        ws4.row_dimensions[3].height=22
        if c_kat and len(onay_df):
            kgrp=onay_df.groupby(c_kat).agg(Fis=("_t","count"),Net=("_n","sum"),KDV=("_k","sum")).reset_index()
            ri4=4
            for _,row in kgrp.iterrows():
                bg="EEF2F7" if ri4%2==0 else None
                oran=row["KDV"]/(row["Net"]+row["KDV"]) if (row["Net"]+row["KDV"]) else 0
                dc(ws4,ri4,1,str(row[c_kat]).title(),bg=bg); dc(ws4,ri4,2,int(row["Fis"]),bg=bg,al=C)
                dc(ws4,ri4,3,row["Net"],MONEY,bg=bg,al=R); dc(ws4,ri4,4,row["KDV"],MONEY,True,bg="ECFDF5",al=R,col="059669")
                dc(ws4,ri4,5,oran,"0%",bg=bg,al=C); ri4+=1
            tr4=ri4+1
            ws4.merge_cells(f"A{tr4}:C{tr4}"); ws4[f"A{tr4}"].value="İADE TALEBİ TOPLAM KDV"
            ws4[f"A{tr4}"].font=XLFont(name="Calibri",bold=True,size=12,color="1B3A5C")
            ws4[f"A{tr4}"].fill=PatternFill("solid",fgColor="ECFDF5"); ws4[f"A{tr4}"].border=BDR; ws4[f"A{tr4}"].alignment=L
            ws4[f"D{tr4}"].value=f"=SUM(D4:D{ri4-1})"; ws4[f"D{tr4}"].number_format=MONEY
            ws4[f"D{tr4}"].font=XLFont(name="Calibri",bold=True,size=14,color="059669")
            ws4[f"D{tr4}"].fill=PatternFill("solid",fgColor="ECFDF5"); ws4[f"D{tr4}"].alignment=R; ws4[f"D{tr4}"].border=BDR

        buf=io.BytesIO(); wb.save(buf); return buf.getvalue()
    except Exception as e:
        return b""

# ─── AI FONKSİYONLARI ─────────────────────────────────────────
def analyze_receipt_pro(image, model):
    bugun = now_ist().strftime("%Y-%m-%d")
    prompt = f"""Sen Stinga Enerji şirketinin mali denetçisisin. Fişi dikkatli oku ve SADECE fişte yazanları analiz et.

BUGÜNÜN TARİHİ: {bugun}
- Fişteki tarih bu tarihten önce olmalı. Sonraki tarihse anomali=true.
- Yılı dikkatli oku: 2025 ve 2026 karıştırma. Tarih formatı: YYYY-MM-DD

=== KATEGORİ BELİRLEME ===
- Restoran, kafe, lokanta, köfteci, balık, yemek → kategori: "Yemek"
- Benzin, motorin, akaryakıt, shell, bp, total → kategori: "Yakıt"  
- Otel, pansiyon, konaklama → kategori: "Konaklama"
- Donanım, elektronik, kırtasiye → kategori: "Ekipman"
- Diğer tüm durumlar → kategori: "Diğer"

=== RİSK SKORU ===
- Yemek fişi → risk_skoru: 2
- Yakıt fişi → risk_skoru: 1
- Konaklama → risk_skoru: 2
- Fatura net okunmuyorsa risk +15
- Gece 22:00-06:00 arası fiş → risk +10

=== KİŞİSEL GİDER TESPİTİ (ÇOK DİKKATLİ OL) ===
YALNIZCA fişte açıkça bu ürün isimleri geçiyorsa kisisel_giderler listesine ekle:
sigara, marlboro, winston, bira, beer, alkol, şarap, rakı, viski, vodka, içki,
çikolata, şeker şekerleme, parfüm, kozmetik, şampuan, deodorant,
netflix, spotify, oyun aboneliği, oyuncak

KURAL: Sadece fişte AÇIKÇA yazan ürünleri ekle. Tahmin etme, yorum yapma.
Restoran/lokanta/köfteci fişinde yiyecek kalemi varsa → kisisel_giderler BOŞ bırak.
Genel "yiyecek", "yemek", "food", "meal" ifadeleri kişisel gider DEĞİLDİR.

=== AUDIT ÖZETİ ===
audit_ozeti: Tek cümle, sadece fişin içeriğini açıkla. Kişisel gider yoksa bahsetme.
Örnek: "Şükrü Bey Köftecisi'nden alınan yemek fişi, standart risk seviyesindedir."

SADECE bu JSON formatını döndür, başka hiçbir şey yazma:
{{
    "firma": "Firma Adı",
    "tarih": "YYYY-MM-DD",
    "saat": "HH:MM",
    "toplam_tutar": 0.0,
    "kategori": "Yemek/Yakıt/Konaklama/Ekipman/Diğer",
    "risk_skoru": 2,
    "audit_ozeti": "Tek cümle özet",
    "kalemler": ["sadece fişte yazan ürün isimleri"],
    "kisisel_giderler": [],
    "kdv_tutari": 0.0,
    "odeme_turu": "Nakit/Kredi Kartı/Havale",
    "anomali": false,
    "anomali_aciklamasi": ""
}}"""
    try:
        response = model.generate_content([prompt, image])
        return response.text
    except Exception as e:
        if "429" in str(e).lower() or "quota" in str(e).lower():
            st.session_state.current_key_idx = (st.session_state.current_key_idx + 1) % len(API_KEYS)
            return "ANAHTAR_DEGISIMI"
        return f"HATA: {str(e)}"

def apply_business_rules(data_ai, data_store, user_name):
    """
    Şirket iş kurallarına göre risk skorunu Python tarafında hesaplar.
    AI'ın verdiği skoru bu kurallarla günceller.
    """
    kategori    = str(data_ai.get("kategori", "")).lower()
    tarih       = data_ai.get("tarih", now_ist().strftime("%Y-%m-%d"))
    risk        = int(data_ai.get("risk_skoru", 5))
    anomali     = data_ai.get("anomali", False)
    anomali_msg = data_ai.get("anomali_aciklamasi", "")
    audit       = data_ai.get("audit_ozeti", "")
    kisisel     = data_ai.get("kisisel_giderler", [])
    uyarilar    = []

    # ── Aynı gün aynı kullanıcının fişlerini say ──────────────────
    bugun_fisleri = [
        e for e in data_store.get("expenses", [])
        if e.get("Kullanıcı") == user_name and e.get("Tarih") == tarih
    ]
    bugun_yemek = [e for e in bugun_fisleri if "yemek" in str(e.get("Kategori","")).lower()]
    bugun_yakit = [e for e in bugun_fisleri if "yakıt" in str(e.get("Kategori","")).lower() or "yakit" in str(e.get("Kategori","")).lower()]

    # ── YEMEK KURALLARI ────────────────────────────────────────────
    if "yemek" in kategori:
        kac_inci = len(bugun_yemek) + 1  # Bu fiş kaçıncı olacak
        if kac_inci == 1:
            risk = max(risk, 2)
            audit += " | Günün 1. yemek fişi."
        elif kac_inci == 2:
            risk = max(risk, 6)
            audit += " | Günün 2. yemek fişi."
        elif kac_inci == 3:
            risk = max(risk, 5)
            audit += " | Günün 3. yemek fişi."
        else:
            risk = max(risk, 75)
            anomali = True
            uyari = f"🍽️ Günde {kac_inci}. yemek fişi! Aşırı yemek harcaması tespit edildi."
            anomali_msg = (anomali_msg + " | " + uyari).strip(" | ")
            audit += f" | ⚠️ Günde {kac_inci}. yemek fişi — yüksek risk."
            uyarilar.append(uyari)

    # ── KONAKLAMA KURALLARI ────────────────────────────────────────
    elif "konaklama" in kategori or "otel" in kategori:
        risk = min(risk, 2)
        audit += " | Konaklama fişi — standart risk uygulandı."

    # ── YAKIT KURALLARI ───────────────────────────────────────────
    elif "yakıt" in kategori or "yakit" in kategori or "akaryakıt" in kategori:
        kac_yakit = len(bugun_yakit) + 1
        if kac_yakit == 1:
            risk = max(risk, 1)
            audit += " | Günün ilk yakıt fişi."
        else:
            risk = min(max(risk, kac_yakit * 10), 20)
            uyari = f"⛽ Aynı gün {kac_yakit}. yakıt fişi!"
            anomali_msg = (anomali_msg + " | " + uyari).strip(" | ")
            audit += f" | ⚠️ Günde {kac_yakit}. yakıt fişi."
            uyarilar.append(uyari)
            if kac_yakit >= 3:
                anomali = True

    # ── KİŞİSEL GİDER KURALLARI ───────────────────────────────────
    KISISEL_KEYWORDS = [
        "sigara","cigarette","tobacco","marlboro","winston","camel","philip morris",
        "çikolata","cikolata","chocolate","milka","toblerone","snickers","kitkat",
        "şeker şekerleme","sekerleme","candy","gummy","haribo",
        "alkol","alcohol","bira","beer","şarap","sarap","wine","rakı","raki",
        "viski","whisky","vodka","içki","icki",
        "kozmetik","parfüm","parfum","perfume","şampuan","sampuan","shampoo",
        "deodorant","losyon","kisisel bakim",
        "oyuncak","abonelik","netflix","spotify","red bull","monster",
    ]

    # SADECE fişin kalem listelerini tara — AI'nın serbest yorumlarını değil
    # audit_ozeti / anomali_aciklamasi taranmaz (false positive önleme)
    _kalemler_metin = " ".join([
        " ".join(str(k) for k in data_ai.get("kalemler", [])),
        " ".join(str(k) for k in data_ai.get("kisisel_giderler", [])),
    ]).lower()
    _kalemler_n = _kalemler_metin.replace("ı","i").replace("ş","s").replace("ğ","g").replace("ç","c").replace("ö","o").replace("ü","u")

    bulunan_kisisel = list(kisisel) if kisisel else []
    for kw in KISISEL_KEYWORDS:
        kw_n = kw.replace("ı","i").replace("ş","s").replace("ğ","g").replace("ç","c").replace("ö","o").replace("ü","u")
        # Sadece kalem listelerinde ara — serbest yorumlarda değil
        if kw in _kalemler_metin or kw_n in _kalemler_n:
            if kw not in " ".join(bulunan_kisisel).lower():
                bulunan_kisisel.append(kw)
    # AI'ın kisisel_giderler listesini de ekle
    for item in data_ai.get("kisisel_giderler",[]):
        if item and str(item) not in bulunan_kisisel:
            bulunan_kisisel.append(str(item))
    bulunan_kisisel = list(dict.fromkeys(bulunan_kisisel))

    if bulunan_kisisel:
        risk = max(risk, 65)
        anomali = True
        kisisel_str = ", ".join(bulunan_kisisel)
        uyari = f"🚫 KİŞİSEL GİDER: {kisisel_str}"
        anomali_msg = (anomali_msg + " | " + uyari).strip(" | ")
        audit += f" | 🚫 KİŞİSEL GİDER TESPİT EDİLDİ: {kisisel_str}"
        uyarilar.append(uyari)
        data_ai["kisisel_giderler"] = bulunan_kisisel

    # Güncellenmiş değerleri yaz
    data_ai["risk_skoru"]         = min(risk, 100)
    data_ai["anomali"]            = anomali
    data_ai["anomali_aciklamasi"] = anomali_msg
    data_ai["audit_ozeti"]        = audit
    data_ai["_uyarilar"]          = uyarilar  # UI'da göstermek için

    return data_ai


import re as _re
import html as _html

def clean_audit(text: str) -> str:
    """AI_Audit metnindeki HTML tag/CSS/JSON kalıntılarını kaldır."""
    if not text:
        return "Analiz tamamlandı."
    text = str(text)
    # Önce HTML-escaped tag'leri geri çevir (örn: &lt;div&gt; → <div>)
    text = text.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&").replace("&quot;", '"')
    # Tüm <div ...>...</div> bloklarını kaldır (multi-line dahil)
    text = _re.sub(r'<div[^>]*>.*?</div>', ' ', text, flags=_re.DOTALL | _re.IGNORECASE)
    # Tüm HTML tag'larını kaldır (<tag> veya </tag>)
    text = _re.sub(r'<[^>]+>', '', text, flags=_re.DOTALL)
    # CSS style bloklarını kaldır
    text = _re.sub(r'style\s*=\s*["\'][^"\']*["\']', '', text)
    # JSON/dict kalıntısı varsa temizle
    text = _re.sub(r'\{[^}]{0,300}\}', '', text)
    # "Proje: ... · Öncelik: ... · Ödeme: ..." satırlarını kaldır (div kalıntısı)
    text = _re.sub(r'Proje\s*:.*?Ödeme\s*:[^\n]*', '', text, flags=_re.IGNORECASE)
    # "ŞİRKET KREDİ KARTI — ..." veya "HARCIRAHTAN DÜŞÜLECEK ..." satırlarını kaldır
    text = _re.sub(r'(ŞİRKET KREDİ KARTI|HARCIRAHTAN DÜŞÜLECEK|Genel merkezden düşülecek|Personel şahsi)[^\n]*', '', text, flags=_re.IGNORECASE)
    # "Kaynak: WhatsApp" veya "Kaynak: Dashboard" kalıntıları
    text = _re.sub(r'Kaynak\s*:\s*\w+', '', text, flags=_re.IGNORECASE)
    # &nbsp; ve HTML entity temizle
    text = _re.sub(r'&nbsp;?', ' ', text)
    text = _re.sub(r'&[a-z]+;', ' ', text)
    # Birden fazla boşluk/newline temizle
    text = _re.sub(r'\s+', ' ', text).strip()
    # · (separator) kalıntılarını temizle
    text = _re.sub(r'[·•]+\s*', '', text).strip()
    # HTML escape YAPMA — bu metin zaten unsafe_allow_html=True ile render ediliyor
    return text if text else "Analiz tamamlandı."


def strip_html(text) -> str:
    """Herhangi bir metin alanından HTML tag'lerini ve entity'leri kaldır."""
    if not text:
        return ""
    text = str(text)
    text = _re.sub(r'<[^>]+>', '', text, flags=_re.DOTALL)
    text = _re.sub(r'&nbsp;?', ' ', text)
    text = _re.sub(r'&[a-z]+;', ' ', text)
    text = _re.sub(r'\s+', ' ', text).strip()
    return text


def odeme_label(raw: str) -> str:
    """Ödeme türü raw değerini okunabilir Türkçe etikete çevirir."""
    v = str(raw).lower().strip()
    if v in ("sirket_karti", "sirket karti", "kredi_karti", "kredi kartı", "kredi karti"):
        return "🏦 Şirket Kredi Kartı"
    elif v in ("harcirah", "harcırah", "harcirahtan dus", "harcırahtan düş",
               "harcırahtan düş (nakit / kişisel kart)", "nakit", "kisisel"):
        return "💵 Harcırah / Nakit"
    elif v:
        return raw  # bilinmeyen değerleri olduğu gibi göster
    return "—"


def detect_odeme_turu_from_whatsapp(message_text: str) -> str:
    """
    WhatsApp mesaj metninden ödeme türünü tespit eder.
    Kullanıcı 'harcırah', 'harcirah', 'HARCIRAH' vb. yazarsa → harcirah
    Kullanıcı 'şirket', 'sirket', 'ŞİRKET', 'SİRKET' vb. yazarsa → sirket_karti
    Hiçbiri eşleşmezse → None (mevcut davranış korunur)

    NOT: Bu fonksiyon Railway bot tarafında (bot.py) da kullanılmalıdır.
    Bot tarafında fiş gönderilirken mesaj metni parse edilip Odeme_Turu alanı
    buna göre set edilmelidir. Örnek bot.py entegrasyonu:

        # bot.py (Railway) — fiş kaydetme kısmında:
        msg_lower = user_message.lower().replace('ı','i').replace('ş','s')
        if 'harcirah' in msg_lower or 'harcırah' in msg_lower:
            expense['Odeme_Turu'] = 'harcirah'
        elif 'sirket' in msg_lower or 'şirket' in msg_lower:
            expense['Odeme_Turu'] = 'sirket_karti'
    """
    if not message_text:
        return None
    txt = str(message_text).lower().strip()
    # Türkçe karakter normalize
    txt_n = txt.replace("ı", "i").replace("ş", "s").replace("İ", "i").replace("Ş", "s")

    harcirah_kws = ["harcirah", "harcırah", "harcırahtan", "harcirahtan"]
    sirket_kws = ["sirket", "şirket", "şirket kartı", "sirket karti"]

    for kw in harcirah_kws:
        kw_n = kw.replace("ı", "i").replace("ş", "s")
        if kw in txt or kw_n in txt_n:
            return "harcirah"
    for kw in sirket_kws:
        kw_n = kw.replace("ı", "i").replace("ş", "s")
        if kw in txt or kw_n in txt_n:
            return "sirket_karti"
    return None


def detect_anomalies(df, model=None):
    """DataFrame üzerinde kural tabanlı anomali tespiti."""
    if df is None or df.empty or len(df) < 2:
        return []

    anomalies = []

    # Mükerrer fiş (aynı firma + tutar)
    if 'Firma' in df.columns and 'Tutar' in df.columns:
        dups = df[df.duplicated(subset=['Firma', 'Tutar'], keep=False)]
        if not dups.empty:
            anomalies.append({
                "type": "duplicate",
                "severity": "high",
                "message": f"⚠️ Mükerrer fiş: {dups['Firma'].iloc[0]} — ₺{dups['Tutar'].iloc[0]:,.0f} ({len(dups)} kayıt)",
                "count": len(dups)
            })

    # Hafta sonu harcamaları
    if 'Tarih' in df.columns:
        try:
            df_t = df.copy()
            df_t['_dt'] = pd.to_datetime(df_t['Tarih'], errors='coerce')
            weekend = df_t[df_t['_dt'].dt.dayofweek >= 5]
            if not weekend.empty:
                anomalies.append({
                    "type": "weekend",
                    "severity": "medium",
                    "message": f"📅 Hafta sonu harcaması: {len(weekend)} adet işlem",
                    "count": len(weekend)
                })
        except:
            pass

    # Kritik risk skoru (>70)
    if 'Risk_Skoru' in df.columns:
        high_risk = df[df['Risk_Skoru'] > 70]
        if not high_risk.empty:
            anomalies.append({
                "type": "high_risk",
                "severity": "high",
                "message": f"🔴 {len(high_risk)} adet kritik riskli işlem (risk > 70)",
                "count": len(high_risk)
            })

    # İstatistiksel aykırı tutarlar (ortalama + 2σ)
    if 'Tutar' in df.columns and len(df) > 3:
        mean_t = df['Tutar'].mean()
        std_t  = df['Tutar'].std()
        if std_t > 0:
            outliers = df[df['Tutar'] > mean_t + 2 * std_t]
            if not outliers.empty:
                anomalies.append({
                    "type": "outlier",
                    "severity": "medium",
                    "message": f"📈 Ortalamadan 2σ sapan {len(outliers)} anormal tutar (ort: ₺{mean_t:,.0f})",
                    "count": len(outliers)
                })

    # Sahte şüphesi durumundaki fişler
    if 'Durum' in df.columns:
        sahte = df[df['Durum'] == 'Sahte Şüphesi']
        if not sahte.empty:
            anomalies.append({
                "type": "sahte",
                "severity": "high",
                "message": f"🚨 {len(sahte)} adet sahte fiş şüphesi tespit edildi",
                "count": len(sahte)
            })

    return anomalies

    return anomalies

def generate_ai_insight(df, model, question=None):
    if df.empty:
        return "Henüz analiz edilecek veri yok."
    
    data_summary = {
        "toplam_islem": len(df),
        "toplam_tutar": df['Tutar'].sum() if 'Tutar' in df.columns else 0,
        "ortalama_tutar": df['Tutar'].mean() if 'Tutar' in df.columns else 0,
        "en_yuksek": df['Tutar'].max() if 'Tutar' in df.columns else 0,
        "projeler": df['Proje'].value_counts().to_dict() if 'Proje' in df.columns else {},
        "kategoriler": df['Kategori'].value_counts().to_dict() if 'Kategori' in df.columns else {},
        "durum_ozeti": df['Durum'].value_counts().to_dict() if 'Durum' in df.columns else {},
        "ortalama_risk": df['Risk_Skoru'].mean() if 'Risk_Skoru' in df.columns else 0
    }
    
    if question:
        prompt = f"""Sen Stinga Enerji'nin üst düzey finansal analistinin yapay zekasısın. 
        Şirketin harcama verileri: {json.dumps(data_summary, ensure_ascii=False)}
        Tüm veriler: {df.to_dict(orient='records')}
        
        Kullanıcı sorusu: {question}
        
        Türkçe, kısa ve profesyonel yanıt ver. Sayısal analizler ekle."""
    else:
        prompt = f"""Sen Stinga Enerji'nin yapay zeka finansal analistinin. 
        Veriler: {json.dumps(data_summary, ensure_ascii=False)}
        
        3-4 cümleyle en önemli finansal içgörüyü ver. Türkçe, net, analitik."""
    
    try:
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"AI şu an yanıt veremiyor: {str(e)}"

def predict_monthly_spend(df, model):
    """Gelecek ay harcama tahmini"""
    if df.empty:
        return None
    try:
        df_temp = df.copy()
        df_temp['dt'] = pd.to_datetime(df_temp['Tarih'], errors='coerce')
        monthly = df_temp.groupby(df_temp['dt'].dt.strftime('%Y-%m'))['Tutar'].sum()
        
        if len(monthly) < 2:
            return None
        
        vals = monthly.values.tolist()
        prompt = f"""Aylık harcama verileri: {vals}. 
        Bir sonraki ay için tahmini harcama tutarını TL olarak tek sayı olarak ver. Sadece sayı yaz."""
        response = model.generate_content(prompt)
        pred_text = re.sub(r'[^\d.,]', '', response.text.strip())
        pred_text = pred_text.replace(',', '.')
        return float(pred_text.split('.')[0]) if pred_text else None
    except:
        return None

# ─── GİRİŞ EKRANI ────────────────────────────────────────────
def get_logo_b64():
    # Embedded logo (PNG base64)
    _LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAABEAAAARACAYAAADu/yZPAAEAAElEQVR42uy9d5wkV3W3/9x7K3T3zGxSQGmVA5JWIhkRBCggMEECYUDghElCxgaMwQEM/r2AbcxrI7LBfgERjQCTEUgECQRYGSSUUc55tWmmu6vqht8fFbq6Z2Z3JSGkXZ1Hn9Wd6emu7unpqnvu957zPQpBEARBEIT7yMqVK8OKFSvI85wQAiEEoiii2+0yPT3NzMwMnU4H7/1GjxPHMc45vPc45+j3+6xdu5Z+v49zDq01ANZa5ubmuOWWW5S8+4IgCIIg3B8kiBAEQRCErZh99903JElCr9djZmaGJUuWMDU1RafTIYoidt99d5IkYWZmhqVLlzI9Pc3U1BQzMzP0ej2MMSxdupTly5cTRVFz3FrYqAWK+rYQQnO7Uqr5fjGcc83962PVj1FK4ZzDGLPo41evXs26devw3jMYDLjjjjsYDAY451i/fj133303w+GQ2dnZef+GwyFzc3PMzc2xdu1abr/9domLBEEQBGErRiZ6QRAEQdiC2G233cL09DRLly5lamqKAw44gB122IFddtmFVatWse2227LDDjuglNqocFCT53lzX601IQS8981jrbXzxAnvPUoplFIURYFSCq31Zj3f5uCca45ZU2eZ1OJILbYkSYK1thFnrLUYY1BKjd3eFlXqxzvnCCFgjGlee1EU3HXXXdx8883cfffdXHXVVdxxxx1cffXV3HjjjaxevZqbb75Z4idBEARBEAFEEARBEIT7y6pVq8L+++/PQQcdxOMe9zj23XffJvMiiqImo0JrTZqmRFHUiBF1pkRbvFiIWsCoxYQ4jufdp53dMZnBUQsIm8Omsj9qsaMtsGzsWAs9t/eeoiiI4xit9byMkVrkqF/7Yu9NW9SpH2etJU3TBZ8/z3PyPGfNmjWkacratWu5/PLLufzyy7n66qu5+OKL+dWvfiVxliAIgiCIACIIgiAIjxx23XXXsMcee7DXXnuxzz77sPfee7PzzjszMzPD8uXLSZKEqakper0eUGYx1AvyOjNjc8SBhUSCWsxoL+7bWGubx9XiQP19+/5t4aSNMWZBoWOx51tMAKnv287yqMWexcSUOntjod93sZ+3f1Yfv/27ba4Y45zDOdcIJACDwQCAbrfbiCRzc3MMh0PWrVvH3XffzS233MINN9zATTfdxOrVqznrrLO49dZbJR4TBEEQBBFABEEQBOHhzy677BJWrlzJypUr2WWXXdhzzz3ZZ5992Gmnndhrr71I03TBco568d8ID+UPoSUcqJZQMFkW4pzDaNM8xlmLiaJmDN6jjF5w4V/THC9UTxaaO5bH9POf98FkoUyMtkCy4P21hhAICjSqeR/rzJiFxJj6+3aZTPuYbRGo/dix7JjqtsKWZUCRGR0nEFAoVH08V/4t5v39jGEwGHD77bdz4403cu2113L1tddw0w1luc21117LddddJ/GaIAiCIIgAIgiCIAi/G3bdbvswPT3NNttvx8EHH8zTD3sGBx98MPuvWrXgbBqaSVaPRIUFZt72j3x5b9p9U/TkQyYTISYEk409V/sxY54YG8nUCN4T1EgwuC8lMPXL22ggoubfb/LXCdUvpapfyIdSoDDazHvf2iOAsxbN6LUv+vsu9p6ohd92Jv5Oi739ZrEHb8bnoLybB+dZfc89/PKiC/npT3/Keeedx4033cRgMODWW8W4VRAEQRBEABEEQRCE+8n+B6wKz3rWszjkiU/kqCOPoJvGdDodTBKPsiEWWEQvvM7Vk4rIZj2uvQhecMIOm/e72MqktPm3uZkctSKhVPPktZ9GOyNjoZKbdklJkiT3SSyZFExcGJXotDMsmucLvnmvglIYFI6AqUWPBd6rMNGtZv4fbTPCIzU6rN+YABI2EYVNfC4mHx+8Hf+bhUBRlUgBDPOM2dlZrrz6Kk4//XS+//3v8+sLLpL4ThAEQRBEABEEQRCEkh123TWs3HkXjjjiCPbZcy8OPvhg9th9d6ampsrFtjboOMLp9gK3LGsIBBwObx1pNPKA0AtMsvq+Trhh/iLbL3DshbIPNKNsidFtauGsEcZ9NRbL7HDBb7QE5YGyOa9h3muqWufWYlSY0BI8AVUddtIDRd2X0McHCGrxCEq1/rATmSL130yFB/j+qErk8eXfQSs9+g18mf3jnSM4h6nFJuu4/bbbuPPuu7jx5pv5zdVXcfbZZ3PhhRdy4403SuwnCIIgiAAiCIIgCFsrqw55YjjwwAM58MADOfigg1i1/wHssssuREo1JSbehqqsQo8taje4DBUZNBrNaAGtFhAfJm9fUPzY1II4zF9kh0VEjzD29ahMxLeeXwVPa8m8UZFhUkioBQql1NgPrbPl4luXQlFQ5UK/Fh48oXxOvWkj1EnPj/sjoDSvrypjaVrdEuZlvKgJ4WWTgkhg8eydTb1kNf6+cl8+BxUOT1BNHlH5+wUaMcc7hzZmdLwsL+8Tx2A0QQV87XfiLHfeeSe/ueIKLrroIq6//nouueQSbrvtNq675lqJCQVBEAQRQARBEARhS2KfffYJ+++/P8cccwzPfs7vs9MuK5vJzltflVIoQoAiK0jSeHwx6gDvywU+ARINE4vosRKKjbSc3SyhI2xiQWz8gtkFk4dxVXZK+7bag6L96idFFE9oxBMPqBDwCnQoH5+oymS1EjxGq3k1rpYsNt7Xt8X7sawN3epM0+7wUh+6KUdaxGfDVX9SWqJQ+/GK0IgK9XM2HWmAvHp8+3F1Zk97XPANVuCUHnv8vNc/GZCFTUdqbSPWRT87IWC9Q0XRooe0zhJVXXDuvvMuzjvvPH5w6mn84he/4JLLLpUYURAEQRABRBAEQRAeTrzmNa8Jxx57LEceeSRJkpBlGVNTU6Uo4Bx4hfe+bCdr5teUBFt1EVETs+JkJkZelJkNUTRPENnkDLvYcVnk9rHn9pMr9gXFj/qHvvWzQMCMHTYs4lcxXhoyZiKKw+PRaAymOY7DoVDNbe3b6yNS5cxsTCOpv9uYdqI3+va2hJ/gGnPUQNllB2MW9IVV1bOaiYyedhmRA4q6q8yE4LGoCBJab6wGh96oADP5+00GZkVRfu4AIm3G2vrWQkjd2afdzlgp1XzUw8QHLYRQvuYqi6SdFeMK24grOjJcffXVfOlLX+ILX/gC114rWSKCIAiCCCCCIAiC8KCz7777hqOOOorDDjusaTG77bbbNj4Qtenm2ALROrQ2EyvAgM0ylNEjz4TWjNhkWQSPzQvSNC0brYSJlWoAWzii2Cwuaihwan6Hj80lBpTzVWrGIsKKYrRkVyMBxOHJnSUxyYLaymJZJKVIEkrNh0BBRMGorWsgMMsss8NZrLUYYwghYK0lyzKGwyF5njeL8cIXTYnMQqO3rvxbKN3criNDGickJmImnSI2EbE2aK1J44Ru3CUlbt6CkQgTxoQcU/1brCxJAzaUJrF1kVNTakL1eVKGzbGNXUi4CYvKNxsJxlp/rFBlsPhFxJPgAsZU4lGYL8I471GxGpnEtstnqiyXPM+JomjMDLadaTMcDun1egCsW7eO2267jWuuuYazzjqLCy64gB//+McSSwqCIAgigAiCIAjC/WXPPfcMj3/849lvv/049NBDeexjH8uOO+7YLMg6nU6zUFuoi0fjK4EC60CbakXoIDJNl4/CFpg4GsuaqJfR7QWzav3c+4AOoLVCK8Z3/BeYZYNi7PiTwsPGBAkFJPix56/vG1rHs61HhErIKCjIsKxet4bCOYbDIYPBgLlBn8FgwCDPsNayes29OOfIi6K8T16KGEVRMMSxjoLV/Q0M5/p4BZHS2ODL9yCOCNYRdJlJYUPZstURSsEijrC+2HQwUnt2eI9zDq01URQRmwg/tERKgy9FFh0oF+yqzOzpJClpmtKJE7TWdDodpqamSJMEnOdRy7ch1RFpmtLr9Zjq9eh2u3SSlMRELJmeIUkSep0O3aRLh7RyeynFk6gaF5My1EaCKQOYoDcrClvsI+Qmju+qz18tfHjfSkgK4F0YderRE+rMJK0Wyj6MymoWypiZV3ZTCSrr16/n4osv5rzzzuOSSy7h17/+NRddJF1oBEEQBBFABEEQBGFB9thjj/DYxz6Wo48+mmc961msXLmyXOw5N5bRcX+oS0HKSU+NCQ2uKu3wjbAQWl1USgHF4YiqJbACbCgX4amOxlusBkqfDF/d0Pgz+MpQQ42tlj0BGzxB6UrIqMtTVCPAFEAfmCUnI8NW5SjDPGPd7AY2zM5y5z13s252A6tXr+aeNfeyes29rFm/jizLSjHEKHx7oa3GRz9xu59YlDvNRjM4NjWCx9/vVigKJgSEZk1fv81h/PvJAGfy55OPrzMiGs8TEzE9Pc2KFStYPjXDyhXbs83MUrbbZlu2Wb6cmU6PJI7LDBVlWKqnK6FEzSuNSYDYVm+EVvNVk1bWkWvlj9QClwEir+Z3oqlTPlpFRGg11r43UH4WtdILl1cFX6onUTSWXdQW6ULVTnhjwWLbdLYoCuI4Js9zfvjDH/KNb3yD8847j8suu0ziTUEQBEEEEEEQBOGRy1ve8pbw53/+56xcuRKlFEmSNAspay0hBKIoekDtVxf2QJi8z7gTBmMCyXhJxZiZqHcURUEv7cybVGuDVE9ojDhHC93R6/KMShw8sJ5Zbr39Nq66/lquv/56brn3bq66+1YGqjQG9apcrHsChbPk1tKbnqLwDh8CypSGlrbazScyWO/GRI2xNfCE6LFQx1fjH0ArV+WrY96/A3hU2YZY6bHX0P5as7AAUgsfk7/7pCDinSuzVVDl363KHorjmI6JKdbN0tVxk3VC4fDOYTzESuOzgumpKR61YltW7rwLe+22O3vsuhs77bgjK1hCD0XU+nx4fOOfotFE1W+gqs9iaIlwUYAkxK30oDAewelxwSOEUGkto2OOhDlfqSpmfhS4gCgWWMTkdRHq1sSmZVqbZRlaa5Ik4cYbb+Tkk0/mU5/6lHiJCIIgCCKACIIgCFsvhx9+eDjmmGN4whOewO67785OO+1EHMfji/JqJ7k2c9SbYzS6mQLI+MRXOzPUq71Qv4BWt5NW5xeblxkNShHqXfbKa8O1RI366/ZzZjgshlmG3DN3D7fcdhu33HEbt991J6vXraU/HHDv+nUMbM4gz+jnGTZ4lClLQIgMmc8JRo29Tx5QRoPRzPX7lVKhMXEpGLngm/ez3QJ3UuDwi92uRmJB9EAEkAeI15DrMgtFLyKAzGtHvInXOpkJ4m2ZZWTqehFXluF47yGUpTxNV5jq81H7zkRaY4zBe1+Kds43PhtaawyKZb0lJFHMVNphSXeK7ZYuZ9cddmLPnVey4/Jt2XPFrnTQdNBEjEpuavEhbKJNsHeu/Hn7fKk65igfUMqMRI/6o14UzedDtTJAvB5lSdVvo9mMYHEyU8taW35+J8/HUBrT5nnOb37zG375y1/y9a9/ndNOO03iUUEQBEEEEEEQBGHL5MlPfnI48sgjefKTn8yznvUskiRZUNCo/TzqhVF70WStpSgKut3uA3otzjnQowVkXfJQLwidtZj6eRuTDT+2+qtFjTpbo8BTECpzUMVa5rh7/Rruunc1d917D3euvoe7713Nhv4cv7npeqwqfz/rHFaV7WSDUnijGBY5GE0wGuKyY4kLo0V4ZNTIMLTyyAhVtgfV7noIAVdlfdQdYepFs7K+WfBPikELZXy0MyZ0ABMeOgEkVCayTi0esEwKHmrCRHSxDJD2Z6HJIpkQjAAKZwn1Z6d+T1tCnbW2EUS01uNtdEMgqNJRxPjS0DZkBdp6OkHTCZrIBlZ0p3nU0hXstM127LjNduyw7XZst3wbpqam2GXlSgyGhLjqxFOLN+UnskM83+y1VuIUBFWVeFWfKXzAKD36zE9Ehe1SrPuSAdIIRK3zzjlXmQiPl8kopYiiaGSUWxRcdNFF/PjHP+bss8/mV7/6FbfeeqvEqIIgCIIIIIIgCMLDj/333z88/elP50UvehHPeMYzmq4R7QWOMaYRQQaDAXFclhVMmiu223k+kPKXkaDhx2e9EPChVS6gdVlC4st2uJEemaLWIscsA7JQUOCZy4bcctcdXH7NVVxz601cccO1DIIlC45CA5HGKSiCJ3hPFEylqfjSY6TK4KgX951el8I7CmfJnC3LZqrMgkhrgrOjSoWJ9yNUx63fs9r8tb0Y11WGy8iTY2GxY6Nv4QMJNML9z+TRgAm+FCruxwsJCqxe/PcszUQNwbWEper9M1WWj1VlQUrzmazfc+sIztPrdEaimh/9HWozWIcCo4mrFrZG6TLTqHCl0ObLLBOjNKE6Lj4QVX9DV+QsW7qUXXfcmT133Y29Vu7Gro/aiUfNLGcZXaaI6aJJqlKbiNIklQBowxBdCXWOpGr6WwsmRZ6TJMm46y5UXYDUZgWKte8Hrc/ipG/PcDgkjuPm9vZjarGzFj9rY+Nf//rXfOELX+D73/8+V1xxhcSrgiAIgggggiAIwkPHH/zBH4Q3velNPO1pT2sWjbWAUQsexpixxU5RFOXisloI1QujerFeL4DagsgDFUBcPiy/1AptDKHqcuFa/zyQ4wCDA4Zk3HDnTdx0x2386KyfsnrDOlavuZe5bAixQScxTkOOJ8SGHE/mLIUOeF2WyoTyzSB2ZQaAUgplyt+z9vIICoZ5Xv7+pvznqctcAsF5tHdjglDjseBLX5C2T0q9AK3/Dt57Yn3/TWR9lf1yf01Qaxnj/maQaCDy/n4/vskgqUpovBofa1GpFpDGhZvKw8XoRjBTSmGqz6+hzBZxRdEIH015jdZESpefuTghd5Y8z5vFfVsMnMxQqj//SimMKi16tQflA9oHTOGJXDnGNrA07bHPyt144qrH8NgDVrH7Nrs0WSEeGOCISMosFEDh0AQSDBEKTSizXtqZM60sok1+RqrPXPscrsWkWohrZ4A0z1F5hBhj5pXDDYdDvPeNkApwySWX8KlPfYqPfOQjErsKgiAIIoAIgiAIDy6HHXZYeN3rXsfjH/94li9fzvLly5vyi8UWS5tsU7uRBZb3nqIoSNP0Ab5yPyZ42OrfHAVzFPzqqku44c7buOrG67l9zT3MFRlDWzA77DM7HOBTDXG5oHWEpjSlFjTK8pPx90CpuuuHorCjBW2diVAvuuvFcmhlpDQTdHWsuilp+30uF9ujEowoisoMgnqRbkzZNtaXi2t/P8UHqkX0AxFAdHhgQpZXflFxYzHRpvkdJspdFnp9bXGp/XmcfL/b/9r339jPAWxmx0QPVbXvtdaSOztm/lv/XqO/vQJboFqflViVn8WoEmFcUXmPFBacJzKG6elpli1ZyvLONI/fdT/22Wl39tt7H7aPtyECYmj8RpzNSU1UZaY0v3z9IdysaNF7T57nRFG0oPdH+z2tBdH271wUZavkOI7HskEWyhTJsozVq1dz1llncfrpp/OpT31KYllBEARBBBBBEAThgXHIIYeEpxz6VJ76tKfxtKc9jR23f1S5CCkK4igeLZTaizdfGng2VhrBY5Qud7V9QEcGQrnDrAKlF0Zr5z2E0h9jLP2+KWPRoHzZFlV5ApqyqGQ0Bjy+crsIaBweC6wJG7jjrru4+Y7buOmO27jhtlu48fZbuXP9GnxiyPDkOlCoQK4CTpeLP6d92YUE35Q51C1Tm0WiK39HpSqvjpY44VHoJMKFRbIM6mMtslgs7+THPSWq2+vnbAxjfRjfeQ+Uhqqx2agAoh/kz5Hy9/8ZggKn/bzMjY2V7oSJIEdvRvlM+/1tZ9HU/iB1JkP7Pu3WyXUGVJMBQYDKEDUmmieSAI0gYoMfE07qzi7l58gTVf1h6u4/KpQeMfjQlFRFUdRkBtWfAYCuNyzLI6KhxXjNsukpVu64C3vvuiu77bCSFUtmOGCvR5Oi6ZJW2SAacJXfiCKZfFMnsEVBlMRjn9taIARQWuO8K9+fybCz8v6pRbvJp6mvH5Mmq+3ry2Aw4KKLLuLUU0/ljB+fzjnnnCOxrSAIgiACiCAIgrBpnvTEQ8KfvOJPecELXsDOK3chKN10N9FU3gGTrU7qfpnOlfaJSYRXqmr96knQeOvKFqJaj4sm1dc+lOUcJorGDhtVAgPBl89nAiiN1x6LweIpUFg8OVBgGeJZ3V/NNdddz5U3X8uZl/6KdVmfQZ6VC35jKF9pwKmRUWagGtVEO9AFZsp5LVc38p76B+lvtbmygn+IP1MPRGBpd6mZzODY7EDnITBwHS8B+u29f5UMODZuDBPKNsalkW1ZTqO8rnxVqrEI7LjNNqzae38OfPR+7LrTSpbPTLNUL2EbuizDk7jqVXhfCjRqwpV2gb9HU8alRue0nri7HhOuqu5G1U89geA9iTajawwLd3YqbCnKauCeu+/m+9//Pl/76v9wyve/J3GuIAiCIAKIIAiCMOLFL/qD8OY3v5mnPf3pAGTDIWmnA4DDN10k6u4TzcKuFkKshbqVZiUe2OCxKqB8IKH0wKgXTYuVxUyuF62zWJvTTTtVZseohCXHk+OxaO4J67j8+mv43/PP5VeXXsxd69YQd1LiTkoeKwapZtblWGsbX5L2jvtkVoYgbE2MlVWFcWFFe4iVJuQW4wIdZVA+EAXFTLfHdskULz/8+Txh7/3ZccWOKMCSExMRoSl8TqqTplOMLpWP8fKWAErp0c8ngs8mK6bK/Gq34A0EolbeSGgJQZPXChUCcRSXJrW2LFPTkeGa667l/R/4AJ/8xH9KzCsIgiCIACIIgvBI48D9DwgveumLOfoFL2C//R/NdG+6TOCwBdpDFMXlSmOYQRyXWRv1iqkSOeqskPZipelUQoAQiFRl8ugdQSm0Kktjhi4HIDFJuXixOWlULqLKDhgeHcXgIUSeO7EUlCLM3dzLJZdfxq8vv5Rb7rid2XzI3evWkOtA7izD4FCxwcQxhbP08wzSCFul5NdlBwuVowjC1shkBsnkzyJ0KXqgymwL67HDDFwpVAZl6HU6THd6LOlNseM227Fqr315ysGPZ7d0RyICnapUJiJgvEcHqlKaVjpVYKxcrrmxziYJpVdPnWHiN/G7LBjIVm2tVQjoKBp7vnww4LY77uC0007js5//HOeee67EwIIgCCKACIIgCFsjRx9+VHje85/PIYc+hX0P2J/ppUvwQD8boJSik3SqDhHVQqVw5XftTI1qCzeokaFoe6fW44iazhKt3WZVehZY7zEmakSTUD0uVFNRHnISlTRPd/fau7nm7ls4765rufLOm7n2+uu4Z829uMpnpAieAk+UxASjsQQK7whGoavOKN77xmNgIcNK4IG32hWEh7H40c76CGp+WZGtBQMUHRNjKv8YozQh0gyMYt1gDm09M50eHQx+bkhiA0tMymP2eTR7bL8TB++9H4/ebU+2j5c3bXYNnk6IwIeRAFK/HufAO0iSRlil8c4Z+ZuYym9oXvZIbWJrLcqYBcQVypId5yFUWSJGY6rnu+yyyzj73HP43qmn8q2vfV0uAoIgCCKACIIgCFsyT3/608PvP+vZ/NUb30i32y0D/wVmgCzLmk4r/X6fKIpIqvsWzhKZaGw3dt4ubLVo0WriJ85VxpxFaRKZJs1zjpeywBw59xaz3HDX7Vx45WVceNkl3HLn7cwFi+8aslYnEK/Ah1CWsyQx62c3oKMys8OFUBpEtoSNSJtG/PBV61gRPoRHigCi/UjwqAWQ2vcGoNPpUBRF2YnF+bJcDYi0Bq3JFJi49NewwxzlPB0TkyqDsZ5UGfwwxw1zOjpix+2259H77sfBB65i3x1WcvD0znQpM00AfHBlxokp7V3zPG9K0xb/LZhfIzcvgavsiIRWmCgiEHCFJQ4KzKhErzZtdYANVYtpbfj2d7/NRz7yEf73xz+RC4MgCIIIIIIgCMKWwKpVq8K73/1ujj322DJ9HcYMCMtN1oC3rtz11RqMwVmLC544SSqvDU/hHbGOx5Yh8+r4q5Rz02596f1E5ohvXsdsPkAnEYGYtfS54qbr+Pqpp3DjPXdw+/p7ySOF60QMgsUp0EkMJjAcDrHWEscxcSXWZFlGnudMT083YkYtcLSzO9qeI86V1qztjh1SCiM8UgWQMtlr1KbXmFIsDK5s0+ucQ5lolD3lqm5OlMJirA2+sBjKDkh1xkmkNFEUMRUM/pbVPHn/gznyiCN43KrHsoQpFB6Fo0NMBE3GSHCWYF3ZuafurBMl84WP9ilbFJDEIw8iSoG19i0ygTILRFFel1RlsFoW6zUMsgG9tIsGzvrf/+Vf/uVfOPXUUyVOFgRBEAFEEARBeDjxnOc8J5xwwgkcddRRTE1NNQv7LMtQcUSuy+4KSVWeMlbqArg8xwNREuOVIncFVgUiHaNRzQKlvf4YX4+ExqbQFmW2RxSVae+z+ZBB5NBRh3V+jvMuuYifn38ON9x1O+vskLX5ANeJmM2HZHh0GhO0IncWgCiKcK4giiLSOCaEQJZlFEV5WzftMBwOywktjLI6GnGjKr+pBZk6+6Ne6NXfC8LWKoCoMOqcs5AAglbl1+0WvZTig0YRvCcfDFFKMT09TWwihsMheZ6jdSl0VCfXvDbC+EA37aFdIPaQoFmSdNljh505ZNVjeNze+7PP0l3oYeigiKFsu9s2F2qLqWEBAYRQ+YaAV4GiNmBGY9BEwWNC2V7YE3BVtlr7EN67UbeZwpYZIwH669Zy+bVX84n/+k9OOukkiZkFQRBEABEEQRAeKtHjZS97Gc985jPZeeed0Vrjfblrm7RKXepSk7qLpKmCfaMUZqJkpSnNJ1RrDIVqd2FYoBVlPfrqcbZqmOuBW+69lStuuo4Lr7uCq269kVvuuJ2BK3BGYWNNpjxzNi/b5ppygRJqjw5V7QCjUMGVmRs+jHVxqRdcURQ1XR9U6/eqjxPUaBFV+4EYY5r3a/HUe0HYOgSQ5hxfQABxwY8yP1rdkWoBQ1tPN03x3jMcDvGUZTNKKbIsa0yFqcQUo3RzLKcp21R7R+Qpy2ZcIM4c3aCZ0Qk7TC9jj+134rH77s/Be+3HykftyAxdYiDgMWUz7CYDzTAhgATAu1LYiKMm+8NXzkQdTJP5NmmuWtiCJIpKkahJHwkjrxJnITGgFXNzc5x55pl88Ytf5OSTT5b4WRAEQQQQQRAE4cFk1apV4YQTTuCEE04gjuOm3CNql58siB9fK7BA+Xx7cVHfsRI8vCpzPPTEpBGAovq3hj6rB7PcvH4151x6EWf9+pfctvpusljRj8GaciGmq93oBXpBjBZlClyreYQKfpOTld5IFYuSChfhES6CtK8E7ba4Xm36XKrPH72Z59ak4OIqQbJ97htflqYYDxGKWGnILMp5tl2xDQcfcCCPe8xj2WunlewarWCGpCyRwaLwdKm6RnlLrE3jW+K9R1ViZxEcsTKooKESP+oLnXMOdNmZqp0NNy/DREGWD0m7ZVtway1RVGbAfOc73+Hkk0/mW9/6lsTSgiAIIoAIgiAIvy2+8IUvhJe+9KWkaUoIgTzPUUqNZXosSqD05GivhpouLmWHlMiUhoExrcWA82DLRYJLIixgcRhM0wFmQME6P+Cr3/s2F151OVfddhPDWOF7CXmsyIJjgMOmZpSBUi18tB/tTrcXXH4iTb9eti0kmGz011ajhZgJIoIIj2zq8yEsEvW1zw/Npu9z355cj52PkwFoPhjS6XSItCG0Mr0ipUlsoDfwPPdph/Oco57NLp0diQCFpYMhLht4owko6+lWfiFjv6bX4BzeOXRk4L5mfNXJMFXGWBRFTdbLYDCg1+tx8sknc+KJJ/LLX/5S4mpBEAQRQARBEIT7wt577x3+8A//kJe97GUceOCBzM7OMj09Tb/fp9vtNunlm9XFJAB5Ado0xh8hz3Eaok46WiNU5SST5SABmPU5QSdY4ObBXfzswvM557KLuOGeO9jgc/JIsW7Yx+rSRyR3paFqFEUYY3BF3izCfKuEpv6+FkDau8Mb27WuhZJJMWRyZ7vecRYBRHikix/t82wx0UOFhYWO+nFj55xqHXviedrnJIzOv8lzuH6eTpKgAnjrKIoCVQkgwXtCbpnSKZEvrwu9KGHlo3bkSQc/jkOfcAh79nYmwtNBEwMRZVaIzwoSbTBxAmFC0imKUpCJotI7pIqGJ8v6amLAOzd2bWwbK9dfhxC48847+fa3v80nP/lJEUMEQRBEABEEQRAWY7fddgvPec5z+KM/+iMOPfTQMY+KhTI9BoMB3W538wQQ7yvFIYwF/KiyBj6KYxyBPDicCigiAqVhYJ+M8y77NVfefD2XXH0lN91zJ+uVZRDDhlCwLh+geymWQBzHGKUpsozgPLHSBOfpRKOFQ13aUvsPjMpcqiYNLcGivWgaEzcWETwmxRDJABFE/JgvXugJoUO3BInFzpOwgPixMeGjfV63z7+FMlBsXhonq8pHJNKGqPLlcM6RmITgPBqI0cQO4qCYUhE9FfHUxzyBx+y7P6v23IdHpSuIKzPVMtnNE1G22/XWjUTeumtUUaDjuMmKqzvI1CKIBqLKB6n2R6mFj6IoyLKMOI5JWx4pvV4PgAsuuIDvfOc7nHTSSdx6660SbwuCIIgAIgiCIBx22GHhjW98I4ceeig77LAD1lr6/T5LliwpBQPnmqC7HXzX2R95nm9GKcxIAAnO4RVN29oCT0HAYSiAAY4783u5+pYbOfdXF/DLSy9mg7L0cWTeYoPHRRofaUKk8VrhVKAoivL1URqXdpIErTV5nhOCG006YXzB1R4XW2hpv+nF2ELiiIgeggggmy+ATJ6L9eOdKs1MF7tPG71AmYvxakwAmRRQ0KUpce3j0XSjqVtaq9L02KDAB0JhS2NWE7Mk7aJzh8kdU0Tss3J3Dv29Q3jM/geyw8y2dEnwOFJiNKXHUXtsmzvXosfIRHVcAKmvt8AmRemiKIjjUQvxH//4x3z4wx/mlFNOkbhbEARBBBBBEIRHHv/3//7f8IY3vIFer9e0YtWtdo91O0mty44Kde15zXA4JIqizTBALTsp1AuO8a4tMMTj0Nw0uIOfnnc2Pz3/bG665058GhGMYl02oIgUrurSAhC0Khcy1Y6t9WUXiW7lUTLsD0bdWbyHWDWLo1oAae82t1PsF0qtN2HzFnkbkX8E4REtgtQCSFug2ByBMChwOrQbQC34+LHjTggiuqW+tFvxNt9X1xHvfSP4GqWJ45goirDWUzhLCGWWWWwidCgzR0Jhiar2ulNRQk/HuEGG72es3GlnHnfAKl5+9HHEaFIMCZqYQBQCHRURo8e7yqjxDBAqAaTdYru+PjtXdp7Ksozp6enm91BKNfcdDoeNX5NzjiiK+NrXvsbb3vY2rr32WonBBUEQRAARBEHYevnLv/zL8KY3vYkVK1YwMzNDmqbzdgrroFpr3dSct/0+8jxvWsHW999UG9eAxxEY+IKgDQ7DGvpccsNVXHLdVZz5y3OZ8wVrsj7r8wGFBtNJwOhGeKmfv+48E8JoUdDpdBgMBk2LWaUUURJjjMFaiw92o6JEeIAzkWR6CMLGRYyFzpXFzrt2OUtZlhYWDBjDIsHkvNsnnqgtyDRCidZEVftcfJmpVre+1hjiOEZpTR4cuS3KlrdRWSpjlCYfDku/kE6XVEcMZ+cIuWVpd4qkUCyfmuHgfR7NkU8+lCetfAwp0AVcMWRJ3BlXPNTEL1R1jXHOzcv8WMiLKYRAlmV0Op2x27Msw3tPt9slhMANN9zAN77xDf7mb/5GYnFBEAQRQARBELYODj300HDCCSfwzGcdxY477EhhC+KoFDycd42hqdYatcAl2TqLUgqjTfO91hqtNM675vbFcECfHNCsZ8il117J+RdfzKXX/4Zb7l3NmnwWm8YMlcVHESrWWBSFz/FookgTco/REKkI8ARHVTRjUCpgrScER5J0iGPDMLc4X4AyeG+JWgJNaC2yFvL3WEjQ2NTPRQgRhPsvjkwalLZLVDSgfViwM4zfzOOqoMa+n3c+K9USO1RZ6gLoSkwNuS09kVBYHF4bVKLL7215nUo7Mb7wFNmQjkmJI02wAXJLL0pQw4LIaaaMYXk0xb477swRTzyUpz3uCSzVU8R4TNDlbxU0hOq3V+AZtdYNtDLz6ta7waOVprBFWcqjRu9WfZtSauz6nhc5SZw0X//gBz/g85//PF//n69JXC4IgiACiCAIwpbFHrvtHo488kj+5i1vYb/99wejH1AZxn1Z13uqbi7KEAisY8gvb/oNZ150Luecdz73blhL3OkQEkMWPAWOoDVeBZxSBFXv+46PpUHh+O16fJXTjH7e4xdfJP1WJzQRQAThAQsgv43zapFLwybP+YWeS7cO6lUgqPK6Uo8LXafqUbfMlbXXmOCJvCZxnsRpUuc57JCn8bj9Hs2jd9ubnbbZhp5KUXj80KLThKD0g15GV1hHHBlW37uGj3/843zpS1/i6isulxhdEARBBBBBEISHL4953GPDO97+Dxx77LFjZS0PdF2+OcG382XpTJ1d0h/0Of/88znnkl8yXBKx3mdlaUqa4FRgkGUEBSaJsc6NCRIiJAiCsDVQltuoVjnPqDNN5CFyEFlPR0WkQbHnTis59PFPZNdH7QTWgTJgIn4Xl0TnAoPBgOnpsovMz372C97z7ndzxhk/llhdEATht4hcVAVBEB4gb/qbt4TXvva17L/f/i3RwqPR5HlGL055UCPoEJq2jngPut2ywROUo2wIWbd5DDg8GtPctrGJQfQQQRC2RBxg0bhFAuB2RxhPgbaeXpSOjFEDsGAB0G/7hQYwasHv777zDk4++WT+/cQTueU2aacrCILwQJELqSAIwv3giGcdFV7xildw3HHHAZAmKZnN0VoT66jx94iURvmWkqAehLGwEEej760brz5RCpQvxRECRKaZAXxRoON4U9UqD87rllFGGWV8kMfQEkB8W9eoKCiIiYGAAiIUGsizDKMioiqr7sF6fa6fYXqlSB6sQ/kAaXk9d3kBOmCShCLPOeOMM/jUSZ/ma18TrxBBEAQRQARBEH4H/Mkr/ywc/5rX8NSnPpVIG+bm5pjqTZU/dK7KvhhdWu1wSNTtPOiBfiDgCktQlGUwLUEjH8yV3QuMKRcEVeeCciEQUFV2iF5E86gXDlpGGWWUcQsaFYDXo2ulqkWR8bbg9bUQQNcmrJXqEaEeXAG7FkIKj47KZ/cBgvOYqHoV1qLqlucKrr/+ej7xiU/wta99jeuvv15ieUEQBBFABEEQfru89/3/Fk44/nUsXbKUPBuSxsmY27/PcnSSlMGsdaUYkiRNnvVvw/BzsdjZFgVRHI/d7oIHXwoddUVMoOxYYEPZxUBvJLVbt4QPQRCELREDYFsCCC1lpBJCsqozVy32Fq4AH5ouLQ92AUyWFaRpPO/2EMA7T2R0I4DkWVa2DE5i8jwnSRK+8N9f5L3vfS9XXn6FxPSCIAibGU8LgiAIC3DkkUeGf/nX93LAqlX0er2mPWN94fTWkec5nU6nuqJWEbW1UO/WUflz/DautpvcQbR4ArGJQKvmdu891pdJ4LVJanm4gLV27LbFJgfxAREEYYsLctspHoyEjwUvdAtc5FxuMWn0O4mWQwg45yrRWqOU2vjFt/WaNmzYwLXXX8eHPvQhPveZz0psLwiCsLG5Qd4CQRCEcd74xjeG17zmNeyzzz70ej0KZ4lMmQid56XPhzFmPECtvqzv6wHrbLlbpx7YHuKmLtTW2vHXQyl6KKWa2yZjaB88GjX2mI12rhUvARlllHFLGxe6+C2k7rZrAG1LtP4tZvAtRl7kGGMw2swXRbzHW0cUx3jn0Ka8j3cOHRmsLecYrTUBWL9hPatXr+Yb3/gGJ554InfcdrvE+YIgCCKACIIgzOfAAw8MxxxzDH/3d3/H8uXLGQ6HpGlaCgS+1WVl4upZx86eUBntKRxlpxVQGCCmSsW+nzzQFOxFL/Rhge9FAJFRRhm3llHXna8Wvx42XiGTgkltoqof/FLAwhalabaJJi7JLQ+S+rZQvsC2eF0UBXEcE0LAekdkIlbfu5rvfve7fPTDH+HCCy+UeF8QBEEEEEEQBNhll13Cf/zHf/D85z8fY+bLFFmWkSbpRq+i1juUNngCWbAEpZsWs9Zn9HTykAog7e3L+isVFhFEZFYQBGErIShwatNJIDqMbp+8NnodHlQBpCiKxm/EOkue5/S6vXnXfu8cIYQm46P5HUMgy7KmFHOQDemk5ddZnpFEMZdccgmve93rOO+88+QKLwjCIx65EAqC8IjkiCOOCP/0T//EoYce2uyeWWuJoogsy3DO0ev15gsEUGaEtL9XqnSsa6VN51lGHMcorR/y3zWo0Q6m3tjFXwQQQRC2NsLGo99N+RupzbnTA36NYTSXVDhrKZwFrVDGkExmh3iP957IRATvUaa8ultrcc6RpOnYrzo7O0tRFLznPe/hQx/6kFzpBUF4xCIXQEEQHlEcf/zx4U1vehN77733yLyU0ttDKUUcjxuCeu8bEaMJhH2dG10FooMBqtttP6gZnS0wvc5D+jt7vYAA0groVRivcQ8yMwiCsDUEuaG61oWFo96gxgXits5R24JEYZGMud/ai1RNdoepzLOLPC/nIl2WVLZfvnUWFUqza6UUNi8wxpSlMVo12SHOObIso9frjZmrDgYD7rjjDk4++WTe8Y53yNVeEAQRQARBELZG3vCGN4S3vOUt7LHHHk2mRx0kWmtJq92yurWgcw7nXGl4GkXjF80w8a8SO0KeoyIDSTx2hQ1jFej3naB+e1kkdSC/WEBfLwjaC4CN3V8QBOHhHej6eQJIaF3v6iYxkyJInRAXAeZBvP7VXh7e+7GuYtZaclug4oSgy/voABqFUQpTzQvZYEhaPcZVhthNJokaGXdH0XgGSZZlpGnK+973Pk466SSuvvpqWRMIgiACiCAIwpbOBz/4wfCmN72pdMkPoay3Tsp669nZWaanp4GyDlsp1XRT8d43O2khhPFuKW3BgNLw1KPKDrhAgaNsSFveFj1AFw//AB8/WfaiF5gIGh1nYjGw0P0FQRC2DPyCge5kpsek8DF5/fxdXAMdrjI9HQkVpYm2bjykqqY05RgWaGNelWKOZZQoGA6HjbCybt06kiSh2+2SV5kmSim+9KUv8Xd/93fceuutsjYQBEEEEEEQhC2Nb37zm+GFL3wh/X6fNE3Hdr9qzw+Afr9Pr9djMBjQ7Xbx3peBY2WIun7tOpYsWTK2o+YASyDHkxOwaHIUF9xyCV8/9RSuuvE6Zrbbhmiqw7333kvHmAd0sQ1o/AM4gA5+XACZiPDrEhhfVfV4NV4Go8L8xwiCIGwREogClB+7htUZbQtd19SEUOLVgyt/FEXRiBPee6y1ZFnGzjvvzOGPeSIvecIRLCcmbvrBePCBpGqbG5zH5jkAcZrO61hW2HK+c84tbvRdZUBmWcYll1zCa17zGi6++GJZIwiCIAKIIAjCw5lVq1aFd73rXTzziCNZtnw5w8GATrdLsBYVReRZ1hjDzc6Nsj+89yilUKgmWKwzQqIoKj0/dNlTMcdToBlQMEdgDstXz/guv/zNJVxx840UUUD1uvTtkDwE4thg/P1PofZjKduaUWHK5o+bEj7aAkj9XJMCiCAIwpYZ6fp51736etgOgvUC5YFliYx+UH2RnHN0Oh2GwyHWWmZmZnDOMRgMWK5T9omX8dQDHsNhT346++2wBwkGTUGEJkFhwvj1PliPMgqUwfmR90cttmgUJo4oioKo2ggYFjlpHFPOgnDvPau5+eabeec738kp3/+erBUEQRABRBAE4eHE4YcfHk444QRe/rKXl0HecFgZyBlwvrzUhQBGN11aCKHx+Jjs1JJnGVESo7XGEwjBUWBBJQzw3Ok3cMPaO/jJL8/nWz89jSKOyCKPVRqnPU5pgioDUoXH+N9SCnXQZTB/X8fFJoAJAWRMeGkvDOQjJgjC1hgEb6a4+1AZQ6sAHR2RrZ9leTLNUw56DM950mE8YZd9WUGPLp4EhfYeMKAqASTWWOurzMfqCt7yq3JV15igdWOyWpfXmPp9qSaBs/735/zze9/LqT84TdYMgiCIACIIgvBQsnK3XcOXv/xlHv/4xxNFESqAt64UP1RVz2EtVOUvvur0oqJo1LKWynE/ScjyjKAgTVI8MMyGxGlCAWQUzFJw6c1X87UffZ9zrriYQaJQUx3msFgNTo1fXHW4b0G2IAiCINQEBVlwdNMOPQx+3RzRhoz9tt+FY494Ns970uHMEBMT0M6TmmRMsHbOonzU+Fg1ht7VXOWCh4kSHwNjRih1i90br7+BV7/61Zzx05/I2kEQBBFABEEQfpc8+oD9w1e/9jVWrlzJzPRMc3vb7NMVZatAncQUoSxxierWtQGwrswK0QoiQ1BQUJuausZ2bojlXgq+esa3OP3nZ3LHunsZKI/rRNBLy5IYHcbazNatF2sBxJrxjApBEARB2Cy8JzhPjKYbJyReoQtHaqGjDM947BN51Yv/iG2YxpMxQ4rGoX0g1RH4kegRFPjg8d4TUWU/esa75FRivm0JItYWpFFMlmfEJuKP//iP+cpXviJrCEEQRAARBEF4MHne0c8Pb3/723nSk55EZCKG2ZBO2mEwHGCMIYoidLWbFaqIzqMoCChU2c6QVnpvK+83szmZDjhjCERY4Irbr+L7Z/6Ic6+9lLsHG7AETCehUIFCBTI8gyIjqrrKmEr4MJX4YUIZSOZmPDtEEARBEDaFBro6xjuH9x4fQjl6T2oielGCmx2wLOrwe/ut4pjDj2L/nXZjhZoiRZMPZlnWXQJeYb0r5zBj0EqPMj0mu7SbkQASgMJmpFGMRhNCaRLunOOKyy7jE5/4BJ/6f5+U2U0QBBFABEEQfps86wXPD3/zN3/DEx//BJb0pjBK4wqL0Qa8B2Oaq1nuCoLRKGWa2K7Oz/A4DJCgy2yQ3BGKoiyJSSMy4F4yLrz9Wr75kx/y88suJIugUA6nQWmNDR4XPDoyoDXWVz4iYWEBxANZJAKIIAiCcN8wASIPvrBYAqoTo5MYrxXWWpy1bL90Oetuu4upYFimEvbfYVdeeOSzedIBj2U5PeLg6apoLAtREcqymEBjkLoYDk0A1q5fy7Ily8YWED54rrnqKt785jfzw9N+ILOcIAgigAiCIDwQ9txzz/CNb32TfQ86EBc8HRWX6bhZjlEKbeLS48NXpSxJ1LSpLfBVgofGkhMRNRe92uxNMdoA20DOTy48ly//4LtccvsNsGwKO5WwPh9gEkPhbNMeN2iFc+W2mTG10epIAKnLX+rjSwmMIAiCcF/RQIpBKYU3iiw4Bq4gDw60RmuNHWZs05shtWDXzrJ9Oo1dP8eyuMufveTlHHXIYUwTU+ZweGIUHTajNXuoBRBA6aZKJisytNbEJiYrMozSxFHM2tWreclLXsJPf/pTWVsIgiACiCAIwn1h7733DieffDK/93u/15i0DfIhaZygQiBSpvTvsK40OA1ltOaLAh1HoMtuL56AtZYkMs2xPYGcgEPTJzCH4/Onf50fn38WN6+9m86yJWTekhV52Q43eIoo4DSNkZzygTzPAeh2u+SurJbWi5idigmqIAiCcH9IoojcWgpncQpMHGGSGLQihIDPCnCejjJEDjpoTFCEYY4vLEuSKZ79jCN40ZG/z+5mO6aAyOZ0vQYTjastlU1W3VQXIBvmdDodApAXOVGcYJ3Fe08Sl+WfeZHTjROKomDDmrUcf/zxfOM735Y1hiAIIoAIgiBsjMc85jHh/e9/P4cccghLliwhz3OSJMFbhzamMS313jdpu9ZabF7Q6XYrhcPjC4uO49YlzkNwFCrgTUQfuGr2Vk75xU/40YXncMdgPWEqpW9zhq6gm3bQLuCyUgSZ8xlxmjS113Fctsh1hcVTiiK+eqpa62i3TYy8iCCCIAjCfSe3BSaJiaIIH0pRP4RyQlFKoYAoisj6A7ppB6xjMBiwdGYJOoC3AZ1ZtkumOGTfA3nBUw7nybseRA9QeV6KGLUflhq3BVGACbqce8snxIdy/i39xEt/rVIoGdLpdCDAYHaWiy++mH//93/n69/+lqw1BEEQAUQQBKHNwQcfHP7iL/6CV77ylaRpirVlRkVUZ3d43wrQAkGVnhoe1dwcgiNRVVqvDeBC2d5PAcEREkMf+PVd1/Hd837Gj399Hrf21xKmUorqmGMXxzHBYuECFt36SVDg6yYz9deqPE7kwUgNjCAIgnAfqOeSZs4Jo7lp0rnDMxLe67lIAVEwJGiM9YRBzjZxj9/be3+e/9TDefqej6dLoINCh1AKHUrhlSr9s4IjJQZbTWBGNyuH4MvblFnAQ6Q1f/70zJ/yD//wD5x99tmy5hAEQQQQQRAe2ey4447hH//xHznhhBPKoK3KsDDGlF4bIWCHGXGSlD4bWoFRYz4foJodKAO4IieNkzIa9AEiRR84+9qL+dZPfsj511zOHcUcRTfC9RJ8GpHbgqBA+4UDy1rmaIsiiwWfdeBJKxiVDBBBEATh/oog9fyjWXwuaTYGWkK8CkDuSUxUbij4gHKe1CuW6ZQVJLzj9X/F7su3Z1s1TQIoPKaa4RQBHUy5oWAtKFWKINCUmi60qggLLDROPfVU3vzmN3PVVVfJ2kMQBBFABEF45PHZz342vOIVryDPc9I0BcA5V5qKTgZ2waFUebmyRVGKCnEMQJZlREnMoMiJk5SAosABBovjytU38/+d9GHuzGcZDAaoyBDFMVYFht6S4XFGYfX4hbHu4gJQ6NFuGox24SZFEBXGS18k6UMQBEG4v9SCx6ToUc8zC8037dtUKLue4Ty58hR6lFESrCPOPenQ8Xt77MexT3smRx30ZLahg86HpF4RJV36/QG9qanWE5VZIgDFYEDc7Y4mx5aHSPNavQMfiKIIay0//OEPef7zny/rD0EQRAARBOGRwfve977wyle+kmXLlpGmKc65xldjJHiERhApnIXUoNEYFCp4jNKjLaYq4CqADBhgcUT86pZL+eapp3D2by5hbkXCrHJoH9AognVNFxdiA5HBtdSMWvzQvgwmrR61sa3FkTq41K1Ac6GduVCVwogYIgiCINxXASRyo45iXm3evDImzluPMppgNFaDVQGvFVEUkWiD6udEg4KlzrBqh904+olP55mPfxLbmmlsPiROpslDQcgtvbSLAXxeoE3UCCHtFUX9uupNA91abNRG4nfffTdnnHEGL3/5y2UdIgiCCCCCIGydvPzlLw8f/OAH2WGHHZibm2NqaophZZrWNjXN8xytdZmuy6jUZeAzYqXpqJgIGA4GhBDo9npkVTeXAYpf3XYlXzjlG1xw5aUwlaKmO6zNZnEKQqQxUYQyugkkDQqbF5iwcBtbVwsgEwIJbFwEqe/jFXgtAoggCILwwAWQej5ZzLeqnTXiNQzx6MhUGwcBX1hCVVJqTNViNy/oes1UMKRDx65Lt+WYw5/FsU95LoYIA9W/QATEKEwAOxgSdTrjq4qqRLWeJr0t0CiUUk2Gp6/8Q7Is4x//8R858cQTZT0iCIIIIIIgbB085SlPCZ/73OfYddddSdO0ETvqMcsykiRpSlyg3CWq28zqyJSCRRXYueDKQI7SpC2nzPw488pz+dL3v80lN12LWjHNIIJ7szlMmhCr8jjBaFzwDF2Bc64UWrRpTErNRBaHDhsXMNq12e2AdaF0ZRFABEEQhPsqgLTnkya7YoHofVKMr+8f0rhsoVtYTIBYlfMeWpVWWZTlKaGwFLMD0qBYmvaIbcCum+Pvjv8rDt5rP3bqbI8hgM9YojulX4gr29S3u8hMCiBN1iaMGZzXGyAA69ev5/Wvfz1f+tKXZF0iCIIIIIIgbJmsXLkynH322eywww4458p2tt5TFAVpms4TPvI8x1pLmqbjPiAByDKIY9Bgg8NqwwC4K6znqttv5gOf/Dh3zK5lEAFTKZkOuFjjtWKQDenEEUVR4AuLMYZOkpJU9chZljXZJpMdXWqvj2RiB66dfjzp0j8WvIb5AokgCIIgbC7t+WhS8JgUS2p8+zZfp4YolDForbHekec5RVG2fQ8hkMYxSZLgnKM/HBJCYJlJ6dyTc8COu/HMpx/G7z/pcJbSIcZj8pzppDOqd2EkgLRFf02ZsdmtWtVba0vBpSqHqZmbmyPLMo488kh+/etfy/pEEAQRQARB2HI4+eSTw4te9KKmrCWO4ybogTLttc4GaXayVCt8C36UzlsZrgVbYFNDjmYNGWf8+hxOPetMfn39VbheQpFo+sFSeEeINEoprPdlei+OOIowqDIYtK4KzFS1CxbmdXGpHfRNgNiVY1sAcWpc/Gh3fmkHppOeIfXPZZRRRhlllHFTY9t/anOF9XZ2YuRBKUVQUASP9Y5Qe4CYCJsXpShiLVmRQ2RIprplR5m5ISvyiKlcE3nYbdsdeNZTns7RTzuSbZkixpOiG8PwhVYYAShsQRzFzc1FURDHcVMKE0IoO78Bw2zIr371K97yN3/DeWdJ61xBEEQAEQThYcyb3/zm8P73vx9jDP1+n16vR1EUKB2V9b+6rRS0xATA4fGl9ECMJkajQpnxkWtFhqdPwblXX8QpPz+D86+8lPU+J1k+Qz9YrB4P+mofD6dLD49aqJj07xhz0W993e76YvzCweVCmR9hIxfWuqRGRhlllFFGGTdnZCPzyWK057t2OcxC81b7OG1Bv5kbvSJCkQZNnDnSzLPXtjtyzNOO5HlPOpzldIhwxAESZcYzQrwjxDF5cETKVBsPUOQ5SRQ3JqrWWUwU4SnnfKUMq9ffyynf+S7v/T/v5obrrpf1iiAIIoAIgvDw4Zhjjgmf/exnWbFiBbOzs3Q6nablXRRFrVgo4G3Z6lab0bZW3TbPtUSEIhRYa9FxygY8P7v2PE758Q+55OormdUOM9PDpoa+L7CEsYCv9vIwfuTh4eRqJwiCIAibjVfgTVlCGqGYNkmZEdnP6XnNMp3y16/5cx63+6PZhi4RjqhwpDqiKcoxejS3+9LDSynV2gwpPb+STlpWvOYZcZICkBc5nTjhr970Jj760Y/KLC4IgggggiA89Fx11VVhn332AWi6uwCNyWgIAe8hilolLlVZC9ZDYSFKStnDBGywrA85Ku6RAb+ZvZl/+o/3s95m9G1OpgM+MeQGZl1O3+aYNCnFD0aZH7oSQCZN4QRBEARB2Dxik2Ctxfmyna5SCm8dobBo65lWMftsvzN/cOTv89yDnsFSDF0gASgcmHj+akOVXiCDwYCZ6RkIVUwQKEtUrYOknNe9L9BJzIYNG/izP/szvvnNb8raRRAEEUAEQfjdsuOOO4YPfvCDvOxlL6sCFN8YmtW1xFrrpttLCAqtVZnt2qSDMGGe5glYBrFiiOFXd/yGL//wFH75m8tYEwb4pOziMgyOobf4SBN1U+Juh7lBvxE5NKXwodqZIEEEEEEQBEG4L+gAPq9MzGNDjif3jmBKD5FYGYr1c6VZah5YtdPuvOSo5/LkvQ9mhpgumtQBNkDwkMS4IicoRZSUniBFURBpUy5IAjTbFtYRihw11SG3BVEUobXm3HPP5Q1veAMXXHCBrGEEQRABRBCEB593vvOd4R/+4R/odrtYaymKgiRJxju3tLC2qu11Dm8dcZxU4oeDoih3gnSgiDW5ibli3Y189Uff45wrLmZNMWCdHZIumylFD+/xWkFcmqVlRc6wyJtWenpC8FioJa0gCIIgCJu5UAiAVhTBU+AJWpVlLc7h8oJe2iH0M2Z0QpQ7Ok7x1Mf9Hi949nN53PJ92ImYyDLWJrcmyzMA0jRt4oXIRPNew2BYdpEpilIIcc7xyU9+khNPPJFrr71W1jKCIIgAIgjCb59nPvOZ4aSTTmLXXXcly8aDlknBY3Z2ll6vR5IkZbvbNIUQULXhR/BlymukyIEBcEtxL9/42Q845edncEd/HTY1WKPQnYRhnqGMKV3iVZlx4hWY6raiKMac8SczPuq2tYIgCIIgbB4eMHGEDb4pg4miqMnGCN4TrCM2ESEriJUmUYb+7Cy9pMPjdt2bvz36T9h7+U5MRV08Do1CAzHl4+tuce1Wut6XpuhGm9IwNUnw1X2Loih9xLTmrrvu4m1vexuf+cxnZIYXBEEEEEEQfnv8/Oc/D0972tMAmu4u7Va2dVCiKkf3ut1t3equDKPKAMc5S9AKqw0DPGsp+ORpX+Znl/2Kq++5DTXdRXUiLIEQAtY6OnGCKzzOOXC+KbXRqMpMbaR2TLYIrLvAiAmqIAiCIGw+QXmGrsDEho6JiYyBovT/wHqMUs08n3Q6DLIh6wdzdLpdkqkuYf2Q7YYRz3z8U/nDF72YHcw2TGMIDFlKhxgwPuCyvMzgTNN5mSLOuWajI4TQiCHD4ZBer0cIgZtuuomjjjqKa665RmZ6QRBEABEE4f7znve8J/zt3/5tJWKA1qUBmrW2ETrqTi/1WAdDw+GQJEnK3Z3gm24vOTAE7mKWU8/5Cd/6+enc2l/LwARmlcMZhY6q9FrnSEyEsgqDagQPrVTpOeIDwbV2kBagFkAkA0QQBEEQNh8PRIkhLzJcYZsMDxMgCgpjDJExDIZDhq4g7nZQaUy/yMiLgl6Uskx1WHvzXTx65R4csv9BvPqYl7GUmGWkmGDpEY38wBRleWxkQIELHqNKM3UoM0zjOG42YurR+3Jj5LTTTuN5z3uezPaCIIgAIgjCfePFL35xeM973sMBBxxQtqer3NjvfxTlsHgybdiA4/Qrz+N/fvx9Lrv1WkI3IVMep8qWe2W5yqitLYBxZuyCpVslLroVqEH5+PqxQS50giAIgnC/0AEUvsmsnDe1q8ooVS3++OBgSToN/Qy7ZgP7brczxzzjmTz7SU9jp94K0nqybk/Y1YaJr+b4zZnDazGkKAre8Y538O///u8y9QuCsCBycRAEoWHvvfcOH/7wh3ne857XBBMPmNrzQ8PVd93K1370fc65+lLsVALdiLX9WaIkbtrXllem0BI0FAQ974I1GZD5OmBS454fKkDkWTSAEwRBEARhgUVCgCiUAkh7k8Kp8c2GelNCt9vQB/BoigCDwYClSZftlyyjuHcDxbpZnnjAwTzviKM4YI99MBMhQ7tBnNnEYiXLMoqiYHp6GoDZ2Vmmp6e5+OKLOf744znvvPNkrSMIgggggiDM561vfWt4//vfD8D69etZsmQJg8EAY8wDywAJwCCHKIJY4xRkeDIcBR6FbkzRNH7ebk8ALLoRRBZ7ClpBUx1AAZXZ2ihTRBAEQRCEzVsk1D1ZXDWvOsBOiBT1/UxrNM38HeGqOTjgy4wSFDGmmevVxFxef62A5D4uVobDIcaYpnT3ggsu4IlPfKKsdwRBGLu2CYLwCObpT396+MIXvtB0d6lbyta7KA+YOqIJAVe5t6M1AYea9O5oOsS0lAwFaL1RBWNSAPGt200rgBMEQRAE4T7gGStLcdW/yRKVWvRQ7UmYav5uHlt2g6vvkvmcRCcLLkzq426qBKb2H1vsZyEE1qxZw7//+7/z/ve/X9Y9giCIACIIj2S+/OUvh+c+97ksWbJkQa+Pubm50t39AWSAWGCuyvRIAO0dSR3ShICzBSZNCK2ISU3mwFYB1GJXsDARq00i2R+CIAiCcN9obyYsNsfqiSl58nt8AO+rjQyFA3JXoLXGKDP2OFU/WdsTZDNMQEYd50aiR/193UUG4PTTT+ed73wn55xzjqx/BOERjJG3QBAeefzhH/5hOOuss9716Ec/munp6WYHRVXdVay1TelLHTg8kAAqoLB4DBCrsuQFa/HOYZKkamOrcCqUMU+126SaVnjjsUrt7xHUeGDWBFJhvJNeUKNNLBlllFFGGWWUcdNjYP4c28740EzMtQvoFwpVbWIoClugtCLWEUZp8mxIEkXlcUJLbWlngKqNCyDD4ZA0TQHI85w4jjHGEELAOUcURTjnANh111157WtfC/CuM888890SDQrCIxNRQAXhEcbFF18cVq1ahVKqCQpqkaO9i1IURdPW9gF7gNgAprzcBOsJIaBjs/DVSI1SbOsgrF0DPJld2w7KmsMsoIqEVoAmo4wyyiijjDJufBwTJO7PqkKB9x7rHbEpN1kIMBwMiLQmqoSLhebr+7JQaccueZ5jrW1M3L336Krctp0dcs8993D44Ydz2WWXyVpIEEQAEQRha+Tv3/628Na3vpVtt9kWAFfXzVaRRvAepTXOl6KINgYbPFrpzW5Dt6gA4loRVaVUhOp5VWC09dNymK+N1moTUxNGmR+LCSCqfUzGo7ggVztBEARB2PxFQrifk2dLRQkafPDgA1rrMiOknp8nVyItn5GaTXWBqUtcQgjzNmxCCKAUeZFjjMHocuPFOovWmhACPzj1NI4++miJEARBBBBBELYWDjvi8PCRj32M3Xbbjamp0tTU2oI0KlvP4hwoUyoSCtAKrwJ9V+CNJiEqBYgHIoBs7OqzSLZGW+hYqDPMJi9m93ULSRAEQRCEzZvDN7W6CBuZezcxiY9N32Hh+435koQyWtBKj0KLSvzwY5HEfDRww/XX8+53/xOf+9xnJFoQhEcAcqILwlbM+973vvB3f//3zA76dLu9UVziXdl6VpUBQcgLVBw3tbZ1woar7h/LxUIQBEEQhN81i2xm1DcXtiCO4uq2gLeuzDRRCk9obaroeXGMrh6fRDGu8HzlK1/h3/71ffz68osl5BGErRg5wQVhK+TAAw8M559/Pt1ul0E2JE7L1rYqeEy9C2IdmJEruydQBIdBYQL43JbdWuKYgJcSEkEQBEEQfrcLlZYAEhYoaa03chRl1odqL20UhIkcENXU4dYHLTNFgi+TYTfM9vmLN76BL35WskEEYau9rshbIAhbF1/96lfDC17wAtI0JcsykspkrLAFygXiJGntqASCVmRVm9r6ohADyocqBURtughXEARBEATht80CXiGTVTG2KHDO0UnSMmbxvhQ2jAE14RoWxkth8sGAOE1RRpMXjhCV2SNn/OxMXnf88dx81dUS/QjCVoac1IKwlfDyl788fPCDH6TX69HtdonjuDEHW6ze1hMonMMTiE3d/SUHoBMnqFCapWppmC0IgiAIwkPIQqFM3eFl7A62slGtOtwtvNrRrYcF1mxYz5IlS/GUez+zc7MUec5/fvjD/NO73yPrJUHYipATWhC2Aj7+8Y+H17/+9eR53jig9/t9ut0urrBESlc7IUAIeAKqChp81enFEfDegzZNAFDg0WhMlUAqbQFllFFGGWWUUcaHapzEUhARoQGLxaAxmMrvzJKqaH4WiRoJIMM8I01SfHX8/nBIr9MZHT/POP/c83jta1/LVVddJesmQRABRBCEh5KDDz44nHHGGWyzzTYURUEcx6WIQVkL27ScK2wjgLhQ+nnU7eDqi4AnUKAY4rltcC83rr6T9S5DJzHaO1Qoa3GDklFGGWWUUUYZZfzdj5OUBqYFwXmUUiQmYmpqiu2Wb8u2TLMMiDyNyTuV0OEZmaNab4m0KeMl5wjOE8Ux3jlc8MRxmSH713/913zoQx+StZMgiAAiCMJDwTe+8Y1wzDHHEEVRU+oyHA7pVDsX9df5cOQD0j7zfTX1exR5sAQVcaddz39/++ucdvbPWEeBmu6Q2QJDQAd5zwVBEARBeGjwivmxSAhkWUYv7dDrdNmwYQOEwNOe8lRefNRzefLSPem1j9H6Wlf/VwDBY4c5aacDzoPSBFugkpi8KLNrZ2dnuemmmzjwwANl/SQIWzByAgvCFsaznvWscOKJJ3LQQQdVc3/AWtvsUABYa4miCKAyQo1x1hKZqDQHcx6SCAv0cWQYvnPBj/nmT3/ITffeRZ4ach3IgyMouUwIgiAIgvDQUmeCtNGMbtNh9LUCpnN41VFHc/TTjmDbZDkBR4QmwaAIGED7UHaSCRp8KP/VviLBE3T5vG2vkdtuu41PfOIT/PM//7MESIKwBSInriBsQXzgAx8If/mXf9n4fMzNzTE1NVXO0yHgnENrPW4KBgzzAWmSNid8AAY4NlBwx2A9H/7vk/jNnbdw8/rVFB2D6iQU3mHr1E/n5c0XBEEQBOEhwS+wYqmzQdqiR1sESR3MzBXst/Nu/OGxL+aJuz+ODhrnByzTXSLABCC3gBmZpgZfdcBTVcwUCKH0SatjrLVr13LBBRdwwgkncN1118l6ShC2IOSEFYQtgAMOOCCcffbZLFmypLmtzvIoigKlVJPxUZPnOdZakiQhijTD4ZCo0yHH08dREPOtC37I577zNdaQM6scRazR3QTrPcM8A6ATxRjn5u26CIIgCIIgPJhM+n60v1cTAogOIxGkvl05T9fEmLmcw57wJP76uNexnASf9VmRThF5VdbFeA9FAXECcemZNswz0jjBOdfEWLUI0o7Fnv3sZ/OTn/xE1lSCsIUgJ6sgPMz5t3/7t/DGN74R51yT7VFPwMPhkCRJCCGU7W6ryXgyC2SYD9BJynoyAilXbLiJD3/uU1x8y3X4XkIeKzIdKCqD1KBGJqomKIyzIoAIgiAIgvCQCiCwkI/HwmKI06XfWaQNXQthXZ/9ttmJ41/yRxy29yH0gKI/y9K0B6o6ktKlYXzr+LYymQdwzjVlx0oplFJYazn99NN5znOeI+sqQdgCkBNVEB6m7LLLLuHzn/88z3jGMxpxo81gMKDb7Y7tRjjnUEqNfe8UeK0YEphD8aX//TZfOO3bDFJNlmr6Nif3biyQMMagjMEFX2aaGLlUCIIgCILw0NHeh5kURlQYX9ToAK7qfKcCpEGxVKckA0s8l3PEYw/hNcf9Ebsk2xMTIJQtczUK66uNpOo/Zy1KqQVjsSzLSCuj+RtuuIE3vvGNnHLKKRI0CcLDGCNvgSA8/PjjP/7j8Itf/IKVK1eOpV0qpVizZg3dbrfZjbDWYowhyzKiKGrEj/p2pzUbCJx3x5X86+c+xg8vOIthxzBnPDZSrM/6JJ2UOI4xJsLo0ghMWQ+h6v6i5W8iCIIgCMLDQAip5AVfZ6yqUvxoiyL17TqOiOOYwlmyPEPFEU7DrXfdyf9ecB6P2mMXzHRKR08BCo8jUpoYTXCOEBiLrZrjh0Ce5434MRgM2GabbfjjP/5j+v3+u84666x3y19KEB6eiEIpCA8zvvGNb4TnPOc5dLvd5rZazFCtjixFUeC9bybfopWimed5Y5R6WzHLf575bb59zk9ZP7uBZGaKvs3p2xxnFDPbrmDNhnXlBaEyD4s9GBRdHaGMYegyvIgggiAIgiA8RPgFBJCayfa4KpT3qzeH0jTFe0+WZcRK09MxiQO/bsixRzyblzzzuezT2ZEpoAcoT5lyonWzWsrzHKXUWNc95xzD4bApUa4zQq6//nqe/exnc80118haSxBEABEEYSEe/9jHhR/96EdML5khSZKxcpYQAkop1m1Yz/TMEqyzhBBIo5g8y0iTFO8cOjJYPDkei+bSm67kw1/+HBduuI2sF5clMa4qd0ljMJoNwz5RJ6VwluA8sYnoKINxAawjhEAwZS2tIAiCIAjCQ0E7w2NSDFmoPW758zKOyoqcoBVpp0OeZbi8YLrTpeMM2T3r2O9Ru/BHzz6G5z326SwjJnGWyINSVXeY6nlCtXhyRVkmo6rMkMIWRHVmrrNEJmLd+nW85z3v4YMnfkDWW4IgAoggCG0+9sGPhuOPP56km5JlOWk3wftyojXVLJ4XDhdrAgrwGDQGB4Ul0jGEQOZybBpzK3N8/sff5us/+xGzxlNEakEndb9AUFEHEroVUAS5UgiCIAiC8BDi78djJgUSPXks75lJutDPmLKKIx9zCH/y/GPZu7sjCZZpH1PM9omXTBGqx2kg5AU6iiFUGSK6NE5t36c8vOOMM87gNa96FbfdcqtEU4LwMEBOREF4CDnkkCeHD7z/RH7vcY8n7XVGM3IErppBVYDgA8YoLKMJ1oaCVGkiFHmWodOEIYHzb72Cj339i/zqjuuxy3vMuhyDli4ugiAIgiA8YghqVCYT+1GZb+0P4hTkzmJQdDAkhSeZK9hn25047veP5ujHHsEMmrrgJSsyjKcsPW4UlEoAUSPxoy2CKGAwHDAYDHjpS1/KmaefIWsvQXiIkZNQEB4inveCF4aTv/jfLJmZKmdLB4O5ObozUwQFc1lGp1P6e0TAsJ/R6aYEX7a17XS7FHg2+CFKJ8zi+OrPv8tXf3AKs9riewn3DmYhiUDr+7VzIgiCIAiCsKVSZ4BEfrxMxutSAIm7Hfr9PpGHmSjFzGaouYzdl2/Po3fdgxP+7Hh2YCkplikMkVNgLf1Bn96yJYQq10PBqE2NGm1W4R1Glz0nhtmQ//qv/+Kv/+rNsv4ShIcQOQEF4SHgu6d8Pzzz2c8iiSNc4VCAURpdtZt1AYZFTpomOOeJlS6n2AB+mKN7CU7BvXaWEPW4YnATHzjpv/jldVcyvd0K1g1myWzB0qVLyW1BoRABRBAEQRCERxRBlcKHqbI/dCizQmoBpFABjMb4UiTpBs2U06hBju0PedR2O/C3J7yBQ7bZj5Sc5SRNQGVdgYlTPFVbzbDwCmt2dpbp6WkCUOQ5115/HUc/7/lcd911sg4ThIcAOfEE4XfIM57y9PCt736HNE3pTZddXur50jO2eTDqPBsAa1E6AmtBK4gNAwIbCHz2zG/wxTO+y1xH0zeevitQ3hMZQ+oV3nuckQwQQRAEQRAeWWxKABl6C7EpDU2tR2WWxMFMnDIVpWzYsIEpFfHcQw7j+Be8nB2YQhVDlsQdXJZj0qQK1apMkHZQxyiYy4ZD0m6ncVFdt2Ytf/4Xr+fLX/6yrMUE4XeMkbdAEH43/OVfvjF86QtfxChFd6oHAVxm0ZGm8AGvFP08w3pPYkxZO1qU3V+U1vTnNhD3uqzN+rg44Tdrb+O9n/sop198PqsZsibkDEwgmuqSdFKc9xTWokxZhSoWIIIgCIIgPKJQo02lyqqj8QAJCnQc4RVltz2jieMYFRuyPGf9YI6k10WnMZdeeSWXXn0lK3balkct3xGNIorMaLOq2lNWilLkaIQQRwCiOMY7h9Ka4WDA9MwML3nxS1i+dNm7TvvBD94tfyhB+J1eFgRBeLA5+ev/E4499lg6Kip3BXwoZ8kswyWGEEdYRqZZMWCch2EBSQKRIgs5A+WxKuF/b7iI93z0RPJezNqQ46Yich0ogmdgc5yzdLtdEhMx6PfpxR0xQRUEQRAE4RFF7QFSZ4C0u9tZDUNbYDoJaEVRFKgAHRMTKw25RVuPzXKWzizBDTKWqYSX//4xHPeU57AtHdIqZhstrHQV51U3VMpL8B7nHCoETFJmjThrMVHEvatXs+rgg7j99ttlXSYIIoAIwpbNE57whPCt73ybbXbagRiDCQHlQjkjeg9agwGrAhkOQ4THoQpHN0pKF60iZy5y2DhlDs97PvMBzvrNJcwaj17aY302oAgFSadDURQEBclUh7lsyMDmLF+yFDdXOpcLgiAIgiA8UsSPuguM8aUWof3oZ06BSiKGWVbGTpUwkQ2GaKWYSjoo5wmufJDygWkd49bN8bid9+Ftr3sDe3V3aESQGFBBj9rAABiY3bCB6ZmZsdfmiwIdxwTvCSEQQuCVr34VX/ziF2VtJggigAjClsk73/nO8Pa3v51ur4cbzYOjTIwwOgODarVO845Ym8ZRvO9y+lHE9y4+k//58Slcdc9tzMWQRZWDeRyjnMf4KrUTsCaQG8iqIreOVSKACIIgCILwiBJA3EQGiKlisLodrlOjLJEFjxEC+IAJpYgSB0WsNKlXdL3mlS98Gc99wjN4FF1iV9A1cZUB4iAyWOuJoqh1wDLzw1tL3OmAc2AMzlpc8Hz285/jhONfJ+szQRABRBC2LL74xS+G4447rvTvUApjDJO+WDAyOlUBhoMBymiSNGVdf5Zub5o5N6QwMR/61klceMu1XH3nTfRjCFMpRfAURUGEIvGK2JcTuwdyEyhMKZJ4BakIIIIgCIIgPEIFEEVpgGoCC5YELySCOAWWgAJiG8osEqUwKLTWxEFhhp6jDz2C457+LPacfhRLiEigbJPb7TWRXmELjNKEEIhMKYgUw2EZH4YAWmPi8vZzzz+Plxx3HLfccKOs0wRBBBBBeHiz7777hlNPPZU999xzNKk6X9aEKggackb94Q0QAcVcn263B0B/2IdejyFwU3Y3//xfH+HS1Texxg8YKo/ppqg4onAW5QOJMhgXiP34joY1kOvy+8gjHiCCIAiCIDyiBBCvRgm3uu4EwygmWiw2GsVS5R1iW5bPmABKKag2uLCOubvu5Un7Hsyb//Q1HLB8VxJbsCKaguDQKh4JMNXovQcf0GbUiyJ4jzIaay06MgQUz33ec/nRqafJWk0QRAARhIcnf/ZnfxY+9rGPEccxaZqW7WedI47iUvFQEEwpgNSGp6b6l1STsMOTKViD5WdX/JKPnvxZ1uiM2TQwiOsSGU/hLMF5Em1ITQTWlxN7Jay0W7x5hWR/CIIgCILwiBNAYOQDoluiRy2C1AKIbgkh9f29LjeTQggYr1AhoH0lgFTH11oTBUWXmLCuz7GHHskbj3klS9GokGNUhM0LppIu3juKLGOq28N7Xz6/1rg8xwNxmjQCiK1iyK99+Su84hWvkPWaIIgAIggPL0466aTw0pe+lOnp6ea2oijIsozpqenS8LSS/y2jDJD2SZiFHKs0s3g+8NVPceq5P8dst5R7hrPkcSDEGqN1U49aCx7tSVsFNZbGKbqHIAiCIAjC/DKXhYSQ9n1DmeiBIzQxG5WZqqkOZtKYYZ7T7/fZdskypoeK/bffhdf/wR/xuO33JUUz6K9nqtcjJsIVOZ04QQF5nqMCxJX5Kqrc5FJa4wgoFMFaLr74Yo4++mjpEiMIIoAIwkPPXnvtFU466SSe8YxnlPOi9/jK0TuOy7THLMtI03j+RFwJFJYyK2RA4OrZW/m3T/4HV959K8V0wsAE1gzWkyQJqjbyqmpPjdbN8ymlyhTPSgCRchdBEARBEB7JtEtdarN5p1s/X6Akpi2UeEBX0odTo6VTbabqCcwNBkzNTBN1UgYbZlmiOsT9gp16S3n1C17KkY9+AttFMxQuIzUJEYq5uVlmpqbQqNFuGDA3O8vUzDSFs0QmwgePUZrhcMjs7CzHHXccP/nJT2TtJggigAjCQ8OTn/zkcMYZZ9Dtdsv2syGQJEkleKSN+BElMU6VE2wEZfpHqz1armE1lu9eeCYf//p/s9Y41oYc1UvJgqObxnSDJhQW6x2e0NS0Bl3WoIYQqslc4VU5qfvW2S0lMIIgCIIgPNIEkKhqw+eqsmCnxzSHMW+QSQGk/HmYVx5TlhqXG09aKRITkc316SUdYiKGwyFRFNHZkHPcqqfxlj95LT2VMttfz7LekjKDBCAEvHWYKKLIc+IkoT/o0+31Rs/nHKblFfLqV7+az3zmM7J+EwQRQAThd8tb3/rW8K//931oyg4vrj1BhcrMCkCX5qe2UjwaASSUM7PVsAH4h//3fznv5qu511j6qcKmhsxZVGQgy0gdRL50CSfS2OBx3pf+HrWD+EZOaR0kK0QQBEEQhEeWAKKrDSCvRxkgYYGFkJ6IkRrfEF96rGnUmKmq1XWaSCBWmshDkeVEUYKKDMPgWE7MdqsLdu8s4y1v/Gv2ftTuBArKYhiNwRFhGPRn6XanCcGiVATKUzhPpA2qFdfZoiCKY3525pkcdsThsoYTBBFABOF3w1e/+tXw4pe+pDIxVfjg0UqXJ1O71+3YZGrL2VUrCgIFiiHws6sv4L9P+SZX3nUL/RgGMWSmFEbqiVaHsotL3eIWNt6zXhAEQRAE4RG/yFnA16MtbiwkfGzqOIsdY6Hnih1sZ7rM3nEPuz9qF17+ghfynFVH0iMQAzGWLgmmKrahNQaquNJXio13YEyzcjvv3HN5zWtew6WXXyYRoSDc12uDvAWCsHnssssu4Xvf+x4HHXwwHpjrzzHVm8J5h9GmaW9W1qa0/gEoD9rj8RRKs9rPceovfsp3zvghN997F3q6S27KnQnf2mGYPFHr0hYZZZRRRhlllFFGGR++I0A+GLI07WGsZ0YnHPb4Q3jJ7x/N7lPbExFIJ3xA6qCvbN2rCYVHmQhcJYAEX2YDA9dedy1ve9vb+NrXvibrOUEQAUQQfrscdthh4ac//SkA1jvQunTnBlxwGGXGhYr2yRUAb3HWYtIEB2xwQ25bfTeDPEMlEZYwXnNaTYZV+/kyoaQyN5VRRhlllFFGGWWU8eE9eg1Jr0s+zLD9IUlQLEm6bDu1hOVpD23i+btdVcfA+uY8K0jTdKzUut/v0+12QSmyPOMv/uIv+MynT5I1nSCIACIIvx3+/u//PrzzH97B9MwMhLKMpRQlQjkJ9abmPWbSkcNQdmvRuvQTX9efpdProdA4HAZTz3ljI0HOVEEQBEEQhC0NBwxCgVam8vyozE/rMulFxA9XPbaMH8tsjyzPUEqRxAkE33QBTOKyhe5Xv/IVXv7yl0u0KAibgZwogrARPve5z4WXvfQ40rQDtkw/LIYD4m4Xl2WYNB1LR6zTFhuvjnqsfEIKW6CUIjIRAchtThIljejRnJRh4gCCIAiCIAjCFrXK8ioQUGhKE9PgPEnVKdBmGVGaghqJHp7JZoGOpJRNKFwZQ8Y6Gj1H8IQQiLThl7/8Jce+4IXcctutsr4ThI2fmoIgTLL33nuHb37zm6xatWqk0vtAyHNUmmL7faJutzqL1Ei5r86qWgRpJ3BoaAQQozRFUYAu1fxNnogihAiCIAiCIGxxWGdRAUw0Ei6C9yijx+JF1xpbj8Z4iCrRw9oCrTWxNuRF3sSQw+GQTqfDvfes5iUveQk/OfOnssYThEWQk0MQJjj00EPDqaeeyszMDFmWkaYpNi9bj/miQEdxNXtRlsQUBaTJogKIArQDnxeYtHqs9RCVOR/OeUykm8c5xrMjFWXKpJysgiAIgiAIWwbtcubgoSgKlNHEcZnREUK5h1az0F6Xh7LbYH0f54laQkotfFhriaII5xzee17/+tfz6U9/WkJHQVgAOTEEocXrXve68LGPfYw4jimKgjiOGzHDB0+kdNmHXRuU1qUgEsfzzqpaxS9b5VZmpo0qUvmIeHC2JYqoqvUto8eH1iSq5c8jCIIgCIKwxSyyDMzzc/NVfYvWLF7yXO2AOWcxcVSJKB5VlVw35TNANhySdjuNUWqgzDj+6Ec/yt+85a2y1hOEBc5NQRCAT3ziE+GVr3wlnU6nuc1ai4oi7GguwuPQAYwylbJf+nvUtP0/6q8XNDcNGzkr1cKPFwRBEARBELYQFjCzD0x0CgwTX7fjw1YAGIoCpXTZDlctsJqbiB0BfvyjH/KcZ/++hJCCsMApIwiPaP7nf/4nHHXUUSxbtqycZEJAKYX3nqA1cy7DmIgIAwQUAWstSRRXc5MaO5nac1c7o6MWQZqSFs8iM+IiZ2h9HxlllFFGGWWUUUYZH97jREwXFgvzFhJCFJDnZZl1O7i0FrwnaIWK45FAokZxZ/1wZwtuuuFGjj32WC677DJZ9wkCEMlbIDzSufzyy8Nee+1FkpQTTJZlxHGMUgpdpRrOmLSZqLwPaKWJVYSqZ5sQxmY11fra6LIXvJ+c48LExNcWRNQCSopifqqkjDLKKKOMMsooo4wPz3FiD6stUozdrcr+bSsUGtCdpIoxA9TZxnG04B7ZQs8XRzF77bUXZ599Ni960YvC6aefLiKI8IhHTgLhEcsee+wRzjnnHLbffnsAnCt9t40xo8kjBFxhiUxUtsGNDMVwSNztElxZHjNvthlTOWgMPNzEpDZP/BAEQRAEQRC26lXXZKdAPxEKtqtm6oxh3T5EZcIfnBuPQxcw4Xe2IDYRSilmZ2d505vexGc+8xlZ/wlyKgrCI42XvOQl4ZOf/CTLli3De4/Wuil7AfCVQ5XWerw3WT3zaEWWDTFxhI4MoTqV5k1QzWzmx2o0WWCi85vxunV1PxlllFFGGWWUUUYZH77jYgutjQkgbQ+PAGQ4Yupya48PjlRFRO3S6zBfYJlc6NVdYvI856STTuL1r3+9rAEFEUAE4ZHC3/7t34Z/+7d/myd41EIIlNkgcdxqd5sVeOfQcQRJhANspYvkODRmrNxz0utD4ZtJsU4OcRMTX9iMk1PKaWWUUUYZZZRRRhm3DPsPvZG4btIvLiwggLjqCGVcGYhRZXfBKraM0fM33SaeaDgckiQJWmvyPEcpxbe+9S2OO+44WQcKIoAIwtbO5z//+fCnf/qnDAYDut0uAP1+n16vt+D9vfd4AkqXAocFCuDau2+iu3QGG2s2DPvoJG7qOlUAXfl7GF+faB7whHar22pWDGr+vFUfY94JG8r7yyijjDLKKKOMMsr48B59FRvqUMV2VSxXf82ECFLHhaWmofFKE6wj1oZulGBcQOWW6aTD0niqJYawsO+I9yhTPlOWZaRp2mSDnHrqqbzpTW/immuukfWgIAKIIGyNXHHZ5WGf/fbFBY+J6uyOqoVtCI3goY2pzE49+ECIDBsoCBgGOC6+8XI+/Mn/x939tYQ0wUdQhHrq0mW2R9Co4FFBg/LoUE5rXo0Ej/rr+vZ6gpwUUQRBEARBEIQtizAR36mJ+E4v4gHnVXuhZvDeE+sYgiNbP2D7Zct40e8fzUsOez5dNDGeBF2WW08W4SgonCU2USN8tMWQ6667jr322kvCTUEEEEHYmtj5UTuEX519HstWLCdeOk0BuEomj6jUcWeJ45gADIqMbpw2anoWLBt0xHoKvvqz7/KtH32PPDIMjePuuVm6S6fIXZnhoUM58aig0bUgEsbdPYJaeJJro8UYVRAEQRAEYetaeLXiO73IfdpRY2wMgywjmIikE1NkFoZDdluyPQftsJJ/fuXf0cMyTQTeQ9CgHE6pqmuMHntepeab8N9x++286lWv4rQf/kDWhcIj4zyUt0DYmlm1alX44Wk/4FHLVqA7HYIpS1gsHo1ufDoMvhRCKg+QoSuqiSdmCFzPOj77/a9x+lk/Z60d4joRRWLwiWFoi8ZLpK3oKxExBEEQBEEQhPtBXS4TFNhIE6LSwyNkBctUwo7RFPsv3YG3vfLP2bO3PZEtMChMleUxzAs6SVnuPWm433QjrH6QDwa86jWv4UtfPlnWhsJWj5G3QNhaecITnhAuuugiojgi7nUbqd05iyKUpS9A8A5vLUZHFOtnMUmCMYZcQx/NFbO38rEvf4bTzz+bPIKZHbZlzWCOgbfE3ZT+YNCkFJqqrlO1JheFdLoVBEEQBEEQNh8FkFtibTBa4yqz/iiJcQQ2DPrcu24tv7ziErZbuSPbLdsRo1XZISZApOKq3rrM/KDyoGubrCoCCoWJYo49+miC9+/62S9+/m5594WtGRFAhK2S5z//+eG0004DIElTCuqKyIAOEBmDqVuIeY+3lshEmCQBBbMhJ1cxV2V38P997EQuve16skSXpqf5kNxA1E0JuswlVNWxdGip6i1EABEEQRAEQRA2FwUYFFopvPdY51BaESUxKjIUKmCmOtx695388tKL2XaX7Vm5YhcUjkRp8KEsiakOploCSKgiU60UrijQKHQU86QnPpHddt31Xad8/3sigghb9bklCFsVb3nLW8KJJ54IlK2/4k6HIQ6DIaZS/Zqe6ZUO3iq4XDecxfemOeu2y3jXf36Ae/wQNd0hV4H1/Vksge7MNJm3FEXB0ukZimGGDqMMkDZ11xdBEARBEARB2Bw0EJuI4DzOOTwBp8AbhTMKp8AYg+sP2T6eYjtSXnTIYbz28BezDAVzfZJ4plQ+THlASy2CBBQQhYBBVz13PWhNPujzkzN/ynOOfr6sE4WtEvlgC1sVH/rQh8Jf/dVfNd/neU4wGkxVolL9U36+UYfzjjkszqR8/fwf8f++/RXyJTF353P0fYGPDWmng44M/WzIcDjEGEPHxGgfRuUvYSR81KMIIIIgCIIgCMJ9WqgZjbeOCIVRmhACubcURhE6MXNFxlS3ixlaekPPiqHi2Ccdxhtf8CcsIyZ1QDBgyoxlR2sPEHB5ThInZP0+adqByguPEPjBaaeKCCJsneeVvAXC1sInPvGJ8Od//uc459Ba0+/3mZqaKtuQVfdpUv68Q/sAaNAKpzxDpdmA5+OnfZGv/+QHuOmUdXaITyN8GmGDxxGw1uK9Z2pqihjN7IYN9Ew8Jn7UwoeTM0wQBEEQBEG4jzgNOR4doIMhVQbtQ1kOo8Elhkx5MmcxAaLMsR0pS3LFE/c5gLf86WvZQS2hh0KjRqXfTaAKtUGedw5tDHPr1zM1NVV2kNGK8887j0Oe/CSJZgURQATh4cbFF18cDjroIKy1GGOarizel2mDsYkIqhQlSr+OuvQl4FRgqDRrcPyfL3yAn11xEXqbGe5cey9pmqKTiNkiozPVw1pL7iydOMFbRygsaZyAdZj6kNXz1AKIVzQ/EwRBEARBEITNEUAKA0opYgdR4UkcaFQjjthYU1S7b704Jc4cfkOf5VGXHbtL+NTbPsA0ii6aBEUCtQlItRLUjWN/VuRlTFv/zDkwhmE2pNvtyppREAFEEB4uXHrppeHAAw9c/A4BQp6h4ri5yFtrsd4RJSnrKLiLPu/+9Ef4xbWXEZb1yA0472GYo7UGrXGthu11q9u67KUtcHgFXpfiR50BspA3iCAIgiAIgiAsRC2AAMQOYl+OdQzqFWUmiBqVXavKjF976FrYe2Y73vHnb+LRvZWkWKZySEwCWQFxXN7ZgEOXGdPeEWuDCmV5jEkSCluQZRnPeMYzuPDCC2XtKGzxaHkLhC2VRz/60eE3v/lN2GeffTZ5XxXHpZJd2FKQiCJUknIPGdfM3c3fffR9/Pw3F9PvamaNZ30+oMATdVKSJCmNqHwpdBhfiR6hnGCMH7W7bdPOBBEEQRAEQRCEzV6kBYhc+a8WPVwruxjKGDSq/qlQxpyFhmEEGxK4+M4b+bv3/ws/v+lCMhQkCTYbluLHcIjNhoQQcJT+IGjDsMhx1mKSBFsUGGOYnp7mtNNO47jjjpPtPGGLR5ZmwhbJXnvtFS644ILShyOON3n/fn+WXq8HuYUkwgLrsJx7+5X8y2c+zs35ejZEgWTJFJm35MOMTpyAK8tcEhU1E8vYCRTKSam+OVAq9q5W5XU1gXnJABEEQRAEQRA2UwCh3GiD0Yaab/1sMe85p0uBRAVIXWDKKmaGgde98GX8yZOexxTAsCBKUtBU4oemwBITYQCbFaRJ3AS4zjmMMWzYsIHjjjuO0047TdaQwhaLkbdA2NI46KCDwhVXXEGSJCilyhKVjRAAE0cESjftOZcz0PCjy87j3//7k9ySr8ct7ZFFMMiG+NyS6lLwsN4RaYPxqlHf27KhovYVKfFVKojXtDxHxkUSQRAEQRAEQdioABIg8WXGcZ1VbCtxo/5eqVE8qloxah2PmiQmLwpcUXDj9Tdwz5p7OHj/x6KjGKd884CAwnuHUaVhamQMCihsmQFSUxQFr371q7n00kvfdcUVV7xb/kqCCCCC8CDzvOc9L/zgBz8gSRKcc0RR1HR92RgFngLFEMd67fnmeT/mE9/8Ere5OfyyLmsHc+UEYj0dFZEagy0KHKEUWlxAh5GwUU88TJS5TH4No3pMEUAEQRAEQRCEzRJAGHnItcWPOtacty9XbbpVPQ4ByPMMvCdJUwbDIdffdjOX33Q1j33c49FEFDgKl5PqmFgZfFGUIohWOB9wzjaxtjGGJEkAeOELX4i19l2/+MUvRAQRRAARhAeLV73qVeG///u/mwtxHMdkWbbJEhgPFDgGODIivnDGt/jEt05mfVeRdQ339jeQxDEz3R6JV5hQqt+BgDKGoiiIlB7v7KLHhY6xTJBqUqonIcXCHiGCIAiCIAiCsJgAUseXvvWPVtYHVamLUiMBpJ0VsnRmhuFggIojrAZrFLfcfSfX3Xojqx5zEFP06OgIrCVRpox3rUMZAyjiOALKzI8oKr+enS3Lyo866iiAd5155pkigggigAjCb5uXvexl4eMf/zhaa5IkwRjTXIA3RaimgXuKWf7zeyfzlZ//iLXTivWxZ67I2HabbQj9jNgGitk5lA8EAkEr0k6Hwlk0upmAardtp8cVeCYmHdrih/h/CIIgCIIgCJuJUiPxo445G5GD8RLrUP1TlU9IqISRwdr1rFi+lDVzs/SWLWGDy1GdmHvXreO8889j//324VHd5UTOE9kA2qBUWVsTVJlBEkURWmuUUgyHwyb2ds5x5JFHsnz58neddtppIoIIW865JW+B8HDn+OOPDx/5yEfodDrlBTd4iqIgTVJc5dFBmP/JdtVkkONZQ8bbPvRezr3pN7gVU9wbF/jEEHKLtp6u18RB4bxlanqa2SJjw2CO3sx02RbMlrZTQY23G4NRR5i61KWNlz+fIAiCIAiC8ACpRY163BQaUM6DVpBEzGYDZqaX0F+/gW4e2LGzhN06K/i3t76TnVhCl0AUFG6YYdK0jKJ1GfB678fKzfv9Pr1ejwAMsyE/+MEPeNELj5V1pSACiCA8UN71rneFt7/97SRxMhIgWsKCqf4RPD7L0dpAbErVGigI3Bo28Lr3voO7GbDBZuR4QmIgMuADeI/2433VJ308JgWNSQGknmgEQRAEQRAE4beNr2JNv5kx50KbcDrosmQGSBx0MUxZxYnveDePTnemR8FSYiiAoEsHvSRGMeoEA2WTAK01AUVhC+Io5uvf+Dove/FLZG0pPOyRD6nwsOXtb397eO973zuW3eEmBBANxIAKHlypVAcVGKhAhuaqNTfxrv/8ENexnvXGYa0tH6c1Wmu893jvy3S/+ziJCIIgCIIgCMLDkbJtbimVmFCVy4TRzwBcYdl+eindoeetrzieZ618PKkdslR1yuA30swO5uh1umXcbB25Leh0OgTABk9W5HSSDoNBn5//5Kc8//nPl/Wl8LBGPqDCw5ITTzwx/PVf/zXD4ZBup1tehKsaSEv5DyCq/hnvIFCKHzrQx3DV4Db+4cR/5V6Vc7fvk0XVh77VOtd7Twhh0wKInCmCIAiCIAjCFkRAj7oRtjYUgxplkoRhwQwR26iUN7/8lTxnv6cyDSSAQo86GdaP99W2oNbNcWbnZpmemkYD5517Lk9+8pMlchYetkjWvvCw4z/+4z/CW97yFpxzdLtdiuGwcXeqS14a4aP+EGsFRjHrCzIMv773Wv7i/7yNe0zBPSEjVyOhQylFCAFfXcA3JX4IgiAIgiAIwpbGpPihJ0UQrdDdBJ9G5Inmn/7jg/zw6nNYB6zFscEOq8zrQDYcVgctO8XgPQrIhgOmK2NU6ywHHrSKa6+9Vuz/hYfveSFvgfBw4tOf/nR4xStegfe+6TWOdTQNixQLynZBedYVA1Tc5efXX8Q/fepjbOgobs3Wo3spmpH4AYyJH7UgIgiCIAiCIAhbC3XIrCbKX1yVVT0YDtl2m20YrF2P6uesXLINai7jeU8/ktc8+w9ZQUIoBnRMTKqj0lQ1AFpDCATvyq+NJsszTByjVZk1csZPzuCoI58pa03hYYd8KIWHDf/5n/8Zjj/++DGhYna2TKkDWg6k41f2oCHDsw7Hjy89mw984VP0u5p+qlmvLCaNUUVRXfwDLgRUCHjAKEVQqvm+bS61uSZTgiAIgiAIgvCwXOzVnQpbZv9elyKISRPuueceVixfzlSUMHvnarbtzpBvmOOEF/8Rf/Lko1lC0mRea2vL7ov1ErIaClsQFJgoxgePtZYkTjjr5z/nGc94hqw3hYfXOSFvgfBw4GMf+1j4y7/8y/IiWhTEcVyKH9NlG1pXfVhN/U2ob4ChgtVknHHtrzjx859kNgrMYVmXD5hetpQsy4gr12p8mNdGbFOjIAiCIAiCIGxZlDuHmjAvA6SdBYJSTE1Ncfcdd7L90uWErEAVjhU+4Q8PfQ4vfNpR7NjbDo2lgyFBoXytpngwhsIWRHHM7KDfZHAbY4hQ3Hbbbey8886y5hREABGEmo9+9KPhDW94A/1+nzRNMcY0Hh3GGGbzITrpAJC2RRANVsN64L/P+x7/9e2vcI/K0Ut6DF0BWqOcR/lA0KOPejvDJISwWSaoWoQQQRAEQRAEYQvCK98IIDBRBlNlgbhQCSVa09ERNsuZ6U7h7l7PknWOFx/2HP7w2BfzqHQFMZ44BFJlCM6jtKbIc+I0wTlXevIpjata49btc4uioNvtyrpTEAFEED796U+HV7/61Ru9Tw4MgQJPF00EeFugopi7/RzfO+/nfPzbX2J9CnMJFFXdSuwhdmUZS6EfWCcXyQQRBEEQBEEQtiSC9mNxbF0K49UoE8S37l8bpqoAXQvbFQnunll+/xlH8OqX/THbmyUYHCkaE0BNlMLAKEkbSs89rTVFUXDnnXeycuVKWXsKDznyIRQeMr7+9a+HP/iDP9jk/SwwS4EhRuMIwROrmAGOr53+PT797a8ymNLMxZBFVU1jgMhD4sqLeGHK2+VEEQRBEARBELZ2AoDyYyXdbS8QGG0OBjX6eU3iNN1cscx0YJDzuP0O4NXH/RF7rtgZ7W1pilo/0UYC5VoECSFw2WWX8f+zd+dxklX1/f9f55x7b1X1MgPDooBAUFFxxzURjOKSqNEkxDUaDUZMTOLXaJRsJj81X42aYIxfo9EYNQZ3jTu4IIILggiiiICiIvvOLN1dy733nM/vj3urunpmYEZlhumZ95NHPU4vVd1Tt24X9fnU53w+T3jCE7j66qv10lruMDr55A7x1a9+1X791399+2+QDJwD1+x+GVjFV791Fl/9zrdYiCV+vkPtm3I+o3ly99Zul6GpADGd7SIiIiKyBxjnMpJbTno4tj4S99ZUVaTwgXLTEus6szz6Yb/GIx7wYOZCl2rQZ7Y3s/VfOnn5nlZsPY8xcskll/DEJz6RK6+8Uq/M5Q6hE092un/913+1F7/4xeR5DsBgMKDX6932M/jSCDod8JAskrLAkMSAikCHmnrFc+74xB5Pcxlf/C+4ioiIiIisJnYrQZ9vP3e3cbsIRDJGjJilQw4MBxvZp7eGAkddlXTyYrt+qZlR1/Xktf/i4iJHHHEEV111lWJR2el00slO9Z//+Z/2whe+kMFgQLfbZTAYMDMzc9s3SjT7YBJt2jpBESb7FmuLZG1m2U/dZPwcnKa+owSIiIiIiOxpNs9LeFa+Ybj56+ca2EhFAjp4OgQywKVIhsM7vyKatOngctxslWbYgPd+i3/PRRddxBOf+ESuuOIKxaOyU+mEk53mpJNOsmc/+9lbPAmORiM6nc5tP2NHoD+Amd4kZV2XI0KeN2V1FreecZ7safS32/80RERERER2+UBvKithrNwO7mxlsmLz19DRQ+V8+4agYbGmF3IwqAdDsm53i9fbWwswFxYWmJ+fB5p+IFVVYWZ0u13OO+88fvd3f1eVILJz/y50CGRn+I//+A/7oz/6I4qiYGlpidnZWfr9Pr1eb5sjaA2o2o8DMBoOCObodJvRuKPhkE6vmFx3827WzYmuBIiIiIiI7EFBXlr+ZFw5PV2p4bcSDE6/li6rkuAchc/aNyQT+NBeMUHwK372ytfeTc+PEAJmxmg0otu+djczyrKk0+lwxRVXcOihhyomlZ37tyGyI731rW+1F//Zn4Nzk1nhK55o2zniWzszx8UfZZt/zmiSIAFIdcSZ4bKs6XLN8hO72+JJWAkQEREREdmDgrxbSYBs+Tp55evd8TZwZ+0Xx5fU/kDvl5uItJet9RuJ1m5Edx7DcDgWNm1izZo1xKomhEBZllx77bX8yl0PU1wqO+9vQ2RHefe7323P/8PjtkxwbJFhSCvPSLcyA+3bp2i3PdkIhzIWIiIiIiK3x+tk+wUiSQeDekieFW1CxeOAqhzRzTvLP7eqIc/47nnnc+TDHqTYVHY4r0MgO8qb3/xm+/3f/31c2DJLvEXr6enP3ZYnqbOpLPQv+iQtIiIiIrIn+0VeJ2/rdfyt/J4ia7eom0223Iwnx6Sqaq9mYPDABx/JF77wJb2Klx1OWTbZIf7mb/7GXvOa11AUxaSHx22V3Dni5LaJlZm5pklT0EEVEREREVkljIhzjuFwSJHneJ+viELHgxDM2jjBwTnnnMuvPvyhilFlh9HJJbe7F7/4xfbWt76VwWBAr9djcTgg7/YwmsTG5mO3mv2IaYsTMyw/e4JTsZKIiIiIyOqRll/LA5hvtr17P+lJsmFhI2vn1zIYjeh1OmTAF075Ik/8rScoTpUdQieW3K6OO+44e+9730tVVeR5zmg0ouh0iEwlNaa3srjUjLRtCuAm1R/ZeMtLSu0NA7UOr4iIiIjIqggy66okOE8WsuXX/w5SMpKH5AP9ekg36zZBaUwU3uMMPvb+D/GMP3y2YlXZIeemyO3imc98pn34wx+efJ5Swk83P510kJ66kU/goQTq5edFMqBofkiTIHF+apOMiIiIiIishmAzB1JdE1xo9rm4pgJkSMQRJm+A5sBwoc9sbwaAE098Iyf87d8oXpXbVaZDILeHZz3rWfaud71rxddGoxG9Xo9qVJIXxfLsLQNIk/0wJTDAKIFEWpEAyX1ztYqkBIiIiIiIyCqQgLJ94T9DRifLlr8RK1wWyGne4DQig3KECwWzczPQH0Ge84oTTmBh0Ld//Md/VBJEbjc6meSXdsQRR9h5551Hr9cD2k7Pzk3WZs53tuXWlwBDYAC86+QPMgpQh6Y8JEtQRMgjeIPSOUxnq4iIiIjILs+AxWrEoQfdhaPu/yAO6u5DD0cBxFFJKApwnkFVEvICSOR4XExQJciyZu+8gxNOOIETTzxRkYDcLnQiyS/lkY98pJ122mnkeX6b10uxwodAVVX4PGdITU2gj+Of3v8Wvv6D8xhmUAYwByE1SZCs3S5Tt82SRERERERk1xcw5rMOa1PGv73qdRzk9qaLkcVE8AHMSJbwoYkjLCac902VSEpNz1QH3nue+cxn8tGPflTRgPzSdBLJL+xhD3uYffzjH+fggw/e5nXLckheFIxSTe09FZ4FIv/2v//FmRd9lxvj0iQBEn2TAPFA1u57iU4JEBERERGRVRFkGmQY9EccumZfDl2zL6/605dzAPNkqWLW5211uIFrZj9aXeOcAx+Wo1QHi4uLzM3N8bznPY+TTjpJEYH8cuemDoH8oq6//nrr9XrMz89vx7UTNUaNY4nEEM+7vvRBPvnVL3JLHBLnOowyqH2TABk/cXqzZghMcs1UGBERERER2aWZa7bFd1wg9EvmKsev3f0+/PUL/pw7s2ayHcaNt8gb4JbH5AIMR0O63W4TSaTEDTfcwIte9CI+/elPK4aVX5hOHvmF3HDDDTY/P0+3291y2stWJUoSIzxLwHtO+yifOONLbPQldTdjiZrKQ3LNZbnaw5oMcnJ4JUBERERERHZ5yTU5jSwEslHELQ45dM2+7GU5//V3b6FDzRyerjlcasNS55sxuZaaLfHOU1UVeZ5TliVFUTAYDLjb3e7GtddeqzhWfiE6ceTndsEFF9j97ne/n+s2o2qI5TkLJD75rS/zzk9+gIXcqLoZi7HE5csjsJonTWsm5rbJEKf9LyIiIiIiq0YIgU2bNrH33BoKPHHTEvsWsxx58N35/57/MtbRZQYozLBRhe8U0CZOIkag6RPiXBMHjLfCjEYj7nOf+/CTn/xEAYL83HTSyM/ljDPOsEc96lGTJyCAGCMhhFu9TQRqEgOMT3zrVP7j4x+gPx9Y7ypiJ2OpHNLJcoKBT83Ul+QMc812GNOpKiIiIiKyqsSqZs2aNdyysBGXjL1n53H9knyp5NijHsufPuU57EePGcCqiswHCJ4aw+HweCwl6roGoCiKSdxx2WWX8YQnPIEf/ehHChLk56ITRrbbSSedZM961rOIMdLpdCalaNtSAhsZ8vUffpcT3/ef3GBD+msyBl3PwmjA3NwccTBqx94aod3qEh3Uod0SqAoQEREREZFVYTzMwGUBujk3bVhPnufsM7+W+uZNzCzVvPh3/4DHP/Dh3G1ufzrTsYNFvAsEPGnqjVZrq0HG22LOP/98HvSgBylIkJ9LpkMg2+Nf//Vf7elPfRpZyMhCRoyRvCgwoI41eVg+lYxmetXYiMR3rv4xr3vP2xn2AnXRYX25QOjM4IKnrCum60c2z3W4qSdSERERERHZtTkD5xyLi4t03Czr9t6bTUuLbBos0e1k+FDwzo9/AIDZox/LvszQAQKQmcM7qOuaLMvo9/v0ej2cc00MkjUTZI584JGcc/a37GG/+nAlQWT7z00dAtmWl73sZfavb/rX5pPUnjW+2doSaTO8JMrRiKLTYRgrYvA4AhUVP7r5Gl7wplexoZuIvpn0Uvvl5kjjE3E8+nZa0uEXEREREVl1bu3Ny2bAAXTNkw1r/uwZz+X3HvB41gBzQCgj+LzpBxI8DqjrRBban2htBOEcmPGNr32NRx7zaMW1sl2CDoHclj/90z+1V7/61fR6valnLSYNitoPibGiyIumNM0HRiSGJL53+Y949Tv+jevyEUtF29PDtTkUW3mB5SlY44uIiIiIiKw+42EG40SIsyYGMAfJQ+UMlwXOv+B77HvQvtxl3zvTJZC79p1SIEXDB4f3TQARk2Ep4kOgGg4JIXDIrxxGryhe/eWvfOU1OuqyLUqAyK16/OMfb29/+9tZt27d5GuxfT7ytIkLmkt0CXOe2mIzwgrPpeuv5rXvfTuX9dcz7Dii9rCIiIiIiOz2zDUxQ5qqy3BTXzMHdYwUWU7VH3LBd87niHsczkF7HYClskmC+IDHsBhxwRMB8+BCwDkIWQbOEUcjHvnoY6iGw1d//cwzlQSR26QEiGzV4Ycfbqeccgr77LMPVVURfGieqGgqMzzt/ql2j4r3jirVOJ+xwIjrqwX+4S3/zA82XAv7zjO0CvUxFRERERHZA0wlOtpPm4+nvj4zM8PGDRvYZ90+eBznnHkWR9z/CPabvxPOGxntNAQMC46IYfimB2FdYSkRfCCWFcHgsb/5m/zo4otffeFFP1ASRG7r1BTZ0jXXXGP77bcfWZZR1zUhy0hs1rMDIKXJE1PtEiMfuImS417zMi4fbsD2m+OmxY108kIHVURERERkD2BuuedfSMvb32G5D6BzDp+M1B+xT3eWbOOQ/TtzvPUf/4WDmGd2VNMLnab3oHdEoLSK3OV4jAxHORjS6XQhGZhRVyUP+dWH873vX6A4V7ZKmxJkCxdddJEdcMABhBAoyxKfhUkz0nHPD2jzHuMvJCM6uJkB//e9b+GijdcxXNul6gQsD6r+EBERERGRiWE5InQKQq/DjQsbqWcLRvMFf/KaE7hkeA2pU2CZY6kaLfcGTNbOY3A4IM9zLLY16sGTFQWnfvGLPPD+D1A7QdkqJUBkhU9/+tN2xBFHABBjpCgKPG7S62N8MRKJtJwAyT2lc7z1Y+/lyxd9m/zQfbkhLXHzpg1kXjutRERERET2JG5q0MFWA1HfhKIpOPI1s/g1M1xy3ZWsdxVvfNfbuHD9FSySGBXGUhrgSRTeUZAIJAZLfXwIuCxjWI7aEnXPfne6EyeddBL3vOc9lQSRLSgylYn//u//tqc//enUdY33nhACS0tLFEUx6drsJpUczfOJB2oHA+D17/8PTv7eN6nXzXBDtUQ+2yU4T4bHmWm/lYiIiIjInmCqbHy8/WUST7RfzPOcfr+PAZ1el02LCxS9LtE7NmzayPXXXs8RD7g3s36ezAUCRoiGx0MVyTtdcLDUX6I3O0MiYSnhvWf//ffnHve8J+9///vVD0RWUAWIAPAP//AP9vSnP32S+HDOMRwOmZ2dbWfSJqgrIOFI+JQIQEViCfjnz76X0y77HjeFiipzdJLDLY7ohpw6RR1gEREREZE9xLj6I1jTAyRMTY8cf4060ikKQgiMyhJXZKTMM/LGoHCccfF5/ON/vYUlEgNqwBN8BoMSXGjiExLd2VkqwJyHELD2jdrHPvaxfO5zn1MViKygChDhSU96kp144onMz883T1htmYeZEUKgrip88O3sqvEoqqavxy2UfO57Z/LBM7/IT5ZuJvVy8l4XqropeQuBWEeCc6oAERERERHZQ4yHJoxjgHEFCG7qa1OfG03z1PEI3Wymy9Kgz0+u+BFH3f9XKYAOvk1+GHgHfvz27PLP8841DVa95+CDD2bdunWv/tKXvqRKEAGUANnjHXnkkfa5z32OtWvXUtc1g8GAoigwM8wM7z0+BMwZKSU8Aec8o0GffuE459pLee3738Flow0Mu4FO0cHFRBqWZM5D5kjO8KYEiIiIiIjInsKz2baX1uZJEG9bJkKSgzqDjYsLXH/1NczOdrj3wXej6zw+Rghh8ktS+5Omp1V6YDQa0e12edjDHsaPf/zjV3//+99XEkQUk+7pbr75Zlu3bh0pJcqypNvtAjTJjnFjIoxRLClCTogQyxF1r8MF66/kj9/0D1zrR9Rrurgs4EY1cTAid54szxmFRGWJgMepAE1EREREZI8Qpl77b2siZGK58gOaZEZFJKsSd/I9ZhdrTvj9F/Ab93sEa8gJMRLaJEjEY5sFuL5dq6oiz3MAjjnmGM444wzFv3s49QDZg33lK1+xdevWNaNuvZ8kP8aVHwB1XWNm1JYwoIwlda/DNSzxt//xz9wQKqqZnLquoWwuIQTCbIe6cMQYccmU/BARERER2QOZaxIc05etBaXj3iDjipAiZPg8Y/1wiX5mnPj+d3HezT/hRob0g6OkSZQEIGsvYbMAN8uyycef+cxn9GCIEiB7qpNPPtmOOeYYFhcXKYqClBIxRqqqoqqqyfVijATnmcs6LPYXqYqCG6n549f/LVfWi5QdT/IOqoivm6oRnwXKFBlUJcmMoih0wEVERERE9iCTfh5TH0e/MiEC7TaZ9jLeNuMN0rCklxdYHri5HsB+a3jp61/FVSywBFTjn2FMfqCzpj1IpKlod84xGo0AKIqCn/70p3pbdg+nBMge6JWvfKU96UlPAmBubm6S8AghkOf5JGFR1001BwauhjUzs1zBJv7yvf/Mj9MCV4420ul0oD9iPutQuIALjorIsCpJWNPN2W277E1ERERERHYf4yTHeGtLdMv9PaK/7QA1GMxlHTbeeCP77Lcvbq7Lj9dfj9t/LS9/w2v4aX09S9TUFiGOMx7Nxbdxh/eefr9Pp9Ohrms6nQ6HHHIIZ555ppIgezA1Qd3DPO1pT7N3vOMdlGXZVHe0e+fGa13XOOdIKZFlGd57RnWJzx0LGG/+7Ac49cJz6BfQ22sNCzevZ747Q0qJ2hLJNYOnvGu21DgcS0tL5CFTwxkRERERkT3I+E3Qzd8MnW6M6sbNT1dcx8hDhs8yblnYSOgW5J0OVVmxuLDA5T+9jEc95Gi8g9x5QjvFcvxDvWv6f3S7XcyaYQ4hBJxzHHLIIXS73Vefdtppaoq6B1JMugd5yMMeaud86xxG5YgQAlnItnySwiblYt41qdk+iZtIvOnkd3Hqt75J6mUsWsWgKimKounz4dyKxkV+s7xqcqBUq4iIiIjInilNRZ7TscLWAlKb+k5ybV8Q82QJehV0a/i1e96Pv3/ey1hDABuy1nXJUmpimGjNpJjbiHaf+9zn8v73v1/x8B5GD/ge5Mabb7K99tqL4Ntqj1g3TzBmZFnG1gbVVnXFzQz54Hmn8dmzz+DKK6+kmO0R2zK2rMipYk3cLL3h2318Y9ErASIiIiIiItsXpiY85lb2BskS5Ak6NawJHR555EP542P/gLV0yK1ixgI+GvgA3t9mtDscDnnqU5/KKaecoph4D5LpEOwZvv71r9u6vfamrmuyItDv95mZabau+HDrm/BSHbFqxNc+eQrDwSbWRY+31DQcyjyMIsO6nkyNgeWM7vQziZIfIiIiIiKyPcbb6qdjCt9OicnaC8NNfOfUr3LJr9ybo498KH5U44sc6qqJU7bR7bLb7fK2t72No446yq655holQfYQeqD3AG9729vsz/7szyY9P6bnYW/OzFaMwU0psRhLRjlANmmyPNVsmYhNqkf8ZifWeFUCREREREREtpfhSZvFF46mieV49K2jgqpmbd5rqs/HAYoDwm1XgMQYSSlx7bXXcuihhyou3kPogd7NvfSlL7UTTzyR0WjEzMzMiuRHSs3Y2nEPj+mkBzD53IB+GpH5HAMqIp7QJj8Sbiq9Op348DrBRERERETk5zTOZcCWhRyhjTEGoz7znRmsrsmcB+eJ/QGh6FCPhmRzc9sVjJgZp512Go9//OMVuuwB9CDvxp70pCfZBz/4QdauXQvA+vXrmZubI89zRqMRnU7nVqtCzNrWQ21HZbM4mQ5jDnAeA2oSfuppaTpDqwSIiIiIiIj8vMbV49MVILcawBqQEnVVkWVZ0/x0HJlsIxhZWlrCOcfMzAyveMUreNOb3qTwZTenB3g3dfe7391OO+007nKXu+C9Z2lhkdm5OYDmySHPGQ2HdLpdLKVm20uYmops1sykarfEuKkeH7h2q4xrt8OYTSbGTJIe2vMiIiIiIiK/cKSaVn5uW0ayk3im/VIChtUIzJgpZm4z2B2/EQwwGo3w3nPcccfxwQ9+UDHy7nxa6RDsni688EK7z33u0zwRtOOgqrIkLwrqqsI5R8jaHrjjZMf4uaUdgzv9NWzc2dRhGHVdE/ItJ8co+SEiIiIiIr98pDqVALHNVgdmCec9dYpEjOAzIlATyQhkeMJ2/JrhcEi32wXg8ssv51GPehSXX3654uTd9bTSIdj9nHLKKXbMMcfQ7XZXZDYnzx/jqg7nJltctsfmeY3pnKzf2kllK7+ovIiIiIiIiGyftGXAalsGHQZElgc1WPut6QTI9sY/KSWuuuoqNUXdjemB3c0861nPsg996ENUbZVHlmWUZUlRFLd5uxjjNn6y32YCY/OTybU3sDb5kXS2iYiIiIjINiVwCW/LccX0ONzJtdo4wxxEt/y5Azpk+G2Eu+OhEOOPoRkE8alPfYpjjz1W0ctuSA/qbuSYY46xr3zlKywuLjI3N4eZUZYlnU6HKtbNA95OexlvXTFsRQ+PX+oksds+01QBIiIiIiIi2ydtveHpVqpApq84+ba5Sfyz4ke01SDjaZghBOq6bhqoTjnhhBM48cQTFS/vZvSA7iYOP/xwO+2009h///3pdDoMh0OKosB730x02cZWl2Tptk8UW67o2N4Tapx9HWdlTWebiIiIiIhsT6BqW26Bcdv5jqo5cD7b7t813TZg48aNrFmzhqqqeNKTnsRpp52mKGZ3Oq90CHYPp512mh1zzDGTUbV+ampL2TY/bfbF2aT6I6Zm28t0RcjW/MKTXcbZD51lIiIiIiLyc0nbf9WtxCmWtnwDdroHyHTzU4CqqsjzfEUsddVVV3HwwQcrmtmNZDoEq99b3/pWe8xjHjMp3ZpOfgDkRUEZ60mpV/BNdnPzrTC3JjKVBNme5x839ZTVfuz1MImIiIiIyHZy7RTK6STG1lIivo05pvuEALgQthq/pJQwMzqdDtCMwO10OpPEyDiWMjPucpe78K1vfcse/vCHKwmyu5xXOgSr23Oe8xx7//vfv2LfWowRMyOE0GQwQ5h0RZ5WxQozI8/y26wAufWnm9u+5ubb85QEERERERGRnydQtakYwzb7fPpNWjf1uQcwN+n3MZ4As/l0TNiyEsTMqKpqRZLkxBNP5IQTTlDsvBudV7IK3ec+97Gzv3Emc2vWNn+sscZlGTioUyT4gGF4XDOFhUTEU1Jzy2CRGzbexNKoxAWw5JpZ2+bBJZx5bOrz6SeZ21onTxzjKhDXdGoe79fz2/lztGrVqlWrVq1atWrVumeusHKiZNosvpiuCnHGimkxvl0LH6hiJFqzpWWu6LL3zFr2mZ1nJnRwlvCumRMz3Rh1HEdB0ycxuGYa5tOe+lQ+8YlPKH5WAkTuKN/97nftAfe+P9QJQoC6gm6Oeajb7GhoL6QEPnFTNaTOC/7xvW/ha9//NjY3Q+UjyRyOhLfmqWecAEksJ0CgeXIZJzW2tm7Ltm6vVatWrVq1atWqVavWPXvd3OZxht1GQDu+fe4D/dGQKnP0Ol3yfsXhex3A2/7qdexFTt7GSR5PXVXkWc5gOKLodajNyJyjrEq6eUFtiVhV3P+BD+TSiy9RDK0EiOxsJ338I/Z7T3wyM8VMkxJNCbyDzBHbBIgHXDQChvNG31UMyfjouV/i3z92ErbvPFduuhk/223G49py1pTpUjO3/R2XRURERERE7miprsi7HUpnVMMR+3fm6S6UHH3ofXjNC17OHI4enmAOh2M4GNGd6bBxqc/s7MwWgXKyxLe//W2OevivKoZexdSWYRV69vOea8956jOYmZlZrvnKHOSO5I061gSMDMi8w3nPkosMyTj7yu/z3o99iDRTcHN/gazbIWRZs0XGLZeMTSc8lPwQEREREZFVJQvElJgh0EmOTZs2URaeU797Fh8687NUeBIwGgzBoDvTYWGxz9zsDKmKeMBSs3ogc5773ec+/O8nP6HoaBULOgSry0Me9GD7wsmnUI1KMJoGp5ZwmcccjKoK7z2Z87hkYBC9sYTnZ/Em/uHNb+SGaok4W1Bmjn6qKbodUow4INjKRkLO/fzTb0VERERERO4o5tpWhga+jMx2eiz1+8yunccyz6WXXsqvHHwX7rLuzszmHRxQ1ZFur9PGQB4sEXwgxUhZV+QhI8sz7n2vI7j62mte/Z3zvvMaHenVRxUgq8xnPva/YI6i6BBCwBzEzNO3mgoIeUbum5FPqa7BOyrgOpb45/e9kxvSgH4OVRHYVA3prZlj4+LCljOyTSeJiIiIiIisTmVZUuQ5qY64ZOy1zzpu6S+w0dfc7Ere9L7/5JpqIyOa7S15HhgNh82bwh6882CQ6ki36IBBcB6S8a8nvokHPvCBep94FVJsu4q8953vsgMOORSquinLcI6KRI1hrpn04tuBthYTyTuigz7Gf33mA3zjku8x6nj82lluWtxI0esSLRHybEX3ZBERERERkdWsCBlVVRE6BRv7i9Q08dGIRJwpuLZe4nX/9e8sAIvUGNDtdEl1nJTADwcD8qKYfL5pw0ac98zPzfORD31YB3kVUgJklXjR8S+0444/vvnE2n0qDspY4/BkODISrq4g1lSxJGWeBSJf+ME3+dyZp1N3M6oicPPCRvbeb182LS2S+UAesubH2fJluiIk6fCLiIiIiMgq4QyKPCfGSMw9Q28s9vt08oL5mVkWRgPC/ntx5uUX8y+f+k9GPmNgJZDwwVGPBtRVRbfXA6CuKurRiDVr1zLq96mrinvc85586AMf1FvIq4x6gKwSX/nK6a/udDpQllAU4KFKNSHk1NQUuGaMUxkhyyALDIBvX/cj3vjf72CxC2m2w8Kwz/y6vVhcXKTX69FfWiLDE9rKEQ+TJiDjmdsiIiIiIiKrST0a0ZubZbEc0p2doZMXDPp9vPN479lYDnFZ4Pprr+POB+7P4fseglUjspDhQ4YZeN+MyM3yHJ9lAGRZhg9NGH3ooYdy9VVXv/p7F1ygfiCrhEb4rALr16+3vdbu1XxiTKo/IhBJTVdiAwYjyDsQK/rdwGVpA69+91s568c/wOa71B6iaxMb7SM/3vYS2nXc+2Oc/BhXgmgSjIiIiIiIrJ5I14gO6nbPQxFdM/DBmphooRqydn4NftOA/aqM/zjh1Ry59lBm8TB+U9k76romhObtYksJ5z2kBL75wRs3bGCvdXsrrl4ltAVmF/fmN7/Z9tprr6k/5OZiUw9gAIgGRQeqCrqBPokPnPIpvv2zH8K6uckfvqdJdoTUXozJE8GtJT9ERERERERWE5uKZzbvdeiBvfbai4Vhn2HHM5wveNW/v4kFYNOwD3kgWmreew6+/XmG2fhd4+VAaX5+nuuvv15vF68S2gKzC/vd3/1d+7d/+zdijHjvm8RH+7eWJg+gNWU85kjDAW4mp+/gs9/7Gu/+/P9yYxEpO75pcsokfzKZZz3+fPJEMb3tZeobyoWIiIiIiMhqYA6ic5hzBGsuziZzJEjAaDAgz/LmjWLvKMuSq6+9gsc+8ChSrPBZ0V7f4ZwjpQSumQ4TY8SbkWLEZxmzc7N47199xhlnaCvMLk4JkF3UgQceaCeffDKzs7OklPDeY7SVGe11mgSGLWcoOp6+gx+sv5JXvePNLM3l9HuefqooCJPMp1vuodokNm6l58eK64iIiIiIiKwGDlJbAZ+lJugdx1CpDXB6eUFd17jgGZUjvPdcfeWV7LP3Wu55l7u3nQccySKu7RsC4HyzFcZnGS4EYl3jg+fXfu3XuPQnP371Dy68UEmQXZi2wOyiPvCBD7DffvvR7/fJsozNa6oC7ZaV9hu1iyzi+Gl5M69655vpz2bcNFokLwoCjuSW+39sfoGtNzz1UxcREREREZHVYrzF37PZVn9nJGd4wEYlXecpvKd2CVvT5S3/exLn3PxjRjTxUWmR2H6Mb4KnkGfUVUU5GhGyjNFoRFEUvPnNb9aB38Uptt0F/f3f/709+tGPJsZIrx29NP2AeVhOYbZriWMJ+K/PfYwfL9zIzTYiWzPLhvXrmS26GBB984e7tUTIuOfHZJ9c+0Th0/ITh4iIiIiIyGow7nW4uXG8s9hfYqbbI6sSRYQ1c/Ms1COudgP+6X3v4Pp6AwnIfbMVxoBoNgnDsjwnhECKkU7RIcbIfvvtx7nfOU/9QHbl80KHYNdy1FFH2Tve8Q5CCE23YeeoqmoyagmmKj8MwIgeFoic9I3P8Lnzz2RTB5aCkTx0gselqaambnlLy/hHpM32uUz6hNjyJXnQX7KIiIiIiOzqHM2UTMdmwx1cE9s4oDczg4tGVkZsVLFwy3pm917LhhBZv7jA/nS5650Popd3J6GSc4AZlhKxqsnynLIscUDIMpIZ69atY8PGDa/+9jnf1laYXfTckF3IpZdeane/+90nfT/quiZrZ05PjJuBAIREH7ioup6X/MtruDwtspQlylizZn6eeqEPQMwCyS33//BtSdikqepUh+TpUrGQmu8lv+UWGRERERERkV3RuPqjniplb+Kfqbd1y5o1WQcfDbxjQxxR7z1DtlRxl+tqXnv8yzjqvg8jI9IlIx+HY6mJjMyMEAIYkxG54/Ue97wHl156qeLtXYwekF3Ixz72MTv22GNJKZHn+eTrqY5YjISiaP5ORyOKogMONqUBS6HgWW94Kdf6ETemITU1uQ/kNDOqU0ok71YkObZ5YrTX8SwnPjQWV0REREREVkWga9uOYdxm22SiaxImRYLe+pJ7rL0T//7K17GOnH3p4Az6CwvMrJkn0gypWNGb0Zaj7J/97DIOu9tdFUHtaueFDsGu4WUve5m94hWv4MADD9yy+iPZZNb00uIis/NzAIww+hj/9NF38PkfnseNWc0CNXhH17kVCRBzeqhFRERERES2JRh0yJmr4J5r7sR/vOyfsMX1HDC3NwB1XUNWkGjeMJ4kQWCSBBmNhrz1bf/OCSecoEBsF6IHYxfx05/+1A477DCAW9n+ksCaREgEFushKevy5YvP5p/e83Y2dCILHRjljhQcwXkyc4RouJhwSoCIiIiIiIhskznwWcFowwLrYsYzHvWb/MUTnwf1EuuyWTIMIzSNUTcLrgNtOwGgqiqe+MQnctpppykY20VouMcu4Mwzz5wkP8qynHx9uvdHSok6Nn9eZawg6/LDhat45/9+kNjLscyT+UDuA5kPzWSXlCClLcq/tGrVqlWrVq1atWrVqlXr1tcELJZD8r3m8PvM87EzvsjXr/4ePptlkEY4HMEgaxMe40RImloN8N7zgQ98QAHvLiTTIbhjveAFL7BHPOIRkySH9x7vV+alYoyE0HzdUiJ6GJH4709+mKuWbmZUNN/r4MCMMrZ/cmZkCTyOyhmVbye6OK1atWrVqlWrVq1atWrVuvXVkecFyTluWlxg77kOb/2fd3PE376Ow/w8FiPOmpgtC023gnqqwmBcATIYDLjTne7Epz/9afud3/kdVYHsAvQg3IEe/vCH29lnnz3Z6mJmk60qMUacc5NkiLWtSGuMAY7/Ov3j/PeXPs1oTYdbhkt0s5zkoPZG5ZrRuM5gNnk8joFPRD3aIiIiIiIit8kchKxDv98nd4E1Lmdmqeax9zqSf3ruy9gLcLVvyzyaS3RMmqICxM2meR577LF86lOfUkR2B9MDcAf68pe/bI997GMnn48rQMysrfoIOOfakqpERaTC8Z3rL+UVb3k9G7qwMYvQybGqwqzZ62JACg5n0EnNWnnDdMhFRERERES2qTLXjLgF8jKxtyvIb17ipU99Lr9/1BOZNb/cAMSxXPbRGr+5HWOkrmsWFxc5+uijueSSSxSD34G0BeYO8m//9m/22Mc+dkXVxziB4ZybZAvHI2xHREoc17GJt330JBZzYzEYA5fIAiRzBHOQjIAj4DCM2hnJEk0diIiIiIiIiNwWA7pFh6VBn/nZOZb6m8hIHLD/3px08id58AOP5B6zBzKTTY3AnbptAnwb44UQCCHQ6XR4/etfz7HHHqsDfAdSTHwH6ff7lmUZeZ5vNu1lSzWwSE0fx9s//34+e/ZXud6WWPCRmX32YlN/CZ8FnEGRmr4fIbW39ct72URERERERGTbUruZJfPNW8muinRr2Cvrcqcww7/9zWu5M3N0iRSxSXSMYk0KHocnB/zUm939fp+ZmRle8pKX8Na3vlVx+B1EB/4OsHHjRluzZg3AZKvLbamBAYnPnns67/3ip/jJhhvwa2aoChjUJaNY47sFDujUTRKkiE3mcZQ1+9E07kdERERERGQ7mKNKhrdmoERwDosJqyNdn7HGFzzrcU/h945+HAcxRw+wOmLBUTlPItHBE4DRaESn0wFgcXGRXq9HlmWKw+8g2gKzk73qVa+aJD82bdpEt9vdZgLEATdeew3f+MJXWLrmZvab71KPgNqz0I/MrFnD4sIQXGoSIBHy2DTvGWVQBh13ERERERGR7WMEn+MNMlzTywOjNk+MhtU1p556KvdcdwAH3/vhOAPnmiYgBgzqCp81SY9OpzNJgszMzOC954ILLrD73//+SoLcAXTQd6LHPe5xduqpp1LX9Ypxt9tTBXLLhvWcef65xL1niJ2Mxf4SBCjLkjV7raVflwBkCfKYyFOz9WUUoA7N3jRzWrVq1apVq1atWrVq1ap1W6ulJlTOcJhz4B21M2oMn4xqQ5+H3+O+3GufA5lJHuq250ARMLc8BaaqqhVtD5aWlpidneXFL34xb3vb2xSPKwGy+/rqV79qv/7rv77iaxs3bmTt2rXbvG2qjYiRMk+NTR46IxFJeDIg4WlGL40f2JpmK0w+9YCbVq1atWrVqlWrVq1atWrd6hqBksR4xmazNhUeqa3zyAl0gA7tDQyoarAE3YIYa0KeTSZ9ppRIKZFlGYuLi5gZj3rUozj//PMVkysBsvt597vfbX/4h384qfRIKTEajej1epM/itsSE3gPVRUxM4piaveSGaT2z9MBpHZtH2RrHminRqgiIiIiIiK3yRzgPYkmrxExXBs6T6bdGlSjkm6n29xofOWyhOChyBiVzdaX6cmf0z1BTj75ZJ785CcrJt+JdLB3goc+9KF2+umnMzs7S13XOLc8U7qua2KMkz+CrYnAsH2wfB3JcGSZZzhqyqmCn/qD880fbNXeNmc5mykiIiIiIiLbEyQ3YzUnCRAzgvNNcJYS+Ax8M3Citkhd18zkneWxuK65lGVJURSTrTAA69evpygKZmdneeUrX8k//dM/KS7fqY+t7FBXXHGFHXzwwSuSHcPhkKIotln5AcsJkEhiHj+p5LD2EXTjYdPGZP9L2d62aK9obvwnLCIiIiIiIrcmAMS6KcEfR8w2FVQlAx9IdYl1CyoSVarp+YIQE845kgfD4YHBcMBMt6n8dwZu3Auyrtm0aRPr9t1HcflOoikwO9irXvUqO+igg8AgCxlZ1hzybre78oq3VqLhEgHPDMBU8gOWS68mn7jln1W4lT/XWfuHLCIiIiIiIrfNZ1vGaY7lPTAOfKcAg+A8XV+0t2uSJonmjWwHFN1eU6zvffPzDKyqCFnO2tk5Lrv0x3bY4XdXEmQn0EHegR784Afbueeeu/IPx23972irCRDXNDltPvTaxyIiIiIiIrIKouxIc/FTMZ9nqi+jAaMSigIc/PaTf4vPfv4Uxec7/qGRHeVHP/qRdTodDjn4kOXkRZvHGO9YWfGHsKKaI212Pa8DKiIiIiIisosLsOLNa2sTIkzFf5YS3nkWblnP/N57k+qK+z7wAVx88cWK0XcgRdU7yBvf+EY7/PDDOeSQQ6irasX3xkmNxG0XdYzLprZ1PREREREREdlFWBvI1c3Hbiq+q2kGVpQpgoP5dXs3PRzrmhNPPFHHbgdTdmkHueWWW8x7z9q1ayf7vADMLyc1YHkLWZhkONLkUanbi6NpZqoHS0REREREZBeXgNovB3wB6rYKZGpADHU5YqbogEEdm4EZT//9Z/HZT35Kod8OogO7A1x++eV2yCGHTD7fuHEja9esbXt6rKzo2GIv2FQCZJw4dGxZRiUiIiIiIiK7IAPSVALEN29sT8eBRoKU6PiMmCKZb0ZWLA36zM3MKk7fQTQF5nb2jGc8ww455BD6/T4zMzPUdc2atWsne77a838rjU/TZn8wzXUmD9DkRiIiIiIiIrLLchB92mLwRZgK96IZwWdEjDolgg84IHeBz376M/aU3/ltRX875qGR28vd7353+8EPfoCZ0el0MDOccwxGQ4pOM/Z280qO6JpsYEmihhWJksByAmT6eyIiIiIiIrJrSu1l3Mogb9fp6gPDNztlLBKcJ+AgJlxMkGU88pFH841vnql4/XamA3o7+sQnPmHHHnssdV0TQpPfc85NEiGTXiB1DXlGP5aMsoxPnnMqb/vfD7BUQD9vkiKeRDDIYtMfxJn61YqIiIiIiOzyQbZB1znixiV+82FH89d/8GLm8eQx4glQltAr2pYHTZwX2tuNeyDcdOP17HfAnRWv3860BeZ28qIXvcge8pCHNAc1aw5rjHGSCOkvLjEzO4tVFa7IMQeWFfxw4+W857MfZ30PNnabBEjtm5M/S9CJENrdMeaar2vVqlWrVq1atWrVqlWr1l1z9SRCbazdb46vXfJdHn3FBfzqIfdmXSigjFAUTcsDfysVCQb77rcf//zP/2x/9Vd/pSTI7Zmc0iG4fXz/+9+3+973vgCklPDer0iATJrdpITznvWjBejM8I/vewunX3w+G3uOhQ6MskR0TfKjiNCtm9tVASoP3iA5rVq1atWqVatWrVq1atW6q64uC6TFPgdYl4Pzed71yjewpoS987nl/TG+2QozCczHEzNovjcYDjjqqKM4//zzFbffToIOwS/vXe96lz3+8Y/He09ZlpjZii0wY8OqxGcZIyBmOV+84Bu877P/S7bvWhaoKQOYM3xb/ZEnyGPzRxT9ckbRoVWrVq1atWrVqlWrVq1ad8U1ORj55o3vImTcct0NdHzG/e55bwqXk2K1HC/icG4rZQoOzIwHPOABvOc973mNom5tgdklPOQhD7GnPe1pkxPYe0+WZZgtdzqtqgpLNUWnQ5+KRRLX1Bv5ny98GvZbwzWDjcSZDGszhrC81n75DylLKinTqlWrVq1atWrVqlWr1l15BajqRJblDEeRvfdfx4dO/Sz3u+99+LWD7kc3d2Tjnh/jyzjp4ZZzIKPRiKOOOoo/+ZM/sXe+852qAtEWmDveRRddZEccccQk4TGu+Egp4ZzDOUeqI96DecctVPTJeP3H3s4Xzz+Lem2PRauILk3+YIKNG5+ubKQjIiIiIiIiu7bkIeWefr/P3jNzFGXCNizx4Lveize86JXsQ8EcnsJoSv2hGQPqm+mfCcgwfBuu33DDDTz84Q/nZz/7meL3X5IqQH4Jz3rWs+zwww8npWaj1njri5lhZnjfnMw+BIiRKlW4LOfCWy7lqz/4DvW6Wa7YcCPz+67DqpKQmiqPsNmY3OlMooiIiIiIiOy6PE0/R7xnMBjgii7MFnz/qp/ywa9+mj961FPp4inG6Yy2CsRYToA4s6YoxDn2339/Xve61/Gc5zxHB/eXpAzSL6Hf71snL5pEh3ckDMORLBGcazJ2sd234h0LqeQ6N+JPX/dKLh2t55ZQ09lnLQvDfjMXum18Ok52JN80PoXlLTEiIiIiIiKyCwfZBrlz4BIxNS0R5ubmcP2SzlLNh1//H9yVtawBQvRN5iM0b3yP2p+RA1bXhBAoy5JOp8OjH/1ovvrVryqG/2UeGx2CX8z7P/wR+/2nPwPqCp/nYEb0jgqoqYllxdqiB4t96HboZ4k+jref8n4+evoXGM4XLPhI3QlUKRJwhLS89cVcU/0RpxIgqgIRERERERHZ9ZlrKjgCDvOOMiVCgr1CjwP8DB864c2swxHKim7eoU6GC66dCdME6sP+gG6vR4oRM+OSSy7hvve/n2L4X4LXIfj5HX7Eve2Zz3wG3oPPc8qlPqPRaOqgZnSLHlWsmxnPwTHCcfbVF3LGuWdTZ47QyVka9OlkOZnzeNv6g6HKDxERERERkdXDXNP/MbqVMV2dIot1yS1ln09951T6QCyySePTQPumd4RY1XS7XVKM+BAIWca97nUvXv/61ytC/CUoAfIL+PIXv0QAhsMSgGJmhk63izcjYAQgUVHGEgJsrAasZ8RHv/g5rl5cT5k5lmJFb+08N2/cQMiWW7GMqzxcmxBxpuoPERERERGR1SRhmGtG4gIU5sjNEeuSxarPp07/Epf0r6YmtE1PHURoy0aaGNG5Sb/JWNeELON5z3seBx54oKLDX5ASID+nF73oz+zAAw+krhPdbkE9GlEOBwC4mMhpRhpZHcmLLmXwUPT4wne/zvk/+xF+zQxVL2OhHpH3utTOiJZWPihTSRARERERERFZXcaDMcwMklGYp+MzsiyjDo5Lb7qGT5z+RfoYI2JzowQxNkGg0axZnmMpTd40P/DAA/mXf/kXHeBfkBIgP4e73e1w+4+3v41Bv0+eNYcuKwqKXg+qGlyAmAixppMVJGAIXM0mPvTFzzLoem6q+sROBt2cW5Y2sc9++1KWpQ6uiIiIiIjIbsI5h3OuSYLEhE9Ggcd7T8o81XyHL51/Nl//yXlEQnMjM7xvpsCklLC2+sM5Bw6SJWKMPPvZz+ZJT3qS3i7/BSgB8nM47rjjqMqa+blZqlGTtEgYsa4hyybji6ibE3VTGrAE/L8PvpfLN9zIkk/EItCvSyzzhBAYjUbNCS0iIiIiIiK7hXECZJwESXVskhpmlCRSN+f64QKf+soX2cCwaQCC4ZwjxRrvPc57sCbPkVKT/IixqRZ585vfrIP8Cwg6BNvn6F9/pL33ve8hBA9tCVJMTUMaHzxYgsEAQgbBM/KR2hVc0r+Kt3/s/aS1XTZWQ3pr5ulXI5IZMzMzbNqwkZlOF2fWTIChbZrTXsZ7xpQiERERERERWQVcs4XFO09mQDKCOfCO2kPljMpBHgI3X3c9a3o97nrwwXSDxzuwWOOcx3mHAc43iZQQwmQs7v77789NN9306m9/+9uv0QH/uR4a2R5fP+ub9rCHPIQi5Aw2LdBbMw8OyroiOE8wA+eb6o/cs+SNGxjxwjf8LVeWm9gYKobBwLkmsZES3nsyH4h1TWZuMgIXILWlTxqDKyIiIiIisnqYgyrVFCEjdx6rIj6BD4E6wIjEMCbmsoJw8xKHdffiXf/fG/iVsBczgDcHeJI1MeNwOKTb7VLXNVmWkVLCOceGDRtYt26dYvqfg7bAbIcXvPB4+7Vf/TXyLAegNz/fntmQZwEfHATfpJMyj3moMT5/zlf42c3XUOZNB+BAk+TIEhR4sgTUcZL8mP6DSe063lUjIiIiIiIiuz5nkPsA7dYXcxCDY+gSVTsAo0gJZ4nuPvPc5Ia8+cPvZhOJYV2B0Yy/9U243u12Kcty8rlZs1Vm77335n3ve5/CxZ/nsdEh2Lbrbrje9t1v/8lY2mnm0ooDOapKLC+4wjbwh3/3Mso7r+HSG66mOzdzq38c43U8/9kcxHb7y/jXBVWAiIiIiIiIrJJI2yZV/E185ybtDUKC3HlSahIeRYRDZvfmT5/8DH7/3o9mFgCPmVHXNXmeT37suPqjLEvyPCfGyGGHHcbVV1+t2H47qAJkG974xjfa/vvtv5XvJHBpkvgwYClV1HlgkcR7P/VRNsYRV910Pb35WVx78vu08uI2S2yMqz9W/O3oYRAREREREVk9gbY5wGHOkXCTN7jHQXiWgCqSz3SpO4GrBxv5zDe+wjXVLdTQDMZ1Dp8Fyrqa/NyqqnDO0el08N6T5zlnn322Dvj2Pi46BLfuAQ94gL385S/HTR+o6WzEVOKiX43A52yi5qyrvs8Xv/UN4mxBygO0pUrjRIe35cvkR7mVW182/3Wq/hAREREREVldxn0dxzGeb98UT1VNjJG5uTlqD9Yr+OG1V/Dx0z5Pn0RF3VzfeZxzVFWTBMmyDGAyDWYwGHDQQQdx1K89QhHjdlAC5Db8/d//PSGEqVM3YW3lx3TphwPIHSWJJeCdH30/w15gUyopZrrNthi3fNInt3wx1zQ8je22l+nkR0jLFxEREREREVldol9+39wZBDM8hmt7R45GQ8pyhJEYWMUZF53HpcPrqYHhaAhAFjJc8BgQQqCua0II9Pt9et0eDsfpp5+ug70dlAC5Fccee6wde+yx1HU9ya6N2VY+cWQsUXLGJWfzo+uvYoGazl7zrF/YRAihyfy5lSNuN6/6mE6MOGsenGDNRQ+UiIiIiIjI6pHacDH6cfKjqf4AsODJugVLS0vMdLrUoxILniv7G3jnxz9IwtPtdImpiUW995RVCTRVIGVZMjMzA2aUoxHOOf7+716pKpBtUFx9K8bVH957QgiTgo/Elj06mq8bN4028q4P/A9rDroTi1axqRoyOzuLmRF9c+KPKz1WXNqvbz7xxZkqQERERERERFYjc8vbXxzLb24D1CSSd2TO08sLcnOYd2yk4ivnf4vvX35xE7B7j2E43Io35p1zpJSaPiHek+U5r3zlK3XQt0EJkK146UtfakccccTyCWfNWTruBTLdD8Q8DIEBjs+ccSo3V32uWn8jea/LqB1VtLyNZsvqj81treFpUhdUERERERGRVcm1/R/HE2ESUFtiVFdkWcaoP8A5hwuekUtke83xnx95PzdWS5Q4KprbdPICDOqyIs+aCTDD0ZCsyMFBVuR85EMfVhXIbT0WOgRbuvbaa22fffaZjBuKMVLXNZ1OB4uJOkXyPFADfSJDAt+84ULe+K63ceVoI2m2Q0kkFDkbbr6FvdaupW6zdVs74HYrD8x0k1QlQURERERERFaXOB59a1vGdzYVH46/lxzMlI51m4y/fO4f84QHPApHyVoKCmj31Rh4R1VXZHnetFtIkdwHhoMB97vPffnJZT9VBLkVmQ7BSq997Wvtzne+MwB1XZNlGSEEQghNJUhK5FkOJCoiRqACTvnSF1m/uInOTMGgrLFqRM/n+GKGvAYfjV9mJ0vQQyMiIiIiIrKqhF/g684gn+3x/k98jF99wEPZlxkSTflI7A8IvV4TzGcZZaxxweN9oIw1vV6Pz3/+89zjXvfUwd8KJUCmPOQhD7Hjjz9+8rn3yzuEDEgpEdqxQ6msKIqcmsR3vn8OPzz3u/SCMdfJyGOiLD1dF5lLgbqsiVkgasORiIiIiIiI3IZOhI45brjmGj79yf/lj499LqmuIBRN8sOMWEd8nhFCRrQIDvIQGCz1Ofwe9+CYY46x008/XVUgm1ECZMqLX/xi7nSnOzXNZFiZAFnxuYFvi5XiaMjCdTfxyCMfSl0E+i5CFnDOEauKYE3lh+XZVnt+iIiIiIiIiEzizuTxg8h+D1zDws3rgZrgPKQEePCO4DISNmmnEGNNCFnTtqGuefu/v43HPO6xdu211yoKnaKD0Tr66KPtlFNOYX5+fvI1MyOl1DSk8Z6YIj6B9wEsYXVFzD0bywF5d67tCTIka3Zn4TEyXFPGpH6zIiIiIiIisg0GVG382AGKdq2WBuS9HqTmTfdRrKlTpJN3qOuKbpavaDD5guNfwHve8x7F/FN0MFqf+cxn7ClPecqWJ5/ZpFNvwBHLplMvZhAj5AFz0E8VzuckIoYR8ECiwANGMCVARERERERE5LbVDvoYg2rE3vksxJoOHmKCLAMHw9GQvNuljBV5yAlArGusqpthHiFw2c8u4653vati/ik6GMBTnvIU+8xnPgM0fT7quqYoisn3Daip8Xi8GdQR5wK4ZqDzYNinM9PDcMudfC01ZUrRmp8QMqIOtYiIiIiIiNwGg8kAjTga0iu6kKyJP4GyHBE6OThPbZEYI52QNW/CtxUgZgkXPG95y1t46Utfqri/pQMBXHDBBXbEEUfgvV/R9yPGSEqpHS3U/JeZwzkH0UhVhS+K5ig66Pf7FEXRVoi0Z21K4D04R1IRiIiIiIiIiGwHTzMrN41KfFGQyhKfBQjtLoTRkE6n24SjKeEMXDIwI3mHzwJXX301d7nLXRT3t/b4A3Hsscfaxz72MUK49UGzTS6jbYw6fdBs/JXNjqRNraYjLSIiIiIiIrdTBO9WVolAO1Z3HH+65evWdc373vc+jj/+eEWkCsthw4YNtnbt2tu8zuYJkJUHbuWo3JXJkc1+iIiIiIiIiMi22Fai9ql1nABZEZ9O3cZSYlSVeO/JsoxDDz2Uq666ao+P//foMbjPe97ztpn8GJ9M/jbOx7HpE3CL1JLybSIiIiIiIrId8SeJLao5thaj3uob8N7R7XYnn5500kkcc8wxOrZ78p2/6aabbJ999qEsyxVNT7curTinlsuN/FY+Wv6+Cj9ERERERETk5wnSw+bbWW4terdbj/DNDOccVVWR5zkPfehDOffcc/foHMAeWwHyxje+0fbZZx9SStuR/Fh5Lk2fY/5WzsdmCO6ypL9jERERERER2QY/HVy6lTHorVZ8tN+ctKNMCe89MUbyPGcwGPDa176WJzzhCXv0sd1jsz8333yzzc7O0ul0GI1GdDqdbdxiZQpjec+VX7nfaipLZ25rtxQRERERERG5df5WIlEPuFvZZmBu+brT22PquibLMsyM3/zN3+TUU0/dY/MAe2QFyJvf/GZbt24do9EIYNvJj601oJk+qaanvdiWVw1u/HWlQkREREREROQ2ODA8ccsvbz1OvZV0xsLCAkVRrHjT/7WvfS2nnnrqnnxo9zwLGzfZzNwsznuGoyGdTmdcxzE5j6YzZ5FECUSWcx3T32/qQG67B4gjbdFIVURERERERGRanLr4SXOF5dWRyPDk7RpIuPb7hicBDsO1MW5MkeDDJOI95phjOOOMM/bIXMAeVwHyP//13zY3Ow+uOaFcp6DGEWLEWbttJQQ2lX1mihnqasCCq3n75z/OUtGkNdIWp4rHb7Pbqao/REREREREZNuib7a6eGsSHtF5kmvWYImicqzLOzz00CM46h73o7BEAOqqJC+KSSLEMMy5SQplNBzykQ99mDsdcOc98rjuUQmQOx9wkD33uc8Dg9HSkGKui8MzSiU9FwghALA4HNDrzlBhuLzDuz/yfj59wdfY1NUfooiIiIiIiOxY44SFT+CtSYjUHirfTIjJRpH5En54wMU8+B73JzgIgPfNrZMBzjdVIFO9KTudDmvWrOFxjz7GvnzG6XtcFcgelQB56UtfClnTGjd4jzNwLtHxBYm2xCgmenmBkehTcU1/PV8+/2yGhVHr71BERERERER2gkTTTzK2zU1j+zUDer0OkZqLrryM0y/+Fo854qHUJHpZG+Inwzfv7ze3s4R3Hucc3ZkeH/zwh9n/znfa447pHpMAuf+DHmwveclLGCz26c3MkHWK5syJETKPAbVFQkxkeU5NxHB87Ctf4JZQMwxKgIiIiIiIiMjOE6fG4CbArPl4oRwy4wMhh/ef/EmOPOIBzOMpgFiOyLPepOeHB5IZztmkM8h+d9qf33/ms+xDH/nwHlUFssckQP7+b/+ObrfAdQvisG62uzhHEXIqS1QukbuMLHdAooo1P7rhcj5/7pncUiTq3JPUxVRERERERER2sPGo2837TzbRKiTv2DQoyeZnueTma/juNZdy9IH3owa6Rae5oS3fJrhmhqlhxJSIMfJXf/PXfOgjH96zjuuecCfvf//729lnn0Oe52SZb06EKoLzEBx1qkltqVBBYlSVxLzgtR95JyedfwbDvbrEWOmvUERERERERHYam4rY3dTgDVdk1KOSLEG2VHLvvQ7gxJe/ksOKvdmLDJ8yMKNpBeIn0b+xvB1mOBpywgkn8La3/vseUwWyR1SAvOENb6Bb5Djv6ff7zMzMQGgTIWZkIcNIDOuSynksLzj/2kv52oXfIc51SN0cG8XtmPQiIiIiIiIicvsYV4AEW06AmIPBcEhnbgaHo5iZ4bLrr+fbl1zIXe//GEqr6foc0tYrHrzzlLGm6HR50Z/9GW9767/vMccz7Al38r/+812vzosCq2uKboeEgQPnx6eDw1kiCxm19/SBt37qJL57/eUM1xQslkM8KzNuIiIiIiIiIjtSGkesBljTz8MMyAO1g2GsGA6HrJud5+qfXMZvHf0Yui6QEdoJMK6pBDHDUvPTnHM476lSzdq99mJxaenV3zrr7NfsCcdzty91Ofnkk+1JT3hic1dds8Tm3Gn2QpGaLwB4Y5NznHndxbz8rf/EcK8uN9UDfCfHVZGQ9AcoIiIiIiIiO565qQqQtDwWN3kYeiPlgTLVdMwzu1BxcJjj+Y/+Lf7gUb9Nl4wcqMqSIi8YDYfkeY4PgbIq8UU+SQdcd8N1HHynA/aIbTC7dVvPe93rXvakJz2J1Ga6mKrgiLA81cUBZUVtiSHwoc9/hltcxfo0wmeBwaZFbX8RERERERGRnRuwW3thsx0JqanqyLKMLMsoZnpcffMNfO0753ALfTbZiApweUYdazrdLj4EMKMoCkKb/KiI7L333vz6ox+1R0S8u3UPkL/6q7/CzNoHevnrTfVQmnzsHNDJiR6+d9MPOfeyi3Dr5hjUQ+ZDh6I3S0oJDYERERERERGRHWl644Gz5cqPad57LBkhzxktLDETunTXzPHD667kE6d/gecd8wwqIHOeZHUT+BtQR8gzUoq4EMgJkAc+d8oprJmZ3e2P7W4b09/vfvez5z//+ZRlOXX2bP3kMgfRwxKR933qYwyCMYhV8/W6xpkp+SEiIiIiIiI7xXj6y631ogzeE8uKLBreoCxLZvfZi7KX8bEvn8J6BpRASSTPi6ZxiBmEDGIi+IDDSCQS0OvN8Mzff9ZuXwWy21aA/Pmf/zkAnU6nLfNIk7E/jpWZnwoYkTj36os595ILGa7JqEaRmdkeoUqU5Yg8z1ELEBEREREREdmRVvT+2CwlkTzESXLEkYYla2fmoKy5aWEjwTtGKfK5b36ZZzziCfRwFIDFmtznbUDsIUaGdUnR7ZGATYub+L+vex0f+dCHd+tju1sWNhxwwAH27Gc/m7quGQwGk+an0/L2Ak0vkAUSn/jqqdQzOWWK7D2/BldFUh2ZmZmZVIqgVatWrVq1atWqVatWrVq17qA1tQmQce5jevxtbL9X1TVr5ueJo5LRUp+iKBimmo2pJKyb55NfPIUbRxuAjIpEnrfRb1U1P9gHZro9AG5afxNr5tZw0MF34enPfMZuXQWyW47Bfde73vXqI488khACeZ5jZs30H5b3U41LiZyDIXD65d/hvZ/7X0YdTwyO3AVG/QHDsqTodYkpkdpBMlq1atWqVatWrVq1atWqVeuOWpNvPs5sKgHim9YN4+t1QoavjdFSn5BlFLM9lnzEO8fouo0ctGYd9/6Ve5JhFLj25zSjcYebNhGKHLyj15shYXjnuP9978vb3/b23XYk7m5ZAfLkJz8Z7z1VVTUnitl0D9Qm69POwq1oEiCnfOvrLBUw8MbMmjk23HwL6+bXMj8721SRiIiIiIiIiOxgjmbyi7PlnpU2taPBGWRZxo033sjs7Cxr165ladAnYTjn6MeKbJ9ZPn3GqSwyBDylRSCB95AS3bVrGA6HBMBZwuGoUuReR9ybo3/9kbttFchuVwHyute9zo4++miyLCOEwMLCAp1ul0gCHFbXeDzEBMFxU1rkwqWreOtnPszijKfOHFVVMZPnxKoiOsMFD7hmJ42hVatWrVq1atWqVatWrVq17rB1MvqWdmqpo41o2yqGZHSLDlVdETGyPCPGpvoj4IjOMRj0WdfpcsQhhxFcwHnw3khE8L5pjpogxIQFj/eBxYVNPO6Yx/Dv//7vu2UViNvd7tBgMLBut8toNMJ732yBgba7bSI3D6Nm39PIJxY6OX/x7n/ijGt/xKZO8zOyBL3acAZVaPZZOXOIiIiIiIiI7OqKoku6YSOPOuie/MtL/o696eKpCdHohBzMNW/xRyBFrAjENkEwHPT51Yc9nAsvvHC3C4J3qy0wf/3Xf23dbpcYI0VRTBq9LGzahAfqwag9G3Lo5tDpcDOLXPSTSwEIqblAk12LynmIiIiIiIjIKlP1h3Rnevzgqsv4wnnfINFMjXG+SQG48YjUABYcLhlZu/Gl25vhP975jt3yuOxWCZCXv/zljEYjQgg45zAzUkrMz8/jgF6vB1UNHvqxpA/8v/f9J0uxBJoRQ8GacqPxPqukJIiIiIiIiIisoiA/xUjoFgw7ng9+7pP0qXGEdrrMuLlI21sk+HbETMJSJJI46hFH8dCHPtR2x2OzW/jLv/xL22effeh0mn0sMUZijO0EGEc9HDUZriwjWmKQwcUbr+DL556F9QpgOQESrOmuG9ujoxyIiIiIiIiIrBa9Xo+FpSXKXsaPb7mOr1/0bao2TvbjCDdFamu3QFiTHgg+4PGMyhEvfOELd7vjstskQI4//nhijIxGI8qyxDlHlmWYGdWoJOt0muoPB9F5IgUnn3kGbq9Z3EyTNMki+NQ0njG0BUZERERERERWHzOjtEgVoLvvXvz3Zz7OAhV5KEgxNu/yB48zaypC2kar7UKn6PBHf/RHPOhBD9qtqkB2iwTI8ccfb4cddhh5npNSoigKfLu3Kcsy8qJoynvMiClSAd+/5Sd89htfYTEkbhosAMvVH86arS/WbosSERERERERWQ0S0K9LspkutSVSJ+N7V/6E0y46BwOsjk0mwDsyF6BOJIsYCYsJYsRSIoTAM5/5zN3q2OwWCZCXvOQldLtdqqqi1+uRUlPGY9akL+qqohoNcZ2cGBxDjNPOPYuNVKTZgmymi9EkPvxUxkPJDxEREREREVlNzEHIMyx4RlXJUjlk7SEH8OnTv8QCJXQKqlhPgl4fApYFUvA4326DcR4z4yUvecludWxWfQLkyCOPtCOOOAJgMvVlXP3hXLOHJSty8l6T5BiQuDqu56vf/TaDDIYkBlW51ZNmcpCUCREREREREZFVIrY7GrIsw3nP+uESP7jqMs694ocsUuNCRqojqWxi4QiUxObGdbOWZUm32+X//b//t9tExKs+AfL2t799kvC4LQbcMlwkkvGJL53CDYsbcb2C0iJ4N9nyIiIiIiIiIrKaJYwYI3meU9YVvpNTdgP/e8aXGJHRjyOcc/iioO0CgicQ6xpCAKDT6VDXNb/3e7+32xyXVZ0AOfzww+3BD37wNhMgRpPRsm7BBkrOOP8clqiJmZ9MiTHXZMnGmTK33AdGREREREREZPVIhllz8b6Je2NwfPl75/D9xcvJQodYVeAhOUciEoBoaUUQ7L3noIMO4vjjj98tqkBWdQLkla98ZZO8sNt+LBLQ7HAq+Pg3TmF9PYRewbAckTlPcB5zTePT2I5A9gYh6e9GREREREREVhdnEJxnMBqRdzuMypLoYFg4/ufTH2cEVM7Gm17w0QAjLwoiiTrWk+RJXdf83//7f3eL47JqEyCHH364PfvZz56Mur0tBoxI3MQCn/ryF1ikxrwjxkjAEaxJkiS/XAUyrv5Q/w8RERERERFZLZxB7jzBOZxzlHXVtH0Ijs6aOb7xvfP4ydK1WKdLScRhdHyGq2LTC8QZIcuo66aMIMsy7nznO/Nbv/Vbqz46XrUJkOc///lNGU+M29wCk4CE57yLLuCm0RKbqiEViW63i6sTPjWPY2wrQNJUBYizVX6gREREREREZI8SEri6iXkX+31Ct4DgGbTTUT/y+c/Qp+n70d+0CAmC99QYjkCiGTJSVdXkZ/7FX/zFqj8uqzauf8YznkFRFIQQ6Pf7t3ldAxbo85WzvkGxZpY6c9Qp0ckLqCOe5R4gaWq/k1P1h4iIiIiIiKyyIN8BFhMe8FkgBkdpERebpMgZ3zuXS/vXADDr86ZqwPmmeWobQ0NT/TEcDgF43OMex6Me9Shb7cdm1fnT//Niu9td74ZVNcTEzMwMESb7l6YZUJO46NrL+fr3v8NC3eyBSilBSqSUVlSQ3FrSQ+1AREREREREZDUwi+ASw+GQ+dk54qiiLitm5uZYP1jkRjfgm5dcQJ8a1+tCSmCGx5Pa6Leua5xzdLtd+v0+zjme//znr+rjsiqHnJx7yYX24Lsd0e5tSdDJGLlETU2PgKfp7+FDU7pzIzV//u43cPHCtVy/aT0WPCFv5h47s6a0J8YVY3CdLWeHlPwQERERERGR1cScEV3T87LbtPNgFNq2D6OKe83diQ/+5eu4S+rQqYA8Y+AThidLRjAI7Ujc8fTUfr/P7Ozsqh2WuuoqQA65373s3ve8N3gPzkOWYXVNmUpqDGtzOqFNfixZybXDW/jpLddww2Ch2dPkHM6aB7FOibJt7uJs+QLj3iEiIiIiIiIiq4c5a9o7OMNhhNT0uYweag9ursvPbr6Wr11wLrXPoJNBXRPwxFQSfMA5R1VVk+THaDRiZmaGN73pTat2G8yqS4B8/IMfJuAYDPqTf70LgcIFuuQAlHVFrOtm+ouDkz75MW7ZtJG6rgkhTOYgO+cIoXlgRURERERERFY7m+ptOU581O3E0/GwjxyPOTj5a6cxABIGWUagGZ/rAe89KaVJvJxSUx7w3Oc+d9Uem1WVALnzgQfYQ+77AKwq6c3OgIdhfxGcI3OBABjNrGLvPRGoyPnmD75LDI48z8myjJQSMcZJAmRbU2REREREREREVpvQ7nCIvrk4a75WlSXduVl+eN2VfOOy81lKFbgmQdBxGTE1HTazLJv8rG63S13X7L333jztaU9blVUgqyryf+lLX8pwYZFOVjRNTx105+dIVUUAYlUSUyT4QDKjJPHRs05mo6+xzOOcw8xIbfNTYDJKV0RERERERGR3MO5vOW7vENupp6FNgPhoDGNF2cs46eRPYr6grEocEIDMN70/xlUg40qQLMvIsow3vOENq/K4rKoEyMte9jJ6M3Ng0B/0iSRwULcdazt5QeEzEjCyyHqG/NcnPsyCj1TWPGhmRgiBPM8nW2HGyRARERERERGR3YW5lRcAn5qemWWKbEolF19zOT/r3wBFQd32x8SY7JoYJ0Gm3e1ud+MRj3jEqqsCWTUJkJe+4uVmMUFVQVkx15shxkidIkW3Q1VVQDPWpooVlhV877IfsclH3HwPC26Lvh9mzeM1XdYjIiIiIiIisppNBnu45cv0151z5J2CjfWQNNfh/Sd/kiHggm8mgbSx81iWZdR1TYwRM2M4HK7KkbirJgHy1GOPpdPpQJGDdziaspxhXRGBvCggNlkpH3IGwNs//D+4NT0WqxEWmrs6rvio65qyLJtxueoBIiIiIiIiIrsBZ02zU29gLDc/dVP1GlWsqS2R97oMXOIzXz+Nn/VvwJwnjkaY2SROHo1GQJMEGRcTdLtdjj/++FV3bFZF5P+4xz3Ojv61o5pHzwF5Bga583SLLpPH0TuIiQScfeX3uXJpPdbJqSxOqj+AyYOZ5zkhBPUAERERERERkd1GsCZ0jr7Z+jJufmoOkm8qQKq6xuMY1CVpbY9vXPxdBiRCr4MLfhIndzodzGyyg2K8ArzmNa9ZVdtgVkUC5K9eccLyJ669AOPMRwLqFME5CJ4RcOp5Z7HJ1azvL1AUhf4CREREREREZLfnWa72sPYy/pq5dodLu6siz3NCkdMPxqe+9mXWM6SmqRCBlTsoJiF5W1gwHA551rOeteqOzS7vqKOOmiQ7xg/g5oIPGFACP9l0Dd+85ALSfBdzjuC0xUVERERERET2HH7c74OpniA0VSFZlpH7QD0cUZYl1iv4wZU/5ds/vJARQAiT7S7jRqjjxMe4Iar3nnvc4x48/elPXzVVILt8ZuCf//mfbWZmZsXXpvvPjsf0OGBpNGAIfO7M07lqw00sxpIsy6gGwxX7nURERERERER2d9O9P8y1W2IAK2u6WQ4xMaxKUu7J163hU1/7Mn0gsdwAdTxIBJg0QQUmOy1OOOGE1XM8dvV/4J/8yZ+0R33FMplj7GkSICSj6PRYIHHWJRcw6nhGsaKXF3RDrjNfREREREREdnuJ5ZG3RYQ8NbFz7Zd3U1SDIT7BTLdHURQMRiNcJ+e8n17CD5euoqbZ/jIajVb0zDQzQggr+oDc+973XjXHZpdOgPzGb/xGU/0xNX4Ha/7Rsb248SOMIwFn/OgcLttwA9lMl5mZmabzbR3RJhgRERERERHZE6S28WmWmou1CZDomxi61+mSyoq6rimKgpQSC4M+ZS/j46ee0sTazpFl2STZkVKafDydFOl0OrztbW9bFXsudum8wItf/OKm1GYq/0HbzbamzXtY+4HBxuECHz3lM9w0WmoqRGIklhXDpb62wIiIiIiIiMhuz1y7zcU11R9ZWwFS+eXKkG6W4w0GoyHREnkI4B2xk3Hm+edy4+AWgEkfEGh6fowTIOOvjUYjsixb3rmxi9tlEyBHHHGEPeUpT8E5x3A4XP6GW+5iu+JrLnFL1eeSa68gzM/QyXOqpQHdXo/ZtWsmD7SIiIiIiIjI7iy55dWakLlpH5GaviCLi4sUvS553rSLsJjI85xbFjex5CMXXP5jbh4tNVtmUjv+FvBZIFoi+ECq4+T2zuCpT33qLl92sMsmQI477jgAyrKk2+lONitVrqZqBt/iUgkukUJk0XtO+tJnGc4Fhr7GFvvsNzNHn5KFNCQqASIiIiIiIiK7OWfgzZFwjIKjDA5vkMemGsQZhF6HpWrUNNRMEe+hShWdmYK+q/i3D7+HYacgkfApkohUJJJ3TTbFDO8DJCNWNT4EXvm3f7fLH5tdNgFy7LHHAk1nWUsJ2j1GDk9tNTmewmfEGBmSuLx/A9+57IcsxpJut0uKkXI0oj8YQB4mGTARERERERGR3dm4BcR4eIizpvpj869vLtFslbkxDvjudT+in0pckWM4htUQAyIJcDAq8SEQsoxUVdz//vfnEUcftUtXgeySCZBnP/vZdvjhhy8/eM6Bbx6dALiYJv/wuq4xAmdffAGXXnMFuQ/0soI6Rix4vHMULkxmIIuIiIiIiIjI1hlQWeSMs89k6NKknUTmPI62RadjxSSYsq4JecYzn/6MXfq+7ZIJkBe+8IXUdb38ADjAe1KMTcfarMABVV0ROgUlcMZ3v80wg05REEcltTNCJ2em04U6qgmqiIiIiIiIyDYkBykPnPXd81hfDRiSSCR6WUEAQpP9wHU7bcBudGZ6jKqS33nKb+/S922XTIA85CEPIcsyUkrt8WxGv6SU2na2EOu6mfKC4ztXXsR5l/+IsG6eqq4Z9ge4PGNYVwQD6qizWERERERERGQbzEHljOsW1vP1C86jxJNSwhm4do0xTnqBJIwIjKqKQ3/lV3j+c567y5Yf7HIJkFe96lU2NzcHLI/WGa9ZljUdaA2CDxSdLiPg5G+czs31gLqXM6xKADq9LoNyRF1WFD7btef9ioiIiIiIiOwCzEHtIdtrjs+deTojIPgmFndtfB7ybHJ95xw1RndmFgyOf/4f7bL3bZfLC/zlX/4lAMPhEOccZjZJgDT/Ygd1DTgMuLpaz7k/uRhmOgxiRcgz8jwnuSZh4qwt0RERERERERGR25SAKnNUhefSG67mm5d/l/GeCueaFIIBbTakrQBJ7Y0TD33wQ3bZ+7ZLJUCOPPJIm5+fJ8ZIt9slxjhJggBUVQUO6tQc/qU04sIrf8qNdR8rAv3hgKzbITpjMBzS6XTw3lOWpXqAiIiIiIiIiGyDOUjecfPSAmmm4HNfPY0SJoNJLCViik0SxAwzIxDoVwNwnnx2ln95/Rt2yQh8l0qA/N3f/V1TThMCVVURQph6FJrSGgNCt8OoHJJ8xpv/+z/p51CnRJHlDGNFjZFlGbGsIKZm64yIiIiIiIiIbFN/OGB23VoWUsn3fvJDRsAwVYw3VzgfmpoP7wi+idtn8t6kKmS8s2NXs8skQA499FB7whOeMPk8z/PJJBjXHuUszxnWJRXgul0u23At6ylZiCUeCE17kBXzjFP7sWkXjIiIiIiIiMg25XlOnRL9VLNoFaec9xVSaAoOnN8yjeCmkwsOfJ7zoAc9aJerAtllEiC/8Ru/wdzcXNNNdmvaSS5ZljMiMQA+8/WvsOQidHJCgqK96XSyw9xyEkREREREREREblsIgTpFKm+MAnz81FPoA8NUM53VMJqeIYG2OMQ1lxQjJ5xwwi53v3aZBMhxxx23IvkxGo2Wt66YNdNfAMNRA5fHW/jiuWdS5x7vPd4gS2zR7tScqj9EREREREREtpeLiVRHfJaROhmXXHsFF99yGXneG7c7BZh8HNpLc+Omb+eznvWsXe5+7RIJkAc96EH2iEc8gsXFxUnfj+n+H3VdQzsCt7aaGs/ZP7mIa8tFLA9UoxF5arbAhMSk4WlyzQNijhUPkoiIiIiIiIhsydn4YrjgqZyR5jqc+u1vUsLyxBeYVIM4az9pCxCKogDguOOO26W2wewSCZA/+IM/AGDNmjVAk/DIsmxSEZLleXswHQFHBE777jmUazqUsaZDmDxIwZoqkPHWFyU/RERERERERLZf4ZoYm2QM6hK3ZoZvfu88bq43rkgjpOkPEkSaiwFLS0u86EUv2qXu1y6RAPnt3/5thsPhZORtSs1hdG66m2mCZGQucPmmqzj/J5cwyCHGSC8v8Aa+rQAJqa3+mEqCiIiIiIiIiMi2kwTBe0Kiic+9YylVXLPhZs75znlstaSjrQAxmgTIqK6YnZ3l4Q9/OIcddpjtSvftDnXgnQ+wu93tbnS7XVIz65a8KCjrCu89o9Go/Zd6BoMFRsDp3z6Lm0dL1Bjee8qyJDmInsk6uYNtVYiIiIiIiIiIbFuyJojOcHTynKXhgLoTOPuH32fYXifcSkLBgDzLJ5//xV/8xS5zv+7wBMhnP/1pqlHZJozcuHKGkOUY0Ol0xg8BJTUbSXzlwnPozXZJscJnjkE9pPaJyidGGdS+qfoICfLYXJQEEREREREREbltCaioSS5SADYakXUyNrmSM6/8IT+xWzASxERdDqhTCS5BLEmkyWCSsiwBOOaYY3aZ+3aHJ0Dufve7k7cNUoCtl9OQINZ05ue4nk387KbrWBr06WQ53nu6M73lfh9TW14cTeLDK/khIiIiIiIisk3moLQI42mr5gghkHLPTeUSp513NktpBGVJt+iQ+3aASZ7h8RgJD5PWFve///25y13usktE5XdoAuQpv/VkW7P33itm17qpf5iDSSdZisAI+NjJn2RpNCTkTUnNaDTC4yZJjuWOte2Dp/NXREREREREZLuZGT4LxBgxM5xrkiDDWHH6Wd8g+Q7WawoZrI7jWzVFCG1Un2UZ0CRCXv/61+8S9+sOTYC8+CX/h5gi0dKKSS0rkh+t0jnWW5/PffmLzO61ZvJgOOcmpTWunQCzecVHUhNUERERERERke0SQsA51yRAUoKUMDOyouBn117NdSwwAKqqIsQ2gDcDS3gcKUayLKOqKsyM5zznObvE/brDEiCHHXaYPf43fgMfAji/4h+0IvnRdpIdYlx4xU8YBUiZZ1COiBi9ThePmyQ+QmoSIeMkSHJtQ1olQURERERERES2nSgIgZQSzjm897hoxKom7xQMg/G+Uz7BEk2lCKGp9CD4ZidGjIQQMDPyPJ8kU4466qg7fIPGHZYA+c0nPmHS8NTabMfWKj+gHaOD4zOnn0pnr3k2LC4QsmxS/ZFlGX5c/dH+jMkWGKcKEBEREREREZHtZWaUsSbLMnIfyJwnxkiNUeaOz33zDK5jgVAU7ezbZhuMdx6Pw8yIMa74mccff/wdfr/usATI05/5TJaG/eZYWbr1K7omAXJVdRPn/PBCNsURpUtkRY63puRm3FxlcqemtsKYU/WHiIiIiIiIyPZKGCklsiyjrmu8c2Q+UKfIyBnrKfne1T8mAsRIsmaLDIDznlg1yZPRaERzlchxxx13h9+vOywB8uu//ihmujOUVhNcuM3rlsBZP/guaaagzj3duVn6/T7OOWZnZ5skyHjyi63cAgNtFYjOYREREREREZHtFzzDqpxsZ0kOYuZwe83yrYsvoMSAhC+aISXUEQyyPKcsSzqdDoPBgBCamP9pT3vaHboN5g5JgLzmda81gJpE4TKqWIE184LrsmqapzgoY0VFogJOO+sb2EzBKNZUKdLtdokxMhqN6M70mj4ftMkPnaYiIiIiIiIiv5hkZFnGsBzR7XZJKWF1O8DEO24ZLHL6t89iI0PoZE0s7prmpySjriqKopkSM14BnvGMZ9yhd+sOyRU8/vGPB8DZuPeHI7i2TCbPqVNsSmmynNo5ruhfzxU3X88tg0VqmkYsZgbJiJYo6+UKENusEmTcZ0REREREREREtp+5LVp0Yg66c7NsLAd8/XvnsmAVEZqtMkUB3pHlTUVInGqImlLiMY95zB16f3Z6AuSAAw6wI488EtoD5IA8BJiq3DAHg1hSASWO08/5JjcvbWKYanzb/NQZeO/x3mPOYQ6iSj9EREREREREfinjNpqJJs42muEiy4UGxijWfPm8s6h90RQweNfcsG1+amaTviDjpqj77LMPz3nOc+6wbTA7PWXwZ3/2ZwTXjq3F4QHfppVcyLC6xocMQsaAyAD45vnnkvJA3ikI7Tie2CZPXBbAO6LbesNTNUEVERERERER2T7TPTXH8fT0ZFVnkMoK38n5wVWXsZ4REQg+YClB1vT7qOumEaqZ4b2f9AH5oz/6ozvsvu30BMgLXvACnGuOXjcvJpmlcbMUYNxGhUTgOlvPVbfciMua2cGkBFXEpeVMUhnrpgLELW95UTGIiIiIiIiIyC8utXH2uKjAGwSDTjvI5Oaqz+fP/xoRsJSoYt1UgUwnTNr433vPpk2bOOqoo+6w+7PT8wR3vvOdCT5Q1RUAFhMkAx9gOMJlGTWJSDP95YtfP4PFekRtiRgjZkbmPXlo5hEbkMya5MdWqj2c6aQVERERERER2V7jKpDx1pdxJcj461mEajgidXO+fPY3miIE78mLgmhNF87xCF2AsiwBWLNmDXme88d//Md3SKS+UxMg/+fPX2xVVU1+saPNBrXVHLQlMVWqicB6Fjj9zK8ztAh+ufFpCAHvm3+6c46QZ1ttziIiIiIiIiIiv5jx9pfpwSJ+vAXGe8oAV958AzcNbgGa65V1PZkKM67+GK/QbI3567/+6zvk/uzUBMgrXvEK8qJgWJXkWd4kNJxrmqVYgiIjxUjwTSJkQ38TV264EZspSL5poJJomqdUVUVZllhKZD40zVimfpcmv4iIiIiIiIj8/IIt76ZI7cARa3uchtR8b2Zmho2DJRat4pwfXsgCFUaimzdjb+vUTICpYk3eToWp2vG4hx36K3fI/dppCZCHPfBBdsjBB5OAPC+IKS5ngYLDckg+4rxR4Agkvn7m1yh7noXcqLyB99BOffHtNpiAgzoS2lKcrW150TYYERERERERke1LErjNprSOe4F4a7p1dvLAoBwQikBVwL986F0sEujXfVyKzW6PtmWFDxlGs2Mjz/N2X03id574W3ZH3Led4tGPOaZJYFhqJsC0W1gMw4CINSN2UlO70bc+5194ARuqAbEIxFuZ5DIew+NMiQ4RERERERGRXzpRsFl8ndokiGtj8Ng2oDAHIxKDjueKeCNF1p3ccDrZsPkODRcy/uT4F+78+7WzftGxT31qe8+tGX3LVO8Pa74WcLjgqTF+es2V/PiKn+FC0BhbERERERERkZ1sOgkyboo66b/pHck1W13MwTnnnUuESW+K8TAYz1YSDw6e+Du/zSF3PnCnljHslATIPe5xD3vIwx4KDoLzzUFs76ZzTXPTyYF1ngrHuT+4gNjJ6MzNMG6cKiIiIiIiIiI7z3jyy1gaJziCB+eIlqgxzvrOuYwwqnGw38b940TIdF2DpQTecdQjj96p92WnJECOferv4bKMsq6apqcGjBugQpMAcY5YN9NfhkTO+v75+PkeZaybZqkiIiIiIiIicodJbTg/bmEBzfaWFByXXvkzLlu8lpLNJrRuHs67JnkSU+T3n/Ocnfrv3ykJkCc96UlEoIpx+QDE5V1Ark2K1HVNAq5auIEfX381S1azOOjT7XZ1pomIiIiIiIjsRI6tDxuJlkhYMxkmeJJ3bKqGXPDjH1LjmuTI9G1WFoWQMHwI/PpjHr1T788OT4AceOCBdu/73peaSFbkTd+PZigwjO942xC10+1SAV8771sMgjEgYsHrrBMRERERERHZSdLUfpXp5Md4Igx+6grBUzsjBse3L7qA/qRFKlskQca3qmIkAXPza3jq05+207Z87PDswpFHHsnee+1NIBBcwOq6PYoOzIixOTjWbn/pU/G1876FzRTUmaM702MwGOgMFBEREREREdmJvG2lf4drprpamwQxYBRrUub5wRU/ZRMjatjyhlNCOxq3tshz//APd9792dG/4C1veUs76Lbh8rz9wIFz5CFrPs8CIyKJnJv6C2wY9Sld0xtkMi1GRERERERERHYoc81l3AB1vA1m/HW8p65rvPeMRiOyImdEZGM95PPnnMGQuCL5UW5W1GDjHiLOc5/73Gen3a8dngDZd999m/G24y/cShYoOkcJfPbbX+KmwQKuk2MOqqoiHydNRERERERERGSHSbRJjtsQY8Q5R5ZlhBCakbh5YEM54MLLLiURWBj0Jz+naPt6WkwkSzgcsRmay0EH34V73eteO6XqYYcmQI4++mhbu3bt5Be56d/qVh7gGqgJfOb0U1mKJRZ8MxkmRjxOZ6GIiIiIiIjITuba6o/p5IFre3r6ttihTgmKjIHVXPjTS5tJMJ2MSFs10g5Ecc4RnCfGmkAgWiSEnJe9/C93yn3ZoQmQF73oRc0HZoT2l9kWG4iaPUMDIrewxJW33IDrdYiWcOMtMilt0XVWRERERERERG5/qb1sLXGQgJBnuGSkGDEz6hRxwWNFxobhEhff+GOcLyhJRKyZCGs0yQADUsIDmcswjOc+97k75X7t0ATIscceO/nYtc1TEjSFLptVgEQC5/74QioP2WwXM4M6koegHiAiIiIiIiIidwDPlmNwnXOklEhVTeY8zprJLpYH6uD44Cc/TgWURBJuua2FGZjRyQvMmhSLw5F3Ohx88MG2M+7LDnH00UfbzMwMAKmO7Z1tkh+RldNwoNkC85WzvkHfaipLTVfZ2FR+hBB01omIiIiIiIjsBONmpyu2vUwF8bHd0kIy8tD0AYkxEi1Rk/j+Dy9mgQGOZtrLpKdIareAGHjnKWPZ/D4zfud3fmeH368dlgB5ylOeQkpNRsd7P6mfMbYspzHghngLF/7oElLmGdZVk/jA4ZJBTDoDRURERERERHaycQ+QabGqKbKc4FzTssK5phGqgzJFknf87NorSTgS1uzwiGmyDcbaLTAuGYbhneeFf/LHO/y+7LAEyOMf/3i896SUcN5PtryE6V/afs2An153NTeWS4Q1M1QkUkoUIWsapFT1jh9XIyIiO53fjstqui/TJmPiRERERFa5zV/TpJQm21qqqsLMCCEQQiB5x0Ko+M6lFzfJDxKEAMFNcgDOexaXFunmHQKOUTnifve9H4cd+is7dBvMDnltefjhh9sDH/jAyVzgmCJ1Owc4s0RhkdoqIgksUpE4/bvnwD7zXD/cSJgtSFZTDvpkwZHlSn+IiKze/2Haisv4bQSP4Wzblzv6/wCbJ2LGiY3xC4Hx96e7o5uD1F6UBBEREZHVZvySLW322me8NyOE0CQ+nMOFgJlhMUFMWObpz2Z89IxT8BghRUZUlN4wanCJ5IyZ2Tk8TZFEr+hgKfGbj3/8Dn9dd7t74hOf2Iy3aXt3+BDAB6IliAbJEdqtP3WKlCQuvvynbKgGpOCoUjtTOARSSlRVpTNQRGRVJj+aNU0lAWwr39/W7e/oFwDj/2Fu7f5MX2fzEXEiIiIiq11iyzYWt/W6KToYBdhUj7jspitw3rdbYYAsTF4NetqCkLblRcJ42MMetkPvyw55nfbbv/3byy9erblz49eKFuvm89SMxh26xKU3XcHl11yFmdHNC4iJ4DxZljXXDR51ARERWeX/85yuiqD5n2N0UPutX6Jvr79K7p/GtYuIiIi0r4ucY1iXnHvh90j4NvHQ9AoxDN/2AcGWG6om4NGPecwO/XftkATIEUccseKO1+Okh3O40PQDCc2OGCwUnHHe2ZQpUhQFmQ+kqsbMSGbUJMJ4ZI6IiKyu//m1I9CDNRffXn4e02WXO/sy/v3j/ynDltUft3a/x/dViRERERHZ0ySM5B3nXvx9Rs0rKnJ82wO1eTfMJ5u6fvOa67C73ZVHP/rRO+zV0+2eAHniE59oe++9d3MX2+oP7z3JEg7XlrzAuCZkCJx98QVQZE02qKpxdfMys4w1g6okeu2hFhFZjTwQkiOL7SUtX4I1F3crFz+VhZiuHtmZF/2/R0REROQXSICkhC9yLrn2CjbR9P/0baJjnAAhAgZZNh6V6zDgyU9+8g59bXq7eu5zn0uv1wOWS1mC88uN4QBSBN8kQjYy4urF9dQBrI5YTOQhUIRsMkYnJm2AERFZrdxU1cf0GLXp3hq3Nv3Fs7KaYmdf3DYqOGwrvUDGfUBU+SEiIiJ7cgIkZo4b+ps47/IfEM1wQG2xSYI0mZAmQeBc+7rPU6XIYx732B3278pu7x/4qEc9avmFoHPNGFzntjgYnoABX/ru11mwihojJaPwod0P1IzRyUIgYSu75omIyOr4n1/7/7Totp4w2FZ62wNZ8j/3tpnb7d/vlv+V3jVvVEx/b1v/LiVBREREZE/kvaeyhGVw+rln8/hDH0hKCfxUDYbzkzi/JuFo2mUcfvjhO+7fdXv/wP333x8zI8Y4aWI63goDEDFo+4AMrOa0s77BhjiisvYFJg6rI6PRiIThgqf+/9n777hbrvK+G/6uMmXvfbfT1bsQAiEkIYEEFqIIYcBYYHDBjsE1T/zGLcVOnMdPEr92niRPivPYbwpxTTDBwQWbYmwwxhQDBlMEBgQIgVA/Ou0uu8zMKu8fa2b27H3fp0jnoNPWT5/9maO77Hv2zJq1ruu3ruv387ECJCIiIuJ0xYywqZi+XGchOlIVyEldvOd0PMQWGiZNz+o3fYGNiIiIiIiIiDhNoJSichbbT7jrnrsxQOktIOqukE6O7z2i5gyUUPT6A+589au+KdtIJzQ+e93rXue997MWuFIipWzVXgUikCDeUgjHlx68DzXIIVGhUsSGihGdpThCG42UMYyMiIiIOB3ha5cXnydMhKMSHrRiVEzI87xtFTncC8Lvl+rkvIwMxH0vzcL51Iu1UmHNMsbEmxwRERERERERsRUJkiZsmJJ9o3UOMUaoBA8YZ6aVIAIQAiUU3f2kn//5n/+mnNMJbYH5oR/6IZIkCS0uUmKtRSmFtTZY2TqHkAIvBV4pvrbvQQrpmdgKsUXQDNNe6oiIiIiI0xNCCMqyRBhHnmYoB7nuIYcFuXFoZltc3Fy7zMnsInECkJL1tTVkmiC1pDLBxsx7T5ZlOGNnzrf93XjrIyIiIiIiIs5SeO+pnCVb6DMeO/7kQ+/j+259BQrIpa77pKcvwWzef8UVV3xTzuuEEiC33347ULu+1CRIGwAjEEIggVJABXzwc5+kwuG9CA4xWwS+gthDHREREXG6QnjoqQxTVrjKkRnHQCY894abSS30VIJ0HuVmSQffaGIBysHJMmMpFNw33s9HP/dpEp1gJRhjkAgKa0BNl9Gt2mC8iGtYRERERERExNkHJ8AaS9LLqSYFb//L9/Kdt34becN8SKbBHpuLHlaWlrniiiv8Pffcc0LDwBNGgNx6662+qfgAKIqCXq+H90HM1ANSSET9gQ9R8oHPfILKWbTO2x20bhDpxaxzQERERETE6YdibUgvSUllSmbhouUd/Pgr3sAKOSmQMtX8cPVa6Ot/d/VARGedfLKOG8Dv3/sB7vrSF3AerAttnqlOKMuSqqpqq/fZhftwuiAREREREREREWcDhBAgBVVVoaXg6489whBDD413NgigiikB0ub79SaYMYbbb7+de+6554Se1wnrLnnBC17AeDwOyq5AkiSH/VmP5L7xPu7Z/3DQ/jCuJTp8vfPXbAYqH14REREREacfJJAmCalU5CjUxFAdXGcFRc6EnpuwgmObd6w4x3bnWCG8lnEs4lgA+sDgJBwzHDuXVqiKElsZXGUQnpbcbyodm3XLdcgPJ6bVLBERERERERERZxOMs2RZhinL4Oq62OOL+7/Wkh2+Ecan0+5c74KFDmTJ61//+hN+XiesAuSFL3whCwsLFEURdsfStNUA8e3n8diqwicJX370AcpUkiaKopygtJ4hP7oaIMKHKDr2U0dEREScXnCAkIJhMcE7yXKWkumEjIQ+igUp0Q1r0JAm9Zzf9r34k6cElQsohwXOWLIsDWuRNZiyRCLals923a7P2YmZqs6IiIiIiIiIiLOLADGGXq+H8LXF7UKfv/zER3net14edEIJ5Ad12NeWADctMVJy8803n/DzOiFR5fnnn++vu+46YKqMD8wGhbUDjBYSg+PTX/oCRU2/6LntsWbnDLa2HIyIiIiIOD1gJVT9hFEmWFeODQwbvmKVESWuZv0d+NmX9w5bv5x3eHdyXg4H1pEnKUoEsdZG5FtKOeMCs6UGSBwCEREREREREWchlFLYypDp0BlSCMfH7voU42ocYkSmrc+uS4D4aQAlhOAVr3jFCQ2nTggBctNNN7Ft2zYmkwla67b9RWuN9z6QOM4H/Q+tmGD4wle+xFoxZuQMIks2V3wwLR12cQstIiIi4rTFxngEWqH6GTZVoBUJGRqNxSKkBC1BTV9CSqSQIGor9ZP0Uki0UmilKIqCcVkglEQoFVzPjJ1Zt5q1q0FcviIiIiIiIiLORmRJSlEUSCnxUjD2lgf37WVYFZRlOSU+6FSANMFTHUAVRXHCq0BOCAHy0pe+FIA8z2e+LkRwfsGDsAZwlDi+/MjXeGTvw6R5wn47ZJR7rHRI71AOtJtWfVgBLra/RERERJyWkB4WRII2PtihWYOSEoEhA3qosBQJiZey9oSX4CUCiUTOLoZP9guQ1kNlEYnGJ4pKeKyEsixJtW61qubFupWvHWxiGUhERERERETE2RT/AbaYkAgoqwkilayZEclKnw996uP4NMG2ymkWvAXhQLigDQJY78iyjOc///kn9NxOiAbIUVkZ4RBK47xhIjxfuPcrlNag0xwhEyrhkR3b227LS3MBYgAZERERcfpB1O4o83O4ABQdmQ864tdMvyhOgRKK7ik0gl1Nqabwne/PVSw2nz0S+BERERERERFnZRwoBEIqClMhtMIAX77/69g6egqSb2Im3mvaYpQIOmvPetazTug5nZAKkKuuuuroP+TBeY9D8tmv3E3hA+ejEQjrkV0lWLk5gI6IiIiIiIiIiIiIiIiIiDi10ZiaCCVRSuEqgxIS4xyf/fLdVE2eTyAkuum+bOkDjxCCwWDAq1/96hPGCBw3AfLKV77SZ1l29B+0FiEVJYL7HnsEq4InsDCOpN4e83W7S6MHIn0kPyIiIiIiIiIiIiIiIiIiTicEEkTgnEMpBd5jveOh1f3ce/ABoFP1C9P2Yx9UUCVBTqMsS+68884Tdl7HTYB8//d/P1Iew9sIgQUeqB7jkY1D+EThnCPzktQJlK87gOoqkBN6khEREREREREREREREREREU8KHB6kwFpLlqR4Y7HWUiSCj/3tZ3AEs5RGCYSOKYpqTFGcI01TnvGMZ5yw8zpubuGWW245th9Uwf72Q5/8OPsnGxglUEqR6wRlfVvp0RAgTkz7pyMiIiIiIiIiIiIiIiIiIk59eFHn/97h8CilkM4Hl9d+ysfv/hwWj6+bXzx0ROgFgkCcNDjvvPNO2LkdN7+wffv2Y/5LG5R8+NOfoExE7eziSZHIKtgI+tr2tqsB0tgLRiIkIiIiIiIiIiIiIiIiIuLUh1SKypigAWIMWiqkUmxguHfvQ1Q1AeLxs4LxdWGEtw4pgxDq7t27ue22206IOMZx8Qo33HCDn7e+3QoeqHBMcHzt0YfQC32sCBaC3lhw08/SVdCPpEdERERERERERERERERExOkFKSXWWpRSFFWF1hqUZN0UrFYTKjymrQGZIw886CQBoKoqpJS84Q1vODHndTy//GM/9mNIKamq6og/54ASzxjDWjWhcAa0Ik1TlFII76PgaURERERERERERERERETEGQBrDEmSULlAgjjnMMYg+hkbvuIdH/ozRqYERKcZpoYHV7fAKKUwxvDa1772hJzXcREgd955J957kpqdORw8YFH81af/JrS3aNWSJhaP7xj/ykiCRERERERERERERERERESc1miy/MbltZG8MBL+4mN/RSk9DodCYJ2lLf8QoYUGQGuNEOKonMOx4gkTIBdeeKHftWsX3h+dsbDABPizj3wQJwVaSpy1WDzWObyW7ckIPyVBXBwzERERERERERERERERERGnDZqcXvpAabg5l1cr4fPf+CpC5hhcrX0qQAi8Me3PlWXZ/ltrzc0333zc5RJPmAC54YYbsNYihMB0TnIrWOAAY774jXsRQuCMJVUa7z1WC5wUreuL8rEKJCIiIiIiIiIiIiIiIiLidIXqSFx4ptUfEAod/CBjL+tIFBKQTb2IFODszHtJKdFa8/KXv/y4z+sJEyA33XQTVVUhhAiCJkeABe499CBF/WO+qEh1gvceJwWmPot5AsSLOHAiIiIiIiIiIiIiIiIiIk4XtJ0dTPP6prujcX91g4yP/u2n8QgEtJ0lQmtQCgSkaYpzLlSHAK961atOyLk9ITz/+c8ny7Jj+lkHfPaeL6EHPYQQSOvDRRGCSniM8C3ZITpMkRORBImIiIiIiIiIiIiIiIiIOJ2g3NZ5vaqZkLF0/NVn/ib8T5D9wDkX/tHhAJybCmNcc801x31ex1UBYmtl1u5JbQUHfP4rX6LAIj1kSlONC7TWlN5ilcALWieYaH8bEREREREREREREREREXH6QdSExlYurw23UXjLPQ9+gw0zDASIEHjvGxlUAIwxaK2x1rbyG7fffvtxCWY8Ia7h+c9/vs+zHJzHW4eUcuZEu/9jgAr4xr6HWJ2MMApEohkXE5RS2MpMRU87TE/UAYmIiIiIiIiIiIiIiIiIODY0OhuPt4tCbvE67vNgeh5NoYPwIec3Esa24uBwnQcPPIYVDi9E6/zi6wKLRmvUe9/a4R5vFcgT+mx3fOtLwTm0ThBSYozB1f9V5QS8A2fxOFbdiHuGD/Lgvr3IPGUsPRu+or8wwBQlPZWgXagScbJ+MVWOdcz2DEVEREREREREREREREREnE04HDHRJT1c5+XFkQmRLtkh/ObX8ZAgDjDKY2vmQ3qPrhP6UjkK7dCZxkrH1/Y9zLqwFATtUOEtUoC3jjzPmUwmaK1xzqGU4kUvetGTT4BcffXVICVYCz5Y0jhCuUoiVe1142tSQ/Ol++7FSrA122PFlOSY6Q1ilugQ0Q43IiIiIiIiIiIiIiIiImJLdDU0j9RFMU+IuE7+7cXW3z8euLoKZN7oxMrwcqHvhc/e8yUcCRVQegNCgJ3+9TzPw+cUAiEEz3jGM47rvJ4QAdKUnTg7tafxTc9LXbaCFBhvEWg+/bnPxpEZEREREREREREREREREXEMOBxhIQ/zUjXRkLip0KgTh381RISVUG3xOl4zkoaLaciPrhCqqAsmhBDc9YW/pWEVmpaXjrhG+By1E4xzjksuuYSrr776CQtmPG4CZNe55/hzzz03/HKagpieqGt5oqB64oXAAffef190c4mIiIiIiIiIiIiIiIiIOAHYsm3FhZfyU0mJ7msrksKzmRw5kXKcjf5HQ+I056KlxHjHQ/v28phfDdUitd2tlwKhJM45yrJEykBblGUJwPOe97wnfj6P9xfuuOMOFgYLM1/TWiOR4Gu5VzxIAUhGFKyNhrGNJSIiIiIiIiIiIiIiIiLiGNE1CTlcQcEMAVK/VE2CaDetDGnkJ5Sb/pzovMeJRKNH0j0/W1edNF9TQlJZQ6UFn//ql/E1r+AhtMEAUsqWFIFpO8zx6IA8bgLkzjvvDKImVQXO4Y1pfXtbKInDUwF/e/+XmTgTR29ERERERERERERERERExLEm634LYqHz73mR0y5h0mhtdqtClJ9tl1Hdr/tZYuREkSKio/3hxFQTxHuPcRaRp3z67s9jcChU+/lsLbeRJEn42brrpCxLrr322id8Pvrx/sL1118fLq5zICXee7YioyrvMELx0U9+grGt8CoO4IiIiIiIiIiIiIiIiIiIYyUOGjQkhxfgut9rrGab789Vg0zfYPrPhiiRbF1Z0hIqx3H+XoS/Kev3aVptmmoUASAlE2+45/77qHB4JBaHFJKtdEAAlFI0khxPBI+7AmRlZQWANE1BKaTezKF4wAhPgeVvv/plRrbERg2QiIiIiIiIiIiIiIiIiIhjStQPZ33bFTBtX2LqumrktOKi+2rfu8MtdFtj2haZb4J+RSuA2vlMOk0YmpL9wzUMwQbXEtxklVJYazHGIIQI/EP99aWlJS699FL/RK/r48Li4mK46I0DTM3ESGjtahwg0Kz6IfuHa5AlMxc8IiIiIiIiIiIiIiIiIiJi6yR9RtuDaUVII1JqvMN4FwgDAVYJjIRKeCrhKXFUwreEiJFgFLhE4lOFcRYvQCiJ0AqUxEsRyJITIIN6uBaahnypqoq0l1Pi2Lt6kAkGW3+9gVIKvUXBhdaa173udU/42h4zfugHftBrpduTgZoI8eFTtKqteCwwrAo2bIlVIrrAREREREREREREREREREQ8QRJhRuNDKYRWeCmweIwPVrFKKVSikWmCT9RMRUgpPCNbMSoLSDUGz7gq2SgnFLbCCo9IFEmWnjATE0eo/hCdlhzhw/lPigKRaIzwfPxLn6HEhXM/hvd97Wtf+4TO53ERID/w+teDd3j8LKFREx8NAWJrJuq+Rx5iaEtKRSRAIiIiIiIiIiIiIiIiIiKOAcciQtq6viACueA83jqccxTOSKAflAABAABJREFUMPGGUjiMAp8oRJYg8xSZp5BqSDUiSxBZgk81RgnGzrBeTo4rfxcdMVU/5/7SfC6pFcNygswSShzv/8iHmGBQxyhTeuWVVz6hc3tcIqjPe97zEEJSmAopFUKI1pPX1YwTNH07gk/e9Wkm3lI+/k6biIiIiIiIiIiIiIiIiIizEk3VRBfST6tAqqJESomSMhyFBOcR3uOcR2mFJ7iteOvxxuDqzg2Hp5wUIEVrNdv+KSkQJyB/V/W5Whladho+pRFetd6FzyhCBcsXvvoVKrpKrkd+/8YV5vHimAmQp1xymZdJgqcutxHhkzSkh/AgajJECoXB8Tef+TSkGitm+5YiIiIiIiIiIiIiIiIiIiI2wwUeYsaNpfl300aS5b26osIjrAfrENaFChB8sI+VAoTA+yAwCgKpFVIrKhs0QJwIDivWObz3SK1QWoGxT/j8m9xfAmXtBtvY8jYonW0NVTywYUsKLI6k5RiOhOXlZa6//nr/6U9/WjzeczsmXHrppUyGw/BLoqlhAZTCWdtWggRbXMFqtcaDex9BJBoXC0AiIiIiIiIiIiIiIiIiIo4JrvPqtqM0bSRuUmLHBW5UwLgkLR19r1hJe+zOFshGhv7IMhhZBhPHYOLIC0cyqhAbBX44QYxKdOVIvSBFkggJ1mGfYHVFFw3Z4QnuNA0h0nw2i0clGudCJYjMUx5Ye+TY318IXvCCFzzu8zpmauKGZ14HzlNW5eY3qQVRIRAgAPc/+ABKB0ETB9EFJiIiIiIiIiIiIiIiIiLiGODF9NUlFSShvSSTmr5KGMiEBZnQR5EbSMaGZKPk+gsu57o9F/P0nRfw1G3ncuXKHi7Ml9nuEvKJZafMGXhFMjHIcUViPKkT5F6SuBPXvWFlxwLXTz+bVCqQH94jhMAIz99+6YvH7D9TliV33nnn4z6fY26Bue6mZ5EvLlCZCgk4a8MNqMkPR6j8cM5hpeTr+x9Fb1/EmBFCihmv4bN1AB8ORxpcJ6p4xsU55ITcw1a05xh/R/jwe/H6R5zOkFvMI49vTgv7F8IrwGEkTHCsA1W7gkz7Q0XdNunqr9H53pN9HAKFdrOLN9PyU1//cLcs1Ynp990W16p7fWScqyNOcixyuBhkPliN6/+xzZMnGnFOONnrV7x+x3Ktum0acf7Y+jo+3me5aU1p38MH0kNb0A5SC8oY9izu4LILLuApl1zGZZdeykXnX8AOtYMMRYpH4LCAQNVVF7DGButmyCc+/SkeeuxR7v3GfTyw9xHWx2OMCtUalZAUSbDQPdLasFUcY2Ut4SFCkKRch9Cpf0Z4yJWmKgocglLBhq24+xtfw95E6zJ7JB2QNE257bbbHvf9OOa6jEcffdQvLy+TpVn4AM61mh8IsDiKqqCfZKwD//gP/jN/9sVPMpGeCoc4i58AL8JA6i6g3QvftQOaH0gn6rL5OAM98eBDEKYPEXrXur7csuPHvYkk8dNEqplM5u/rsQdM8f5FnDyIerVryLxmUfaHWVDm5zSkQEkB1qONY0cy4JylFXbmi6Re4q3dFFR1gyknp6TCk30sFXxjbT/7JxtsTMY4JVCJxnqPtxbnXGsL7w4T8HSDh4Y8aXZw5qfmLmE6JVz8ab6JILZM6I5VXf50X77kSZ+/xZbPbfd5bcaj2yLwlv7sTmKaazH/DM6v4+Jw3z+CheWx/O0Yv53k9essT9qlmxXd7OYuYi5KdSKs11ZMY17tT+/1yyGO8xr6ubhGzKyDKtFYazHGsbS0xNraGnmeUxQFUkp0nrCxsYEAlvM+uvCIjYILFrbxtHMu4v/4rr/DZYsXkALWlWiZYuv3ryhZIG3ndgsY/HSzieAcU/gCLTIqPH/6iffzzr94Lw8f2s+j5QbjHRkj5THG4L1HpwmmjnucMQjjApGhE6SUjKuSSnpcqkE4lPVt7tTEQ76eTxMHurCkaco4FYzLgsU0Z2m14v2/+N85hxz8kTMlZy1SKZ5y5ZV85av3HPPNOqYKkPPOO88vLC2SZlk7Y3SFSbwPtrg6SfDeY4TgwYP7mEiPtRatFR57VrPYyk8H++EW0a0S4hNlH9xM9idjB/VMONoOweE8qLnFlLmjqO93s2i445p8w4nE+xCPJ+MoO2SfCzpa2Ca5F8dGzVVViZUS6TxVZdlrLet2wtfdozhj0FLNzItdQtFxcm3UvQCfJoxNia0rHb2d1qYIpbYkPmYTmE4C1ZArdcS4pcL7HJnS9s6ehuMHaHd+jrSmHSkpPN1F1E/m+ttcf8Xsc8sW5MdWYy9i67htfq2ved4tv2/F1jvlj4d8ievRyVu/znZ01+HueuW2eC6an3Wdlg17Go+fdgE+YWTc9Do1c+1oNKLX62FtxerqKs45nLHYypCmGeODq+xZ2Y6YVBR713jGJU/l+/7Oq7j14hsYIOgjSQFKh5ApCBgXJV4KFtIUVdWOL0rgZdjQhbAmCKAalQzykN97Kbjzphfx8ptexGNugw9+7hP86h++iaIYsrR9mco7SDTDyRgzKVhYWMBVBonAGsNwPMIpQZLmOAllaZD1xChdeP6q2hFG2HA9UurY0DkqBaWEAsf+8Sq7e0nQHT1Sp0RdjHHRRRfxla/ec8z34pgIkGc84xn08l5nMPjwKZrF3TlQAonEesuqHfLgIw/jhcc5h5YJ1tmzdvJQfsqgbhXozSfQbovF77ifwbiAHtexgaWzgPrZe7NJnOgoZBfHGGj6OfuoeIzHJ/OoHOi67LKxM1OdnZ6tyorng32ZaFSiUQ5cYvEeCgEjV2FcSaKSNtrYilQUJ3MH2oMfV8EmTiuklDg8xoZPrqRsNwEOd45NMNgKgPnZeUVsQX5LAtl6uj//wofE8Ej3bz4RadfCmjQ6nXcPT/b9axT3VR22zfeSb/W8Rhz79bFyc3zmxdHJk8NMNTF+O8XWr7Mdtr5Ofo7U665d8wS367y6c/zpTOAfz5hoNi+6RH5bHSZgeXGJ1fU1kqxHohQSQTEas9QboB0s+B75Y2OuvfByXvXtt/OCp9/ICimJcwg83nmEVvhUMjEVmUjo5WnYbJmUkKZTgrb7wepzSPLw/cloSNrvkUqJw7JL9nnpM5/LtU99Kp/44ud44+++iSJX7N13gOVdO6icoFgfYqwl6eeYRGCkRkqJtRZVOPpI7NzF68ZzXoDUCuMc1jqEqqu1vOeBRx/miku2kx11ghZUZcmznvUs3vf+vzixBMiNN97YGd0enAMpa28eat/gcFW9lNz34IOsDzfwA4VSCu/9Wb+j0C0V24oImd/p3Ir0OJ7ARHTGfDw+/uNWk18TtM/fsyZ5wc+WEB93ACbifYjHJ//YBvlubm4SnbHePBOH0xLwHmfqKkDncCLsRIhMI7TAyWmTQPf9/RYB1kmBcUgpUUKGubrzQRvy40jPrnZzO0ocW5l1Q4Kczs//YT8XR+4fPhHz5imVQIuTd/19Tdq31Uhz92I+qWnG3eFaO866+K0zLruBe5e48HPJXjdZfDzXr/mdrRLNeDw569fZDs+cFkU3rvXh+nbnatlp/fJzN+O0G0fHOSY2aXh0CIBmDO7bt49de3YzHhckSrPv4Uc5f9cehqtrJAau2nE+3/GCO3jFTS9kGU3fA6Z+N+cQicJ6RyUETieM69lKCI/KUyp8rbI2JWIkAi0EAphMRvR7ffKFAb7+fl5rhSSkrGR7uOi6Hdx67bN44+++iU/f+yX2rW/gFYg0wUgovMXY4OaSSY0ZTfBFRZLnOH/kAoigEOeD9a4MgqgWzzceehB/ybWbYqetORDBLbfc8rjuzTERINdccw3OOyQitL405EdzQ2u2yOLxCO7+2lfD96VAKYUx5punDnW6sKhy6wejufnzPYlb7SAcTxIgvCC6ET/x4HV+B2iml3Ruopzvqz7a5Hgsfz+J23IRJwkOKBR43XkeDpPEbzXHOA+q6R0jkAfOO7z17TrirNuyBaJd+E6yBoFSqhbt8lgXFmqhZP35QlvMYckPD5ndXDHmtmidO1zyJU9zGzXXDZqba9pN7tzceJobA0advm0ZJ3v+bp/fw4iftm0Cc89vQ+J3CamzHd32oPk5z3XGq5qbq9TjmL+2IkAkcf0/WetXHP/TFv5NGjdsbqHrxsvNmueFOK0ry473/rdrvN/6vZdWlhmNRpSTivGhdS7ecy5mdch5vWWetudC/tWP/COWvaLvFGkTPHgB2FakxQAVngpPiaPCYjCMqDhIhcUjgcwLcqEZkNAjIUWR9nI2cHhfooQkQ+NshUKQq6BPoqqK7fk2/vX3/iSfO/B1/vmv/nvuK1epMkW60GdsK/ykwBuHUJ5cashkaJnZai7t5lT1hRFCIGszFaTgvocfpKqJCi2OzNDpJOHaa6898QTI5ZdfjhAiBHmetv3F1Z1EAoFCYHCUCO554D5EohFiVivkbJ6AEYdPfN1RVNkFxy9CpiKzfVwBrOr4WLvujelM8vP36WjuDo93QYhlmREn5wEIBK7tJLBNSWhXGPCISvFCIJRECBH0c3wgEWynxk1ssTB2v+flbGLwZB3pnK9xDocP1SAyCJ9ae/T2zka5vXWF6la5iCOr5p/u+hfzGiiWaVvMTPLhtx4DTsxWUJyOSfNJvX9bPL8za1RHk2arqo9ur/pZHcMd5t42U9jRYizXIVC2Os4wft1jXPtP6vp1tpMfzZp1uBb++fHdncubda86zXdf5TfBybHbiiilZPXgIS6/6DImB1axB9ZJJ5Yfeu1reeW1L2AnCX3RbAZ58A6UBCUoEKxTYUhYo+Av7/or3v+RD7P3wF6chYOTIUPtMcKjrUdbyIUik5pcJ6RC8fSrnsorX/oynrLtMlKCSGqmknA/rUOh6OsefjghHeRcubiL3/wX/4l/87Zf451/82HW1tcRCzkLeR9bVUyGY3p5TpqlmKLYlPPMd+I4PF4KpBQ4IXDOIrXi6w89QIk/egtMLctx3nnnnXgC5LLLLtvEQPu5/xECBJIJjgf278XgQSmqqgpVIN6ctROIF7MOIFvwIUdloo9HCLAR4YoVIMdxDzuTf6NwvZWa/lZBb7fU0h0jATb/Xs5vFliLiHiy5q/umJad11ZOJltZ4ZXWIJBBrEqGlklX7wxIKcH5qWbOXBLWEARWHDmB+GYdIbieNefeVrI0JIYU7TxwuGtRSbB+tu+8W/V3OE0G0aztnL7zd+MG0JRJN2TGVruwZ2IS4ji5c/dWz+9WbnQNCdIlqGLivXm97sZoXTe4rcg7UbcHVGoqoH6449ESsIiTs35FbJ2ruG7rd6fCQdQXW3VajrbSHTrd5vDjJZHm59lmbQzXzbNj2zYeuOdeLljcznmLO/mZn/p7XL/7KjJgYsY4kZArhVQCrxQG2MCyjuOtf/2nfOQLd/H5++9hbCuEEGgpSGWC74GTAu/BKrDOU2DxwuJV0Da75+uf4m3/78dYEAm3Xn8Tr7z5Np6++2K206OvQGyMkVkPkeeURcFyNiABfuzO7+HyCy/mLe95J48U66A8SZ6zWhSMbUUapFmPek0sHiEk3juEm4rL37f3YR6zqwzUzmN4xj15r8fLv/Vl/k/+9N3HNNqOiQDZsWNHe5OErfU/hMAjpqIq9WGC5VA5pvQWCARImqahDeZsZlHnGK/5RXJ+12W+hAx5/P2JcUJ/ghN+Z8e2SV7mRfvmF9OtrC23vCci3r+IU3zu8sFrvtkt22qcT0tdp2O1W+mQ53lbReG8x1kH3nd0Ndx010jMJhaOUP2h6on0yT42a5+UoYLF2tDuKWtW1Hs/64omZnfSm+DPHIb0mLfXnGlL8GfG89+KcR4mmG6uk52bE+edHM7WAPp4n1/tNq9T8yRVV7RwnnQ/m9ee+ef0SGLHW1WAer91BZPYQgMoViGcmuuXcvEayi2+Nm/S0CT3zb+7JgIuXrtOxcMshusbnLeyg0F/me0i45f/6S9wPkvIYsRC1qfSPSxQAQUGj+KhyX7e+Vfv448/+OfsEyWHfInva2Tao5iM0UAuHLaqEC6cgZRT9sUKqITFSphUJSsrS1Sl5Z2f/DB/9bGPcu1Fl/Mdt93B865+JjsX+mAsKIVSGR5I8Vwol/m+G1+Om5S87QPv5St7HyHfs508z4Otr3dY72ZahJtqopm4R4jp/Fp3mQgleXTjEPuHa1y0dGQCRAhBVVUkOuGaa67hT/703cd0b45KgPzET/yE7/4RlDriYN5n13h09UAoOfOOtJfPlDmfrQ9AJiS2MvUglDOBc1L3WIm6t7x5BXHZui9KSg7XTSSEaHuoDruI+7iVcyICodYKrPP14PLjEdajEEghSKSichZjDDJNcEcRSmzueVEULC0t4SoTrLCcQycp1tpYwRNx0uYvZTy9NMNYg63HphCCLMuorJm2N3Tta7sipsZRjCe11ZvDC4lUQSncWxPmN3/4dgjlT65dobUO6w1ChSqW0XjEYDBo7eucc1smLa3YZ6rbn/U2lHdWVYXWGunCqt9tk2nIgiZQMPI03kGrE3DhD59cWu8wzqHyFGvCdVZpQlEWKCHR7tRv42zIMSklqq5+dc6RZRne2pP6/KbVdGzJOeLDChAyjLHmmW13dWPmPY2fRNDAa8hO73zdBu5RSYKtDJV36NoVqiV4tUb4sH4L/Aw52rx3ExMebpcrVoAcZwJaOQZ5j+FoSKqT9jl1ziGUPOL61fz/2VoN1Qhm4jyJDJqOWmsqZ0mylLXREJ0mM3OHaBynjoHcOx3g6rytmderqiJJErTWlGU5fX4Pg0RIJpMJi4uLrI+G9AcLFEXBpCoZpDkrWZ/q0AYX9LbxP37uV1jCkZuSvuqDDZUbh+yITPU5QMGff+xD/OYf/C4sZTxSbWAGGpQO83nddSE9WOdAK6yq1ybrSIQMep7eoT0IKUikZFyMGaQ5Yx0UVj/+wJf57G/ezVV7LuKf/R8/zYWLu0nwaAQKj/aOTCgy4Ie+5VVkTvDm972LR9ZGkEDSzxh7i0skUniUm41h5ttcEaC1ZmMyZtDvMxoW7Fxe4KOf+STXPv+ysGlEcJZVUlGUBXmatbmylBIEPPP66475vh6VALnppptmApkujVVXK4fE3EMpDJ//6pcZYkAGIRNr3RGT97NlAvGVRft68RSNeJ7DOkfpq/a6ClmrqghRl4oHzFfQNItwQ34cjeCwHZvGeHz8GgCtJSVhl9J2dgaEh37eo1wfkfdy9j26l8WVZXSSIBJNaaqjRjDee1ZWVtjY2KAcT+q/KUiEpKqqNliN9yMen+yjBBKlOXDgAHmek2UZWZYxHA7ZKMbkvR7Wmi0dJqYkoSCXGmk9pjToLMXb0Gfqvac01Sa76fkI5GReh0GvF8T0igKhFJlOZqpCjhpAVYZEaxKtoXbEUQhsWZHpBGts2+6Cn+4aN9VmVkxbE07H+VN1Knpm1rH68y0uLjApCyZVhRCgpcIZG6oXpET6Uz8Tb5LjhhzsBsXGuZP6/Nb5eLuT3XzfU1tc1oK+bWVSl4A8y3kQ4UGq2pnAWZxz7bMvZSByN8oJUkp0pvFa472nqqp6Y8uhRPAR7e6GB8G/2Q0scZiNl5M5fs6E9WvQ77N68BDee7IsI000xhhKa1no9RkVkyOuX2czHOC9I9GaqqhjUWOwNRGSak3pXZhjOhU0zk+vpa3H/um6fjUxejO3J0nS5mVHW/8loJGkckqWFEWBc45BmpNakKOSq865iH/z4/8cyYQFcnKZwrCEVDFSFlSfL40f5hd/5T/wwNp+1O4BX33kARZ2bWfiyhAz2GZd9bWzSj2nSwVCIoTHu6DnmctAWhk8SglkkjAqJqhEIfoZxaSiUIYvrD/Kj/zSz/EdL3kZr7/9u9EYltBkQmGLgl6WIYHvff63kaYpb3rvO3jUjVmbVNDTWCHCJs/cnDo/x3kCkaGkDMUAUlA4x8MH94cWaGdb3TWgvQdN/qvq4oyLL774xBEg119//Sz50eFCXLO71z4onr/+7KeZ4JBpAsLhm4n7LJ9EKmvQUoUg19lAbiiBF4rKu9ZlwAtacUBHWGhxnn6W1uVBdYVI58YDSHV4BjIws7INbOLx8R3D2HbT+yM3PwiPPfYYu7fvZDIp2HP+eRwarrM+GZIvLbBRlAxkdkQGvKoq1tbWGK1vMOj12bFtG+vr6zhjg+10moW/G+9HPD7JRwsUSmCXe+jlZQ6sriIrQ0mJShU+AbOF3XNX26GsDKlKAIHWKWmas7GxQWUcxhiyXt6WIM/swjUBhOu0xIgn9wiwurpKlmWURYHWmjzNsN6hhcRVZhPB2QaC9e/3kpSiKDDDMWVRkKYpCwsLDIdDEiHbyspmSjFz72WFPK3nT4Pb5BjQJZcfffgRVrZvo6cSKmuwRYUAelrhKntYl51TjQABcM61xJgQAusdPklO2vUPz6ZrxbxDxWIn8BRQCT/TAqP8tPKqqUA4mxNDU1ZtcK4ShZSy1f7xTQUNnhKLszbEZgpUqkmUxpcmJFWdCl/vfLtFnunksAmCF+BVEtej41i/DpgxpqdYXl5mWJYUxQShBJVzjIsNtNaHXb+2StjOquonAWNrGOQpZeno5zkaiZn4MLcJgVRyxqlL1DqdvtXBO73XL+kCGd8QIFqHqn1rLUmSbFkBOjN/GEOSJExMiagrX7MkJTEeMSw4v7/Cz/zwj7ENzTZSNB4zHKPzPiSwRsn77/sU//7X/gtr0mK35Tyy+jDbLt7DejlGO0niQNtpzmdk0B8yEoyzKBGqbEtjEE6g6nzUVRXGWxa35YwLg9eaibIYLNkg5YG1DbYtL/Hrf/7H3HXvl/mFv/uPSBBYb+lnGZNiQprlLKL4nptfxtr+g/zeX/05E2nxCxljazcRDV1SvdHIMs6RiuAcG6o8BGMsX9v7EK7eRBAyWPMqQArZzr+2MqRpiveeyy677MQRICsrK5sfiK12BDwoobj7vnuptMDK2h1G1YuFPXs1QByQ9fK25K40Bl+X3nkhcM5DTRTZpq+8TQIEWgXGUDEtwWx2H5rA62gaK90+vXh8/Efl3bQscla+GOFg985dTNaHOOcYlwVGQrawwIarEP0Ma6dB51ZYWFhAC0myKEmUYri+wcahNVYWl8j7ORtlYHjj/YjHk3GcOIvDM5yMKXGsLC2R1juhq6ur9NKs/fktNRtSjZESU5R1cOlw3tPv9UiShI3RsK10aEiQppxWufDsKDe7aD6Zx5WlZbTWIQgyBus91nvSNEUoiTHmiHo+hw4dot/v0+/30VpjyopyUlBOClKdgHMzorEz5ddMbWJPx/HjxfRKzuubNETIRZdczPr6OocOHWJpaYlelnJwdRVVt5DYk9hCckxJQqcMtznXLiHSkucn6fpbGXqqtAuknO44ElkBXomWgOxqU8wLeZ+tSJQK1T31RbXW4ipH5SzOe9I8zH9SSrSU9a55SBDLsiRxTVuAmCHHZH0sx5OZneRNBIg4eePnjIjfEo3As7qxzqiYkKYpiwuL9LVmdXW15a+3Wr/OegckAclCD6skI1PCBFKpQnWTEOgktHv7UHrQ2uWKTtvjyZz/jvfY6PdoHaqGGuKjO+cfjQApioKFhYWQ3AtQIlTEluurnJ8t8Q+//0e4tLeD7aQIW2KsJ1/s4/AcxPDWj/wpv/HOt2K0hkRRjifsHiwz3LufpV6fxHi0C/N708rVOB85CdZ60lQhPBjjSYQkVWn7+TKVYYsyVJ0nCQbPhikgT1g+fw+jjTF6IeUz99/Dj/3SP+Gf/eiPc8OeK5E4elmOAqrJhDxP+eFXfC9fevgb2L338dWDB+kvDxBNl8McCdJWC9UFFN4HGQFfa41aJXj44L5Q/drtesC3myINIdUQTdu3bz/2TYuj/cC+ffv8jh07ZipAvAhiLL5mUFT4y6xpwYt+8cf4hhi3bGBT4if82TuFeAGFqTZpf6h6UfXekyrdllh1NUBkTXo0gaLvCO91f7YZAIf7+wL1uAQ3I2Z3ApoS7K4DDISkTHlIkSgEZVniU43rJezdWCXfscx4Y8iy06RuswtM8z7VpEBKSaZ0cMQwjkG/j0Lw2IH99JaWsVEEJOIkIU81lbWUZYlTAiElhw4eYGH7tnYOa4LGRvCymbOMhKF3yCxBGEeWpGBsWHCtRwlJonUbJHWJkCYhS+3W9rhPFqrJGKUUWmu8FJiaENJpSlEduQfYieBR71yodnHWkslQOuzMNJDqagw1BEjjnKKdPG13IUOSPS3hnydAADY2NsiyDJ2mSCkpTYXFo5OEsixJpTqlP7+1tt2BMsaEis56jQ4Jrzrp118ASU0mJh2HBiuglL7VAIEQSCsXnjtgZnf3bESj69EkPg3ZoZTCC4ESgtKY8Hw7184VzbxgvZuJ3aj13kQdv6VpOkN8SOZFlUWM344DwltErd+QLfQpypL14QZZv9fer8OtX+09OEuvnZEwScLnT42nLzQ9ofGVCTpH3kOmqWR3w5C24iwQCfK0Hr9FMSbP81ZsE2hbLo5FX9FXFSrRoFXYQa0caeXpDw3f8+KX8xPf+ndYBBIfnv1KCCZ4CgS/9Du/wrvv+ghqqUdqBbK0vPi5t6JKy7alZcZrG5tIBeHlJscvK2FtMuKR9YMcHG2w7koOFEMOjDZIFnpMTIXMUya2wopAGq5tbOCMoS8TlnoDEiFx62O2iZR/9RM/w7fsvho5HrKcDgAolWOI5ACeH/13/4ivH3qMIhWYZNaGeiY+ApBBo6SprMGFnNZ7z65C8r5/8Rvsqus1PL6dH501aKXr9mHfavvccftLeN/7/+KoI+6oFSDbtm07hhUWEJKxG+OVxIguoxMIkrN97h4MBljn6pahICjkjAFjkR4SYUKflPPTUlWCaq8k9NyJjgiXECGoah4+Wx5lh8zHzsbjIUBUff0aVtV3Sv6VA1cWDHp9Dq6NyVYW2be+yiXn7uSB/ftZ7PdQ1h31ORuub0wZWFvgjWX/wUNcsHsP1dr4iBUkERHfzPFflWvsGAwovcJpSSU8K0u7cEIxLMYgZZu4yy0sBbMsCz7vAlRlGR9c49JzzmeytgFFhSrcjL1etx9bEpK2k0WAGAXZ9h18/eEHWV5eRiJw1Du5tZ6HTI/MThbOtAlPP8nIhGJ44BA7l1ZYP3iIfq/XBitdcrT5yIcTET1dCBAv3JbEUBM+rAy2sTrcoJ8nTJxhMhrTX15EKnXaCHg3mxHdjQ4pJUtJjlgrSOzJv/4NsdFUd1sJhQKZCioFpXx8Fu1nC8o6/lKyERAk2DXW4vWJ0vSlRqokdN8bB5XH2hJjzJTgqGM31W5shWOxXoS5tt7llJ6ZihDr4404nvXLWcPSUp8D6xMoxmhvWcwXqCzINGNSlkdcv87m0Et5SAtHMZlw/vY9jPYdRJdjMh82LpTWjDaqWRcpP42Nmzn+dJ1LKgVDrZmUJUmShPi8k49NJpNWj+JwSPo548kEpQSp0rhRhZxYbnrKNXzvS1/JIp7cC6q1EWppgQo4QMl/fddb+JPPfwyzvcdGUZCOKi7KV/iJb3sDGZYcQYpEElpDbEuEyFaeoqHeLVACYxxjHAUwxDCi5Hff/ft89JOf4OChddJEYXNNaSHXCdt372GytoGQkvWyAO3pJYp/++v/BfddP8DtVzwrdDQg8UVFkiX0kfzd7/47/Jv/+iuMnGDDWazqEpKzz5f3HqFVW0kTqu08TkKB496Hvsau867EH0EKv2kzAnjZy17G+97/F0cnto/0zde97nX+aOq2XezduxcrPKWzZFmGcC6IvSl1VrfACB9E8KpJAc6RqYTUCzCWzIcAaXhglZWsz67lbZy/aw/n7drDubt2s31phX6es5j20Eq1AoR5npMmKRqNRKKPwmXJ6CHyxO8foOslsJlkXIdxVfXxULHKIFumAB5jxEPuED/x8/8ELyUb3uGOcAtWV1cD8SVBKUFZGYwXnH/uufz6P/xPnE/KIN6KiJMRANXj3xG0KUZYKhRj4H9/4o95x/vew4FiSKnqvtPG2UNM5z89LFE2lGku65w9Oy7iF37q57iAXWRAWi9GteP3NAFj2nJ5stLgIfDWe9/Pb7ztf1Naw3gyYVwWJFlKolJ6vR7V3Po2b+VaTYJieaZTdGnRk4oL0yV+6ad+jqf2LyUDEjoVlcyWiPrTeP5u5s+mlLyZQ00dkJXAY6zxs//m/+Lur3+NxT070f1FKuPZGG2wsLQY1s5TOcmqy6C7wrjNzv6ewTJv+cf/jqWTeH6uvv6qM6YMMAE2gNf+/A8wTEHoqRtMU3F4trcAAEjd7PYG9xdVE0nKQuIEjApSJNoLtIOVhUUuueBCLr/sMs7duZuFNCeViixJ2/gtyzIynaCUYiVZbp8VQV3xO7NtGOO341m/rB+TiIwRJZ6UCTAG/vO7f4MPf+oTrBoOu351E7azEZmB5QKesucK/s8f/0fsZoFlsnbNNlhAYZn1yVCdBNMiT1sv0DXg+974T/jCYw9QVRX9fr+t9EqS5KgiqF6AkwKvZVsJJp3nol3n8B13vJzzxDK5DWI1Sb6AK0rKPOWPPv4+/uBj76fYNWAoLWvFiPNXFhkbG2QVSFhAhGtsmxb9UGlja81CV1936YPIeiICYbKErDeXNI6c6172Qwxf9v38zT1/y5v/5I/46D2fp3/OdvoLy9z3jfsDEbK8QqYTisoyNAWPVgVv/MO3cPXPPpXz5QA/GrLQG+CKCWkmue2iG/jMc2/jXR//EMOjkUzOkuoUZ4NWUqIUpbO1+53kE5/6JM8570oEoo5EQxuMVnqmHaYhU2677bZjurdHzJpf/epXT3c1aucSOqVgXXaplJZ79z/CxmRM6Qr6/X5gc2rRmJM9+GeCSbHZq525789PfN3fbwf1nKWP8LOlX00ZnXSQCHCFZJAucP7O3Zy7+1z2rKxwwa7zOH/3Lq4653IyBDkpGQJVc3phj8CgkEg8ClUHMxJH8Fe2BEeB7te7x7h8ngD4Ogz0sm5ubMLCcPSVYTFbpqxKVCJZQPMv/9nPs5hmHBwVyFxh5oLJGbZcaQa9HhsbG4yKCXm/x0CnrO3dzxKaZTipAXTE2Tz2Q8Zqh0PU8oA+in2MgZyvfP6LHDhwAAZZO1+qZn70dQsHkCpNJgTKekRlcWsjLmUnC4D0JQsinVYK+jm2Qx5Gd+pJDKB39ZdYP3gIryRpLw/tPCpY2wGIphx2LhZqxFR7WY53jqoskF6itcaNSna4hG04ciQJkNT1Jbh6nnFyMyt0OsLKWhVP1oIuEoPDIClxDFjkksF2inNKDlFxaH2d/o4VMmsoh2Mez0bMySZAmiBMeE+qNdvzAUvA9pN7hpueIFMngRJI7FTzo9ua5DvuQ6fMtd4iFptfT7txmzjCjr7vkLTNz7al483z70LiYCsDlSeVgpXBMuft2cl5O85h+8ICl55zIRedcw6X7L6YRTIUEo9BIknw9S7tNF4QqDo9aY2y23giyOF344tYQX3861cGeBZ1SgEcxDDB8IXPfJb19XV8P9ly/TpTCm/kXG6zlR12N+eRna8rD9sqyZ5ScxW7SLGkeFLn0EKDCFuAvqVZXV2B4KaOAafxdVTAM865lAf37+VAMQzX0AUHzzxNSbWe2eyBWQci56GalCRZQlUahHMsOsW151/Gcy++HnwJPgnsqgSXJHz0wc/zW+/6QyY9xYFyhJGe/rZlJmNPKURLKQnflOzpduwqERJ72yWjhJoqsjf3wnowjW2MYJAm3H7F9dz4k9fwvz70bv7ne9/ORrXOrh07cM5RuYpyVNDLc7Ik5eC+A9w3OcQ//E+/xBt/+l+xXOf8jSbICM9rXvhSPvjRv2JopuRiI3o6M5/bIChfNuesFM5UOOfxQnLXPV9ggietq++UoCZBGtFx1cYIxhguuezS4ydAbrvttrY/jraXy82SHya8yyqKd931Uax07F5a5uDGGmma0s8zysmkDRBP1sMvXbCBpd7dkA3b7kRoYXAOr0OwafE4H0qctQgXnMoyGAx47MB+soU+pJqN0ZDFxUWqqkJZT4IgsaAKQ156lpOcnUsr7MoXueM538K1T3kq52+7kFDcE/yUBRJHSUKKbAmPKXEhWhOlZuGU9blPj6J+n8MdI45//WzHjJNN01o9sGpmO0lxzpGKhArF73/8nTyWemwKG6VHHmX4J0IyHo7QaYq1lhKHHI/ZpnLylmiM9zLiiY/h7gI9QzKIzqIopkJw0190UFWohQFVWeLTFEWP3/zA7/Pxr93NaKAxWR0s1A4TQfCvrpZqFMmdRacKYw17ehkVJRkZuUjb+a4RUpuPl04mB6BxyHFFXyaUiWRsSnyiMMaglULU+ifUQcc8ga5q9kYIhZFQScWqcaRJQm4Vi2jUfHrXbDjU3van/fyp6jmsqufP2rUsQ9ZVL54X3/x8PvlHb6JaSnE+Y1wZ+tmA4XCITE9uFYKfC3A3VSnVNvNSKbzzaBlWZzOcIBd8fQvlyb0JYvaR10CfUAVSSclYg6n7LJXvfLZTYPz5jjuN6DrT1O41XgqMdzgXAulMKjQ14UoQGe0vLDEuJsg0oSKI0G+MRgwGA6rxhH6SkSDwRYk0jr5O6emUzMGedIHLLriA66+9nmuf/jTOyffUhIZH4UhJ2rhNTL2roP5/1Raly7nZrHsU7c9v/n5cv0R3IesOZGbHNszn3HXPl7WApgzqQvyHN/8qXx8dZL0vcWkt9tkR3IYzQ/umIfakgKq+KImftpVaWbt0aYUWEm09ZlIG55DFPs5Y8lLwCz/5s2Q1kWedA5nUagwO2pYLufU6dhojwfGaW17En3/wL+lv71MaB1aQy5TECQrrcEnIA9oKO+HrfIFQFSYVk1GJTjN04dg10fz0d7yeDI8WOVaD0sEtdA3FL/3OG1kdSA6sr7GwskxpCvzE4qTGCUEPyRIpoondnJgSqdYhVJh9fGUQIoHKgkrrvqTmYfGQKHAWdLBvT6VkBcUbbv12zr/wAn7hLf+Nh8sxlfBkSjHIUoR3rI3X8UsZh1LPXfvv41fe9dv8g1f8AE4YBqRIoIfgysFOfvK138+/ffv/ZHW0Tm/bEsNigqk1PxbTnHJS0FMJrnbachImtkJoReaDgOuXDz7CI2xwHgP6gkDceI9IgrUvBE0W6x0ySYLWyvESIMvLy63QC3Osu2xXx3AtD1Fy//oBKmtwZiruiQuy4/IkTiTBc31Wt2Hm+wLyPGc4GePx9Po9vNKMq4KNqiQREi2grCoWFhaQiQYhWFQpbm1ED0nmBKo07Bks85wbruX2m7+FZ577VPoooCAdOxbSwIy52h5KqaZXOAkDectFcfNRPM5jxPHD1tdfCTo+uLItMzOuREvNWjniq6v7edPbfo+N1JLkGTLrYW01EzxXc4lmlzm2chqAahfaA+K9jDhe8uO4fjbJAItOUw5RsIrnvR/7ELaXIPIMUxah1cVNd0/NnJaFk+BV2DBqRYQ7QUO3ymOr5etk5WECUF4G7QQfHv9mJ4OOcF7zw/O9zk0g5Oo10MhwDaycPt90SoSnEl9yMyF1GsIdZv4UBKE2jUQIx83PvAH71t9kNHHIVOMFTCYT1Emu/uhWGCi/eb4+7LipScAQb5wC87eXODHdwKKOhRVhLFadh042doqnyLJzuCqO5svj8Zh80A+OQWWFqQxOCKT1VEXJwsIyh4br9Pr9IEo/HJOmKTsHS6ytrrKc9siMQJUWWXj2LGzjpmdezwue+y08Y+dTyAGNJ61TvPCSddVas0HYidNEt21NzhFgj/cYMUOCHI5knZtzfXfclAZyXbdpKB6sHuPLD96HWO5TmjFWh59N3Gz1z5kcCzSfU3iQSlIZExyiktCiFYSoIZOai/ecS+Y0CaFtIgjSyaDJ4MXm8S1m2zZP5+upkFy1+0KeuudCviw3eHiyzkKakgjJcGMDmSVTAfdOftleWzxapBSmIu1rsonh+15+Jxek26gwoUUWT4HDKs0v/tYvc+/+R3C7FlnethJIXRe2v52u16F67aS2gjW2ItFJ6z1svQchg/Cq96G30cOknJBkKUoqvLUIrfEi3D9Z71ilzuO15fmXXMePvPK7+Xdv+TVYqG1/jUfX7zvBsY6h19P8xaf/mpfc/C3ctOMp9Vrvwro+LnnxDbfwf7/5v7Fz5xL71ofkSwNKZ7G1I6pSHYHzOZK++fpYOoY4bHdEWQuJnBnb3odBnfd7x0eAXHTRRT7Lwraec27aArNFROqAISP27t2Lcw5r7YxCrjwFAphCTRPKsIsQTt7WDNr6ZMQgzdEIRsMJY+dwvQSxkFPi8aVjowpq9GZ1g9xLdsuEzEjOWVjh+171Ws7ftpOLd5/LdrlYB7WWDI8kQ/bCItiowrfn5n1c2U6LSbCbDU1X4yYIy6TmUDGiyvv85u//Gl5JhK2onGVYTsi1ihcx4hTIJDqRpJj9uu8krHI+aDEGMklRK5O/8Y/+Jw8c2seh1GN9cDto++L9lARohNGcj2XcZyuaZERuMda6P7Mr3cbzn/UcPvyNu3lovIpLFP0kwZRVHDtn+fhpWlOaAkwrO9ViHpZ7A7AeU0xCYK8UKkvxUiAGKRuAFRmFFthxwfLyMsXqBpk3XJKvMNp7gGc/4zpecvOtPPOqqzk32V47u02rLz0O0fabO4qiRHiPloo0SeONerLXr85c0iXDtpwrRCgrXC2GVFmP3/69t/DA/r1s9BUuAeWmSVgjhNqKNIszRwNEdMgiK6eVLsoBMlQ1lqZCSo1SimI0wRWCZz//BgZp3o5/WT8FAoH37qg6GKf7NVukx63Pvpm73/8OEuERWuCVpLCGnsqY395XdRtd00plfDDA8JMKxhW3PfsWNEGQVNazS+ktX1t7hI989lMMVgZsGEflHFVVkahAeDSFu37u/BKdtDSAk4KxsC0BVVGgVYJG43XWRnkqCW2polmgjQttMVqTGcuORPLaa57HF6+6i/fcdxcbyuKcw+ngfiWsw5QVUqasDYf8+pv/J0//yX9BRkI5mdDLcxb6C2wAL3/xHbzncx9nX1HhKwOSmvioybPD5MFtu6NzjMwY9GLrmtcYe3THnveh2jJPc575rBv8XZ/8lHhCBMiuXbvaPzx/Qo7NgfOwnLA2HqLTwLJqrWdEwU7m/OG28J9uJrbmmPVyNjZGJC44tvSzjDVTUE5KhJL0REIiHHJiWNZ9nrrnQp591TXc8vTruGbX5eSEnbwZxxuvwVomVUma9bBNO1HnpgKnfH9zXHNne6LFFoG8B1TW5w/v/jDv+8KnsFqw0F9gXzUmzTOcNe3uYUTESZkHm2RiiyWhS35svVIoEIH8uHv/N3jHB/4cfe42pJ1QOUOCDLpHfvp+0bggopk/23VRbg5uGlgKvuOOl/OhX/40QgYbvMpakjShclW8kGcxlJtNSBtitZlvpPPYokIBeZbjlGBUFoxthRcSbzy5TrCVRVWOnvT0rOLKHbu5+Zrr+K4Xv5IlMnIUobDfoXFoRGiJRgYtlHY8S5Isn2sV3JyBH6miLeLxw2/BcBxOIHvmx7KEwpXobMDnDtzL+/76r1C7FqnsmESnMzbubc5wBlvf2jqRa2IBXxl6gz5jQqW7t45+mpFUjkXV45Ybn02KxmERKLwLvyyEREh5Ro//ppnthc9+Lv/9Hb9HfzmjspYKR77Qr6stOh+/oyfUuNmZqmKQ97BFybWXX8XuwTJ4SyaSOi6TWCH4zbe9laqfYJ3FVQ6fKvI8x1cldm7Md6tDXW3NK5XAKVm7u1QINIUIkqh9NB6LwLJISgIoYVHOkyCRWuGrAlHruojCsDvLed2LXs6Hfu2zTJREprLdtNdCUpgSn3iMEtz70P18/Iuf5nlXP5OlvI/CUzmDkZ43vPx1/MlHP8D2/iJrlcGrYA2eqiR8jrqFdCtdTQDrHPsO7Mft3l2THGITaSIIpEpDEN1yyy3c9clP8YQIkCuuuGJ6InMDvN0drFcDCxwaruPwSK0RIkwm1gcbXCnEEe1rvukDeMZaUMw8pK3wl1L0lxYwlWOtLHGTCVJKVqQkLQVqNGb3wjKXXXA+33nHK3juRc9EU7KNDO2B0jIVevBgqsCwZimZ7gXNjo49nrU22KqpuDSeFkF8p4e62+cvarXldVuwqiRvfu87We8J7KggsRKVKipbBevjiIhThAThMISH3Cp4DCpaHCw2WM0Uv/3Hv4ffNuDh9QOs7NkVxFHt1Kq1TVDi5Y7ozp+dmEHMBcsCWCDjKdsvwB8a0tuZ47OMxx57jF07dsAkEiBn9bx1OGveeiEuy4pEKhKpcA5s7cqUpzmZDpW6YlyijWdXfxtX7j6f13z/y7n5gmfSI/T5BxcmHyxqOymbQFARWrXAIVqhW0ISGNf2J30Nm+ec5te1rZiQYVlQZpr//Z53YZdyTK6hCDbbjWFBZ0hhxZlJ4vs6jvWyYzFfC2I651BpAj5s0C6IhBue8jQu3XFhqOSz9SauEHjrQEs4wzf2BOB9yaX9c7nh8qfyiQfvwS6kjG3FoJdRFUVdkdCJoxoNmU5ApaWiJxJe/x3fxZLISWmc0AwKxRcfuJe/uOvjuB0LTIYjkkEOWcbGaERPhCb4RlOtSfJdTS7JJKnLLB3rZsxvvO1/8Z6/+QjrrsSmCqsETkA/y3n6lVfxolu+hadecCnbGLBD9ughSBGofo+yLEnTFKVSlIOnXXA5mRMIG8bGuCoRxiC1IhESL6CSniKVvOuDf8GtVz8bgyUhPFspCQsoLtmxh28MD1IJcMJRmAorVdCJmYsX5x+7CsfX7v8Gz9v91CCsbwn6aEzdsprKkHBdHHfccQf/7f/3n58YAXLddddtbs8Q0xKq7tcs8MAjD4NW08nDWoT3OO9BypMaDAs/u4NgOwFYMz4Pra0iE02iM7IsJfOSnpfkhUMXlhfeeBuvf813s5NlMkC6gkUSpAdzcBW9uBxKiHBBbShPkSIM8CB42hHftRYpp4RIt2Uo4lRcMTrHDgkiCNaVFWBUxq/+0a9z36HHKLVgedd2Dq4eCg+7lrH6I+LkLuAwY0M6P7TF4ciPZgESDpUNePtH3slHPv8Ziu19vEwYjkZBdKvWdWjIj2bnoxEBlZzxcVLEkebPZu4U06CtC2sMWmtWyPn2227nd+/6EKNxcMMaFhO0OLutKM/2hFfM7ciLTrUZhJ0/pTTWOqpxgdCKpV6O9Q5zcMjApgy84plPuZq/8+rv5OrlS+kBKY4eMih7uOmANdbgvMfrsMPdBNcKiRSdDUAPzlpkjN+elDVsft1yW/xcu0Hb+YWJrdDZgD/5/If584//FdWuBdbMhFFVMEgHSOfapNWLKfnhOuK7ZxKsrFvKauKnn2YURUVFxWBhGVOUTNbGbCPn1mufVYufgq31HIWQQR/I10zgGdwCI4GBCC1ur73jFXzmjb+M0xphbUji8WgvZohZ5WdjoSxLKTZGbNcDbtjxVBIA65GqJnDxvPejH6LoayYYBkuLDIsJzhiMMZAkszqBc2PfFWXIJ1OJ1j0mGvb7gkkmEAsJo8kEL2C13OD+T3+U93zyo+zsL/KK57yAv//S7ydDUJQFvTQjSVPG1pIr1W5M7FxcZu9oL8YFu92yquhpjZZB71P0Mg4e3OCue7/Elw/cx1OW95AqiVYJEyokjjtv/1b+/f/475gM0qUeLs8BKIqSROkt8/buZ777q1+hetbtNBozdObcLvnR5NTPfvazj+nebombbrqpTdSttbMlPt25RYYk/6v33wdaUdbMuzPhd1sx1JM8gLUTJE60wj/tBFoPUt3v46VCJIpUJ9gD66T7hrzmGc/ld//5v+dfvObvciXLZMN1lvEsyyywpoDethyopExCT2MzRSmCxdwQxxhLaU3bEqS1nml7sdbGFe40IUFEO/5d3RMcJqK/Xf8Gf/yR93Oo3MB5w8GNQ+hBSlGWbFtajtcv4qQHkPNOKoctG/azSWuwyxQcpOTdH/sgyfYlDhUjduzaSVWWaKmmLihianfWBAJd54aIsziL9SFws1sMQOlBWEsf+IFXfxfLOseMi+DQUccUEWfp0ivASE8lPV54pPdoN9Vza+adiTM4Cb1ej4FM0OsFC2uW86uc77rh+fzaz/4i//EHfpabli9hB55tQD6pkMMJFHY6HqVAJAmkKV5qHBIlVCgTJxAexpiwAw6R/HiS4q/5NWRe9LRttfObFzehEiZ43vHB9zHJJGu+otICnaUkSdI6oshOwuXO0Jy+qc70nQoXjSBTOrhw1LoTGsE5vSWuv+JqBAaJJ1Gqs3l8dlQ+NeMqwXHLFc/gvMEKvgwVEJUxrQaFnF/P/CyZhnU89dyL6IekDyqDMyUazQYVH/vy32IGKQfNGIOnqiqstWxfWdkkAj0/xGWaQu2C4oAyEZh+wkZfsd+MUYlC4HEKRD9FbF9gvy/4wGc+we/95dvCc1NUCBcE24WXFCYs2BK4/KJLSAhcgEwCWdFIOHghGLoKvbLAqi140x++Fa1yHJ7SlGRo+iTcds3NrKQ9RGVrgfDQAWHxM2NxS8JOwJe/fi8VNsQPElByeg41x2BdLbcuJXv27OHcPef4J0SAXHXVVWitj5oTWkKA/NC+vTglMJ1kXgl50qs/5hmlRpymy+j6+mFeyvuI9QnusVVuv+ZG/vM/+0X+4Wt+kEvyFQbGkjnY1lsM5TcO0ApTFa2VgZVBzbfCY+vhqZGkaLTSLSHUDJ6qqoL1UBoFtE4bEmTeShsogP/05t9klEsWFhdJlCbJUtY2Nti1axePPvpovHYRp8TYna/06BIj87tmTbWTASYIfvUPf5svP/QNSulZXF7i4YcfRiFQYiooXaqaAJFTAqRxhomIOGy/vtIo69HWsketcMH2XQySjGI4otfrxQt3lsPKaVLazFeyk2zoLA07pcaRCYUcFgwmjtuffiP/5u//DP/oNT/ClUu7WAAWEOQeKCq0zhBpHtqXrQuJyaYNO49zdqr/oRRa6aB9cITp9qhBdsSxr12dGL7rkdO8WrHaLSYXA2xQ8Z67PsjdD3ydZGWBKpUMq4LewoDV1dUZ55d25/4MmnO3IhUdU5JnPB6T6oQ8SakmBa4y7FpY5rKd53Jp/5ya/At2rs654P7p/WGrP86o8e+BypIj6SO5aPtu7GgSSLNOJf+RMClLlgYLPO+a6xF2HCr+0wRRk6jvv+tjfP3Aowy9IVscYESIsbx1DIfDY0xwBYigVbRejBl5Q5VKSi0YVQVplqHzjJE3jIWjTAQPHXiMj3/qkwD0+gtgwRYVWgu0FiCh8paLLriQLEnxAnSSoFQ9DmoCohKeNVsgF3t8Y9+jHHIbeIJYqkKQAgM0T7nkMlb6CwjrGK1vAJDM5b+iQybJDon0yL7Has8c0bq9zGiUeqiKsuYeQiPj8vKRN58Pe+cuuOCCtjLB+6mChzX1boz3rb1dgeVLX/sqaEWapjjnyJKkc1/ESZ8Agm2hx1eWgUqhNJhxgUChZYKuHL2h4XK5yM995w/y/7z+Z7lhx+WhPNIrUEndqyRBKbyUWCFxeY8SSYnEIvEoJIoERYokDx4wm0p+pZSBeY4CqKfJJBiaB7w3tUBaqAKZAG96/x/xuYfvpcoEVVkiKou1NpRvb2zQz2MAH3EKEHcNCeI3B49iTl68mEyYlAUWGGL4mtvHX3/l81S5pvJhUd6+vIJ0HmcsVgbiw8hAgjSBle5og0REbNqcbcedB6WhNAhKnnfdjfjhhNRLbBH1P852KCHbeLRpq0Yrqlr8zxUWWcG2ZIDZv8H5yTL/n1d9H7/wup/klgueRh/JQPZCC0s964k0C+/TvJQOcR4K5WUnflOkUm1twXAUxOjuxE8cYu6lAOHmLDLqsTIpJhigQvF773kXJlNs2JLCGpIsZTge0e/3p7nCGV7U0K1y6SLP82BzWhl6acZimrPx2AH+/ht+mARHimyvbeMIKqQ8auvLGTP+nUd4WCTlR773+1nSGaI0eOewzrXtGnILu24IpCmV5Y5bns+K6tUiskFIFOAdH3wfatcyY1NSVRVeCiamQgixdXvIpufD17luIASyLENqRVmFFh2ZJsgsYViM0f0cnyo2qoJs0OfA+ioFJmTKCpwUbYuNt4FMGAwGaB2scE3NASilkFLi8IGgTjUT6Xnw0D5WizEGaLxoEqCH5MqLL8UVFa4y5GnWVrkcDRUOEsWDhx6hxNbjzqO0xvkpCZJlWWjV8uE9L7/88sc/Pm+88UbfEBdNy8ZW5KqtL/chu07hDKW3GD9bkuJPESGhyWSCShJ6gz6HDh2il+VsX1pGVw5/YJ2loeUZK+fxu//Xr/B9z7oDPVynD5RFQdP02XyWpozXzq2HTTIRxLRAd8u/YwJw+s+BZYmXAgcUoxEOz8NujXd87ANM9FTwqPUB95sZzYiIk06CzAWQIWB0UNng2eY9WS9Hp1kdPEr+5f/777j74W/gUkXpbZjXjEMiyLIszItydqc2jv2IY4YQYAwqzRiQ8q3Pu41d6YC8dheKBNpZPDQ82LKip1P6Wd037gxeCrTWpFKRGM+5/WWSjZIXXvMsfuP/+8u87rmvgGKDQb0DqZm1Ym5LSRRTNlh05staMOlxxW9RD/Wbv37NEfjCgyurcL+MxUwmoBQbxRiRZxTAH3/8vdx36DFG3uCkCLoFQqCExBizqeKjWxFyJq5f8y0VE1NRWcPiYIAZjklKxwUrO1lJeltu4J5141/revPIcvnO89nTX2ZRpSRIki06JWxHUii0w3hW+gsskAThzdqWWWuNB/ZNNliXFnQgFWzdZL/V+BN+i0stgiuKMSYYjjiPRtJTCanSFM6wb7gWWpTr9r1cJdjKkPd6GGBkK6z36ExjPExKg0jC5/jq1+5lUhatJIZzbqajoSHEVicjTCL50w+9P+hfijC9ajwJgvN37kETnrtGt0McgaBrJCqQAivg4PoaHjHb/iamVTjBcdYjRXCePVr16JYEyNVXX92+WbigHfGpWu3Wex/cTAjtL2NbBZ2L+uJ3S9VOBfR6PYwxbEwKksGA0jrMsGQwcVyWLPN3X/BKfv3H/yXnec2y8WzLF8FYdJZRQs1mTfuXG+Y57bxmCA/XecXg7fRHVSHTFINgaAvyfg+P4j/+r1/nqxt7sWIq9OhEZ5fdR/Yr4hQKIg8XWDpAKnxZghBUeEqCjtGHv/opvvTI/fTP2UG6OEAIwSDroV1QhS/Lsp3mmtLa+UUs4mzPYmlF1MXhvi8bT0bLhXIHTz33IjIDudRxJ/0shgQyoVAehsMhk6qkPxhgnWO4vsGCzugXnv5GxY99++v412/4J+xhgKgKdmRLaGRLYnRbV5v2bUOoELbzg7Mbxz3O5G++SiHixK5fYi4ZlG2Fjkf3cozwJHmPEsH9k3287f3v4bHxOqUK84xGhJa7urLIyalg5eGSsTOJ/JixrBeEdn48pqwQhWFRJLzk5lvZJvpBsNOf5eNfAM4hjWOJnGddcTXJxJJ4ETaP5sajFeDktCokk5rz95zDgKy2Hw6CnUJKSix7x2uMfRAClwisd62MwkybR40uZ9vcG+8smU7QCBLjSSaGbGLJnUBmGbKX019aQiLJDKyojB6KnTt2YFH4NGOsLCXgsGSpBgEVhi/dew+FMyRJgnOuJT5szQMIEfwwrRIUCt77kQ8ypEKjobaYVsBll1wCzofKkc77tG6sc+Os1VBRQXT34cf2AhJbezibRh/ME6pqCARPQ6tccsklj58AufTSS8N7er+pRaMRvhEiqN5aPA88+jCVd5j67Bt2dUqS+JNeBSIhCPiowDwpBKnxbDOKn33dD/Pjd3w3y6UlKwGnsMMRvh6kq6O1mYe67TlsblqX6HDMCAjGFfAMQZJALTCkVEaJ4K/v/Rwf+MKnGeWyJT2cmLVPizvgEadaEropoPSEEnDnEHlG5QwVggN+xF6/xv96x9uQCz1sIjm0tkpZlu1OQFMG6cVscNUsZlacuXaCEY9zAe6snfNj0PtQ4uqqElOU5MCLnnUzSelQxsUKkLMcWmuG6xvkWcYg77G6/wCZF5y/tJ3JQ/t45nmX8m//wc/xXTe9lNQXJKZkJemFJM/PznXzYtBNK3f7Y1uKI22R6InDzKsR3/T1y3cvuSdouNTkPQKGpqAg6LO9/QN/zlf2P4xf6uGkCFoWDigMElGX8YeEtSHwxRlWdTa//ioXXq14uRbILGEyGrOU9dihe7zsubeRI1udh7N6/AswpkLrhBR4ybOfR6/0JPUm0HxCbWsheOkhcZB4wbk7d4e1z/l23jHAvfsfYKMqqIQP49H5Nvfumoj4ZgPBb5Faeo+QEkGwy81RpKWjP7YMJh65PmHgFdW+VZKDI3aUEvXoGlfvuZB/+cP/BF2/U4JmON6gJxSKujAYz761Q3gVKi1kLV7qnGtJDOE8tjIMVpaYOMMjawcZUoQ5tdaMAc+5288B68K0KsKzJ5zfVNHVHbdOTuPIBx99eKb7onLT9plGm6brqPrsm59z5HVlqy+ed955LQGilKoXhqAD0lSACCmQSDyCh/Y+ilMiqGcLkLVM7akS9EqgHFesrKww3lgjVwkrOkeubvArP/dLXLdyEQMnYeSgr0CCWuzjBRw4dIDtK9unffLdSdfPJRJi88tvJkcjTs8lBFQoryoRrLLB//M7b6QcJFSiapM+I5uJyrc7CG6Td3RExEkIHrcKYFo7rMDcOmfxWjPGIUSfd37wj/ji/V9jtQ8FHuEd/byHtZaiKsl6Odb7NrFI7ew82bSFifgMxLHH5uSzGRNlVZFmGaSankwpgRdc/2x+58/eztr4IKKv4iJ6lqJW30KnCbasSHVCL1+geGyNfjrg+Tfeyk9/z4+yiz6Jr+iLNNhLOqgmJUmebupVFoAS00R63lZVis3jtjtkhZg9EuO8b34QP5fM+/nrnaZQu0RInXOQMfurIW/7y/dQ9DViKaOaTNBOkDmB8gJhXdA28D7oefrZP+f8GfYciemmXPNZHVA6wyAN7RI5iotXdnFJbzeaYHnbHdxn3fivP5RMdD2PGG668BqesucCPnHg6+hEhRiozhGlgLK+ZI1blTcVK/2FsEduKqROsTgmOD775S+CktPuCluTCkrWebjDd8a/8lsI/gqoKoOWiuFknYGT7FF91soJxnl29pbZeHTI+b1FFgc5WQUv+/ZX8p0v+nZwY3oyRzpLJjSJ0Ejr8M7hEs09j3yNiXA4KfDGhLGjJMZZHIEjoKgCAaMkhfT4TLPhSpB1dRYOh6UfPHACtyCCK6tEtISjDzRC6MQW0+fPCY8V8NCjjwSCkqm5iqOuiJEyWJLruk1bqKNa4W5JgFx00UXtSVKTH9ML3R3xoRfn4X17cTL0AIhuM507NfpgHJAt9Nl/8AB5ktK3Ar+6zq/+8/+bpw0uotf80KCPGQ3RC30QwRd518r2dmBv+WB4Zm0VmBIfrjNJy7g4ntawpkKpFIFgguEPPvo+vrj6COVKD5lpqGyrg9ANqhqrPhtvfsQpQn5sImUFoQLEW9CybveTPOwP8Ad/9i7kYo+sJ0AYEiFRQtaaShq0YmNjnV6aodzU8cXVdrhVTcYnkfw4a+HnAhXYTIalWYah3vUitJTu0Ss884qn8tDnP9EGmBFn4fgRUAqPylLMaALDgp35IucOtvPcp17LT3/Pj7JCTg5kIgHrcJVFpsGNDesD2zHDYNCStmouiduq60UeZkzHUfkkrV9+Nq7edE9s0E9AKgosFQpJj7e8882s+4oikXjhsMaQJIrEC7TQlNYhEoXzIcFqNjpbIkRwyrhYnkioOZtWmaWMq5IMcMMJNz//OnLETLWeP0vHf1MAIJSCupojA573zGfxqffei0S37SrtcBTTFirlQBrHQq8fxpIUrZisQ/CF+76K0ArhPaI2HFEEcsACXgq88CGdbl7NRCWneXmSJCA823tLfN9rvouXfOtLQQpUkmCMI8syhBDsFMsoHJPRGstIlOwBHi01VA6paleWRDLE8aY/+SMmmmBDXhqECmt0t0NEekiVZm1jg36iGBeGsa0w0oDXbbApgH6/z6ovkErirUMJ0bYKOWYLJ1rnL+fxQvDYvn2h/YWOEG9zCaTEGYOAIKyq4MILLzxWXnWKK6+8MjwkdSmJqMsZ3BYj3gMHDh4MDisNQ+h9UGVu783Jzf6shKE36EGP3fki/bWS//BTP8e1g4sQrmBSFVgNVjvU8gCvBGVZkggVJsLKTO9etzRSglfgZXg1ZTq2c3Ej+XFmLMAqS1ldX0MAX9//AG98x1tR525H5SnVeNIK5HrhsdJTqakV6KmmhxNxdo7hRsB5S2kiAZWpcAgeWd9HBfz67/wPVs0ksP8+9AcX4wkbGxtY75BpQummvvDahQqQ1E37i6vaFSYSgGc3mpLVTS4a9f83u5CWYLsnPGgst93yPHoqievnWQwnYCwd66ZgeXmZ3AnyjYof+bbX8n9+z0+wk5wBIKuq3nSTyEaYUAC1naOXWwRjdUYtmpef7rAqNrfKHO51lOwp4kSsX8x2mLtuTC4F3lQhBkOxxogH7GP8+Uc/RLZ9ibF0jJ1ByiCqLCtLLnXbvtAIeHf1quarQc4USL9Zc8ErGSQMarHOW296DnkjeGBMsBw9i8d/0AsKzi1KKVI8z73+RgZ5b1OZUJPEWzmdTzIkg34/7JfXlSQCqBB87ZEH618M2pmJVEgE3rqWZOhe45l8omWiQtuMsxbvDCuiz2UL53JpfzeXJju4ureLK+R2zhfL9JAM0Ozqb0cjccYi5suChWcDePtXPsaf3v0JhrKmeGQQMLXe4UT4LIWp2iqOqqpASUoca5MRvrPjFjpG/Iw1rTHhmeyKDjfEnOvYUTfVSwfXVhn5Uet203AUjf5Hs4GitUaJ0J595eVX+MdFgJx//vntvxubMTHzG35augocnAzDIHAe64N6rRHhKIRAC3nc/XRdN5muB7wXm7+n/Oz3ATIv0MOSfFjxiz/xM9xy3tNZwLMkM/IkCwNcCApcEHtJU5QOAi6io/I747TVmQDsYZ7zmPieQjQuWzjZ+dnv2fl7GZrY8EC+uMQQw++/+x2MVfC9LqoSqdVM76jcon/UxTsQcZKHvzvcfCTAFBOSPGdtvMHK4k4+t3ov7/vUx9DbFhnbCutqS7YkIU1T8jwPotIbGyz0BxxpgRHRxSOCI4if1i2DSoZgRgkVnIhQPPvip5NNLLIpjz3M+n+4r0ecIklXk0jIaZuoZFoi3sRs3Xhu5h5WjpVswOSxQ+zUA370Na/jzhtfwgKQ4lDOhR3Q2rFvXBXTkn03mzxvORD91sf5li25xTHGeE/uOBKde9kqACiJSHQgUPEI+vzn3/ktJpnk4bUDCBFK7nUadrcnZdnqAUq/teDpmTR/dHfWRV2V0HXXGg6H5EnKYppz5QUXc1FvDwngKwM6PevHv0BQlGV7DbWxPHXlQs4f7CC1m+cu4QWqw9BJPJnUrUkCzmFcMA3Zv39/0MVoiJS6Hca5hgA5+qLqnENIGdpNKksCZICyFX2CBW3iHAMk2lYoa0gBbwyJTvB4jLOBLE4E69Lz+eH9/Ie3/iZ2R59SBNeXJEkgUaHipdYCMUVJohTeWnpZhiTEiaPRCFAYV7XX0OEZLC22RRGNm0yXnNtqWkYKDJ61csy6KUKjm3cktam5FaEAQ8hQVZLU5J1zjosvveSo61KLiy++2Kf1JOG9J0mS0Atkg7KqcbatCxPABHjg4GNhcnHh5ldaYJXAi9DEI+zxpX9NmVbjtNF6LjMVnRRaBVE+48mEwk8qelnGyJRkUrO0Ybmwyvl7d34vz7ro6tbfXVau9nwHTdA1kTJMpLarTC9m4rVuEchUFJVZN7WogXoKZX+dnUa7FSFRf68CJlhG9UOLcTApqYAhjvd94a/5wGf+mlQniMrilcRq2Qo9KidIrCC1AuUEVgiMFDEYjzipw78p0ewqwHdJCd1PqVzJQq/PKuv8i//y7xlvS9lfruOTQPA1nu9eQGkMwsNC3qMaT4IAsIRCw0RNtT8SC5mNScLZTny0u+nzwnn1YuktZGik83VII1FIFkrHt153M8qD8Q7pwI1LsiSlNBVOClSahIoRC7Ke2K2YJtoRp0DS6j1eeAodXrYWDe8Z6FfQQ4Xed5VQeE/pPFZIJsaSqpTBGJaHsNv1eMPLX8Nrn/1yBmgSJMpJhAgmt01SnPeyacwmt6jEPcw43CTwuMWPyC1+RRyF4Is4AfNIbTjQkGXNZpUVNcvlLBbHiIovTL7OZ+7/MmuqgiwhRaIqh7WWEodYyFk3BSoJVSC6077ZJVAfjwnQKbv+i+mOetM+oVrCJ8y1mc6gdMih4ZUveCkpoLxEqBRX2Rny6ewc/xKUpkAyKiuUS1koLT/26u9jUEoSkeCkYgKIJCUxgqTyCKmpvCNNEqpyUhN3jUCkRCJZSPqkaIQT7b0qfQXKo2UgS+Tc/XRzc5hUKnQqlIZMBqHWDMitQHhHRdARsdWEFEikBgda6dpJRWGlYl1YDuD4+IGv8g9++Zc4oCpWq0kgGoyn8p7CubotB1zlyJKcsjBolaJ9cFdKnUBJiUHik2kLjEQyHA4RSiG1ruNJcTR+B6UUJY6h9vzt1+9Bo1Fe4suQp1VSMKlJidBKVEsPWMuNN9547ATIFVdcMW19qU9Mylr9NfzPTKXDI34fhXAtq+i6pStiM2N4PA/xzEQ4pxRr6xIY0QirCEE1LuilGaKy6PWKVz33Rdx67bNYIUc6g50UQTO3ppPnyx2PJbA73Cvi9EgMW/LD2vq+h+IqLRPG4zF4D1nKkJIDTHjjW38H10/Dg+s9QquWhGvGufJhMW17LGP2F3GSoWp18MNOTnUP9AR4/12f4IG1/bDYw2iJytKZeXG+HLMrptYkns3z0KjNxwqQSIJsSj4731BK1UF1KMadlAXFZML2tM/33/laEiPpJxmJkORZFmKLRGO9YzKZtMr43QSmiUMiTi5c5354uruk04pJaueD9eEG23bswOEZjUZcdMGFrD92gF16gd6G5R//4I/x6lu+FeGCXoEZT5C1SGO3ysN1x9ecBe4TDeDEYY4RT2Kw5g+XV3sqZxhTYdH861/5jzy8eoBkadA69KkOOdq83Nw61l3LzqSqXScOkzvVR+1AG08PxTWXXkmKBGNDypokM8/P2Tn+PYlKQpV3rYMpZcJlO8+lVwpkZYPjqajbOqxHehV+VmpG4zHrw43p22mFVpoUwXk7d4dCgbrqozQmbLxLGeiSOqGXnXvptnoItA6vOkd3gE5ThJBUVYXDoZM0kCXOBd0cD4kI+f66rxij+K0PvI2f/Y//igOiYs2Og4SHF9N2qUaXozN/J0lCVYVWGF2/enXlkK2FMiuClMTaxjpFUWB9qHAxxkxbXcTW865zDqEkpYJHD+6vK0CCjkqY4oOeSrhAvm1LSpKE8y44/9gJkOuvv3760HTsj1S9yAgEtrae8cA9X7t3xqbnm8n+zpMh3XYDYwwq0TglKJ0Nfsoe+l6hRiXPueY6XnH7HexSy2iCL7NEhO2Bus2ne0HmPZZjH+cZEH2Lw+ix1AyqIqgCh45zG3aREkEpHQWa33r3W7nvwF6GylFIH3YYpaBsvKjnxqnsjNGIiFMh+Wx2zDyh37d9WYeQmq9xiN981x+QLQ4YHlpDJZqiKmMLV8STAmPCXJrneRtX7O7t4JylbajChEBOCoamDMJxzYaH3xzod8mQiJOLhhDVLlSFybpirFJgFMFRKstYXl7moUceRmrN9pVt7P3qfZzfWyYbGV738jt58VXPZYmUBZmGOK5OztgiTPNEpuKMw5yddlIfkSDSFEfCR+79Gx589BGyfi9sUok4ALZ6Hm2HiEyRLMmUF9x0C9tZCglmkyhHIBAoBBrIk7Qtpblw6RyecdlTkDaQFJnSWGvxMtiZeF/bwiSK0WQSWmVq95Qgduq46qJLQ4KvFV6IQAx0fl/Maba0Iqgzbf0e74PTSqUcq5Rs4FkHRlj6SYZ2wXa3AIySoGSocPdh0+G33/I7/N1f+Gl+50/+iAO+YOwNvTSD0rRrqegIu3Y3HLQOnxvCZhuVZbk3QAO6JiMEAotlPB63DrNKqZkKpWa9Fl3B17qSo2mVuf+hB6cEkJRtnDCdI+Q0ruXIQqibCJDrrrsOXwuqALP/Zvq1IIrjueerXw0nIbYmK050EO+2IEbaMpk0wUgovUVISV+n+LURFwy28T3f9iou6u1BYvGuChc4qXc25y7gfDVkJD/OkCzwSAO/Zg5TJBmSqphggUPjDcZIPnPgK/zhX/wZi+fvZt2WuFSBVoyKyYzv9Ddr/EdEHBf8dDO0JUHqRRYhsHgK4Lff/TbuWX8MgyfLMtaGG0EcLSLiyZimO8lKnudhR8yNuf1Zt5COQyBmcBSmwuJRQqKl2nr9juTzqTH1dILbRii5UfwvVW0ZWZNZZVmSp1lobx4XbEt65CPD7dc/h9e/4DVkWGw1IkcGfQKxOUlzW4VsMQc+rdeu7n1sco3GsQUPlbMUwGMMeeOb/wfbzttD0ss5cPBgm5idtXPqvHsItRZPLdQvfWiV0BPDnS/5VjSgUCDBObspPzobUwcJWGtImqlEQWFLEuA7XvIychTCOhKlUSLYwVp8q+OhspSNySjksMaGKg+C29l1Vz0t3JuaEBBiGpO1Qqidyp2tHUkFKIlTigrJY+WQu9ce5O7hQ3zl4MNtTjxyBWMMhjr+q6tZJlXBuCp5ZO0ga8LBQo+JcKyurrN72652/pY16aE6FVUQ2lNRMnxe63BlxUq+UEtLBGtMhaKkCs+jFEGwVAeh0k2aT3O5lKivB0Kwd99jgYoKVjRtzNC61UoZOIn6ud+xY8dh7+0mG9wrrriiVZ5t39gH0VPvHF6JcDFE6GW6/6EHEUq2qqzftAd4bkFtoOpaH6+CMEvpLEmiwHrcpKQ3cbz8xbdy0zlX0wMUMnxoY+seHVn7h8+uk3JKrUWcQevolr3Avh4PUtU74TDIepTAugyKIf/199+MXe7xjY0DiF6KUwKvFJP1EdtXtmHGxczENGOj5uMuZMQpEkC2JIioF4BA5Vut+ci9d/Fnn/ooo75CKYFCkeZJ6PesFcojIr4ZMMagtSZJklYYrTkuih533vpi3vWX72F/YlnXYJ3FOYtGTYVPO7GBcnHOPdVIEOED+dHoBVkBpt6JToRiNBljtEJIySDrIYoxyXrBLU+5lp/8rh+sBU8VQiQIBwgNRQmdFr15IkSJGMOdScFbQ2WodiFztRujZgS862Pv5+sH91KWCbaXoFJJ2kvbyrKznQhp1nDbtGZ70B70uOKc/nYuGZyDri+297U+Q6ygCdevzg3wHisFTioUnlsuu55zlrax/9DDqDwk600lh6rbS0pneOzggaCHpRSuvqapl1xz2ZWhysEYnPRIrUI+7VzQlhRiEwc1f0csvtbEcexzI970jj/kLz7xEca2YlEk/OS3fQ/f9vyXkMqMitDxUFpDpnQQE9UZb3jDD/Chf303j5WHMInEO8WiXmT//v1kSTZdW+fyHC+gsoY0STDjAudg22CRZfqkzeTvHSg4MDows+Y75zCy1ggT03m7rQIh5E9aKVxl8EJyaGO9ZhumCbqkqYLxoYlWEMau90ckQDZthO/YsWMTY+rmGHYhg6qUwbPv4IFW7+NIcCfw4Z3v2wsbCIqyLKlwwWbIOhgVPOOCy3jNC15KDjhTkCHQQuHKKpAfor6QYjZXmKn+iAvoGZUDbjlUlapJMUJZGDDBkOTL/K8Pv527vv4Vyp5mIj02UQxrQaM8zymK4rAbTPO+1hERJ+cB8JtKxRtbcw8Mgd965x8wSsKOrJUwHI/oLQwoyzJev4hvKlo7u86OjhAiuMgBF+kdPHX3BWBcqPxI9HRerUtsXSeAatoPI069JEzNJWJGQonD4FlaWqIYjnDDCX0rODdf4h+8/kdYIUOUEzIgVQmUdUIro8DW2RbH2ZngKvyzAh5jnXd++P3o7YtBmN45er1evGjz11BMLX+9CM9jWjpe9cI76CNJ62jWOIvSOujjneUxrIBQaehhMplgAIHCl0GL6LrLriK3AldU7dplGwtbARNv2bd6EINDKYWWuk09l1WfbfkACoMxBl9rgTjnZqrL/RHcq0Tds+CR5HKBKpUc8iWrynIocfz2n/0Rn77/yyFnIcFj6/YcMMJjcfTp8frv+V6UcQzXN7DWMRqNWFxcxMnD63A6AbZu9ZEiuN9ce/XTSZH4JqfygaD5xgMPIFWo+rDWzqz3bot8qcnvlRAI5/FSsD4ZBSMLQUvOifoa2M5YlXVV4Z49e46dAOn3++0JzRMhbWCCCHkilo3xiKrWBDlZgmPCB5LGifChq6oiEZJdC8t85+0v4wK5QoZvdx/wBGGfupLFWDMzsc5wHlHV9IxYNF0n4Zu3vg3310FVtF8fmxKL5kEO8vvvezdqecCjqwdY2rkdR7B49taS6gRf98jNVylFK8aIUy7RrF+hgdGCtYwwvO2uv+Sz++5naMZsW17k4OoBsoWc9fX1TS1eEREnfA0XIgR/nYCoqQLBWwbAi254DqI0IXCqHdqCuK9og/nufBtJkFML3TWyCaCbZMwqgUwThusbXLBjN8W+Q2wj5Z/+vZ/g/GwHZrjOQpKH9bk0U+Ij0Xhn6pbszn1nTsMt4oyAbW9pR6ZUwDolf/AX7+aefQ+wasfky32csGyMRkwmkzi/+tn50M21pZ3bX+bOW25nEY0GnDNIFdoM4gYIbVcEPugOtSKjUpHguePZz2NPvgjGtpIRktqdFI9Xkn1rh9hXrE7zTGeD/S2eS3edS0+oUPWoFbjQ9qKUwtS57UxuMUfOSMB6Q0EVrKAViEGOWOyxqh1/u/4ov/vh97BarJEAuqzItaIEjFRMqoIEeMGlN3HHs57LrnwBSsM555zDvtWDrWbMVmSar/NuYwyp0iRC8pLbXhiqOupcyyuJBe69/z5cLXre7TTx4vAkj6hFsoUQeCUZTsaMKILoaW1lHZzmOl0rnaKNlaXlYydAsixD69nOGCVnXWE8AoOjxFI4g/GzN6grjnKiGmMaUdItPeKBqqrIk5Q0TZkMR+RCc/HOc7jjGbeS48kQZGkWrIKMAa1aNkSlSWuN6uZpv0iAnB0TXFVBL68Hm0TqlDGOX/ntX2NDOTaqgrQXEkLvPYt5HypLNZqwMBjM7GrNWKhFJ4KIkx79bJ7SQnLgKLxlA8Ovv/2trMqwq7qxusa2bdvawCdN03gNI765iY2thctr0qMoipYEUQ4Sa3nxjbfQlwnaC5yxgRxxQcC3QZTsOzXRiOZ1d/i6CdnEVmHHzjqKg2tcvv0cXvvCl3LD+U9DY1kZLNY32EOig4Bfbd0nlNpUqLtJ7DwSIWdGnDZ/K+tYq0Lz9r98L9m2JVymObC2SpJlOOdYXFyMF24OrtPOkFq4YHknK2RkgLAGLUMjzGQyIY9VNKHCwnswBqE1DkfhCrTWKAQ3nfd0zusvk3iBEqLNB4QQVLXo6f71VR5ZP8iopvGkC7mGQnLJ9j1sVzkaEcS9AU3Qyaismcl5N+UTHoTzpEKTkYQYbjxiVBVUwrMuLfLiXfzxxz/Ix7/4WRz1mumhchUO6CU9BJYejh9/1etZqSRLXnNw/0EGCwtTd9c58sPWJqpJklAVJbnUpEhuuOBaGtsUBFghKPA8sPcRKkIVZ1dntEt+uC0qTZxztaOOYL2csGZGLZFkraVyVdte1LxX07milOKWW27xx0SA7Nixg+Fw2Kq0WmsDm+LDIHAtUSBwCNZHQ9DqiOvLkxGUaCFxxlJtjNm5sIwoDd9352voAbKspjOn1mEBrXeQvBQt+dH8iGMLIa2YxJ72C+fhRqfHIRKNm0xaefEDbp0Hq7188LMfxyTglUALiVYqCLhVhkwoeirBzel/0An02jEVx0/EyYT0IBzeWxSOtJ7SbJLy5ve8g0fsBmUmUF4wQOPKqrU/j/3TEd/0BLlTZSSlJMuy6fekQhrPrmSJG5/2DDIn6KkEX1nUXH/6kYTUIk5S8tARqncdi+ymHUb6QGKgJIkTLMuMp+++kNff9h1kgKZTgSbFlMFVda+3FLPW3Ew7myPOkITdFAThF1vHbIAParpj4Bd//d/iBwn7JxsYCdlCn3ExQWpFaaq4/FNXyeNRSRIcOBCkTqALy0tvfQHOjIMAqvVTq1slY+4zXZhC7kgQ9uzLDFtf2xzPK577AjIn8JMqiD0LxfpwgyRLEUqS9HLe8q4/oiLk1VIFoVnjDT/w6u+Cg0N6OmVjdY1+kiGblg4ZKiRa555uTt0ygqFVNAjYQj/L0VpTmAqRJTw0WaXaNeA33v6/2Wc3SHqhmi6XCR4DOFIEfTzn0eNnXvej7PQpifE4IXA+SIwKFWLCSVlQGoNOUyobqjJToZDGcd62naH6o2VnPGMsBsEn7vo0xjuU1sHBrW6FORxB187pLlSLTEyFSBQP7H2EsjbYVUqRyWSGzBBSBqmO+hRuvfVWjkqAXHnllR6CrkFXAFXM9Vk2BMFauYFp2gGaUpgnadFx88GO82QqsE+yMDz9kiu4/LwLSYAkSTeLe4hpP2GX/NiUMMcKkDMsGjvc5CaQgzzYNOERss9P/fw/JV1Z5FAxmgZsLrwSVyshe5ButiduKyYzIuJkoyxLqtqyuRpPkEJxf3WQ3//Ae5joqSJ8M6ZVx0UjJpERJzcDCpaXL7/txahxhSot23sLVEXZ2uPNxwebApyIk5c71HOJa8QX5wiQXCdUwzHbsj56o+CnX/8j9IGyHE5FbsXml6XbFtGpbos4s8ZPkjCZTJC12sG4KrBaUUjB5x69ly8+dB+PTdYpcZTeYp1rRZXPdheYJo9L0xSVaMbFBFOUJF7AxoQ9g2W+5fqbWNQ98GAr0z5DMmrsbMobYWq/3JAROYKbnnoNalyxbbCIMo4sTcnzHOMc1lrGruKR8RpjQCiNnRQgQArFtqTHjVc+jT6KXGrMpCDVCWVZkqZpu541FSCbwrE6gW2IB+mmsZuV0Nu+zEFX8BBjfuG//wojwHobyBI00nkSJJmFzBlecOV1vPjpz2ZJpojSgPc4a7GVaSs+pFZYH4TKfWXoqYSkdHz3t70KWTMDaIVXAoPis3u/BL0UK2A4GQd9LyFJpJrJn5ib79v/r10LDZ5vPPxgm7sLIbZM07uk+NOe9rSjEyDXXnstxph2N6YRYmkIBmoWKLwEDz+2t1af9TPJnvRs+YG+6SRIaRgIjRwWvPDGm9kjVkLNyjGQH1txHVvaqUWcGbyHaOqYpvPHRjmhVLCB4y0ffjsH7YShq8gX+u3D0iU9GuLjcOMyIuJUQVEUJGmK1AkTW5EkKQb4D2/+NR5w6xg1O38nzfiOKtARpwK0xpmKG897OttVjlgbI0sTetvrIL1LPHsRSZBTEc0Opu9UgSQOpHEkxuMPbPBT3//DXJbvRnvDcjrAh1CaZv+/+9qqFz7iDAzeBGit67pzSJOMNVewjuNNf/Y2HhmvUSlB2stRCJy1dQWjwMQKkDafsz60HmghyVHkBm555g1sV4uhzso5krr6rhsbn/UEEpvFOXUnyVbeccW2C3jqeRejxxVmXOCNpSzLtpVYpglf2vsA97EvOFSlCSUOgWCJjBde92x6pUcXln6SUU2KqQ7WkZKaTuuCasiZrvWxh7V9B8gHfR5gzIce/BJ/+bW7qFTtemkhERLpPRJNRsoKmp/6rh9kdzKgZwS5CBVDSghwDtERMtVSsZT26BnYnS1wy5XX00OTIDHesmYmrFLwu+/6Y9bKMWQanactOdkQbnLuendJkKZNxqlAgHzhni8DsrW+lXPXZf6KXXvttUcnQJ7znOdsutiHYwAtcN9DD0xtcb/ZD2/D9swRK7b+sFmSYEYTBlZy3sI2nnfts0jwwYO4q2I8X0LErNRH3QExs5jGCeDMmsi2qgIxeESaM0TwleFDvOntf0Bvz3ZINWsbG+2k1z1KP1vxtGX5NbOaOBERJyV/TBNKZ7CAUAlGCz5896f40N13YZd7WDHdoe32B8+P8YiIk5UEee/JEdx+03NZFAl+Y0I/y6mc3TTvuihAfeolYHMEVZcEcetjzu0v84zzL+UVz3jB/5+9/46T7Lrue9HvDidUVafJyIkACZAEGAAGMOcoiSIlKtKyJEu+z76WZFuW07VsWZZl6Vn36drWpSxbli0rByZJFEWKIhjABAIgCBBEIIgcJ/V0d4UTdnh/7HOqTlVXT8D0TPdgauFzUN093VUn7LDWb631+xETAJFJH7/pC7oNwA9xtGBhZmek5WURFEnyElUDYFJz0xPf4Ibbb8a0I0QaWjuEB0o7VIcSZ7mMqwQipSmKgtwZVByRqgiRlSyqhNe9+GWAoSgrslMlqxYyN6sAYUJEQTQWoYaQgvCQIPiBb3s3rPRpxcmQ6LPdblfSt3Ag6/IXn/sURXWfCx+qkzTw2he/lJ0+Zm8yh80LdCXv7ZxbB+y6Dda1afGG8NCJEsqyRCy2WU48v/K7v8khCvo2gIO+NKG2qqKL0E4wj+L//L4fYt5K4tITCUmkNDg/xAmcscRVJVF+aIW3v+K17CRFEDg5Su8QOsWR8OChJxkIR88UeCEYFDk4TxrHUwsmas4RL0BWssyuAkHue/ABHEFqeBqYMZm2O+ecc44NgLziFa8YI7yTUgZCVO/HtKDrZ//gQw+Fr6sepWk41alUwhiWA4kwwSMHUW55/YtfxrksITDVzfDrwA/fuAG1MkLEqIdqpnvwjHO/AIerOD8mzeIYUNIDfvsvPsCyL3hs+SBWgop0yD5NOG7NYLFJfFprWNftMmrGzDezLd7AERKEILc5Blih5Nc/9HuYxZSuL0eBiAyylE54vPBj/fszm9lWgR/g0VFEiuB73/bt7E3mmNMJUow7QU3nxzUC7Jkbv7VWV34MW2Eae6ZysCtu4w+s8pPv+1GkG5ASSqMlYEy5LklVH6Lhr01W8M6U2J45+5eOQlySCIUC1vJVSgS/8cE/IJuPGWgYlAUmL1DWo6VA+FAtMinqcNYCIVIitAIhsHmBzEou230uV5x/ERE63KeK86MsS7x1SDFbOWFUcTa2wTT2FV9RE73q8hey6DTKQeksshp7EkG/yJFzLT701x9jmT4G0CLCElpRdkfz/MBbvx2O9NBSkRV5kHF2fggQiMmStykgSL331a8SaBHUalb6a5TtiHu7B/ivf/XHlHFCiRkq3CDlcCG1ZcHrL30p73jRq4gKhxuEtinlwFuLRIR12DjmZMRzzrmQt7/8NbSBxFaVKCqixPOFh2/hYNbFxoqBNxAFcEhLGbi83PrrG6sUrGRwXaWoc+DI4Yaajjsm71MURccGQDbqk/HWjr1x/fwfevSRIUHLqY7xhhwjjNRg6gDUCoIEj9TowvKm619NAiSEwRcm/fp7s67qw48P7tr5n+2fZ77/LCegkPHosCb1VXzirs/yNzd/Eb1jDjXXIreGNE2DVJ9sEBFNOSZbX+qWmZkc48y2fgN3CKFIVILB86Ev/w1fP/AIKyZHxdEQSK7Ht5GNFoJZADmzrQ6CqpRWjOBCvYcdukVbaPL+YKpzY+WscnPbPLuGryaqPbEGRVy1P8eF4/qrruH5Oy5lSab4inxfEAjua5BDToAdsun4No460TWtVWZmZyYIYqwBpcFBK5nji9+6hdse+iblXEyhBaV3yIoLMFERwgeFiA1bCM4iy/McFUXISGO9w2YFS0mb1177MnaLqv1FyJAedA6t1BA4Ks/yFqIxUYxm+0AjPrTV2hY5w3VXPBdTFBTWEFXcNcYY0naLXHoOZz2+cNvN1NTyGlWBKY7vfvnbedbuc7FlGfyx0hCJoISm/JS2JDF+DBP8oimX6zFZPgQvZKSI9i3y+5/+GF898iB9JUEpGORhfgGDIqMVtdA+54e/47u5/JwLiLzA5SVSBOl5XSm0KesxKz2+/9vezUXt3bTwtGQUeEPw9Cn4td/8DfavHcEleiiaEscxSkiKQbauaqUZR9U8H75q30IKirIcJva89+vwiUlrkqxvCIDs2LEjbEwVaVDN/j8sIXO1Anf474mnnqx4WE/PBjptUbSyctytRSPYM7fIs3dejCB8b/H4KSVw9U/WAR+TEjCznfMZt5FOInmi+rIEfucDf4yfS8kU5MKRtFIOLh8eA9uaQEjz8BMZxxrRlLP9d2ZbvIF7JEcGqyjg8ZWn+LUP/B5mVwcZa7CBVd+KUeWHkaEaZDZ+Z7YtTEosHg20gJe94EVIU8ngSrGh31CvxzPb4jWoUe0h3cinM5UHupS0+ekf+XtIn5NU5c4i0vT7/VB23UxK+UbVx1EoijwTmduZnbkBfJmjK/ADDyWWP/6zDyEXOyyXA6wWqEgTa01cZaZxHuF8ULE82/3eKng0zgbySQSLSZtXXftSNOBxlM5QOjumoOFnAdBR15ERTYLEA3NO8d1v+zbm2h2iNAEpgrw7AqUUuS2Z372Dj33qkxzKVgGPAkxZEiFJgZ/44R9Ha83c0iLdbhclZOAedKMKOjEtoG2QQtuJWGVpaYne4RUu3LGH3uEjPL5yiGw+5pd+/zc5RJDZJUrAOpwpEUnK4WyVBZFwodrBe972Ts7fs29IlioR1TosiIXiwj3n8LprXkYbiayQ7aI/AOCh/Y+yZnLSHQuoJMIrSbeiFXDO0Wm3h9ckjxH31z8TWrFGDyqp4GPF6mmaHhsAqSfKcNOqFw6lhtGirCipchzLvbXw+86GNxKucqLHnY/NcEAmS1qHz73eECtZp4t3n8McihRVsfR6XDU4xcSFyxP58FM4uaYdUwGZsV8Y/3t7lGP0Z676m7PrtSY7VUzwBgnwEjJgAPzKH/waqzanZwtK4cmtoZcN2LFz51Tn+kRKbCcrUCYXseZ8UQ1yVSOhqK7lTL7/Ycw2XqeM3eE/bWf24Yk5aKrDT0IOU693615LPElrnlVKfu8vPkSeKnrSURgzBLgnUXg7KyF/RjlxY1vHxA82+jfPRAbdb/y3/ij/fnLnLsOaiUQ6SwS8+01vp42i3W4zyPN1foacoso1s600UT2j4KnXvlzsoF3Cd77yDZzDIh1RETBWTm273aasSATXyfT54/nU01TBe4yEmT/GwVFeZwZxFJQwrCsolecvvnwD39z/GIUCEWvyshzGL0VRkOd5KLHXetYCA7TTFs4Y8kFGrHSIjwrLhfN7q9Z/SSQ1Uqrh4PNlkDdVOtree9lRYrXNoHCv1xA15R+aVWilLWjrmGsvu5p5GSRkeyuroX0jiTi8ukIrSelnGXc8/C3++tYvkiPIseg4AuvQpuC6Pc/mO1/+RgaP7ufcPfvIi3IYV9ciDKNLdZWgQ4i/XSO2cYLhmntkdZX5xSUO7j9Au91mx86dZMJw58Pf5E8+9+dkVIGHs8hIkZuMHekCGEMCvOVZr+aqvRfQthrtIHISYSC2sEjEL/2zn2UXKWKQgYMyGyA6bfoIfvm//RomVnSzAav9HmmSBIUmqYKyU8WPMtbdM7F/1xX4qqpicVqyVg6qKhAx/rCmASBxcvS47LLLLvPWWrz3w3KR4cIhCJrrUhIhiNHs9ytkCWgt0cYSWYfyHis9pfJD5/mEgIZjZA8865U3tJUoL8nykiRK+ft/68eI8cReUBXqkDd6hJsDVkzetGaDqdjkHXQDx7Au08xxlNXXpbOjSNk6MDa8uurAgbcYV+JwlDjMUY4Sh63+c2fhq8eCC4fyFoHFYOljWcGyjOWrK/fzhXvu4HDWQ6cJg8GAdruNxWOMQXrQDWlQcZSjfnRehCy6O44JUHpHlMTkgwGRF7RURJqmHCr7DAgZjzP6/vvRM8BbvA//brCUWPKqqozSQBkWYl8xtmwHR9AN66odbq07nIdZNY4so+sau14fjq28/wbLGhl9PJ958Hb++utfJhOOsiyJ07Ax1BkG5QXKi2GgUqhw2FkQeUaDH+uII/04sG7w9Mt8HJUEMiwlUGLw3uLKPIz9ai54oLBmqKg2DUA5WTxTVOCHQqJFRARcnOzjORdchrWepNVGeolycti2Wq/VNXfTLAe8tRYpTZkVxCpGSk1uLMIJzhVzPEvv4P943XuZq5yvHImTGo8EL4midL2fJo7im4lRwBJVh9iEObQumDoagrGBn+c2OJrzxOKx3o3lvZyxeFtVPPkTA1ZG5+CH+6kHjLMYZ4/7PY75i6foEID2Acg/HBsO4vkvH/w9VmNPv8iIkERCBnJGwGmFUBqkxDmPszMoyTlQaGKv0QXMy5Qf+e4fpINGO4cO8DJyJKSK0HGoutkeVzAcwZWnOL6vTJmH0wr6n45JIAHi+k0bQWT9bYSko9KhruQrr7yGhYGngwYpWLU5sp2QWIcvC7oLit/+zF/wBD0yHHkZ2kCS3LGE52ff/EN8x7NfSvepZUSSUESCAkcrSkiEqghGS8Dh8gxwWBciQXAMygJblHR0jDQOqTW5ANeKKbwnzzI0gjSJ+MCff4hvrj5OoRyFNLgiY1HHpDi0lkhgCfi57/tJLmrtpF1KEiISp2gVgr/z7u/nCrGbdmGJVDsA3J0OR5D80e2f4puDw/SVQyURSkhc1dZjiuB/DlyJkWDUKDmsGhLpTkDflagkJiodunSUWgROkYl1v9mh1OQEcc7xohe9aN0wGI7uvXv3jsnfNitBahRGCAHegYDVMkiGeuGRPjQRSD9OTNocdZKTc0Jq9GtSkUASHPdWq01sIy6eP5cWBNBAClAaPUHks+Us4WK9fJstS2SUYHBoqShMSaxCH1W4Dihw5KZECIVAY4XC4CmrTtqwuY2WiOF25wLZzNlqksDtVFtdcVAAOaEC5Ff/8H/xRLZKvHOBXHqUU3jncMZC5McyiieyZDfPYXIONFU3fOXwLMzN4wvDkUOH6bTaLOzZyWMso2nRPYOfQfP+j8de4R8GNmeRiHNUu1oFRyWHU9H305+/xFiDRiAjDRK6GB5yXTpyni4lSgSuDde4Xs/WB19hdUh5xO7nL778WY4oy8Ablnbu5MnlQyitgwwaDYUjMeudfyYCIWKawyiCrH0UJVUvYNVr7z19WxDrFjGBHE3qqIrS3DCzISf7a/3m76muIl0TIkj2xcDrX/5KPv2Be8h9QQe9TllEMk5KPbOtM1s64jgN/osMagKJk6w88gT/4u/9NC0ksYdCjIIWNTlwxVEQstPt0h3PeBLTgLwN7k/VmuCFQFTtCqLaS3AeKcSwKmb4N94NpU2bVQ617y4bX9eJTVkJm+M9Sqph+693bjhP1r2ytfPHC4fD0hMOL1J+58YP04shWpzDdZcp+oOQQa/GjWwQ08+Az2BFUdCZm8MWGbIwLEYdXnDFlUQVuEylczHj+po+jdVkwNZYj+p8ubUWqRQawZuvfzV/fdPnaXUi+giiJKbsZzjr2Lm4yJFuj2+tHuA//u6v8wvv+8d0IjVcU7QTLOD5x9/zozz6W7/KrQcfpEgVOolYXl5GG41BYkQAQZIkpu8yhIwwCHLARwqvJdK6ADb48ZLzukqkVNCP4Od/8z/x7//xP+fcZAFBSQxkJkMoiRAxHkmblB953w/xH97/n0g8tITmu9/x7Xzn1W8gwqJVBNaDFgzwHCDjv3/4j8haiqKqTmFiLZmmZiOmkHW6alJHDqTzGOFYznq4+fXgh5+y1iqtueSSS/jqV786HQCJ4zgsukIghBiTP3LOjXrCfODUWF1bG2No9c2TP41gwrAFxkMnSRsOnsc7v+3kXLyYEth56OgE7z1WSAwepxUljpVylU5rkcMMkLTwOuGx4kluvvUWvnbn1zlw5DBOCkprgkyQqzZFGzIGzgUopFAOO1vdxgC0ugfZSHhq+RBxu4XF0htkRFEUygK9gNKecim1VqtFb61LJCTttEVnYZ40ijm0vMy/+C//jnK1d8bKkdZcPXUVwbDfz4N0ob8xlZpXPfsa/tm73seOqEUp5NAJFtvg/B2ArprqdMyaL/idj/85v/vlTzIQjqTOmOvRPKuvW/hQKrhV16Ec+F5G2mrx6Ooh0t1LWOfp25KsLNjRbuNLs27zEQSnaBY7PnMcuXXfVI6cwyHrBldnUUoHtQDvWM5X2JvMh/lYR0xFCa2k+nMxdASngW+bMn6kwFo3TGY4Sl56zYtZ/PDv0S97+JhR++0UH2FmWzjuQoRN2moxWO3ickur02ZBRlxy0aW84uqXjh7zxOupRzKextyZxqbf8HunkvAdYwxqoRp/LNb3a1dmyjKAGZFGSolUssFDMPLHh6dW77nOVI3ggRPDGoNUAQApynLUIy+Y/rq1XltVPaN4tDzARz72UVaLPqIIRJ1xGgeJzwk/u/n12bwGuMrXLa1BC4np9XjeNS/knGhXBYeJMT9n24UJTZ8k5IJPv08ijv2hzVbi5593Feft2cdKdz/aegpnUEKC95TW4hW0ds7z5btu5+tH7ucFSxfREdDppKEVRmound/Hz/zQ3+Wfvv+XeTK1HCl7dDotUtXmCDmKBCUi9vsuLTmHIcSWB+hTtjR6vk3eH2BLQ6w0ygeew1KG8ZCrsGcWEm5dfpSP3387b7jsetqEKkuhw5qQk6NpYYBnX3w1r3zlK7nz1tu47tnP58de+Z1EmOA7KChcCSLGIPj13/vvrBQDerENQiQn48M2ihi89xjvOHjwIG6PG0nwHAMAv/rqq/nQhz40HQBZWFigKAqSJFkX7AU0mmG/jUCwsrKC0GpYZyAaHyf95i44GxK/NMdmaXj2FZeFyes9SBnQK8B7C0Jt7QT2DLPaNQCimv/mwJQ5ZSSxOmaVnD6GtLXIf/v0B3lk9RD3P/4IDz7yMN1eD91KSFopTgoKaxBShgxYnT3wNfrvw+In/VkLgDQzAU1y0uHt95DsXkJqxUqvi/CeRCvKskRLGaSavD+l2QRrLVIrpJBYHyqwclOStlvc/8SjJJ02/gx9frZCmp0Yn8/Khd7T2ILod1m2GTrpNEqCLcKKQIK35XtwABeDUooAEdOTlicGqxQtTYLASsiUo5ANsFPWGvHHdoJPlWlgx2KK8Z5o1wKHij4mUfQGPeYWFzYkkdxG8cfMTtJ3O1YFpnU2yCQKgaxV0wT08owv3HwT73rdW/A4WvU+2qgWFaehulAKifF2LKjeJea57orncuM372ANi1HrfdSTrTyd2SatQVJhjEFoRaQUkQXb7fG97/tB0pEbWhEybsNFyB/jnCaANzcNv/BHec8hOZkfoRZ1y7OUw5/pCcUjh8d5jxTB162rgHF+pJogHFKp0XlIiYjjoeuplaI8jpDSNebT6XwNsu05JYo/++TH6fqSdHGetbLEKU8+GEzl+ZhU5TubLe60WF1dYylpE6N48yteQxQcTxC6AsHXj81tVRXSqCxch0dMASiaVQCnehgYY9BKB95MGWK717z8FXzjz/4IpRKkdKRpisZypLdGLhzSaloafvm//Rrv/6f/AYcIDUjWoApLq5Vw7Y7L+NWf/lf8m//5n/nWckFeFLiO57v+yY+wI2kjtKIoS7rdLkm7hTeW9q4lHuovkwlHK07QaYK1dqyNufbFjQyLrhOSX/29/8Gv298iRdE/soqwjh0Li/R6PeY6C6z1ByRJQktFfMer3sCPvP37SDG0qvX7SNYjTjv0gd/97Ef40l23UyrQSUzpTk5JqFaBGZ67dTz51FP4qxjnAGGDSjsRAJBpvjEAV1xxBXEcD52aZjWIUmoIcgghcMDy6gpCyqAa4PwQNJF+lOndTBBk2nuJBkCiSsdrXnp9IP5zYZEfPmznhg7blgeD1YI+VtFZRXxRkmKlo4vj5sfv5Ut33c6HPvrnmEii51oUzuJiEO0FCgEDbymdxYnRpilknY0TQ8fRA0bYs3YzcGKcQFc0QLr6npTOQGlRWpMkCbYoybKMTqs1NvlOlQ2KnE7awpaGwpQIAYNswM7FJRZbCd0iO2OfXxP4mIoNCtBxRL/fx+Por/VI5ztoVEB+LeO8PFsQQGovAohqS6y3uEiH1hIlkUmEzc26ktua0BYqtaotAkC8gINFH2st8UKHgXXMLS5iBgNEpDlyZIX5pLUOoJGVP+5mTuQzwtaBAQ3pPC2jKngSCKXAWYwX9Mqcj/71x3nL617PPDHWuzAnVUAbnLWIWE7fmDd53EitqkkkSEWEBb7t+tdyy523s5aOSHuH67qfgSDbZuxJSbfXo9XpsJC08IfW2KlSrn/ei0kY748Up3AMbRr4MeWf3QZzLlyHO3oW2djRBqFU8OdEw48tC4iiwAEiwJdlmItxRCRlBX00Pk+OPEyLIgdySgoKIiI0Ghs0EgnyAeKY4IdnvNjldL0aoCThIbefT93yJXIFvTInExadpKFK0bmhXHvTr5tZlQDUEqRAGscV513Eyy69GoUJySU3ysNu29s2hRl4XVH2cVQBnMzHi6PM32a8FbRHDa9+ycv533/+AYQLbSm2NMRSEbVSLJblQReZznHHE/fzr3/zV/jZH/spYmI6cYyyUBxZYWlpnuvS8/m9v/cL/Mv3/yKfuedrlDqlWEr52tphkoUFirJkx64lCicwh9foDdYQsSZNNMZaMgw2FmN+nGx0TigPwnmsEKwUA2zSon3eblYOL/OE6bHngj3sX15Basd5ccp3ve6tfP8bvoMEB4VBqADCttMWT9LjjsOP8L9v+HMOqQwXKfJBhk7USe3BNZg7lMQVnsefevK4n7Bzjuc85zkbAyDnn39+kJKtkFSlVGjJqH5WAyK1k7S8cgSkwFXcBXLC8RdTFtCT3kR9w29rzFbhITKe51/wHHSzfNB7pCBwaWyjeVzfr0kE00tYJuef/P9+gXtXnuLJwSpyIaG9OM+h5eWwxSmJrrJ0FolHgBAhE9doRapRM1dXOrizVwxOTlEHaGZlHJXGvBDDVrCsLNBxRBzHDAaDMUDtVDmITkBv0CeJYpYWFlhbW2O5u0phSlqdzhl9/4/WvuOAKI5J2i1AMjc/PyI38h47yFCd1tYCl8agtAalUCLomNeSXllZ0KpadqDaUDbQNN+ScxdgIwmJIsOh51qs9LoYY3AuoPwuK2ae4jPYxIQPOc2Zs9jgEggRynR1TNxp8ciTj/PQ/sd5zt5LEPjAb2PCBiyEOOUAg6/W5yCDabDGIrRG4XjhxVfQNkFhxIr1aluqwWkzA0G20O8xliSKcXi6q2vsLuF73/Zt7Ao0gZWKICPioe3Q+3isCTUxhSTHEYCJDb6O1TqHvf4dh8emUXVrwrxzSRTmBZa+62GtpbSGLMvo9/v0ej36/T5ZljGwJY+vHeLJ5YOsra0N29vLqp0mSRKMMcf2W7foeXgB3SLjYG+Ve554mPbenaA1SqqhYICqVIam+hdi66ovt4M5Af1ej06aotYM11/9IhaI6QyRLT/k3ZFTHrwQ22S+Taug2iZrhFSKYjAgbrWIUaTARck5PO+Sy7njqYfQXuC8J7clQkmiVkrUadHrDdhx/l5uuf9ufvtjH+Dvv/195C5nfuCYm5uHwgZ+kLLkP//9n+VPbvsUv/S//ityZ4eFnUt0lUOmLQ70e7RlRJRIcixea6zwWOEDR1clBDIUcGh87QCVaIyHMtUccQV5CWIuBWM5sHqESGkuOmcf/+x9f5eX7Hs2Os+IhQYVQ0Ve2sNzkJyf/qV/i9y3xMqRVRaTBYQtN2ccVzFurSL25P6nKghXDAsKmr5I08+RUnLRRRdtDICcd955w8qPOtstqsC6/ro+CS8Vh1aOUPqg8SCUHAbap3L8TwIrdVClPOydX2KJVgBAnAWlcM5V3PHbZJ40iIbGzkdBP+9x2BT85C//HI8zYL/PyOYiZCQZDLqk86G8yfhACFUag6meTZ0Za5JV+YnvI6XO2g1A+HH1FhgnyLMC0rhFZsrgSBQFUkqSOCG35rT0RdYAo4ojlNb0izwAA0nMwtIi+SA7Y8kYmvd/Uia7nhCrgx4DW1LgKYDMFQgP86oVwA+xpdMWIk3tIvYx9HD0bYnVAqfAueCA6aqrtlahqB0Qu4VyslaAiyReCvp5RtRKSaKYVpIyWOuSG0+MXOdAOmbVH880EGQat8KoPXD0r1ESkwEyjhHthM995Utc9s5LGoTigR1RVcGUPJazerLnXvkf1rvQqgNE3rNTpLzxuuv57W/cOHvA29Qk4K2j025xJO/jBwXnz53Le1/zTlJANXfYSQUVsQGQsH0xkfXIyAbAiWvgPatFjzTuDOejlyIkt4AMz+Osccej93LLLbdw/4MP0M0GgcTQWfKiGAIYllHry/AQkMy36A56FEURyFCrtl4pJXEcODSOBkBsJYgvPLRUTJ7n7Lr0AjJvyYsBuQtrgdY6kDLPbEMzzqJ0wryMecXzXojGorwKlUdanRkTS2w8v/xGc3ETCbn9tDneODetddgTBZVaKrzlFa/hm3/6uySRxsaKrJ8xGGQo6dBJzFo+wBnLjvmUj99yI2VR8E/e9eOIOdvQs4WObmOAN151HS/4D7/Gv/yVX2S5KHhisEpX5iy0UoTW6Pk0qDvJoCzpECRI0nIcBKy5skzFA1JKQb/fZy5tEUtFr5fRVhGJirGDAe9+85t49+vewkV6iQhIdTosCxvkGauJ4kG3wr/6L/8Rt9Diocce4dwLL6A36BOnyUm3wIR4dpTNcMaz//AhCiy2ah0SGwHR1T8sLi5uDIDULTDWWoqioNVqjTkedYmc8Q6LpDfoB7kuQSUZ4MeQJbnJ7S9HC0Clh2edcwEJI+brUD0YFnnrg7LKdpnLaiI48UCWav7Nr/wHvtU9yKHIYhdbeC0oShOqN4zDWzfs0Q7KxCJIDTcIapsBPhOAyFlrPlSX+glpuuZhq2y4kgqUIooibGkY9PosLi5SFKc2Q66UwtqQzRBaURYBKc7KkmLNkCh9Rt9/4cbvtxN1xiEsa3GaIGJNiQEiIpkMQzNTlkOW9627BF8VGQc5OYnDS4GTAis8Bo8QQU4UP96eJ3yo7vJbuOb0BxntxXkSEvr9PlprtFQkUpMojbBuXdWKF0f35Wd2Zs3BjSRDQ9DlxgAQD3TtgK4dQKz5wtdu4bvf+V1EhEy0GqZc5MY8IJsYNA0BEBsIWq0xJFoTA+9713fxB3fciHbMCGu2awwjBKYoEdZx7tIuLttxHotEKFeSyGiqhHIToNtS720S0BMnAHpM+buaB840Dht3WAVW6XP3w9/knm99k0efeJwnnnySx44c5KDPKVUARpBBLSYQ6nuc8jjhhkqNrnoVQ7DS4W0PmQpUp10l/UN1d/ALc2Ry9BTPVgLhyoVKl0I51gZdnAyqEJGKkEoF/8CLERmsWK+2d1YDkB7mW21k7tjdXuQ5+y6hXaXtXVEgo9bRx/wWV1o0VNlHCrRTzmeSd2cz21+OCTAZg44iXD9DtlOcK4hUzCtecC3//Q9/G5+XlMqhIk0r1jgtWVlZYc/ePWRrPbq2ZNDv8smvfIFOlPBT7/ghEFX1R2FCxZNznJ8ssRv4vZ/9f/jgjR/jr2/6PPcfepJu7ugXBSKJ8N4iIkXkwXmBsn7Y6jJcT31oy1YV9myMZbEzhygtLstZIKJtFS+84iq+79u+k+ctXMw8uoqxHZRVy14a43RKF8s/+qWfY1kaDpsB51xwPt1uN3AUBQ3mkxq/k61tRni6/R4WF8C9Kr7fKPfiKrB0QwDkwgsvJMsy0jSl1WoNy02iinTJlCU6ioLUDfDQY48iIoVxeVWO2gi4xQZIzMlO5A0uTni4aN+5JAi8Mwit8cYgIo1AVDK92wD58B4twCEq9N9REqRYf/bX/2/uGxzmYGTJ52NcAlk2QBpHJ07DZqWr0sV6qlc32Hk3do+ai8EwEDuLHUMrIJfjAWlzQrlqmRNKYqsRZosS4aHTamGK4rT0kdftZsaY4OQAKo5C5YQ/cyXKvIBBdfJWjoDteu4qD9qGChgdargQOKKqKkErvaWbsKCq7KhKqgyh5LYGhJ0U1PCYqRZC69ZXHG2VOSBNEmxWIAiKUxBAKYVEWId048+rrv6YgR/PLBCkDmTUOpn6sLsKBN45UJJYpTj6lMLz8IEn+frj3+S6855DjEKpEaLnbQMAOQXVH2H/Cu8fJ0lVeRJ61wVwbrSDPe0FMjFgYLKQIXIOpRSR1AHwi6PZ899CM2WJVoqOTjCrPb7/h99DQiDBHhJ9bvM9bENcrxrvtihCgqomKp24LmsNUodKwj4Gg2CZPvc/9Ri333cPt95zJ9944D4yLDKOGNiS0llUEmOrEwhrsscLP6ykdABKTqnSqCq58URSIp0fa4VWTSa6Yzg3W5l+cUAhHDYJUtf1Gua9D37+TKps6DtKKYfAlmoQVdt+zmI8x3e/7dtZIAnxWlEia/WfidhBTQ700+R/NQPV+muDZ9UMsBI+/+nP8vY3vAWNRyEwZTmMUbd0failp6MgE59UrCBzxFxyzvmsdZ9iLVujtTBHXuRID7FS9Ne6KKUoyoJOOyW3ng/e8DGiKOJvvfk9dChZilOEt8SVVLiyhraS/Mir3sn3vuptfPAzf8UNt93EA4ef4qmVNVpzKRSC3qAPQLIwx6oP9097MfS5hRPESqFRZP0+7UihSs8O3eFlz72Gb3/dm7hmz7NpAWm1BlhTVXKkUbWOwZ1rj/IT/9+fYzly9LTHJ5puv1cl+8XQKThZDhAqkQ+DQ8cRwsFy7wh7O/vGQPJphUL187nyyiv93XffLdatawsLCyRJMobYN9tehoMShwF6RRbkWlVAXicX39MhO+Ua7SQLUUqKGAYmXortlwwSAunB2UCY6JWgQHDH8oN87amHeML1WY09VoaAREWa2Dm8teMLP9PJoup70kS96+fizuINwspKhYRRK0Z9f+QUkscmSFIHhu4U+WfyBEgmz+Qiz7rVaBhgiwanjx8RhmoIElz14lTXCG+lf1xlJEW1yupq3RkbR4yzztfggaiuc6vn32QWrAkCblT5UX8/kxE984GPjZC9ZoWwb/yuo1EdJ0CmMX/1uRt44fdeRYlHe4/yIR8nT3HwuiGm4kdOzDte8wbe/+d/RNSJGBQlrU6bsiyHynbWz0rkt9JarRY2K0h1xLPPvYiLd5+DwofK3Ca4L7bn9Nkwu1yNwbr/P/wogPleSayzlD6osBit6VMEOde1/Xzkk3/FDV/+Al1fQjvhSDmg50pEEoG25BocEi092jaqeSfbSDn2mu2qBb9O+pzIa+0/btUW7EVQq7Bi/bowbW87G00pNaQsqGkLhvQF1rErncMd7vGSq64OfpXzQ7JdV5bIOBmTLB8+e3965mUdY9Z7Sd3SJWXgORS6xWdv/Sx/8tEP8/o3vIkkSHKgIl0Fx+KoW93JAjhHa4N3VOpMPoCgSqYoIdDAIik/8n0/yE/+518kbUWjqqvSkugoJNWcR2iFTlKW9y9z0bnn8L8++kG+dvc3+NWf+HccpGBJxEH+N8vRUUgCmDxjPo74wde+g/e89p3cu/ooH//iZ/nCbbdw4MABFlSM15L+8oCdCymF8Ejn0R5iBMJ4/MBAWfCcc86jrWPe8ro38I4XvJEWEOOJAcqcRFWJByRWK/rAEQo+cceXeP8f/W/WYk+uK7VRQsWzcpV4gB9V8Dxt/7V6eDXPpcVjBPSLHN+Z4utMrhXV+Gi32+uB3XPPPdfXLS/1YGxK4dZoYv3BOSXdXg+bhoHnxAh9kX40WDabjbnZk+4bwLX0sKuzQMQoM9usopRbvauKwACsqh4xrTRaOHp4DtHn5//fX+XRfJV+R+PTCINDWU8qFJGQ2NIgopGMWR08Tr4eDYCSZ/HmYOux0yjjEH48UJ3mfwl/YgDFiYAXzblxvBv4mfwMnR+tEY4JKVwHsa0OAtAQNan1twNFuR2dSxwFWdvhOTuIbC0TXnH4TQALchvc/2ZbnDgGKa0/BUpeM9vemEhTMrC511gJmbDc/I07MIjgzCiJzQuUjIIE+zTU4lT5zhOIiALecf3r+P2PfphSSHJXDEmlTVkwH89hzQwA2dIATWtM2UeX8JoXXMcuMRdkODdq3hfbc75s1PISt1pYZ4dcT4WzCKkRUlGgWKHgSXeIj//NJ/nLT36C1UGPdHGeQWTplTnGFOh2MuQds94ipEAKgTUlEXKshF1MnpPfGACpAYSnq8ICIzWzLdt+p4AfzQTKNB/rbNq/aqXCZqt7HcNpLygOHOHqPReyLw4cDtTKmGJUtVATRdctJ6oJgpwmAKT+uhbj8N5jhGANy0du/BR3P/EwT2ZHOD/dE6qE/Ujp8rQvCGLcZ1KCKsYLxKPCW7RQXL3vSnZ3Fshdl9yFyhW8pxUnZHkOAtIo5sjqKnv37OK+/Y+xsKPD1w8/wnf/3I/xH376X/K8+YsBhVWelnIYa4nSICCuAW9Lrl64gGe/9Qf4u2/9HsCR+5IHH3+Uux6+n1u/dTe9bIDPSxbSNhefez7PufRZPOuCi9g5v4RGkBKhK85Mj0EhkQh0lODyLFRhKonB8Yhb4bc+9iE+cduXOKgGoCOMGPUvukq5W9pGAlCc3PhWQuCEx7pAveGAtX4Pu8OhjuFh12NrYWFhPQCya9euMbCjJmScnEhh4RFkNmdQ5Pg0CmRNbrz59lQisusyrNXP9y7uqDLHtXyYaJRybf1uqqIAfvgsRyQ1c67g4eWneODQk5Q7W9i2RiiHKAqEDXKDdSWOYJRFVhu8bhRoOWYM+MqHao660qC5gW4EcDSrbk7JpnWCAOGZ/AxltWk0wdG6imIoxTVx77eVHKJgQrt6vHJlEpCc5C1y2+H+T4Afk+N/GtmdOMscybMF/JjMaDVVsdapwwgYYFHCcuvjX+el5z2HFglxHJNnGUmjjHor5qUAzmOJl17+XL788F0kUlNU7QhCCIyzMwWYLbayLImkYt5qXn3NtcPSY5cXyCQeLatyfXAht+skmhiHXkkcghIwQmGAVXrc9uDdfOK2L/Glr9/K8soROosLlPPzrBQ5MolgroN3FqMExpSUZRnaz5O48sndkF/PM0oy1v7DZHuzcON7lKuU5DdKnB3rFQIIKrZQxh0/vh9JP9uTxu5RBXzUsVtdTeG9JxKSJZnwE3/rR+mEprNAfOrAlgUqiUOFazXvmpLHasp+cCpMaz2Vp6EsS2wcc9faw9zy0L3svfQC/vqLn+N9r38PEh/8xwblgpvY3wSn1n+sP884g5QaEUeVpLUglop+OUBEiutf8GI+9OUbiOKEUgTJZlW1m0ohMMaglOJIb42d5+ylW2T0Bjk6jfnpX/q3vOqal/Bj7/0BLox2VgFNSOtn3S5zSYuWUNU98ESAdhIl2+w7/zm86Pzn8D3Xv33snmhqoS2HzXPmkrTa+z3eWkRVWZMVGTJJIYmHrXs3P/gN/sfHPsgX7v8GdrFF2Y5C5VEj1grdGaFaw8kgW3vSsVJdWSf8sDJ1rdcdkkVPA6hcYw8RwNLS0nQAxFob2NWryTOJvuDBC49EsrK2FshNVFh9nQ0ASHORPpWBm5+oAhEe9izuqAKm+ofbZ9sME9IhRXVawtMvc0yU8KWv30bnnN2suDWyMsMXnshDjET60OcpY423NgxYf+L3SgqwQqxD0c8WE4RM/UZl/06Ost5OjFqr3Ca2XmyUUT+e5+nFiDvjTL3/uPXEvM1gzFZZqoLRqwJiWTl0YosvIAD7CBE4e0pCW5WV4ZB+RO46bf2zW0yCWt//5jicBD5cE3zyZ7d04NlgtUNUO46iAYA0sT4vwCWavrf84Uc+yNV/759QoolEIIteNzcbTsBmcSeJjX5YLSoJljdfez2fv+Nm0l0dlssclcSImlxazNhRt9KKomBRaa4+/1lcNn9uVUekgoJdw2ENRJ6ccZudwQdOD9dHyJQSyZ3L9/GHH/oAX3/4mzySrZDFnrl9C6wIQb+fE88naK1ZXl4mTVMoDBKIpQr8V4XDWhuIf6No4062o/gUstpgNQIhnz4Ash3Apvp8pokiuA38LHmKYpDtCoAME6ZVO4z3Hmk857Z3cN05VxFhkISKCVEphSgaG4AYVYBMCjaccj+lqmJpVoKEuQVfufdOjkjDUkvz4U99nO94/dvZSQs5Udl02sbjRPWHlBrrLBpV67ZDrOhEKSUlb3zZK/nLz9+AsQEyGXs+MnDdxXGMzQtWel1yDWknZeA9+8s1Pv7A7Xz+P97Dj73ju3jX1W8gAjKfsXduYdRfUvmnEYpEgjMWqSASoXrMO0fpQ5UaSqKqGo800di8QBQWmbaCpkqvQCQxLd3GOcP+co0Hu4f4q5tu5KNf/AyH8h5yqUNmS7LMQBrAWukC4WpqQ2dIoTbRCaDBHygFKMFat3vcwYG1lt27d68HQPbu3TvG+dHk/mh+X9vy8jJCK6r2HlwltSX8KQXbRs7YlMG+c35xjPlzu+2fuTdEQhJVG34cJeynx1995lM80T2I2ZninUdZT0fHaCTGGHJhQUmiKX2Z0+7N0aoKzlYXsK4yqIEPOSEHJaaUZzVBkNNBgHos9FNy5kqSyqOVqDbndWPztY19RkrGZK62yP/CitFG644xv7wY5/3YrFaqk1k3bT2WxAZr6iQXDhN957M47hkDfKjjCKSa46PAUXrPfQ8/SEIbj8B4R6TVsAd7XdbtVO89jX2+jeTay69iLk5Z8aClwlXZ0LpveGZbN96Qgtgp3viSV9BBElfZT9EEQBj58lqMgLjt9Oymju+K2L5LgZJtnmKNX/v93+LGr34FkyjWbEa0c4E4Vaz1e+QVL02vzHGDPvM7loKv7T3aC7SU+Czw1ygESZJQerdufZ7m523kI8rqQTSrL4/3dbus+8pvkEQR0+9Hc/+Sz/D9qxm71cF1zTchreR5lz4L5XMSEZMXGUprtNREdfVeAzCaUKI+7ddQB6vee+I4ZgXPJ798I35Hh6e6K8wPPN968hEWz7kChwBnN1b53KQF5HjyoHlZoqOgIom1YB1aSVI0z9t9Oc869wK+tvwYQgUnzBiDqIoOIh2RDwZhDYhjCuVZyQYIr1k6/xxWswIjSv7v3/qv/Hb0v/knP/kPefkFL2Q/OW0E0noEEqtEaAfxEilDixPOQ+kRUhJXzvTIx/aV2EIcyPc84Cx0QsVHj5LH1w7ykS/cwMe+8lke7h7GtSLytkJrgU5SdFlgfC1iEHgu6ydpZCic0JvEJTds9ZJhnK/2usfV4VBL6O7bt289ALJnz56pRGZjgEhDPvTg4UOISv4OWfXXuONzpjZ1YxWjgdlOW6Msgve4BinOVseNoT9M43E4Z5FKYoGBLTnS7xJ3WlgpSNCkDhIrED70OpU6SJwhxajn8RjApN9g89BnaQQzJOAU45v7ZPBnT2Xf+sRmfaLqIMKd2fd/MsCe5Meo+UAUI7CjTvBKtr4ApJ5DogISVPW9cqMWmBrosFMq1LZy/nkBpRyRsvop7UZNcKYpO1b3fhvFzM7oRfDok2i4Jm4gQ1pWYIcUEXc8eievvOBqyrJAR3ElDy1O7/lXC0QdMKcIzm/t4voXX8fH77kVUSnBaCWxzq9L4szsNAMHUqKRXPu8a1A4JGrqqKmTW03Aezs8ubFz8euHZgD/Yj5x++f4nY9+iMcGK5QLKWuuIN25RNbrQ+7weUY7jphvz5GXBd1BP5SaZ9nwPoXPcAgtiOMYnSSU2WDM793wPvuN/YshqCFO/HU7tHBO467ys2k9FhgO70tFZaC1piVj3vPWdzInIiJEaMNAYKoAsCxLIpGMjevax4HTg4SUlZpL3QYjhBiq2Nz54De499GHMDs7ZLJg1+Icn7/pS1z9HZfhp+gTnc4hMdou/UhpTABa48oChEbJoArz6pe8lK999E+Rka7usQ9cWnjy3hpzrTapTji0toKbi1ncsUR5YIXukRV87knmFkh27WC51+Nf/+/3EyN51eXP5x99z4+zUyWBP4+g+oSxIBUWkFJUuqNhMluCkqjFV/yYAotDSMFaMUDGbfrAx++6kT/5xEfZn6+yv7vCQJRESy2s8JRS0TcFeT9j5/wiOi+HFcOeUB1dys2bn5OcpN57nIduv3d8z6niMT3nnHPWAyA1MUhdktMso5rcnAyw3FsLMp3eIbxECrHhwrRZC2eT6Kh+b+VGAa2OozGqWdUYmNvH+ZHISIKAjJKnDh8MUmdSU5RlmPgOnCkDuBSpMFmkwBs7JI/0RwOEjqJqclb3TE62tohx5Q4ntl+2aaM5cEbGX3K8hHWa01vvHapamIa0G9sBxfTj2SQ58TyalVdig2e3pT3UT+PfjsfhntmZs/7VgcQkr93YuifGf7cOqHQc4aRgf2+N//Gnf8D1//B5tOIEj68kdKePfXd8+MuJgyBiohLLeWIpeP1LX8ENX78F6R0iEmipyMt8SKo3s1MXoDINzBAVybWBjpfsTXcSITHWECkd2juqQGdq5nmTVCjWvc0GzOf+WIHUhDKGETDAcYiM3/rIH/LxL97ICjlFquh7R6Gh319jXmmK3LA4N4/QitXDyzgpaKUJ/X4fVUmY1r6qisMdLY2ht7ZCHMdTr2fyXOs9dlL9T86G6DNmnk0jKZfVg262tUZKMadidqk2zz3nWcQIsiwjThMEUPoSLaLQxmhGg0g0AUixOb7XOsJeMf4PWms8UJiSJE6G/BErJuNjX7iBeL5D1xREsebwoMcnb/k8P/od7wtgaUWCuqFS00me/yQhcPOejERePVoobGmQlaqLjGOQgdi/heGVz7+W3/iLD+B1RFma0IgkJN465lodTFGy3BsQzbUw3tFfXWMpTXH9nKXFeVZXVminHcRiyiOrh9m5c4mP33kzn/kXX+KqPRdxxfkXcN0LruWFz34Oe6MlBI4+nghBkigEDoOgxGJqVRVkAEeE5NGVR7jlzju5+Z7buevJxzhQrLImPH1RopZiVtcyFuJQ+FCYkiRNkFJSliVxQzDCNbo0vPfVGH36D0F4UDKUVAshQseJ91jv6ZU5BY4EuW5tbjIYWOtAy405QGr0ecPR6wwoSQ/PwysHQUc40yeJIvLSAkGHvCZossJPbDib40UPwY8mAaEDgcaJCj0MXUzVqYstrwIJi5YfKtQE50Dw+P6nEGkUgkOtA8CkHAKJ9BIvQbjwVIUXxyR+apYtTg6gaQSHZ5X/79cHrW7i344Gup0q8OJ431uewc+vqd7UVEdpstOriuekXjOiOgO13UApMe7cT5MknPbMt3L+OcbbvsSUr+XEebtqzNWVH7P2lzMc/PDjzq2deP6iyb9QWbPoR3k4knVZ2D3P7Qce5IniIJfHu0fZarGePdcjh58jN+saaj1yyfh7S4cEXnTxlcwPBINUkpUFyVIHYxzMmmA2xe+ato+FKr2QBPMmPBURa4w1GOHpyIT5bsGPftd3IREMnGFOxThjUSqq+OUaY5JGm9YxEYnjC77chI80DVg7WhBlsJV6Q4UqGINNI1aB+9x+fvE3f52HV/dzMClwsURIgXEOhydJYgpjEUlE5g2UJgCKhD79REcNtKIKZK0bnkMaxSc0fKc9I9fwI9zTeN0OPsRGPpDb4Lrrv3km7F+1zkldkVNWbX11JaooLVEUkVmPkZ7COHa15khWct72xtcQVat5mgYZ0AhQIhmty2o0xfQkAHKS1Z9jnCJ+AhGpuUekp8BBrDAYIg8WSaZjvvzg3az1u8SLKVZa8lSwX+bc8PAtfPdF1yKcxysxXD+GXIlegtwc8H0qmCJGFcr1/0UldjEWZHtY9Jpnz53LVZddyc1P3gelJ40jekUZyrttqMCgFZN7i/SBy8MZi9KaNTOATkQPg3eO9nybrMwwsaTUcFv3cW678yE+cMcXaGvN7sVdnHvOXs4593yWFudpyYRICZAapKf0sNZb5amnDvDUwae499778MJhhcRIh5GSUjmMCGuZLUo6SYotSiBwVPq8HA4No8D69TF6HfOedPLPBY7Sbm+NpaUl1no9UBGPry2TAfO4scxOXXWuqgGglMI5x3nnnbceAJmfnz+OESCGFSB9b7D4UP1RD+opmuRObB6RXrMX0YkK+Kg+O4AuAifFUG1iMljZ6gruKLgIFRQSTm5t0Mc4h/HghJhKONXMHB9rEB3t92YBzNHv33avrjiTn5+fopzS7Nu1Eqxr7BvVnK7J8LZFBciUZ+HOIAWf4107ZmvGMxgEqebW1ILIiczJ2K+I0LIQJQlH8j5REnHznbdz0YteQ4JG1W+47g9dw3XfrMWEDTm+vLfsEC2+/fVv5ndv/Dg75uZY63bxWjLr4Dq1Zkyo6NCq4l6h4l7xDl8azm3v4OXPfQERiqjq1xdCBDlOKTfu496kCpATAUmm+98GraLQ2y8lPo1YxbCG5F//l1/hgf5hDpcDSmnRMiZSfiQK4B1HS4KeTt/DPc3XM9UHeqbsYRsR0tYxVidOWe2uITptdu3ayWOPPYbpDXCrGW+7/jXr1r91uMYGwPdmzjt/lB+OAMqqnlCGWPPmp+7isM2YW5yja3LQEiPhwKDHJ774ad510bWkjY1nCEh4wIf9Z1PjP3HUbzf+Ye5YbLV41xvewu2/c38QXnDQipNwG4wFL3EicP1oO54sdEOWCVf5zRVBrHQYCZkvQYe/6WM51N/PffcfQN7/jTGA2lqLcbYqY5ZYaylMSdRKKhJ/VyXr3NjnHWuNctPuzSaua8NWeKUwziF8YC9ZGfSwiHUDbDLlIRAIKdi7d++692XHjh0ntKAURTEkPmv2nW3VwlBPHMFEDbrfHlFtWGwqcGaIqCv6RTZsNVJuJNVafz1TYZjZzGY2s5ltffDhMcZUwKTghq98kbzpbG7g+Ek2OfmwQUl2fQoxire/9g0kBjo6prAGr+SMK+Bkb/sxqk9rNQMnBRaPww+lOIWHWEecv3AuPtAWhveU09mdNvtRreOSEsf+wOHvVtfdkhGiKCFSFMqzChzG8u3/8Pt5oneElUEvVHtEMZGQoVy7BkDcDEqe2ebFOjWRf1U4UcXXBhVHZGXB6uoqEZLICa654krOTfZseQvUhi011fcSKIsc74ISUm4K+lh+/0//mNWK5yHPc6SUKKWQSnH73d8gB6wKVxdNTuXttObHGofj2kuuJOkbyrLEVDK+rjTYZgW0b7R+c/REW11ZbGsZ5Loi2fshuJEVBf0iD3QLwmNF4PQywiOTiHSuM/x8JwOHXbPCebvsnd57oijCGIMXAocPgiwcP8n5ueeeux4AWVxcPI5dZHQXsiwbMqx77/FCnPYb5dchcZMU2X78IrcBCCKq82q2KimlUIgR+DHRqlEvcnIGhsxsZjOb2cxOt+NBRYRaliy0O5TW8LUH7uWB3lMYHDi/oWqAYNPayNc5tQ0VXDweJSQxcH57Nxfu2EOx2gvyolrNeGxOZWAGQc5WCqx3GGexPrQjx0oTKcXenbtQgPNlqIO1duRXej+d7HoTB47Y6L2P1vrSLNM3IFUUyOuRPMgR/vlv/DLFYsrBvIeMI5IkIVIanMeWJhDvsfVJwpk9M9bgseHrx9X1CmdZ3LUTrTX9bo+9SzvRxvE93/6deMotjYHqqVa3Xk6rhhKEKhZZ9VHk0rPs+9z2zbuQacxgMEAikFJinCVKEzJn+OIDt9FvVClspBa45TNQwmDQZzctXn7584iERCg5lLm2ckSer54GXhrHQfI9gCEu8GQoiU5i4k4LIzxOCYgUIolAK0rv6JU5vSLDyXGxghpIcUdp8T7t+4xzw1aWmq5jtbu23jfY4FytteuKPSTA3NzccW50obRxMBgMyVLtaVzcN9L+PtogF9tsFfPeD8+71WqhhMCVZgh8NHk+auBDzPbPmc1sZjOb2ZYFuZ52uw3Ok5clh13O39z8RRySslKwcFP24lrVaVPBj6kVIOGHEZACb3vpqyhXuug4wsw20M11Y6ZkBqWUODzW1yXUlQqfdZAbXnHdSxE42iIJ48E12NydGwMo5Klw4qYwrNbfumk+pp/yS4OMNV+wBvzC/3w/X93/EL2OJk8VhQiyo2VZYotyqDpQJ7hmNrOTmW+u0fZXq8o124kHtqRX5gglacUJqrCIbs7VF1/BHNHWJoEbc6/mA5lER4o8D9xvKsJYg1MR33jsQXwrxscqXFerhfeewpSU1pBLz29/9IMcZDAKCCc4g7bFyi/ACU+71WY3CT/8Hd9NJ0pC1RxB6cmKUH0hGpU9kuOvwCitGRYjCCVxSlA6Sy/PWO33kGmMkWGcDFxJIVyg09IKWfER2QngYyukkDccQhXOIKUcVhtKKcnznJzyuM9zkuf0+CpAhi024WMGeYZzblwm91SPoaNcYWCdnfiFbS1755B4ztu9l9hLRGGG12fF6HBiNCFmNrOZzWxmM9sKK8uSTqdDb60LkUIvdvjs126mAKJOa8T6zhTG/030onzDqZWTPra1KA8LRLzztW9kT3sBW5rgHM4e4Un5XnWFxkbOuHMukH5KgYg0Sim8tfjSoBFcf+1LwJtRpUczcVapwMhTCX5M+dFGvCPT1V8cpDFOxPzPz/wpdx14lEezFY6YjFKEpFbtpAutUJEeOuvW2tkgmtmmgCBMgCDDf9eS1UFoFZEeegcO86oXXsc8Ecn6+vitCHmGL3bKZIukwmUFeCisoUDygU9+jL6w9G05ZLMtigKkxOCxseLm++6ihx+953aK2hunVFqLRpICL9pzObvjDj4vQxFBg7Szfn5wYvw1zrkAPlcJdqEkMtboNCZqJSAFBo/BhzVaKaiq9pogizsejpMtBEBc1U4ohMBLgfGONbrrTtYf5T0uuugiPwaAnAgHiKfBASLFsBLktM8nMe4MuUrX2G9TECRwpoAQEmWDevVl513IUtwiQQ3VcwoFuQ6vRs50zmc2s5nNbGZbHwTneR74qrSm0IJHlg9w1/4HMA1nqWTE+H8qHMk6aJ3WWiOEAOvRHs5RC1z77OeivRg6TTM7ueff9L2caPKvhR50hwclEVrh8DhjaeuYfYs72aN3EgmJrGp1pJDjiFmj5VdMAiGbWQHSiMcmAbt1rS9NLzmSFFryFF3+/MZP8eRglXTnIot7dzHIM3QtY6skIlJ4LSlxY4nCmc3sZEGQOh5oVoY7AUm7RdxKQ9uZF+xuzfO33/M9LJISs/UiEM0KkKmV+kojCaTIUZxyf/4kX777DnJN4M9wDpzDeo+OI6RWgc9iIeVL37pjuAdtV1NKh9a/3oAEz8uvvJqWDS2DuTPj68+U530si+M4tIcQKmQGeU5WFBTGUDhL4WzgZdIKVd0/7wOvV1mWo6XONzoP2D5dCDVXprU2gB8Vz4kQguWVlWOCRb5qm/HejxGhSjheFZjRwLXWHteDOV28FUFCzU/CH9vKhhUqlTRQiuA8tYOdcZt5FQ/vlZEB/CjVqCds1gYzs5nNbGYz2ypL05QjR46wsLSIlbDcX6PnDX/5+U/Ta9RfjrUUnEbcocSipQqfmRsS4AWXX8m8iomcmFVRbiIIMi2IqZWAgkpsVfXgPItpm/P37EMDCRpblpXueRWSFUVQVpkGfGw2CDLVL5sSqE1+roRMOfaT8ZP/7l9yxGTYKAAcjz/6GDuXduCtxRhDbkPAkVuDqZz1JElmg2dmJ40fTI2xqtesLHDOkXf7RBYu2rmX5yxeTEJVbbXdVQ6LAqIIa0oKPDd+7Ra6yhHNdyhcWB+klGit8d5TOstq1sd1Ev7gox+mqOdqs01yGy36AjBlQZS2SJG882WvYUnEeOuQcTRcX+tY70Q5q8qyHHJjxHFMnCboOArtMAK8FDgBxjuMMRTW4JxDa02r1Rp+fvNQ7tjk16fTpJQ454bgh6uEWPr9/vgzn9YiW7XNAOzatetpACCMiEafPLCfJAk9THXP51YBIfVnG0wtoDQqOdwm5FOekCHxEPQ+kZDlxDh+5v/4B0S5JXEB4Rq4EhMrcukxsupZcjP0Y2Yzm9nMZrY1lmUZnfk5+tkAqRU6jikUfPr2m1nDUVTtDQObhX1uyJ3gNsUR9dN+0OyiqHOc1oKQaODtr3kDrA1ooxBmlI2vM1/AkFRtZscHfrgJ/o+asN17H3rZvaO0BiEEaRSTrXR57qWXExGy0EkUj0cFSTwCQxpO6abHLo3xshH3RzYYjPzGyncsTUlGUHz5i6/eyAE74OBgjbm5OZRxpFLj8xKJCOoUWuFkKM/2OgQfzQzrzGb2dAPoo5lSClsadnTmoZfzosuvIkWAKbdVGflGZMQyjsE7VCuhQPDJL34OE6sQDzkbsvfWBTWymmC5ldLD8NChp7j38MPkthi9eRUD5kW+La5ZEpSwwKOA6/Y9m0vndqGEoPR2jNdlCAqJo6/HY0cFQNvq/hgfOiKcYLjnaSHRQg6rOpQICl2mKMfaHKcdW21KKYqiII7jIdBTV4UM8mwqvje2vis1bIFZWloaB0CO1wEI/VsjJ+K0IT/HAYJkLh9uakKIamJtn4kvpQoDM8vAg44SWkjmnOLSXeeQWgGDgoW5ebrdLlES4wX0ehWT/cxmNrOZzWxmW2BxHCOlHJbWWmtxSjBoKb7ZfRwnwi6dqrBXlf1BvfFtWvLRHcUfGArhaQ1Ve0WbiLe8/NV0nAzSpIyY5JMkGZbUThKjzezEgBFROdNlWeJEpWwnJNJ5lHE867wL0TWI4qdEQmIEUpzSxO3R5CSdCyX4BL9xkGehslgJMjyP0+M3PvJHrFCy59xzWD50mJaK6OgYUdoxgMjK7UciOLNnxjybVOQYCirECRgLg4I5GfHG619FXAXdNs+3zXVsCG46C1pypBxw14H7eWrtCKXwyDjaMNb0hCr5XMOnv3oTNoopvMU5G/gtgCROKPJs64tBrAtyrcLjrKUFvOP61xJ7SaIiVFVxIRvXtlEwf9T769ffo8mfT1Z1NEGOdcCKH3FAbbV5MaLcqFswLZ6iKEb3R2xwv2p+JiECmXt97ZdddplX6vg7xPr0KybfcXTpdC0CzQfb7EPtdruj35sAPrYDDOIIfa+6vvm9Acp5Ll04h1/8mX/FvFPsjFqIbsa5S7s48tgTJDrinL376K2tzXaAmc1sZjOb2dbsX43qiTiOibWmcJYnekf44A0fp0TSL3MkoAFvbOX32XXk/KfCgbbOrtvwExRvf9mrUUcGxATW+Fp2EBiqdsw4Gk7ACZ3wu2rnWik1rHTwUoTnVFqiwvGcCy/ZmINAMN1jZwpQskngR5P8dDienCdpJpqkoASc1GQI/vTGT3DP4ScoY0k3G0BpmY9byNIhnEf6cU66sbkzG14zO8m4RzZkb/0UadJ8kNFSEbHxXHPJFVy1+zJkRQ2qqhaL7YB8jKmCNc7feIcBfNTi41/6HIfzHlaCMSa0NjKuhiMb4hBWwKdv+wpdwAuFrKr6SmsQMPz7rVw0fd2VICVSKZQxvP7FL0NnBrPWR7tGi6ForLXN78XGvCBjgIafvrw2wYzmsRHosZ3aX5oYQ+2H1B0o3X5vBDZvsJ80uUrHKkB27tx5QidypL86fMNpBKhbUUzqgMMrR4Y3QVTa8sPBtw1aYSSVik6942oFPsj27WWen/+pnyFeyehkHrmWcdkFl2B6A5567HEW5uaZ5ahmNrOZzWxmW2F1W4NSiiSKiJTGesehos8nb/kia5Sh5N+WKCCemwMBuTWcrAZG7QhuSFIJRFKHPb/y2jyeGMUVu89jj2qhbZAlbYIgvpEVmtnRfZemU9100IeyjQhwVXbOOZyxSOtZiFL2xjun+y/iKA/8VLhsU8AUUQVkUoaqIZvlmLIkSlIKQlD2uD3Eh2/4BHsuvRCnJb1ej3arRdEfDOeElSNyXuWqcvbZ0JnZJplqqL6sa0PzkGcZbR2jMsPrrns5CUBeBlBSya0djI15J/x0QlalNSWeDPjSXbdjI4mTgqw/II3jqYG4rO6JFfDg8n5uPXAvfcJ+YxqAeBRtPQAktQ7PzlcRsnWc29nFNRdfTsfJwLfRiJ+bS1UTBDkWUFbfJzkBiAx/fqy9VowDLtvFaqyhCYI5AvH2SvfECgTGAJB9+/YdVwtMvSetrq6GTa6xQ9XSO1tyY6qHdPjw4Q0Ro+0AgCgqMiJRDek0Biko8gGu7PGSXc/ml3/yn3GR7JAc7sPBNc6b38lC2sYUsx7Smc1sZjOb2daY1jr03FpHOciDY60VydI8XWH5ysN34kVE4gSUvtqbPV7LdXwLJ42G+PX+tQxIRvD0tMCI0Gu9hxZvf9mrITehD1oHidb6muI4nj3c441j/Eiu0U9UgGBDX7aUMlT/OE+iNJecdwEper2qi2hk7MSxH/mmBGEbYSsNORilNTqOcECJowD+5KN/xlo+IHeGPM9ZWlpCa82hI8uoNMbqkZSl9KDd6BDM2mBmdpLB80RwW7d+NOfhXNLCZwULKuGFV1yFxtJKUpACuw3GoK/nebUZDOfhKGuNQXDbwXt5rLuMi9QQ+JHWr1sLas4MVa1Ja9LyZ5/7m0qFzKF0VElxu+0xASuQ3RiD8w4ZBdWb73rD29gtU6L6nlQBvqxaYpox7rRKsml8HcNKD7++JcZNOZqVJRsdW82SNQRAGq+1ra6tjfkYU1WGapJu71lcXBzdv3PPPfe4e2A9ntXVUAEyZGOVWwcT+cbkWl45MjzH7QJ6NPdeU5YoEZxBIzzWWxDQSVrsjDqoss+rLria//TP/y3XXnA57cKTHzyCtp687qee2cxmNrOZzew0W80gr6XCm6p6QgqINTZWfOILn6UEhNSQh72t8BYlNzn7NvTaxgNnQeBxQIXC79I7BLCA5r1vegeJ0hQVd4musnF19n5GUnkSwVnlZtVcKrXfpYUkkooXPvf56AlnqAlCbFi6vNkOWPXabH+pA4j6H22vD1W5fOYKQJPj+MLNNxF3Wqz1emEIVtVDrXabzBkyHKbi/VA1AGJHAczMZnayVhNk1i1ozUy98JAoTdkd8IoXXssFi3tICRXmQmkGttja4BVGIExjAk5ODwP87of/lEyHvUYIQSdtYfJiXQVIHejXcyxPJF978JuArhRBoSgKRB3bbmU4KKq9iUAC7a3DVef1okuuQC33hy10Vowq62QDEDlWG90kX8c60GxKZYcXG4MiNcBmtwH40QQxfCUm4r3HVxQcQxWY4wCgnHPjAMj8/PxINeU4rDfogxxHYOw2gNhqDpBaDHdbVYB4iJQGDwaHlRKrFKbaTG2WsVO3WQIujhf4tX/4C/ztd76Hjle0Vcz83NxMBndmM5vZzGa2NeBHxZeR6ohIBDlCJ6CfDchNyS3f/AaPdp+s+gAUWIuTctP4P/y6LyZ/IUivWkLm3lcOj7Jwyfw+LjzvfIQQFEUxlNMzJlSFzFRgTtIxrQpbtZRgA0gmhcDkBVc/+8pQ/boBjnXa7vwUUtIxPhkPqiqVt5WCggVuuPWz9MtQ8dSOExId0Vtdw3lPa2GOTDgGWIwaBS7KQ1RVgMiZ3zazzZhfDSJKPyUg9oVBO3jL695Am9Ay4osSB2gVb2kQu26uT2kPs3gO21VuvedOMuECmCgEEjGcQ5OVZ6ICG4WHQgsOF33uOXg/smqy2U6tjV6EPQoXKg9De51nUSZce/lVKDfO66L80WXHYT1/x/Hs403+GCeOr/pju7TCNCs/6lcpJUVxHABf4++aqrdycXGR4yFBre/BoCwo635HD2pigDZv9maSP9WaxFaAkaNzctVn9Yscg19fBrNd5kADgRIVZ72hZqRPISsQDtpG0cLxQ69+D7/173+Vn3rvD7PYcyzkMF/AXAntEloGEguxDYuA8ttHsqienE9XRmmze9DqSWxlOKZNbMn2loGa2cxmNrMtWcs9RAIoLVZAUfVXSw/eOlQccWjQ5Y6HvwWmACVAKxQS6zenukJMfiHW7691JavzDo2i1p8xJuPqC68IJOPeI5QM2TVriYQkjeMNHT3JqPf+rAXAGs7zsPS88jmsDP6YV0GWEOuIRKgEyQcZ5+zeu632UNE4F9GMLpyDKALvKMuSRCas0eO3PvRH9HQY984FKeXWXAdrLWu9blDo2wBAG2bop/gik1LCG92jmf+xuUFo7QPW81pOPJfJZ7EdAkBfxTy2UWHUrC6qr6UjI64+91nEPqzPItY43JaHQGP3tDEBZSNQK3Dcv/9xyvmEI75ARRrlwGR5JR87fj/cujXKc2TQ5fc//AEGWEoscRRjbTHig9zKdUdKBnk2XAzyMkcj2Kk7/J3v/cEhWCr99P1mksz0eGIlOeX3NwJGjva6XcCPGtASjHinnBQU1oxf7+R4q8ZHbU1VVd0sB9l4BoI3BhlJ7nvqYfLYV+22FmkkqdSU3k38iRg6SaKqln26KKR0EDsoJRgVFgNVLQI1IPLAE49SIogQOEIvpweKPCeJk63fdauS3VjEQ6dCNre26qEkUYrCESPpsJPLr3o97/351/LFR27ntnvv4rY77+Dx/U+ROUMpPIVwlEFkeogqgg+tSc32JOePOcDWg2YN5l0lh5myWoPZVdJxteNT/004RoNWSF/hP34d8ZytfiaVGmbkZKRRQmJ8IFPzFct8uGNiWGrrXGBgH4JkSiG0wlpLURTheyXJigE6iXHVSlAzSNeoehNtnbZYeAFSCGZ5wpnNbGZnmykPLSvwkWIgLaSS2DoU0I40zhmyyHHjPbfx9ue+lJ0eXFkSxRFeKEqfE4mnvwfXDo1gfTQoG19JFSCPloiHpKlISEXKj7/jvXz+lpsYRG16g5y2jEnw+LzE4XBR+FtFzXURPBhZFcc6Kc5qVktXZbmUD1U12gVCuqKWdbAWYSAVEaIEhCLpzNHSKRHjSICYeIyCDUCto//4aY0jxUQ2ddg8bqlLVTQag2N5sMpDdoXeQozxAoHDCU/hCmQk0IDNClpCg2VYxm4qz3sYqPlQHVMYg4g1gyIPstIefF6SximuNNhqkPvKV6qDXWDYYjOzpx2BIrSi9AZvHbGXKOtx3iNiRVE5flH1HOtnYeuMPFuniGFFiH2E8+xUKWU/w0qP05KB9yAksoTXXP0S9pEQW/DKY4XAeU8s5JaCaALQPtAAiFiFCMWaQFztws01KN7/x7/DctuTRxGt0qKNYWGuQ6/XQyk5BsTWSXdVxZ2xVDhnuOPhezlAjwuZQ+Cw3qCkZjss3mmrNVxy5qPOcL+5Yu+z2LOwxJHBYRSK0hi8lOg4oiwL4jjC2nJdfDJUshLrAet1z8CPKkZO9HV7gJcCZy2yikNjoXClwUcxq8VgHNAGdIPzCREUWGtrVc8BQDe/2XgEO4QKaGLhLFYJPGJ4g7wNAsZ1hqAOGjfz5olGlYlrbKK2+qxHDzyFAUoM0oJUofN0OzAAT27+9YK60Q6vkWggqZYtg+Q1F17Fqy+8BvPG7+UIXb52953cdMdtPPzk4/RcyZOHD1IKT1mBIMMSp6pnCinXl80dExkdgViyDIkO4TxChClmrQ3lVyKMBaUUWgaSuSYbvHMOT9DndniklAGgkgLrHaX3FCajnSbIOMI4Sz/L8M6h44gkSTHWY7zDW4vxZijBHMkRs39WFlAGgEZrHZ69FEgZspZqAxTVielZllnb0cxmNrOzPnZoEM419xDlQrEHgNWSz972FfJ3/5/kApI4CrxXUQBBTtYB3yhIFhv83rjUomQPczx7z/kcOPwgq2UP3UpwpsQ6S9pKWfPlMGBtZuAknPXAdzPjKhgR9DkBTtXcFwLpqopgAtmf9JJEadSU4GNLwpE6ySEaIIxvRBGVv1RX637l1lvIEslAg7QuXPMkOOimBB9iPUjnjEUJgfE+yJIqiXDBMS+zHK1U8KUbQMcM79g8y3t9RBrjYoXSCu0lkfOUwmGFwFcpW7kN1XucIJBJ5wZblMRSUXpH4TxKKyIkC1rzg+/6LmIPDHLEfIIFYqG3/nqqKishBKb6UaRUmCylwUnJfcuPcSjv0ZMG6xzO6aFMufd+LElZt1X6Bl9G1h+we2GeI0dybrnvG+y9/DoSZMX35NgudVR+SvwngRc9+7k8fOuNGOFI5+boDwbkeY5QElOUHI2m83j3p8kk7/G+bkd/RIoAioRWxSm9Qj7c2NG+5ZEEzq+xCpDdu3ef0IdnWRYy70oMERc3yUh2CjbgmhxGsr630gHLvTXWGDBPC1yJUDVzrDzjdxLtYYkI7x3WOZZ0hwuvfBlvu/KllHgsEoMjo2TFdDlw8CCPPPII3/rWt3jwwQd54vBBVn0R7uExuFGalSD119rBledexFycsnNxiX379rFv3z52797NzqUdtNttIqlQKGIiVDW9PZ66mzYloqSkbwas9ro8efggDz32KPc/+jCPHz7AvQ89gHGQdXMKHJ0oRiYRznsG3RynFEpWlUVCoGSoLvEyTAInIE6TIElnLVmW0Rv0iZSmncakZnw8NeWUbAPp30hDe0blPrOZzexsDYDrNbLOcteOeQ0QCCHITclHvvgxfvD6t2OzAe20xVqvS7szt7UOU+VwvuK6l3Ljh++hlUR4ILMlWgTOh0m1DjXRd+5n0ei6bXCY5BKVjKwLxIUCwFjSpEUqku3hfvnRYKj9yDFQrW6fwuMQ9Cj5my987rgFAo4ZxBpL0kpxwqGFoMwLPIK5OA6l/sZMEB5WqhGyDkRmDsjTNelhR3ueEseaszjv8F4glATnAwlypIYEo5M+oN8G56+9wAlBZkrm4hTnPA6H1Apyy57WPJfPX1gFDGHMamoCUofcBgCAUqqKCQg0ANYglMRLyRfv+CrLWY84gcIFENIpgfHuqBQNdZCutUYqxcqgxydv/AxvvvwlWAIZM1WCfisXooYAzhjw6isw9j1veQefvOnzrEhPr99HaoXHoYck3bMNaDJOrbk+hxymxzFZhRBjym96165dx/eJ1uKkZ7XXDRl9KUEJpBc4Z0/5BdsKGdc2fC38aDB5Ablw3PnwNznvomvQqkmCM5LOOaPNgXACXZWNRRUYEKpeLBKFJWGPTrj8nF1cf85z4CVvCvcOGOAph76AH3ulWpAq12X4WlsEtIB4zJ/wOG+H9zZCNd5h9K6hlkUjvMV6yS7ZgcUFrlw8D3PpNRVpncci6FOwSp+1QZ9+mfPUoYPcdc/dfOO+e7n3icdCaakUeCVxUmLw9LOcfpGTtlvY0gxJ7trtdkCcixJbWOJqYxuCHtXYsRM9dFasX1y3Y1ZgZjOb2cxOy9YjwMlG6bEfXzNDq4lAtVM+8YXP8o7rX8d5aQuso5WkW+17DusoX3/dK/j1j/whOo3p2TIEO0lEYUrQcizjVRPLuVnrwfTxMCGDK6UM/gAeiUR7wXm795ISbfnzbzrFdRCi6m/GFGKCL+OAx44c4J6HH8DuS0Nb7klGwZFUoa03z1BJPCR3dHmJr4hja9DNNuaVtqNTnIFwTzPwdqCcCfKoiUAohSkMwoXWNi/FOJeCHx/b22H+C+eJpKJ0ZeAD8YFXwluHyEpe+7KXoCjwxIg4wg5yVCuh2C71axWo57GBk0MoLB6tJTmOm+6+gzWTE8WCVtXOaJSg9I5IBRCjlnZtqt80AZBuPkC1Ir563908lh8kSXbQcXJdVLItzDNUpRLAs9sXcNHibh5wXR4frNBJ58GH6vk4joctMDMbBzMcntKY9aN8XaWoGNIzNEFt3W63jw8gkAKPI8uy9c/Sb1wBUg/Sk1m8fYP4tCb8nNyQi0jw2Zu+yJsvuiYUYdogiTcCQs7kJ13t2FW9scOTmxLnPSoaVVwowFULjBQSTZ3VgBgxLBXy6/0CZOP5iSkOZDh8AyARSKHHWnu893jnxnk+hEBIiRUKL8J5Wm+HMleRkMTV+c+TspeEPG6HPvOFZ/HuS19GgWOFjP3dI9zzwLe4+a47uPOB+3iqt0KqExYXO3SLDJnEZGVBVuR474NSgXMoJfHO4YQYW3zq8VlnMf0GflOd6Zy1xMxsZjM7G81WVRBqCgEbQG4NPe95rLtMXgHzZjAg7XTIBgPSVmdLt88Wil3M8YrnvZBP3XMbVjtEGgeuLC/XBTr1dRoxDorPbIqfWQcm3uOqfT2VmqsuvbzKdm6DDEIj4FhXUFGdW+ksWmos8PWH7qOM5bCN+KTvk1YYFypT51spnXYbWVrKw2vM64SomkxpleCrA3ftKrJ/MRuET/ve+yBDaq0hThJUFGMHfbKyRKcJURJTejMMrt2ExOxW+37CgygtcRzTV4JCOAwOKTUYy6JOec/r30pSrWFKVUC1Ddx6bjssEtX8E87XWEgAniTc89TDfOPxhxh4Q2R1qP7QPnCfOBta6DYg+ayByawoEEKwY8cC3dWMG279Mhde/47w3LYRcjjJP1QT76Y43nTt9fzmF/6K+fl5MmcpjCHWGjWbwoG8vFoDm/Glryu4JsdaY4y4Bk4hhBhvgYnj+DgBkFBEVVqD0ApfrQiuQWx5qswReiPrXuR64DQVX2Q75c4H7mNYi+I8W5962EQHtJJmC2QugkiPpK0C03OQjIpQIEY9ZqYssWVJq91mkkL/WC0wk5PWTiw+riIgqglmhAjo+jTLTYHSUagSEYqkAkOct0hvMUVZDUxBS6Wjfi7viZ2jrdvsmWtzxdX7eMfVr6YAlunzyVs+x2e+8iUeOfAkpZUcznIoHVGkwAuK2vmoEORpi6irxpWfYD+e4R0zm9nMznZrtglGQ1LQiXVUK4wTHBz0+fhXP8cPvejNLMx1wEEaJZziLtmjuy5AhEBS8u2veSOf+PKN2KUYmab08wGtOKk4Kxr8C822YjEDv9eNiUbLqBOBB05Kibc+7NlIrrnyuZXe3fYay5OgCKLikmNUHXrr3XeiFzo4UW5K80BpDbGKidLQElQMMlIr2Jl2ePsrX0dL6pGaohiJB8R23CeZ2dMbqzKN+eI3vsZtBx+hzAuUlKg4RmhFYQxUXDahna9BMFmNk60EEYJilUf5wGFg8BTC0xKCyHiuufQKzlc7mCcoS5q8JEkSKMsN/fHTHr84i1KhCqqOMZyAAY5Pf/UmjpQDRKpJkEgBGYH3RFgRQrnJ6rxx7DK8Z6RYtTnzieKzt97ED1z/bSEQyi1UnJDbAQua/EYD3lve8crX8Wsf/WN0eweDfhedxrSSFv1uD6VmC0BzPnshEFJgrae05rjuu6uqP+bmRi25OoqiE+pzNMZUAzgAH6JSAfGn+IJtFdSrqiSwVCPHpGYjXy2zEKQLAgxqLV5VVQhnuAPKBJNzYQq890RRhG7Wb1hbEXlJkBKtIrSKpjqgx1sZIyqnwOGH5WSBfFYMpYitt1X/r6jAmKpNxgXgJtZpdS0e6yy2GoxaRCAgSSOsd4FQF9BSDdnKtFas5T2iKKIjJa4s6AD7ojZXXPtWfvzatzIAbn7o6/zp33yMWx64h27pWS0KSlciFyJKPyrhln4c2VcT3lENrvkp7TAzm9nMZnY2mROjFtR6zazJyOs9KYoiyqLEtCJ+4w9/h+970ZsprSHyMuxFW+x0KoAs58XnXclFS3u4y61QVJxaFj9qd2Q8s1i3TMKsDbK5TTYz5MqHACfWGu88GIeykisveRZgEUJvj5OeVCKsfuZr30aGsvwCzzce+ha5JpCV+pMLfx0hWeSkIEkSnHOUg4zUKXa2dvCdr3szl6nz0AyFaIZjNqZOas406J6uFcAqjm424BsHHqWf58EvTiKs8ORFiVDRKLhivd93MiqWm2FaSLyxQ947U3EvLnnNm178cpwZIFQrVGBVle81ma/xhkhsDwAAMaIkcFJwiJzPfv1WTKyQkQRjiZRiYG1Qm1QKV1qEG+ccGoJD1fNJ0oS+K+maHCFi7jvwON9ae5Td8xcgnNsWy0+9vzQofoJCDpCIiPNbu3nh5Vdy88GHEEIEAvFNaL97xu0/zQoQ70YcIBx9o67/rimMoo+/PcSP+m28xwkfMudSBhWO0+CACTeSLDWN6g4n4HB3lR3pDroUpFYQ64gyL4l0dMZn8gXgnR3pIAtBS8eNJ+NHEIhSDW17gnSLcxCpqYvRiZ2HmBgRo5YYJdTU31cyNOh4V1VgSBHkr5rP1zlEJa2LkkghazHfatBKWkkbVTuqMg7XV4bFQ5cFaSvmtRc/n+t+9Pk8yRpffehePvg3f8Wt99yJyDxrKkiJyYajy5SNrl5kZVVKNXM7ZjazmZ31TodorI1NDpDq6+6gz9zcHP1uQaw8dx96gBftupDIia0vpfOA8+xM51gDXnTFVTxy360ccRDHMcYY9Dr2qvEWSDVzQtcF9bKxjxoXKkCc9IHcsPTsEzvxLgept7QC6HgBCiEUJZYcx6OHD7CWOkQUb8r4rcdZ4QyR1rSSlKQA0xvQcYqWCjxrqrqntbrwqBRhpkf0tMEDoEQSO0Ei1BCoy/McGWviOKZsPGTH9qv6qsn9pRSBf9ELhHUs+ISXXvE8FnULCkMc64A35wakROCJtwH4USfZ6wRopDQWQQ/45sHHcfMa5z3GGJQAKw1OB64/a20Qs5gAPcaUHJ0jNyXJfIesX9B1JX9146d57lvey0KSbOm1+8bMVU00RDS+dx4vCr7vne/irt9+P2UrYmAMR3prLLQ6GG9mE3kDs002y+MgQnUNQExrrSnL8thysUJgscOBLIQA7ytFmMAAfsodMBptMIzLl1ot6NmCv/z83/ADr3wLEaCTGOMsSp7hXVSeUCnhpz/gdWolTZIPIQOL6Uky2cspoMHRxA0nz0nYiZWg/qVao9mDFCLoPQ9/VQxBmjGK1cnriEPLTAwgHBf4Nudf/BLe9qMv4WsPfJ1P3/EVPvDVz3BElKAkppLXitopqIBIdwcDlpaWyLIsoNPGEMcxRa9P2m4FSd/ZWjOzmc3sLA14p67t1b9FUUSW57TaKXmRcdu9d3H19ZdQehuqQLZahcA6hFQkAn70e36Av/r5W5ClQySS0tphAsE1Wl9mbQcbjAUx3jpaBzjGOZxzJEIwn7aROFKZBA6QbWwS0EJVSTzFZ2//AmquhWUQWr43gX/DORc425QKGfAKSJRWkIiIBEhrn8oxnsSSDi9DED+zpxMggcYxp2NsVqBaEicqBZU6IDpKi8HJchhuhg1jNAnWhYRn7CVvfPHLuSjZHVavqAHj6sDPJ7YJBai3LlAnENrkPZBj+YX/8f9QzieUPvA4SB2qsCKpsB6stRWHX2PvmWiJcSJwRGghw3SJNb1+zs133QFvfx9FnhMnOoArSm0JJ6RvnHsdz1bhCniwRUHaSrjmkitQ3ZzO4iKDrKTTaocKh9nUH3JKah2epZaSVislz/Mwz71DVHMDIdYl+Ovn3mq1msCcPO4B4VjPAO43iSTquG7AFDeqLsVtdTqsDHr8xac/ybIZYAksyVLKZ0bg6icADn+Mm9RMI4hxzpQTfZ3AK47rWHfe0zznKcBNffp1CWjTdfbNk5CNXxBAUQYQxEvmiZj3nh3Aay59Pv/oO/42v/dv/jPvffkb2JVLdvuYy/aeh+0OwDis9ywsLXJ4bYW1rE/UTvFCYIwhiiKibdI/OLOZzWxmWxUkNnkxJoERpRQC6Bc5Vgv+7IZP0KXEK43fDlxcSmGzHO0du/U8izoldYLIVVKJU4J8P0Ua/ax0PP36e+AmlBiEEKFsUkmEkqQ6IqphJSG2/vnXBbJNH7LhrEjAYsmxPPTEY6xmfYx3p8y39ZMqOozLY46f9zYAEM/wtUsTSGWbijrN+S0nKtvcNpO+TjttBoNBABKsQxqHzEredN31tKuW89HFjcZ205/e0megVJB6JiQ2C29xRNz9xMOsumK4ntRcUzXPh2isxXULu2iAUkNVR2uJoog8z8nLAhNJDg263HXgAUgSHH4IfngfiDPrSgB3OltkqjjHTXyv0lClspMO11z0LLpPHsRlBQqxaVLczxSz+PDsXWh/qVVgpJBIpUK1XFWgEbojqltdreWdTmccADmRG1xLyRxtQT9Vm/BRb4q1xJ0Wdz/xMAMdyC+tdxWYfoZ7MJPoQrUf1ouCbRxmg4PGQniir0zBXo73qMEXH40O6kM3PqSxeAtXS5cFGTjtw3sV1WGbm1P9t1E0RpzqjKXIc4RztJFcQod/+ZYf4SM//9+5/qKr6D7wJK1CMB+3kVKzvNYlandoLy5RWIdBoOMUJSO6a/3ZqjOzmc3srA6Cm8GBm3BMsY5YaWxpkHHEw4f3c/vj9zHA4aTc+h3YWVSS4ErDHBHvet2b6RiB6WVjvsWworTaT6WfAuqfrYGkn/6zJl+Kk8Fhb0fJ9grZGz6QmgA/Rq6VwCN54MnHMMKj4gghTj4AqcHDSX/K1X6bBIukRAbfRoJV4PXITxJVcmd2nPhRu5pqguPNTQB8YlIueZtMegfkZQGRIooifGlJcscumfKC868grhhi/LpBNzb0t379GLbBQCEEX3vyHp7sr5BXC4j0FdeUrOZpA5ewNUGtn/JzESpkEh0haknpJObgYI0PfPJjZIBhXOijGfeejoqQScBDNr+vHtKgHNBG8Q/+1o9y8Y697Eg7FEWB8bPacyYgYO891nsKawKYRaNCXzQAD+fXxf9JoyVKKqWOe4GvORm8ryiZ5PjAOVULRhOhdRPobf31YDCgPTcHcylfe/iblEDhbCgB88+gFM5RFEp8Y7I5QumfnfxT//RexYm+Tiy8ZQVelA0gwzTBDLH+GptVL6rexBrXWD97X3k1TjiMMyAFKtJDwjGf5ywimDMFu1H84o/8DP/lZ/89L7joclqFx6/22d1ZoK1jsm6P1dXV8Ld4ClPS6rRn7S8zm9nMztrAVzT236Ys7HDPdx5TlHRarSC3Pt/ihlu/TB9LuQ0oHPMilMkqF1ol3/2at9K2krk4ndq+62TFOzbhiM+s4ceL0T7vRWDZRwqcgDRJKgnCbZB+miAzF4wTudY9+BKJRXBgZRkVR8RxxbPm3EnzQYjJALvpwzKuQFNWvlEB5NXXMwTu5AIniRze/6bvWANTzcqC4fNh+1SB9Iuc9lwnlPkXhr1xh3e87NXME0FhxoPsdRN1e8QtUoaKMIMnR/IHf/kRVCcFrcbmRfM6lGsoTYkGSXUd51SJYC0VCoFGECmFw5Mpz033fJ0D9LCNCSREUNMZBslbsBYNQ54q02wJbaQx8Jyli5kvBTozKCFBq21VjbQ1KOAIexBCjEniGu+qfaaevEGJTEwUd6gpikjyhKo/pkjeCiGwp2EQ1U5YXfHQ3IS9gCSK6A36dKXlTz75l/QxyIrZWZzpGup+/YImGkfd6aKro1lgMWQWn+QGOdFX30BWTvAQfryNpfl2NUhTgyG2kRXxKhwIEAaUCSpAqnq/snIQ+kAXz0BKCq2HToMHtNYkcQKZoWUEcwhSSl644zJ+7Sf+PT/89nfzrLndcGAVudJnb3uB3a15TJYzGAzQrQQ7q0Cb2cxmdhab8o1AV45vR05ApBRFf0BLx5RlyXLR5yv3fB1HRLnVQbCApNPG45BxjAZ20uLKcy9iIW3jjB1uV3UAVMpRBYjyszaYob/RCFTW+agVUb7Dk0ZxpR7n1mentxIEaZSg26ZDUtqqDSYQ6vuKkHGjiuenCRmN/+QobRZ+wjeacfCeROw0eV/leLJWVuIKNdDpGi0wbpuAIK1Om9I7enlGJCR7SPiht38nMSC1Hg8TtmG4U7e/eGcBwRP+CJ//+ldxVRK9llp3DRCkBjqEH+0zwofnVVsdL6RpiitKIguydPTzDNFKWKbgvkOPBz5BGGt9qVthTkd8uK7trl4Uq4vMvUGjETgkJc+/4FLigWGu0yG35WwSrwcuAqBWUXioiuvBV7ykY2tpzVVaPeemaow8sYVktMHVwMLpSI4MCaMai9MkT0WqIrqDPrQT7njwPp7oHhhuOeJMh8+PdfoTfSe1o9KkyNj0czme14kyUzVxTHa2TrbPDCtYxPpfaPLBiIlnbDAUpqB0QRLYGxNIoXREttpljog5ILI93v3i1/Gf/+nP8fpnv4DOSoE+MmBRJURe4K1DRREra2szBHZmM5vZ2Rn8MsrETfbH+0YFyFzSIlvtopRCzbW4/8AT3PTw13BsPQ+XB/KiCE6P9yQ43vaq17F64NBUB7j2M2qHeyaDztR71PzaiFCW7Jwbkur7bdKAPOQPq/wHy0RZugoEkgMKutmA0llMXqCl2rQe/GbFh5sAlWqeimnJq3XB0+x4moR0I86IZovLsMJtg/Htt1gC1wtQccRqr4uOI5IoZqdIOI8FdAjERoDARmDZFk9CpTWlKYd8S1+663bcfEovGyAbYOCknz2ZPJ1UIauTpc45in5GW0WklfKn1YIVV/Cxz90wrISXUo4BIFsWyw0R9wBwKqGrGEnSJuKH3/VedukWxhjyspwRcjfAjOGtrCpBhBCjTqJGdUjzb5p/Z8xIUUceL/p1okOlmS3YLBCkiRSu21Rs2HRlO8Enmi/fejMOKG155nOATCwQ/gQXfb8Jm6eXRz/s5NGo5vACFA59HIeaOEQFuY2Xs4SL0h5SD20PHefpAC3v6aBp6ZhIqqBepARoKFxBujCHKUtSYJ/qcB4pV6rd/PKP/gzfce2r6KyVlPuPsKc1TyQVgzwj6bRmK8/MZjazs9Zq+Xknxvff4Z5cGuZbbWweiNvaSwsMvOGDH/2zsTbMrQp+M1OQpCkIcNbSQnP9NS9mPmmNkVzXfkadJW6qzp3VINhEC8c6h7xWbxNgvSfWUVVRsfU3z7O+B3/azwCWzSpGhDJrbx1a600Bv5qEjc0gT/nAcxbXc8yEQxuIHSQOYt+o4J0dT4OIbgMfesr4llN+dztYludY51BJjETw5pe+EllkeBeYXeVGY32bgI9uGLwKCgw33PQF9GIgo5TWjwHOTdBJ+PUqPFNblaxDOk8kFMpDHEVkpmQgHZ/+yhfZv3YgzLeGCoyUcmpbxKnEPYaxmBsBsXU4PigGlDZoPV2+4wIu3LmXvD8gmWvP9p8GTjEJaHjvGTAYxvliArCe5HHSWnPppZf6AH4eBwrmCe0GFlDWExk/dIiaIERTn7lZKuk2sYXAyoaEECAqKM17TxQplldX6JcDPn3TF8kBrxRmW2zDJ7+ANImbNljjj/oeJ3Mcq9PlWH+7DknzGy8Uk3yvVGNvKl9I8+SsR1QrovDB0XXeoWToCaw5a+rsVNbtkgCJM3S85f/6wX/Av/7//BR7SDCH15gXEcp6lJ+mPzSzE7Gaw2fSqahJrXQ1SAoq4l55kgjsKdjBvAhkdYZRWbKcKJ2ddHpnNj4GtA1HvVc0M1ZNAat1nsMWr79WuNG+M9EWUY9pJzd+5nUbnSSM9cSEV1tlgOzpH87rVrTamXRV2aCrzjW206ehbGSwvfcMipxOp0U/73Po0CH0XMJjvSMcoIcZ271Ofz410mnlhDukCgLuC6LNK17wEua8Iq6UBiezvX4DNZjjCY6G91NszTVPe+gSOVZRUFdZKgeRHQFdfqLnfvI+DCuCqn1ZChGCeeeIvKvU99y2qb6dHOvDyoqGotwAw1qvi0IQI5HOo72gLPNNW/8Eo7aqySBunaPWdK5mdlLrd4EbqotoN2qrGLZbTxnfw319G4CPGEsnSRF9g8xK3vum76IdtxGNigaxjYeLB1wkMAIOZKt87d67WBn0QutKXZEhxkHoJgg1SVDb/F0BREmM0hprLf1+P6xN3tNanKdbZNz3yAOsZd2w3knX0EvwYy0Rp3Kvlc2xNAHoKDwLcYtURdXvel52zYtY0ilp6c9qHipZgVUSAc4jXCA3lc6jbHg90lulZ/PGOuqH8utSyKH6Tw2C7d69O/xbmqbHHAAWKhqZBLPWZ8lHAQQRgsKUqDgalZMxYvCVPvTSlhKcmCCuOIHDC4+VnlKGVy88Eo/ynth6pPc45ekXA2ICH8jdj93PFx/5Gn0cay7DnsE0ls0JJDlmwcdRAYWne8hjHOoohxwbGSd2iOpoXvvYAjJ0YCoZvsbPlVIoISunQxGrpGJMlXgknbkFQJLImDmR0EHylqtewS/81D9nDy1aPceC0USlqxagaed4ArvYWXoIRgeE+eurlJZ2jtg5VGaQQpAhWCNwuwxXJbf1W7p1YIXgiMnJkOQo0BG6hI6IhwCIq8ZWs4/Yi7P7+SM8Eoc3JS0PS3ECRUFMxKrrU0BFk+nWkwVuA4fOA6WCvOqFSLVGWYcrCtJWQuEMptqfXCNVK324GCt8cLo0KO9oGcdC6dFFiWgnrGIoccM9SjQWbr+ZF9EklZ64zwjIy7ICmuVw/kW5JR6EHJUZ650Pb6h8uM6klTCwBS6SCCVopzGZKXgwO8QtB79FMWw6ONZx6iw3Bic1Vij6zmApec/r30q0nNMuJdoGFQ4jQXiB9AIrBEaK4TpWD1DXOBAe6wyRVnhn8Tic8KENU7oQYG0HJ2IYMIS9yyArhQRILMwVIxBk6HxWwaKVI1BaANqH+2FUOLCWlhfMOUHHSy646HwySvSIrnBLL70mdK+dCO1D1QU4rHQMpKNE861vfRNZlrSFpC0lRb9HO01Pbg/HYwqD8pL5OEVkJZHxzLfaZKbERxE5DiOqStdJx0tssjN3lh0WyIAjZZ/UQbv0pEphvSUXljyCQoV1ul7TIufRbnzOn9RxkgGgtpKOUSwOBG958Svx1dz1SLyl8pElqvKXR1KRW7+BOkJbukFwBMcHPvdxSgyRCFXZXoMX69cIJ6BQIX4U1XNxMqw3rlqP6/0nL3N8LLHCoWKNrMKBftZHtRJuuOlGVJqyZvt4Iah1Q7xweN/Ymyarhjbp9jVb3bxgWM2ugMSDtgpReGxuENXa/OpXvJr0cM7OfpWEOFv9RzzeOnABc9AV2W1soWUgcYLl7iqoGIvEewEy3OCynI5t1BUlWlYkIseaggoo6bNISpavQhJROoEvJbJXEAuN8g7lJAhXBQEOK8L30geRsSGcdwKvXnisCN9rK5BUmX48XgisCNmnpXYb6R0dJ5FK88GPfIgX/oPnsCjTMx4FE8+UzxdP7w3FyZzUhKNfJ1dkwzGSziO9JdWa555zEf/vv/tlfurn/xVHXImyYeGdToYqjv8KxNMb/8+E16ZUYjMjJl1YyCIraFWlYpbQB+2dHQJWW22qYupu64QMcBhUaekYQW81o9VSGB/KloM6ghjLkp/Vz7++hyJC9wryfkE0txMJtGW74pSaEu03svFyC9dAAcxHKXMyIusXRGnMfNxi1RiinmHeKbAygHrVPier/Ul4gRfQMwVJqki8Ji0hGhjmich6fUS8e8PwUJxG/zWOYiywmnVJ0jaCELDtTDqs2IzSjs6oeX1hfjuEgQ4a4RW6b0hUQtYf8MGPfJi3/p0rcV42hOPcBrn5U/cMldDgw2hLZYwgZpdM2UPCahnOJENQVgRUssoWhxD+6OPclIIIjzUCqQSRiMiEIfYRidv6OgjfnD9+NLdq9zC1AmtCJUepRk9H1cTzLqi7KNcgJmyQFpbO0rIClRlU3yALS0IU4nln15Ulb6kj04xqRM03JrDA4w8/gi4dwltaWiCUpL/cpdNqjViAT3D9k0CcdALhYr9kkQg3KBisHWGBeD1EtAGP2sxOcn1zgrlCYK0hEglK6tCCUJbEMiJyodpXVK0aXnhc9So2Yf8brXfuhF6lh0jFMMhp546XXHE1lgJBjHUWrXWoehYbzPftkMGvDovg7nvvYU6mrGYFS+02a1mXRCcI75A+xIsOia/jR1x4Lk8zfowdfP3rX+dwtsLOdBEPrPaWmW93SFS84X3yp2ARFqJRNVdXuzggyyFJUEISaGIF53fO4bXXXMenv/ZlWIzIT7ak+Az1H2WjKsi60ZiKLLSMILWghcRigVAFpKu21mPxN2k5IRWz0Z6RILGrPS6a283gqR7KeJyTLMgOSip84YYASOBtmARAqnaVaoCfyKtvvI+2EolD+DAxrJBY6Yhai+AsvW4XVfbZE82z+tCT2NUeyUKKmrUwnL3WiCKaLTmOkawuQqKrf5+XKQ7N//q37+d9/9ffRdpQVXT0zzjODNfTGP/PiNfqAfhqu3FiVBKsLSRCMV+Ap0+LNik6PBvjgkxlo09/K6IH0+2j2y1aKizIBsdukXJJa5H9vRVEGVX8N1UW36tRS+BZ/vytdBgfOJp05EnnOlyy+zwcOZpWFV6qdXN2u5Tyxkj6D+3nvDIhswpyQSwS5myJzCMsEqwaXq8YOq5hn7LSUUQx0klclqMLS2o0F+7ai+9lxDsm1Loms/abCTxvcFN7vR7tTiesf+kcFigoOfDEU8SlYHeUMFeO9t3aQZfV95GKGZTQ7ilKk2BKR7vTwoiE/bffi8ThHEglTyvwMQQ/PCgpq+qXUNYgpeCSxT18/1u+jQ986uMMZGjBC0C3q1o6RjKMR5u/1ng0Ems8UkHsU3Lj0T5mruurMHd7gSAVNRYRMC8iIh+ykcZRPdtm2084/7qyt95Sa76vwpSkUUyqWizMaXbJNtobYuKAsgi5PbI4UyLDkJAVQEF2aI0l1cZla7Q6Ced05unLAbKsdCServ+aWQrrEAra7Q4+bpFEkn3tBbIjK8RL80TNlLMYXwdnGMhJJC+QdICLkyUuae3iibVDdI2lk0h02kJJjy8c2k3EL3IUx8h6XX86++BJAiBeQLc7IJWaK845j9dccx3zxESAKy1oSQNZ3j5Z08ZpxF6jhOOJ5QOsPPAk8wikkMRxwk4p8SWjBDoOJyYBkKcfP7asg37O7bd+letf+jLausWOzo7QMuSryoLTxAXSNEdYg0OAnVQRvsUAmcspIsFLrnkhDzzyEEfyLifdqXOG+o++iq2GAigizCPtoeUhFbCYtImrQg3r/DCwE8d4rGJ5edkvLS0d895ZHwbjsskwWuBJyLH0yLGUtEgrEssAUIRScBecw6EIkXt6C0C1IIUCc4mqXl31/g5Hgcdi6dBG40ImymbsUB2wdksG+My2kflmtqfhGFf/ZnsDVKsVJKmwFEJxmD6SNhklUBxjP3HHwMCf/gb4zHgNIEgNgPiJZ9ChjfEDdog2vsxRZVB2CqVsjEqDt2rsVIPGljlOS8pIsmZyCmVJRIdVBpXv6saSd+vHwNn3/EOxqcZV7CkCC2s5F8zvI2YU/IdBMaqzt41YYCsrQAxBavsIXWJaDOgDEQaDJMJhsVUHthzudyMgxOFYJUMT0SamFUJlnM1ZVC1SFFF9fZMOszgFY3mjfxKQmxIjPJGKh/d/uVyjiDRF4/poXCc4YhIy+igUMZoCg8DTpoOjzw4fkRL6cY8OgMhTM38n0W9rKJ1BtGJWKOhhyaFqx6Ii4a6XHdn40+njPAB4Dlt5KoqIkhyQ7KTDPJBs4fI1rHj04/thIRw5sMxgeP2Vu9lQbqvbUUcJA8G4UptA4ytvjCJnXzzPAhpRVmlPtbUBWXO/Ga8IDRVBA1OAjlk2XaROKXCUeCyCVbfGomw3WnlP3H9NaWFx5JQ4PCUFGoEsSs6Ld5JSteQ0+4zkuCToDAQ5ued/YO0weaRI0jkKFD1yCiwFA+ZpIyfiF1vFHX7KundirycHgITds0WXFUQ34+K5fagyJ/ESraLR207bM/02AENq/7rdopBwyPRwOsWhOMwaCoEmqtbcBgBVPYd6LRJP8/5HOHbSwrkBsYxw3pKKBFMWREqPFwD4EVDb3C7UJi5AzWc09HuMC4INZYFsJeTAgNAKfsAcpqU3Q4jhzPQfa91Zi2twxoUZGVf7amI9O1Sb2EswFrSqQJMQc0gA74fdLi972cu46aabxPEBII2SQSOh6w1OaEosEYq6pSlivayp5eQR7OYGrhrBxaRWusFWesC2CqoUFBat1ayUcGYjNuqx7MCoDK1YXSOenwcVJP0GOEpC1UJ80ovg2V6BZMb2AVcFFaLxHHIzYF63EN6h0FWNtgNrIY3+/+z9d5gs21XfjX/W3ruqunvSmXODpCtdXWWUhRJCoIhQQgIhcjIGHMAW/BzANuaFF/9s/GC/fo2NTTBgEEFEkxSRkFBEWUgooZx1k+69J03orqod3j/2rurqnp45c+LMOVPrPnXrTE9Pd1Xtvdde67vW+q6DBUC8j9eiVWtdNLonwjt+vtKqM+IqbeRHU6J+9jObfg44X1OgMErPOv6yc60epP8UgHFV4rWQ6zz1pRIcLsHxO1e5zJm7dbt5WwyKHEEnt1l3gY/5zfIyAiBIDHQEiW1rz4w3WBouY/G73ufUtArJEpj+zmIxiQUia4Mg6TISIdm0/FZd2gEk6ZLaQZPN52P01CpJDu8UHG/AjwZ33W8r32meW5wfAYlz/GIY0Re470Hi9wizXpKFyEEBab/b2ba+AUDUnP1lOzZY83eNUWocMLGQm4MFsNlRBbugLNZTeosow4S6LYnRKTfEpOji+X73mfoMo2yJEtt+UpbmhasnrGajhYzQThpArjdfL3gBpG6GFbCdYGtN1rYg7rYcDh3/wh/g2m32z5omEz+uLxMAm8BF5yDXhxoAwUGYVPhBjjVwVzlmWAyxzDZ47M7zbonehVy+SvtuWU5YLkY4b8mVaT/T1jXGFDv23EsBgLRlg/P+R2Njqni1lRI2Q4kVhSEjO+A5ePBL2Lf7THeONy3Dmz02GuOh5YO0aQFoZCEAYs6lF7KvA7oQcqtQGdjopiQDp1PP1J1IDYvvxd7AO7O7URA+TRFJETVb1mR9K4Ze9sBCu5KvrETm4O0KUxiWjMKmmavb2dWDH+f7DEIbW1zsi+Vm0K7gmcd2tvKjy6KBfXSYlESjwzqUErRSuOB3ZKD2WmdWf2/aGmMiY32hsjjOYthPG/YDz5wPkfsl2OTEp821tjGy3R37eQBkaviolCdikETqpbWO0YrDwo/QtIJMF780HCWQRu+IQM9fsa09WaZxtcd7T5ZHszZ4CMEhRs8YBzsBkEs5gPH7RHei686DToERa8lNdEsbndQY5TTk1/sIoDSm1PSedPu6HBaFIHMeno8Z9LSG5PRtDQg0M9pz/Bl5crq2XI1OHQwgID4Zfbk5NLVsM3t+l1wnjY+fVAxHEeqQGcBLTTO0zlOWstVkxGczHZ8MgMp2RmYuNgnyUZeUiSQajASWdCQrb4ACAzuTeOUiZuBcwEA6gW1fkasMN5lgiiHUyVmu68NhH+1H9ZicJhF/xRR0rzprfLtL5j8aTDFsuzD5VE+itT4rBcSlwuOaW4wc22FqB3hPpjTLUrAZSgZCmyl7dI3IWBJjOwCI74BmZn5vSko9Msfsvvj290wbf6SIfRxMFgcqb/aROhk4c5vjzJ9fgAJoJslChZJqWUkPpgo1mWSt/aETQdDuK2i/7OS9E3ule2EiDRHT7OhrAWcrtMnx4wkqEZ6pAGEyISvy3X2U0M+L/S5iaerIZ4YlvlY5i9EmMoZ7SxCLUQavp5EZOcBrx6ippwCg9DR7RdRiPdJbr+3jW9HFtAytttFJ1GnPcG5heKObDCmHYf5mKl4r8bp1Kr32rSm9YMzTnijeY3xoUzMhMU0qtTvKIxdV/SXnfJe5mbz0pkRFEcjToFR1RZbtQhaXPidL+7+WgPc2Wu1aJ+qHNM4yXT9TkODSj2zMVAgoEZSJIzUpSwZZ1PPamFTGwaxDNB8CDGcDyjosdx3UQw4BAjJX+TJLBhoiiZyeA/123P8CG0zS/jnUWefxCF7FNSECrqrQqefKQcoOPTIDBikynSNBkXVsPV/XOFchgwsk0vcgLmCMwohQ2xrvPTrLpwDILuDHQWcgXBWSTZ9r12LzocnsNLP+i5oFAg9yL9cBVlQeryEfzhmwvl3Vh9bcEKjKkqCEXHIq5yg6lASaucy0i+0/pjH3LhZTaIllL7aqYR8cmJcDDNEyaz8E68iMZlmy2f3oqIqLi9IoMIuAydQqGlFtgGo6neTCvXrnXdPEErxLf+gxQJ7p2d11vz1az8WCZucG0d3EhRBTqyWLBk41ITMZ1rs+HNvLjHLplic0i0TnOeVkGzUaEqqSemsLAhSDQaduvZfzf/Zq5zmodhkbpabjoRUqMbTXBMoDbmLdpHpP8FSdVDxPiCUDu2XRzehBf7THv6zbfGLR03z4YC1tWKizX0jH+DwM4Iet62StTUEBpWIZl7V2H5+hwMX+9Tg/dZIBb+vFc+ay3uNsCwFvXexmA+TG7OtyyskEtCYbDCKw41wcX2LErbtOGgDkXDJQL2gNi6IMlomLrX7NICck8s7au9YYmnGSL+DSvN+Z7XLoUMl5G6rbeXBK29SZFumexM/ptJgdaX2J9XWq4RdKVxIEdJEf+CLec0lJnJ9ZnmPLcmpMe1AqI8svQhfBQFr30TLPTEaRF6mEwS9c+wfbPPgqkg6yGSSCoa3zLRK7qOwxUQ7F6vXApJpOiKqK50HRKq2ZrIL52zjgm8iWCvJhTp1mdaNrq8l4Cu5dKv8xQPCeTBuMShkfAUxbCul3cxXOzUm+kOEN6RpCykz0HgmRgF2cS8+mo3/P5XyV+G109iUJETSTmZJOacGP/YrZbwqq0oLzNVppJDgEIQMm5Ta5ydBKz86UecJJOf+J1K0F0037oAUp5yGBIAEY5Hms1FZyFg4S1fmW3b69d4CvigXU2Ry6ILpL86pYGkVDcFiQLaCsCzMKC5xzKBXi3N/3TD6K0iE+bDJBFixIl9olKhSBgMUhqbf9Qdse0ReQmeadLnjEB0TrHTrCL3oCRxQEkWYD6EbCUzmCZGZf63b6QQcjOteRhCtER97kpoXllFHTf893bmkczCpAZqYlVB3QQeXZrvbpgfGeOI8HjDFomZav7XWB+TCfkgCntE1pGDSMxjmHiMxE3Lz3l42gXItB2mSUyGmiTdaJtPtW14SmXlglA/ks5QiBpr44xZyMJKg0oniHpQHKzA9ytkU7VV5OPA6f7K8uqf30XCiduNgqDDlG60hUHxy5HHCa/i66pOVLSBlMPgGSymQRmBAVSx/1+UfY20eZZ8yHuL1zqAXzP8jldcCudvOvEjvtNCZQJ06yLNkaIiz0XxoMUF+GBey8a21JHzxKFNZZMp2IBop8On1Ggw5gI7hFc26XjIpLKWVZkmUZSsVyV6UUtbMoHYMFCtJ8jzpkaVBM9cgl8h8jrhDSgMYaT1/HDCxTFHuWoF7UtbfHGEgCZZprUSZrx0+r2GHOBUsIErNFpNmxPNLMX3z7+uz5CrcfhWngdFdQLLQ2ZbMfN9kfsg/PZB8zKCT+Px+PYBE8uVZoLR1oxidm7bhhNv/t3C73f55XZn6X/UwvOPZ/g/02c1R2w0VrqCXXSelV8yi09Q6Lx4fE9C8KnYC/853XR+28cDE3Sl4UudKdtSsUiUDRBXugqzMCZi5RX06jK5lojDYx1bvz7qZbRLw11QWuj+645yZG4LxttbfzkWd/Pvt1xmEJh0JltCSuWjQ6dSaywVG7emb0/dzUbjenXCcr2hOCb6ORQaAOvp0fBxb1DQFbltQpspjlOUbP8jecy3iHuQNogQ5rLS5lhFw28INZkj3FLF+L9w6VivRigEVw1ibDQu3rvl0yu5qzT+aXHBIrdKb6SfY4mHtI0h1Ln+wvP9P1yrsKTaw0KFKXo/heiz7kudsBKOsqcpoMClSWTcseFdOOAuep/0ID+KZnWdkalzKTVZegf0HUu7dKL5YTFQHsGotLXV0MCzpu7eJgXfj+3fhBu59J7bfrUCMSV088h4VrstlDHAuqdy7zkqtThmSe5y3A3WS+RRup4zviE6eeTyElv9B/9BfRf2z2GWkAhiyL4Ae0e94lBYrmSnrUorUtu9nHPoGi0X6ovMX5OvkhpM6KvjPP5s8X9vwOhf2o2Jmi2V2vTeaHTEHB/ey8Zr87REw1iZPIBNNCc750aMkW9qHugqpq7rVzPTfzYa865B3zSOb+8KwRRcVOc6/fgq4m8KOZB3oOYW5+Pc863Yx+g8yLRES7dhXKaAIxsqkz3bJuR1KjufN5zvur6byQH2HGy/QxIqYUOlMgghE1wx1wUI7DoEPZ5azFh4CIQkTwIaC0bgP7qi0QnGo/1WmPe+TGvR1iIWiDqMYp0IvKrneCHxfaRuwijL9GYgQ8jWRzb0btzE9adKk++Gmme+KhmGYeCJeaEUP28UJjELbD4COhqTamjcCdy3jP4iuhzf6Yr7m+1BQZQqox3wUYIQFbbae72mEyjZYGAAotd8Du951G1MdonBAjdYcB+pAFVkzYr7Eusx0LdnurSl0MtFJz1ilXRBQyz/KpugnTSLYW1dHr56f/AtOadUTQeT6j9+rakWV6h8N9hTy6K0IaZqOG30WIXGPiA0qbhQpo0bo5/31Q7fn74D3WWpTJ0KJTFmycL8GHHTzEzRwJu1zbWTekiyxZKidpKgp8cz9KEUJAKWYzeQP7yvC8GP5j833Bx+uq65qiKFqzMsvzhSDFpXA/mnvZq6PazEWE6V6slSApSNgYRdGu9oioq9t+PE9SaJVsLxG18G/3Tywb1JSwLgQwESXPBmbPmrlFE7HJFD2Xs17wuQudqDA3eWSPHX/P9LBdiA17uYLDAHOgnjDniMyOfpPXND9/M21mrJs809P5yrT1mLT8NLTgyfnO/6vhfFYLXCuM7rQD9X5a13fA1qD46QCbpuSpJTmcddQbsEdLV8WoC9J/V/S4C1jn0cnx9yECjSpFr8qyZFQUh9yAjjtQ8AGlVMoAUvuakoljcQbHi93aUkf7OUf5IOA+62yshQ+RsC7LMkSpqUMb1N7jfI7LMwTwPrRcGfMO4EW3PM9mbNvQdorxte2Q1TYogJx1nqu08Dum6YGDd/O3u2iv24/VIzNu5ALVrfM4qJVL3XXSQwkhdnEQDqU3392axuOS4TA6RoPhVB95D6pZ6+epBy1QO0emdQsNldahtUZneuEzV5fXh72qxfsIwGqZAiBa5WeNrkrYOVkuxT4pohjk04YNwfmYiexBdThK5jPg/aJ5Eg5u4jRlL0opirSniyzwXP0eCkp2gk8X4j8mjABRoJVCZ1NC9sl2yWB06W2PsMva3mEaLyxPVFhbRfBUKUSp6Eu0v5XujrP4fLb9+5Dbj+4s+1UgzHSWa1oURJLu3S2q/QMgbd2RSSGbnRM3NHve/CDPgxRyjufuV8likGVh1HAuEHHuyqDP/rhqgI9dADDdqfVcpHC74Ieta8THCE4752tP7R15kza7yNGXBcrtKJ4XrcH0el1WMYqg5lzBRYNzuaVJDdJzsyM5cdZ68lzv1D8pA2geeDuK498lEvad+aCBUVEsHt7DZPWHZGgkY87Z2DgAoJxYisLsavC0ZXXN/YfEV9UBPi63szOvprQxeALeOXRm2paxIQTqsiLPiw7Z2ILzPsyHxDWKUo3RL5cW+GAPXTz/IHRKmA2QJ+fXu4DSMguinPX+ZdYe8RHoUkYOfP62WM4CCyeEnfPPdy5ZB9V+SJgPJDQEdN6DmFmkTx2SZSxn/9VwUCyMl4Wuw3Ye+i9I43zpaDqnD887WYOLOKMELjuHw9Vq/uWS4uYuZtzkWrf7uXOgdOdRhz3AhEu1T3aAgch3GQM/ovaeA/OZEQe1fzZA9nx2n3Mucm2YbEpYye72eDfKL3u857yfL1O+zAAUw+KS40XdZlpqN0DnLOOWZYOOopaW2FWIfFb4sPcKuILtxyCzzTbVgooPFQ0KXLNHNZ35zqI/1X41yNZ2OV1lOrWZcdGo8SFeoJd4NIi/MBcd3RWe2sd5rqBYOvU+YT5vZv44W/Qh0LesPAq74C5jL3PTZKbUsjM3TJahswxCwFlLVZagFXmR7fyA7gfJBcz7q+msdnk+CrJRnvpbxRZgEbI/JIZf0+uxMVA6CK/SQl7oxTX08/rrCI+/UtPnpolDbStLXdY7jYJL3BL2fKQsPXW9GGjPcrNjXs/vT03zF59A11wSJZ8L2LKe7cBxQHtRCAFtTEvKWFUVIkI+KHaOw6KgwqLa3G4lhMzyzYUQqGuHtf7yrOHdeC5SpNVOKmxZtazrKpfI0eLY3/3vUrql1OEAP+b3u1kDcqcJ1fClNGcJQve/kCZ6UIKzaZE3GXwOrA3YOlDbg+/EF3Yz89KL5biK08CGiOMEKEsXsz/UPsZ/j3NTPqYlcqTYyk3J/IFyUs2Mx8KS0d4+vSCZbNt27jcZuzjAJt9RpkC136uW7xLtjyGldTiX9g49BVIn47rT7XJ3M2oeMGtu0V2G59stbXTOUVVVy/GUpeBgUAv8tDnfLaR/Cxf3+TdN2CZVfBrOg3PxgU4m1WVVwzsIasPse2Y4XdIAexeDLr6DSotSU+UkssdxZduPTQVIsxfNu/bdxi+amP0hDeHtWfSn2e/ADZaLyPCduhYaAZd6a3cJgiQhNAHQvgOAyGyU4Zx91w6E1kwON98JZgHS5yVej5q7lp0LmEOTrtrLJVI+sns2UtfHXWTEtmztyZhVotGpg0XYj9Ny1A0Y2Zl+3Y0k1tbGdpsmEsOFBJvLYeHiUWnDVEJD2F3agHcBk6tprWroAL9zoYUQju7YN3aOdzFQrDUUmTm7oXlIxr8oFGWV9r8s7q3eQqZTo4guWajsXPaq4whJyiAB0FoSqLobYHZ57r/2LpYiDUfx56okT1wFVVnHDLdwdhUnYfHbmvv3npQxpdBK0FpTVXaP+7sI958MiMjDsLMWWCRm82idt49/e7uiSLwQtgEB9rj/xlFumh255HUYc1hsip3PccbPczut9JnOC02XxgRkuvSLZuRUNk0Mruu4vrUSvIAkGtQD62hEpAKM16AWtgjNUgqAqxzZMO7rudGxTM927MPzlKr0FIUCB6ZjNDtnGRT57jO8aY8pfTbyhchgaKaeZYPuuegBmVSi5DsmoWK2iUK41DZc0xzNTPVJXUNRwGCQzazNxulj3qGes7U8l5dGy7lU0pUOSC3iRRO0TMvJZRrFl46fFroAbRMQuAjmc2iS03Ts5ubTYitLx0AbxOhz0psXikHv3DRpMx08sxkjzTc7D03X3hCiHdL9+fwv4Mrx3xrbelF3IxtiifV0T5eUIePZ2eLpHAGQ5jtvu2OLd77zfdx550kEjSeQ54ZxVWLyDFAt6CF4VGhSJyNY4cWjgjrns+DRPp69KJzEcxBPaKuA4lmCSizC8WcvChX8Lh1hpnyz3/PdXx+jNWEegezs/HIRNMlBgywXIYXsSnfG9izH2C1VSsd2ic47vLdkJiMgfPQTn+a9f/N+bE1qwmymWjdZzB5QQRIYp857HVzRZwI+OXNhHgARj3M1X/ag+/Owhz6I9ZVR0lke3U0nO0j3wXt8AJM28s3tio987FN8+jOfZ+PMGGOiESth2gdGgkobvST9dzTHHzxlNWF5eYSzFePxJve+1w18xeMfyzXXLO8sEzuEm+/HP3kb7//ARzlx4gTDwQjnHDbENrF1XZNlGUF8BABaq02189vakiI3aDTVpESC427XX8MjHvIg7n2/G84CAl16ZzhTGlUU0x5GShEQ7jqxwav/4vU4n7WplvsZd2b2ZQfekmeauordX+57v5t49KMfydqqwrStkP1srqtcvPvdrXtca7yntpMEKOvAS1/5GjYnJYhBq+KsdGCR6C+SIgfrUov0WAc/GCi+6YVfc3BTu9tGsHEyZNaO/N3f/bOW5L6rm336W5V+F/VYmJZ1pTeGkPg+nGNtZYkHPPA+PPjB96coog+v1EGv4b0J7lWmsTX8+ctfzYkTp1ldW+PUxibLKyt4D8Gr8/0o5K8AAQAASURBVNbfpFaVuTaMJ1tk2pDnhtOn7uLa647zNV/7VI4dW+oE8Pwh1IW7s5RcKSbie9/3MT78oU9Q20AxHGC9p3YVwUReBd8BmXSIulyFbunm+e5/c09SdlJgTiYTlpeXseUkdpsLnvXjKzz+sY/hXvc8trDRw+V87m1780VsrEmstWg97fl0x13bvOfd7+WzX7yNfLCUfLGOn5bOQTxBpvuFhPi66rwvSNpXkn93bmeLKI9SsLGxwerqKqdPn+b48WN80zc+N2YEtc9yXjeoi6aCFz432d+2bxLF0qlTJe/92w/w+c/fjCiD1hmTcYnOd7YZny8VuVLtzxgJDUjwbXuBWW4exze84Nnkg4yiyHYvXVv0XMM5hCUHxZBf+B+/yemNgDIZTkBUHY0HMggG3aaCVDswlgt5AN2F0AAbXeBjt3OsBIpwu+oYWL5l9vYgjte9+W/5Hz//U2QKBkUCiG2FMTF/qB5bsmFOqGOZazNLva/n+rirnXP8QlMZ5cJAPNl1A9uv9lPnwMDrd7mGg+/kseeCEPbeYTw47zFGUgd5j0dxn/vej5/66f+Xm794mqpWZPkqk3GNyQ2iAkp5tssJhSmOMAASlXHc+P10Y0r6QQVLkIonPfHLeeITvrx94EqYKzg/OInOjSMkMpBimPOud32I337JS4EBnoZJ3MYDUF5F/Zc24yMJfKXxjzWSFqMtwW9z7fERf/S0p7b1uDv0o8w6Lgc5CzzwmS/cyX/5+V9HyHDOIyonKI13DVjgWgBkfi+In1GjtUasAu/IlCeEbX7zN/47vuX5tRDMDn10UQzdmYySaVu8JndJAhjR8RqAzGRYYFLn/MKv/D4+jAiY8wJAVIjgx6lTp1hbvYYQApsbd/HOd/4RFTECC/X0mc1snqptP32B8ckZCs/5bFDnLGJMshKE0xuOX/u1P0SpJbwqCHPWZJDZfU5Cap3rHLk2WBvJALUoHvSAu/ONL/wa9GHYAEMXBPEpKK74xV/9U3xYwiu7yz4/BaniODflzs361FhXUWgw2vL8r3saj3rU/cHFLKnDEPeYruZOo0Ux0+eRwa/+xh+zXXoChu3Kkg9GVJVt0/vP2wEOsU5fG4VzJdp4gh+TacsTnvx01mbsND8z/7lQ++2iacFZTRSmFvSuhJWXHszdp/oTeN2b3sZLX/bX1C4HpalxqMzhXI1IFn2MxnGeSTmfvna+4y/pHJIdEGb2N5X+A4VFU1NNTnL9dUs8+9lfE2NpCyaz7GG/7koiep7gh08pYk1Hm50kph6jpQWaJxXcdUr4v//9ryJmCYvq+BJTgEK8Iii/8HXOG/CYPzsy7XC+xtaJpFUCSgfqesj3fOfTYlaqAnyV0FoNEkHRC83i646F7OF/7Jan0Hx17Tyr6wU33PM+/Luf+R+IDDm9MWEwXKF2YWYsoq7yVwkAMq+Hog8RA4wKkTH3f8gD+YrHPzxl1cVMKlFxDsQAVTHzzJu2zfsGQDSgJTaEjPkdOQ6PYHA4xKfXgwEsQWJEjGDwIqgQc2N9UozndN4tWr9v5df0255uKD4ZaBEU8Xzq83fyD/7xT/Erv/gfcB6Wh2BMjrMVWgLZIE+2YaQTrsqSfJChtKauqmkrpYuRnbALCeyF7QALkNv9ns/NVZwDog5xJsg5vid2KIkLShcxlVob+JpnPIvf+e2Xkw9WsLVGmRWCwKQcozJPVoywjvOf/1fFWSdej8TvkQwLCYKTgPLb+JAtUHaHJfW3iYD4lOkTM35CGOJlBL4gqMZQSAm1SqF8Ma3hC+GIjr/giftCBMY9njxuZZ2ykMMsTnIcBSIZVkAxgKBiCYBMI+DNPJkavkKQgGOQIuyCBIcQD0cKJDTzqwEPZgjhLsI6CF1QabaHgIRd0JIAIWR4CqwMCWFK/ny2cZ+PeNaTkuHS3bAMqF3FcPXuvOIvPsizn/UIhkVI5YdxL96ZBXIx9EC61+b6ZnLDm/TZ2C4vy+G5z30+f/wnr2c8zji9WWIG2UzkNsw5pQ2pa8BhxeDwrXPjKA48TL579W8E8i1LOJbxuA6ARRs08gmlVGmOh5RVgzexzEVpgmRYPCIVzkpLMn4YUgTUjlnUXFR8tZkSx6+7gdOfv52gBzgJ1GFALY6sCZydl/0qSDAEBC8BJ4YgDg8IFU4yPKlVs8yv1Q5ot4+A3yU777g2tTDstaOS/BCMfVNi4UTwFLgwxIvBBovC4sRgfEFAEgjiCcql+9LTzHDCefsvKkgCPWIIhZCyGpLbm+ucjc0NlodDQghk+ZB/+S9/jGKQyh3Mud+3XGplMtftQhtDVTtUpjEZ/N4fvAxl1qh9Fss7ZIF/sFsgWxrQ6cLPCsd2XaIwDJeWGI9LlNFUk03+7M/+gu/49qeRtyQSGrzFeY82esofeoHzeL+tiWWXvSsSyRo8cO311zEYLnPilMUzIkjS2+0zTAB11DhToO1KtR9bBMKnOZTAOAHlY1tg0XlnH45/0XS2a6gKvPdordtnCaCc2z9FjlLSpnkChCulqD0oJB2EDtMOOhnmhtOntvjc52/j61/wfXzmMycIAbbL1N5NckLwYECyuLDyQYYPlrIaRxK8Frn37USW7sIRHyPDKp2bnztHYPF5x2ee43ExMYP9HapzVgee/XFRp1LYaUo++9nPpqomTCZjrLUp80PFiC+aXnrppZcjLSHqQ2stVVljdM5fvPovGQyawEqXGa+7eV0sEF2dxdqcJe25xz0M97jH3fChZliYVJeu5j6nQ8fW2hVHfJhDbG1clvWVdeESsbcHPeCB+Nq2JU0hyEWxX2LJUCcDcrfoUuhS/Jk5yKYBai7vub2WkKWzmVmranbl7jQYezmrPvLekxsTuxJ6Swiepz71UUDKQDgMC2Q+R2FGR6dDR/d0Ywte+epXoHOD8xVIiZIJSsaX/YxUiET+Nutjn5DaWbJBwSc//VlOn4pXb11DdxCSE+1niEgP9Ol30mTX1+Bxj3kMti4ZDXKsq6dg9VHeexptpWZ9NZkbvBBCJOZtAJCmhdHZByEORHcwrhgQBNVG4EAng8VA0Hg0JhvhnUFUwYte9KP895//Q/I88SY1qZ4SN6+77roromyi0HoRO5ZfYGQ1m55jJ9dvSIMXOk18pudwkSZ2m3Z3jud+E5vayCLCoOlvTiR1vNe9hjzykQ8neEueZ1TVBOtiW1djTCSC6qWXXno5omKMSU6xwhjD9vY2n/rUp9nYaHbGDv3/QvT+YuyBsqtjZlNKrA+Wqo7JPE99yldhq01EOeZp6WRXsONogiAhBEKQyJXkPePx+JA7n7MTQFIC10Me+mB8cC0A4r2PLRYveOp1gQ+/APTo4HAtoHaIMuO6JDqdjiR7Pxnfue8jbjp2Mnq61XQqxKMuJxSFIeBQCh75qIdTJdLtg4c+mhIds3i8U6S3qku00gjwjne8j8lkkvTCwU9irTViNJvbE3SeMaksJs9ZO3YNr3zVGxO/X7oZrRKtgSdQHwJYISYd1LZu5843ffM3MhoV5HlGXU6QVO5yVCH4SKZezWzzahfiqS5+oc5lgs4DIOFKamsQOhGm0EXXI6I9KJawTnFmw4Ie8YpXvYFv+45/xRduDmyO459sbG3HNMlrroldIOoK6zxu5jl06kvboFKThiS7HjsbWs33i7rw2/f4WIN4judAv4EBWOvS+hKcd+2aAPju7/kOioEmywXvLSE4rK3w3vdZIL300ssRtv6nMWKtNcPhEOccmxvb/PGfvYrxhD32uYu19+yVc+wxTURIFFkWM7Cf//VPZ3VtSDnZmroCYbpXSzezdEdmyMW+/sMPgIhIdHi8JOdnZ2e+wy4PeMB9EAmo4FESEH8xbNwGPXB4abhTdvLIzLw9zIIzBw5+zF/XAiBkenSzoXvb8eziMZliMt4iWItRwt/77u8iyw5L9kdHp83zunQyQJTR7bR4zev+itHyKrVzKJ0hvkDcAPHDy34mFFivQDRKG0TFc1UHxpXjd3/vD6lKyLMmKN7t/WMP3Rz2Hh71iOu46cZ7UFfbaBNiaXlbfnv0VtDx48d3AB5a610BkK2trSkAck5mxIIdTa6kXW4GbZcWGNk4M2a0tMrS8jp3ndzm9GbN7Xdu8O3f+QO8+rXv5vRWYLi0TOkUDsW4CphsRJYvocS09b6zKWJ+ro3hXoeZAWWkc744JST+Ag76zWyHEorPQseSUR720PuCr9k8c5q80IyGA0IIVFW1KxLZSy+99HIUxLnAoBhS1zXWWoqiQGWG17z6rxgMWtNtDqu4mOUvewQS0veNJ2MadjAFHF+HRz3yIeQFbYRt1sCUnd+zMDPk6rdIRXQkF1QGpRR1XR9a8EPmxys5ckrDPe91D7JME4KbAXYuHorgd87txkY8RKShuzyws0xlv/j+eln8WFvwKDAaZJTVFtoE1taWeOzj7ovRqYeWC4d3EaWspsrViBjGZcmpMzUfeP/fYa3DWcGYIq0108m8v4xnVGor7NB5wda4ZjBYJojhzMY2p86M+eSnT0WI0gEpYOlxKN1k6x/8fM5NNoM0Pu7xj2ZSbjLIdeycQ0j71NHzN+52t3UyE4MY3i3WUw3WEULgjjvuOHcAJPK8TT/kisoAaZRyF3VvoPUgrBxb5647T2Md5MUKokY4BpzaqPgvP/eLfM/3/CM++OHPYV1S8z5r0c5JHTfTsMPImmWT7kaMdh7MHKnzz2zaYZNVch7nCPOoNp3tXM6yHxClExFYeFwFYhKdvQ++XWzR+ItI/Y033oOqHmOMwnubiPWEcAQR2V566aWXqfEgiGicc5GBXQl5PuD22+7krrvCzm2i7cRwcZTnroEEmdZ954nI3IdY+z0p4TnPeSqDXBLrPDMgiJopEO+27Tl6BmjXFlRKYevD7gDPczDEfXxtDfJC77ini4sisNMG7U6lyEgwnZsyzaw4iPMMg+6u+J6fs3vPAjoeSbQjsDNfJqDwOFezNCogVDzxKx+P1o2/BUbL4bn8haAXKMlilzAp+Mu/fDOnz4wh5ARvCF5iJrmqD+SA2JKcIDgbAIULQsBQDJZZXbuO177uTdFOF0A0tbUIghZ9WBRs+0+tIDh41jOfzvIow/tqpnNRd5zaGP9VLt0eJF28uglUd4FsEeGWW26JlLzznB5XvwLyncni29fPnDnDyrG1qLBFgc4pq8DKseuofcbNt27woh/5cb7zu3+YL95Skw2gdlDWkGVT+pmwYHM9t4nO3l1Y2k3xHM8hGYDncSZ0O7rskt4ofvbofvdVIHVldywq72LNmRZYW4WnPPWrWTs2BF9z+vRJALIsuwKBwl566aWXi+f8icQa3aIoEh9ISW0925Oal7381VPGd2HGmAuXKRRuvUMrTVVXmNSZaDSAJ33VI9jePpW6n5ytHOEqtaPE72KQTJs7xoDYFAxxzk1/9gc9+6ZXK7vZXOmNeQHGRL63dt8O6iKNbRdBUFMPZd6hVC5NNJdI8C07Mysu41n8Luz6fkrmP2Mbdu+xB0H2hOGSW7K9dYbl5QGTcpNnPuvpmKYR2GFRKYu4mVq9IGhVEIAih798zRtQUrA8WodgsHVY7CNcpsNLLIcwxuCtZW1tjbIs2d7eZlCMuPOOE7z5TW/DWlrAL3jdzlvrDwMRi1BNyparCIEHf9kxrrv+OOVkC4WbyVKc6rajsfacg9pGHi/pqOvGV+viHDMcINbafafoaw15nrcfaozBOXeFACjdiJJDYRNrbjyyTONcjfUelMGGQEBhfQAGKLWKyDq33LLJC1/4ffzQD/4sb3rzJwkBfIhJUjYZcc6DdRFhbPhlJ5NqOif9WYCPRcGvDmgTcKmZ1rmUsHQ+81zP7Wa2M7Vj/lraf3cYz3248klAs8K0i8roSAZltKGh1bUOvv/7nocWT23HFIMMrWOtmcoMvfTSSy9HFgJJNoL3kVvK5AXOBTIz4Nd/4yVsT+KWUpVxv6rqirIq5xypS+iIJBuoBbfTvuc8fM/3fAfWTdBGReMqxHZ6DUnmXp30whEhEO8+vxAEax11nZyJQxEA6DrnO7v2KA2TSUAFuPvdriMEnzI7s4tCYm6MoaoqRBRG50zGJcbkjJaXcN5PI7WtnReBj2hfBS6shPnCD+cqUB5ry07nQo93NQuJXdN6vYoSgC9IIpdeF/hQEfxIc1PweFty73vdjUc88p4gTTkGOHeIOioFZu37tnw/3ktVwkf/7jPkeonTp7bJzKBdc/6AjkaRi/PkRjPe2iA3iiIzBO8pigG33XY7b3/bp6nKeI95llPb1Eb3MJSwB0VeFCAw3ppgkuPx977n21gaZoRQIwlNawicjTFUzhKuggQH7/0MztDwe1gbO3ZpzUxWflfvLrJB2nV4Ll1g5tGTK8f6mkezOx1ZZlrRzmYueBEIioAhy1c5dcbjwhLr6/fkIx+/mZ/4v/4j/+Sf/Cy/8Zuv48QJ2B6ntDUFqEgFFVTkBxkMcpx3TMZlNKKAcjJhc2MjXVpM1YoRgHSIB1fj6/HMxiIxOWtmo7HW47zDuRBTeIOkjfPiLF5rY29s5z0+hPYI7ZHQtaTSAw7nHT7Yq4gDYxZQ6q6E3MTo0WBoGAwNxmg8Ka36okWQeumll16uVCcAupHlgMQ29CHn81/YJAA6i6BCnuUUeR4N50vlQXUyTZzzgLTt8Zqy0dEQnvrUJ5IZRaYB7wjB4WyN955M6c7fHF0QpAmEGRMDZKdPn26j11eKzZjlsdYky3QK5ERjeTcyvXOznzx5PmgBoqWlFaz1nD69QVX5jmUReeACkWvOt80tG+DGHMhZmTytz4za+Q4J/AIenD7jYw99s1OZCZ5r11fZOHOCb/vWF5JnkcegyCBcpPl30exfWczxUpYePLzkJS9HoSnLmiLLsHVJZlQsQfAHdAQXa0aCRYUK5SsIFRIsIdQEYvfG17/+9RQFlAl3NyqjqmvMIWhiEBof3QdWlgdR59bwmEc/ks2tk2gFRoTgfOuT1bW9cn32BbhDS7LdZGOKoPXZ919gx9+12qqqqn0j9EpduQ9zWgvlZ4AQFUJKb20orWd7fXmJKqr2iqJYAxlw8lTJ1jZ4P+AjH/sCv/Xbf8zzn/+9/If/+L943wdup6wi07cDyipi+ZOyRmlhMCooqzG1LSmGOcurwwi+GCKTb5gQ/BjCBKSCzKPyBk1Xux7a5CiVRcZlMSAa0ImXBEKoCFQgNSh7bodYtM5QqkCpApF87ojf5xEq6yhri/OgVLyWqyMCcHaQUAJce/w4EuI8qqrYCvdcAMZeeumll6vXCQggNu2rEDAENK959euSMwCizEznMSUXwaGaT16cM2NipEihVTKmvMfVDgU84mH3Ym1tme3xBtZNUCraQqFJ5XR+jhPkaIpSKpV8CmfObJKZxK1xWP3hOcMkUzHqfuO9bwDvUq29R18kB2gwGOCcoyxL0ArrAiYfsXps0E5LBzgyHEU6N4ea2nIHcPbAuI4UwUrnKFVEkMZDVdeLgY+wF3dErwYj91/0Nza3zlDVY57wlY8DYrZ943M1LboPj/27054tckUI8NI/fSlKAkuFYZQrcm3BbWMkoBUHchggE0emajJVY3RJpkq0LtG6RitL8BXvePtbqetp2VEM3B4W5aV2/OSd44Z7DnjsYx6JBIuogCSaBxGNtTa2/71KAJBFYIbWuuXuOhsA0kjbLpdUArPfByQSJ0WDxFw56MeiFnWzQMiuuyIRzCAoUBqlC0w2pBgskRVLeDImE6h9xjvf9QH+1b/6ab7ma7+bH/2xn+ODH/4SOo+faIq4iZUOTDHEZHEDmVSOSVUDuuUeEZODzhKIEfCBOYLV3c+ROGn27BFEqZQiBQGLDw7nK5y3WFcS8ATsLudpP7sQolHTHM6Bc032hyYzBXk2SIZkjN5V9ZVfArPrmmiVURznhz/iwZTVGOfqSISqdQ+A9NJLL710911ibXZA48Xwmr98Pc7C1iTuv3VdU9XVZTJAFYImhGavlVjeaISqrhCBZ33t05BQow0o8REA8ZZdiLrm9uarCBTZlddLRbLDZJw2RubhsL3nyoHn6zLaUo7oeD7+8Y8lBNdmrl4MWzeWMXiUitFwW3uCVwyKJe68y0YOnFTfP60+nhZpu4M8Uok3Ws28vl16tMkh5PR8H+eyZubKrwJoEe5/n5u4z00r2Cq5JsDW5plEvn9YbMhdSp0C3H7LacbbZ6gmW1i7yclTX6DIJuS6BHcG3BnkIM5+E3GbiN0Elw6/iQ9nCP4MwW9h3RitAx94/yenhJoBtKhUinmw81okPXoVKCfbCJ680CjgH//gP8R7i6stmYqAh0JaQPpq8T/mMzia+9pPhUEXt9jc3GxfN865fZcoxMiHmkFlrhh0qWl31px3LGO/B2gCDoeta7TWDItYFzrZHqONYmm4AnjuOnWSIhPyfJm3v+MDvPXt/5L19TUe/KD7ctNNd+fpT38Kj3jkvSljZQtFASYvIirlEmCRdmYjoLUAGSJxE9rLoGhYzBe/R+Lft2ziOo1dY5w1TBa7HQYXIndJCA0QNv08EMo6Gg8iCRhJNVuZUeRZztUhfndEVsV68Sc96av4oz95GToftnXiEZXtwyC99NJLD35MQ8IJKAia22+7k/e9/1Ye9Yh7EAhpzxACUNeBPLtUdsYUqPA+2kLeOZTWaFG4EEthvvO7vpHX/NWb2NzYZlJPyHTs9JWbAoLCWnfExnGnWGsJPtZkixLOnIHjq+nt+nDfUQgulRUrHvPYL0cplcp68ovj/ARLWW1jTI7WwwSGGDa3av7o/7yc9aUCwcYsYCInSAQII6igfXzuEhRBLveZlsdsvHWaIldkRsgN/MN/8O2EYGY7xbDg373ma6HWRQBJbhQ//m9+lBAgy+OM29w6xfLy8iG5A7UnCGPrwD3vtcZv/+avocyAQVFwemPCaDRgc3tClgtBQiTxFC7rGQTt4/x1KnKxBPF4FVoSWh08dTVheRD9sc0zE3QuDIcZ5rCAeq1/FyFRI7A9ER77mJs4vr7G7XeMGQ6Xwbu2dE8pRVlaMpVd0esnpCqNqd9KClj4c0jgELz3MQOvAUDOheCpAUBEhOCn/OxXRjaIQNCz4Ma+u5QEfKjJcoP3ljNbYwZZzrHj6zjn2NrcwnnLcLTGaDRge2uD2gWKwRKnTlW8/V0f4p3v/gC/+/svZXl5mQc/5Mt42IMfwt3vfj033usG7nvfm7j2mkjE1R1MB1gbAQWjErAhi0EQtaBXu/fTVlo6M4CJkYWWrX16NmcxUCQBHyG9P6TU1iYjpMvzqTVoPVUaZR3rGa/c/dCfVS/pZLA84pEPIC+yaDiLUCdlpPrASC+99HLEJUiKbst0rwoolBnwmte8lsc/9nuZjCtGw7ghVZUnzy+F8pwFs631GDPl1Ypt8wLGxBLO666HG2+8Fx/4wAcRD0oC2gjWWrwDpfQe4MrVL1prJAjOluTGkGeBT3zi8zzhsfc+NBv/TEvXubmgtcLjAMX11w24/vrrueX2bYRhBMUuxMQVT1FklGUZbeVkQ5t8QDXe5rV/+SZ8NUZhW2PLIzhRhGTsSwB1gO10KudZGQ3Z2t5gUBgycRw/vsQ/+IFvx+SdqrLQAx+7zYFFHTkaIlSl4MsfdT+0RD6ioAPLS8sE7xB1WNBDtUB3RqLQLBO8g3vfuIb1UNewfm3kqrjWD3bAKJfz3JRiEcApTRBwata3yQRUGKVUJ1heHbT3FxJb1YHq87SubF2SFxnWVWhdMBgYKgePe9xjeeWr3xyz2byPGWxGaDk+r/CtKO7HUyAj+rcxTWo/CRzRB1MtcXkLgNTnWF92ZRJaSssGrNpFG1Kr18UpXY2FFpJWNxn4UKOMZpgX2Mpx58kTiAiDbMCwGDGZTDh9ZoyzAWMKRAoCBpGY15Hl62xu1bz7PZ/hQx+4GcRjbYVIQHDc/R7X85AHPZCHPuzB3Pemm7jmmmsoBhkSPNceLxBFW3+8CAQJIZakWBsP51xcAC5QO43zQlVVVFXFZDKhqirqusY5x8c//vE9nl6gGm+jjTAYDFhdXeb48eOsra0wGo0wmWY4LMgyzXA4pBgojJlep7mqm6BMFaMILC/D6uoypzYceE/kClMcnhTGXnrppZeD2Ia7beibfVYlBzrjHe98F5X9XgaDAh8qCGoGSL+0ulta51hJ7GIi4mLrxAC2hid99RP427/9W0yW4VyNMRm1tdR1oCj00R5aiWVD3kViOoLjve99bwRADoUFGM52A3FGJEfj4Y94GLd+6T1459FGLjiBU2nQBpy3BCeI1hiTk+cDtrdKMp3hwxSAa8CPgE6z1J1DwO7SPMFJnaG1pq7LSA5pLM5Og2dB5rEPfyTBwL1BEL1QDz38EQ9lMIgEnHke/RLrKow2VGVJXgwPLwiSgqNaR//DmHgEYolgEIeRfKrr059drvMU/YxBWz+v/9PvahvIjeBKjy7itTqfSuMOGoRKekklPaS14IJFxFBN4AXf8A286U3vwXUAAhGhqutIEnqFJ6DHewoz/B/xOeyPA6QLpHQ5QEzs1x72TCMRuhluoTVgVIgprBLi4havO0aF6yyWKz/93zmHaAg+UFUBrTNWV1cIIdYrl2VNCIGiGOKNITiPrQPWQZblOBdiSpLOCMFRB8FZB17IMk3tKr54yxluueXdvP5N74qpctKk9zgyHVLtsWoPaXk5QnsOIcRWg75BAgM+KKzPpszeaYxiaUb8OXZx2Q1DtWQ64F0VW/7JNL0qBJdaAVqWlodkWYa1FXme8/CHP5TnPe95PPlJX9bVQQu3RNmxYXKINs8ZPJmZlnqppCqEmIXjA6yvrbC9tU0VDF6B6Ig6qr4KppdeejlyEggyTQInhNQtTOFF0F6jdM7Nt93Fu9/zcb7qiQ/ClsKgiMnHVenJi0uxB3Tq8FMKpXPRkM8yjfdT3qs8h697zpN48Yt/HVsL2+MJxTBHa4Me6Lks2DBjYF/xJJC7lA1L8LFEIyjqukbnBc4GnApMQs1HPvIxJiUMBofJkp4zRwOtjYNILOEN8LAHfxlvftP72K4swdkd9y+pTGWaaq/mrNzZL9re3mY4HJKJpq4dHkU9KcmyjGK4jKurtgTHI4k5NgIgMYAm0yyCy30mrofxZMKwyMF7smKNtdWVSFMns07m7PqyPQjSrKEF66fpQPkD3/+9IDAooConZIXBaIOtbWx/etDPbkeJUyTlbeZ/NYF8ANZbFArBMZlskRcGI1lMGT/rPVwigE9UCyCkXSf6POk7Y2J9LNcH0IWKdAdGDh74YBbNUVmBd1X0I13kpFpegi9/9HUsLedsbHh0Ak89sYSzKAbUpb2i/Y8GAGlAkOY1bRTFnl1gfOfvY5nmZDKZAiCTyWTfNTTBwdIop67G5IM10Bm+LtMkSXlGzUQPzWsp4hPMga7eWY6PMPObxU88zM0/HVNdSVkwAWxZTdW6xLIPa6t03xAIKKNwwYGKiF1Dq+FDSMSkqc5Y5Wm4Uls+6DDWe+o6XOAT0El57ZIsJnslkWmc9yB5avmdmsT5jlI0sFF6KIk/TOCNb/s4b3zbx1HBcve7H+O664/z1Cd9Nc/42qdy/bVJGcmUMs7aMmaM6MhLUk4mFMWwc92HwwgMaYTatLgEfljvMErz5Y94KJ/9zLuwQVBFaJHwXnrppZcjaf9LN3ASneaGXDwITJxi7dob+IX/9RKe8MR/jzLJqAl2Skp3IXuA7A5+RAMp7fMdM6XJdm2SUNZX4Klf/Tje8tfvpdY5zgr5cInN7QlaG4SAF9tpr6tS6cKVP35Nh5v2Mc6l9BsdgzyZGaHE4sOYd7zrPWSDKW/YYfDfFk4FFN4HlFZYHwhOeNJXP4Ff+sUXY2SIdyWiG9Kz1GUhRLtWwrREJTpaTeZwYz/FTGOTDYlc8A6CinakMbFdpbWIKJx0AMPkpKkZ8taUQXW5zwRq78hzjfMWQeGsx1oXb1e6z9XPOvzSZ79Og2SBunasLA3ZOH2K4UAxGMDa6gqPePg1rUYaFKadqSYrdgUgL+viCUQiwizWiEyqGk9Onml8gJ/9f36ZH/yhH2BtPcfHGYKIRssIV3v0vtLAF+h38bsCSOes/9t5qnZ8kzD7jE2ezYEyB+x/tPpWodQAAmRGAzUKjQuKZz3rq/jzP3sLiozagc5yCBOqyl7xVWlKqdjVJo+a3E48w8EQV5Xk+VLswtYpq529YY/3MWnBGMN4PJ5+7smTJ9kvD4hScO116ywNCpxzjMdjgvUYYxJeMJ/yJtPX5KCVYLcH3vkpscZw23uLDef0/YH59rZm5xFyPMUFHdPPU+d19uSdY5fvCcOFh5UhN9++ycc/eSu/8Mu/xbd++z/iJ3/613nnu29hYxOsi7PEmCFGZ4Bme2uLYjAgWDslKzlU4lvC2mYgJcRCq+FwQKYylDKE4AihppdeeunlqIqXBO23e6ikDNKo+wdLS9x6213cdWocc0dVY3Rb2sLsgzTA0jU999lfy+bGaZRSDIdLbG5sk+cDgkx5TXz7/6s060922nQKRW0dWmdU1pHnA7TJGW9z4JwQYc6VCbLzDU1ZuFGQZXDdtQOOrS2Rac+wMG08rG3rKn72TDdmphY7wXvYj2GHhdhtjBsW2JeX++zx4vfhELoe8Fj0VJwnzwuyLGNra4PhKEdwSLB8zdOf0q4R1YBmzTMOhyRrxpOifHE+FnkRAbEAJ0/Cq1/7Fn70X/8UJhMmdc3mpKQohmxubu8P/Nit/8LZfn8uxxzg0T3OovkPwcNvQJDu0mzAxdiy+3nPeyabm6dwLjbrgFgikmXZVbKKupo8bbipGmGm3faibMUOAt+l/VAbGxvtw9qPGGOobRk7fGSxJtA512u4XvZUInXt2NicULtIRvv6v3oL//gfvYif+In/wmc/d4ayhK0xBAzWCqOlY7FVnNIcegr5OVlbW4utqC5iG71eeumll6vGh55TiT51Ddvc3uKNb/wEbZaqNtMWaAcsWsNjHvNlLI+W8N5S1yXOVhiziBzwagOw9vGe1Hmg6SyolOLk6fGs7XpwM253JEbAZJra1u07iwy+7uuew8rqkHKy2QJc7dENWnVeb+aBBN8xyHtA4MhbwEpRliVaC0oLIoGqmnD69Gm+8Ru/4Uow4dMSCjhf4/FYB+Mx/MmfvAZthnzg/R/jA+//IkWWMRosA4rl5WV88OcEThw+AOLKkPvd7zgPf8TDGC0NGE+22N7exBjD9vb2lQ99JJqObvlL8++dAM/u+lZEZrrAqJMnT+67BMZaKIoCa+0MAHIunWR6OZpidMHy8hpGD9kaO7J8mbVjd+c9f/Mh/tE/+mf87H/6VQYDmJRQ1zraS5Ih6sppH9Oso2PHjs3wslwxraJ76aWXXi4jCKKSgzgejzl27BjWWv7n//gFrG98ZhUZJA8DACKQG/i65z0XkymqasJwWODqMpZEnCdwcMWBIIv4DERmyPfqyuF9YDwu41heAc/BOdeOo1Lw9c9/LqdO3tlJ+2hKXNQux4IJ3oMfvRADx7asYstVAVtNWF4e8YD738SN98pafAEOubmrNVrpWKqvYbgEb3jjm3EWlpaP82M/9n9zegM2ty21j9TDXc6FXs5HOoThu5RyNljrt33bN1FVY4oiY3k0QImQaXNVZSI2HWFCAjqKQYcEVWbJeed9M4Ctra3p07v11lv3HaE2Bu5xj3uQ5znOOaqqIoRwThkkvRxBCdGIPbMxpraQZ0vU1uBDDrLE5ib85V++jX/2z/9fqgpECz5ETpWqvvIMiLW1tZaAdn7x9dJLL7300jHh2rbtHu/hs5/9fNu+3daByD5+sFHARoNbB9/5nd+KwpEbIc9Mm1K7WMv7qwgECTO+fXdMnLdorXAudh0o64qAYmtc4Q/NFq52DqpEo9k7x6AYIBJtD6Xh+DWGYiAUg2kpd8z+kLmyZTUtX5GQ+G4arhCufBLcXi5cf/jAYDAg4HGuBgJKe773e797D21ziMRDsLH0IKS24Qr4zKc3ueWWLyEMkTDi1MkJL/2zN7A0NGgleA/D0YBp44BuGVUPDp7XJrTgFw0s+8hHfBknT95GkQvDQcbp06fOqUvK4QY9FmeADPbBst19/wwAci4ZID7A0tKwTVdtOoHEvrr9ZO5ljwnshcwULC2voXTO9thSW8VgcIy8WMOYZf7qr97G//WTP4cxsLUNtQeTK6y7MiyI3TJAeumll1562V2yzFBVFSLC8uoaf/3W9xFSacKC/poHqOPhhntqBgNDoGYy2SQz0iFZl8gVNuNsX+k9CBfYdiESvDaGt3MOrRS1r9FaENGIaD79qc9jzAFf/lyGhixwJuI+HclHdRqyQQ4/8sM/SDneTJkhfgF6ond+mMw7eb1tfNTF2ggQhuAYDHO08txx+y085cmPJLjpOjqUWGkzzduW5EJVxqz/n//5X2RcWjwZ4wlcc82NvPjFf8AnP7VFbaH2IZbALEJU+vVxEYCQ1N2kiuNx/Bp4+tOfwnj7FFsbpygy3TbruBoAkBa4UCrpbE9RZPsqpWp8sW5JkDpz5sy+L8L7iLaE4BAJcwAI/UTuZVepvWe7LDl1ZhOPsLq2DpJx4vQ2WxOHDTnX3u0m3vf+j/JvfuKXMHmMwnhi6u2VZEKura21NdC99NJLL73MgQKdKHlj4GitEeLxB3/4J2yPEw3koUihiNkpIuAtPO/rnk3wFRJsaicfIhN9OEpjqWaej6RUB+sdKjO4ILzpzW89NHv3Xg6mNoaqjo6CCFSVJ8vhW7/lSRxfXwJVgzi8RHsk9mfpHCIx00cWRbp7OfI+q0jKmHdoA0p7Hv8Vj2Y4ijRHO+emmvq3hwj89T7O/dHSMtslvPd972dleZ2qdKwsH+eWW0+zvQ2/9dt/iA9gjICYBTrA73L0srv22n0+mIQyFzm86J/+AKNhhnUl62urh6WC9IIBkO65+XcIYd8krw1OMcMBcurUqX1zeBgDd7vbdSwtLVHXNVtbGzhnr4oUm14urWRZxrFjx9BG2NzexjmHC5HfbnllnUkV2Nqqsd7w1299J5//wiauiTApuRJWaItQDodDlFJt2lWfBdJLL730srdx4lG4EMHyD3/0Y9x1sozNB/ThuD6tIgeIMfCDP/TNjAYZw1FGVW4tbK0aDdarxahvepOo2XsMsTzEKBAJiArUdY0xBuc8737X+y6o+d4lcSJaj272xzzLKasS8AwLRVVC8PC4xz8CsIBFzXQkUrs4dWf53l6OpP1LiN06Nk+fwKjAT/zbHyPTcY4tLpM6PPPGu8b5NvigEOBTn/giLijObGxhsiHbY8vS0jqihrztbe/lQx++DQ+U1i64l35NXCydJihUogfxAe57nxXWjy+T6cDGmVNoudKf9Sw7zjQbxCMCWss+dPBUZrrA3HbbbS16dJYlgPdwzTXXsLW1gVJT9tWeBLWXs+MDnvF4G6WEYpBRuQo05MMB47JCmxyPpqo9Ohvwohf9CEjThf4KABA66VmjUT7TGanPBOmll1562V3yvGBra4vM5OTFkIDmr17/ZpxnpjHkzn3l8gDMSqlI5ifRYckMPPABNxF8xagoojFGaMtC2usTrkIizJ2kn865mYhcQKF0gckG3HliOlaR52U2End5AgRqL5u6hXhyU7Q/FwXUFn7sx34YrSxGx3H3tsQohQpgrYeGA6TJDhHf6QrTd7HoBYJ3iAh5oVESyEzgnjcsIcTWy3Ee+h3r67Bghw0IHUI0dbcn8DP/8f8BpamcA61wQRBdYLIRZzYr/sPP/Gc8kCf/MrQEwovW5C5ZIH3scO950Cm3rOoIRA+HcK8briP4iqXBcMeedEXuOEphrW0xh5CAaKUU11577Tl91m233Tb93I2NjXO4CFhdW46pqiLtkGgj+994eumFMJMq6gWCKEQbUAbnYTyp+LuPfJFAQ/FzJYI+PQ9IL730crSlqWDZqxOItZaVlTUmk4qtzRKC4fVvfmtMMtjjs7vdRy7hHUBwDAY5jdWTafjaZz6NarLFZLwx2wVm7kZnW6RegeM3d1vz4yEBskwjKrHzB0dV1Uxqy+Z2xSc/dWs7B6a129PAwMGShHeczkTqKslmVcCwgOURfMe3fAOTrRPU1RbLSwXe1uR5zvLyaI/9PTknoSdBP/I60Htyo9k4dZJiYHj2c55BkcUZ4l04/I5+CFSVp0nG/uIXTvDFL3wJrXJGy0uUVcVgVDCZTBiXlsFwmdtuP8Vv//Yr8cB4HDuHjscloHAupEYBgvNulzXZy37dKWctgyxm5pQV/PN/9k9YX1vm1MkTV8UtNvv8bMDDIyowHBVnQy5m9puZEpjPfOYz56Sd19cFUfuIvIQe+OilMwXb+ugIegRxBOUICQhp0ppENCEIk0nNG9/w14eIQf7sG8R0se4EQXrppZdejq7MowANAB5jW+PxGJTgCOSDIYOlJT75qc/yyU9tHxoHptmnmoDts575JNZWR5gMVOr4MW94XT3afy46m1Cd5p5FBGtrwLeGZmYKtM74xCc+FY30lBUpInjv28zjy7FHLi6d93NGtk72R5g2iCFmfXznd76Q9dWCTBy+mmDrCXU14fZbbiXL9O42cDjU9Ja9XGYxxpBrxQu+4evQJk4LadolB394gRAlZIWidrHM4vWvfweoAo9hMq6QZOAHSQ0yVMF42/L7f/hSvvDFTYqhgQBaTfka2jJx36+Nc9bFc1mFUedGHZ0beOAD11g/vsxgkF81vFTdTi6RhzTOm5WVlY4+DwvBj+bvAU6dOjX726qqzv7AiS3gsmwWsQ+4aBzIWWofe0CkF6KhOM3+sAkMSa0CRcdcu2CwHt7xznddOQDIXBSrQSz7Fri99NJLL3PqMnGgNtpxOBxG7ggd2a+3xyXlxPGS3/0DticcuGOgdHRyra3QOjoBq2vwuMc/mqXhIDaHbFueXn1R/+Z2/C4NeULweG9jarJWIILJc7Qp+NjHPz0DdIhI6/x0S2IuP6DTtVNns1G6AT4FXHdc+Jmf+b8wUlOVW6wsDfG25n73vw+bm5sLCCyTvRtMb/v2kvynipWVJa6/2zU88EHH45wJAaUS+DGzoA7ZtdvoI2oNdQWveuVfoRiytVkyGBWYXLG5fZJ8kGFyw9Z4wnC0yq233cX/+aOXQYDJJJDniuBBKd2uE601faXABe5PonCuQhF5ZsoKnviVj2NpWFwVFBXdNritTyUeaytWV1fOii93QfYTJ07MIhTdtjD7+RBjDErTdrrwvk9Z6uVsBu9c9E+aI8wYDsErRAxGD/ncZ2/F1lfBvfcgSC+99NLLrpLnOd57REeuDe8U+XDI37zvgxSDaRnNgRph3sVIJ2BtzAR55jO+hu3tzY5TrWZBg06nm6sLDVGp3W/auZuuZ8lIdS5QV46yDnzk7z4KRM64eVvRe58coMsJfMyDH7OOp4hEe8UHgotReQN8xWMeyOpSzrHlAZPtLSR4vvC5z7O6vAyJ/6Wp1Y/d64QgiiC9c3fUpYlYb21v8JSnPCmWvoT4epx0HTv40AFmHm0Uk4kleHjH2z/Hl750BqUHqCaLi5pJvQXG4YJlXE5QpmBYHOPVr3k9b3/bxxgUsZ1jVU391bquO/Zxv07OWX8100Y8Rmsqu40PE4Y5vPCFL6AuJ4SrwD/33rfzpBtYttaytrbWeTZ7AygAd91116z276aE7PXgm24co9EoIvgqXObNq5erBQSRTh1XO3FDSocLBqMHTCY1k/GVuNn1pS+99NJLL/uVra0ttDJtGcLqsXVq6ykrx803723YXBZdG0CUwmiD4Mm04D08+jEPYTjIUSmCe/W3wZVFGzsB1xIlumRwWx8AxW133Mltt23OGLPtnx6S4IB3HWciEo8hSiEaRDwhwCCD3/vd38C7GqMCa2trXLN+HGttHHeZA1aaEpgd5I+9HDm31ToGWYara573/OfGaWYdSoXYVzv4RSrnLG7dZbVqEaNB4M/+9BVoNcA7w/H1a9nY2MBSMlrK8KGkDiV5nlM7UGbI6ZMTfv/3/ijZxbGkrLlB50K/Ni7C3tSAHIUxiMRSw3vdM+eBD3gAwzzbpcvQFbbzzHGAND+vr6/v628bufPOO2cBkJMnT+7rAnSap8vLy6nm0/bR7V72aTQpdjKiT6NjU4RO45wHIgiytVlzJWEJIaRe6Sm1twdCeumll17Opjdj9xDnHEoMznrOnNnkzjtO8Iu/9L+wdtppZDfD5tJeoKNbY6x1IoVfhWc961lHfvwaW7AZH601SinyrMDZwFve8pZocKbSlwYIObgAmm8tkEAcy73eO94+gwDHj8GP/+sf5dpr1jlz+iTb25t4v1uK+SKbp5cjaQGnOf+kJ38VN94z8mD4YJmN6B9eW3FSTciMUJbwkY9+kq2x5cxmSV07RktDQnBkA4XzJUqlTlAevDNkxRIf+MDf8dI/fyPOza61wSA/u+vQS6uxdpshohS1LWkalZd1jQ8xC+Rq8NHnSVBDCCgVy6fW1wezT2kB2tMtbdyRAbK5ubkPA8VHKqcARZGBc9iqjmh+UARMq+iDTHkefIddu5ejbOHGeSBhNnV20SS1CTwwxjCpxocOAInkaGq2O41EsMMmAMSGiuDrtCZMP/699NLLkZQmK2K+PWy3KrIoCrz3TLbHGGOoa8e119wdU4x42zvejXXgvEqtFLsmoY8kgpc6TirRkbV13bbgk4SLfMu3vgDEIlgQN2t/haupvl3NmOB+juxbKdXyGoDC2UCQmAny9re/n+1tn8AGBaTnhY+1AN1H1hnbS3svi+xcdtobIbC0PGIyHiPAM77mYXzLtzwfwhaDIlAYEGIWiDSfnca9aYsbS2L2fz8yd+z0CM92XuBAJbv8fM6qsdvm5vNsxpPf89le4Qbs2dtZ7/Z78WQ6YO0WL3zB83B2bh4qs5Cl9zAVhgzyFbYnnv/9Gy/jzMYWo+Ulrr/+Wr50x10EUXgCdV1jbY0xhqqqMCaHIHinCbLEL//q78SOj8T20tbF5+Vs1foHi1oA9yHEPeZBM2+CIjMF1jsUCqOETOApT34o29t3glQz89OLtMeVsPaUgqAE8QIubj4KMDqwutJ9QnpmIc3PHWstt956q8w807NngCgUBgnxD46tjVDaY4xCMIgYCJqAIjRMYNLNKewBkKMskfBOoYNCvCQQJNbZSohGU57nWGtxrkIUaOMIVFx3/erhmD4yD+ZkkeCsazmJRhTceeIussKjtMWIxlUB1bfC66WXXo4sCBJQHc8yoNKRxDmCdywNhgQbCEEoq4BIQV1r3v+Bz6F0pw9rgHKyDbgENHsCfqfBfFEsaAVkgMZkeQK+oxdjMrj//YYYU6Ezi/MlAUuWZQgG6xXOX9kAeAx8qWj/BQgSuz0E8QQVCMTuDr726PQfQUUgy9UoU/CFW7YwA0VlIeCBGu+2Uu2JtOHNCGU1Y3kpQBA1c7R+Z/qHqIR1tb9QkZydwGCUUZaWIoO///eezA+/6HsYmAl1eZJRoQgu3v94q6QYjlB5jhMIRtrn5eeOhgtNaI701WH+aH6j9nmehU985/vP9dwYQOIzxOcony0o9bJEz0QW20xXvBHrkdAZo9Z+jdrAaIlzWTy1LQlK0HlG7T1aCxI2CfVJHvvom8hNHFNtCgIZIZip49aZEzCdFxcDv5lFFPwMyOj97O+nVWo+Zt8B+UDx6te+FkuFyh13nryNYpCRmSGKIYSMLBtgrcVkGmurSHQsA0q/xka5wk//zG9xZgIqA9EWKNHNZArTSwoh/tNh8dRJFxxhH4pp4LXVX3PL3DkgFAgZmdYogZVleOQj7kGmajZOnyIvDGIMiKa2Hm3yFgQR/MzRBUsOWhyWqirRKseoIoafnWWYBwbZ1KaIwWYzB8r6xLUzbYU7A4BsbW3tY/FMGbEHwxwlHTbWYJht9eU7WSBCn8fUi2o0WxOtm0tTKssS5y1LS0tkmaYsx4gEVlYjWnzw4mchxTC7OTb3I8BkMokLThxaFEb6DJBeeumll71Bkq7NodrMUk/OK1/5utQyMpkiIVCYbIeO3tVlvhhhxNDpFpKs9cZ5/nvf+13YakyeKby3qVxHMCZDRF8VY9MdnwiChDmwSXUiucyAXV+85QQ+gDZQ1RXg0LGl4I6tlRk465BIiOhMMTDxqgK88Bu/lh/4ge8gN44Td93OwCgyrbj7dddz1113sb29RcAxmWy3VtB85tP8s5LAOdbr+13OF2vS7/690gE0pVPOTFBzJv+VXwYkYe8ul015nkigKAogdtf0QQjeYrTjec99BipivdR1pxRE5rPEZsGJS27TsrgErKoigaY2Ch/gox/dpJw4dKZwziIaVpbXOH16IzqdnY5HoZnMySktreb2E9u87Z3v5fY7x0wqj5LYujV4u2e6R58Bsvu864rWGUYPAIOtasrJhMEA/uk//X4GQ81111+Dcy7qpRAohks459r25M08b+d6678fsOeVWiuLhMgR5jUaQeEZJvRDzQN9zF56Uz4zz3eqAE6fPn1OF9T03e3WfPbSyx67B14sXtWQ0oTbkI94QgJFtBacr6mqCQA33HADAmSHwH4MZ93AVWsnnT59mhBCDG7Rd4HppZdeerkQef0b38Cdd9pEnJdIphN/hKiDd66+9VufQ5ZlKAXGKAK+5Ybo9X9MPf6d33kjAHk2mGlv7zsGeMdo4HAFzhS2LNN4xgj52ip80zd+Ld/397+HG+95PZPyDFW5zelTd7C+MmRghGPLI0aDPGXQ6HTuZqFIAvrS0WTWzBzRhhJCykLonPc8fAImQAWFhAzls3M+x5LlAKoGselwBGV3dainjsjVwoGidgcQghCCJKAzZj6FEEtCcpMRQmB5eYm//33fw2gUQUDVsWmdd7t832JA8VLcUwipEi0tO6ViVrbrLNQ//uM/YTwu8U5R1zVFUcx05NhrnEXgxhvvyS233MYv/NKvkuWKM1sVAYOYYueSv0oTiS6Zf5L88IYfI8/zNtvhsY99MPWkxFYVEqKfFYLDqAjcGWOu2HtdXV09p7+75ZZbdq6A/XWBmcrq6vIMweO0BrcHQ3pZMPHEE5QjqCqexcf6YZnGeESEoijY3txAgmNlacRTnvzV2BrUIdGAYaE27qSKpX+eOnUaUK0y6ttE99JLL72cv4zHE9769nelfaMBPRTBH5zNoTrnlWV44APvy2SyzbDIYnGPD9iy6gEQFMPRiN948W+16fVKFQQbSzbE6H27nAc20qIQk8VuCxKdWOdhaRn+4T98Dv/mx/85IiXDQUApS1meYVQovnTbF9CJPy86sovYPbo166rNEJlminQyAmYyaP1Zjvi3TYK7SsDLuZ7jgrMEsQTl8OJSaYxvL70vUVAoZfAObO1ThxdFkWmCcwRqbrhhdTbLKf2glXD5e73MZpyE4NsskLqu2/eICNbDxga87rVvJIiJJP9OUMpw+vQZlpdXdrWXGxAPFdjY2mBpdY03vuGtvP4N72dpaQSkUvJ5/pPQzLxuaVcvu0njY4hI++88z3HOMSnhcY97DFsbp0EcwyIj+GnmR9yfuo0qDuka6xCZNudrrrlm3wiZiHDrrbfuXAUnTpzYz1+3Tt6xY8falK9uf91eetkdBLFt7bCX6WYfIyIC3uJ8jTaKwaAgBMeTn/zViBxSBHj+ohJKLwKnTp5Ba92i45Htu5deeumll3PfO2C0fIzXvu6NkW9C00aXIxByEDvE7HfWNTz3Oc/E1WOcszHClunY1aaPC1HXNWXtOL0BW2MLaMQUuMoiC6K+h80MDw60zhGl4piquNdXdbzuJ3/1/XjVy/+A669ZJlMlywNFpmoUFYVRSAJBGjJRtYMMXjruaMwG8TtKZs4GduwxW8P5HxBazpL5zJRdv/kqZbAM0jj2zaESCai05DFNR6TCZHhn8bbi6U97CplJ3W5DzGqOauygH1Bq360SEAKIjqSm8T4VIvC+936CzY0xRudolZHnA6oy+oBuYQbX7HfUoWRcjSkGI7J8iV/8xRfjA5w+E7CuswTmsEHV8E/2KvSszn0DfGitmUwm7b+LAn7kh/8JS6McW5coCYTg8D5yVU0Br8OqfZnp/tINLF977XH2E19ucIrbbrtt551228IsfrpTJRtCBEBCcIiKGq4HQHrZr6rtgh8tZ3ogLsRyQmE0zk5Q4rjnDXcjN1DbQ9gGZh78kOmGf+bMJkoMIjLT9q+XXnrppZdzV7jOw4c+/BFOnYz7SFVPTZhpCYW6zFc1ZT7TAl/zjK/gbne/lu3xGXywkSS+V/0ATOqK5eUV3vLX72W0ZPBOYauAzvIZh8e343i4WsiGpj2M0LbuNQqKHCZjhxJYX4eXvOS/8rM/81Nkqma8eRfrqwWu2kJw06Ots/fMF6039lF77gAhszlH7PKM1B7Pz5/nMXX+/QInKex6LVw19QszvC2yM9s9AgEhdj5BMFqjJFCWY9bWlvgH/+D70tyZdhmKpVR2lxKYy26ZR6cyWLQSBKGqLFppEPjTP3sZw6U1tsc1qAxlMqrKcuzYcba3JnMPaOezy/OM0fKQsqpxYvjSHRv88q+8guVlYYYiSXwssZojBe5lb1EJmO3+3IgGbrq35qYbbwBXIYmf07oapdi9Ffkh6t46X+ITfSrPdddddxZ9M4uOLCyBufPOO/dzBfHzVQRAmj7v3YvrpZfdVawhoNuDoBN5bmNMaIRA8BUKy3Of80yWR3E+51oOaRbIbM1rY7ScObPZmg5BjnpyaC+99NLLhUlVOsal5RWv+quWTLMx0pTOLr2jPGOJqx0giMng2Bp85RMeixKHwuFcvbtxecREa8PG5jYvf8Wr8SESmzuv9ufdHALzUhlN7aZl38FbAg6FZ2momUxqMg3DHL7y8Q/kFS/7dV74gucwNJ5MVWhKhCoddedoOi40QEgnBB66tpLudE7a69iFEURCKkM+98OLjwGrVBITMx5UsunmV8LcerlaskBCvO9d54cY6rpGS+qWhBBcbPX8iIc/hGuPg6sDwYGtKnwCwBSCmanxvkTcH/PZFXNj4rxDJOBc3bqNOjd44CMfu533vu+D+KCpa49IRl15jM7bQF/TuWi3BawzxR133IEXhfeGLF/hd17yJ2xsdaaIQOw30yXUpWdW2KdkWUaZeIryPI/j6hzOQTWBpz31iWRGIDhyHbPSrLW7cIAcviyQJqO+G1heX1/nXCjAFpbAbGxs7BsAUQpGowHGaLQ+CwAS+rqtXqKab9igCXk8MKkdblScVblNbkBJzfHjy/zIi76L4QDKcRkRy8MMfjT3mXT31uYYW/u2TGxRxKCXXnrppZf9OR9ow2C0wqte8zqcnfJCVWW9wwVTuzkAF8cV3gmChORTOHjMYx/F0jBH60BdTWa75R3V4RNACWIyPvbRT/PZz4wxORS5ZmNjHH19WfSc1eHZNgNkWSxrtbaKJTB4KjtGYVka6DgHLIyGsDyEH/9X38VfvOJ/84xnfBVKSjRjNCVKxigpEYkgCBIdZcS3raKb8piGMNWjUjtK2eWs8ej2ve05zLeEPJ8zHftNdw51doep7QZ5Vc7s9hkZY7B17FaBjyUGwdUMC8MjH/EghJjtoTUMBnnbltP5+mD1aho33xivxLIclwhKaguvfd0bCGKYVI58sIw2GVVpGQwGnDp1iuFweNavGo+3WL92nbwY4pwwnkCWrfJPf/g/4MM8g2RvL5+rdEteuryDIoI2MBzA85/3HIpMGE82MJlmMMyxdUVVVWncDyvXipoBP8C3/x6OigX6bTF4AizuAvPpT3+6RYsg1mt2U05a5AOoKrj3ve/F5uYmZVmitca6qp+BR1yccy0y1xDkGmMSIzbYWhgN17AlBKvJVEE1rsEHhoMCLQ5bb5OZwPd/33czGEJZeYqBJgQLhyqPoktKFl+xFoyJWSDvfd8HMMa09XWDwaCfIL300ksv5+VAK3zQbG1bvnjL7aikZ62FvMguS6Bl90Bk3AMkgNHwtKc+lvH2aQiOwTBjPB6TZX0WSFU7lpZXKWvPO9/1vtjJB/blPB02B7qJmApQGNOWtigsmYngnJJoXA8K+Omf/H5e/tIX8+9++se42/VLlOO7WB6CZowKY3SoMcZDqCE4xHuCsxil2lbPWmsQjQsSudN0hjY5QTTWhVQmo3Ah5pQ0Z5+yNdCaIBpPfE/8fVNeo9vXFh0hSCS9lILMjAjWoMkQr3B115WQvW2lK14RTfWMT1x2PpUFhRAYDAZMJhPyIqOabKNVQKj49m/7WhAYFFlKwPAp4xmytpX3wZZ7NXO6cZxDSmc2Gt78pnfgnSLPBwQPVWXJ85y6duR5TuiWXszdRUjPqCgKxuMxMYMoZoOfOjPmtttP866/uQ0nMfejdi61yU7S0+ftSxofw5hp6X1021XL5XPjPRU33ng3ikyxvXWayWRClmULshQPHwgS5+WUTqDh2RmNRgv2i8X3Yq3d0fFWAXzuc5+TVsmms7Skp35mI8oyGC3BYJjHMhjNPtM8+0quq1m6Ua4YJbFsb29TVRVGFMbk3Hn7CYwuWFteo57UDIuc1ZVlNk7fhXcT8szxwhc+l69//lcgwLBQaAF1wBG00NnKw47fRANYpbRCZ2E8LgGFtRalVFL8vfTSSy+9nI8onVPbGJH85V95aSRDTcHtyAFyOY02tWODCD6mgawswbd/xzeztDzAlhNWlkZsbm7ukR5+dAz0O+64C20GvPkt70Qlx1GbZk895E7yWU2QpuWs31FtoIgtc5/5tQ/l937353nJS36FF7zgmVx7fIki93i/jbfb+GobrRzLw5zMCOX2JpPxBlocAQfiQDw+1NS2pK5LfIg2RgguZRX4GVvMe4v3MdU9to8G0IgoQlCx/akXtDZone9yjtm6dVlRbo+xVSz10Dojy872aK6CGoawF7gTxTkHPpBpQ7CO0TDnmvUlnvrkr2Q4aPMsDu0cb4rM8zwnoNEGyhre+vaPc+r0JgGTCF8bYEOdg65MzQA8OOtRYtBmwMrqOidPbfC/fuU32diMT0bpAVk2ZJJKOdCwcWa7Twi50B1LYGsTvv3bvgmF49rj62RK2pKn6WqVQ+mxi0jL/wEgKmZZrawsLVyPi+ahUmpHtUs7i7e2thY6s20GSGcHHwymyH23HW4vR1vquo51kFozGo0YDoctOJYpzTXH1hnkGVsbZxgUGq0Cd91xM0UeGBWepz75K/iRF30zRjds2YHNrc1DCJ7thEKaOrTbboOyrBHRCJos1zjXw9i99NJLL+frgATRFIMlagd/8Id/PONXScuid5kjV4G4UYlH69iwUSt4/vOezebpUzjndomwHT2pnCUvBpQ1fOjDn+DDHxkz3o6/sy5mePorMktgnvyz4+RKMz/B2kj7XuTw0C+7nh/+p9/KS//8l/g/f/ibfN/f+xYe/5iHctO9rkGFMZunv0Q9Oc0wDywNBBVqvB1DKMm0J88CuQkY48lNoMil5RbRyqLEkmlPZgJFBrlRiI9tcDUaI6Y9dPrP137PY5AbRrlmVBSMBjm5Ubi6QgVY2ATkqmOvTCBIkwUyl5XkXMyGUEpRVRPqesydd93Os5/9VIzax/wJc8dFV1V+cavi9F0+ZV40vpxzsaT7T/7kVZw4uY0Xmel+s+MuAnt2uwohoNOEMMYwmUzwHoJo3vu+D/Da176Tso4qtXIwKEbYusZVlpXVUb8HXhi6BQLLS/DEJzwOb0vuuPM2ioGJAeo5DpCAzAz0Yehi1nS5aQhQRQTnHNdcc03bTnpPj817lFI7SmDaO7/tttu4//3vn1p8qZ1gSGLADiESoa6urnJy4xTe+7ka157y8ShKLHWJtWeReMe1E8/7QFFotrc3UQqcrcnzAqM9la5ZXcl5xtOfzL/519+Ld5FRX1SsS1werSRL8xAxEs/v7WG6RD78ob+DIHg3ZWJeTDLUSy+99NLLWY0XAVt7VleWGW+WnNnc4PM3l9z77gXGcDCdVoKa2jqh4Tkw+AA33nuN0VLBuNRMym2Gg3se7fFDEQgUw4JqY0JpHX/6Jy/np37y26gs5JnBs5ML4fBmzaj927rJ+RgNBB88uVIRIvFx2lx/nfAPv/8bER3Bs+0xvOUtf8dfvOq1fPzjn6SsSsChlcYFj7cTbB2zDaKPEmtuBvkAHzzBe3wIeBebmvoAQQSjiwQUTu11kWkbVm1kzxGsyzMYY1DeYm3J2FryXLctXWVX3oarCQVRgJuOuvgWQAghxPaj4y0yoxA8g4HmsY95wFkegbosK3A6V9VCg1ZpnbKtTaQhTe1+3/+BjzAYrlLvSOTxOz9E5r+n40B7QSmDcw6jMia+wk1qVpaXKIaK//bzv8Szn/MEag0EQ209WVZMPdS+gOCCnRat4NprhPvcdCO33rnJiZOnGBbLhODPTacdxOWH6HWJCKIU3tc4b1lfPxaz2qTLW7Rzjje4xsmTJxcDIF/84he5//3v32kxw8y/4240/fDrr7+ez918IipPrXfCwD0B6tEycrwny2I9Y1mWbfRrMBigNJw+eYLj62s4Z6kmNbkRNk7fyf3uc09+8id/lIc/9J6IxJTY8bhkNCxo0HFvHerQYQjNgsta8KOu4G/+5v1k2ZCtukbnA8ZlGZ+L6+dIL7300su5S2RZGE9iZt3qyjFe8fJX8aIfemEkRFUHbB9LwNY1JlNIUBQZfNu3fQu/9ZI/xVmHtT1H2mAw4MyZM6wWK9TjLd7+jvcA3xZ5XLLufuo5nER85+i4zmUICIB31M6S5TlKwWRi0RiKAqoytkhdHsGzn/lQnv6Uh6IUjMfwpTs3Ka3jC7fczEc+9GH+7mMf5dYv3sqZrU2CLQkiUJeE4BEfPVWDRowm1waUoa63CV5mCBIjCCJtevle91IUAY3DhZo8Cwwy4fixIbaOPBE7DJwUJQrpOQlXuD8QVCw/ktlbRCL4EzM/KqytWF0bUNdjvuM7v4mlEfEZZefoE110hRYWf2ggRhsTyOabpsYCf/W6D7C1WVGM1qmrenrvoTP7ZT9lPbFFa6ZzyrJEFZqlpSHVZJu6dhA8SjT/8Wd/hf/4738QrUC8oq5rnK0YDIuuq9rL+Qy9d6A1kwn8xE/8OD/8L/4tx9ZXueOODVbWBgv02uECQiIWoVoAxNkapRSjkczd6C7LKeEYn/nMZ2RXAARm+we3fximZI+NnrzhhhuQv/1YjPQrtft6TYqjl6t8jYVAVVUtaFYURbsplJNNlpcN21t3kGUG77c4tnYPfuRF/5yv+7onMiymqvnUyTs5vn6M8XiM1po8i2RLhTnkacSpx/unPvkZ8jxn21Zt2pZoRXD9Guill156OR8p8iFnNk9zzcoA5zd42SteyQ/+wxeiufQZIF0OqIW7UPAYE50I5zxaG77ru5/PS373z9CJCPuoy3g8piiGiGiUytjanPCHf/B2vuO7nhhbgsoOTGlf/tuBgyBhL/BjuueXkwnFYIDWmrIcUxQFw0E0v8uyZphnWB/t60xDlrhhhwUcO7ZMAB76kDWe/YyHtp9ZVbC5Advb25w6dZqqqhiPx2xvTxiPx1RVhbMR8PBEkMPZEDuPJLtciUFUoK7cNKNh0VlFnp2qDBRZTqE1g4FnUEBtx2RtnYde4Fxf6dLpSJTGNrTZHz5yZiiFd57BYEBtK8bjDb7phd9AbWGQHSZPmNmOy81cKkvyoqAsa7JCU9fwv3/9t/DkbE8sQTcZb/P+XMr8OYuPV1U1KysreB8oyzIGRjODtbH8bTQc8fq/eguf/f7v5T73GeItDPKMLBPKyZhisAT0QfXzFi2MN7YZrYx44AOuJzcZJze3WV9fZ2tckev8cLtXbQeYKR/I8sqIPG/2/735P4xRbVXCngDIPJll3N8DEnyrzJyDe97znmitqa0FVFvf1csRNVCLItX1+bYHdVmWKKVYP74GboNr7rbKIx/5cL7xG76OR3/5vTE66uC6tvgwITeaa9ZXqG1FkWcolUGAosiviGcgGu48cRe1Cy2QmGVZz5HTSy+99HJBPoiOKfhKUY0td2ye5n3v+xRf+fj7H7Bj3OVI8yitcS5G8h/8kAfx0Y98cUdQ6SjKpK44vr5KdaaiMAVQ8Vu/9Tu88JufCKqmyA97gKApd2o4INij4kPtMMiLwSDxOwSKPIcAdT3BOcdgOAQ83lWt7RSCMB6PMSZ130i+t0/+ptaQZbB0HDg+gnuN0t91pmRaF77D2xBJT7uOxfTY6859ujVrI0CjfMxsQECrQGzX0XGSj4J0SmBEBYL3mMxQjsc84P735drrFOWkSZFRh2D+qoWqK9qpOSDkWZxrH//4l/jc577A0tI9cGKwWIJMgcnFBTV+l1WjMBLJYQuT4esxk3KbPM/QRjBqSFluc/z4dfyLf/mv+fM//Z/kOTgXMDrEtdPLBc5Tz3BlhA9RbzzkIV/Gu//2o3F0mv0pKA5rCYzWGu9jJ5csdbk5fvwYWqcGFHuAH809felLX9odALnzzjtbhCR+mZ/duEVD6imsPFx33TpFbghegejk5IWkaOc2i4vxYGXuM5nmYwk+9lIPIfVSn12iXmIdWn8+v/Nuym1G7JhcWQajAXkeUd7VUcajH/1ovupJj+cZT3soS6MInmkVeT5qC7mBQaaAAvBMygmDYgRovA0oJTgLBwlQNrbOblzgQaCsYhLo5taE7XGN6BxrPTrLqOqarE/h66WXXno5Lyknm6yMlrB2AmrA2rGcN/31O3nC4++PsynFvKufmwincMFOWVf/L/xtcnrH2xMGozU8sLkNT/rKx/HB930II9cf+fE7trLKqRMnWMpWqL3Duwqo+PDffY5HP+am5ECHmT21zQIJ6irgAFCxREXivYTakuWGLOStbZVnmi6vzGiUR5vaVyAarTRaLcbgmhiL6pCPtq+lteCJdpfXs1CFT/jdrvZf59lnaprjMSji2op+QljoEl8d4md9kBk/ZJr9XtclKkzIcuHrn/9c8DAaZJTVFoN86cDn35SVdw68k6YLjAIVAbJ3vuu9XHP8Wk6cKimWImA3zcrqZHyERd+z8/ktLQ05ffo0g8GAoiiwIWbK+doyqS1G59x+x0mWBsJb/voTfPVXPpAsk+gvaNnluy4jgHDWe5z3kQ7ROggqOV5Q24DJhR//N/+Cb/7W76cOmlyP8E61ene2ycPhULxKGby3WGvRRqG1YmVlZbr9Jn6e6VpVnekSsz9uv/323QGQD37wgy2PQ1VVs8zlHYjY+/hVd7t2nY2TJ1g9dm/GY5eMjcTmLc2FGAiaRYj4uU5ACSn1LC1knyaYwiHUVNsnefxjH8l46xTDQrO5eYalpSVOnDjF+rXXYEs7o/D78+xZS6Sx8tbiQmB1eZm73eMe3P366xkOC8py3IJjKysrXHPtcdbX1xmNYlrnYDDAGEVRFBRFTlFEpFHU7DIyeqoamjaG0/pQzaCY9kVveD8OQ3aWr+L1OufRGTjvIqlYYsdWOfz2i9+EMiMGI9gaT8jynNpGLpTIItVLL7300ss5GT8hMFAeX01wdWA0WKOsx/zJy/6Sf/Evv4uUKBgDIVbN2J421GglHSPp/EGQxS9MHaHhqGAythQDw8jA3/vOZ/DiX/1lMtk+0snbKoC3jiIzBF8xHA6wVU3ta372P/8Cf/AH/wUlpi0xcqFGSYbzAa0EH84W5btcDuTcRDgnckuFqGnJjJh8JzDX+TzpNiJQZ5m7u2RwzL/WuMALz8mR2O2s5x9D608rIGcmJ0D2+YiuFBHPlOPEz7hO4uOTURqKXOOqDY6tDvj2b306SsCWJYOiOOCHoXbOn/lrkYyqDqgsEvH+zu/8IbUvyAYFdV0jWs19QtjpZAtzPCcBlZzpqqwZDjICIWUgGWwd545WJnL/TbbwXvOf//PP8/u/9wusGlA6Sz5n4OCI9CLJddxjNKHTIUVmQPbYaaeuLVk2aJ95OGj9JYDOqB2IeATN3a+Fe9+wzh13ek6cGpMNVlKmmAVxCBYVFASDBFnY+eeyYjheUdeWYVHgqm1MVvOgB96XYKcoRrxEx5TaeToXtda84x3v2B0AGY/HbcaH1rr9d9N7V0QIBEQFFIISx/JwgC0rvI/vD6Lw4nfWEF6E7V9CigoArl3QviXhuf7adf7Lf/63rK/GyVaOY7veIImEqA/A73+5+wZ1i2fnaVt5+ZQL2U0O8ik1UyRuBN2Nt5k/Z08DPtwmok6ouUobYHNPLi27yRj++q3v4MTJUwQ9gKDxPq4bbx2qr1/spZdeejkvB0RScCZYwTrBOaHIh7zuDR/kKU9+BEtZZxtJEfHYWUPj8Vx8BqnZaF9dO7JMkecGIQL9mcCznvEkbrv9i0e7QFg8eCE3Clt5aldhnUMrxanTY974po/ztKc9CAWUVUmRZwRcZFcQcwU3FlS7I2nSBQsuj20gF3je3YFWXPVlLy3ZZ8MHIjN3X5UTVlcKGAx50APvQ5ZH0GiKBxzS0iCJ12WtB6XY2oQ3vvk9eFHYkIA7aQb8LEG8sEt6UirBmF0T02foJUboi8Eyjoo77jrNS373pfyTH3oBofkTJ1Ghh86lXK4zOnK+JEiHbl+QBguSxAcTAllm9sSaLjt4AFR1zPxo8pWchac96Qn8zu+9htXla5nU6UoFvHh0COl8KCZpTM4wBSF4TCaE4LjppnvHEpgwD/ZO2bqaWaeBm2++eXcNvbGx0Za+aK1bopGu8xpCoImlDIdDikGeGM5336HO1h/6nDeUhVCU5447bmd1LY5hVUVHvOGdNNmU96c/Fh+17bBaq/j8mhRikQ5vkpqdbA34YUzMkJiPOnTJa67sDTA9C63xzqV7iil6HrAePvLxj7X323B/KKVmWNd76aWXXno5x51fqRh5a/edqFNf8YpXkWcN98E0St8A9eDxlyFy2LQ6V2oaDAD42mc+ja2tzSM/frVz6DzDh4BzsaBIScaZ01u89M9fgbOwteUp8hgtr+uaLNV7+NDvn70ctBepZoAemXFsfezYs7nJZGuTb3zBN7SkKVmWEQ6r/dflMEGhNQyX4XWvfwPj7ZLgYzODReSRFyYJTJJAkEAQz6SekOc5xsRMsD9/6Sv42MdPM55E3e5UxECCHMy5toraKqo6+pe1jXw4Lnakpum95BsnCk9o9p1D4P60fEDB44OlKOD5X/88grdU1aQdjyksIIcKsGtwCefrWJ3iHQ964P1j6Z5n7lp3Xre1ng9/+MM79+3mH01/3Ib7o8n6iI5eorLp1NWsr69R1yVgUJfFwVW7oJdRlpeXAZhMYDSYEjY1xoj0Afg9JctnQY2GaAsiGOJtyvBQ84Zp+n2nbEzSQuqy9l4VAIgHTEAFIUhcOsFHYPqLX9jEucBotIxFE7zC2oDuCfB66aWXXi7I+VBaUjeVDKUMkqKWH/voJ+JbAljvMVrN7F0tWHKhanhHSfRsWW9LGJ/2T2XiD1/2ZQ9ifX3tyI+fc7HzSAhhmiqqwGN4/wf+jje+6YM87WmPiNHKqmaYZVO7Ux0hYs1erjC7MLbuMUrIs4zCjPjqr34womA8toyGaqac6VBdd9eON1DVcGYLPvqRT+J9QBNLH5QyF7z+dwdB4k95nuFw4AL5YMSX7jzBz/23X+Bxj/1yRqNBzFDBI0ER5PKewSdKiGn2j0bHagixIDXf9M3PYlBE3a+VpCKMVJgZQB+wA5plEjN6fORb8cC9773Mwx72YD78kS/G8tFmLjRpN6FJOGj2uoNJB4l8RIK1FvEB5y1GwfXXXzvN0gkNyrP4OTvn+MhHPrI7APLZz342VnClxVrXdctI7ZxriWissyhtWF8vmEy2MWYZGwKhWwPY7Xoh/rI8t0m5TV3HshckXkKT+eE9R75Hzbncf9NxtukZ71wgn0vr8n46zKKmwEi340kIrgN+qCv+AQZv43MUIQSPkzjHHPDSl78MkWikBy3Y2rWRKxFNL7300ksv5yfOBYIHbTJc8Cij8Sg2trZ469v+jqd81UNRQRFcCnaoKXJh1KWsf/Uz16iUwmTSplBfe92IZz37a4+8ASIiuOAJSlBKTzNHdcF4+wwv/o3f5Rlf85+YTEDrDBHwweHqkiwvOFLdRXo5nBJUW0LgUztYldrzeF8zKAqe9cynMGhaGA8j540cPIHNjIO7wzEIjQ8Fv/Hrv8fpU5sMhiv4kLG1uU0xHKVM/wtyY7sPchaAkQBKsb29RaYEyTQrx67jne/5IB//xOcZj8dkWZ6IkQ8CAAFfxy5H0gFzRBxKWbRUPPGrv4qbbloGVeDxBF+jlULw6dkfPJGh95Y8ES/aymMyxdc//zl8/FO/QWkdCp9gjoar6PCUtmmdMZlsM8g0wVZkhSbLp/w7sxjDzvWmlOKjH/2o7DUrmUwmbblLN2ofWhRIEZwFoBjAyuoSiCMEd8kVz9kMkOPHj8/c+3i83bn5DuLYHwsP52q8t2ksE9oqAa2FPFf4MC13aTlCdCx7mWkWlLKGRMLVkfnRBUDUNAzoAwSv8cCJE/C61/0ViKZu0gWVpDKy3m7opZdeerkQsdZiTI7WmqqqklFkcE74tV/9TXwCPkRN9yfvLS5cHuK8mDE73fN8pxXpM57xtCM9dl6mbQy1NjgCHsE6T8CgzIhPfurz/Pbv/CWDAvJMRfpE76OR25fA9HLAwMdCpypEZ16Cp8iEz33+k3zdc58FMi2/F6Vw1h7u2xMoa1hagje9+W0oleNcQLShdvailHBL6iIjXRCkbS3jYwcdrRksLTOe1DhvOH7tDWxs1sCI0maUNmNyEOc6I8hKOpbioYaIGgIDggx493s+EMlO1TxwEFBy+NrLmkTq+NjHfTm22kRRTcckQQMqHB7/TUQwSiV8wpMXZtodCDqAmiyEQMIujtgMstBtE5Nl2Qx60tSxNa+HAA94wP3x3s4phfjg4jF/cRcqahdUyrO9vUVZQV3HV0ajARDrnXywHbCkPxYdWguRpNzhfI11Fc7XbR1bU+rSPUSmJKfxcB0AZRaguvIBEI/S0qa9BMlaZP8d73wPpzfGaK3b9WGMiQSo3l9dQFAvvfTSy2UWlbLooj4NtET8Yvjghz7KF754pn2vc35qt4TA5Ui/EAmdQEsEX2obu+MNR9mRHz+RKb+cswFBo5ShdoDkLI2O8eLf+D1uublsSWyNNnHspAdAejkUs7j1P1T7ikdwIJb73vdGbrppLRYMeJiUkxmn7OBl3u5P7m7isPjkp7aYlB7RObWLtv1F4wDpdE5ZdD1aC6PRCOscXhR3ndqishovAwbLx1BmhNYjtBle9rPKhtiQ44LBYrAofFDUXlEHhfXCa1/3Ju48MeurSjvuBx8FDYROO+GATs3SrrtG8RVPeDRIjUiFwseEnCbpIRyODBBrLXmeE0KsRhGBPN9tae3ECLa2Ns4OgNx6662pznaKmrQdYMIsr4NzcL/73xvv7Vn4NS7P5lVVE5aXY+tV6yzOOXywKFEJgZu/5f686KzEoJXG6BytNNLhzw+BHRkNzdxool/zzn4g7Iq+XTniIUUSg8QUsYYPp7bwrne+l+Fomcp5UE2tusd7Twh+tqV0L7300ksv5yRZFg3xui7Ji5jG62wgM0PWVq/lM5+7hbImlb9G4CHg8S60TfEu2Pc5iwnlg0/OgscYQ5ZFvd+1qY4keJUev/cerbKWZ06bHBfABaGshaqC//mLvxbjkBJJBceTEuj3z14OIRwSQIJH4dneOsM/+//9EKNRDMIWGQzyDFvX6EPagjIk3eUBncF//bn/wZnTW0wmgUExwtqoxyKwe2E6dCf+4TuddUg+pePOO0+wduxa1tevYWNrgjZDNjZLakc6wmU/OxdAaYLSBGJ2mhOFF0VA8Bj+5r0f4OSpccuUEXzkDPHOHY6xDg5JkEwIgckkZlEOhvBDP/QDCHUE8vCLoIEDl6qqyLIM51zinLQUgzSTXFi4H0sHCtnY2AcA8vrXv54sy9oU08j/EHkclMqoyrJ9b27g4Q99CAEXSS/FI17taHvbpD5dajEdJaO1jhkNHQLXyNqsjuzBBR+JTFZ2ggNqzxb1+qrIgAghxCyooPCpK/n2Nown8Ka3vI268igxOBdilMu5uHkECNbRSy+99NLL+YlzLjnNEss1JRqllY3p2//tv/1Sm8AbE0B8awtEEP8iGHRnAUGUqBRlm420drNpjzQQQgwOaK2xwWOtRWeGIArrQHTBW97ybt76tk9TW4h+RwzzOedaexRiRLAJrNhDXmLQy5Uv3ntGoxGTyQRjDHVZgQSGo4LxeIPlUc5Tnvwo6jqWhcdgYcBk2aFxgp31LRjcgB9lHXXmibvggx/6GMqMCEEBhuCTD3gJMxiiX+bbdXzs2DG2t7epao/JhjHTWhdoEbSEAzlEBMTixYJ28VAeVMBrwSlQpuCVr3oN3sFk3PAoqhQwPnj/Z1qGE8tIBkUOAbyD+99vhWuOj1BUZEZFnyUEtMpwvm5pMQ5SBoMBW1tbaBNJa9ZWl2e7y4aQSFsjD5i1sXQrEPDBt5jGngDIe9/73vhiuuH5yLVWkZnbh+j0Xnf9NWSZatNmZj9WzaJ9B54G1L2m/nxu515EJLJ5i2ZrKy6m4Qh+4X/+FmXt8WIIfalLL7300suldEUi+NFaFZpAzom7NvjC5zeYTGA4jFkGta0RlbV8HJfGpmAfe2VfwqEaR6AtZ/F4adLwhXywzKT0OK/5T//556httBhNNsB6QessAmDJJjXGUNc1VVXNBL966eXiG39R05TlhOXlEXVdMxqNwAc2Tp1idWXEjfe6B8HHzA89EyhUiQT/gG1pL2hj2rKMqq5wRBJ/D7z9He/HO0GbAtEFPoAyGrTi4pi1iUNyTzBg2mUlKgY1lzYSDuicdJV4vLi4/4gntPpLIcrwute/GVGxEYcHxhOLNjn2wDMA/dwONf2XFtAKHvqQ+zOenGFra5M8z3HOMZlMyLIs0VwcfBa/SAzlO19zr3vdEMlPmXYdbd+nVMpcUggBJXDzzTefHQB5xzve0QIgO4hvQgcQSbwfN9xwN3IzT/AyN8HloB7c4etlfKASLvA4K7B09gySKxwBoarivB8t5UnBwV++7k1IyFv0cao2QzLt+g5EvfTSSy8Xx5ALdGvYweDRlJXjrW9/99R+VnKZHA+1y17Yy9wGigoRCFFN+rv41jTf2h4zGK0yrj3jCn7mP/4yVQJBqlpwPmYBlWVJmTKRmy6FvfRyyWevCkzKbRRCcBZFoK5LtBECjm//9m9maQmsI2VZxDKvrl14kFLX9Ux2fnQOk1b18IpXvpraCcHrtsV47ALpYgvrC3U/5kpgFHQi+LG7SpcodZYwNer+yLcSLvt56gRFrpcmG8SLJyT/1picL3z+Ft74hk9GXiNSVUJQhyKDYhZgircjErkdRcE/fdE/4rpr11ka5UjKWpRUQVHV5cGvPy8tCOKt45GPehjGJC9f0s34OESLZFEL3B0AyM033ywQ01+63WDmORwaxGXt2BCIPAeHKsoR5luv9gZJLxeuOMrazgBKv/97ryV4Q1BZjESKalMLe+mll156uXjgRyQ99Sn6RtK3AsFQ5Cu84fVvJs8jP5mgE4kmiFwOEKTPntzTgG2dm2h9KxxNO0wvnmI4YFxWHFu/GydPbfH61/8173vf56lqyAuFUjEAVxTFTElRn/3Ry+WQOM88tY0ZR2VZovCsr61SZMJznv0VccWHmuBKvI+gQfAxIn3w1z8FC631GB1/Ho/hk586w/vf/2GULiiti3pVwAaf+JT8RSAinguSN2DMDECkZssa8CABFboVBP7ynyXyvOy8l2nHlNoHVleP8Vu//XtoBc7GMpjgPerAOQBV5GJcEMkWASVw7xtXWFkdIhLY2DgdKVxT6s9hAZqbXIoQHI961CPSCKQSpcA07SoBPCGE1MTD87d/+96zAyBxQYzbmsoQIp9BA4A0Z51yT7SOyOiOxRGmiN7MlV8kQ2gx2KJ2MUp6Q6SXCxdrPctLyxBijd/NtwR+5yX/h6oWvNVTZY7gZ3IGDxk42EsvvfRyRUo3+8N3jDuNC4pPfPKzfPRjt2P93F9cNvW7CATpbRBhjguuIT8UCwkICUpAFBtbE0y+RO0z/vW//XfULnKBWEub+aGUwlqLtRYRuThdKnrpZa+VrWg7JTbttvM8Z3u8yeMf/zhMFnVNZjRG6zZwfFjmZoMBR466CCBaD8UQXvbyV1LWnsFgCWdjI4SmhEDpacvUC/LahLbkrQuCqJkj5eyHQGTasyhs0hPTLJLLfTQ+bXONU5CmbUVGObEgGZ/59OdxLoLw8bmrnZ0jDhAI2ZnRbxE8kzLwVU94HLaaMBqNIn+GtxGAOiTbV+Ri8WSZ5n73v08LhrS2gMhM2lAIoaF95X3ve9/+AJA777xz5gPaL25+7qBxAVheXoqvN0SoB5nuFXpF3cslXHzApIo1fr/x67/FeFxRlQ7RhoBO4Ef/rHrppZdeLr4Sjo5z6ARUfJOdVzpA86pXvZosix1ZIu1mX4J4WIxvxbRzRgtipVr606dPM1xaAVHUFrJihc2tiu/7gX/LpARlYhR+Mpm0IEhD0N93WevlUjsWZTkmhECWZUwmE3JtUAQ2Tp3kBd/wfADK0kKwUSuljohK6cPhlyS3zdsYMa9r1zrpb3rTWxgOlvAonI+ZdR5wzhLE4y/ZDUy1c1dPx393yk6639/4l5fzHNT0CoMBDF1wO6AwOqeuHKD4oz98C03ShPP1odiApO0Bs3NuKxyDQvj27/gWlBIGeYFzNVmWobUwmWwfklbksbPmaGnI6sqsXxZJUGc3eyWRAyQQeNe73iP7AkBOnjxJURQzH96clVJtmknAYQxcd911CYWZndaXbhXvdhuyDxDEH91DLvA44tIYWXUN7//gbbzyVX/J6to6g9EyBN0xxucN9l566aWXXi7Mggudc3dPSmWHotEm5x1vfyfeQ+gg0RcngnWWTL7z4s86QvBHmLPhpHk48ee1Y8fY3NxkUlm0KTC6YG31Gj7wwY/wX3/u1zlzpmxLYJoS7aYUpu8C08ullth+U7C2SqSQsYvRwx72MB71qBsByHKND65TahBLt/xhSAJJAEiWNWWBGmPg7W//CCdOnEBrzXg8bqkPvPepXbVg7cUj8YyZILTgQawUEMSno6MnBItgW13vUXiRy35ugA/xBtDR3g+mw6miyLICrQ117fj93/99El0gVT05LBr4LP4w3HCPEQ9+yIM4eeouvPeJAPXwdDFrsqpGo9G+YQLnHWdOn97/U2kQ9rquZ9MLhQ74ERAUmYL19TXopKHE/u2SuBAUF6UEoMu42+Acwo6Wu7thjL30shgYmjNg97JtA0xKGCzBL/3Sr2N0wV13ncSYPK2ZxfNwZ//zXnrppZdezkuCxCzTMEs+mg0KtsYlt991hhMnanza+esJ6MsBQsyHMHvjY87x8TvGUXU2xqZV6Pr6OnVds7k9ZnO75p433o+Xv+J1/Mqv/C6xk6FQpdad0Tb1CeCyncPvso+nDKK+LLWXc1zcKsvJsyGTqmY4HOKDJVDxnOc8jUEROR8UAaMKBIP3grUeJGYvHbjoNOUFtjbLeE0BfvEXfw1Uhk8tqrOBSa28o8OZmwH4S5HVP7sG97aT/S6K9XKdVapsUAuuLO5FVe2xPhbubGxO+Pgnv0TlYDBYuqjP6Xy3pl2dGrqAE3zrC78eoxxLo5gFcuLkaYrB8qEg8gXw3lIUWftkpGl0Irrdg13wLW5QTSpOnjy1fwDkb//2b3HOtahP0w3GhybqIigMGsHV8LjHPhppU1PjhHASmdmnhJAXK4vg7J/RpFl2HVqZ+W1/nN9xZYMfoXPMgCDduRLAVq79ua5dG6MaW5Ac/tt//zPe/d4Po8yAzBQ458iLDAmRKKmpEfRpLfRmVi+99NLLheIe064vKmRob9BeEkGex4ZAtrTKVqn5qf//z2FT5z5XTy5SJsY+98EeBFlsa6uAUyERLCpiGrlBBYMERbCW3Cgm25sYE7N5kIyNjRrR6/zZy97C/37xq6kdiBY8oHSWbMsaqHFuC6gJvsbZqh33qoz7vsel86w9EFKSf5+s08viCawI3uDJEHKUMpgsoNWER335fVACufEYaZxlE3k0tEoZ1PaAATcPoYIsAjKDYYFzUFv40h1nCGTU3pENNd7XVH6CBEeuC8otT66W2izn89aeIcwczeIMqZPK9KClSZ7qCrXwMy7XsXPsQuoGE/ceL0LtPErnlFbIhiu84Y1vQwy4tvWvmssKXOCD7Mga3CNgex4giDT/UAkQU/EHQZOjyAN81eMfzvIgYOsNnLesrF/H5sTO8RoeBICncCHywtzn3jdQVfEWqtomACTemw8WMVPeGK0Kbv3infsHQD796U9P62qg0w1mmtYVQkxVMgrudrfrMJkiqMjmHVoyyCZ96OIt/NmW0GofkNfFmTy9XB3iz/KqqypMprEuvmYyTQix3a0y8KEPT/jDP3452gwZDFcIAaqqQieGqd24fvvZ10svvfRy4SKhkzbdIdV0QRhPLI6cz37+VlwyJEfF4HKyoPayFwjS2PZBzRzT8phZGCLEmHp00OqCl/zeS3n339xJbYnZIArK2qIkZhprHTt1SOoY46zHOchztWMvng1O9ET5vewtS/8fe38eb01W1/fi7+9aq6r23md4pp6bqRlFpmYSUFRABEGJEY2ZvMYkhqvGjPfeXBOTm9GE+HPCG43BeBMVRA2KgoAYkFGmNA3I0IxNQ0PP/UznnD1UrbW+vz9W1Z7OeebhPOd51vt51VPn7FN77xpW1fquz/oOq/t54P6j9KoB4/GY++67h0c+6qF83eMf3ramMLMBtw1iL4EYGKsEn5II+5jGvq997dt54MFNVIrpZHXy1JrPueEu8Oy/nvg3vXTKCC+MJ2U2ruwEm9XVVUaTml5vlY3NMX/05j9mYzNt5XWn0PiLPC5dUnh1mtllJs5MxoF9q/DsZz2NGCaoBmIA46pdvwYpHCuFPa6s9ClaPa5sqxtNbzuJrX6RhCsnJffddwYCyEc+8pGF0rdd7oPpa3NuM8bATY94RJslOOvnmUuVJMgZDLKjV0ur5MakGjpnOHZso233KenpV78KP/x3foTV1XWqqs+9DzxA1V/BWptqrGdDKpPJZHZHGBHBGIO1lvvvv5/P3HZPet2RbZPLgKIq2doc8Q//0f/Jxz5+D7bNrVAWPSaNMhoHopqU2FEFRLDOEHxoTVaDocDg2mVmD3Szo7LTYCdPoGWA0dYmg0GPsnKsr65ww/XX8jd/6IcoHDShYS+4esWYvA2MgYmH1//+7xNjVxVGOHGeiMypKo/6JmLEEYPim8j9Dx7hllv+nOEoXIQy7GeuhWx/MSXHLQr4q3/1rzIY9FhfXyWqJ9T1rucz7PLRiMB1112Hsym/h0ib/1SXj1GJGjFO+NKXvnT6AsinP/1pQghTAWTbDSQCcfa3668fULkCIWJUkXzDZC5FA/mE4sfsAVD0e0zqCQqsra8RgeE4cHwT/uE//j85dOhq7rrrHnwTOXTwaiaThtFowqC/cuLOIyeRzWQymQv0YJ/F+xZF0SZqt/z+H7whWSIKmFwlZK8TglINVih7q/zIj/493vtnX0AM+FYE6fXWMVJhbUUIMaWlU6WsUhWOhRy67TpHJ2VOv/0F9q+vcuTBBxiNtojR8+xnP56mhtKeKknkpTEALsoKlSSAfPxjd3L/fUdYWz2ARrOQ0HO7/ZqTDM/6Gl26pklAHY/H0wTNRVGyb/0Qf/y2t+M6V4XlEuDMJYG+RMYHVWWIEW56+NWEpma4tYEGjysuhfYbUVXqZswjH3VTJ0ikUyo7CybtwI8vfvGLp39n3nXXXbK5ubngBdJtqrpUFQboFdAfVCe5kHlWPHMJcKIM/a1AoSiqgrUFTZNiESc1VD3LP/8X/467vnYf9993hGuuuxEVx/33P8ja2j7KXsWkqfP5zWQymQv2/D61K/RkMkHEUhQl73rn+zhyZG6uJo929zCG2kdCFHwQrrrmRn7in/1LfuVX/wgx6RprpHXjd1jXw3tPnbKmIq2DcheeIKes1BNP8HPmSqVfFWxublBWltCMeeQjHooroCrnxsfzP8ii2bn75m/aIe/T/rz5rX9M2RugOGJoc/Kondv55TK0V7gX3QlEivQsSaEZ3nvKsoe1BceObfLxj92Gb6CJbfoGOaMLduEO5WR96BjW1uBbvuW5jMabWIlYE+fCFHcHJ4YYGkLjefQjH4WQolOmOoewoE2oahsayZl5gAB87Wtf20EAYSE3yFSAEbjq4MH2BC3HSM19jWYhJHPpPtQmdT11m8UIIYIr4e+84p/yyU98jkF/H03jp0r52to+Dh8+zMpgdYcyfNnlOpPJZC4W1lqG4xoxFoxjczjkYx//VPIS8PmRvJdRoOz1CSpsjWrGk4BxK/zGb/wu/+GVr2Fjo51LDTAaaSr3WZWtN1DEdyGqeoLuWTiJ4JHt1mwnRlQjvpmwb31AUQp//x/8GFWRmk5TN3PexcvtynApqK/SlgsXC/fd5/nQB29BxDEa1Ygp0JQVc7t9nD2YmZXrXq5kNbu2xhiapkljATVoFDaOj3n9//gT7LYMETr3rNkpyepOz5/z/RzangIAhcEArIFX/J2/yfpKj6KI6CXQgYpJ52llpc+1166lI5DZ81o06RHCopOG94E77rjjzASQTjGZFzyUWT4Q1Zn/oCjccON1pDI6KbXUYjLI3IFkLqHnWHfDy2IbrXoDap/UWiwMR/AjP/Jv+Mxn7sT7kq1hw8raAQ4fPsr6+n58DBRlxfHjm1RVf1sZL6O55Wcymcz5IMqyQa/bjNGyLAlBAcNgdT9v+eO3p1J52ftj719/VQara/iojMaR4TAwWDnE63//zfzyf/kdvnaPogbKntD4LiFeGpQ657aPQeUEdsGC3WraRKy5J7+SMQpoYGVQcvzog+xb7/PkJx1IVSfibFx08sHmbrah1IZjWy3j3e99H0ePbTGpI8b1kmf/QsJRnU5ii6blihdBpHUd687D0oR+CGEqgoSg7Nt3CBXH7/7uG09POzhTD5HzfnyC94GoUDfwsIcKN9xwFUY8RbHb119TlU1Rrr/+Wnq97lWdFmmBrpS6SVGv7euHDx/m05/5hJyRAHLPPffMvlp1Wz+hXcbj1v3w6kMHT5j7I0+8ZC6dh9i8+BEXOgcfW3dpA/feq/zkT/4cn/70HYisUTcFRbVCXXsOHryKe+6+F2scvfZOnN0jc/fA/AMjez9lMpnMeRVBkGQYQarGtbKyymg0IaogxvHu9/wZW8OcAuRyIITI4SNHOXTwGsSWYCqGI491q7zhDW/ln/6zf80Xbx+nhHhCWybXEuJS9cBpxtO4NMMdT2UWZ65gfD2hVxRoHPOt3/Ismjo1GWfAWTdrM3MZdfUSaUtp7twwaZQIvPd9H8QWPXyjrAxWmdT+xPuZ89edhNk5c87hnKPq95j4hs2tMULF1shz9Niyj8dO4/EdzrNeKO+P+f1Pn63e4wqLEahcevWJX/9YHIHgJ9N+dtf6fg1Ya7jm2qunbVpjKi8dGt/2EWGbZnH8+PHTvIJzfPKTn2xFoe1hMDEEjLVoiMm5K8JjH/doYgxYEULwOGeIMeK9pygKguabKHNhmVVimTeawjQOOLkxelJMY5wKH6MmEgAPmALuvRde8Yp/yC23fJqNDSXGFaCPbxTrCkbjCb3BCjHCZNJgC9e2753buFGTzalMJpM5J042RZYSoA5HI3qDAT7AaOIpqxV+9/XvYDhOpR/n+4n5RO87JXzPXFqoEcregK3RGFRwZYWPBmwPNQO+8KW7eMWP/AP+559+FmyqCtD4tlcWSbODJgINiCeGBojtZF6cTupNv6/NKyJ64vL2mSum9SFGmYw3KSvLy7/3ZZRVSia6tVWnp9K2MqOnHGZd3AEkgBGOHIUPffhWVJO/fmgrJKFClGypnrAFSETnxCDTlu/uPL1DCDjnUh4qDFV/hfHEY2yfn/u5/w+A8SRMPcrq0EyfOCGG0+vmLqCQ0136EFIooQD/+yt+kNAMkdhgTBrTd/2mqhJCmJanveDnXwPj4XEe+6hHIkD0DdbYafWaySTQ65U0vsGIQRAm9WSqZZyRANLFzEwzqTIrMyMsTqdYB9dfczWm7UTEKDHGaVm67udM5kJSFMXUwB2NRoQQsNZSlilL1Xg8RtoKAcPxJHl9qMFYl+QQA6/+1TfyD//hT/K5z3+NsjrA2vrVTGolatm6057olomcqOvLZDKZzAWQROYetZPJhLIsqWtPjFAUfYbjhg//r1upeimuGZJrrPceay0ismDjZPbKYMSgCIpFxRGxTCYwmkT+xf/zb/mJn/hP3HVPwJTJS7n2YJwlqp/GQ/k2rl2A8WQ0jR3v+nchh05lZvTLkqoQHvOoh/HIm9anmT0Gg3LJ5Iuk6bQ0KXapFFH2HqyFV/3if6Y3WMO6lKxzNBmncdqJslxm7+WT9UDT50WMkag+jXudxdoC40oefOAoH/zwrRw9BlVpUz6qtiJlNzVqLwUXRREgYl0a0yPQ78Gzn/00jFVCbKb9p7UW5xxlWbaJSOMFPsspDGttbYWrrj6YxnuuQFEQwViLcxbVWdqOxjdUZcXHP/7xMxdAPve5z7XnRKbqTozLCVF1mlH7IQ+9kbJ0SAxYAQ0N1lqMMYQQtokmmcz5f8D7qRDS7/en7VZVaZqGXm9A0JQpvuqtEEi10DFwx1e2+Nlf+B1+63feyG2f+zLXXPswxpPIuI5Mmjit453yfMSF5YTix3xMZe5EMplM5jyy3eiytgAxGFcSIljX41Of+jwPHE4zW2mbRYMthJAnaPaG7AESid2CSf0xgoqjjlD0V7HlCu9495/xw6/4cX7v9/6MEEFc8vsM0RCxBAUfFEXwMVJVfU42S581soyIMBxt8PznPZfCzapLicwmhxftwEtF+mifey7t1Yc++FGaWjl6fJP9Bw/RNA3ihMUCFu2xZe+nU/Y5Hc4KxOSN4L1nXE9wZYVxPR584DgfueUTRKD2KQzJ2gra549eGg0cP5nM2nSA9VV44bd9M77ZQjU5MnSTBjHGqefkxZhEiOpxVrjx+uuIPokioWmmQy5roaljG45GlyGV22677cwFkNtuu026eJr5B0BqAkortUxv/EMHhAP714HYen0kTxAjjhj1orjIZK5slhNRdTeniKQSVSEZQKFN8hNJbtEf+vCX+Zt/++/yh298G8eO16zvvxopehzfHFK4itXVVSJxMY54m+CRy4RlMpnMRTdIWws9eX/UWFtQFBWjYU1vsM64jvzO7/5RErxbA6/zFgSybbLXRJCun5VIlJQXRoGi6lE3kSZYqmqN0Rh+/hd/le/7/n/EZz97jKjg1SWfD3FUg1XAYU2FDzGFBERmhv18QvPcRK5s8QOoJyOcFV74bd8CEaykwZOGEw2SLx17MAI+wOe+sMWDRzYxtqQsKx548EEOXnVVGybeCTY6FT9mOSIMciVP4smJq7/Mjzdi9FgrGCtMJhNUhZW1dQYr+/jjP/lTQgDbDsyl9R6JsfNDu3jndzFjoZk7hjYcMIKz6QifcvMTWemXWDELTg0p2Wu4aH2oMYbNzQ1ufMj1swIsSxMXZTmXk8U4Jk3NAw88cOYCCKRSuF0ej/lOoIv/6W4YbRWjhzzkxlQFxi4OPtPPeYYlc4GfUa3yOp8Ip4tRG09CinGzbfaPAr5w+xZ/5a//ff7Oj/wjsCtMmoKiXKHxcHxjk8HqChglaEPU1EFE0aWl7TBOWi5McvvPZDKZC2nUqU5nprwPYGyaaStK3vyWP04GnchCXwGnU8Ehcylg2gqDhojRxQmHKIItKo4e20olPaXHA0eH9PsHuPOrh3nFj/5j/slPvIrJJM3djcfp3ZMGjm9OMLYEEcSYmVGdvT4ycxJCaCZ847OfxQ039mjahEJKwBhmOUAWJIdLqwH5Bv71v3olvWqFJgpl1aeqKu655x5WVwfJo9kEtjf+3a5gc+mISHEHM76rkqMa8KHBOkOv18NaSwgB3wTqELj1o3/Ol+44gmsjppqQEtNeSv1PUZakyj9heuVvvH7As5/1zFnhE5GpENJ5hFyMYyjLgsZPuPrqNVz7dda5bapO45v2eqW8YHfdddfZCSAf+9jHZmEESwPNeUMihDS+e/SjH73NIMnJxTIX0wBeNmiNMRhjqCrLcJSMnw99+E7+7t99Jf/gH/8k992/yYFDD+XBw2MaLxhbJWMI6K/0GNWbjCebYJL4sWMHl7NkZzKZzAV+wHczMDuLySEE+v2VNDvlPYPVdYajCUGFo0ePc++9m22Z3O2Z4jsxJHOJMg01nXfTT5MPnRQiIgxW1/FBEFOiWrG5FVg/cB1bW8q73vm/+Mt/+cf5iX/2n/naXROignOwslrR5UlftG91p9iGzBWKc46Xf+/3ED30SoPB45vkURbDaTy+dnn/Nzbgi7d/mSNHNxEcMaZn5mClR92MW5ePHaohpbjv3ABOYeeXVQFE6npMmE8BoRC8cvz4Jrd+5KPThtBFjRhxc/mHTvSd53GMoSeQd0JIbk0wXTc+CQQve9l3oqp47/HeY0xbWrztdy98CEzKL3rNNdfQ77f9fZuTZBrbuhSK5r1HEG777KflrASQ97///YjI9EC7i2atnWWHEpn++KhH35Q6jjiLC1JN4S9ZCMlc8OfTkhLpvZ8KdQ/cv8Ett36Of/JPfol/+k//Nbd+9Dbuu3+LycTiG6Eq1xj092FcxdZ4xHAyYjQaUpSG3opDTNhugLVLF5O8wx6RPT8ymUzm4jz/k6lkcK5EJLkXixhsWfCGN7xhup22meM7A27exslcotdXU9z34iBtVqZlazhKlVuMIURhdW0/W8OGzY2alf5BinI/D9w/5h3v+AA//uP/hP/j//gF3vu+LzAaQlGmRKmLlrG2s54hT3JkOHToAE95ykOZaaU6PwyaG1jOCwmXzrjnve+7hbIYcOjgtYzHYzY2Nti/fz8h+LlUrTt5MZuZEJLZefgsimrAOYNqqjwZYkTahKFVVeFsybvf/WdMxqlZGJPCktqWdBH2NZ6qA11ovxojzqQjvPkpj+LAgX045/DeL1R+6SrDXGhC3fCMp96ckpnr7JjmvzvGlBwVkmD5lTu/csrPPWHP/6lPfWr2wUExVhYF8RjBRqw1BOCaqw4hkvojg4EA2tYIiqLnnEwn3X9mbniZ3CFFUznerg1FaaudmRM1AENXBjWvL+b6DB4op9smln7xXrA2tRWxFQ8e8bzjHe/gtb/1eu4/3LA1iriiwpbrRK84K4wngcFghVE9mRrGBw8eJDQTvK+xVlCV7VqGmtRZ5ASnp2O+Lp237uXUPkT9tjYyi7JsQ+h0tv1utV8hovP7P/1bBPGz9ttto/OfkZmemx0GN7P2sNhOZNszfDeu/6wPmpoJEhHMXJs8wTNsW8eXBm3Slt80tH2jmHaAt3jLCF2StHPov+ZTFbXHoPN7KssPUzu3G2k/RXV3zXmJGAWVZC4u2+MiwubmcYwIZVly9MiD9CuHdQXj8XH+x+vfxA//7R+YxgmLKOCZ1EMG/cH2e5e4g92wV+2Hufto7h7TBSM+lYgXUonHKMIll9dq4dk7u71UYHVlhaNHj3Lo0CGOHDnCSm+F9f372NoYEoO2/X6BoeG+B4ccveVT3PrRT9HvGX7sR3+Yb/nWZ9KvoNdLpS2RFHCzPdr/TAYZcXEQeaEHMadtY11p9udcW5cI0qRrKjFNU6lJ4wY1S8+V2F7/hmuuuYp+CbYANCKSBluh8ckVXxfP8fzdM2tDF8uOXuqHFP7Lr/43jOsxHE/o9VdwlXL48GGqXsW4nlDYmc2SxpgGUZmej6x/nPx+HA6HrKz0cS7lotJWHIhxgiVS2QG3fvQ2jhyDa6uUtLNuPFiDr2vKsncCkzmen/Fr16epWer/28+1BkJKIWCLKjkvOAgReiU85LqriHqUptlEoyLGtukuzCmcHGRpxHYW/YpENIx5ylMen2wABWeKJE7b9FkhJo0CwAePs473vefPzl4AOXLkCMPhkMFggGpAcCn2pkuOY+dPrOGaaw9gJVDXkV65znDSDh5NIBIQcZizHSwudXyd8CFo+lkiofHTc7w1mrC6UgGRrc1NVlZXT9Bw8/rirZUYG4wptqsYO2bhNztcs66yS5sOrfNUBcYjKHupANlb3/ppfvXV/4277rqb9X0H2BoJQQaYMsWFx5CkTVGDM1DXk2Tyi7TZkGtAUlWB7uXltqvmhAO6HeSZLIEoOwwqfTK8pcGYhtLJ1G/GN4o1ihHP9pHw7rRf7z3WlWhKMZBqvpuIc4pvwoLJY6JZ2t8ruz1IFEQc3tc4MagmsdIohBiwVlOHNn/vz99bEnfx+WXACGINzcQzaTy9yiYPAwRrhLiQ62qWqFGnx5EmEKqyIPoh9WSTXhGox+nZ04kgizPcbS4tZMn4PcP9n5880Jg8N6fBA22/KW3JeiyNj9giJVK3EpA4AbOyq03YzBlYuq1/SHZAWSThZjzaYlBVAEwajzErjBrPrX9+B1//+JtYWUkGaNAhg76hmRyjqFZB7bZ7VvHtVXVz13ev9b9zgv30uCIRi+JaY7hBqYmxAhzGtGU82za5qxLutoEpC7ZkBPxkwmq/z2Q4ZFBV7Uysp6hcenZP20+BlAWNpvx2kw3lP/z0r/LvXvn/8phHPZwf+IHv50Xf/iSEFKdfFe1cn84lQ20n+ma5UrVVTeaf8/P3sWuTZZpk6sw9JnSumsjy6zJ3yypN+v7OVlp44w6/Tycvk4u6K4or2P5NIrAPQojjJGI4Q6g9wRu8T8+6olewublJ6YT+oGA82qIoHCaM+KX/919AjBjtvN7TebbF/Dioe91Mn9v2XJ7bJ7BDdhLDUk0jy3A0pN8ftLtjqGvPLR/9AhubI4I6xJT4qKAmeSz4QGEd0yq4Oi/4K5diUtfdeP6YbRM4uiAMlVWfJsR2PGNnRbWNpAkFWUGM5Zd+6b/zr/71DxEC9ArDaHSsFeCbtN3cWHch9cR5aUfzEys7iKrWYdsqKqatDGSNwQj8ws/937zku36I9ZU+9z1wmJXVGxmPJxjj8LqYkDS1JZmz+bvztJOQ107v7Oh51D49VVmplKc/+bFITNKD9+BckaobacDambBixRE8/OEfvunsBZC7776bwWCQPtDaE7T/zmg1HDqwRvB1Sv4SU8mciGKMEkM8Zyef5XkYQRdO5GAwYDyGQR/6K1WrNRn6g1VC0Fl4xPzsbF5fvDURY0p88NMZO8FO1UMRu2MT09YYqOtIr5c6si7zOwpHj8P9D2zxpjf+MZ/69Of4ylfv5vixMVvDCSsr+/FxwNbYU/WrbR4jZzNbfXriR2ZnAQRUTDsIm5+WjsToGQ6HjIbQq6ByNnUI+BTn17q27Vb79c0E50qQVEZRFPq9FVxhGNVjRPo7ZAfP1YE6ytIxmUwYDAZsHd+gKEq2NqF/sMvbM+9BM5cVXZcH8Ltz/YOPWFNQrhbYSZNid8XS1LGNgTUnmKVNxrEzPcbjMb4eUZhIv1/RLwtWVlbnJkbamWLxS7MnZvtkwJmst3X5i4aRAlFj8kIRwbZCgg/pAeycwV/yUQA6M+KXB89YbLGP9/7ZR3naM25iOFH6lWJao62oesxcvc1UENK2TSqzrP17tv+du4eU0PoemVlLEBaCqCWa9lSES+TympOIYyebgOgM7DREnN5RalFs8ijSAlXDrR//Arff8fP8yq8cYH1fn6c+9Sm8+EUv4HFfdwgjrdDi02kqbBJEpm1j/pklcVF7QpH5SgndxI0sCh8pb14aeIiZDYTEgsx5ZWk3AyQyO6ITGNjGGozl3J4fl8F60oArehSFZXO8RawbxPbolT0aAxFh0tSsrq4SYs3m5ibOQgwNj3j4dRQFOBsRXHvx3MwDeGF9otfPZf/NVOZNo6okVGsbPKHEdowjDPpr6b4OcGxjiwP7VnjHO94N6tD5PnWn+2fHeyzbL6dt659wvCCMxgFrDR+59eOtPZAeEUn86LycZ9dHZVF2tufl/j3T9hqnu7a6Alcf2seRwxscWF/j+NGjYIs2z6LDt4lwpB2fdW2q89BL67McL0nEULO+UrahkDud68hwNGTQX23zOBm++MUvnb0Acvvtty/6qsjJ74f1ddi3bx/HN4QYYhJC2tmlC5IDRKV9MCSjpG48tq3PbQRGE+hXBjHpYaALfY/J64u89rXBlUwVRkiZqefrqEdJY+PliQw1UFSOzQnc9bVjfPFLX+LLd9zJl798J5//wh18+c676PfWOLaxAaZkZWUN1xtQN4FxmLC2b5VJnRPdXVrDFeYcOwzWVKys76PXT9c+hJRczBqLaYvYq+zO/YsajK1AUrhfjMkDJKhh0kSaGspqOUu4zj0248z4uELZGo/Q0BBCmao1qaWsZt5cMt8VzXl2LRr2u/X8AmMcw+EQaws0pplcY5TxuKYsS6w9+fWt6xrnHL2yQnTCcHiEY0eOs7k1IWjKOzbrYt02KcWcQ/+1/cYzMzFybq66G+QpEHwa+iqO4GU2Hb1HGY89b3nrn/Jjf/fluEKAEjA0saEwRTIE5/RKBaK4qYl0Lud/99dzdlsb2jHvrzDvpbR92unyyX8hC1bgzI5UMVT9tZS40MKDRzb52t338dnP38Gb3vjHGAtPe+qTuemmh/OkJz2JRz3qJg4eMHRzara1YYx0TqQGI4tey0GTyJ9mhUHEImjbL7RCpEREzHRSQFEa71N53uAwxlJYg3FMw8BiAB+TIJO8WiFoKs+q0j5X7NxA5Aq0P6Omsp7WFG0fYrGmIKjgfSCqUFQ9jh/fQkMECfT7fXqVY+P4Ub7pm74ZV7R9lO7e/RslpRfoQrPm59OthSZCbCIYqBys71shRnj3ez5MlCJrGbuJNaytrrK19SAfvfU+nvb0axiPPGv9imWzUWXe12exEMlu3T8i8OIXvYTXvOa3Kfs9tkZb2DbHyWhSY7rSLAuuCnICN4bt4lrn0RdP4imxb9++2bN8h+06JweRVIb41ltvOaXfxUmzf91xxx084hGPWEgadpLehac85cm8+z2fQtUiJrnAxNYb5HzffElVSqq7iqMJE9pxEuNJOv3zg5WoJ5wcyOuLsDZFugYbmynO1lqoA4hLM42dEOIj1DVsbgYeeOAB7r33Xo4cPc4b3/RW7n/wMMePH0esQ9QwnjSEELFFxdbxLfq9NVxZ0MQUrtD4BgkW50Iy4LPHxi7KHfNuePPRsUkcaGLD8Y0hkzrds6u91uVYJbUPu4vtFzBi8NomnXbdEVkEy2B1QN2c6PmYE+gBOGfoDVYIfkQ1WEGj0B8kN3OvULjtPhQLLqC7+PyOCkXVR8QymTT0qj7OFThXEoOwurrKcDhmm4IzZ8D0eskDZDicUFUwWFnDrVaU/TXocofLohG00HrO5fiXZ8m1m/Gf+56YIkC8n4nQriAZzrLXk4QaxPR48PAx/vhtn+S7vvOJKfgnOgyOJi4aQpGZABflPJz/S2C9QyaaqfiRfrYLM45q4sIM3GnPgl6ynEhOTH3MZJLCc+vGI2ooinXEKFvjmhg97/vgx3nfhz5K+M3X472n36941CNu4uanPpnHPObR3HDDdZSFpdcrWRn06PcLej0oy7aoghikrTA37xfoQ4qjL1yquuC9pywLOj3VuCINJeYuYGiLNogkO8qaNOFXljOPEnGzp5Be4fZnF2U9mQRGkxprS6r+CpujmrquMa7k6NGjrKz0sWIQE0EbDh89Qq8s+J7veWkaBHYfG7d/z0LKnAt1/+rMHyu0OzP9vRXWTJvjyKdIR377d9/FcORBq2yE7CK9Xo/N4RZNM+E//vTP8tu/8x/p9UsCUE+gX7b98k59z7INtAv3T1PDS1/6Hbzmt/4HGxtbDAYr1AEKV7E5HLcCyFJY1lwIlU5rRZ9dbqler2Qw2PmdIkKInqqsUg4TaxkOh6dnl57sj3/yJ3/CK17xihOIH4Zl98jnfes38653fhRryqlBF0NywztXAcTs8EpSypJOZWyPB4/AgQPgynSaA8n1rapyPY7dJnX08OARpSiE0ajhs5/9LF/5yp184hOf4N7772M4HLK5MWQ8HqOqOOfSjKtYpFjh+EZNCAX9lQHOWKKtsc4yGAxwIZVn2hoOqeualZUV1lb20TQNde2x+RLsLt3TW5alEYOghGgoe2uoTUmXwpxuYu3uO2L7btdtWo9rkqonBcOtCa4ckJ8yJxbAysoRvGc8HpO6A8vd98L+9RS2GOZ7E91BANnlodP99x9m//791LUnBqWuG0JQRqMxdV3T6w1Ocvgpv0ZRFGiA4GtG9YRBv2Bjc0Kgtz3H8g7j7nO6Aksnc15Y0TZMOSopjHRl9t0PPjjEuD4aZM/OICqGSdOwtnY1H//453n+85/IyqDVxAH1EMyi4DUdqMrlcgfO/SCLkrROTcHlZJ1tkMxenzmWWencWe6QRS+o4COlLVGEGC2uLNOAUivERbAQNIBTrA00Knzujvv4/Jf/J/A2YozYto8TVVSVqKligsQU4lsURbJL1tbo96tpqUxjDOPxmF6vx3XXXcPDH/5wDh48iKqmhIrRY41w8OB+rrvuOg4ePEhRJLGyaVJpyl6vR68H/T4UxWxCSdvBub2SCx2156FwPYIXGt9QlFCVfbxRggr79u1DRNk8vkFZWUIzYnV1la//usdx6KpZsP+udUZ6gt87IcRD03pUV1VqGzHC+z/wccZjkotQNk92p/kJbI2GGIms7j/Al792D1+9G64+lCZ+XJm8tqYTE2wPntbdtIE0TYY85CHwhCc8iU984nNoFOpJg5H03Ap6AsF8Xrk9jaIRs4kaWRj733jjjQvRAsuShEYBY9o8k4Z6vrb5ybqGk/3xu77ru/RNb3rTCW68+YRalnFtePAwfO/3vYKgKwSxqIBnhHGCCfbcZhAWZiEsErtEQx4kgDRcdXCV27/0eVZWK5yxrKz0iTFy5MgRBr1evhN37QYyiK1aZW5EVVUYY5hMUl6F4XCIKwpEBGMczrlpbhDvI00A4waIc0nt0zC1UlWVEBpUUumjJNalttL4Cb6JOOcoTZE9QHYRo7POYKE/7wxTiUzqLa6+aj/333MPVWlwJlXq8d6juyphaZtkiRTyEmDfgauYTODY8S0OHLiacR2WjldTlnlNQ2jlyi612cRA00w4sG+NejKkqcdcfXAfRw4/gAg4s1iCL7LYXuwuDsIijpX913DnV+9hdWUNVSEEZd++fRgpmEwmhLCcnE4XO31TpJJxzYTVlR6C5/AD93LtNQfZ2jreDp66c5CivM/b8atBdNHHQWX+fGsKw7EW28b1Hju2AWK5/vobeODwcUT6e7gNC2XZo5mMGY+PstI3hLBFaMasr66l0n6hzTnU5mFRia0xGtoKNHHvlkNVQ+yunaSKW2o65csRtcBIH7ScPme1bYezfDR72INSIrE9DqPzuV5meV1EBGst0Qdq32BEUQENHrEpH0ryOo5tjLmkahmqs9AW0daGMW3ywy5oTwh1WPgegBCbqXd1Z7Rrm3PHWjv9m7WGenQMV6TtutKTqkpRFBRFSt5prV2wndJAQRBR6np85ZbzVYc1PZoglFUfW/TYGo0ZjmqKoqLsDVCBzc3jFNZx6Kr9HD18H84Zgm/oVYL6EaKz69UtTNtAvKD3b3I7lbkB5WIep1T6W3CuJPjI8eObHDx4FZsbQ4pen0k0V3wY7m4So7C22mc4OkpTb7C6UjAZbeJaAVR92+90GZbxRBOmz18T3e5JIG37i1iiWoIaymoFMY7NzS16K/02B0icJmvthOaueIQuP3tO2ZfMjtUx5G/+tRfyo6/4vvTZc/4UMlexrWkaiqIgBOVjH/sYz3jGM84tBOYjH/kIo9GIfr/PtnKO06zirYFmYW0dej3H1th3FXcw1qDn4dmwows0rYIvSozCXfccZn3/1ays9Dl27BgPHBnT7/fpr16dYvsyu2V/MpwEelVJf3UFVWU8qQkqWFOx/9B+hqNJelCoptCY2HYy1lI4y7hWbJtIbFwnZ13nDCIGJVKVFXUzIcaUEVhEsKag6CdBJYxzDpBd7gLm9Nz5V2eJn3rVKpubDb3BPtQ3jJuavqnA6Fyc+u4Y0IEUSlcWBUw8o1HEuoqqWmE08W2Yn7lyjcxTdTTOEFXauu1Cr7/K/Q9uUJZr9ErXKvZxboDOQsepcfeMNxXDgw8c49DBq1Juq5ByfwyHQ5ytGA6H9E4hsDdNQ7/fxxMJMVVnWF0/wLGtMb1iMB1Ede1NWax8cU59qKRBwLZ7satSI5FBNWBjY4MYA66yrKxfjYjl/gc3iGqxzuxpT4DNjXSN1g9cxWS8QX/lIKEZ00RoQuqHpoOM9tzPRICI6h7uP8TMxKtWAEleEaldiLhZee+55HXzlZf2+uDJ6EnKiErEe6WJgcJYil6FaMDHlMtORfAxYJxrK1ilENvQ5rorql7r6RHbwY6mkDKdTdJUxYAQlBAjPnTbJUHEiCFEkKIHqskbbl5HjZGyXEXbfkhFkyKqShPB10LZ358mg1RpfFzIuyeiiOlfwfanIYgFEYbDCUFrXFmxf/9+vI9sbG6mSTmEwWDAcDhkOByzvj6g1+uxtbXBoOwRKehS6Xb/tPX2uaBj07lchzP7YrEtexV84zFeERy2XOP4Vk1/ZR+bwzG2uLInYHa9CYqwuTVkNB7zkBsfyn3330Xp1qiqHlsbQ5zrdcZGur/FoyZVSEyPgV2cwBUoioIjR45x3bUPYThuGNce69LkpDEGCWFOt9jZ1t+pGs3smGRxnL8w4I9887c8N+VZMjPhI8ZOADFL3xl485vffHp26cn+ePfdd8t4PNYkgOzYrdBm2sDaVE6q16/YHE3aDjOpWyGcowO7nMgNszPK4lQBtc5y7PiYXn8fUWEyHifXY0lORdI2sLy+mGsYDEomjSdMIkEjViqqfkFT1zTBY0zRZg8WAnGaSN2qoCIUVUHtI6KBXq+XWl4IqAaKoqKeTDDWUlUVMYZUstQWiArDzRGVK/JTeBdZmLzY8TkCxlo2No6xf98+gvX48ZBoKpomVd3o8v5c9PaLIuoZjmqq0mCKAU0dWxd5R+EcwesJlPMIV/rMi0RiVPr9PqPhCOcc2kBZrWKtZWM4prDVbJagEz3mGkwX87w71x/EpI7++PFN+v0+VVUxHtWURTKaT2cGUETagZNHiOzbt879D9xLr98j1vWi4NHOpXSeU6GNoT2b/Z/v6o2mknOpZXaDW2Vz6LHVKn1XsLG1SYipUk1dw/r6KpN6bwt7ripRI2wcH1JWjrpRYnCphF4xIMRugOEWDVEiaEwDKPZu/9tVgRC1rT0150M/9YqYK7csl2PGxO0hPt3z2VVJhG1ixBqZeliIc4gRSldSB8+48RiTZtxNWyJ3Utc401aVmRceOi8QEWoNRCKpKqZJ97VNSS3FGULdINZiEbx6JApqFImpxLZvK3toiIi1OGMRK63HiqdQQ0TbKgA2VTgykgJMTSSGeMXanwC+8RSuouy5lCQ2Ql2niYvBYEDTNFRVhRI4fvw46+vrOGc4duwI6+treB/SmKaNRbBtFRaNStCAM+6C3r8STWv/2DaZbZyrrgFEJYiAOowr6FcFTYgc3xzRXxnQ+DwBuKtYmEwa1tcOct/9h3G2z6RuqMqSKJFAtSQgSBKl2+dVwKU8Qrty/6RcSOv7r+Lw0Q2CkqoNCXO2T1wQzGeCh2ntujOwH5b6HiHyhMdfN/uIJbExhIBI8oaDlEfpD//wD89dAAHY2NjgwIEDCwd0InolrK/1eeDBEU6S0SZqUo3fC5AEdUHhTclGaHzE2JK66bIjD5japtJeEPL6oq41JfjCpOyRYiyoUIeAMZYobcUBBW0FD9os6lHbMnJh0lZKMGhopkmpRBSNnsKkiuuxSQ96J3aaUahyJTkF9qUhguxU0Dq9apg0StlbZTxJXgLW9vEeRKo0I96+b1fasVqKoiIqRG/AuOTBZFkQP0xuZicd/NtiVqI8nUvFuqqtBmPbttANTOLc0GUWu78b1985Q117qqqf+pPW/dz75ZK1OzcA64RJncSfbruNzTG9/hp1owgFi87NZiFPg9Gz779m+2foAmyizgQmJSImPWsntVKUK6CGyQSq3t4XP6azpDHgqhIIhBgRSd4PKZrStUlBZGpTpPPXhcbo3rUfNFng6TeZO5a5pqGLVUtkWcjdy/kD5vd/aRY1SpvdIYaURF9CGn7YZHhH6jbSOyUjtTbNwmvwhJjcnFNOo7D0nGIqMsUuw5GJ09j+zsMoqEFjJJkrzfS+FCPTNUBU1wobaQCcBvGKiMEWZfqqtgqNyqy6lmrynpPd7j93cQ0RZ5OXUwhd8Vhm3hvEaYhrqBtWen1UY/LaG6zSeG2fDyCmDY/T9PkiBmMiqhf4OMTMvCHb+1Xnesu0H6lxhQjjukERirKi8SEbILv9CFKlKMs05nE9UCiKivEYjB3MPJwlTvvqrgz7VEHB7OL9kyomOVdixaSwdEkebOrD9Pm6bP8u+HHIfA4mFiqPaWhDC60wmdT0qx4hNHjvufGhVyfRcSkKvvMGsXN/8N5TVRW33nrrafVYpxRAHnzwQR760BsR0e3ih5qUf6P9iyvgsY99NF/80p9hTC/FxoV4AStwxAWdaL7mcYrflamLvelmddqZjry+iGviNLnPdN2q6dOsxzI/SDZLv8fpnOiyWT9fK9PoLOZs++A7ZhHk0h4iTwe43cz3otu1TEvQ7V57ns/MLTsa1JkTDUCWqzDIghA2O7+xfR7IQn8TkVmb2IXrf+4DwOU0Zl37sa0PhmxLgobO3pW84+xZPn/nYnHb66HT89z1i2m01O3Hzvu/l5nPL5OGqbogxJq5y2IWnkV73n5gFmrY+f6Yk2Z3nZPeLpfn20mPY9HzbD43Trp3upT6BkNo7Y+lte5kk84GBXEp99X0PmRpMNvarfNr0a5PNAvPhTNbXwr95261f5nNLGuXW2mH9rDgGcXieEftXD+1G8ehc+1l7rpO25VZqPakSzmkMpemrbtj/9yNedTN2uButr+F/tFMS4efmjn7TZZeX3oeW2un1Vz6/YroU6jmykqflZXBDqr8Il31F+ccd9xxx2lfiVMKIB/72Me4+eYnn9K4i5rCYJ7z7Gfxx3/8HqR96qum3A16jvaTypIdqSfqyjq1f861c2kuMa93Yx23dzan/bhQDOmGEOmGT/OfP7uZROO279ecl2H37c8TzMB1huOysLDwwFx4zy61Xzlx/PhOyvfsyZgFksXruFTvfSmb/WInGRfeG3fx+ptTdmB60oGX0eVtFsWQ+Vw4J+rrz+X4zU7nV+IOz2hziomGPdnwUkLiuRMcdxQ+lgb/c9cvbutr9tp68dovXvm4rR3PBk7mMhFCdq6z1J2J2LppSysELgu23XYyJ98ur090vpP4UJ1GVNHJ+hjdVhbzjI5dr1z709BNgM0L2XH7UG3+GT8v2qvZ/ry42P0PJ5jAO+F9mSf7LrU+SObt2y5FxNQ+XrJ1pLvv7SVw/8RW8IgnLEdjTtbkdNbHLjgkMFcc0iStIPpAUZVsDLcwFgrb44brrjn1E26uLMyXvvSl074qp+zV/uRP/mTuw+MOBy8LD5Sbb34CzqWTZcVgkBOU0T0X4qy02Vw1mpyA8NLnRANFOQ3DQLoZmW0D0lRtY9HAiJeJ8X4FD5bVZA+Ly+fO37m7WXiGz+7VdC93NRQuxQHUmb59vr86jfOycG7Oc395ws81l5nwsTwg0B2OV66AwUPcYTGnd9xX6jN47phTK4nt5MqJ1ydeNNk9aufu950Wt/Prasg1TM+t9Z9cJNDc9jMXZdxj9DT65fmKp5dAG4yyg32gJxLYd3punfrZpaqEmCa5Y/RT0Xc43OLrv/7rTn1+23iYyWTCV7/61dM+tlN6gLz73e+ee0jI4rNjh+O65lpYXVvh2LGAOJdOQpQLb1R05TQXLtCs7nDsGlYXS53XF3XdzYDKSUSQU89wyAkfBrMEivM3abyMjdq9ZoSYHR+X05m4bfW/d5ij3PX2eybtSC7zQdXZGQCze33pBQnbzl33rJjNhrC7z7GFa6pn1Dd1nmmpalncNqgxO5nr0xnJQBQ5x+PXsxBydDqA2/vy22JIx+x8mqmxtdNzZ5tQtGf731Nd/6UJpKXs/JfX80t3eD7tYFe0buenkw9W5VTDbT/nd3M6w3M9j1KkXhrPz11cx6VzOs2ptByeua3ds/gc2LXjON3nd7ZzL0naxOOCacMP40mvm0G2WwW7ZvfM7e0Jcn2crOedHp/O/74YAi2qiEJZVXjv6ZclUT2TyRZPe/rNS9Vldh5JQPIE+eQnP8mZ7t0Jufvuu+XIkSOnNcTpzsn1119P1JRFm3g+a2THJcMs7mCgzWek1amiZha8RPL6Yq6NKjJ1ATuxJmh0pxurfXComYuJm1+WjZCItuUd1QRUdKGaRGZvyie7245PYFyc1ux8106v9Bm8ne767lntZ8t0xlQXZk2mISS7ev3Pj0i/XWDYqX9LOY+6JQ1gz/E4xCehqVtmKRmX2vdOr+1t43O5LZrpzNqZzLDv8f64u/472U87tnO5jJ5fy219+c+tfTG3SCyQWHBCz4wuKaGcTmjU8jmPtKnfty8Sts0In3tybeWKtkOXxwjb8r1wad//Z9TPZi65LmjOQ13Q9uf5+33OBppe8zj1ANn18ety+1oux6xLz8CuTz1N75UuIbGIUDfjNieIsro64Kab9p16hNBqDEVR8N73vve0r8tpFYf+9Kc/zTd+07NPqoaLyFSkfOQjH8FnPnsnsT0gRC6sGTXnQi3LSu40UZCmRqRmUfnN6wu/nnsInPpZHXdW5dQuiR3bB6ORFOsed+jcTHZlvAQGwDMvn25GUrYN/szSsCvOtt/l9jt3FG0jjDuYSIv5SqZJB3MDWJKF0vM6tqLlsmuozAmcOv+eXbv+3WAwnkMfNXsGLs/vyjYbw7eFcLv+y7Vxtmd5HPiT2sqyULL5VO16bxqgCzKmpvjrKDvnqDIs3/u6p/th6a6/Lvani0kS4+nOi+09CV1OYFtoNwO5OCOJmlliwlY8OtFEyg7ZJBbWZm7gc7J3zvYpLtk683mEztFOvgLtzziXUL/zxOtm1BfPqdnZGu08qHbR/pBtN+vJcjXJNmM7TwFeOn2QSvpFumfLdCJoltVWumf0XKLb3bJ/4lz1IdEz6yniNKl4ZKfk9/MCyLSiTIhE9agqNz3iEaz0T8OeNIYYYXNzgw984AOnrQKeVk/3+c9/fk752T6hl34UbOvZc+MN1+BcKlEq1mCsO7eWoydL6LVD1uYFwyXOJcZku/Kb1xd+PV++i1Nnpt5mZ+yYBFN2HD7noJfLEd3d+1fiqY3WE3cB+fKddTc0Szy468+xc72WpyvAyswrRHfqz856/81pttPLsL0ui5jK0gzWyVxrdfefP+e8jifpZPVsTMK9KIGc/nNoKnzMXk8ihGxbUiiVnFD8OPv9mX3GeXFgvYLtz2kCWz35eT7pedvt87hj+8l57vZUNyRnd9/KfJTDrtw/XZdhFvfjtPOIxR3608W8b8YVuKIkisEUDlXFiPLIRzz0pGkT5mmahnvvvfeMTu9pKRO33voxfuiHfoimaXCFI2qqwStA0wSKwqYuICgW4eabvw77m57jw5rBvqvZ3JrQc+eYB0TNiQdFc3XelVkp3G1PvewFsPtDWJmF0+tJjZDlh8d8jPZpxPDOlY7LXAJD3BNU0dBTGt6X1v27MBMunCIL+/yMo+YnwPLzXOY7NnPi83xJXPtz67u68t7b24zuYByZudQL7XvP9fhlh3tMr5zWuXjftgaYzOcM01OLAHu1L5G54z+ZbbRj24yXRQs5lR2gc8LjYk6I+XtzJ7vjbJ8g5uT36o7Xg3N6Bl3R9seJzvsJz7Nuv767eA71hArOXLhEnvK79Puf5XHPdOxqFk1dzMJkyG63P3M6ttAO99C0ytb0c7pJreQR0lWX8Vg0RkSVqjdgvHWYA+t9HveYh6U6OHKKZyZQVQV/8id/cv4FkNtuuw1ItXqBVOK2degzXTIxH7Ctp8ejbnoIIoGyLBkNJ/T7fYIfn4c4xvyAv1yGQXrOn3DqtpDFj8yFbczm3NtqPn/5WE+yvZ6vZpSfhVfuvXm6117NlX2edjhWlV28Xvmezef5IljUmV3udU7a/vQyOeL5B+l8QRKIWFDFuAI8xOgpioLh8DgPf+j1s4iukzyLY4zEGLnjjjvOaO9O685/+9vfLjFGjDGEGJD57PVt+Zn5RKf798PVV1+NKww+1LnVZzKZTCaTyWQymUwmkwFm+oGIEkLAiiF4z6Mf/cjTer8xBmPMGVWAgTMI+Pzc5z634+vSCiCmKPB+MvUkfcrNTwIig6qimVxg749MJpPJZDKZTCaTyWQylxYLldcSRlMRldCkxKdIJMSG6667hquvNqctUhhj+MxnPnNGu3PaAsj73vteVBVr7La/+eChVW5EUjjtk570RFQDRWnRabmvTCaTyWQymUwmk8lkMlcyzhmaZoJISrXhfc1Tn3oz1kA8jRy/MUbG4zF33HHHGQUtnrYA8t73vhc0le5Z3iFt3T5cUaQPFXjIjdcjGqjHI4zJtakzmUwmk8lkMplMJpO53Iknq/7Z5Ws0BhFBRDAmLd/wrGegbcGVU2GM4QMf+MAZ79tpCyBdIlQAbRWQ2JaXdc6hMWKtTBOV3HDDNRirRG2wkr0/MplMJpPJZDKZTCaTyZDyftgULRJDQ78qeMyjH7FQPfRUvO51rzvj7z1tAeTw4cNIW/Gl8/gIXokRBEm5QNrXGw/XXiNUhaPfc0yzvmYymUwmk8lkMplMJpO5ApnJD9F7nIHoA97XFKWwvr6KCKclfgC89a1vPYc9OAVfvP0OaZoGgG5tC4uPIW2gOt3RooC6hpd953cQmhGoz9c6k8lkMplMJpPJZDKZK5hO2yiKAlVFRKmcJTQNVx+yWIHgk/PEZDKZvi+EMHXEABgOh3z1q18941wbZ1QA+61vfSso9Pq96WvJbQW67KchNGiAsoInPvHrKQqHEUU0e4BkMplMJpPJZDKZTCZzJSNtFZgYA2ggxIYnPPHx7R/B2jbHqHOz97T5QiBFpGxtbZ3Vd5+RAPLrv/7raIy030tUMGKmuUAQQVVT4hKFJz/p8QgemyvAZDKZTCaTyWQymUwmkwFUk0eHtQYReOELX4CzEEPyCoGItbMKtJ340f38qU996qy+94wEkN9/wx9KCIG6TiEtXXGXEEKbGDXinME5QODAATDE9gAymUwmk8lkMplMJpPJXOmEEHBWcM5RlY7nftOzQJIwwpwDRWwLsIgIIYTp6294wxvO6nvNmb7Be49ZqktjrUWMLuyor8FZuO6aa9DocxLUTCaTyWQymUwmk8lkMsSYPDxCbOj1ehw6lF63Zvt2yz+rKm984xvP6nvPWAD56le/inPpbUoqhWvEAIL3nk4EMTalBXn2c74hxfZkMplMJpPJZDKZTCaTueLpQlqapuG6666ZvW4Adnae6Bwx7rnnHu644w45m+89YwHkne985447DhC78rgxYG368Be+8AXbPEYymUwmk8lkMplMJpPJXJkYhBhTuownPOEJhADJwUOBQCeCdFqC9z4VYFHlM5/5zDl87xnyvve9D40QgmIBQVAUMBjTJv+ICiGVuPn6xx+kLCIpDcipRBqZWzKZTCaTyWQymUwmk8nsJUybGUN0Nq5XMai08SKiKBFioF84nvC4x1JYaANNUNVpoZV5AQRSGMxXvvKVs9+3M33Db7zmN8UYECJNXSepIgoxCs5WNI3inEuxOwoFcNPDr6EoHHXtKcsKYwy+aXBW8M0EZy2igqhpdymF1KSPjsScPySTyWQymUwmk8lkMpk9gY1pAYgYgrSLAZVI6QTCGKuB5z33CVgFXyflRMQi2LbCbHqt1+vRNA3WWt71rned9X6dVWzKpz/9SawViiKVpTHGtd4fnRcIKYNrWzL3ITdci/qANQUxRmKMdFExnRvLFN1hl7IAkslkMplMJpPJZDKZzJ5gvhCsSgpomR/VR1/jDDz2kTfhbHJ/KAoheE/nDCEiC9VfjDGEEHjve9971vt1VgLIJz7xiTZeZzFUJdXxFTo9o/v7zTc/GR9qjAUNEW0zvqbti4VyNlnsyGQymUwmk8lkMplM5nIkSSHeN5Sl43GPewzOdVrALNQFFkvgduu7776bL37xi2edM+OsBJDbbruNyWQy/b3z4JiVpUmvW5v26+lPf3r6OYYU60OcqjfGGDTudFJO4RWSyWQymUwmk8lkMplM5pKnG9F3niHGgjHw6Mc8kjawhPG4pur1AN3m+QEQQuBDH/rQedmPM+I973kP/f4KkESMBQ8OYLlszSMefoirrzqI4hFRjDELByQiKSGKxG3vNWe7k5lMJpPJZDKZTCaTyWQuPqKt2iG0g/0FvaAoLKPhBk/4+sdNa6A40yohpLVzbs7JQrHW8ra3ve2cduustIV3vvPd0okX3Y4A03XnEaIK3qd8H1//9Y9D8AgeiKgqIpYYdFsoTSaTyWQymUwmk8lkMpm9ybze0f049QJB0egpSssjH7nalr8FV1iauWovnU4QQkBEMMbw0Y9+9Jz266ydK97znveknXQueXBMRY95QUPpPDqe8fQnYl1ENaLRo5o8QSKKyvxuKNvCX3IITCaTyWQymUwmk8lkMnubNuojRs+zvuHp2Db/R2hTf/g2Cep8tEinL4zHY2655ZZz8p44a2XhD/7gD6jrOu1sCFPXlFThxUx31lqLAo961CMorEKsCcFDVKxxqGbvj0wmk8lkMplMJpPJZC4ndOl3UTCaPEJKJ7zgBc8DwBoQA8F7+r2UaqMoqqmmYIzBe8/tt99+zvt01gLIxz/+cZxz0xCY5TCY6UGKIAI3PeIq0Bq0oSosVVUwHA6TF0iMRGFuidNqMAJIFkkymUwmk8lkMplMJpPZE0wmo4UUGdYWGGPa8b/HGHjUTQ8lNO2gH8VOy8GwEGECKfLkF37hF855v85aALn99tun4sV8qZoTfcn6GjzuMY9ENNJMxkgbAkMQrCnaMJdZyMzsyLP4kclkMplMJpPJZDKZzF6hLEuiegI6rQDbNA0xNFgUjQ3XXX+Qopwf+yexI4QwFT66qJOtrS0+8YlPnPN+nbUAcuedd8qdd96JtRY3p9TMiHMLIPCCb/sWytKgqnjvp/lDFpjz/shkMplMJpPJZDKZTCaztyiqcuooYcRB1KmoYZ3wsIc/hH3ryfkjBsVYIcQGMIid6QudXiAifPCDHzxn74hzyi76X//rfz3tbUOE533rN2Es9PrpZFhbYK2laZqlrbeXw53ljs1kMplMJpPJZDKZTCZzSSKp6ut8JReAyqUwGFHlL/6FlyECo1GD0rSOEJLeMzf2d84RQuD48ePnZdfOSQB5zWt+a3pgM3YSL9JL119fMpmMsVam4TMpCapJJ0mW3ps9QTKZTCaTyWQymUwmk9lTdOJHJ2CAwbkC2tdf+p3fiAF6/QLrDBCx1rbhL3OSQFsF5g/+4A/Oy36dkwBy++1fkMlksj2MZQcKl0rbXHvNVYQQpicihIAplkNo9HzuZiaTyWQymUwmk8lkMpmLhLY5P0VmJW1jjMQYOXRgH/0SmgY6KWEyHgO0hVbAGMN4PJ5WgfnlX/7l87Jf56ws3HHHHaclgAgQAnz3d78M0bhjXd8LtIuZTCaTyWQymUwmk8lkdgknBl83lNbx6Ec/kqgQQlcERTFGgNgKHky1AoAHHniAT3ziE+clJ8Y5qwuf+tSnpmLGyVCgX8H3fe+LMQaayRgrYJxNoTCYdncM8/k+RBfXmUwmk8nsLWSxX9thWdy8TQYufu7nOPvbxVzveDhx+36dZJuunP38clqfu9P3nMP+y1muM5lMJpPJnCHaen5oyulhRXA2VX6pSsMjb3oYxkCv3yU+jRRlSfCR+ewaZZlKxHzuc587b7t2zgLIK1/5yml938lkMidiGLa2RoBBY8QCFqgKuP66qxj0CpSI93X7fgMqKA7Utb+nWCCIWQHJZDKZzJ4UP0RNu0grBoBonFvmxZCuz2uIpgFJRoFoRHZjTVwSa2b7Peuf49IRd+/3aR0FG83ceZg7F/PCxzSHmE+LNNB9xjnt/7mdAyQSFxbaRU6xpO0ymUwmk7kSGdcB5wrquiaGtJTWM9p8kO/5C9+Bbe0Aa1otAMHaEiNzThCtB8if//mfXzoCyC233CJ33XUXAFVVAUw9QrrfmRp1YA08+lGPoG6GWAmImU+i2oog0okfmUwmk8nsbVRY6NN0QTAws41UWm9IWtlhlhZcxezKuuuXWfLamP3a9dczX5blzaeftEO/rrBDf28WviR93tnuf/d5Z7tkMplMJpM5U6Kk0rdGHIOqR1EUhGZEYZVH3vRQDh0slq2BbUwmqVKs9563vOUt523f3Pn4kDvuuINrr7126gmSStxanHPUdd26rkRUwTl45jOfydv+5M9wVR9BaaJHbDY0MplMJnPZyR+p/zMpvpUTiR/zg+2kmIDGVgRJAkPnE3kx14tCQGxFDkOc2wKRJRFDAItpU7gHYe64Y3uIOjWQ2PY9y8h0izPf/x0ElzNaW+I2w6zd6RNN1OQKdplMJpPJICJ47zHOYMXQqBKj5yk3P5k5P4kTUpZJJDly5AhvetObzptP5XkRQD7ykY/wzGc+E2AqfEDK/LpgeGiyG26++cn0+hVNiIgoBgU10xmj+WiX7rUcAJPJZDKZvUgUbfuyOO3fTKtxyJwQohhQiGJAI9q6fYqamXSgF3mNWRISDFE6eWFZCDBzv8ep94qRVvwhLJ2XJUFhp8+aE03Obv9nnzS1J85w3SVny2QymUwmc3oYTaGiTdOgQekVlqoqqcdDbr75yXgPtjqVgAJ1XZ/X8JdlC+OsecMb3kBRFBhjUFWstXjvEZFp4pJOATHAwx/a42EPuTGFxcSAiE4NQ1k2hlKwdG5FmUwmk9mD4kcKeenyR6hEVNpsF5KG1TtEmLTds2sXs4vLCc2Sba90x6ECikuLmDYXRty2zHJ+6A7HbkDteTmGeE7L/PG2i3ZhP4vL9CddXDKZTCaTuSJFEGOwTlCNGAPOKkLg677usdipG0bkRGnHVZMXyTve8Y7zul/nxQPkne98p0wmE62qKrm5tPV+O7z3ODd7zVi4+alP5Et3/k8iDUZKlNi6+WYymUwmc9lIIFMRZLuWH1OfPxX/F30uos7kANi9sIpZb552yChLYR7aHWn768yLQ7u/dh4w7ZZJGJgPrZlPfN69KXnCzBKknsNVkG426szXmUwmk8lkzpwQAqV1BA2E2OD9kIc89DquvbqclrldtDbM0vsjRVHwwQ9+8NITQADe/va3853f+Z1Ym8radvlAIKk/kNxYOp75DU/jjW/+n8QmgI2tyReJahYMq2myuJwUNZPJZDJ7ENOGY3RJQEWZS3WaRBDRmdLQhbyIdNvvck6JbVVeaKdl4uwFNUjrHZESmSfvjdgdl85EEIO2h9osfW4Kq1EMSOdPEc/5+FVmATtns56JM1M5JTfqTCaTyWROZT6oTr04RBuCn/DCb38+vT4nrjXf2QyAc4bNzU3e+c53ntfpiPMmgLzlLW/hRS960TQUBpLnh6pSFAUaPWIMqtA08KhH3URhlbppECxQzpxgp7NA2Xc0k8lkMnta/kiD/iXxY1ZiNWDUw9RDRNvZgrlB964m1TzxdxvVNpRlJoKkpKEOsCglVqXzG5l+XleW1ky9O3TOOyYdu2Jmr53j8SvnkgzWoO3xMHccitnh3JgsjmQymUwm02KtICgaA7YUTGn4tuc/r9UJoHI72UwzcSTGyK233nre9+u8CSC33HLLNPlpCIEYI2VZoqqEELDWMJlMKKuKsoDrryt46EOu57Nf+CreN1g3b2BsdzvN/h+ZTCaT2XPyh0ITlH6/z2hzi0G/z3Bzg9VBj8l4yMu/97sgThACQhJCZt4G7rSEiIsngpiFgb4h5fIwBprgcbZH40Gl5M6v3MNHPvJJxPbwAfr9PlEDzgrDzSHPftZTeeiN14E2iJn38pgJDHIejr0rQaytl81O66/c8WWMs0Qf2NjaZLi5xbieMNoasjWcYMoBYitCUDQKRVGCgPcpqet4XLO+vs54UqOqrK6u0jQNw+GQXq9HCD7fCJlMJpO54nDGUjcTQpiwVpQc3dzkqqvXk+erJIHDGAjeY52jaRqcc231mIhzhle96lWXrgDy4Q9/WEREm6ahKFLJmq4EbgqHiVRVBQo+QOnghd/+Ar54x2tx4mjtCIwsmTrzMz85DCaTyWQyewxVxSI447BYKlcRmprrrrmKH3vF97C6mvo+gQXXzwUnyPkwkou9XjiYLgfIbL+atpt2Req/x2N44xs/zJdv/xJHjo7plauIWJrxiEDNoFfx3G98Dt/5kmeyf72t9nsy59Zz2X92SjC7JO9EprHI0hplIcDhw4H7HniQn3/Vq7nts18EVVZWVzl67BghCP2VNZrGc911V3H48FEmk4aV1VWOHz9OXdesra3RNA3G5EQimUwmk7ny8N5TWIPFsLW1yROf8Hh6/TaluEaMnfecNKj6ac5Qaw3j8ZhPf/rTl64AAvDWt76Vb//2b5/+XhQFo9GIfr/fmRkgZloO91u+5Rv5lV/5/yh6A3yIdB4g86dCT2EXZTKZTCZzKVMYhyp4HxnFEaWzjEebPHDfkPU1sG3Ei5nTOi6lCNCFEvW61DELdCm/fEg/DyqIfsTW5jGQkrquKXBpEiQqVhrW1/rsX0+9vpidO/qF7z2X/T/F34OCXZpfsRauvcqyvnYNr/6Vf07dwCc/eQ9veuObef8HPszm5hijEwyRu++6E1f0uOaaqzh6/BgiwoGrDrKxsUmv1yM2Tb4JMplMJnPFEWPEGsFaS103vPSlL6LXa+UO22oDRIxxaIzTNBpd8vTxeMxnPvOZ8y4FnFcB5Od+7ud4wQteMC2DWxRFWwZ30X21KNKBHbqqT39QMRyPQaq5Wa9uKiaclvGSyWQymcylSgiepjFYK8QQqaoBVbmPrY17CWleABNBpU49ngJY0EsjBGaWjNwtCiEA4sEqUSO+CRR2AJISlyXPzwKtLSE2iIDGwNZog8l4CwHGE6jKTuSIM6/PNmzlfBgBKif3HnVzp9n7SIwR5xzGwKCfhJ1eAc946nU84+a/zee+8P285rW/w7ve/T4aVR5643Xce/8Rjh07gitKmjowmUyoqorRaETlXL4JMplMJnPFYcUQgqcwkbWVim9/0fOxkqrcOQE0pc0wtmi3Tx2/Dx5nHe9+97svyH6d15iSt7/97eJ9cl1xzlHXNdZaJpNJMgK9n35p8FAW8LSnPpnxaLN1qdWFMBcVkjEkOalYJpPJZPYmpnAEIv2VAbZwbG1tEWOk16uwNoVfGANGBCMRaYugLCx2t5Y43T+ZL84yt6h6jECvV876b1WaJjAe19jC0QRPUI8pDVKkz1WBXq/9bBvTMne80+89x2MwcvIlZUlVQuMRjZSFS69HpZ5MKAzEWokNRA+PvmmNf/HPfpj/6//6B+xf73PHlz7PddccZDLeoq7H9PolW1tbFGU5nc3KZDKZTOZKI8aIiKIEDh06yIED07+kRZK9kAwHg4i01WQFiPzcz/3cpS+AAGxubgKp3E1RFKhqcnuli7GNbTkcKEt46Uu/o80Zkg5edHsC1O40ZTKZTCazpzp/gbJX0ISaia8xzjDxEza2Nqh905kABJRIQNt/qM7CTTQZD7uzXvzHfBhMu28pXjcSidSNpwmAsfT6A/qDFSJKIIBRispRlg4fG0KAcdP2/qoLi6qiKu1y9vufYm7DyReJYBRbGGxhZhMvRimLlMOsrJSySBM3hWvtl5c8lf/yK7/EDdddzX333cPq2oCisIQQWF9fZ2tri7KqdrRpMplMJpO5nBFSYZTCWqy1POYxj5qm5zICQZNjhLWzIija9v+CcPjIYd7znvdckB70vAsgf/AHf0AIIe28yFQQQRWZxvXEaczw0572dVxz7VUnNh5z+8lkMpnM3pVAGNUjgokMmxHRCiv71uivrlD2esmDos1+qljoFinaLGGtpUDrzrAba7rvl9k+tXlLMJYYkjmhGIrCYSyIcTRB2ByOoTBIYWliw3C8lYQfAXFQlN3nyZzbRtGuZfb62e7/aSwxduZQWuraM5k0qEryeKEBP2a0cYRQ19NDbybw8IdZfv03/gu9yjEcbrK6usrm5iYhpBDe0WiUb4FMJpPJXJFYa4kxUhSW5z3veYR2bgKSZ0jnFaExojFlRO88Jy9E+dsLJoC86lWvwlqLiEyzoNe1X/i6LiG6Blhbh+uvPQSSXE+RmMJh5ncxV3/JZDKZzB6laRr2r++jchWT0ZimaZjUIyZ+TIydM8XSPw0LThGopNe6v009RsLUc2Rx3XlOzK9B1RBJv29/v6Jqlr4vtFt027DNA8RYIaL42EwnLbpwWGsto9EIMYoxBo2CMS4ZRaQ44Nl3d14fSw4h0/1dXi+eB6UBAqphzosknHIxhunPAGVZUlVV8myJ7WyNUfprq9jSMR57QoB+L52D666Bf/Uv/gnXXrXGvXffwcEDazTjCUVRURQVOZV7JpPJZK40FKh6Jc1kTGEN3/SNT6CwbZG2GLA4QohAsg3AIO2EiA+ej33sY3tHALntttvk9ttvnxoRnfqj0YKadjYkAh5roa7h5psfi8YJVgLOGVQV7yPWlBSuj6pciF3NZDKZTOaCYtTQswNGxyfQRPatrKJ+AqGmV1qapnOoEATT/q+IRMT4tEiaKRFRDBGRiMFPlxjGGDyT8Vb6OzF9BpLyWkw/ufNxkNZfY/YZGn37epuOK0AMNQaPpUaoEdK+TJ0lpjm6PIZAacAQCR6MjSANSKCy4FQw6jAUGHVIsIh25X9Du7+2PQ/tVwgYie2+KtE3GJTJaNxuH0FnxyDdIn56vtLSvRYXzqW05yrUI0QiMTTJPonJcoshGWJgwbjWG8XQ6zmsS/tnLVQGvvW5j+VlL/pGrt1fMdk8TK9wjDdGEO3SpE4mk8lkMnuDKEKUORG/6/eXlrjDgkSaZoISedrTnkxZtXaRQAwBEKwtIBqMsakEfeszYQy87nWv3TsCCMAnP/nJmfqjijGWLr9JWZYISowToKFfwnOe8w1YCfgwYTwaUhQFxjhCUMbjcU4ilslkMpk9iiBR6Jd9iMrm5ibWWvr9Ps452hRZc12yXeqaWy9INaTCbV1ISlIhFIOzDjDtpIMgCk1dA1C4ou2Ml5Zk2gCRoE3ydlDQ2BkfgnVFm9kjfV9Etu8bQghQe0/dhpLYtuhJjBEzF8gqapCYjk/UgMz7RuykEqT9U5JnhnMOVaHf77fnxWGkICoLi8Z2CaDBEFXavykxCjEaNJhpjhFbVgTvk62hAVA0JgPM+/ljbcN4ZfkMQK+CH/6bfw3CiMJC6WD//v3EoOQJnEwmk8lcphIJOyWs6PrJwgqFVZ7+tJtb4QNQT1l0SdMtBPBN6ky7If9nPv9ZPvqRP79g7pMXpFd+85vfDKTEJ9KqRqbzIm2PTESIraX1lKc8ghtuvA5rJW0vkRjTDI73ARGb21cmk8lk9iSTyQRVpSgKrLWUZUldew4fPkrTQIjgFYIavDoaLQixIAZHDIYQ0zYhQIiGEB0hpm28VtSNo/EWY3ooDqRATFIhgj9xJi1tBRQjFSIlCtQNjIbdFg6hT6BKizp8uy8xdIvD2hUKt4ZSzUJgGgEtsEXvnM9f8Drd361hjWoSJupaUAwiFSL92WIqRFzKO2YMImW7TZleF4MY0/7sAId1fRofQSxNqMEoPkQ6/ehkYSyddDMYCD/5kz+JiNI0NUePHaYoHTmbWSaTyWT2IkYVo3MTFLqcmiJNDhhNy9TPtN1mMhkiEnjKzU/E2uQ1KTI3G6MKwrSvbZrUX7773e+5sMd1IT70rW99K9576nYGanrOYpoRUsBKgRVLiMkV5jnPeTZCpNdzhKZGQ0NRFJTWEWM2HjKZTCazF1EGgx51XU/7sqYORDHsP3A1tmjLvraJRbXN+6nmBIssLiJgCosYQ9MmF5vUim0FEOvMSbr/5FUyGtf4CLUHV0I5mNVI8e02Suux0e2nmeUo3Rgq48bgg+nSgkArREzG5+4B4ZybfsbKoAJN+Uutg6DQBGji3DL3u2+XJqTjqwM0Pr3WHWPTJM+RouwxrhtcUaASsVaJ0Z8yhYdIChsSA8/5xqeABkbjLa6+6gDej/MtkMlkMpnL2MzZuY8XjTir3PSIh3LdNYMFS0BjTMZCjDuaCO94xzsu6C67C/Ghd955p3zkIx/Rpz/96YvnR1Kccwi+LXljqOsG6wq+8TnP4rWveT39lT6KpShsm4AtJg8QzUG0mUwmk9l7FNYxHm6k0A0x1OMRRVkQg/LOP/0S/YHD4AE/G2trN4NikivpSZKB9yqLDzWhGfJN3/j1GBGkdTVVjScRQdK3Vb0V7rt/k09/8g5WBodoYkoIGhilmZrWS0KiI4XYtAnNCSAerzWD1RUmtacJlhhLbv/yvSB9jI0puaic/URGjJEYI9Ylk+Xtf/oxev1VxFaEEDAulapd9rTo9hNjU0LUuXq0xhisATGBG667mquu6lFVILbzYvH40FA4x+kkMa0bT1E4VlbgGc94Grd9/qsc2zyOKOQQmEwmk8lcvqLHXN+6NFwvSuFFL34+q2vJi9VaUvL0LtbFpfDbehwxzlCWhnE94UMf+l97TwABeMMb3sCznvUsmiZ5cgRPGxdspkaIAaqywBh49KMeyurqgI3hJtVgH2JKxvWEEAJVVRFTnb1MJpPJZPYUk3qMiOCcQ0SY6BhrCu5/8EF+5md+kc3hESCmSmgzMwKjqYuOC0aGWTRAJNJMhhQusn/fgN/9nV9jdSU5l4YQKcq0zc4CSsrQMWmUt//pe/kv//k1NI0lRMFYAddQliWTOoA6RCXl3ZjujQcCXif0+302h2MaD6vrV1PXhtFWZLC6Rh2H53T+UiSLwwfY3FL+40//LBtbDZOJMmk8ZVmmcySLIkhniEWZueymintJILIigAf17Fsf8I/+jx/jpd/xHJoAENr8KTo9VzsLGen7JpMRRbmGBvg3//Yf8+0v/htMYokr+vkGyGQymcxlipzkL5HN4w/y7Gc/JeXLshCDx1pJfacKMQaMMRSlSWXnBd75znfytTu+ekHLp12waYm3ve1treEyMzqmRp0xqKbEZrbdg/374Zuf+2ycVYRADClm2rkcApPJZDKZvYu1lqpXMh6PGY/HRAVX9tl/4CoOHx3iQ4EPBY2mJYQ+IfRpQo8mVNPXay2ptaCOFXWsaGJFCAUrq1dRluscPzai32tDaJSTe17obFBfFoI1FaMJROkhbhU1fSbesTlSfKjwoSKEKuUd6fY3lmkfYo+Jt6jpI26F8TgSolBWA2p/7pMXIXogYi2srAjDcU3TCGorqsEBVFYQ00fnF1tNlygl0aRFbQ+kQqnwWuK1R9Hbz133HOOVr/xFvvf7XsG4Bmcrxk2DR09q9kkrgqytrbCxsQXA6ipcf8M1lIWhVxbn5P2SyWQymczu6RtxqQ9LPd+sOoxZWESZLkaV9bUBN15/aKqTODuTHprgCaFJn9q+7H3kv//3/37BD+uCCSAf+9jHZDQapRK4qtPkJt77VOpODBojhpT5tbDwbS98HmtrAzRMGI42KZ1h0OsTfchl5DKZTCazJ/G+RkQJMXkrlGXJ5sYWTRNAXEraaXqIDBAZLA3k+4hUYArEVKjptdv3wVao7bE1bIg4fFS6+QIxUFYFGv3O+ofMdJAmgmJxtsIVg5TsVEtEBhi3Ok0sqrYHpkRNhdp2bSqwfXx0IBVF2QcpiBjEWXxr3Jw9ijXtbBEpf4e1BRhLDIa6kWny2KiOQEFUS1SLJy1RbZvwNeUyCWJT4d7o8NFy5MiEAwdvYGMjcPfdx3jJS17Ol7/6AFXRx+BOYSqlmrnDrQ3W11aSXKJw9dUHmExGNH6Sb4BMJpPJXAF0Bogiban6pz3tCaytw2QcptvUflalrihLYgj4EEGSTvBn733/3hVAAF796lenQ20TnypdMrOUC8SaAgGqIv3t5qc8Fl8PiaHh4P79yXhoGoppqZxMJpPJZPYWYg1BI64oqH0DRrCFa8WPsh2wm3Zxc0tBjI6olqCW0JZz7X5OvxuMKxnXDf3BKs3CRI1HOvFA4rzLwgLWkEQCKWhqJapDTIFXR1Azt39tqVlktkZQLDE6FEsIgrYunyF6bGGAc53BiGhrWBkDQSM+pkyoxjiiuFbccEQsKmZa4UYxYCS91iWPTRcFjEVMj97qAY4cb3DFPsZjYTSM/Nqvva7NKmJOYiqlfYohMFgZJFtHkwfOvrU11gZ9mqbJN0Amk8lk9qb90npzTHs9SUuXmV0l9fll6ajHEzQGqsIx3tqkrAz/5P/+R0zGgUHPYkh5yUpXTPvVxjeIBdfmKtvY2OBrX7tTLvRxXVAB5Nd+7dfY2tpqE56e/OQaYNCDhz30BkKo2dw6zkp/gJ/U2YDIZDKZzB61HuLCYHnW9XZLW1IFQ2wH8HH6sxDFtG6m0OUASYP47j0zb47lbCGz105XgDDTfBmRJGgobsHNdV5I6N6TXmuPR00yjiSi4jl/JWDj8oEBMt2PKLNvSsKMObGpI9qKKgZFqCeBldV9DIeeqn8AYwa8770f4oEHPMGfar8U09o4RgylTd+2urqS4pqtXTAeM5lMJpO5XIgx0sRUSq0oHUVhib6mKC3rq332r1f0KzvrnUWQLoxGlcJVeD/raC9G+MsFF0A+8YlPyDTL6wkNmjQrZYB+D777L7yUlX6B+gkhNogIZZk9QDKZTCZzmRkOzGZT4km64/kQ0PnBtE5jcyNxPk53VkqG7T+dQKeZfehMzKCru2vafW2/pxVVOs/ObiYo0oof7Tcm/5ALlcBciUvKgsqicTMfizxbOrtD2+NRxDgm40B/ZZ3GK74R7rn7QT7w/ltwbvHkneo8dvnOVvoDYsjJ2zOZTCazl5mbrOkmbHQ2ESJGERFC9BgDSkNdj7BGedRNj2DQAyMpIWoqfdt91qwoikjyVD1+/Cg///M/u/cFEIBbb7116g67/WTOWRTtJi/+jucyWKno9UuapqYsy/bEZDKZTCZzmXDCsrY6q2bSeiokEcRMhRBR5iqeaBJCztDTYtGnZCrHsKO7wkISND2JDCBLwo2yU3naszfAtn/1vAg0+16m5WdFDaLSLnPnj9iWHgbrhCZ4iqKinniCCiur+/ni7XechvNMm3G2EzsUYoSyTLNaOYl7JpPJZC5XRARjhBAaDEpsGpwVqkJ40bc/f2pfqOqcHtAmTBUhxjhNj3Hfffdx9913X5RB/wUXQP79v//3J/96TcaLCGiA9VW49pqrUI0URcqcOhqNcgvLZDKZzJ5kwdNDZr/vsNVU1EA84DFtlg2ZFyPakrkyL1rILNTFdBqFWsCQZIi4FLoCiyEzKfDGaJx5ScwJIkZny2w/5oUPaUUamW4nJ/VrOROpxiBqZ7HI3QzSUnWVRe0miR5GBRNdWjQ53s4LTAZPjJ5ev+TIkSNUVTVN3q6qNP40TCiZyUhG0tIZhta6fANkMplMZm+y3XBYwHvfVncNGAtKoOpZDhxY43nPfyIhAhpSf2jmw39BxGKMaUNgIh/44J9dtMO64ALIW97yFjl89MiSqbV8cpMOYm2aSHnOc57FeDxGRNokqAUnqzOcyWQymczek0Xmfu48GcS34kcACUBYKivXdtyt4GHoBIvFPnVa5nYnw2UpZ4jMCzDzHhWtB4qZztroCcSPea/O1uOi/UzR8yCCqG2XbQcwE4Ha/UrCy+xcJQ8Q2y4GE9syfXOCk3OW0WgLawVXmDbvWKQoLMadhvmh0oogM7a2Rifx8slkMplM5tJn3v7Y2Y5RRBRjQUxENeB9zSMe/jB6vc5e6ba3rdcHbcU6QXWWOP3Xf/3XLx8BBOD973//CYWPjhCadmYGXvjCF6RyctFjRCjzDEomk8lkLktmIS+GkOqWaArPMPiZ2KDSDuZNK0os5gZJIoXs0MeaE4zgO4+ReS8QTZ4lGrcJK9IJCyeaylCDxHY5qcF0NmaKmdkL80LQ/PG35yQ5wxiMzmaajLp2MUtnovOe8YRYM1ipqOsRSkN/UPLYxz2a047AVUVbo04Ujh05CjEiMU/eZDKZTGavs3NFNOccSgAU72uMgaapee43PxsFjN3B5GmtEmmdS1xR8JU7v8I73v5uuZhHc8F59zvftf3LdNENxpUWBGKAR960ysNuvIHSOgrr2NoYLqolEhfXmUwmk8lckvqGWVx3gsd8Xo3pNtv7/kURIe7Ydcvc/0svbn9ZFzeZL1I7LRMrO2y/rGYsezfM9cuzzzi/JoYun6I2GVsSSByibruR1m3TJmlNH9DGH2sSa0ZbQ669+hq2tjYIsWFtpUfUmuc/72lomAYW7XB655LOGkvU2T5uDcdom9gtk8lkMpk9acJM+/NlL9HWGjGGUAdCiIQQ6FWOqlS+9XnPxLZl5zUqMQQ0Mi0V300uCELdeN71zvdcdDnngvOzP/uzsjU6CjQ0zSQlBfPJmlGfTqLGFP8TNWIEXvKSb2My3sQi9MpqFrcrcxdgwW24zSVC5/4qC0smk8lkMrtByo1h5vJo6NzS/d0BDrRIa9JgvjM+olFiO37v/EWmEsaSGBEXxI82iSpmMfJlLhN7eo8hiBDaMrfLgSvTtCLbvm+W6FSNRyWi0laIacvjnpsEEAmxQW06rq0xVP0eTQw0MabzFCvQCtUCtGi/sysRHJHSEK2CMagtERwxWGIUjDoGvRWGxzYoBNYHFX4y5G/9jR+g58DZ7bqS4IEG2lmvFAJjMRaigTrCV++9GxUI5Bq4mUwmk9mrAoimxczCcmdV3oAgoBaJBieOyWiDb3zO09i/n7mJGAti06p16IzTmRelKkre/Oa3Xn4CCMBHP/oRNofHKYoCY0yy8zyIdcl1tM0W3yvSLj35iU9AfcN4tMXOlXTj3DrPsGQymUxmL1sZ5oRLV+kkJVGNzOfbmA7Ll0SQ09L9l4WTVnBZMA1Oy9NSl5bZq+cjD0Yk4INHgV4fNofHcc7iCiFGnwwxjdPqLqIpl4ohgHjGkw0mfkTTjGmaCSF4DGDFYIiMtjZYX+ujfkQ92uJhD72Wv/EDf2HBl0QXbI/F6jZRdfqKCmxuwR1fvpP1A/tPlT8uk8lkMplLG5kbc89VoAMIIVC6grLsMehVIJHnPf+bqYq54mgqCxVdRZiO7UWEw4cP87u/+7sXtae8aALI7//+H1AU1cJrTUxnRkPAGJmeTO/hsY87xE2PfBjOgQa/aCCy07J8sc5bAHImk8lkMpldorAOawURqAqo6w2EMVZqYIgx47mlxpoaKzXGTDBmTFUFVlcMKytCWTRYM6FwNc5OgCHr647jx+6mrISbn/Z4/uuv/TIKTOpm5lU6FT+2C0IiltpD04bLvPKnf4VDh67ia1+7e8Hoy2QymUxmT7M0qWGtJcaI9zWNn2CMcPPNN7d/60JeZFtfOCuJC+9+97sv+mFcNAHkzW9+M1VR0TQN3TG71rdU2mCg4WgIRMSAc/Cy7/oO1lZLotbbT/62uOpMJpPJZDKXocWVksMCk7GytlrgXIPqFoWZgG4CxxHdQOImErcQ3ULiEIkjhlv34+vDaNxEwwYhbAJDSttQFoH1FcNf/+vfyy/8wr/nv/znf86gD70SepUlxDHbShQv5GJJ9otrw2XuvRc+/OFbOXz0OFddfS2+yR6qmUwmk7ncSIKGM0LwDRoa6nrMzU95EmurjhBniU7nxQ9VJcY4LTU/mUz45V/+5Yu+9xetvMrnP/Ml+fRtt+kTHv9EvI84Y1Jm2BjBJu+PsiwBsCbNpHzbC7+FX/7Pv0bhBjREwG4XQpZcf7f5fGQvkEwmk8lk9iw+eJwtkn3g4Af++l9iXEdidFhXEYK23X2qkCPTvCcpP8fqvv0cOHiQffvXKApHr1dy4MAB9u9fpaqgdBBqKAuoxzAYgMbI1ugYK4OSFPPcsSR+4FCd5WZ517v/jEkdcK7H1uaYqj/IIkgmk8lk9qjOMd9/bXc68N5jrVA4gzUF3/d9L6eq0lh+eQTeiR7zHiFVVfH2t7/9ortKXtT6sj/90z/Nr/+312zP6SEQvMc5RyQgpMzs114DVZkSn/k6omK3v3FhncWOTCaTyWQuJ6x1bU8fqZzhb/zgX8U6S1Awst006+Y9TCtXjCbQ63UGGDRNij92brZ9IH1WWSZTwpjI6sqAncNe5hPQJvGjnsC4gVf9wn9iZf1a/Khha3NCUa6Q85RlMplM5vIh1Y8zClE9ZSn48YR9B3s88xmPp0iFXfG1x5VuQfwwSyLAhz70oV07govGb/z310pdpxrBCPimmTp1dLFAMSZDocsS+73f+50YmlThRWnL1s2WxUOYr90Xt5cazGQymUwms6eQVmxomobGT6icxQCFsJSjg7lKO6nrNworvdk2ncjh3Oy10bCmX8Jwa4K1IBIZbm0yGQ5PYDbNSummRK/p11f94m8wmgSObmxSFj327z9AE5RttYgzmUwmk9kL/W+XUnMhD2fqW7uEqNaBD2Me+7hHsbIye29RnNzPomkafuInfuLyF0AA7r333unPoU0PqzHiioKoEWuSYSMKzsCPvOL7MeIx02Qp8RSHkQ2NTCaTyWQuH5LhVRQVpStQAhobUI/D44g4IpaIkS7mmC49RwqJ0ZrgRwih3TbNvxhVVldKfPCsrKQM9ho9g5UVqqqPxuWE6ynsBTWoJoskRPjN17yVN//RHzNYWaffW2U09owmTb50mUwmk7mcJJHpWFsUNDYY9Qx6Bc959jOw7cTEeFgveGguU9c1W1tbvOtd79qVgftFF0B+9md/dvpz1euhMSLGoOhcRtiY8oMAxsJNj3woSqAsHFEDITRAxFrBGUs9nmw7lGnKMonZ+TSTyWQymT0sfsz/bkh9fzK0NJXBneb+8Eten+0MlQilc53PRrttRERBwTnTih8RkVYesSViCsBS1wEomNQQFYLAuE7OH+98z2381utej48WW/bZGo4pioIY4zS3WSaTyWQye75Hbj0su/61KC3HN46y/8AK3/+Xv5VJW7dk0CtR1amzA4Axpi2GohRFwf/6X/9rNy2Li8urXvUqOXz4cPpFUsWX4D2CwZrkKhNimOpLzsILnv9c0CR6xOipeiWIJnfYpqHXBfdidj60HAKTyWQymczeRk9muujSj3HpbzvYAa2wEuPMU0OMofUNgWjwTVqKskfjQcWhwGgEXuGP3nor//bf/UfuuvswK6v7CT7tX6/XpygKhjuG0WQymUwmc4mzbfwsO4yvI4cO7uOxj3skMaQKaobkfakhTpOddikuiqKYJkH9jd/4jStHAAH49Cc/BRJpmmR0iDXT3YkKInbB4fTF3/58qsrR+BGiEdH0XmMFiBTFLDlqlLRkMplMJpO5DFkQQoTlvBzTv8myudMlLu2WJIIYa6knk6koMqkbGp9+tQ7EQRPTug4w8VD04J/+5M/zr/7VKxmN4aqrb0ClYFJ7rCsZj8d476mqKl+vTCaTyVxGmGn3Ohwdp66HfPsLn0dhU08b2jkFYy3GGGKMGGOm3iBN03DkyBFe85rXyO4dwS7wP/7H/0hf3oa+GHGozowXI7Pd8j5y1dUlT37KE2mamv6gwvuaEBqKosBYoa7rpW/ICkgmk8lkMpcFO3Xp8wnZ5tfLydGXRRBdMnsU6nFDWVWgynA0pKwKjIOtCdQNDEfpc8Y1VBX8yqv/kBe/+BW8/wMfpezvp6zWCOo4dnwLW5SsrKzgvSdqWnKFukwmk8lcDsLHfB8qqqz2+xSl4dnPfkbXpeIcxNbJAWVa+tZai/dpduGP/uiPdvtILj6/+J/+X9nY2MBaQbDENubW+9hG90JT+2S3aAqH+Y7veCFiAsYqSKAqHKKRGOP0ZO5ENjsymUwmk7kMRJATCSHSmjPzJWHmttc2U8h0G22XdpuyqmjqGsTSHwyIJNHDWJACyj584fYJv/t77+NbX/C3+N3fezMTX4Ab4Cl54MgmjYeyWsGIY3NzkyjK2toq3udEqJlMJpO5HDrh7fQHFd/2/G9l/z4IPpWaB9rQlxT2Ml/6NoRAURS89rWv3dWjcbv1xe973/t4yUte3J4kOz1ZGlMJXO89ZekoioI6wNNu/nr271/n8OGjKRN81aOux6imZKo+RLbrOSaLIJlMJpPJ7Fnijv36qdDT+cNchdqoStBIUGE48lS9ggcf8Hzggx/h1/7767jzq/dw8OAhtsYBEYMPE3rVKnUdOHDwGkQMqik32XAypqoqnHMLhl8mk8lkMnuvD17Osdn2yxK5++57+It/8S+kLaNinBAazzS7hSpiZFropKoqhsMhb3vb23Y1XGPXeubXvva1+OCnbjEA1giqST2aZk7XlEjlwAH4pud8AzF6UjLUQFSPtXb6/nlDSSUtmUwmk8lk9jKhXdr6bkseHp1Hh85+bLeM21Kh6vJ7gRAbqqrASLJDrLX8p//0a3zP97ycn/+F/8T9Dxzl4FXXc/joEFesACXr+w4RotAbrDFpIkeObbCxNUKNsL6+TlmWbG4eb6vWZTKZTCazd0UQUbYtAFdffYhHP/ogAGUpSNun0or/nfDRJUEFeM1rXrPrR7SLAsjrxNlUIkdVqesUxiICMaaSdKqKxogzUBXwwm/7Zq7av05hhMZPsNbinGM4GgFKlJn5M7tABtE8A5PJZDKZzF4m9e4mySGaqrB4ZkvQmVSSFkPAkJxdlxxe53KnGmPwUVGEEJVBz/CP//7f5qf+7f/D2kpBz1mOPPAgKysr00mX0XDCaNLgfZodGwxWKcuS0WiE9wErioaGflXkSnSZTCaT2Zv9roCKWRJDUsl5o57v/gvfiRHwbbVbVU05tVjMx9VVfplMJvzqr/7qrh+X280vf9Ob3szLXvYyQgiUZdoVY1k4WVhBSTWHn/Gkh0IYEupA4VZQ49gajhnsW2MymdDN+aTcIQ6jgsT0uWpyIrJMJpPJZPaeBWZAhEBg3MBLX/q3GI8LQjSICYQ4Zv++Fa46dB2PeNjDeNrTn8jTv+EZ3HCDYIBCWocPgRAagnoK54hEDBbEpbVC2YUtK3zbNz2N5z7r/+PH/v5P8rnb72Yy2SREA7ZEgaK02MKkCRxVRJWqqjAKofFUtiD6gGCy9ZHJZDKZPdf3BhTnCkCJ3uO9Z32tx3BjgwOH+vzvP/xyQoB2GJ/G7pGpl2VX6VVEiDHl7rzlllt2PUZjV10jfuqnfgpVxVp70u1EwbYGzEte9AIMStOkknVFr+LokWNJ+pgmPAPTOr7K3GdkMplMJpPZgwIIsxCWoH0863hW8axSDQ5x9wObfP7zd/H+D36Cf/fvf5Hv+77/jZ/9hT9gPIGosLmhxADGFiiR2o8BiOg0TGbqRNoaD9ZFKhf51Vf/FM/+hqdSFbAyKKhKQ69XYq2wtbXB9jwli/ZLJpPJZDJ7EedK6rqmrmtc6VhfG1CPt6h6ln3rKxQOCpvG6d3YW2XWl8ZIW+xEMMbwi7/4i5fEce2qAPKhD31IhsPhqTdsVYyVVfjRH/1BVtcGlK5gNNykV5Ssra0tHdJ8mZ6Y3U8zmUwmk7nsSP39eDyhLHo4VxJjZHV1ncJVvOmNb+HlL/8xHngQVlYFMTAe15SuonTl1EHXdJ6jy1iHLQ3Gwr/5t3+PH/iB70fDhMlogxhqQtNQVVW+DJlMJpO5LIkxeYAURUEIAe89k8kE1ch3veylFAWzhKddJIbIdOze5QEfDoeMRiN+4id+4pLI0LnryTFe//rXn7SMbUeISUVaXYXBoMfKap+qqhiPx9STCUalFT3M3IyLopIuhuZpmEwmk8lk9h4nmcSIgHOOquoj1uJ9pKkDASEEZXNrxCt+9B/x5a+O2BhC2SvZGjeAZTgeQ1sgd5oXdTlJaruuCvgrf/mlfNd3vQjUI0TQhsLZNt/Y4j52idhzQvZMJpPJ7FWapqEoCqy1+KahaSb0BxWDXsl3f/cLulRaS4JCXOhHjTFUVcVXvvKVS+a4dl0A+amf+qnTEkAglcdF4MUv+jbq0RAngm8arCkAg1Hbih8mGUwSQUJaMplMJpPJXJZsbW1x/PiQ2oMtelhTIRRgCr5y5z385L/8d/QGKTlqv9cHHCu9FVLO+s5TVBdCYGjFi6aOGAPra/C3/9Zf4xlPexKlUw4dOMDm0SNTt9+p6SeLIkgmk8lkMnsRYwyTyQTvPdYKZeUwJvKMZ97M6tqi6KHTGmwzQgjEGLHW8pu/+ZtZAOn4/Oc/L1/+8pdPuZ21szwgP/g3vhtjFR9qnDOsDgZz1V5k4WLMyublMJhMJpPJZC4z84y6Cezbt5+rDl2NYBiNJqiCWEvdKCvrh7j9y/fwd//eK1GgbqBpUraw0XDUJheJbem4toyupkUVqspM84Mc3A+//Ev/DPyEIw/cxYH9qwgBQ5wTQnQhLxm5El0mk8lk9hxCVfbxTcQIlM5Sj7cYDTf4nu/+bupxGnWnPjOeYPye8nxubm7yUz/1U5fMlMAl0Su/7nWvO41LEIkRhiNlMIBveOZT6fccBuXIkQcxOm9ozJ1fCVn8yGQymUzmsrHJusmN1pAxhuPHj3P06FGsLVhZWcOaghjAFT2Obk7YGgc++Zkv8OY3fxRbgHEC0dDvr27LGSYCMhe+IniOHbs/JXkzoB5+9Vd/iWuuOsDw+FEEjxAAxehiuG3MHiCZTCaT2aPEGKchMEqgmYx59CMfxpOfdAO9su2DFUzrT9n1pUpM/8eIMYZ3vetdl9RxXRICyG/8xm+c6vQzngwxJjIYCI2H5z3/mwmhxjpYGfQRbS+ALtYqblWR3IIzmUwmk9mLtP36vJYwH1qiqlS9Af2VATHGlBvMByKC99Af7CNSYWyf//Aff46vfrVGgMYDwexoI8SFV5UD+9cRYiqr6+AxNw34vu99GVWhWA2tCBIXxBnFoJgsgmQymUxmTzIej3HG0DQTQvCs7xvwHS9+IdaAkSQkpEkDWZBAUt8cMG0W1N/6rd/KAsgyX/rSl+T1r389AN57YlyKp42RqiqBSIgpGdm3fPOTMXgKqww3Nxa8PIyaaULUHIObyWQymczli4gQYyQEBbEY6zDGIVgwBcNxQ9VfYzgKICX/5//1zzl2LCVWxwBqAcOkrUqngA8Rnf5ckyZittL3afIE+avf/yKe+cwnTUWQph7ijECMlGUqHajZAMlkMpnMHqV0Du9rrBVWByXHj97Pt3zrs6kqmNRtH9xuWzc102qsgJG03tzc5HWve90l1RleMoGpv/d7v5cMD1WMMdR1TdM0aSdNcqwRoDAQAgxW4C//lZejsebaq/cjBIRILvaSyWQymczlwqnMlJ1DXBWIkuytXn+F8cSjWJpguf1Ld/P63/+fuBJGo856M1QrqwAcOXoEZ82sIEz7OWXpUjhugOhhbRV+/Ed/GNUJxDEH9g2oJ0NGoy1CCKyurjOZTPIlzGQymcyepbCOph5z3/138XWPfwwHD6wiAv0q9ZExerz3lEVv4X2TesJwOORVr3rVnrMsLhq//du/LRsbGxRFAaSkKd3PCcEHP7VsDPCdL30Rm8cfZGt4DPBM43cVpkV5sgtIJpPJZDJ7GzU7mCxJ/IgSiRJRiaiJ27r8yWSCMQbfgHMDkIrf+u038PZ3foayn0yEEGAyaVDgwP4Ds29QWi8Og4bkkeqK5PJrgMc+Zp2/9+N/h6bZAG3olYZDB/fjvaepPetr+/K1y2Qymcxe7HjR4PFhQn9Qcu01h/h7P/4KDh6A0LQJUNswF2n7ae9bb0wMVdlnMBicVq7PK1YAgRQf5H1SkbqYoaZp8D75qUavCGBd2v7GG/s87useyXDrKIaA7JjrI2dfz2QymUzm8hJD2gSjXfna6ZKMtpR3IxXla5oGay1eFedKyt4K9953lF959X/j6CaM6jSFUlYV49anVxQsoD4gOMBirSOGAICzUNfJCPzel38TL3j+czl+9H5ElM3jxymsxRrD5vEtTJ6EyWQymcwepKwKRCPNeERVWZ729Een0bVJJW+DTyGi1pbEEFIkR1v5JQIPPvggn/rUpy65TvCSUgde/epXY60lxjh1ObXWJjFEzdS1prN3nIO/88M/xL61Pohn6gXSHZq2hekkiyCZTCaTyexNweNEtlOcTXvILAFpnK8SI8pg0MM3Dc45Gq8c3xhx8Job+MIdX+XnfuG/YSsYjVPpWuccw9EQIynMpbAWawsUB1LgioLokwhSVVCW4IAfecXf4sD+NarSYhR83aBRKIoqX79MJpPJ7ElCCPT6JT6MefITH996fYCzSgw1rnDJVZKUV6uL3tjYGBEV/uW//JeX5HFdUsrArbfeKl/72tcoy3K2g8ZgjEmzLq0oEtVP3U+f961PZGW1h+BT+MuCF0hrNE2TomYymUwmk9nTyFyNFtGlUvc63Ubb1zV2Xhsl49pjbEmIhrK3zlv/55/y7vd8lv6qcPT4EGMdK/0BKDT1JJXDBSaTMDObJAJh4Ssf9ah9/I0f/Gs8eP+9HDy0H2MMw40h6+vr+XplMplMZk/ifY33Dfv37+OH/uYPUlWp+kvUBmPTODvM5exEkxBSDfqIwC/90i9dki6Ql5wq8KpXvQpVZTKZUNd1e/J9isHVZNh03iEhgrHw7Oc8g0X3V+bihQ05DCaTyWQymcuZ1utj6gkye62ux7jWbjDGsbK2j+MbWwQ1YEt+8l/+K44cg7X1QbItQsA3Db1eNRU4jFiaOtkYxlpUA8Qa7yfENi/ZD/y1F/Lwhz2EzePH6Fc9Sldw9MGj+dJkMplMZk9SliV1M+aRj3oEj3nMOgYIvqFpJtM+1rbCBwgxQl0HnIW3v/2dl+xxXXLKwM/8zM/I5uYmzrmpJ4i1FutMm009YlAM4Eya/PnB/+2vYvCYaQhMMoBibreZTCaT2eu0XoxG25/nkn8CqHQ5L5aQU3TzMpMJ0pZdEvEL03vqXJgKEtPxAIKZ7UznuClzUxrTY7Pti9p6fQZE49KEB7PPVYMo9KsqJUK1YK0wGo2wZdkmanOI7fGv/+1PE7SdSokWVxSgEQLUjWJLcKWkeRg1aVLGOFxRYF2a8RID/+GV/wYxgc2to6yu9ShLN3cQs6WrUDPNYyLL4s38ezKZTCaTuUAs9T9RZn2Tb4asDwqe++xnIEDTQOEKqrLPZNKWvTWm7aoUY6DXsxjg//fK/5gFkDPh/e9/PzG2hp3q1OOj6vWSAGJA8FhS7O0jHjLg5ic9kUJAQ3LJcc7QxAkUgnGys3GYyWQymcylbpsAspQHIxU407b6SRIX5tOBzvKCGoS0TBUFMxMYZoEcsRUNZmLI+SAK0/0Dj5oGJKTv006sMKBuQfzosngEzEwI6Ww0BSG0kx5JGFEtiJipGGKjwWpagleqqiI2DcQGYxURUm6P6JiMDB+55bP89uveQ9SUaF0jYCLYhrKUZc0IYwpgLrzWwngCX/f4ffzF73kJK6uOreFRjI1MJg1F1ce4Hj4afATEYooiXbe5ZVbNTloBx2QRJJPJZDJnL24sWhPMi+tVr+D48aO4yuEqhzhh0tQogmqg7yaY5jjf/dIX4CKUDrwHKLC23+bblDSp4ZTh1hYCfPpTt/Gn73jbJdt5XZICyM/8zM9Mk6iICN77pYuo2y7lS178QjQ27FtfwdcTjm4cY21tjeFwyKie5Bsgk8lkMnuTqfgx32XHbT+fdtV33f4J8+KHdKIE5yt/1nwoatwmtizuSFzYhuVeX5M4Y/BIt4060nRIO2vViiASJS3zhqAsfr7iEFPh7Aqv+63f566vpWSo44kHAdUk3GzbmbmI2xgjBuj1wDfwd3/sL3Ho4CpFofhmSK9X8sADDzCZTKiqiqKoMMZQN00K8aUTiOLiZE3OX5bJZDKZ8yqCLKaGCCGwtraG955jx47RBE9RFDjnMAYKG/lLL/9ODu4DDe27JU0SOFcSVRFjptVgBqt9IPLbv/XaS/qUXJI969vf/na5/fbb2dzcBFJW9qkxMHWV1QUD6Tte8nS8r3nwwfvxvuaqg/sZjUYYcdPqMZlMJpPJ7D3jRVu1YDlEYm4TQNDZvI7MGz6RU4W1yHwYibq2v5XzYyaoTQvJo8FMPRt2FmYEM/cv7YaQvEXSXsWZeLIgEi2FjEzP284+FLE9tqrsc/T4JkeOHec/vPKnUaDoO5pGkNbTY8EPQ5bOrzYIEe8DVQH9Hvz9H3sFpfH0S2FttWJ1ULG+2oPo2Tx2nOgDZVFQFb05T4/lwznx9c5k/v/s3Xe8bFV5+P/PWmu3mTnnNopiw45iQVSCJQohCqIIilGsiVEUQbHHEo2CiuZri6JRYxQ1v3xjj1ET69deo8ZYEhBRsSAIAveeMjO7rLWe3x97z5w5595LkXbPvc+b1+a555yZU/bsmdnr2Ws9j1JKXVPj8Zii6LfLbE1ClqTEGKnrmuAblpcXedDRR+LcyvuedVDPFD41RtqVG12tTgnwwQ9+UBMgf4h3vOMd9Hq96VKYHYvTE5EkgSf+5ePZsLHPho0DFhYWqMYl8/Mb8D7oEa6UUmrdkh29fYtpl7jsIJmwUssjrNxbVidMVqc47PbFw6+V2QezSzm2L0puZG2ypv09V0/UDWtn7U7va3eyRETYwWyYHfw9EctwXLJ58xbqRvj2f36PD33kK+33TxwhTn7nyPaZlAgSsc4SQ02eGOo6EAMcfvhtOerI+1OOt7Htst8RQ0VdjbAIvX7e7vto1pyf2DV/wSRxJfoEUEopdS2cSaxeEhPDpNmIMDc3h3OO6BusgSSx3OY2t+IWt9i7XR46mwSxtMXAu++XZhkYB9HwiU98gvN+/rNdeu3mLpsAef/7349zDmstTdPspBaYMLmGIwIPPeYolhcvI7FQl2M2bdpMOSwRL+gaWqWUUuvylGU6E0BWDf5Xkgrd7IhuwsNk1kRba8O3c1Ul7iyTsnJCIBYrpq03IrMpkmv6B3QVu8TNJFm2T0WsWloiM1uXCDDTvzu2y2DEdnXXdpAE6ma+yNoZFKsSMe2JRZ71uPiSy9m0ZV9M2uPMt76Tn/zs8rZ+iZ09XVy77Ci2+zjWbV0RAnkqZA4c8LjHPIxb7b8PvcIy308plxeQ2DDX7+G9x3uPw9EWd022W/IiRhCjyQ+llFLXpjh9Hy6KHr6JxKYr+l2WFFlGkSWkieXJJ/4FWQ6mG4eHbkiepinGGKLE1SkFY3j3u9+9y++BXTYBcsEFF5iPfvSjK7+m7OjBW5nWax0UPcsd7nhbLrvsYm68377ExjMelswPNkyrwiullFLrxswSltXFvCeJj+6sZO1MCJkdpE+SJ6uTIOZKTgeujfdNOxnQd0U9V35q3MGFjbjm95/9IOzolwYxa2aw7Oz77TyRE0JkfsMmtm5bwiYFJAUvf8XfsoPqHzP7M84sSzJAxPsxzsh0KdKt99/AM05+EqEZkSbQ72X4esRweYnEGvI0w+BmZrHY1Y+7Ln9RSil1HaYBrE0wxuJcQgiBEAKps9TlmCyBB/7p3Qih7XI2eW8SWTmfmKzUaLolMZdfdhmf/I9/N7v+X74Le/3rX8/y8jJp6mZOiNb+yu30mxBgr83wmle/kjxzhKZhYdtWNm3aRNRzCKWUUrvbqcs0uWCv5O18pgbIDdVZtUsWWFk7BXdmwH+N9sUfeEexiHE4lxHFgckQMs7/9cW85e0fJ8SVuaayKkEx+7sbom9IkpSmKbEEDIKz8CeH35X999+PrVsvxbpIkjjEN8z1B4QQpieNkxk8dtXjpUkQpZRS18Z78PYzSaEtggoWZ9oZlVniCL4mhpoj/uT+GMA3fnrqkCRtzQ8AH2qcc93bYNu05E1nvnl9nD/tyr/ct7/9bTM3NzepqbKTk7pJQbT2atJNbmrZ70Z7EZoxc/0emUsYLS3qga+UUmqdizvPLUwTId2tzOQtvtsmZy8z76fCjpaNdOtOjBCt55rXn2hnSliZ/UVXEh7T2qXY9hLT5Hdc9bsawK20952cvnR/r4l/4KnM5P7GsTQc0+8PsC5n21JJmg34yMf+nfN/PZru9TjdZ7NLaBLqusEmPcRH0rQgxICE9oSx8fC2t72Bm9/0RjRNRZpYjBW89ywtLDLXH+hhrZRS6nokM2Noh4jgfURCJHGGphqx1+YBT3nyCdQNFL2uy1r00/tF8SQuwWBofEOSJIxGJa985SvXRc2JXb6/2tOe9jSMYTqLo2kCIQh11TApTCZErOsmkBp46EOPIoaGJLEsLmxlw4YNeqwrpZRan2kPw5rlL6uTCMYYsqygiYLpVoyKgEgCpLQtYt1KIoTVpTaCAMbQhIBJHGI9UWqMiVh3zWcgWAshVogETGqofNMmEgz4OPld1iRrVv2dKYIlSHtbm6TYJMNa203FXTmhmZ0JsrO2wJPZFhMhBIqioAmG2gtp3qeqI8bmnPqs51PV7c/1sT31q0Kb2ICEqopk2RyQYpIekOJsjnNp200mhU2b4Nhjj53+vCRJiDGyzz77TLvdrdNTNKWUUruwtV3GoonTc4oIiAjW2mk3FyvC3CDnTne6HXvvBVkGkxmJ1s78u1sTI2JIXIb3kR//+MfrZr/s8u+u//AP/2C2bl3AWmgaIU1dN92mPaspqxqDYM3KH3TMMQ9i46Y+Emo2zPdZXLhMnwFKKaXWn0lhzO2WuawkJsZVydJoSFMHkrTr/WJgVBmElSTIpKZoMGsa45r2JCZIJEZPkiQ4Z4jiqarxNVyGIXhf45zBGEGkTTYkaU6knfQx26tmR1sd2q8bCyaFsgqUZU1AcM4hdqYuBzLz+17JKc6qv6tNppiuIGkToCoDw+WG9/3TJynr9ueH2FbCj9YSgCTPEbEzW5d8WvltaBr4iyf9KQcddBBN8DjnWFhYYHk0JO9nO/19po+/Ukopda1Zea+0DkLjMRIZFD2WlxZYXtzK8Q97CNbOvovu+DwgxogxBmstL3zhCzUBcm36whc+T1k3JJlheXkEWLK8rVpb5H0EQ4gNVV0BcKMbGe5773uSJIHxcIH5uQFG28gppZRah1anPmRmLW97EtPvz7FhfhMbNm5hcRlq387qEAMN0Ei7+W6Lsfs37dZ48FHI8x4xQtNUeN9gLeR5do1//8QKSWrwoWY8HoNxDMc1i8OV32mHG23EtQmQAJQlFL2NJGmGMa5bvxx3enKGWASzeiZIl2iwItjufm07v+7LxmBIqb1QVpEPffiT/OAHF9A0UDcrS4xGFTQRamm3Vb97hCjt45Ck7e99yinPIMaITRL2vtHelOWIPE9pC9xK+1eYqMkPpZRS147JVMjJ+8kOCmw750iShOXhIntv2cCdDrwt9733Lae1yle/v9pVm3MpYPnRj37EN77xjXXTcjVZD7/kq171Kh56XDt9dDDo430kcbZbFmPwdUOaZdisXSNc1/DQYx7EZz79BXxjyIqCbuKsPhGUUkqtsxOYyYnLjr88Go1IM8O2hSFvectHEKnI85TEwni8TJp0Jy2SMC2uYTxifHd+lHP+Ly+irANZNsBHIeARIwSJ2Gv43tn4iiTNSdMUSQy1b/ja17/D+ef/CmMjBA8msrrjq52kexACQQzO9rG24LLLF8HmGBcRATNbMGRyYjdNHqzZaZOvr7pdRMR2V7LAuWQ6i8NGwyWXbOUd73gvN7/FfhRFSoht4dKiKCjHNc61SaKV37/dvxiPkJBmG1ge1vSKeQbzG7n4kq3svU/B3IY5FhcXcFmbBFl1UrqqY45ewFFKKXV1TbqLzVaymv23nS6BMRIhBsajMY959F/iDN2Sl+3fl3f08Xpofbt2z6wL3/7Of8qd7nSnacEwXweSSXeY7kqYYBmWgaJIiQInPvUV/OrXl7C8HIg2277gm1JKKbXO3q7jqkxBJHWGxBlGoxH9XkbTVMTGMzfX5/Ktl9LrFe39JVk58TEBTN1+FC3WDShLS5oNCF4wqSC2IsZASn6NZiNEaZfViASsEcpyTJpYYvSYGEhSe4WDfO89JkkwpFiXg6QY48BYnHNUVbn9Mh2xmG7ax8rsD5mZOTNJCrWJBuluZEy7VMfY7nYSmMsLtl5+KUkKc3M9lkaLiATm5jawtDgiTXprTgx9t3/bBNNw7Nm0eR8Wtg2Zm5vHB0NVNgwGAxaWFyh62aoTUtsteTLiut8/ahJEKaXU1T5vmNT/kO49KU6T/0n3fuOwCCY09PuWy35/Pt/4xscpcijSNTNAZPtOaCEIy8vL3P6AA7jk4ovWTV5h3WQEXvvav2XQ71FVDQjTtjuTB6Su2xMNN70KBE/8i8ey7fLfd0XctJWcUkqp9Z38mE18TN7XQghUtSfLe9Q1WNun8o4m5qT5XkQzIDJHNJOtTzR9xBREU5Bmc2T5HEVvA7WHJgrWpaRJTpKk1/jdM03TruVrwAeDswUxJiTJHIMNN0JkjmgGCIPp7xa73y2aApdtJCs2YUyPxaWSNOtRVjWLi4ssLy/NJDVm3+snV77sDpIfq9vMWmumbf1i9MQYp5tEw9KwxiQ9jCmogiPL57HpHMNSMMnK/oyrfv+Vv2Hjxn0Zl57NW/ZiNK7IsoysV7A0WmJ+w4AdL+ExM7VflFJKqWvLSq0sX1c4Y0mShOHSAk984hOYH0CWtrMvd5j8mOGc4T3vec+6Sn7s7KxqlzUcL0svH7SPW/e5ENq1u0lmppXk69gWKjMGHnT0k1lYrDHppq4YnFJKKbWe3qbtTNJjMgNkZSmHiLTTWE1C5jJijDSNx9oEay1B2iWgNk5mTQYwDcZUiInUZYOxOanbQMSRJBaXGRq/TOPHOLJ2VsIfqA7tTJO8S6iICCG0v3MM4IxFbLsEZjrbwaxM0/U+kmQFBCGIIc9zYoyE2JCmCXXdzgCZdHYxYkHcdMaL2Eg0oav3sZJsMN1JnTEpSDsLZFoLxITud4HEFGRZRvAVw3IRaw02cUi09Ip5qrKr2zGZamIboJthYwSJgEsYDUs2bdrE7353CXvvuw9lOaIoMnyoAemW0BhszNrZH5IAEblW2hErpZTa084fTJy8DzbdDJBJvSvX1siqhV5eYE1DbBb58pfeTZq2XVWDH5MlVzwD1Hu4y13uwk/O/Z91lVNYV5cWXvayl7WnQzOzQUNYWQpTVxVRIokFZyC18PSTn0ivAGOaNVNkzcymlFJKrQeyg1McR5rkeB/xEhmWJXlvgESDNQnGtEtGxBrEGoxxGGPBZBgS5uY3kbis/boEvPeMRiPquiZN8mv8G2dZRprmWGspy5LxuCJ4wdm0XXpiHYYUTBvFJN3mEJPi0oIYhSYIvV5XRmPSAADObklEQVSPxcXFNgESAqPR6ArO/WLbIeYKCqS25xENQsSYtquMdaZbCmOwtm1Zu7S0hI9ClvYxNiVxOcalLI1G7X4zFrHdPjYOsRYx7d8SjSUG2LhxM6NRyU1ucpM2aRXav8EI3XIdM5No0pmrSimlrmkOpJ3puP1oN2KBvMhofElVDrnTgQeQJRADWCJZklxRbgUMnHfeeesu+cF6G/3vv//+8stf/hKApgmkqSOE9kqYc46madoia93tvYfEwZ8e/WhG9QaWhqatch8MNkmx1jIej0lTN7OmelIdfm0BNb36opRS6oZ8q5Ydn9zs8OrMjt7eZwfXOyoaujLbJE7rTsRrNPvjDz/1kKtwP1m9H9YkNna433bU0ndnf9/O9u0OC62ubVG8s6Ks3S0MWGn3s93p76PnH0oppa7Bu+1kZqIRZLYGCBYr7YzJIjFUowXe9563cccDeiQWME33XtZeqIixHWuHEIgxkqYpZVlyt7vdnXPPPWfdJUDW1QyQX/3qV+bjH/84AGnqGI1GOGdwzk0fGIAYQns6YgLGwJFH3J+qHGJdO+UWI4QQ2qtj1mLMFc0EsVfjhE0ppZS6tsnOB8E7TU7IDrZJM1lZfX+ZqZPR3caKYIXrIPmxs99t7XZV7sf2f8eq3/dKbndlNTZ29rUd/pxwFfbvzJmFTOIV/T6iyQ+llFJ/eAKkbQTf5UC6Itszw3+JNT5UHHS3A7nt7XrtV6ZzAiISI8aYdvw8+Z7dx5dddtm6TH6suwQIwItf/GLG4zEiQr/f79Y6N9M2PpO10LBSKPX4449n8+aNIKGt7t6tn47iVxdTVUoppZRSSimldisWJMHGZObihhB8yYOO+lNS1+U+DG1NK9NOMABIknY5qLVtwVTvPX//93+/nvfE+nLOOeeYH/3oR9R1W1QthECaptNExmTd7kSMkdvc9sbc515/RIweZwWRgLOGEML0Nju/yhKnB4hSSimllFJKKbWrk9nW72IAx6TLmCGSOuGWt7opf3LE3SkrsF1mQHxb3Nu5bPq9ZsfXv/3tb3nNa16zbpdHrMv+ameeeSZ5nrcF2tJ0+vmmado/ytpVCRKAI4/8UzbMFRBrgq9IEguxnREyO61HKaWUUkoppZRat7rkh8zU/TDRYmLS1jAVwYcxxz30SDZtmtynDSIGuu6pTdMQu6UwkxkhH/vYx9b1rlmXCZB/+Zd/Mb/4xS+wXZqqLEsAvPftYzezVilNU0KAg+56cw66652QWEP0GIkIbfvA2YyWUkoppZRSSim1rnVdYBDT1ZhyGLEYaTu99IqEIx5wX6JAr9fepfENtptgEKUteDoZczvnKMuS9773vet6t9j1+ou/9a1vnT4YRVEA7fqkiSzLVv2RvQKOPeYoEiskTojSYC2IxOn3UUoppZRSSiml1jsxsqqIg5kW4AYjkYc+5Ej23dvhmxrwRKlX1ccUWZkkMJlo8KlPfYof/ehH63r2wLod+X/4wx+eJi4mD8jsx7OzQIxp/9DD7n8AWQpZ5pBQYU2A7Za/aDJEKaWUUkoppdTuxRIxeKxpePKTH40I9DJL04yQGNo6IALRg3OGuq6JMU4LoZ555pm7wT5Ypy644ALzqle9iqqqpjM/JvU+kiShqqrp0hZj2iVNeQIveOFzsabGEMnzjNFomaLIWd3mdmW3GFnJlimllFJKKaWUUru64OO0fgeA4EkzgxHPPe9xEJs2QmIAAnmakriVMbCdaZQ6uf+5557LV77ylXVfO2JdT3c466yzyPOcEAIiQpIk0yTI5PPdo40ECB4Ou/9d2bJ5AzE2jMshe++9hcXFxcnDq88UpZRSSimllFLrWpYV5HkP6wx1NaTILIsLlzLoJ7zohc/GdqPfNiEQd/g9nHPTBMiJJ564W+yXdZ0AOf/8880nPvGJ6QNjrcU5R1VVwGxXGLDdMpheDo94+MPwdUmRtjNFJmudhK5DkFJKKaWUUkoptR6JxUeoqobl5UV6/YwoFRvnCw4++E7sd6NJ8qNNfBjsSmrAtFtVNdNx8s9//nO++c1v7hYj5XVf8OK4444zCwsLK7M9ZswWcbEO0gTKMRz3sAdy05vuR5o6xuUQlzmYKRGjSRCllFJKKaWUUutVCAI2oTfo40PJuFzExyHHPexB00RAm+uwtG1vu0FwF/I8ndbUfPOb37zb7JfdouLnT3/6U5xzDIfD7sHKgbYFbtuvOBJDOyukyGDjPDz46KNYXNjK3Nxg2kYX6NoEzWRAtAiIUkoppZRSSql1xNkMMOR5ynC0yF5bBhx01wM49I9u2dbInA5xHbOLYWZXRYxGIy688ELe8pa37DZTBHaLBMjznvc8AHq93jRLNTGZGWKtAeNJXPtgH3XUkQwGA0SEbKZ9rlJKKaWUUkoptV5F0yY0mqZhOBwyGORsW7iEJ/zFI0lSqMZdzQ+x3ZZslxoYj8cMBgPe/e5371b7ZrdIgHzta18zX/va17DWIiLEGFd1hAHarEcMNI2AgZvezPHQhx5DNS7JskyfJUoppZRSSimldgshBLK0wDlD0UvZe8tGDrzjbTFAr9+lAYTZShBdeqBdQZHnOWVZ8o53vGO32i92d/lD3vrWt7Z/kLXTgqjQtu3xTdP+qdaSZl1PXIFHPfKhzPUz6tEIKwYjtlsHJWBiuymllFJKKaWUUruSyXh1u3GrwYoltQ6ixxJoqiGvfvXL2bQBot9579M2F9ImQKyFN77xjVx44YW7VYXM3SYB8qEPfcj85je/aR+yGKdLYYbDIUmSg6QgjsaPsHjSBG5xc8Mf3f2OpFbwo5INxRw0ARsDaerwvkZMxO+gwKpSSimllFJKKXVdiCYSp6kHM93a/0eMrN7aJEi79MUIWInkBlIT2XfzPAfe8aYY2sYgTVWv6nwbmulPJdIgeADe/va373b71e5Of8wLXvACvPeIyHQGyGAwIEqECMFH0sQBDRLGWODZzzqFXmrZNDfH8sI2MuvI04zxcERRFBhjVpbRKKWUUkoppZRS14t4xUP3K1i1kCWWcrhAM17iuc85lcRMvl0kzZPp0pflpRKXtv9eWl4CLFVo+OAHP8gFF1yw2/VH3a0SIB/4wAdMXdfT9reT7i7WWrDgUgtiqL0nTXLKSrjpzSyH3uvuNH4IJiASETGEIEiAGGDnk4SUUkoppZRSSqnrkrA6GTL5dFfElLXNSyMSG/r9nAPucDvue98DSRIwVvChJvi6bf4CFEXBZMHDYDBAEDKX8cIXvnC33JN2d/uDXv7yl08LoBZFAdDOContAVM1DUQLWHq5ITRw8ilPwiU1c3MJMXqapmFusJHRsMIYu11nGaWUUkoppZRS6oYbxts1/54kQiIGoaqGiJQ87LhjaKr2lgaDxeBmVjgkKZRl3X4nYwlR+No3vsGvfvUbs7vuud3K61//erO4uIj37bolESFJEkyX4cqzHnk2wPs2IZJmcPObG+53v3tSN4sYGzBYsqzA2qTbrD7HlFJKKaWUUkpdbwP11aPQtmXLSuMWy9rkx/S+EukVlpvebF+OPOqPGMy19/GhxnarJUQ8TdOOmQeDjBCExgesdfzVX/3Vbr1fdzvvete7SJKE8XiMMZPFTm3CIwQDAs5ZylGNAZoaTnjMsSSuwRpwxjAeluTZAKLoDBCllFJKKaWUUrvecF7Mysako6lnXC5ywqMezvygLehQDsc4ZxAiwXvM5CK/gRDAOUOapHzp/32B733r+2Y33mO7nxe+8IXmsssuo9frAVBVFcE3QKSp2+UxxkCeZwDkOdz5Tvtx2P0PxTqPSKSufJc8gdBlxpRSSimllFJKqevcpL7HTlrdSretLH1pP2slYmi4+c325YEPuBeNB+8hyxNM999kCUyatXE4HLXj5rriDW/4u916t+62azv++Z//mRACIkKWZdMHuSjauLg4xjmweHzXEeb4P3sIeWqRWFNkCaHxJEmiM0CUUkoppZRSSt2wzNquMCvJD4xg8WAaLJ6nPOUJbNgIzkKSQJo4Gt/2uxUE37T/9iGyYUOfGOEH3/8hn/vM53frDiC7bQLkta99LXVdY4whxpUDZZLL6Bc92k8LxkYinjvf6ZYc+kd3w7qAsYEonrqup8VUlVJKKaWUUkqp65phthdpJJrYFnYw3RYj1iRYk1DXNQQP4pFYMzeX8cA/vTsxQNI2RMUAWZJCNwskSVMwtPc1YC2cfvrpu/1+3W0TIBdeeKGZPIDGGKq6IkqkW9VCmoLESAiezCQ42oPj2c85mdHy5czP5fRyR2hqffYppZRSSimllLphmZXalkVRsG3bNpy19IuCXpERQkWewcmnnIix4MzOUgCW4XCI955+v73Yf955P+czn/mU2d134W7d3uRjH/sYdV1jbdvVxZiEGNsiLwgkiSWECFiCb7DAvvvAkUfen63bLiHGhsbXq2aQKKWUUkoppZRS16mZwqatleQHQF15er0eIQScs4zGS1g8t9x/P4584D2xBqwRrJnMJFndNWYwmCeEtj7mcDjkYQ87do/Yrbt1AuSnP/2ped3rXrfqgbYOYpR2+YtAmuQYHKlrC6KOhvDMZ51E4hqKnmO+38Nai9GnoFJKKaWUUkqp65xZM1TvBq8zvPfMzc0Ro6euS5xrSzsc/eAHkGbt7A9jZLvhvwjTGpdZllGWJcvLy5x99tl7xJDX7u5/4Etf+lIzHA4JUajqCoA0M8TYEINgTQJYjHFIgPk5uOlNUw7/k0OpmgXSzFGWpT4HlVJKKaWUUkrdgMN1QUybDMmLtK3fQSTEhrxIuc1tb8GDjj6M1IE1MJ01Iqu+BWAYj8cYY0gSyyMe8fA9do/uls466yycdWRpTtNUQCTJHDG27W3Fg0Sm9UEEePwTHkXw4+72SimllFJKKaXUDU8MNE1Dmjm896SpY+vWS3nUIx/Oxg2TW60p4yDtZrolMb1ej/F4zLe+9S2+8Y1v7TELHvaIBMgzn/lM89vfXowx4ENN40sgYq0FWekMYy0sD5cxwB0O2JtjHno0xsDc/ECfZUoppZRSSimlblhdIdSmaXDO4RILCHe984EceeQhhLgyvr0iAqR5xpvf/OY9avfZPeUP/chHPoII9IredM3TpLipdW0mrK4jc4M5Yox4Dyc/5fH4ehv1aBvbZdAmx99VjEoppZRSSiml1JUxSDv+NAEjcdXA3QgYIoNBwbatlzC/IaeqFnnG008izyC1k+UvHZG15UO6sTB8+9vf5qMf/ZjZs/btHkREpGkq0tQRYsDZlOFwzGAwAIGy9OS9hOGwZtDPCAJnvOrtfOZzX6dhI8Hk2+9AaacgXVncWQJFKaWUUkoppZRaLXazPaDNYBikG1waBCOQp4bh8jZue+ub80/v/VuyFJyF6CFN2+9R176dKeJc+11ju/IB4PDDD+crX/nKHpUTSPakP/YFL3g+/+f//B+8D4QQcXlkMOjRNBUihqKXUZWBuX5GU7cHzfOfczLf+973+N22IQY//V52kkW7wqkfEYxlNvlhpoevRo0aNWrUqFGjRo0aNWrUuH3ERKKRlQGmWNoPYzcW9SwsLLLvXpt4yV8/myJvbzkewlxXwaFpAlmWTceidV2TZRmj0Yhf/vKXnHfeeXtcWmmPyvbc9ra3lXPOOYckaVNeZVmS5znGGLyPJElC8OBcO1NoNIwMBpYf/M/5eJMgZmXF0EpHoSuY2WEiSHsfJwD2Ks8Y0ahRo0aNGjVq1KhRo0aNe2YMxnQrVyZj0NVtcdPEEkJN5oQD73hzmipS5O1tqqoiz3O89yRJQoxd/UtgNBrR7/e5293uxg9/+EOjCZDd3Ete8hJ55StfiTGGSfGYNE1XHVBLi2Pm53sAXH7ZmM179QimTXXYP3AnO2GHa6+UUkoppZRSSqnZAWQwK8PHHV1yn4xLvRfEN/SKhLoakeVt2YYYDdYmXX1LP50JIiJ85zvf4V73upfZQ3ftnufCCy+U+fl55ubmAAgh4JwjhIC1jhjaWSATZRnIczfNyO1wL8qO9ubKOhmjJUCUUkoppZRSSl0FYmVmVLnjYbsx0NQlWZZ0Y09PPR6T9XrUFWR5b3rbGCPjcVv/8qCDDuJHP/rRHpkLSPbEP/qss97LS17y4umUoElBmLqu6fV6uAS8jzR1oNdLKYr269MjZEcJj2ntj7WZjm7eiLH6LFZKKaWUUkopdaUM0pVUADM7lhQLBpqqJs0TsiwheE9TlRSDHlmvBxGyPKdpGpIkwRhDjJHBYMC//du/7bHJD9iDu7T+/Ofnyy1veUusbet9mJk9MRwOyfOcJOnyQ7OzPiSuSYCsSXhsN0VksnDGIGgSRCmllFJKKaXUlQzSJ8NKYfXF9JlKqfW4IuuvdCqtx8M2ASICxuGDx3tPURTEGBmNRhx++OH813/91x6bB0j21D/87W9/O2eccQZZlrRTh7paICGEti0urGTMppV348pBt9MjdcfaTs6TZIhGjRo1atSoUaNGjRo1atS4fbR0F+hlTeJjzb8nyY8YBOsMWW9ADGAdhBhIkoQkSaZj3fe///17dPJjj06AvP71rzWnnHKK3OpW+xOCkKZp1wmmXe4SQkOaOkajIf3eHITQHoWTpskxEqNgbVs7xCWmTbS1k5VW2hdhEbO2Ea5uuummm2666aabbrrppptuum2/CWDE4JswHZ9Kl/RYVVmhG3Rat5LTMEk7DnXOISIYY6b/fv3rX6+za/bkP/4hD3mofPzjH8fNHDBN02CMmbbKnfRKnmbfItRNWzTVJav3pFzFHa59rTVq1KhRo0aNGjVq1KhR447irNkuLiKCEYgErE3YUfGF6bhT2lSKMYbxeMy73vUunvnMZ+7R4/89PgEC8Jvf/FZudrObEILgXFscJoRAmjpijGAi1iSEYLCmWwzT7bXag/eeNE0wQFyzN6U7ig2QxK5VkT6rNWrUqFGjRo0aNWrUqFHjFcQqVmR5CsC4HNIrekz6wch0ycwk8WHalQfdCgTTfbXpLuaPRiMGg8EeP/aHPXgJzMQDH/hAvvvd75KmKc6lWGsREUQM1lrAdu1x21ohk2lE7bQiMEmC7Q4lBzvMwjlobyN6wCmllFJKKaWUugIG0iTvZnREsqIgEgihQUSI0VNkRXfjSSJk9fyRqm4ospy6rnnHO96h+3Rl16pzzz1Pbn+721LVDXneFkKNMZKm7XqrKB5jIgZBgLprn2tJiEDV1KTp7AFoVk0/skAiYEQzIEoppZRSSimlrniYHiz42F5Iby+me6wBQ0CQrlJIO9psG27YbnjfJUQEnLH89Kc/5YADDtBxfyfRXQBHHXUUPz33Z+RZStOEaeKjnf3R4JwFhCAeZ1KyJEMwRCIWR572WJlyNDlkV/87GrBGjzullFJKKaWUUjsX6RYPSNcNBvCNxdlIkqRAWHV7g+nSHpZJAmRyv3/8x3/UHbqTcfoe7d3vep886Ul/TohtLZDJUpc2AeJALOOyIk1ykqRtCmMTqCu44KKLSVxOmwRZOegQM60LIl03GKWUUkoppZRSaqeDdAFnLb6qSRxgPPvffCPBA7Ftc2vtFY/mGx/46U9/wp3vfGcdhc7uW90Frdvf/g5yzjnnAO3BFELb6QUiIobQgLMG000Oka7uzJe/9H1e8aq/pW4iSEbEIMaCrCQ8omkTIFH3tlJKKaWUUkrtgdYWSdj57SyQRE9Vjtlr8zy3vc0teNMbX0mRX8Gofm0LGQMPf/jD+bd/+zcdha7ZVapzxhlnyKmnnsr8/Dzee6y1WGvbVrhptmqv1Y2QJoZRCU984jO58KIFytpQ9DcxHFW4NEVsJIoQJWCN052tlFJKKaWUUnsC6ZIcZrKgZXUCJE0zFheW2XvvvSnLktpXRAltTUpfIvWILZsGVOUyn/nMBxjkbS2QUAeSxIIxhLoGm+CS9meJgaYJWGs55+z/4a53vasOQdfQHTLjZje7mfzqV7+iLEv6/T4Aw+GQwWCwXUatbgJNgCJ3fPM/f8Hznvty5jftx69+dTE3vtkt2La4hFghyZK2iwwGqzVQlVJKKaWUUmoP0K4maEfdoft3xEhXMsE4YhSSJGM4XCLr5UAkSS0SSnqppx4tcuKTn8CTnnQMJk6WxgjGCITYrYOx7QX62pNmyTTNcug978l//dd/6Xh/Dau7YMUFF1xgnvnMZ06TH0tLSwwGA0Joi8yEIITQZjHS1GG6vXfoobfmTx9wGBdf8lv23mcLi0sL5HmKsyl11ZAkWmtWKaWUUkoppXT43Q4im7omdQnee5xz9LIcCQ2+KpHoqcYjbnPb/XnUCccAIAIxBsykuqldGcqLQJatjDk/8MEPaPLjCh4BNePv//7vzXnnnQfA3NwcAGZa+MNg7cpx1MscjYco8NSTnkxRpPT6CeVoiLUWlxiMWIzo7A+llFJKKaWU2nNM5mLIDofhk74tIkJRFIgIMUasMSTW0DQVj330CcwNurSJhSTpalTG7nuLtLUpZ1x40YW85CUv0d2/E5oA2YEXv/jF1HU9za5NkmvWrrQTKqsRAGkKiYWb3sTwF39+Ar+76Dds3jJPOV5GfKAoCpo6oKuNlFJKKaWUUmrPJqZNiaRpiojBCITGMx4tk2cJWZIQmoo/vs+hHHnUPaiq9n7GgDFttkMktMNLA6Ybo4q0y2D+9V//lV/9/Bc6+NwJTYDswEc/+lHz6U9/evqx9+2BFqfZtdgedDOFbHyAP3vE0Rxwh1tDrDASkRiI3pNY3c1KKaWUUkoptWeR7YffXXHUuvaICGmaEiUgwZMlCeNyiEjkL5/4eJwFZ9pchwWitH1wrZusULDTH2EM/PSnP+Vtb32r7vYroCPznTjzzDMZjUZUVdVW2QWcaw9UgH7RAyJBGrwPSIC99oKnnvgEgh/TL1IsQmjabjJKKaWUUkoppXToDawseUFIrKGXp0Rfk5jIQ495EAfddT8skE2bkcYdftcIIO1F+/e+97389Cfn6uyPK6A75wq85S1vkWc84xmIQDn29HoJPkScE4SANQYvAWMywCJA5eHRj34Ol142piwhLzYxrmraMiJaCEQppZRSSiml9oyhtnRtcGeIBQyJdXjvkRhwJpAXlvFokRvfeB8++uHXk6bt3Z2BEANpV1QyRo+1DoLDdBfam0bYuu0ybnSjfXR8fyV0asIVeOMb39gdZEKvlxACJInFGIc1CT54nJlpbwQ4C//4zr8j+op+L2Np4XKyxDFTOxVjDNbaaY2RGKPubKWUUkoppZTabVzxxW9jDM45nDNYC9V4RJYYTvizY0lcO1B3po2pNV3JD8HZBIMBa2h8mI4nH/CAB+guvwo0AXIFzj//fHPKKafgnKEsPc6t1AHx3pO4tLtlIPgKA4QGtmyGR/7ZsSxuu4RNm/pEXxFjRESmU51CCIjI9MBXSimllFJKKbUbk5Xh99LSEkjEEkkcGGrufvCdOOFR98fISt2P9mJ7ZHVCpb2YnqbtOPKb3/wmP/7xD3X2x1WgO+kq+MlPfiIHHHAAAE0TSFNHVY3J83ZeknS7svSeNCkYVxA8PPaxz2Lb5WOGpWCLPjLT6miS/JjMBtFZIEoppZRSSim1ew20J2mL2H2iW8lClqaMlhfYMNcjSQKLW3/Hu/7xrdz5TjciTdvkR7degNUJkElqxNLUwuLiEnvvs1HH9VeRzgC5Cp773OcCUJZ113sZ0jRHxBBFptORiiQhhEiRw6AHTz/lKSwtXcaG+RyLYK1dtcFK8RullFJKKaWUUnuGshzRHxRsW7iU5cWtPOy4B3OXO9+INJlNfszq+t52yY/xqCFNDe9///t1Z14Nmim6ij760Y/K8ccfv/IJgRgE6yIYoW4q0jQj4hAswyFkOfzV89/Af33/HMqYISafzvqYzAIR6eaPGH0olFJKKaWUUmq3GGhLO74T0xUvnRnuWQGJnsRFHDVpUvPRD7+PjRugSKEqS4qiaCd9rC2iOpnDIHDuub/gDne8jQ4krwadAXIVveY1Z7C8vEjTNNMDzjoDOCRClmaAYIkYYG4ARDj5pL+krJaIMRDCygasKoSqlFJKKaWUUmr3F02k3y+omxJjhKeffDJb9oIshXI8piiylRUvYrcfvgtUVeRlLztNd+bVpAmQq+h73/u+edvb3tYmL4SVjZXZGwZhYWkbEAkR0gxuf8AWHnbcMeR5TpIkq2Z9zN5XKaWUUkoppdSeYWFxK71ezv7734JH/Nm9EN9+vtdLgbD6xmK3G7qfe95P+dCH/z8dTF5NusOupvF4LJPpSBLBdEuxfCxJnEMwCIYYLMYZYmiP16OPfQ6jylKOa2K0GOtwzhFjJMbYfZ8d9YiefZiEaT9pjRo1atSoUaNGjRo1atS4C8bVopkUP52UP6hJE6Eul/joh97HvntbEtcufzF46qoky+Z2PnIXuPn+t+GC3/xCx/NXk84AuZqOOeaY6QFoHAQJYCBxGeAwWCyOxBkskDhwwLOf8STKpUsYFA6LEL3gbI+FbWOMS4mmnQoVTQOmATyGiBGDmWb8jEaNGjVq1KhRo0aNGjVq3IVjNFAFTzAQTIJER54NCE2gn2eEakgst/GExxzDTW9i6RUQQtXlNixJWrS9Rg00vlmV/KjrmrPe815NfugMkOvPN77xDTn44IMpigIRmXZ02RkBhjW8+ox38dnPfpVx7dj3Rrfidxdfxo1uvB/blrcSTQ2mxoiAidjoMJJgYopgEet1xyullFJKKaXULi4aEGeJEUw09IsBw8UFEhOxpmLTfMLGDYZ/eMcb2XvvlOFoyHy/hw++u7AOsasdmWXtx2VZkiQJ559/Pre//e11HP8HSnQXXH0nnHACP/vZzwghkCQJ3nuS5Ip3ZZ7CIx91HF//xnfo02fb5b+nKAp+f+klBISsmGQN43SNl2C7FFXb9zlqEkQppZRSSimldm2T8ZwIEoUQGqyD4D0bN/b43cW/5FnP+mv22jsFYK7fxyA0VU3az/C+LZGQZdm0ZEJRFAC87nWv0/17DWjm6A/0d3/3d/LsZz/7qh3/tGVsIvDhD3+DM15zJnNzN6ZuHDYtSIuU8XiI2IDBAxErCYjDxASxk6UxmgBRSimllFJKqV2ZYJGYYJKU2HiiD8z3cySWEEbc4XY34z3vPo3hMFBkhiyzGCIiBmMMde3JsoTl5WXm5uYIIeCc4z/+4z845phjdAx/DejOuwZ++ctfyv77789oNKLf71/BEwDq0G7ew+Me+2x+f2nJwnJg46a9GI5KkjzpiqBOKv62a8hsbD/fJkCi7nSllFJKKaWU2oVNEiBJnkGIjEdLzBUpeRIQP+TDH34PN96nbahhgaocUeQ54CjHFUUvXzVS997jved+97sf3/ve93QMfw043QV/uF/84henHXvssQwGgyu9rZeKPE1IU7jzne/Jl7/yddIkQ8RM64iImVQNnhzTXQFUI4hZ+5Bp1KhRo0aNGjVq1KhRo8ZdLZquM2jwnjxzRF/ibKCulzjpqX/Bfe99K8ajSJEZDOB9Q5KmbadRAecsw9GYLEvZtm0b/X6f17zmNfzf//t/NflxDekOvIY+85nPyFFHHXUlGcBIJOCJWHJChL/9P//Mxz72/0jSPkk2R+2brinSjh6emSKrYtuZIBo1atSoUaNGjRo1atSocZeLsZu5Xzdj+kVCUw8ZFJYDD7wVb3/ri1lcCGzZ6LBAOR7S6+UgFl81JFlOjIJNzLTW5MUXX8yNb3xjHbtfC7QN7jX0tKc9jYWFhSvJMkUsnhhqDJBYeNpJj+eW+9+I4MeEZoiRru3tzIYJGAIQMF12RKNGjRo1atSoUaNGjRo17rrRAlli6BcJdTUiSQKGiif++aORCJu75Ich0uv1ugvd0s4CAYwxzF4df/jDH64D72uJLoG5hrZt23b6eDw+7YEPfOC0HW7TNDjX7lrvPdaCEHDWYXAIhiyDW9zi9nzi4x+nP+hDjBgESwQzSYDMRKH7WLrEiEaNGjVq1KhRo0aNGjVq3BWjSIWRhiwJVONFTjn5iTzoQffEGbCmnevvm5oYAs4lbSlI59qmoAjGGqyzvP/97+fNb36zzv7QJTC7lu985ztyyCGHUJYlRVFQ1/W0Z3Pjx7jEABEhAxKiQFXDm978z3zkI5+gKOaIM0teBFYXPe1aKSmllFJKKaWU2rWJBOpmzMYNBf2e4/3vfydZCr2U6eyPpmlI05Sm8qTd2DHUNS7LGJdjxuMx97nPfTj33HN13H4tSXQXXDue+9zn8sUvfrGdrgQ45xiPxxhjKIocISBYDILQZv2KHB7/mOPJM8toWCPdU2Gl4KkQu39b0Q4wSimllFJKKbXrs4hAmiaMhls55pgHsKHfzj5oR4OCiJCmKYhtkx8iYAwuyyirkizLePWrX63Jj2uZ7sxr0ate9Sp5yUteQoxxuhwGYDgcUvRzrGk/FwWssYwryLO2Na7rUlFr0xyTZIg+UEoppZRSSim1TnQ1PHwNWdfVtiqH9Iqcuq7Is15b+LQJJKlrEyDWTEeEP/jBjzj44IN1GHgt0x16Lbvkkktkn332wXtPXdf0+30A6iaQpg5DmxAZDAb4IDhnVj1BNAGilFJKKaWUUutbU0OatgO5ha0LbN48T9OMydKs/aQkBB9xiaUsPUUvQYiMxsv0e33+6JBD+d73vq/DwGuZLoG5lj3+8Y/nE5/4BHmekyQJS0tL9PtzpKmjqiJFZimKAogkLhIlEmPE2RwwWGmTHmYmThlZ9XmNGjVq1KhRo0aNGjVq1LiLRQxF2n0MbNk8T92MydKEuqrIshyJbfIDwLlk2vSl3+vxyjNeockPTYCsD5/73OfMl770JfnjP/5j5ubmmJ+fB9oZTcEbyOg6xHhEGqyxWGeAGrBgDAYLZu2sj7Y+CKx8XqNGjRo1atSoUaNGjRo17mKxLfbBwrZtbNo0BwjWBAhClqdIE8Cm7SoA084UESAg/Nf3v8fLX/pKTX5cR7StyHXg6KOPNsvLywBUVQVA00C/3x7H3vvuidHmn2IM7TNEBCSChO23KO0mtOtktouTr7cFdaT71HTrPp78GI0aNWrUqFGjRo0aNWrUuH3c8XhrdlwlCLJdXHV7YNOmDQCMhkOSJCeatrOnSTOMhaZeme4vAgbH3/zN3+iAWmeArD+vfe1rOeOMM+j1eogIznU5QQNJkgBx2jK3TYSY6ZNnUkBVZObj6VqYnSUDZ55pM7NFtru5pmU1atSoUaNGjRo1atSo8Yqj7GQc1X3R4zGAw3RJEYOz7fA6BMFZM51v0B+0iRBr05XxmoG0WPmm0Ufe876z+PynvqCzP3QGyPrzd3/3d+b//b//R9M0GGNwznQzPSYzQCzOpd1D0G4iBmPc9GNjHLZ7ErXJyJ21wo0z6cl28hREtHGuUkoppZRSSv0BzGSMtePRl8VgMN0t2gvXMUYkgjVmzZB77QYxRpqmAaAsS4bDJZ564lM0+XFdP6y6C647Bx10kHz/+9+fzujw3pMkCSIyPeidc9OPAYxZ/ZCICGY6+yMynbQjO3pyyvTpCA7Brpk4oikRpZRSSimllLpS0o2djAVpZ2zMDsGE0M39mAyqEyS0sz+S1FzlkXY73jMcccQRfOlLX9Lx+XVMd/B17C/+4i/kbW97G71ej7quyfN81cEeQuiWxGz/RJgkRowxM0kQO3nGTYvmtDeaTYJMplvZ7aduKaWUUkoppZS6glFypJ1VD+CmCRC6Ti8AMYK14JuKEIQsybHWrB5pX8FoO4RACAHnHN/97ne5973vrWPz6+Oh1V1w3bvssstky5Yt3RMldjVB3PTjyQyRiUnyo018tA/RdnkM6cqCrF2rNpnlIbq6SSmllFJKKaX+sJGy78ZVyXbjMExXtNSs+TxAgKqqyQfZVRpt//73v2fffffVcfn1REfJ14O99trLXHTRRe0Ot5YYV5aiWGtXLYGZLXy6djlMZKUZTASC2UGnFywy+7DKzIZGjRo1atSoUaNGjRo1arzSSMJs+YG13TV9gKqE0DX0nJ0wkufZVR4r/uVf/qUOmK/PvJbuguvHy172Mnnuc5/Lxo0bp4mOSYJjUhtkbQJkom4El5iV56KsTopgts9kme52totRH2mllFJKKaWUukqjZFnTDcbE7W8TGsjS9t++EWL0ZHl6lUfa73nPe3jSk56kI7Xr96FV15cvfelLcthhh2GMIYRAjJE0TVfdZm0CpK5rhqOGk5/5YkRSIAFxIMnKk5LYFuERmdYCMdMlMDrJRymllFJKKaWuKpntACMJYDDTpTARS6Bulpmfyzn2oUfxkGMeSJ6Z6UVukdB199y5yy+/nEMOOYRf/OIXOia/HiW6C64/T3va0/jxj39MmqYYY6bJj6Zppv+erfsBkGUZLsu41a1vx+c//3Xy3hzlOJAXc5RjT1mW9OZ7iASsxGlqclozVRMgSimllFJKKXW1BPEkLgMsTRMIjSd1hn6/oCqXSBJPr78fxx5/JMFDVXuyzCKEaQfOuvbTsd9s7ccYI4985CM1+XED0B1+PTvppJPkbW9727QWiPeeLLviNWIBuGwbPPt5r+Dsc84DSYmkWJezYcMmRuMxPtRAnHainiZAtBiqUkoppZRSSl1lYsA5R1k3pC4hS1KapmF+kHPx737DXlvmiHGZ95z1dm51y0041w6smzAmcY4QI8ZDlhXAygXvpmkQET71qU/x8Ic/XMfiNwDd6TeA7373u3LPe95z+vFoNKLf7+/09gHwwJe/ej4vP+3V1JWQpPM0AcrK0wShNyi6B7RNgNjJWrUuASLbLVpTSimllFJKKbVWxGKtpWo8RZYTfY0zgNQQSpp6G6ee+mSe+MQH4WvIMhA8FmhCTeoyHBawjMdj0jQlSRJCCCwvL7Np0yYdh99AdMffQMbjsYxGI7Zs2bKqIOqOCNDQrkJ797s+y7ve8y+EmNMfbEJIKWtPRIisLHiZzACxXQIkagJEKaWUUkoppa7iSNlNB8xNNaRXpPhmSOY8d7vrAbz17/+KqoQiB2sEX5fkWUqUiDUWgyUEoaqq6cXupaUlHv3oR/OpT31Kx+E31MOqu+CG8ZSnPEXe+c53rloLtjMC1B5MAr6B+x/+aDZsvjG/vXArST4g783R+NUJju2LoM4U8lFKKaWUUkoptRMWmzhCCISmpsgcVjyJa9g4l/L2t7+Bm+2X4Fw72gqxIbGGxjekSQ5ACO0Vaecc4/GYoih417vexVOf+lQdg9+AdOffgP7pn/5JHve4x1FVFb1e7wpvK3TpC4FLt8IJj34qowqSYgMLiyNcmiMzD+faBIglEk3Edi1xNWrUqFGjRo0aNWrUqFHjjqMxBu89DugVCeVwG1kaefELnsUxx9wNX0GvAEQwJrS1F0MA23aNmVyHniQ/LrnkEm584xvr+PsG5nQX3HDOOeec0x772MeyYcOGK7llJPoSZx2+hrk5Q1Zs5Gvf+CbWJe0sEmPARAxgEGwX2y12G7rppptuuummm2666aabbrpd0WYiIXqsiyQ2Ag2OmgcddRhPOfFoQg39HJYXF8nztrFqKGtsmoNvkx911WCsIcsy6rrmmGOO4de//vXpOgrWGSB7tGOOOUY+8YlPXGENkHbuR+imgqSUFSQ5vPT0t/OvH/8cW/baj6oGwRG7YjvTIqjbfR+llFJKKaWUUjsjBkSENDWMh4sMepZ9tvT41399K6OlyKZ5CyJI9FjrwAdwKfgIzrbDLgfYtgPMe97zHk466SQde+8C9EHYBXzxi1+UP/mTP0FEiDHi3OqJOdEHrDOI94hxmMQRgHEDRx79WHwsGI2g6G0AmzMeVaRpSlmNGPQyrLX4RpMfSimllFJKqT1lpLu2RuLqLydJwtatW9m8ZW/qup7WZUyShOFwyPz8BpYWt7JlY0KI2zjzTa/goLvcGkNNgukG0obQgEsSYhCsBWy7dMYYg3OOH/3oRxx00EE67t5FWN0FN7wjjjjC/OAHP5g+SUIIVFUFtGvGrHOE2mOSDGNXKn3kKbzhda+mqZfZtHHA0tISsfH0egOMMWzatInRaESMmvxQSimllFJKqQkRYdOmTdP2tDFG6rqmiYGiKBgPRxRZyratl3Likx7PQXe5NRaPhBJDjW8awOKShODBJoZhNSaIBxuw1rKwsMAznvEM3dm7EK0Bsou46KKLTjv22GOnT8A8z6dPTJc4RARrLcZC03iMcxjgxvttIsY+3/3Of5EmBca098cYQvQIBusSENGdrJRSSimllNojRCMIs0seLLOVPhofAMN4PAZgbm4O7z2IkDgD4klt4JBDDuRFL/wLHGCIJNbiQ0OaFCDtcLoqa6yzJGmCNQZrDMY4XvrSl/KBD3xAZ3/sQjQBsos499xzT8/z/LTDDz8c51z75KOdgiUiWOeIEjHGEiSSWIcIjEdwr3vdnh//8HwuvOgikiTF+wgi1L4hzwtE4nZTvpRSSimllFJqdyXdAGimT+aqrxdFDxGhaRqKoiBJkmkyxJhIgmfDvONd//jq9p4iOBOwph1GG7GU45o0TUhTR/DgEsuwGhGJfPkLX+SUU56uyY9djD4gu5hf/epXcvOb3xzvPWmarjyB2+ccdVORZ5Pe0mBc+/nfXhg5+enP56LfbaOY20RZCS7PGZdlO3MEmRZGVUoppZRSSqndWZzWALE7HPpasRgj7Yx7Z2nqGmOEPM+xxhPKy3jj61/JPe9xS7IMmrqiyBwgxBCxNgcB366EwXVFTyOesi75o4MP4eyzf6Lj7V2M1gDZxdzvfvfDGEOapki3bKUJHi/gRXBpvvKkDh4nUI/hZvtZnnriEyhyQzlaIsssTVMixCvpMKOUUkoppZRSewixIJayLNuZ80Bd14gEBoMe43KJcrSNpzzlCRx4xzb5UY5qiqy9OB18wLp02qQzSSFJwBiom5qFpQUe/OAHa/JjF6UJkF3Mr3/9a3PiiSdOl8CUZUniEoxp2zFZA1VdYYgYicQg9AuoRvDQBx/MIx72ECSWGNMwHi6Rpwmi7W+VUkoppZRSe9xQd6Xuh2CR6ecs8/PzFEVG4ixIIEsNTT1mNFzk4LvdhUc/8ki2bIToYdDPqOp2eYxzSZv8MBCb1T8xSzM+/elP89UvfE2TH7sofWB2UR/60IfkkY98JAA+BrBtuZYQIpmzGCLBe5zL2ieghXEJSQ4Pe+SzuOiSBcT2MWlOOa7JXKY7VSmllFJKKbVHiGb1kFe6a/9G2iihwVowBCTW5HmGr0ds2bKJ9571em60BVIHTdN0NT5qXJJQjyqy3qCtUdB+Y7AQBH5y7tnc+cA76Rh7F6YzQHZRj3rUo8x5553XPkjWTjNVibMIQgyhzT5KBCsg0CugqeH//vOb2WvLPDGMib4iSyzWtR1lpstqmjZd2ev12q4xSimllFJKKbXbWD0L3shK8gMghMDc3ICqGtOmR2qcDbzkRc9j373BWQBPmhogdmMvyHo9VrWXMVCWARE46aSTdLfv4jQBsgt73vOex/LyMgYzfY5V4woJEWtTfFV1TzqDhIhvIC8gz+BZpz6NQT/BV8vkmaNpmq7AjyNNU5IkIcZIVVVaI0QppZRSSim1Gw50d1YKINLvF2y9/FI2b5rHmQZjGo468jDuc5/9qEu6bi8ys02+qwUDVemJEaJAUThOPvlpfONrX9eB1S5O2+Duwn7605+ePhwOu9a4FmcsaZpgbft09iHiEof4BpMmYGmncVm41W325dJLlvnNr39FNa6AhNl81yTpMUmKKKWUUkoppdRuw0hb/UPa2fTS/d8iGCISA1luqEaLpGnkTgfchpf9zTPp5ZAmYE1kJfFhmE75kDYmqSMiOGf45L//B3/1V8/T5Md6OCx0F+z6fvjDH8pd73pXvPckSdK1WwokqSN4j7GC7ZIYgmVhqaQ/X1CO4Rmnvowf/s95FP0b0QSHiOC9x1pLkiTTf0+WxiillFJKKaXU+h/orrTBbZMflsikLGrAOaFphhSJsGE+5R/e/kZucbMe5RgG/cn9A0g7xmovINtVy1+Wl5dZWlriJje5iY6r181xodaFn557jtzu9rfH1wHn2hZMxsJoOKY/1wM8PgbKsqbXn0cEhqO2RMhxD38Sy3WP2jt6vR5N0xBCIE1T6rpua4zoMhillFJKKaXUbjPQ3T4BAmAkgglYEzBSY2XEO995Jne43RbyhG6eiGBMe4FYxIAx3eeZxtF4mX6v4KCD7sqPf3SODqbWCa0Bsk685EUvpB6NSdK0LeDTPcWyvIcAQSLOWgb9gqYe4wxs6MNcAae9/KUUvYzRaETTNDjnpgVRjTHtrBKllFJKKaWU2g20S19mNqRNfBDbpTESsSKU4yUe99gTOPCOW4gBvA9gPBLbWouxvSewuumLEOj3Cl5++ss0+bHOaPGHdeLsn5x7+t6bN5x2yN0PxaYOie0MEOOgiYKdzsaKOGep6xpnUpyFm99iM9sWc359wcWMRiOstcQYp4kQ5xwxRt3JSimllFJKqXWvrdgxU7/DtNFMP4wUheOP73sop576GFILWQLGBAgVNnFASsBMSn5MS6EGiTSh4ROf+BjPOPnZmvxYh8eGWke+/bVvyaH3uRfBg0vbftPYdkXbqFyiX/S7B9bRNIHUpYwqiCn81Yvex7e+9e32SSyWLMso62ZlCYysmRBkIkaY3l4ppZRSSimlrv8hq1ytOJnxYQXAEg20HWEimIglsPde87zrH17HvvuClbbtrcFDKME5oIen/ZqYdulEAEJT433NPe5xMOee8zMdT6/Do0mtI7e8xf7yv/97Tlf3A6qqIctTJEa65jCEsFInBEAEGmmTJUcd9VSWhmOsyxCbYmxC03icTdl+RdRMAoRkWvFYKaWUUkoppa5LK0MP2yUvrno0EsmSlOXFBTZu3EhVLuN9zdx8TjlaJMvgH9/599xkv3k2bmhb3rbjqQBERCIiKZgEa6BuAlnqqJqa0HgOOuggfvYzTX6sR/qgrUN3uMMd5JxzzgFgPB6TJMmqji6zbW1jjEjXq3phAS64cCunPONUYkxYGlZs2LiFgKFpwqpZHlYAM7MsRiaZ06ufgdWoUaNGjRo1atSoUaPGqx5nkiBi23HJ1YhGIHqPRUhSh4mBvHBcdulF7LP3Jk595tN4wJ8ewqDfXQI2EEPbGQZCV26gvUA8GWMZYzDGcMYZZ/DSl75Ux9GaAFHXpzPOOENOPPFE9t133x1+fVLTw3bFQYT29SQCH/3Yl/nAhz5KWQUaL2RpgVi3kgAxk+lidDNAYpcM0RdkjRo1atSoUaNGjRo1XrdxdXXCqz8DBCBPejjnWFjYymi4xL777oWzcMAdbsPfvOQU0qT9kZZ2xryZdn6J3XgKrG0vMk+aRnzsYx/j+OOP1zG0JkDUDeGzn/2sHHnkkauTHYCIMB6PyfN8ZTaIQFMJVeMZzKXUFSRZ+xKzNGp7Xc+ucDGzM0EmWVHTfqxRo0aNGjVq1KhRo0aN12XkmswfEQgBrF3pnjlJpYwmY59JkoO25MfKzVba58YYp2OsX/7ylxx22GH8+te/1jG0JkDUDeEmN7mJnHPOOfR6PWKMJEkyTXhMWtwCRB+wznTP8KRNhtRCmre1kWOX9VybZ93RkaL5aI0aNWrUqFGjRo0aNV7XEa5BAoU28TEcCb2ewRjwAdLJteHQ1TldM9wRgRjDqq6Z4/EYgKOPPpqvfOUrOn7WBIi6IR1yyCHyne98Z/px0zQkSTJNfjRNQ5o6EA9RCE2DMZYQLWlRMB7V9PpZ+0Jg2iUvOz464h8w9UyjRo0aNWrUqFGjRo0ar/4SlmvGMjusiRKx5kq+r3RLYUy7zV5QPumkk3jnO9+pY+fdgNNdsL5deOGFp//2t7897bDDDqMoCpxzNE0znQnSdoSxED1Yg00cxiW4xFGPxxR5xqR9zHR6mIldJiSsJD7EdBkRu2aOiEaNGjVq1KhRo0aNGjVem9Fcs81EfF11S1siVTUiTdK2zgeCxAAS26RHjMQYMQasNdMx0ST58cEPfpAXv/jFmvzYTegDuZv4+Mc/Lg984APp9XrUdU2WZTMJEAMEYlNh05xQ17g0b19cYmybXq8SV9Kgk8NEEhDdz0oppZRSSqnrcbT6BxcCmbS0FUx3wbepa9KsWPND7Mxsj3YGSjuLPuUHP/gBBx98sI6Zd7NDSu0mzj77bDnggAOma9YmrZpWkiAzxO7waIgxEuOk0vFMBWST6A5WSimllFJK7TqkHasYA6tWuMhORrqz7XJpW9+GGEnTlLquSTOHoZ1RX1UNhx56KGeffbaOmXcjVnfB7uPoo4+mqipCCNPkB9Ath7GrtzVP4xDb6R3WWpIko103lxAEIkZTZUoppZRSSqnrl7mSaME6MF1hhyjtyhagTYKs3WI3/BWLBLDO4b0nxkiWJYQQCLGtofikJz1Jkx+7Ia0BshtZWFg4/YILLjjt+OOPx3vfzfxwNE3AOrvT1xMAa9okx2SVi0g7acwYizMWQaZtcjVq1KhRo0aNGjVq1KjxhoxN0yCmvYArdPUMJzNBzEzSw+x4ACQxUpU1vX5BCIGqLsmznBACr3vd6zjzzLdq8mM3zqmp3chrX/taed7znoe1lqqqyPK8zYjOJjwmB8BMXY8YIQTBOINzO7p91J2rlFJKKaWU2sUHuXbHQ5fZ0e8kOWLacc6ke+ZXv/p1DjvsT3ScvNseG2q39KUvfUkOOeQQBoMBPgacdauaSpm1TbZn+mXPHhVNBJFI4nS1lFJKKaWUUur6Ya/k4muUtsCpyMpS/pVWt3bVhd4rGv2GxuOStt7heef9hNsfcEcdI+/G9MHdjf3v//6v3OY2tyHP83bJGytNpYDVXV1CgK51Ll3f60kPqIg2gFFKKaWUUkpdf6zQXrUVs8No1gxlRSBKe+H3qg50R6OSfq/tCjMaDrnPfe7DD3/8Ix0j78b0wd2NHXroofLJT36SffbZB+kyqKsecIHp3DBZKYwaY1dE1baFgEbjirm5vmZBlFJKKaWUUrvEQFW6DjAA1k6v3QLttV2bAPiZe9iZ72u779EmUnwTOOmkkzjrve/W8fEeflypde4Rj3iEfOhDH+raQsV2PdxkLojMTCuLsX3lkK7jizEgwi9/9St+9vPfUNYpkRSQ7VsHiWH7hkKRaGTaYqptu2uwAtGgUaNGjRo1atSoUaPGPTbK7KhkuwiztQrbIevKKCS2M0CMEHxJmibc5Mb7cqtb3oING3IQCBFcEqcXgWeTIGbm35MEyCtf8Spedtrf6NhYEyBqd/D85z9fXvva12KMUFUNeZ4ClqYusdZO17wBhCbi0mTaQeaCCy7gmc9+Kb+8qMZlGwmhwdcl/X6PPM9ZWhiSJj0IBiQhdkWExASircE0XQKk/Zn6gq9Ro0aNGjVq1KhR454Q7ZqP48zH8QrvD2BNJDaB2kNRzGGsoyxL0iJDqIhhiLUlGwYJb3r933KXA2/dpki8gDNdJiVSVRV5nlOWJc6lpGlKjBFrLePxmB//+McceuihOi7WBIjanXz84x+XI488kqIopsmNCe89IkKapl3143TVfX/ys4s48ZRXMqrTdmZH8NR1jTGW1BYkLieGZCU7awTpkh/RtgkQG1OQRB8IpZRSSiml9sihZpfZmMwQ32GR026ZikQkNiRJSpoMWFoek7iCvJexPNxGkgYGA8Py4kW87a1/yyF3uz2ZhVAHkrQd5zR1TZpnACwuLrJhwwYAyrJsf4YxXHTRRdzqVrfSMfEeRFt77CGOO+448z//8z+rkh+j0Wj65J8kPWJceSEqyxLvPbe77X6cfvqLcK7CuUD0gSIr2DS/F4JDjEVMO+tDTADjuxc4i41pm/zQQ00ppZRSSqk9kLDjYoJ2+03aTYzFpSk+CAtLi2zcuJEYPVU5AvH0i5R6vMwTHv9oDrn77XEWQhC8rwHwVUWaZYgIVVVNkx/ee4qioCgKLrvsMh796Efrw7OH0VHpHuQRj3gEv//976etovr9PlVVYa2dfi5NU0IIANPZIiJw33vfgpNPejy+XGbLpjnqyrOwsIRzKTEAJiLWg226JMik9kfSbXqoKaWUUkoppdaQlcTH7OeGo5oky9m4cSMXXngB/UEB4tl70zxLWy/lyCPuzzNPeRTiwXvBOUPR7wGQZBkYCCGQ5znet8VQnXOEEKiqiqc97Wn853/+p87+2MPoqHQP8utf/9o88pGP5OKLL55+Lk1TjDHTZTDWWowxNE27dCXPc5wBJ/C4R/0xT3jswxmPFujlKRIhTXJq3yAmEk1AjO9mgLQvXEYcJqboaiullFJKKaX2qMwGO5z5MZvw6Bol7GhzaY5Lc7ZuW2DvvffGNyW+Xqaph9zjbnfkec95Mk4gTcDE0I02IlVdgW0/Srpah03TTC/4Oud4/vOfzyc/+UkdoOyB9EHfAx177LHy8Y9/nOFwSK/Xw9o2D+a9n75IlGVJkiQ459q6IFlGBLyHF7zoTL757bOBAnEpPoau44tnspbPSgLiMDHpXuc82kdXKaWUUkopdWVD0WjAJUl7UVYiqQEnHqQmTwKf/Lez2LgRxqPI3MASJQJtYdPRqKTXa2eCGGOmRVAn3v/+9/PYxz5Wx8F7KKe7YM9z7rnnnp6m6WmHHXYYzjmGwyFZlhFjnFZETpJkOhvEWks5HpKlGRLgPvc5lO9974dceNHvwAppnhGkQQDTZWzbaDGTqWwm6o5XSimllFJqT2HadrXTbZrwmMzy2NnsD0FMmwQZl2Pm+gV5ZilHC2ye7/H5z7yTPIHood8zhBBxXa8FYwxJmoExbYtbY0iShOXlZbIs44tf/CLHHXecJj/25MNSd8Ge61//9V/lwQ9+MHmeTxMfIYRp0qOqKtI0xVpAIr4K2DRHLFz4O3jxy17F2Wf/Ao+DJGNauIi27RWYaf9uMTr7QymllFJKqT1npLnmAui0xsdkCLqjagztfaIBcQYfavrOMFzaxv433YfXveY07nz7eUTAWYg+YDPX3s/AuCop8j6R9kr/7OyPc889lzvc4Q46/t3DaQ2QPdjxxx9v/ud//qc9EKzFez9NhJRlSZ7n7fIYseCFJMuxtl1Sd+Mbw+te91I2bHRYW1NkDhMFR0JqM4yxhODxocJYnf2hlFJKKaXUnmftOGAy82Nlprn3Nd7XOGcwBmIMgGCJ5KkB8ey1uc8pJz2RA+8wTwht8gNAogeZNF8Qirygbsrpz50s9b/00ks54YQT9OFQmgDZ093znvc0k6JAk2UvIkJRFDRNw3g8xjcNuBREGC0ttS9bBjbOw9+98RX0e4bh4ta2/VTjqWuPtZY8zzFGEAm6BEYppZRSSqk9xQ7P/VdPvijLkrm5PiEEer0C5yyj8TJ1U9IvEnwzxoSKqlrgpS/5K44++i4ApCkQBfENLs+JdduAoWkaDJClGeV4hIhMO1w+8IEP5Ic//KHO/lCaAFFw3/vel9///vdAWxXZGMPy8jJpmtLr9UjStE2iiqG/YQB4LJ66WuZOd9iPv3rOKWzZ2CchkiaG0NRkScrWyy9n816bqJpSd7JSSimllFJqqigyLrvsMubn5xmPxywubWPLlk1s2jTH1m2XMt93hGaZF7/gWdz/frdBIjgDMQpYwaQOX5bYrF2Kb01CXXt88PR7faqqIsbIU5/6VH7wgx9o8kO1413dBerCCy88/cILLzztqKOOIssyALIsI4SAtZa6qXGJ6woYRcpqTJok5GlKCJbb3/6mODPgv7//3zjXLn+JIjjnKKsxeV7Qdp3S1x2llFJKKaV2e9P6f13x02m725UxgUgkBE9RFIzLIZs2bWBpaQEfGuYGKdsu+x2P+bNjecqJR3f1PgSRSJJC9A0xBJKigCBgDdZZmqYhy1KGoyH93oDTTjuNN73pTToIUVOaAFEA/PjHPz4dOO3+978/zjlCCDi3+vAw1uJDTZbm+BBwNsFZQ1PD3e56K8ajyH//9/fZsHGOqqrAOcQkGJtsv/xPKaWUUkoptXtamwCZbkyjc23nyaoek+cpPlRE8QwGBYvbLuW4o4/gxX/1ZCRCYiFLDImLNHVFkqVYlzEajrFJ27RhOKwoigJrDFma8n//77/wrGc9S5MfahVNgKipr3/966fPzc2ddt/73pe6rtsXpK4TTNWUWGexNsFgCD7g69DWDQESBwfe8Y6MxmPO/dnPGFcNdSNs3rwvC4tDXGJ1/odSSimllFJ7BLPS+na77i8GiBgDja/IsgSILC5uZd999+b3l17MPe5+Z1736ucxV0CaQFmOyFIDBBYXFun1+oAlzVKsga2XL7FhYx/T/aizz/5fHvCAB+rwQ+3oyFRqtQ984ANywgknEEIghECaJUQita/Jk5xqXNMreiDgq4YkT4F29plYeNFL/54vfO17FL0bsW3Rk6YZSI0l6M5VSimllFJqjxhpTrqz2B0OPUejIfMbBlgbaXxNnjvKakSvl/H+/+/t3HxvGC6M2LipD3gg4GNNYjPGZU2eDTBdSUvTfeuy9IyGC+y1z946zlU7pEVQ1XYe/ehHm//+7/8mxkiWZUSJSAxkSYbB0Ov1aLpqy0mWgkREIs5CVcLLXvp0bnermzJcuoS5wmBiNXlpYvspcDt4kVwbZ2+iUaNGjRo1atSoUaPGXToauvKBsvpcPmLbzcCGDRsxxjAaLmNNw/LiZey7ucdHPvB29t3cLlXYuLFP8DUAPngSmxEFekUPYyZtc+nqDcLC4lZNfihNgKir7+53v7v5zW9+Q1mWOJPg64CjXcYSfE2aW7ARXMSHgHGWGCF1MCjgn9/9Km53800M7GUM3DJWAtUoUuRz+AbK2lMMBgzLMS5LiSau2trkR8RMt5kXUo0aNWrUqFGjRo0aNe66USISPM4AEvA+EKQ9q/c4rMvxEonRkyZQD7dyu1vszd+/4RXcdCP0TJdQMSBYwCIxAVLqKgC2rTlowCUQRagbz0Mf+hAdyKkrpNkxtVO3uc1t5Cc/+QlJkoBADIEoniRN8U3FeDxmfsOmNiPrI4lLEKAswdm2R/d97/NQTLIB7/bBuHlGoxE2NTjnWB4vs2XLFi677FKKIttBZi5iuilzRjRXp5RSSiml1PrQXsyMMSLGYl2GcTkhQhMDIoHEeHy1yKb5hNSU/N1rX8Hd7nLLNmlipV3XYqCqKvI8v8Kf1jQNj3vc4/jwhz+s41t1hfQAUVfokEMOkW9+85skLpl5gSlJs/bjuqnJ0gJoZ4BEaZMfVQV5Br//fcNfPvlUzv/tEFdsoigKqrqdxmaMZViO2bx5M1VVrvq5VmhngYjVA1UppZRSSql1xjlDXddYm+CSFB8hhADWINHjpKHft4R6gdNf+gKOeuBdcEAMgnUCWDBMu1OOx2NEhH6/T4wRay11XZNlGSeddBLvfOc7dcigrpReVldX6Lvf/a555CMfyfLSEjEEECHtEh4AWZoBEe9rIOIcVHUkSQAD++yT8opXvoyb3Xwf+j1LXS2RppHRcIm8SNi8YSOj0QjErWyTBS/Tf7cNtGabaWnUqFGjRo0aNWrUqHHXjVEsxqZYlyAihKbCEsmdoUgjjobMRp7/7GdyxJ/chdD1S6iqMRhofNMlUhwxRnq9Hv1+vx3EWkvTNGRZxmtf+1pNfqirnpjTXaCuzE9+8pPTY5DT7ne/+5OkCb5psM4yHpekaULjG9IkwxhL8JEstYiBEMA62PfG89zt7vflE5/8BGkCo/GIzZs2EaPQ+NAubzEziQ8sgu26hq+UU4rGYDRq1KhRo0aNGjVq1LjLxxBBcG1SRCLWGpwFIxWhXmbDIONRf/ZQ/vxxhxMjZO2qe5LEYIzBdYkTYwxN0+BcO3QdjUakaYpzjrPOOotTTz1Vkx/qKtODRV1lbz3z7+XpzzgFgCgeay1CQESwJqEsS4qiT1nW5F1NjxhokyEWPveFs3nZ35zO3IbNbL1smby3kbqBNOsRgyWa2QlJbQcY28W4g44wSimllFJKqV2QWMRYjDGEpiaxQppYfLmMSMmGQcJjHv0InnriQ6hr6OXtyLRuavI0ofEVaZITguCcw3vffq8QsNaSJAkf+tCHOOGEE3Q8q64WnQGirrJPffpTp9/4xjc57Z6H3AORLjEhAWcdICRJikgkTVPK8Zg0TbHdpI5xDQfcbh8OvONd+eynP0OSphgsvd4cy8tjEte22DViunkg0lV+bifSydoeWkoppZRSSqldkpi244t1jhhrEgdGGny9zP4324eHPOhPOPmkh2GBxIE1bc1TJIAVnHU0dYO1Dmst1lpibMcZ1lq+8pWvcMwxx2jyQ11tWgNEXS0nn3ySedvb3oG1bRFUibOvO7EtbESk18tpmgZBMAYGeTvd6NA/OoDnP/9U5vopvhmxtHA583M9DB4jEUN7f0ME034OQlsUFdGoUaNGjRo1atSoUeMuHg2RxlcIgcRZnI04avIkcM+DD+TppzySxLSJD4kRRNoxgBUMBkHIsgznHE3TrBqPXHDBBRx++OGa/FB/kER3gbq6nv70k83BBx8k9773oW2L3C75AZAkdvrvNHWMxyN6vQEWWB6WDAYFxzzkvvz2gov4+Cc+y3AE0Y9x1hGnGeN24oeYNhEips3VtZ2wDMaIRo0aNWrUqFGjRo0ad9UIDPIU8IhUBF/jqDnygYfzwhecSGrBN5Ck4GxcdVk+SsQai/eeJMlI05ThcMhgMOD888/n1re+tSY/1B9MDx71B/vud78r97zn3btZH22rqxDbdXkrxUttVxukQLCMKk+eJ0TgrLM+xdvedhYu7eNsDiSsWuhi4jSZEkn0cFVKKaWUUmqdqJqaJDEkTkiM58FHHc5LX/xEmjGkrk1+YCKEBhEBZzDWdef+BiNtm9sQAv1+n6ZpuOMd78jPf/5zHRSoP5gePOoaOe+88+S2t701QJelXZkRMh6NyPO8XfsnEYzD0M30EBiN4eOf+ArvPuufqCqPbDchKc78S8vVKKWUUkoptV5kRc7i0lbmBwV3P+jOvPTFz2WvLRAbSNP2DL8aLZMPetOhqQ+eGCBGKIqCpmlI05S6rrn3ve/N97//fR2/qmtEDyB1jdz2treVz3/+8+y///4YI4QQqKoKROhPX8xa7XS2jCjQ1JBkUDfwrf88hzTNQSyry9KsJEDE2O7rSimllFJKqV2dDzWbNm1g6+WXcPe7HciGeWgqKHIIocYlllBXuCxr18AjYNx0PFCWJVmWYa3lsMMO46tf/aqOXdU1pgeRusYOPPDO8p3vfAdjhH6vP3N0RZq6xhhD0qZ52ySGaXt8176t9mxnJ3fI6gPTCkQD2gNGKaWUUkqp9aOu2w4vzrUn96Plirm5HANEqQHBGgsYyvGYougDjuDbi6AutQyHQ0444QT+4z/+Q8et6lqhB5K61mzbtihz/QFlOaLf72OsJcYG6xzSeIxLwRhiCEQE1y2XGdeBPFm9xGVt11vRI1UppZRSSql1wxkYjRr6/ZSqbsjyFB9qmrqk3+vhgydxSXul0zgQy3hUkqY5SWoIUfjrv34Rr33ta3UkoK41ejCpa83973+4fOXLX6Ku2he4qizJi2LVbZq6JkkSjGu7xZTdbVYOxG6Zy9opHzMFUZVSSimllFK7Mks9qsh6PapqRF4UhNjgrCUiMw0TYDwe0+/NIQGaJpBlKSHAc5//HM488006XlXXKj2g1LXqyAccJZ/93GcAqKuKLM+RGDHWEnzEJW2CYzRexjlHnqc7fMGckslRGrl6C2HWHtq6iEYppZRSSqlrPHyUtafas+fZceV8XgziPSbJwEQEmRY1DRKQCIlLCD6SJBmh8bgkoa4CZ5xxBq941ct1rKquiyNYqWvXEUccIZ///Ocpy5J+v60JIiLEGHGuXepSVRV5ngORGCPWzhY43T4BItu9qMbpARwRLGbmfjtOoAgRM/MpjRo1atSoUaNGjRp332j/4Nu359px5SRcTHf67Vb3LJickE9na0sXLTEarDHtfUykaZp2JrhZPQSta0+WZQCMRiP+5V8+wFOe8mQdp6rrhPYWVde6888///Rzzz33tMc85jE0TUMIgSRJViU5Ju1yR6MxWZZPXjm3z8mZ1flks+o2Zjp9zuws+TH9PpPbOY0aNWrUqFGjRo0ad/Nop+fNVz3Onn63dfvaU2gDxradC8zMOfaqO5jpx9J9P2ssIbYXQq11OJdgjMX7gPcBMHgfyLKMuq5xzvGBD3yAJz3pLzX5oa4zie4CdV344Ac/aObm5uRd73rXSgIjtrM2qqoiyzKccxRraoTsjG+atm6IaRMcIiAScNZijAUJO8h6zIi2TYIIGjVq1KhRo0aNGjXu5lGu0fQRMRFrQYxgRFZfo5TuZNysvZZut/u4m/yNiGCMQURwzk0vhjrnGA6HDAYD3vzmN/PsZz9bkx/qOqUHmLpOnXrqqfKmN70Ja+3MshcIIdA0zVVOgOxwVseqI3n10piVG0bAgWiuTymllFJKqas+UmyXkO9w0Ciw3fJzs/osXGaWuVdVRZIk0+Xw0CZFvPekacqHP/xhHvWoR+nYVF3ndAmMuk595zvfOb2qqtMOO+wwsixjPB6TpinW2unSmMmUtys0m8+YSYKIQAy06wu7RHfbQ3ey/rBbOCNuzSuyRo0aNWrUqFGjRo27cbwG2pIfQsAQjYDE2VUuM6vXZfUSGJn5ku2qiXSJkMn5vohQluX0YuhnP/tZjjvuOE1+qOsnr6e7QF0f3v72t8uTnvQksiybTnObCCFcaQJEQmyn34lFrGFV7SSz/Qt2yzOZy2dItBGMUkoppZTSkd5VIDNn0paVwqh2prHAqqohkqyuD9J9uq7raYHTHZ33f+ELX+ABD3iAjknV9UbXBajrxcknn2wWFhbkGc94BoPBYNoVRkSmawCv8PXbAeJWXssjhNguczQWYoQ489IpBjDJmtf/qA+EUkoppZTaMwjtMnGxq+NVumuCdDVBpvcQSwQSM/nmsyfr3fef/dKatMZ4PJ4mQ5aXl/nqV7/KQx7yEE1+qOuVJkDU9eZFL3qRybJMnva0p9Hr9XDOTdcDXpk2W9y+qA5HDYsLJeNRTRCHs1lblZpuhcwkQ2JmWnGZiBFNgCillFJKqT2F3cHn1pwPG2lPntdEISFKhsFhTE2MDcZU9IuEDfMF8/2cq7LmZlIDJMY4rQEyGo343Oc+xyMe8QhNfqjrnR506nr3mte8Rl70ohfhvd8u+TEplOq9R0RI05SmaXDOEDE46xhX8E/vez8f+egnWRqWWFNgXUY5rsl7fWrfdYSxhiRxDIcjinSgO14ppZRSSu0h7HT29OrR30qywjlH01Q0TSDLErKsoKrGeB8piozGjymKlLpaxNmAtTUHH3Qgb3j9qykyg0jdzsbuFsT4WkiSdFV73BACIYTpzI+yLPnWt77FEUccoeNQdYPQA0/dIM4880w59dRT8d5Pl8JkWYaZKe4hIlRVRVEURAIGw7Bs6BU5IcI7//EjfPBDH2M4auj3NrE8qhjMbSKI5fLLtrFpy2ZGo5J+v4+vw2xxEKWUUkoppXZjtuuCuLZ/7exHgRgj8/PzeO/Ztm0b8/Pz9Pt9fn/xb9m4IWdcbaPIHInz3OteB/HK05+HCPRyMERCbDAC1rnuZ1qCj8QYiYRVHSCdc3z1q1/lsMMO05NydYPRg0/dYM444wz567/+6+nHTdOQpumqNlmTz43KIb2iRx1AoiVJYWkI//ZvX+DdZ/1/LA89xub46DC2YG5+A6PSMxqVDAY9JJRXec2jUkoppZRS65pYwK3U5Vg1AmzPiUNopjMzmqZhNkHiLBhfkqeCb0oeeuxRPOtZj2PQ62rwCbRNXnz386Qt2jcpymfMqpHmZZddxle+8hVd9qJucNoGV91gvvjFL55ujDnt8MMPB5guiUmSBO/9tF1WjJ4szSirkjzNsA7GZaTXN9zlrremKPbi+z/4AUmSk2YFIoa6joRo6PfmKesK4wQxAcEgJmrUqFGjRo0aNWrUuBtHQYxBoPt3t9nQ/TsiCEmaMByPqH3FYG6AGKFqKubncny9RGIDxz3swTzvuY+hyGA8hixpmxBIjN3sbaGua4yxGGeY1B4RpFvK7vja177Gscceq8kPpQkQtWf78pe/fPpgMDjt0EMPJcuyVZ1hJsthQmiTIZYIIjjryBJDDG12+sAD9+emN7sN3/jGN/Ax4kMAA2mWtzVFsgyRLjs9LQalUaNGjRo1atSoUePuGs3M1tnBbGjva9K0nXntvafoZeRFxrbLL2HfLT0ecfzRPOvZjyR6cAl4L+RpV/WjmwoiUbDWdctg2nkkZVmSpikA//Ef/6HdXtQuw+ouUDe0F7zgBeZtb3sbdV1jre2m4K1IkoRqOMK5BGsNTV0iocFI24PcGjjqgXflFae/mFAvMTdIqcolkIrERryvsdIe7laMRo0aNWrUqFGjRo27eTTd+W834DOx+1iw0nZJtMS2S0uMuMSQJZZt2y5nces2tuy1gYc89AhOecYjGA2FLG9Xtcz1DRJX8ipNXWOsnSY/oheMMfR6PQC++tWv6swPtUvRGSBql/CZz3zm9I0bN5528MEHUxQFIsJ4PG4PUmdJXNquLYwRl1ja2XYeZy0YQ/Bw45vsxfEP/zM+9MEPkGQJEiMxBNIkaTvhYjBi24y1Ro0aNWrUqFGjRo27bRSMgDGCEWk/np0XYgSikGWOajzGWugVOaPRMje7yU14wuMfyYl/eRS+qekVCdZA8A0hBNLEMB4NSbOEGALOpoBhXJa4JMVaw/LyiH//90/qzA+lCRCldubzn//86caY0+55z3tSFAVpmuKca7u3VFW73gWZFlWyto0iEecsqWtzJPe4+z348Q9+yNbLLyV1k0lOSbsZoU2FxC5xLV00bQuv7o2j/YyAibOdvLrbiUaNGjVq1KhRo0aN12E0XQPD9nzXyOzXZc157I6jEYuRdpmKMV2JUxOnrXAFQUIgL1ISK5TDbex3o8089oSH8YTH3pvUQOra+1blMnmWkziIUabFU5F2+YvECMaRJJbRqOSzn/0sj3jE8Zr8ULscPSjVLufUU0+VN7zhDdN1gxIioWlIsqRNfhAJdY3L2xfeGMHYBO/Bdim9Sy6JPOc5z+OC3/yO0mf4ZB+amJK4iNBgxHcZcXAuJTYWsAgWiIj1YBowoXtxT7sq2m2hJ40aNWrUqFGjRo0ar7u40lAFwIqsJDNkx+fQ0t1eACMGF5O2/WzmCMZT+Yper4dxloWFJYqsIDGGLIk05QJ7b855/nOfxhFH3K2tImJgXA5J07y7qNieJ4cg3SxtR4zdMhpWWt1++tOf5sEPfrCOM5UmQJS6qh73uMfJP//zP7O8vMzc3Ny0K1dVjciLAog0dU2S5l2xVAvtTD6MgYVtgTx3PObRT+b322qaZD/GjSVxgg8VRhqyLCE0kaYJpEnR9UpvM+NiAhhPtG09EhvzHbcRU0oppZRS6loWu9kfMv0/uO582EjskiIrSRJZE41YUtsjNp4yVBgHNnWMqxFg2bhxIwtbt7HX5nkWt13MXvMpZ77xldzpjvuRJhBCSYiQZQXQtsmdJDpcV+/De48xBmstIm3Hl49+9KM87nGP0zGm0gSIUlfXn//5n8v73ve+9gW2bhMRSZp2Fawjvmnaj2kz0iEIVR3oFW0iwzcQApz0jJfxg//9DTbfiEhbmCk0kRACWZ4jImueCjJTJbuLk+SIUkoppZRS10sCpE1oTGZ8rE2ArLCrZn9MNFXTFSO11L7BOEeaObxvqMZD5udyhsuXc5MbbeJNb3g1t7nlPHkC41Fb9yMiOOdommY6MzuEgDGGqmpnk0xmfQC85S1v4ZnPfKaOL9UuTWuAqF3WD3/4w9P/93//97QjjjiCubn5aXVp3/i2VW6aECUiRGLXhzzNHNsWhhRFhnUQA9z3jw9j2/IiP/qfH1DkOcRIFMumjZu59LLL2LhxE943K2siJ+8yYruniD5NlFJKKaXU9cd0yY+uBMi0o0v7qZU0hxjb1sSbSZxIt6Wppa4b0iwlTTPGw5LgPb0sJUkChCH77jPgHW9/A7e65YAYITTQ67npzI6yLMnzvD0H9x7nHNZa0jRlOBxOv/aGN7yB5zznOZr8UOviuaXULu2YY46Rj374X9sER9olI6wQY2RpeYGNGzYClrr2bUEmgeG4ochTRMA4GJbwmv/zTr781f8ECoYjSLMeea/PpZdeQjHogfG0Mz4sNiaAw8Sky4XUrM6pK6WUUkopdd2ZzAKxsvMBnKypFzK9qRGsCYgIoRKMSejnAwyeulqi8YscfNBtecc7TsO59hQ4dSCh/f6uK303meERQvu9kiQhxoj37Xn3cDjk1a9+Na9+9at1XKnWBS1qoHZ5//7v/27ud7/7dS/mBrppd9ZaNm7YSIiBGCNZ1iYrokC/n2ImR3eEXgavfPlTefgxD6AabaWXW4bLC2y97FI2b97MdKnL7JtLbCtnt5s+VZRSSiml1K5hthdM3MHyF4CqKSn6OWIFiQ1ZaihHS6SJ8IDD78273nka5bAmMZC6yGi4jHNgLSwtN4gwTX5470mS9lx7eXmZLMsoy5LTTjtNkx9qXdGDVa0r27YtyoYNc0BkcXGRjRvnARiXY3rFgLIsSdN8uhYxhvZFPEam0wPPet/n+Kf/71/xpFRBcElG5ZuVmttisdLO/jDSfh+xTVsYVSmllFJKqetlpDapRWd3OoSb7RSzsgymrWeXpo7oa4xYeklGNVwm+DHHH/cgXvSCx5Cm4JtImgnj8ZBeb4DBdfOhwTeBNHXTWSAxRsqypN/v0zQNT3ziE/mXf/kXHU+qdUWLG6h15Utf+vJpRx55FJs2baIocmIUfGjIs5zReEiv18fatoVYjIJ17cJJiQFrLU0Dd7/Hbdhrn5vy9W9+nbzIWRouYZ3rkh0OK7ab8dH2YG9rgwRNFyqllFJKqesxASLbJTxmP45mx/dp64cI4j11U9HLUobLl9MvDM845S95xsnHABBqoSgsMQTyLAPAB08EjLFYw7TgaZqmxBjJ85yqqnjsYx/Lhz70IT07VuvvaaW7QK03N7vZLeRrX/sK++9/87ajSwg4Z4gSscZSViVFXlDXnjRNMZM8nxdIDIsjTzFI+P4PfsML/vp0xiWEkIPJ8I1BomHj/Aa2LVxOniaISDt1UHe9UkoppZS63kZqM0u0p7NAzDTxEUKg1+vhfcNwOGTDhg1YC4uLi/QGBb5psFZITEOeBE5+6l9w/MMOo5eCeEgc3YVCiIT2wiGRRjzOJPiqIc97q36lxcVFHvOYx/CpT31Kx5FqfT6tdBeo9erss8+WO97xgOnHPtRtLZA0YzgaMugPAEtT16RJARGqpibrZdShfR/5/WUNT37qc7j4wiWMnWNcBvbZez8uvewSBoMeiKf2FS4pdIcrpZRSSqnrcaS2NgGy/cyPuq7YsGEDEhqWlhZI05Q0bTu1jMsRWzbPMxz+nled/iL+9LA7tZcFAyQWpAGTzIwIDQi+vQFgSGmaQJqmeO+56KKLuP/9788vf/lLHUOqdUsrO6p168ADDzRf+tJXgLYtF2LJ0oKqahj0B1RlCcRp33IsxOgxRDLnCb7iRnunvPPtb+SA294cEyv22jjP8tJW0sTgrBCoKXrZ6jcgpZRSSimlri/d7I9oVic/YtMwKHpEXzNcWqTIU/q9jOAroh+Tm8Dy1ov5wD+9iz897E5ED8vLSyTOI+Lb5AeAQPRdzTwg4ql9ex49sbi4yBFHHKHJD7XuaQJErWtHHHGEede7ziJJMkSE0WjU9SO3bUtc6DrHQIxCb1AwLofE6OmnjnI05mY3ynjTG1/FH93jQJLEI3FEr5eC8RgD27Zt0x2tlFJKKaV2KXNzcwyXF6nLMfNzPXpFigk1+IpYj7jTAbfk4x/5Z/bdO8cBzgY2zvVofN12S5ykMgRs0jYOMFgsCVlSECOkacp///d/c9BBB/Gzn/1Mkx9q3dODWO0W3vGOd8qJJz4JZx0SI3Vdk/cyECGEAKadCmiMIITuwDfEaBmNPb1+hgi85GXv5Bvf/j5lFak9pHlGFBBxO6jArZRSSiml1HU1UlvdBSaa1UM3X1Vs2jgghobxcJEiS6mbEf0i55B73pUXPv/pbN5gyQsYjioGgxyDp/ENaZICdrrEBSD+/+29d7xtVXW3/8w5V9n7tNu4lMulSAsgKEhiIAZRX1HRWDBWjLEkMaBIosbKm0RFJXaT2PJiiS3ys8QCVhRLiEoRFKQonUuRcts5Z5dV5py/P+Zaa5dzbqNeuOP5fO4d5+x+9t5rrTm+a4zvcA5w6GbJq/nFLy7gqKOOkpxReMggU2CEhwRnn33W21av3vOtBx5wIEmaoJXCWR+KP7RG6VAFkhc54NHa0O9l+NLRnmjhyqB6P+EJR5C0lnDBRReTlQVKxUxMTFHk0gIjCIIgCIIg3I+MTYHxQwKI9p7JiRazGzZg84x2GmOLHso5Hv/4o/nHt5zITksVSoFWEMURCihLi/ceYyLAoQ0oPN57sqxPkrQATZ47vv/97/OEJzxBxA/hobVZyVsgPJR4yYtf6j/5yU9ioqFqDeWwOHpZj8l0ErzCFiVRVKndeY5OE1DQL0DF8POLbuXUf/wXerlhdmOXqfaEvLmCIAiCIAjC/ZipLV4Bon0QRrwrKfIuq3ZdycYNd0GZc/Kr/pYXvvAxGAVllpO2KuMQHeFDUTQApbcY7XEUaDQOhyEGIm6/dS3nnnsuJ7z4eZIrCg+9zUreAuGhxp8f/xz/la9+GVuWmCjCuqIa7RWh0Sg0eMg6fdKJFuBwZYZOUjya0oNVcMlla/nnt76bDRt65AU0ljlDY8g2jwzOFQRBEARBEO4ebhNLTY1DeQfK0UojNtx1G6t2Xclr/u6VPOlJD8cW0IpBKQdlD6IUiMBDacF7T5QoPAUeFypA8Bhi5uYyPvrvH+dNp/6D5InCQxL5YgsPSY455hh/5plnsnLlSowxYfqL1qH6w1qiKGr0ibIoiBKD9wqlFNZB6cBEcOWVa/iPM77Azy6+ll6uiOMUo2OKwhGZhDhO6ff7qFpOb9yy3djkGBFDBEEQBEEQhG1I1IymdA5vLc45Im1IY4NWFmdzDBbvM5bOtHnbW9/CH/7hnmjAVWNunS2JIg3e47xCa421HucccWwoy5woirDWNuvjN73pLbz3ve+VHFF46G5X8hYID1X2339///Of/5zJyUlarRbdbpeJidDK4r1HKcX8/DyTk5MopSjLMggjQL/fp9VqAbButuAfTv0AV11zG+vXbqA1MYn2EZ1eH+UMrYk23oJXoLwLEY9XDuXBK4cMXBIEQRAEQRC2hcI7jFFEUYRWHuU8tiywZR9czmRqmJiM+e+vfIIoDqNsJ9ph1ZkXfdK41TyW90H4MCactBteF9dr3xe84AV84xvfkPxQeEgjX3DhIc8ll1ziDzjgACYmJrDWolRQwOuRufWBoCiKxgV7+HcLZA7O+NR3+a/Pf5ENGzvsvNPOeDT9bkGUxLjS4wGFGxJCHArwaEDLFBlBEARBEARhq3AKCtcHHEopjNJhsqEraaURk62UA/bbi3/79zeTd6E9EcxOnQNb9kmSCND0+zlaa5IkASDPc+I4DlXPNhiiRlHEkUceyfnnny+5ofCQR77kwg7Bpz/9af+Sl7yEsiyJ43hE7MjzvDkoOBcOMqoymXLOgdJYD7mDC8//Ld/93jncdcd6pmeWkkQJWVFWlR40FR8h+koAATCVQIJEiRIlSpQoUaJEiVuM6URKnudV+4vG+3BabWZ6khXLlvDylx1PmgbhI89AUZC0DGXRJ4pjrAVjwnp3vNLZGEMcx9x8880cddRR3HzzzZIXCiKACMJDidNPP92/7nWva4SPWvWuxY7hSpAoiprLqQ5CDnAF6Ai0r5T5PsTp4PcmMvo7DAskEiVKlChRokSJEiVuPnoHeQlKQVLZzTkLWoNRIZGbn7dMTRmc82jj6XdnabXTsGD1uhn7Yq0lz3NarVazxv3BD37AscceK/mgsEMRyVsg7Ci8+c1vVrOzs/4d73hHNf88HBDqnsi6+qMWSMqybCpFnHMkSYKJoJ/lRDomihUmhV7f0U412ocDkSYcuIYjCjxezmhIlChRokSJEiVK3HL0CqUgiofOXHvAgK/WnGUB7bYB69GRwmZdWhOT4C3YAkyL+fl5pqamMMY0J/h6vR7f+ta3eO5znyvihyAVIILwUOfoo4/2P/3pT8myjDRNm8udc02LzHD1R6j9sPhKIEHpUNbRTH5Z5KgFYQpMc7mWo7lEiRIlSpQoUaLErY7eV2KHH4geSi3M4Mo8w3tLnEShNASHzUtM0mbciH/dunV89KMf5R//8R8lDxREABGEHYUjjjjCn3POOUxPTxNF0UhfJIQywbpCxLkSrT0272OSJBxIShfaZ6KkOiKNb0pudCvzUmwlCIIgCIIgbEumVq0nPTRChh/N5MqiIIpNWG/asplGiDbVfTTOObTW/P73v+ctb3kLn/70pyUHFHbczUreAmFH5pe//KU/7LDDgrlUFOGcoyiKkcqQfr9LqxUvstkMKer1hBflFn+iugJEEARBEARBELaKcuhnPbbmBFuUKKPRSpMXOXGkQamghTiHtb4x+r/uuut4xStewQ9/+ENZkQo7NLIBCDs83/rWt/xTn/pUer0exhiSJAljxqAamVvf0rGo6LFgq3KV4OFE+BAEQRAEQRDuBq76N9xfrUdiXdlRFAVRHOPD8MLgD1JVNzvnuOGGG/g//+f/cMMNN8iqVNjhkY1AEIC3vOUt/p3vfCcMHTBq06h76yDmcfJGC4IgCIIgCNucsAUZRFcdMIOTcHNzHaamJvEKiswSpQZD5R/iLOeffz6PecxjJOcThKHtSRAE4IQTTvBnnHEGrVaraYmpp8BsEr91W1O4mQgggiAIgiAIwrYwvH7UY3FgRVc4iHW4dZFZ0tTwkQ9/lFe/+lWS7wnC1qVsgrDjcfTRR/uvfe1rLF26lLIsSdOUPM9D/6S/m1tMNct9RCwRBEEQBEEQhC1maW6RxG1hG3ZROhRgIk3et/z9a07h4x//qOR6grCpTUsQhMDq1av9D3/4Qw444IAF02HuNl70D0EQBEEQBOHuZGluLGkLAkg9tXC4Yvm2227n6U/7M355yUWS5wnCljYtQRAG/PznP/eHHXYYaZqiFttU1Fa0tNRmqIIgCIIgCIKwTQyb7y++9iwrA1S85tprrmG/A/aX/E4QNoNsIIKwGU499VR/2mmnLRRARg5A9XSYsSkxY9QdNH7sNm6R290bG+rg+SRKlChRokSJEiXek7g1666tSbp087tjy/5w0eiDjwkgtiwxUYSzli996Uu88IS/kNxOEEQAEYR7xvHHH+8/8YlPsHz58oEfCMOjxzLiOKXbnafVmkDrUYHDe4/zCqpikG995ydc9pvfkmceHbXIcovSCd6HQ+HE1CQb5zeilCJOI/Aebf2IZLIpmcUNHVq9AuV1iK6KHokSJUqUKFGiRIlbFVUVPSiH9kOnrpTDe4/WmtI5vNKoyKCUoigdeZ5jTMz05BSzG+4iTRRp5MizWfZYtQuv+Ou/QCs2IYKEtVyv16PdbgNQFAVaa4wxAPT7feI4xhjDu9/9bt70pjdJXicIIoAIwr3D7rvv7n/2s5+x5557NqKGUoosy0jTtBFDIPRj1r/XB6nSepRR9PvgHXzsY5/l//vyWbTby3A+wquYufkOk9NL6PS6eAVJuwW4cIBTOhx8h9R/PXaqwdVbcyV6jGzeXsuHKAiCIAiCsE3owSkm5VAewKM9OO1QztOanGC+0wu3NhFz3Q5pmrJ06XK63S5Z1mXl8hlmN95F5Lr88R8/kje/4e/YdeUMccRoa0u9XhtawvV6PZxzTE5OjlzWbrfZuHEjz3/+8/ne974nOZ0giAAiCPc+3/72t/1xxx0HBCXeOUeapszPzzM1NUVRFERRhFIK51wjlmijAI0DyhKMgYt/dQuve92p9DNPUWomp5bSzTKcV6Rpm36e4b0PAktZLCh7bASRWuNQY5u1iB6CIAiCIAj3IFNavEVluFJkrjPPrrvuSq/Xo9vps9POK+l0OmzcuJH2REKaaPKsg1E5f/H8P+fkk5+DqR5jQQXImAAybMZfn3SDUP2xceNGdt11V8nlBEEEEEG4b3nzm9/s3/WudzW/Dx+cvPdN9UddJaKUx+Po9rpARKs9UR3I4K51lre+9XQu+81vUabF7GxGe2oKfISJE5RSbNy4kYmJ1lhJphirCoIgCIIg3LeZkl2QNg17uRkTU2Q5URSFSmDnKcsSgMmplKKcJ8u77LJihtPe9k886rBdcBaKXsH0dLzQ2NTrgfEI4DyUZY5zrml3AfjKV77Cc5/7XMnjBOHubNbyFgjCtvO85z3Pf/jDH2blypVAaHspioJWq4X3nl6vx8TERHVrBxTVz5q8dBRWkaYJeFi3Ab705bM544zPsduue3LHXRtIW9PcuXYjy5YuR2tNZgucqkou1cL2ly1NmlEyg1cQBEEQBGEbcI0A4lX9n66c2HQlgBjSKOauO+9k5YrluKKkMz/PqlW7cuvN19OaVjzu8UfxD39/EktmINKgNZhhU1PvQOlmnRiesDrhNfRqyrKk2+3y/ve/n7e//e2SwwnC3UQ2HkG4m+y7777+85//PEceeSQA3W6XVqs1YoJqrUUpT7e3kanJCfKiJI5TIMKh6fUcrbYmy+GCC6/nrf/8TvJckxUKrWIm2jOs27CBeLI98PgYZmjM7kAUGfUKEQFEEARBEATh7gggDvB4TFiHeY1XetB37D2+tKRJjCsLIjzKW8o8Z8WKCZ7yZ8fwly95NjPT0O/C5ESQOJQDT4mq13ZjAki9bCvykiSJ2LBhA+12m+c973l885vflPxNEEQAEYQHjq9+9av+KU95ChMTE1hrm/LEbrdLu91GKQ9YwFGWBVEUYx1oHQOawlYVkBrKAl78l29ibrbPunUdwBCnbXrWY5VqekMHfh/1+N0hkUM59JhhqhI/EEEQBEEQhG0TQKo1lifCKfAoXCVWaB9G2UZKkUaG9WtvZ5cVy8j788xMTvBv//Ye9t6vTRSHhEt7sCVEEczPb2RqanIoI1tcAPEuVIxcc8017L///pK3CYIIIIKwfXDSSSf5D3zgA7RarcYTpB6Z2+t1aLdTOvNzTE5NAVDaElAYk+CcBg29PrRS6Ofw+c9+n89//kzwhn4ONm7hq1nwrhJCgggyKmyoajxbsFv1IoAIgiAIgiDcAzyqansJJ6DqilztwWiPLXpEzjIzk9LdsJ7/84TH8E9vOZH2RFimFSXEcRAzcJa86NNqpWMZmR74fzAQQBTwwQ9+kNe+9rWSswnCvYRsTIJwL/HEJz7Rf/KTn2TPPfdszFCNMcEMFYWzFh1V0dRGWp7SerwzmNjQ7welP47hJz++io9+5OOsue1OCt8ibk/Tme+RtidAR2RZjjYxWkc4C0r7EQHEe4cizKhHGdncBUEQBEEQtgGtEpQyZGVBkiTMdWZpt1OcK4mMJo0Vnfn1tIxl5U7TPOeZT+MvXvQUUgPWgYnC8qvT6TDRaqMM+CIPEwLjGFcU6DgGpSmykjhNKMvgCbJx4yxveP3r+NSnPiELOEEQAUQQtl9+/OMf+2OOOQagaYmxRXUQbLY8R6/fod1qV0KIZr7TZWpyCgesvavPihUt1q/znP7eD3L+L68iz6H0gIpROg5mqoUjSdv4qidGoxYIIPXjy+YuCIIgCIKw1fIH+Ihet4+KDEuXLWFubj1xYlC+xJZ9nO2TRJ599tqVt/7TGzjwgBUUfWi3wiMUeUmcRE1FrssztDHhbFfF3IZZJqeXoJQK/iLATWtu4cUvfhHn/c9PZPEmCPcyslEJwn3A61//ev8v//IvYYJLVpDGcTjwWdARzHfmmZqaABxFWaCUQpsYhRl9IA+9Ar7/g4t5xzvfQ5xO4rxmdr5PnEzSnpghLz229GOHbIdSNAKI87KpC4IgCIIgbD0GfESatCjKHOdzyjKj3Yooyy7e9milmmc/62n8/SnPRSvozBfMTMV0Ox0mJ8NJrn6vR6vdXvDo/V6PNE1RKhpe9nHxxZfwh3/4KFm4CcJ9hGxcgnAfccwxx/gvfvGL7LbbbuCgM58xOZWS55akVbXGKI/zJbpS/K2HoihIkzbeg7WOKNJYDzfd3OP1b3gLt952F6WNcDrBOoN1GqPjxhukNkEdtMSIACIIgiAIgrCtaVKeOaanl5DlHWyZkcQe5XMUBTuvnOGt//RGHn7QKjzBx01DVeGbUpYFcZRWrc8G71xoj46igam9cyitsWVJp9PhjW9+Cx//+Edl0SYIIoAIwoOTffbZx3/4wx/muOOOaxyt8qLEe0uapmRZRprGze29942JqlKhGqTT6TE50cYD3T58+rNf5Stf+RaznZKppTuxdt0cSTqJd6rZpBWe0AqjgmGXCCCCIAiCIAjbhEkS8l6f0mZMtWNiY+nNr+WYo4/k1Sf/NXvvNR3sUVVtcO+IjKaa70dZWOIoGJ6WRUGUhDVfr5ehlMIYQxwbrrvuBk488UTOOed7smAThPsY2cgE4X7g1FNP9e94xzvo9Xq0qzLIsiwpioJ2u91MjAmEtpWiKMjznMnJSXCKdetmWb5iCRb49W9u522nvYdrrr2FySU7YZ3CeYN3qjJddSilUN6BVnivKh8QQRAEQRAEYctZksMrRWn7zEy02LDhdnZeMsUr/vrFvOC5j0WrkEhl/RwTeaLYNNNbekVGFEVEPkYrTZHnxGlY51nrMfWdge9+9/scd9yTJScThPtr05a3QBDuHw46eH//s5/9jCUzyyhLR1z5gpRlaHMpCksch6qPoiiq6x151iOJI1CGslDkpaFw0JqEt7zlDP7nZxfidIp3Gu8VzrnBBl4JIFCPzRUEQRAEQRC2jKNf9Gi1I1zR5VGPOIi3nvpGVu+myTuh5QVfoGMDuKYKxJsIhcZXzm4aRZ6XJHFEnjvysmByMqXbzXnjG9/IRz78IVmhCcL9iGxwgnA/sutuO/n//PRnefKTj8N7T1FYkiSYX9W6RW0MXnuEhIoQT9btkk5Mg9d4oNOHpAU//ul1/PNbT8d5TekinPPgNQ7VtNGgVDO3vj6ob1U9iJeqEUEQBEEQHoxZjqvWUI7gzqGbFZBecFbIg3KVd1q1IFMFXnVJYnjpi1/Ic44/juXT4EuIDbjCohOFzftorVFxBIQ1WuE8RhuUD+eh+v2cVpo0J6N+97vrePGLX8yFF/xMcjFBEAFEEB76vOlNb/Knn356MyZ3HOdcNUGm9gipqzr0IHqo5ZH1G+C9H/gI5577C5LWNHmu8aZFPyuJ0zbKaKwt6PV6tFoJUWzI+z2MMThfYmrvEK9Hn8cPqkectNAIgiAIgvCgyXIKIlMAJaXTWAtaJZgoRRNTFrZadFmSVJNnXVyZYSJYvnSGjRtv4xEPX8UrT3wZRzzqIKwNwofyYJ1HeYc2hjwP6ykTVQKI9zg8SilwHl2d2XIuVPh+9rOf5RWv+GvJwQThAcLIWyAI9z/nnXfe284777y3HnPMMSxbtoxer9e0xABhFrz3GGNQSjf+HeFoWZ3NUGHUrVaKNIEnPOHRKFqsWbOGDbPzTE5O0un2KK3DOkdWlOy8cid6vS7z83O02hMkcUxR5mhVm6fWx+Ph6EF5vKrPiFiJEiVKlChRosTtPHr6/XkAojghSVrgDVluKQoLXmOMJk1iuvMbmZluN1Ne+v05nv+cp3Lqm17FfnvvigbyrFqrKfDeYSIDCkwU45xHa0NZlphq4ovRGueCJxvAxo0beMc7TuNNb3qjiB+C8EBqo/IWCMIDyznnnOOf+MQnAlQVHylzc3NMT08DkOc5cWWQOiKANEQUZSjW0BqyEv7qFf/Er359JctW7IrzEZgJ1m+YpZWkKO2J44hOp0OSJFhbNGcnFtR4VGN1oQynPLySKFGiRIkSJUrc/iMQRQl5nmOtR5kYoxO8V2itSaOYoszodmaZmU7JerP0e7M8/OB9Oe20t7L/vhPElVdpnudEUdSsl2rz+n6/T6vVAqDb7TIxMQGESg/nHGmaYq3luuuu43GPexy33nqr5F6C8AAjNe2C8ABz7LHHquc///kApGlKv99vxA9gaDrM8GY7vOmWxFEoy/QKvIcz/uPtfPJT/4or1zM/fzud2TtZOjVBKzF0ZmfJe31ibZiemKy8R8JjOjTDHbNOOZxy8iEJgiAIgvCgS3O8M2jVxugWkYoxSqFxlHmX+fm7KMtZli2LcXaO6WnFq175Uj71qfew+6oJtIKyDFP74jhGa421FmttszZrtVr0+32ARvzw3hPHcbOme9e73sUBBxygRPwQhO0D2RAFYTvi4osv9ocffnjze6fTYXJysvH60IxXgVQD17wHpen2C9JWGw/M92Hjxi7v/cDH+J+f/opWspTSB4PUtD3Bxo1zpGlKUQYfktok1TfGqzTPapypjMH00CuRKFGiRIkSJUrcXiOUBURRjNYK7y0ei4lA6QJ8jiv7ONtl34ftyWtf+0oOP+xhWAet0OkytO6qVkWVRxuAtaHVZtjLra4I6fV6rFu3jpNOOomzzjpL8i1BEAFEEIRN8Z73vMefcsoppGkaBIkRKaLecIcFEIe1Jd4pojhhvtujPTGNB6ouWH54zlW8813vx3qNwrBu/RwrVu7Khg3zJGkLpUwwOlUuPKLyVesLaF8JIE4KxgRBEARBeLCgMSbClh7rcrSyKO1QusDbLtb3abcUz3z6k/n7v38JRgdzRMVgXox3VStyHDdCh/d+kEgpRa/Xo91uj7Qvf/e73+W4446TPEsQtkNkwxSE7ZDDDz/cn3POOSxZsqRyFR8TQDygSgZzYMDaEmMSrHM4C1GcNhJJkUNp4W9P/L9ce81NpK1J1m/ssXTZzpRO089yvArtL165qhrEN620xoP28rkIgiAIgvAgSXK8QqsWZV5gIk8aa4pynqLsMDUVsXKnKT74gdPZfdUE/Z5lsm3Isz5JElEWWWh7URHDbcfWWoqiGBFEXOglbqpDXvrSl/K5z31OcixB2F73DfIWCML2y5e//GV//PF/jjZqrAIEoBZAqFpgFGVREsUxwz4hRQk6AuvC/b5/zqW87/3/hnUR69b3SFuTeJXiUVg9VHFSiSAAxssQXEEQBEEQHkR4jbKgvCOKAZ/T7W1g55XTPOfP/4wXvugZTLbBVeNt8SVG1S00YW1VlI44TvHe4/1gpC0Eo1OtdZj64j2XXHIJRxxxhORWgrCdIxupIGznHH/88f5jH/9/7LzzTngPeZHRSuKq5DLFO4dSGqzFWo9JWuAhy3LSVtJs6d1eRrsVqkKuvno9Z37pm/zkfy5g/YYuk1PLWLdhnmUrdqKf5cx1O0xMTlJWE2KUVH8IgiAIgrA9JjNKNaNma6ECwCiPzXskGjwF05MJe+25G3//2hPZf79daCWDdhePRXmaxwnnfzzWBY8Pa4NfWj2tDwaTYJxznHbaabz1rW+VvEoQRAARBOHeYO99Hubf//73c/yzjgeg3+vQbreD+KE1Nssb4cOXDmU0KHBFqP5AF9XWbihKjTbh2P6Nsy7hv774Fa6/8VaUTujnlvbEFEmrTbfbpbAWrXWo/vBSAyIIgiAIwnaWzFQCyLD4oZRC64JYZ0ymhm5vjuc+51m8+uTnogClhk1Ox6bd+aocVoHzoa2lLEuiqiUZaHw/LrvsMl760pdy8cUXS04lCA+WfYa8BYLw4OEFL3ie/8QnPsHk5GRzFqLfzWi12xR5Sd7LmZyZqIQQUBHYwmHSEpQFIqxXOBvhPI0Q8o//9AkuvexK1m+cp9fPya2j3W6jjSFN23Q7GTI1WxAEQRCE7S6ZGar+qAmtKQVldhcH7LeKj370A0xNgregVVjR5FmPNI1H2lrGT/aU1hFFmiwLniBlWQKQJAkf+chHOPnkkyWXEoQH2z5D3gJBeHBxwAH7+fe85z0885nPbA7UeVaQpPFA+DA0Zy/CvxznczwaoxMgwvtqmK4Hr+Cy36zlrW87jRtuuJEVK3dhdnaeOErpZX2SuN3sMLxEiRIlSpQoUeJ2Fp33jSdaZAzaWN73nv/LH/3hw+h0HRMtTaTBlY44qr0+6uoP3VR+OF+ZzVczcK0N1R+1yWme5/zRH/0Rl156qeRRgiACiCAI9xevetWr/Nvf/g6WL18KQLfbZ6LVCsf+6gSGyz06UmAGhqlZnuOdIW21ACgsGA2dbjj297rQ6/VJWil4hbUOpcMCQVViiUSJEiVKlChR4vYQ0QpcmFwXylsVkTZo41i2zOAdaB30jM78HDNT03S788RRRJxUXmleg/d4pai1jzpTqttdyrLkQx/6EK9//eslfxIEEUAEQXgg2HPPvf1//ud/ctRRR9FqJXTng3kpQJHn4cDuQklIlvdQSpGkaXNo99WZkn6/JGlF9PsWrQ3NegCwFiLTnB9pptFIlChRokSJEiVuT7EWRurfi9IS14sY78j7QcwIvwYftXDGSAUBRakFiZJzjttuu42TTz6Zr3/965I7CYIIIIIgPND8wxte70877W20kpRev0O71aaf9Yi0IYoTXGnQZuBsnmV9lIE4jlHViJdOr8tEewLQdLodJiem8eFESjijIgiCIAiCsD1RnchphIvKCFVVl1kcWkFRhql2kTbkWUYSt0bu4wgjbj2DppjSOlyZ84Pvn8MznvEMyZkE4SGCbMyC8BDikkt+6Q877BHMd+aYmpwiz3uVf0cMjsb4NGz9Lox9qzposzwjTdKRXYPHozB198w9WKBAU1A63LA7fP2CvdGwK7sYsAqCIAjC9slwDcam0Fv/OMrd83XHUJbj8KhqtdNMcxk2O11EAMnyjCzLeMxRR3Hl5VdIviQIDyFkgxaEhxivfOWJ/h3veAfLli1r+lZB41xoeTFVJUiYXx8N3WYzi5Z7YyHih/Y6fgt7JNkzCYIgCMIDzj05/KvNPYja9FW1nNLczN/TFxDWP1mW0ar8z+oqEedcM0a3NjktneXTn/40r/jrv5HViCA8BJHTqoLwEOOjH/24euxjH8dZZ32LdnuSLCuw1qI1GKPo9XphgeEcWVbQbk9Wu4JN/WNomszd/OfHFiObuh2hd7de74z/Q6JEiRIlSpR4v0a3zdGNihubOt4vcv/xn/09WYMARVGQZRlKqUb8qEUPAGstSim89xRFwaWXXsqTnnisiB+C8BBGNm5BeAjzohe9yJ9xxhkYY0iSZCuqPe4j3PgeZ2iZoxbrf9Fj0cmHKQiCIAgPANs+ltahRlpiho/z48f30fvf14mKrzxC6p+NMfT7ffr9Pt/85jd5yUteIrmRIDzEkY1cEHYAvva1r/lnPetZQDjbYYxpznZorUM/7H2Jq/c2lfCh/ODnBV4fGjCMVKAsuJ0gCIIgCNs/9fFej6UeY0XoXt+nmUpRFBhj0Hrh8/zud7/j+c9/Pr/61a8kLxKEHQDZ0AVhB+HpT3+6//SnP82KFSvu/ycfFkCUrS4ohxZG9WJIARGNAOJFABEEQRCEBy5L2JJp16aoK0C28v4LBBB9r2Qr8/PzTE1NhZWEc03VR1EUdLtd3vSmN/Hxj39c8iFB2NF2bYIg7DicccYZ/i/+4i+aXtgsy0jT9L590uH1j6qFj7K6YFgAqas/ooX3EwRBEAThfswSNnPiweut6ocZPYy7xZMPv9jP+l4zRm8mvwBzc3NMT0/zk5/8hBNPPJGrrrpKciFB2NF2bfIWCMKOx8Mf/nD/jW98g1133ZXJycn750nHrD78lm4rOypBEARBeMDxKnSubjKyiPZRXe/GRJDhBhi1OfOPe2kR4L3HWktZlrRaLW655RZe//rX88UvflGWFoKwgyIbvyDswJxyyin+Xe96130uggw3ugxf6Ldm4TO2w/ISJUqUKFGixPs8bvUBfjPCiWMw3a0+jg+LH7qKmxyZew8zlXrcrfeeT3ziE7ziFa+Q3EcQRAARBGFHZs899/T/+q//yjOe8Qy01s1ioS4ZLcsS7z1xHDej4+rFhHMOY8wW10YX/PJXzM7Nc+fajSxdsoKs74lMG3xMHLeZm++SpinWFjhfEqcRYLGuAECrSAQQiRIlSpQocTsTQLSKsNY242Rro9FQeeEoSojTNrgcbaAs+ihlWbpkirvW/p5Ie1bvvgv77/swliyZCM/vwXvXPE69LqlH125J8LDWAjRtLxdffDFvectb+N73vid5jyAIIoAIghB47nOf69/3vvexatUqAIwxlGVJHMfAoG+2Hh+nlCLPc5Ik2aIAsm62wxfOPJNvf+eHrLurQxxNYYuIVjrNXWvnWL58BbPz87TaKXFs6GfzlC5HG0ccx5TF8C5LlqYSJUqUKFHi/RGV39T1g+NyHMdYa8nzHK01WmucC5NftEnD5BVncT6nyObxlBhdsmz5DEf+8eE898+fxaGH7Ivz0JntsGTJJIowuaVeg9QnXZRSI5NcapPTTqfTVLM6F8STO++8k/e973285z3vkXxHEAQRQARBWJxPfvKT/nnPe17jmt7r9YiiqFmEQDAUc86RJAnr1q1j+fLlmxVA5no5rXZCWcJr/v6tXHHF9ZSFoduFmaUr6Xb69POcdrtFbnOUtqStiF7WJYoiFLEIIBIlSpQoUeJ2JoDU42XrStEkjeh0OhRFwfTUElxhcc5SFhntiRilS5LIs9eeu/Ka157CwQftAiwckouvGmedH6k0rStC6hMxw1UhnU6HdruNtZYf/vCHHHfccZLnCIIgAoggCFvHJZdc4g888EBarRbee/I8H5kWUxShLGNYGNmUADI8CM8Bd90Br33dP3LNNTdj4kmSeJJuLyNKE7Isw8QaYxSz8xtZunQpvV6xyIg8QRAEQRDu/0TBNVEpRRRF5HmO955WOyHPc/I8Z6LVxhUFRjuSSOMpSVJ481v+gWMeeyC2hDiGsoA8z2ilCXGkKn8Qj0JVLTGhrcUYs6ANptfr0W63m9/Xrl3Lk570JC6++GLJcQRB2Ib9miAIAnD88cf7D37wg+y1115ND+4w/X4fY8wWRRDrQGvIS9Aq/Dw3B1f97jZO/5f3c9Mtd2CiFo6IbrfH5NQMrdYE69ZvxBhDdF+P6RUEQRAEYQG6OfkwbmXuQTn6/T7T09PkeZ/Z2Vna7ZSpqanQEtPvMBlBns8zPT3JK1/1tzz+cX/EzJJBxYdn8PN4FciwF1lddQo07TbeeyYmJiiKgvn5ed785jfzH//xH5LbCIKwWWQnIQjCFvnXf/1X/5KXvISpqSmKoqDVajVnY7YKD+vWrmf5imUAdHoFrYmY0oIy8LWv/S+f+6+vsubmO1i6fBfuunMjeaHYacUuZHlJafOFU2QEQRAEQbhP0V6DGj8CV3WdKogSzpUYYzAmGKjnRYb3nkiVTMUlf/PyE3jc449h1aoWWR+SFCINeV6SJBFZ3sPb4PkVRUnw8FAaFIt6jY2fkPnMZz7D+9//fi677DLJawRB2CKyoxAEYavYZ599/Oc+9zn+5E/+BBgsSvr9Pq1Wa9N39JB3+iQTrbDHcSUY6Gd9lI4wcYv5nqXdNnzsjG/wla98G6Vm6PU93fmCpNXGRIsO0hUEQRAE4b5MFEbaT119YfO7MYr1G9aSpintdovZ2VmU8uy9997svedOvOufTmbpNOR5ED4ASlvS684yMz1Nlmek9RUovAOlNXhNWRToyDSmp3meV0KLYW5ujptvvpmXv/zl/OIXv5B8RhCErd+vyVsgCMK2cPjhh/tzzjmH5cuXb9VIOjwDAxBvgQKMwvoCo2IcGk+MrZZW8/PwmU+fxX+d+TXiaJIst3gT4VVYiPmRXddCXxCnHNqDUx7tR3d1yoNXwy9sU7vE8ccV8UUQBEF4IIWI4ePX5o5jW7PM35b76aHXMCyAuPC7sijtacWabm8epR1PetIT+cu//Av23TsiBvqdHpOTwacjy3qkaQx4iqJPHCeAx5YeYyLAVD6r4XmtC9Wm9QQYpRTr1q3jAx/4AO985zsljxEEQQQQQRDuH1772tf6d73rXaRp2oycq8tSh01Ty7IkMlEoofUelKe04bJeLwuO7R600gOZwcOaNRv5+te+yX9/47t0ioTSp2g0zismJmaY3dhhYnKa+fkuJk7wOOLYYH2JwuKwGA229MTEgEZ5FxaQqq4ocWMLPA3eVAsvjUeF26oSsPKhC4IgCPez8DEm4Kth8cItEEWG5XrtNcqb0Wuq4199P+8tzjlaE1P0+33KsiRJWuR5TqvVBh8Bmn6/TxoZ4kiT532M9mhK0kTR723Eux5/8idH8JKXnsDhj9yrOZZrQDHuITZ4lVmWNQbr3ntUI4CA82FtUVeaZlnGRz7yEV73utdJ/iIIgggggiA8MHz4wx/2L3/5y2m322GCizFEUdSIIsOLnbptxrnwu9bRiJeIc1SLrtDv2+tlrLl1LZ/67Fn89H8vJuuX6LhFd75gYmoJ3U7G0mUrKAqLxdPvd1EGUJY4NuR5H7wnoTUQQPRwO42rhBldiSA6rDCrBd9AAMnZ8hkzQRAEQbh30W50qe61Hzmu+jF/DjdUPREEEFUd39xI60oQQBzee8qypDUxRVmWKKUwJmZ+fh5QFLll6dLluDInSSLuvON2li9dQhJBt7MBm3d45jOfzDOffiyHHroXrTSc6+h156qqj/r4umm63S5pmmKMaaa61MJIbYR69tln8/a3v50LL7xQchdBEEQAEQThgeWwww7zn/nMZ3jEIx6B955er8fExAQQzu4o5Rvhoy5hrQUPpRTWWqIoCou5an1WnyyyQAnceLPn3f/yPq6//kZuu/UupqaXgTdkhUebhNJZlDK0JlusX7+eyckJjFE4B9iqccaDV37I0G2sAsRrRtpgGvM3JwKIIAiC8IAJIAPhgzEBZLFLB4KD3tyhSwXj0SzLUBjKskRrTZKkeO+J4wiPRWvN+rV3MT01QRQZ+t1ZWqlh34ftycmvegX7778LS2fCMR0XqibjiGqI7eYFkPokSC105HmOUoo4jul2u9x8882ceuqpfOUrX5GcRRAEEUAEQdi+2Guvvfzll1/O5OQknU6HycnJqi1msAJzLozNm5iYqBzgR93d3ZjdhvOQOYhCyzBr1uT8y7s/wNW/u56162dZumwlG2a7LFu+kvXrN1Jaz0477US/3+euu9aybNlSvPcLFn0L8Hozu0cRPwRBEIQHYKHuVRDvtySAVMewsYIR9LiH1cixTjXHR2MM3nu8dVhv8c4RRRpb9rAuY6dlS9mwcR1pHLPPvnvyd69+JYcftgsKiAZdK1T1nHhfUhQFadLerABSlmGCjLW2qj4xjTDy2te+ln/7t3+TXEUQBBFABEHYvjnhhBP8xz72MYwxTE5OUpY5RVEQRRFxHDM7O8/MzAwARVEQxzHOhVLc8dG6dcOKp7IQAfoFXP273/OJT36Oi375K5xPyHJLEk9ROIciprCOyYkp0Iq86I0tARdfjI2sL4dFEq/lQxUEQRC2u4W63+wtPU6XI8cyPV7p6AZtMXEcoxWUrkCjULokMQV5fw6U42EP25OTX3USB/zBXixbGg7OzgUBBBUaRzUe50uUAsVQe+kWsNaSZRkTExOcfvrpfPjDH+bWW2+VPEUQBBFABEF4cLDnnnv6V73qVZx88snEsWnKbIPZWVgM1Wee6hLYYZO0snRNy0xpLa1WEqpDNPT7nrSlcMAFF17Hl770DS799ZX0Mos2LTrzGWl7inZ7kptvvY2ZpZODs2KN34caFTeq6TGwiRYZEUEEQRCE7U4AUZu9dqEAYprjmvIEA3PryIserThCaU+W9TBG432fVlxy6KH78/SnP42nHPdHaA+2Fj2aR3LBsFQNXq91wThcq+CptampcUVRNPGnP/0pb3jDG7jsssskPxEEQQQQQRAevJx77rl+//33Z/Xq1SMGZ0qpBS0w1lqsDT3HtS8IOJy1aGMG4gmwYTZjeibFAb+6ZA0f/49Pc+WV19LrlkzNLGfDhnmW7rQT871uJYDoUQHERyPPrRny+1BOBBBBEAThAV6oj7ds6jHhY1PHpnAscyMtnxpdmaIqFzyuyjyjnSZYl5HEmjzvYl3GrrvuwvJlE/zNS1/AUUcdTBqFw2Y/s6SpwboSYxTeWxQercJxNQghg2Or92qB+OG9b6o+tdZcffXVnHTSSfzoRz+SvEQQBBFABEF4aLDbbrv5X/3qV+y8884jY+9q93etFz9D5EqLDqepKPpd4naLsiiI4hSPJis9kQmnnayHuTl49cmnsubm2+h2czAJPoqwRGivcEqjXYRTqop6qPXFVSNvRQARBEEQHqwCSH2fcHCrKyB1M1JXV9GhfEm7FdHtzeFtDxPBHxzwMF78ly/iCY8/kLhqPbUOtAaPIwx4s03VR/06rHM4C3Ecju/9/mCq28ircw5rw/jdP/3TP+Wiiy6SfEQQBBFABEF4aHLsscf6D33oQxx88MGN/4f3fmQSTL1AKsuSJE6gtGBMtaizoBTWlqAVWsehhVnpxi/EObj22o38939/na+f9V2IJnBEuNLjvGaiNU1ROrrzfWaWLCPvF1jvSJIIE6kwTlc7jDHkeU5kEvngBEEQhAdgpV6L8ePTygbChzEGW+RNRYXS4bIsz6vR9DHWWoosp91u451lbm6OqYkEZ3MiXaJ0yaGH/AEveOGz+ZOjDiSKwzNEbixjGJmM5qrjtUVhUNoA4Tje7eRMTCQ4D86FCS/OhdbWXq/Hy172Mr761a9KHiIIgggggiDsGPzVX/2Vf8Mb3sC+++47WMBZS1mWeO9ptVo450JZbVWh4YsC50pMKw4zbQl9xkZHeDSl1xSlJ4nDObP5efj97Rv45Ge+xOVXXsvauzbgvaLfK0mTCVqtCdbetYGZmaWUztLr9TBxRBRp4jiiKAq01jjn5AMTBEEQHgDxgzEBZODhUeNcSaQNqrpofn6WJEmIooiydOFkg4lI0xhX5jhX0m6lGOOZn13Hn/3ZkzjuqU/g4QfvwcRkeOS88LRjNXgJiwogYMsS5xxxEvy9+v2SVpqMZBr1sbTX6/He976Xt7/97ZJ/CIIgAoggCDsm//iP/+hf8YpXsOuuu1KWJa1WC3CNMWqRlcQ6wZUOnepqr+WweR8VabQxFGVBFCUEF4+BR0i9bPQerr0u4wtf/BI///n5zG7sUBbQmphEETE/12d66TJ6vR4ojbWWJEno9XpMTLYpylI+KEEQBGG7E0A0vqn8mO/MsWTJNMYolPZs2LCBiXSSVqtFZ26OVivGaJibX88uK5ezeveVvOXUN7LLzhFpC1wJSoOzJd5b2kmK8pt7TfVJiDi8Eg9FUZLEg2rObq9PWeaceeaZ/O3f/q3kHYIgiAAiCIIAcMopp/h//dd/rYzR7GAcrq8qQPzQms+HMXvajO7CHB6FweLDxBg0eIXWg47oLIfPfuY7fPGLX8I66OeWOGqDislzy9TMUjqdHtY7+r2M6elprLfyAQmCIAgPrAAyPsIWiI2m1+8yNTVFlvWIY8PG2fXEcUQrjenOzbN0ZgneFsHgtMw47qnH8trX/DVLlkJeQCuppuF6MLo2BK+fqRZeaE5ANMfmepia98zNdZiZmWpu6z047zn77LN51rOeIfmGIAgigAiCIIyzxx67+xNOOIF3vOMdQJgIk6YpsxvmmJycxkQ6jNZzHhNHDFeKOOew3hGZ4ckuYeHW6zrSdjg7pVRY6M3Owi8u+BXvee8HmZvL0FELbVroqMXsxg5LlixDmYjOfBdlZFcpCIIg3P8CiKuFkGqCWW1k2pih2hLnLdPTk6H1JTUo5XG2QDlLK9bYIidJNS94wfM5/vgnMzMDSRz8spQeNNNooLQl3pYjk9hq7SNEPfJ7UVjieDAT11pPXvQ588wzeec738l111wrB1BBEEQAEQRB2ByrV6/yJ554Im984xtHjFFhMFKvNlOLogjvQwmw0vUyzmFdcJg3KmqqSbK+xStIU0MvgygGr+B737uY73z3XC646NcUhSJNpimdAiJKZzFjr0EQBEEQ7g8GY2yrlpcRAcSBd8HTo50y35mlLDOmZyYp84zJdsLqXZbxuGMew1Of8kR2X92itOAspCnkRU4SR3hXhOOlMZhmhG14fM/oGN3Fps64akLMxo1z/PR/fsx73/tuzvvp/0qOIQiCCCCCIAjbyuc+9xl/7JOfyE477YRzjqKwtFuTQdDICtI0HlJHwDmP9xatNVoPFohZr0PangQFznq0ibFAUYaCXqXhokt+zyc/9Tmuu+YW1m/skiaTOBTOe/kgBEEQhO1IAAm0WykbN66n1Y6YmppkduNali6dYaeVy3ji4x7D8579BGamhio4VF3p0SeqpqupYYHDa7xTKKWx3qONxTU1HwsFEOWhLD0///n/8qqTT+Lyy34juYUgCCKACIIg3FP+87Of9H/54r8ENM4TpsOM4S1oM7Z4tEV1mQOvKMuSKA4O9b2sJElbeMBW67siD6LIO047g++dcy4zS5bRK8CjqgWkHlkobmpX6hRoH+LAwEQQBEEQ7h7aDx9xHMo7PJY01vR6c2jlOOjg/XnzG1/P3nu3SRIw1e2DUGGJYwU4SlcQaY0CnHdoFYVbehWMVlX1KyUOj0bh0OjKIaQ++l188aX84RGPlHxCEAQRQARBEO5t9tlnb3/yyadw0kknkSRJVeFR9TErj1LVstA5lAqtMt6HXui6jcVX1RzhtkFM8T6U75YOyhKSJGgW1157O7+67Ld8+IzP0s3AFpbSerwzpOkESif0ewW6mkAzfGas6ZFWDihZaKUvCIIgCJvBK4qsJI1j6mpGoyDSDu8KlM/wrmBmpsUxRx/F8593PPvtvxwNZFloc2n8w0dW/IO2lrIsMcY0x89+P0drTZIklGVJUWS02m28cygdkWcZSZpyzvd/wEknncR1110juYQgCCKACIIg3Jc8/OEP969+9at56Utf2oysjaKIubk5pqammoVcp9NhcnJysJb0vhFHavHEWjvyu3Ou+RkgK2FjD879yUX88Ic/4ne/vZbZ+T5apRidUpQepWJ8Ux7MSMd02NO6MRM5iRIlSpQocfMRIFKaVpKS9TsURUYaa4pinnZsWL37Sp7+Z0/iMX/yRzxsr2kU0O3kTEyFKsciz4mTZKtW+7Ozs0xNTTXHv6Io0FpjjAkeW0qRZRnf/va3+cAHPsB5550nOYQgCCKACIIg3J/svvvu/m1vexvHH388y5cvby6vRYyyLFFKDUbqjuHCrNwRwQPCGbGyLInjGIzBVqKGtfDLX17P1792Nj8//2I68xlx0qau/nBjvdFeEUqJfQxolA+XSZQoUaJEiVuK4HA+o591mJpo0UoNzhXsv+9e/Pmzn8mxxx5KbEJ1h16wqHeVmqI3u9rPsow0TRuRI89zkiQZuU2n0+HnP/85r33ta7nsssskdxAEQQQQQRCEB5q3v/3t/tRTT23EjNnZWWZmZoIQUVV+AE1bDDAijIxXglhr8d5joogSyIpgHhdXBvnr18F/f+07fPlLX6XT6TbeIMHzozJfBfBRI4Bor3HKSZQoUaJEiVsRSyanUnr9eQyexz/hGP76r17K7rtrbAlJVC3kHdjSo7Qj0gplAF8JIJuY3DJOPU0tyzLa7TZFUeC955vf/CbPfe5zJV8QBEEEEEEQhO2N3XbbzR9zzDF8/OMfZ8mSJZRliXNuwdmsmvHrx0WQZmE49HNReKxXJNXwmbvWWtauXT/W9qIG92oc/GVBL1GiRIkStyVCt5cTRRGrVu3KsmXBqFv54O+hCB5Y0VCRh7UFxlR+WNaiTbxFAWR+fp4kSZpjYVEUfO5zn+P000/nmmvE40MQBBFABEEQtntOOeUU/7d/+7ccfPDBZFnW+IIopYiiqPm9Pus17AEy7ANSljkmCpfHQ2an1tWjBMGYuvJjVDRRQ3taLZNgBEEQhG1geEBtnkNeFExNxiigKD1xpILhaZ7hyoKJiQnq6kNXCfqhHEQ3x7zFsNZijOHOO+/kk5/8JJ/73Oe44oorJEcQBOEhiezcBEF4SPOyl73Mn3rqqey2227V4jAs9pxzweMD6Ha7tFqtBZUf9RIULEXRr4STCDwUpSOO0mYv6he5lx5Zwjr5MARBEIRtQOPR2GYQe31pOKbkRY5RmmhoyllZliPmpYuJH3VbqPeeKIr4/e9/z0c/+lFOO+00yQsEQXjIIzs6QRB2CJ73vOf5v/u7v+PII49Eaz02DnewKBz+F9phQCnHsKRhS4uJQqlwmCozarLqh9peAJTXsrMVBEEQtgkPlDa0wCjAeo9RmtJVIn4UqkGsrX2t4uZYtrmKD+dccyLgr/7qr/jCF74ghyhBEHYYZIcnCMIOxSGHHOL/7//9vzz/+c8HBqW/wwvD4UoQ58pghmoU1lmMjqv7xEH82GTVSI2u3PwFQRAEYVtX6iVgm+MJ3oNS+Pr4peOqjTPCOagPSdb6xgtkbm6O6elpIHhf/frXv+a0007jG9/4huQBgiDseLtVeQsEQdgROfjgg/2LX/xiTjnlFCYmJijLEmPMyJQYYMQfZBiHx1mwriCJFzNZHVSMiAAiCIIg3L2VegmE0ex4TRTHzfI9iBwG71VT8ZH1g1iSpmZkpT8/P8/ZZ5/NGWecwbnnnivrf0EQdtzdqrwFgiDs6LztbW/zJ5xwAvvttx8wKB/u9/u0Wi08UBSWODbkeUmSRNht2IFqeYsFQRCEbV6kO5zNUUpV1YZq7Kii6fV6tFptvAddXW1rc24Ht99xO9/4xtd43/vex7XXXivrfkEQZN8qb4EgCELgOc95jn/Pe97DbrvtRqvVAmjaXMoSKp85ALLSk0aKrLSkkWnqPcZjwC1yjUSJEiVKlLj5OLJQ9wprLbYabxsNH5SqSsNuN2diImHt2lne+9738O73vFPW+oIgCCKACIIgbJoDDzzQv/vd7+YZz3hGGIvrVdNLDVCWjigKxqb1uNvNxYEIIgt6iRIlSpS49dEPjWbXOlpwvKqnvhSFJU1jLrzwl7z5zW/mhz88R9b4giAIIoAIgiBsPfvuu69/4QtfyGte8xqWLFtKkeWkaYpzDmNMNQFm3G1fL1RBtlIokShRokSJEv0mFuj19Bag8auK42CC+vGPf5z/+I//4Ne//rWs7QVBEEQAEQRBuGeccsrJ/mUvfSmHHHIIURzjbEGe57Ta7cGNPKD0llexgiAIgrBFdHNogWYADAC9fo+1a9fy//7fx/nsZz/LTTeskaONIAjCViA7S0EQhG3gwD/Y17/znafzrGc9A20MWb9PWvmFDIQPff8LIH7zVqtehWk0/j56PVuadDP8/Hf3dYzfX6JEifd9fMDxgwVrs4tV9+/+C+UeUAHEulD9EUUaaz3nn38+//bvH+JLZ/5/so4XBEEQAUQQBOH+4VWvOsmffvrpTE9PY63FO4eJIpTXWO8wSje2/PVEGYAsy0jTFICiKNBaV6MMPdbaEWO7uvd7lNHFePAkicAH138YTAMoLRANOspLB2UJSTK4XptwZhFA66GzjYRpArHZ8oFD+01f7wGnwm2G43iCs6mkpk5OxBFAosT7P25ypXgf9X34TWgPI69rE6/JucHe0ZiR3QvOhf1ZEoX9XVmE/aC1EOtBdUW9v1E+FPQpdc9Wy0VREMdxs++P43hkn16W5cg+f/g44L2nLEviOKbX6/GRj3yED37wg9x6662yfhcEQRABRBAE4YHh8Y9/vH/961/PcccdN3K5tRatNVmWNeLH8GK42+0yMTGxyCLeNQvgUX+RxQWQ+izhjTfcyk9/eh7/89Ofc/PNt6BVQntimm5pmO90iaKI6ekllNbS7/dRSqHjiLIsSZIkvNYipygKvFYkJqrM9bKFgsTQaWLtR4WKRbQNvBqdi1PfciTRWnB9uEZ7+Y4JwgO3UnS4ajsPAqa6dwWWxbSVRojQaB8tuu8biCBBADbGoJTCOUfpCrx1eBXMQdERmuDdlMYJ3ltcmaMUaBTOl3hbhhHo3rJ06VKOfsyf8qQnP54/2H+ne7xazvMcay3toZbJYSHce09RFCS1Mj0khpx//vl84AMf4Mtf/rKs2QVBEEQAEQRB2H7Yfffd/Yknnsjzn/989t133wWVG957vPdorUcqO8qybM4Cjow1HBNMNiWAdLt9oighicN9bQmXX34t3//+OfzsZxdz023zKNMiSRKcg16vh8Kg4wjvFQqNV+EMaekgiiLitAXOk+d9okgvUgLu0X50zo1qXpdbNIlarE1n86X2Wr5UgvCAE7bnTbfK6C200rhtbLHRTaVYM0J8sRaUZn+iq32ow3sfxJA4xVRL3CRJKIoCZwusLXBljokgjhwoi3c5aRqxbMkkhz7iYJ7y5GN59KMPw5hQIZeae/bu9fv9RmC2lfg8OTk5clwARsTum266iW9961t84hOf4OKLL5a1uiAIgggggiAI2zfHHnusf/3rX8/RRx8NQKvVatpd6kVvFEXkeT5y1q8sS6y1Iy0ycRQvssd2iwsFPpR0G938yoZ5+Mznv8/ZZ32LTqdXPW+JI6KVtpnv9ohMCx3FKGUoSkdZulAhojXe20HCot1Yv3z1OpSvLh8WQbZGwFh4Lnj0DLFuhBJpRpAo8f6Nyg+2Z+01To3Hha1tm2x12wybFlRKUBZUOSag6mYZ670n1kk1DUVRFAXWhp2U1hqbZ2jlabdTIu3p9eeII097Isa7jEcedjDP+fM/49GPPhgFdHpztNspGoXCo4m4p2LsYtV+eZ6jlCKO42aU7dVXX8373/9+zjjjDFmfC4IgiAAiCILw4OQ1r3mN/7u/+zv22muv0dS/MuyoqzxqcaQoCowxgwoSzyYEEDeUDOhqPKLG6MENSwsW6BUQR7BhA/zgh+dx1llnc811N6GImJpexuzGLr2sJE3bxEkb50IyEpkEa/1QQjPWkqJc5RPgGH2xIYEKtzWbMRlcKICMnyH2qn4f9KCSRKJEifdLrIXNIEyMxy2bqG6tYWlzP6dHBFen7IigOt4SF0UReZY1AnOaphgdYa2lLHKm2wn97hzWFmjlWbHTDI89+o956tOezMMP3gnroJXWu9iiqmQrAY91nki3gegeHQPqdpeyLMmybKQCpCgK/uVf/oUzzjiDNWtkkosgCIIIIIIgCA8RjjjiCP83f/M3nHDCCUxPTy+4viiKEd+Ppv1lCwKI96q6z2glCISEwlW/ukoQcT6YpK65eZ7f/OYqvvSVr3PnXeuZ3djDK0MUtcgLS1lalE7QKsahF+/ZV6FSxI9nMs3rUCg/mjwMnyEeLW13C4SS8Po10g4jCA8MeostMJuPjoWVHYs+T7Nf0AsqSPzw/mao0kzjyLI+SWyIY4Mi+Cc551AKYqOwWZdHHHowT3j8MRxxxOHstipmamKwP607T5wtiQyApXQ5eEtsUiC+xwLION1ul69//euceeaZnHXWWbIWFwRBEAFEEAThoc1f/uVf+r/+67/miCOOII7jxuej1+s1RnlNe8wWK0DA4/FOofVgoW6tx9qCOFGUro/XBkOMxzRpjXVh8ssdazN+eeGl/O8vzuc3l13FnWvXodDE6TS9nsdXHfVVYfnQqxj04ddnet3YIUaPeX+ExMgN/Tl+JAka7/d3In4IwnYhhNydRhqvNA41KnxuYkE6PhgGH4GPQ1Su2htUFSGqRGFR2tJKI5Qu6M5vJMt6rFy5gkc88lAOOnA/nn7ck1gyY2gnC6dkKQ228CSJCs/rPapWaLyt1JF7JsDW+/GNGzdy/fXX87WvfY3PfOYz3HjjjbIGFwRBEAFEEARhx+Od73ynf+5zn8v+++8/MhJxxADVswkPkMW8QPTY7TJCSXdYyJfeY1SMw+Cr2zpGm1euueZOvnH2t/j+935Er6dwPh55bN8kBQrnVZUehd/92OsJZ3Gr1pnh08ONIOIWJFrjDM4MO4kSJd5fcWTF6DazF1t8vkvwSY4WFTH1gvuxQCgFg7bp4DbKoyhBFU2MtCPL51ky0+aPjzyCpz/tKfzhH+0Xxn5bUNaRJuHZrKWq8hjsUvXYLlWpMBLXO4826h6vlOfm5jj77LP553/+Z66++mpZdwuCIIgAIgiCIAAceuih/pWvfCUvf/nLSZJkxCDVWo8xYZftfajsMMZUC3aPdRajTXW9qi6v68fdqIKiNHiNr65SZqEAUuOBDevgf867lG996ztc+psrKIqSiYkpPBFZv0BHKVHUoigteV6iTYwxcXjayGNi6PW7GBPK1Ms8nME1xlCWZUiGtG5GWFJNyAljLX3lbbIpDwKJEiXedzGIj34zQkjYz4zXbgyVrfkwbcp729w+jJsFj0WrKIyh9R6lVPNP+yCO2MKRRCmxVliXURZ9tHFMthMmJiNe/rK/4OGHHMC++yxHaygsxCYIG9YXRGq8dk0NSTAa7wYVIfhqd1ntJuv9oVK+mdxV76e8980+uPYfqc2tAc4++2ze8Y53cP7558taWxAEQQQQQRAEYXMce+yx/kUvehEveclLcFWbykLcgp+LssDogaGqsxbtTWVoOKRq6LEjwFCS44ZvVwcP1gcT1WuuvY0LL7yIC86/mKuvuR6IcETEUQsTJTinyIqQ7OjIkBU9VKQaoSPSmqIo8N6TJm2cc1g8yoURwcNmsGVZEkcyjUOixAciDnt4jFZqDG7nnEMbKsEyiAP1P+89ykVoHVVTpYaEWVc2Y8GtLQCI4mCYXJZleFxlmUgMvc483ltWr17FHz/6CI486g858A/2Z9lyjVEQx4PdmfMOrepKMl817tQuRgMZpK7/UIyZTVc/OwultehIhVHg9eNXQoj3vhlhXrctnn/++fz7v/87X/jCF2R9LQiCIAKIIAiCcHf4m7/5G//sZz+bxz72sbRarZEzj86FBXhdKTLcNlOWJZFJRss7FKPDY9SWjwplJcD4oRO7WQ5ZH7ICfvKTC7j4kku58oqrWbtuPb1eAV6H6QzWYZI2hQ1JQlEUbNgwy8zMDNPTS1i3dn2YeBMZNKEipMhtEEfaLVqtFnnWq9pkJCGVKPH+FkBqEWREABmaEhNFEc7ZRgyoRY04jomiCJsXgKoqP1zltWGrf56yyGi1UlqtFihHv9OlLHOmpqZYvnyCPXaf5g8PP5ijjz6afffdBeUH3kVKVZUeFsqyIIo1kTYjXh5+rL1msIsbiBree7wLhs5K69GRuypMcVFKkSRJI4DUXHrppXz+85/nS1/6kvh6CIIgiAAiCIIg3Ju8+tWv9q961avYd999cc41wof3Fueo2mSGEhWvg2jhF9nz15frLRwRVInzDmfBRMngcg95CVEUJsso4KZbMr77nR/wi19cwF133cV8p08vV0Rxm36/D14zNTVFWVo6nR46SoiiBFuG5MPESTV615JlBUVRkKaxHKwE4QHAU5mYLuo/FLC2CNUdhqaCq2lzcZXZsbNNVYjGYYwiigwmUkRaUZQZvV4HpRR77L6Ko476Yx73uMfx8IcvoZ3U1RwaX+3IFJVRR/Ma7KB6zIOzPgjFsW7+jtHdnxsRP9TI5KpQ2aGUwhiD8wNz6TzPKcuSbrfLZz/7WV73utfJrkkQBEEEEEEQBOH+4JWvPNm/8Y1vZM89V5PnOXGcoBT0+zlaQ5IkTa+9q8Yf6Pq0ad3H73016cAtkvrUJR8l9Rlh6y229ChliKKYMCISstITV20uZRnO0Oa55847NvD9H/6Mb3/rHG68YQ3T00uwzuO9Ikpa5LkjMjGlC1NsvDbgVTXa16CNaV67IAgPgAiiNjUBJWyXUayr0bPBz6cWP6wtcbYgVo7IaKKoEke8DS0vPuxXTKR42tOO41nPfAYPe9hEo8dGUdg1ZVmPJHahmm14+ep9MDAamzKFr64angquNrXwHZ6kFSpStBq9RZaVpGmbTqfDxz72MT784Q9LpYcgCIIIIIIgCMIDxaGHPtIfe+yxvOhFL+JRjzps5DprbZhk0CzyfTh7GgZTstj0hTHLQ6zNMWYwCtITBAyt6qkv9ZnZ0BdfWE9cjVqwtvIv8bB+PVx7/S1cfvmV/PwXF/C7311LUYItHYUNgoeJEpSKKGwwRIyihKK0Yx4EEiVKvH/i0ILRabx2I4OcvHLggs9G2EQ9RmmU0WgURjtiXVD258myjLQVs8fuqzjs8Efy6D88jH322ZvddmsTJ5Xc6qvOvGqXlReONAGweOeqSo8gjDZzZEpfiS7BfySKTLO6La3DRHorFr6uMlumai90zM3Pcccdd/DpT32G733vHC666CJZMwuCIIgAIgiCIGxPHPgHh/gXvvCFPPGJT2Tfffdl511X0C9yTKwxREMShyWkKGqrDgzeVcnJULuMc5VZYTX5wAwLLX6QxHhfCSHV09sStIG16+HXl17B1VdfzzXX3sAVV13DXWvXY0tQyoSye2VQOg0/V54DEiVKvJ8iYCo9dDAdhqEI3jqi1GCUpnQlZV7g8KRxQiuBA/dbxcMetorDDz+cQw45mFW7xs1+xjqIqqI0az3WZURGYXQQYTyarOuI4xQTjWq0zo22uYxjbUEv6zM5Mcli48IrW+aRaVudTo9rrrmG//3f/+WrX/0q5577A1knC4IgiAAiCIIgPBjYeeUq/4ITns8przmFqZlJli1ZPhiVO9RLr7d0gBhtj2esQjwkQR7K0hPHVYd9JY4kSRQSHWtRZiC4DPuyFhZMZbJ6040ZP/2f/+W8887jxptvobQw1y3wRAsSmNoZYHispR+5vk7cNv0eeXXP22vGPQVG0QtL9Efu7Jq3eLEhonf/QD04g+/v8R/4ALcgbe792yY2X+Gw+fdy4eezyTg0Rja8d1v//g0qKzb/vd3S/Ybjtrx/49817UF7h/ZDO4Ohv0nhcK7EGI2JFK1Wwq677syjHvUoHn/MYzn0kGVgoSoIG3SuVPsRpWhMUYMPRxh57X14DqUS8GnjV+RcdZ8hA2dPbQdSbUvVuF1dVXJs6rtQCyDee6688rf893//N//1hTO58qrLZG0sCIIgAoggCILwYOf4P3+2f/WrX83jjnlcEB7KgjiK8dW4ynqyQVmWYTLL2Ozd8aoOpRYXRLYqYVvkcYdTFWsH/y674ha+/Z0f8KOf/A+dToe0PUFZeMrCEbcmyXKLVhFog8OgMChlBomOs2gUDo9zDqXD3+qcI7c5rVYrjOKtplJ47ZtRnbpOuvzA2yCYJiqMqUwSy7JJukYOsCoY0noXDc6qbyK5dgq09yFWgztrjNeDdoMtJMGLCSBu0XE/WyuNuO1EALknIogaPM4iFQ5bKwBpX39OW4oapxzaa8CiKBd9D/3Q98RXhp4W3zyOqV630bXE5pvvX30/pVQlLhoirSmdw1uL9Z7YGJSJyK1BVYYY1lq8d804aggVHFCNqFa68u+wYcKLVrg8I47C9935EptnOF8QmeD78+g/fhRPfMLRPOYxf8z0FORFGEtrNrN/GL8oTK8yzXeuKAu01hgdV2PAN/9ZFUV1e2Oa1x9FEfWY3vq9GvxsuOyyy/jQhz7Epz71KVkLC4IgiAAiCIIgPFRZtWqVP+SQQ3j2s5/N0572NFavXr0w7a1EkXrcbpZlpGmK955+v0+73R5JxIZHQ3rvybKMOI6b+282v/UDwWH4MucczpvgLmLAerjjjpI77lzPmjW3cP4Fv+SXF/2a+W5GP7NYbzBxjLOaflagtaGdtqi1ieHncNjwu1bkeY7XKozUjMyIEFSWJWmcNglinaw65/BlWXkORKhKNFFKVZ4IHm9d5ZUSL0y2m4TYDQkXfuQ6R0imtWeLJQF+E1M6fCV+OLU5KWpLNQ0DGeX+d6LQY0sWv41xW0SWxcWPUP2wLY87+heoqqqBRTx3Rio1FhXQwBblQAwzZiTJL8uStN2iKIow9jqKaLVaaB1aO7pZH21SIhNXBsiVgOdds51PtFpYW+DKMKnFWxda5LQmiRSKEldmTE62+YMD9uPII/+Iww97BLvvtpQ4gSSGVuXhUdoSfEEUB/HBlg4TpVslYBVF0exzNrXfqLe94e2xFmzr96t+H4b3WwCdTodzzz2X//zP/+TCCy9kzZo1sgYWBEEQAUQQBEHY0XjsYx/rn/zkJ/OEJzyB/fbbj5122qm5bngcZK/XI01TtNaUZUjK6kRjODkZTl7Cmd14mwWQkcSoHPiHFEXwDUjTkEraEm77fc4NN63hsst/y+9+ezVr1tzC2vUbKHKLMRFZ36MwTZKU5zlZnSxOtClyW3mYxBS2JMsyymrEcJq2KYqySTaHk9D678zzHBgIIMMVAYsdaP2IADIQLRYTKbQH5aMt5+93exWwZQ+I8BfoTbZY3Lexapa6J14WlFvuJ/Gb/v4NPpetbJ0Ze/+0N2P3H31s5Txeq+Y7NPr5q1ABoaKmaqkWAEwUtsssy4iiCK1DhUf4PkKapqRpQrc3X1VjeJwv8d4TaU3VDUd3bpYo0iSxqUQPw7Jly9hrr71YtWpnDjxof/Z52B7sv/9y4uqrmPfDlJY4qt6NapBUKDQJ7StNZZlJN/t2DXtwDO8T6n1Pvc1tah9R7z86nQ7GGFqtVrPvuvnmm/nmN7/JD37wA8466yxZ8wqCIAgigAiCIAijnHjiif6FL3whhxxyCDMzMxhjyPOcNE2D6WklDmRZ1lR+1GdZgZFqkHuOo7Q51tpKdEgIfftQ+so7hNCm0s8hSoKponVw4YVX8/OfXcDvrr6J3992B3fddRegidM2zjmyosCWnlZrgjwvyYowYSJOE7SKKMuSoihIWm3wYYSn10HgKFwQRKx1tFqtQVXIUFVMfYZa2WIs1168FcapxRNs7aJNHqz9ZpJ27RcRO7YVr1He3Is+HHdDoFH2Hni1VH4Vm1NZtiCCBAHk7laWRHhiqEWsyjx0+PUNC4Cq+tDq75F3ilin1Wjo8B2z1uJx1X0G21p939DuZUMlki9ItCOKq8t92H61d7hqHO3eD1vN9OQkD9tnL/7oiEdx+OGPYKeV9bZMI5QwZIRcT2yp28O8d8G81Ci8dU0VhqpvvBnqbXuzn6IbTGmJomiBGFILrd1ul6uvvpqzzz6bj33sY9xyyy2yzhUEQRBEABEEQRC2joc97GH+pJNO4sQTT2R6erpJbIbLzBcTPqy11chKveB2W8wZRypCHFDQeGXggscHBlAUhQWtiEw4g1x6cIUjSgYJe56DqYpQ7rjDcsGFv+QnP/lfrrj8KmZn5yhKiKIEEycoZfAunHX33uPRFC60qXin8CokkspooigIJbXwUVb+DL4qGVBKgXIYXxk7Ljji6lEBhMW8KaoWGDaXnC8ujKjqJ+2DWeXdQ6Oc4p55cNx9fCWAcA9ePwsqMMbj5p7f43TJ3TNPBU+E8zEwSPAHLTWMtEKNaDHeowktVA6NteHKJEloJylaQ170Q/WHDpUgUWzw3uJKS+kKcB6jHLHOKYs+tiyZnGxzwAH7cczRf8rRjz2KPfZI6GfBryOKg3gYtt8QTQSFG4gei31z1YiYUVdCDV2vtm6bb0STqqUuz/Om+mx4H1JfV1+eZRlf+9rXOP3007nsMjEwFQRBEEQAEQRBEO4FDj74YH/ooYfytKc9jWc+85lMTU1RFAXWWiYmJoIAUZZorRshZLHy9q1JhoaFAe8zdG0E6aoJDzoepGEeysroVNdKhw8mqV5pothQ2HDGOjL1ZAzodiHP4I675li3bgPXXncjv/nNFfz2qqu54447cA5MFFP6SlCoR/MSznzb0mOr16pVhDKhTUExGHnhvavEh2EBZCh9HBI7FooZaihh3jbxY/gRtPfVa9ic28ZmFgrugVsqeO2b9+7utNA0AshWTOLZpACD34yHypbeV4UlGhJAgiBVP6fyoZLDY9FVJUXzLakUERMlA5HA5pR5EASjWBPHEVmvTxQHs9+yDMLAbqt24aCDDmK/vffkkIP3Z9n0BCtWLGfJEmhXxVrWQZFDux2+M2XpMZFCq1FhpqpFweFw1qOVgiGTYVfaqtrJNFNdavIiTILaEsNVIMNVVDWh2sqSpim9Xo+LLrqIr3zlK5x77rn85je/kbWsIAiCIAKIIAiCcN/y+Mc/3j/2sY/l8Y9/PAceeCC77LJLI3rEcbxZX49tw+Fd8N9QOvS8uCKYmKr6zLACnAdl8HXrSpo2p67rUZn1VIph31GnQiLofTB09Ar6fbju2vXcsOYmbrhxDXfdtZY1a27h97ffTqfTIbcuVIR40CbGe4WtB+J6RRgwbBqxZJRhAUQtFEaGk/WhRHSLAkh1v+FKE8WmTDy3tqXDNwLOA4PGo7dy+srCePdWRm7sPTWbNUndnAeJU26BQa0akqn0UCsL1RQiXSX/kdIo7ZntbMDEGqMH02LSNGbZkmmmpiY4+KADWTYzzerdd2OfffZm9R67snRpVXnhgRKSoTG0TaVJdX1RlGECSyU24Ty2zDFxDMpTlgUmiqoR1nps+ahHovfgbHhurbd+lTksgHQ6HSYnJwGYnZ0lSRJ++9vf8qMf/Yjvf//7fOc735G1qyAIgiACiCAIgvDACyInnHACT3jCE1ixYgWtVmvEG2STKfYWTFCHcscFRy1bhsG3TZ29Gc3pfVmi4tFKB28BrUceylNVDjSvaSAyKAZVIwBzc3DtDbdx6a+v4Lobrue6a28kK3I68xnzvS5Zv8B6H6pClMF6w8go1uHkXg0LH2rk8toENES/+GyWJuEfmK8O/h49MsJ1S0LDJt90ZYES5XX1eu7PCI7oHi1XtB8zUd2qL9rQ++KDh8fg89BDn0v1OWmHcjrEcRPXpoWnelzlKxEiiCFKeZSzeCzKg9aKqBrpqrVj+c5LiBPDTjst56CDDuLwwx/JAfvtxcw0zfdZm+Emm/AFUKqqQRn67F3lKTsqTgTxxZUlOjZDs64t3tqByNh8P9XgVx98booiVGYNGx7bMryurdm0h6e1zM/PMz8/zwUXXMCXvvQlvvCFL8haVRAEQRABRBAEQdi+OfLII/3LXvYyXvGKVywoa19M9BieOBOSteCBgFfBiLG2TajuZm1I5JwNXgUj+Z+rS+kd2qihBLh+HA2VqaQ2ZlRV2JxYM5Yie4LHiK80GOdg3VrPVb/7LRdf/CuuufYGLr3sd4OUukoY6+TaqdBaY63Flr7yNYhRRuMKT2FLkigOyaZTFLbElcH3JEkSoiSmyPJQdeI9tjZiNZpYR2A0RVE07/1wi05tLBl8F8Jo1eF/VMk5yoIq0V5jsWiv8dqjqwqHLUWD2eLthh9POYXFhug9xiTVON9NLGSqv6s2Ca2n8QRfiRytQiuSMRHgm2qDkLgXJEk00nZV39dXr8GoCLyqhA6PcirESujw1oH2iwojSlu0KdDYxiumFjqoqnOU9sTGsPMuO/HIQx/BkUc+mkMOPphly8JXNk42LVcN12FUQ2DHvqnjt9qC4DP82VfVVgvupwa3UyoCtXBs9sgrGHtvh4VP7z1aa7rdLj/84Q/52Mc+JhUegiAIgggggiAIwoOb3XbbzR922GEcd9xxHHvssey///4YY+j3+yRJsmBqTFEUoQ0giqg7OLwPood1jiTRwWmhMmj0PlwOEGkdcram1N+NlnbUfTBaLZ4ELlA99CKp5+KiSP3QEDwW8jIksLNzsG5dTq/Xo9/PWbt2LddefwO33nor11xzHZ1Oh9nZebrdLtY7tIqqJF2T9S3GxERRRGQSlFJYa8nKgrIsSZK08V3RWuN8uL72Z5lesrQZn2qtpbJRaYQCax1xHA/5LLSbsb+hlck1wkcdt1bY2JbbKacWud/iy5XhRLr+rgz/TSO/O9+MkA0iG1V1RfgsizIbecxBrNp/XF1lNJjyMxCJXDWCtjb/De95LQZEBrzrsnLnZeyzzz4ccMAB7L3nHqxcuZKJdpso0uy2y1KSBJIojJH15eB7baKFrU56wTtSV5d42GS7kl7k+1tJJrZsTISDMEgjCOZZFlrJSj8wOY7UoivJcZHTEe4TadOYItd+QDfeeCPf+ta3+OlPf8qll17KlVdeKetRQRAEQQQQQRAE4aHLC17wAv/oRz+aRz7ykRx66KGsXLmSsgxnsOupD3meU5SO1kRrgejQ73aYnJhcIGLUxRy+SvoHyV1NrQD4RWSMYTHEjSaFEDJUqOr/Q0VJ1guVK9rEobVAD0QQRzgnX7qBCevwgTdMoAkP3+vBxo091q5dyx133MWdd97J/FyPDet7bFg/z+233866dRvodDoURUFe2qbiISSYla9ENR3DmODDMtvponRUjREOospwohpHKevXr2diYoI0Tdm4cY4kSZicnGT9+vW0WhNUBQuhFSeMwmmit765fLG4petjE+NwwX4C1zyuDv9hXbFZgSqO42ZEav1e1NVEcZzS6/SZaE82okXaSvDek+c5WmuMGUwcakSi6n01xpKYHFf2mkqaVithZmaGJUuWMDExwdTUFJNTbZYvX86KFStYsWIFK1euYOedd2bpkjZLl1QtSEODjfywD4cFY8b0uG1m+DvrB19pXRnwLiqEqEqo8xgdNzY6Ze5weNLYDHxEhrcUH6a9aBMqs/J81Oh04OcRXtMdd9zFmjVruOiii/jxj3/M+eefz/XXXy/rT0EQBEEEEEEQBGHH5VGPepR/+tOfzlOe8hT23ntvpqenmZycDG0mLkc5RelLsNBuBb+Abq8bKiOiqEkgPVVpvQpqhHdVMt1YGLhFEsfx5pYxAaS+SWPEsdC3g8oTATU0yUap4ZQUgKzIAE1a+SXYTRyMbZ0ou5AcDxfJ5DmsXQdzc12uvfZaNm7cyO2338Edd9zB+o0b6HaDF0lZlmAi+nlGb75PP8+wRYlXCo0aeFWgcdaCMqRxElpwSotXmrLwYUTv8OJhqFpidGrPQobbIhZjuKqijk2Fh3JoU7UyebDVFBWnQDnfeKAUzobhyEkM1pHbksREtNtt+v2cVqtFZ26ebr/HZLsFWuFKy8TUJApHnCS0Wwlpu0U7bdGenGBqYpJW27DX6hUsXz7F6tWrWb16FcuXT5KmVTWShfEhR/XnZmpfjrroaKiFS+sxM1LAFrWQZkK1iQ3eGK12mwXq39aIIbUAoix+ZNrPcA1J2EbyvMQ5R6s1+GPqKhQc5LnDlZ60baopMZWAV3jiJPzS7/ex1pLnOddffz0/+MEPOO+88zjrrLNkrSkIgiCIACIIgiAIW+Loo4/2L3vZS3jWs5/JsiXLKV2BxoyIAQvzQY3HV1MrwNmBr0Ht/bCwosBtbXa58PA54p2hGk8KqudxY8/l8XgHpQ8eJ0mU4PFhtK61eBUqYHSVqKqhl+XHDC3DqN6QiDtH1d5ReaNUt8ky0BGk8UDD2bAR7vj9RtZtWM+Gdeu54aY14DzdXsbvfvtbOt0uCrjhxpsw0USYVrOIR8iwGLIp6uqMTb6TQ20rdStJ8xzKU5b9IIQQ2pbqqDygFWVeoIxmsj3B0uXLmJ6cImmlLJmeYWZmmonJhNWrV7HzyhXsvOsu7L3nCpIWFBnEaTW1RAeRybrwu1dQ+4FuaZiJdYMPSGu18LauGUY0IvrYIogOUaxH2nbCm+JZdCrQogrSZi5XNFUgozdf2BIzPK2o/i7poWqm2tukzIOPSRxHKAW9XsaPfvQjPv/5z/PFL4phqSAIgiACiCAIgiDcY3ZftatfufPO/MEBB/DUpz2NpzzlKey8884hmfSuqhwwA6NVp3C+JI7i5oA3OvTVbf0Ems3ln1W2uEBY8ZXhZmSCIOMZG4mhGqEmvC492ohT/U2u9GgdYYxecND2mzBvdUNGsXUBhisqU81KHMFCbiGNwuW+DPerr68T3rwIVSrOhhaJOvrKWmXjhv5mx8F25nubvV6rCBMpjI4xUTU+WFVjhrFMpAlKD8xOw79Bcj49vcjf7wZVFk3BjoV+4WnFCqfC9BQTBwmgro9wlWBh8URKjX5cbjCIqNJewlhZwNXGpmPTeqytjUE1vh5zq9TI98C7MhiODn2fHB6jNCizdSLIpsQQtfhKb7iYqdvtM1G1mIXXO7hDWTryPG+uv/POtVxw4S/4wQ9+wAW/OJ+bb76VNWtulLWkIAiCIAKIIAiCINzXrF61lz/88MPZb//9efSjH82jHnUEe+65J62WaQSCvMjRsSZCN4luaJNRVZvMPT+Q1p4j3odqB631qLBSJbCuKtVwEDw6hroRvB101Ix6wg57YOjqeXzwc1WgVVTpLaFqQCnVJNS1HeWiSfRQslyWnqgyt3RFyLuVqibtxAuNOGsv2YWvddspy4XVBiPYqgpjG1YsTQeSZuHs4HE1zAWPjDDxprqyNs5VGpxeqJ5Vv3s8ON+0EqGHBLWRaTpqwR8wbJY6PHll9KkWq0yqW1dggSmvHxPHKiFrTPlbTA2shKPQhtPpdLjhhhu47obrufzyy7jsisu58MILuebKa2XdKAiCIIgAIgiCIAjbE0cd+Tj/pCc9iUc/+gj2PWBvppZMsWLpMuIkbQSQehzs1jBcITCe7zdeCU3maYfGsZqmFWfUQ2QoWR3/eTyTVwWoqjxDLzaRZuGhPLTZKBQO5V0oA6kzYleZVOiqwqD5A/ToH1N6iFSoRKl8U5RSjDd5+C20Dam7sdRoHtNVbTF+wQ3ww2OR1SYfaDCep65s8ANzC1+WqCga+6QZVRGU2fRjL/r226EXtvDz9Iy2DdXTVfzQ3+7c4P02ajEBZPD5b25E8CYFkCHuWnsXs7Oz3HLLGn7961/zve99j7PP/rasDwVBEAQRQARBEAThwcxTnnqc//Pjn80zn/lMVq5cOSZkuMr7oPakGOSx49rFIJ/ddAuN9yGRbUxAm4qDqmKg6p/IsoI0jRd9jDBRo2qt8Qq3SAvFpp5bKbNIQu9Hk/MttVY8FFYJfkyAGBvXOvJ+qNEWJrymdBatowWVLuOmrZtrpaqrPaIhsaX2ORkfAe2qCp9BFZFrbm+trabTGMaluIHopkbGMQMUhW0m3nS7fc499wd84Qtf4Mwzz5R1oCAIgiACiCAIgiDsKBx66KH+oIMO4nGPexxHH300BxxwAEk12qMoCuI4Hklkh0f1Dk84qRPXOrmuE9vhs/nDSfJ4Auya1hY/lOSGUcD1CNvh11En1WmajjzesChTT2kZ/n1HY1HBYxHBYVjI0Pegr6cWvurnHRY96u/CsNlr/RltSgwZv+3w8xRFUU1vCR4dvV6PdjU15vLLL+fHP/4xl1xyCWvWrOH666/n6quvlnWfIAiCIAKIIAiCIAgD9tlnH3/ggQeyxx578IhHPII99tiDAw88kL322qsRR4bFiPFRr8OTUjaVTA+f0S/LEmPMgmR5sfsulgg3B/VtEDi2JAw8VBiujLg7YoZSqqoOcs1j1WasSimKohgRTsafZ/hzrB9j+HPOsowkSZr7ZVmGc44kSUb8QYY/Y+ccnU6HbrfLlVdeydVXX81VV13FDTfcwDXXXMOll14q6ztBEARBEAFEEARBEO4Z+++/v3/Uox7FUUcdxf7778/OO+/M9PQ0K1asYGZmZkQgWUysGE6iN5e0QxBPapFk+PbDVQbRiIfFAGvtgmS7rjYYT6h3BMb/7uEWlvq9Hn8vapFqMD55lLIsF7S11HH8s9ZaLxCdhr8feZ6jtW4ezzlHv99n3bp1bNiwgZtuuomrrrqK3/zmN1x++eVccMEFsoYTBEEQBBFABEEQBOGB49BDD/XHHHMMT3ziEznkkEPYZZddaLVaI4lyWZZNm0u/329aGJxzTcJdJ8x1gjzePjGcoA+LGeOJ/HA1So3Weof7XPyY/8fmvFsWu10tON3dFplhsWS45aWuClm3bl3TtnLuuedy3nnnyRpNEARBEO5l5OAqCIIgCPcDu+++u5+YmGDlypXsu+++HHjggaxevZo//dM/ZaeddmJmZmbRtpTNJd6LtcnUgsh4W86Ozta0/GyLyOGcoyxLkiRZ0B5TP59zjizLuOmmm7jmmmu48cYb+d3vfscVV1zBLbfcQpZlXH/99bIWEwRBEIT7CTnoCoIgCMJ2wn777ed32mknVq9ezS677MLk5CR77703y5YtY9WqVey2224sX76cqampxvy01+sRx/GCVow60a+rTTZVGbKjVoMMtxAtVl0DCyty6svWr1/PnXfeycaNG7nllluYm5vjzjvv5Lbbbmv+3XnnnczNzXHTTTfJWksQBEEQthPkoCwIgiAID0L23HNPv2zZMg466CBWrFjBHnvswapVq1i+fDkzMzNMTEyQpilTU1PEcUyr1aLVapEkCVEU7ZATYWAwPaUoCqy1zM7OkmUZ/X6fPM/p9/ts3LiRO+64gw0bNnDDDTdw5513csMNN7BmzRpuvPFGWTsJgiAIwoMUOYgLgiAIwg7IXnvt5XfbbTdarRZHHHEEaZoSxzGHHHIIS5cuxXvP6tWr2WeffQaLhkXG6Q63jNSeGcNmoPdUaFFKNYauw89TX2atJYoibr/9dq6++mq01szNzXHppZdSFAW9Xo8rrriCDRs2MDs7KxNSBEEQBGEHRhYBgiAIgiDcI3bbbTe/2KSae4NbbrlF1iqCIAiCINwr/P94fgUk370OXQAAAABJRU5ErkJggg=="
    if _LOGO_B64:
        return _LOGO_B64
    logo_paths = ["logo.png", "/mnt/user-data/uploads/logo.png"]
    for p in logo_paths:
        if os.path.exists(p):
            with open(p, "rb") as fh:
                return base64.b64encode(fh.read()).decode()
    return ""

def login():
    logo_b64 = get_logo_b64()
    logo_img = f'<img src="data:image/png;base64,{logo_b64}" style="width:160px;height:160px;object-fit:contain;display:block;">' if logo_b64 else '<div style="font-size:5rem;">⚡</div>'

    # Animated background + floating particles
    st.markdown("""
    <div class="login-bg"></div>
    <div style="position:fixed;inset:0;pointer-events:none;z-index:0;overflow:hidden;">
        <div class="particle" style="width:180px;height:180px;left:5%;animation-duration:18s;animation-delay:0s;"></div>
        <div class="particle" style="width:120px;height:120px;left:25%;animation-duration:22s;animation-delay:3s;"></div>
        <div class="particle" style="width:200px;height:200px;left:55%;animation-duration:16s;animation-delay:6s;"></div>
        <div class="particle" style="width:90px;height:90px;left:75%;animation-duration:20s;animation-delay:1s;"></div>
        <div class="particle" style="width:150px;height:150px;left:88%;animation-duration:24s;animation-delay:8s;"></div>
    </div>
    """, unsafe_allow_html=True)

    # Logo bölümü - büyük ve animasyonlu
    st.markdown(f"""
    <div style="text-align:center; padding: 48px 0 28px 0; position:relative; z-index:1;">
        <div class="login-logo-wrap" style="display:inline-block; position:relative;">
            <div class="login-logo-ring"></div>
            <div class="login-logo-ring2"></div>
            {logo_img}
        </div>
        <div style="margin-top:20px;">
            <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:2rem; font-weight:900;
                        letter-spacing:0.08em; color:#0f1923; text-transform:uppercase;">
                STINGA <span style="color:#11855B;">PRO</span>
            </div>
            <div style="font-family:'JetBrains Mono',monospace; font-size:0.65rem; color:#7a96a4;
                        letter-spacing:0.2em; margin-top:5px; text-transform:uppercase;">
                STİNGA ENERJİ A.Ş. · YÖNETİM PANELİ
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.1, 1])
    with col2:
        # Dekoratif üst çubuk
        st.markdown("""
        <div style="height:3px; background:linear-gradient(90deg,#11855B,#2F3C6E,#11855B);
                    border-radius:99px; margin-bottom:0;"></div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            st.markdown("""
            <div style="text-align:center; padding:16px 0 8px 0;">
                <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:1rem; font-weight:700;
                            color:#0f1923; letter-spacing:0.02em;">Hesabınıza Giriş Yapın</div>
                <div style="font-size:0.75rem; color:#7a96a4; margin-top:3px;">
                    Yetkili personel girişi
                </div>
            </div>
            """, unsafe_allow_html=True)

            username = st.text_input("Kullanıcı Adı", placeholder="kullanıcı adınız").lower()
            password = st.text_input("Şifre", type="password", placeholder="••••••••")
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            submit = st.form_submit_button("GİRİŞ YAP  →", use_container_width=True)

            if submit:
                if username in USERS and USERS[username]["password"] == hash_password(password):
                    st.session_state.authenticated = True
                    st.session_state.user_info = USERS[username]
                    st.session_state.user_info['username'] = username
                    st.session_state.splash_done = False
                    st.rerun()
                else:
                    st.error("⚠️ Hatalı kullanıcı adı veya şifre")

        st.markdown("""
        <div style="text-align:center; margin-top:20px; color:#7a96a4; font-size:0.72rem;
                    font-family:'JetBrains Mono',monospace; letter-spacing:0.05em;">
            🔒 256-BIT AES · ZERO-KNOWLEDGE AUTH · GEMINI AI
        </div>
        """, unsafe_allow_html=True)




def show_splash():
    """Şifre sonrası tam ekran WebGL giriş animasyonu."""
    import streamlit.components.v1 as components
    
    user_info = st.session_state.user_info
    user_name = user_info.get("name", "")
    avatar    = user_info.get("avatar", "⚡")

    # Tüm Streamlit UI'ı gizle — sadece splash görünsün
    st.markdown("""
    <style>
        header[data-testid="stHeader"] { display:none !important; }
        #MainMenu { display:none !important; }
        footer { display:none !important; }
        .block-container { padding:0 !important; max-width:100% !important; }
        [data-testid="stAppViewContainer"] { background:#000 !important; }
        [data-testid="stVerticalBlock"] { gap:0 !important; }
    </style>
    """, unsafe_allow_html=True)

    # JS → Python köprüsü: splash bitince query param set eder
    splash_html = f"""
<!DOCTYPE html>
<html>
<head>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ background:#000; overflow:hidden; font-family:'Space Grotesk',sans-serif; }}
  @import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Space+Grotesk:wght@300;600&display=swap');

  #c {{ position:fixed; inset:0; width:100%; height:100%; }}

  #ui {{
    position:fixed; inset:0;
    display:flex; flex-direction:column;
    align-items:center; justify-content:center;
    pointer-events:none; z-index:10;
  }}

  .logo-wrap {{
    display:flex; flex-direction:column; align-items:center; gap:14px;
    opacity:0; transform:translateY(24px);
    animation: rise 1.1s cubic-bezier(0.16,1,0.3,1) 0.4s forwards;
  }}

  .brand {{
    font-family:'Bebas Neue',sans-serif;
    font-size: min(11vw, 88px);
    letter-spacing:.18em;
    background:linear-gradient(135deg,#00e896 0%,#1a9e6e 50%,#2d4a8a 100%);
    -webkit-background-clip:text; -webkit-text-fill-color:transparent; background-clip:text;
    filter:drop-shadow(0 0 32px rgba(0,200,120,.6));
    line-height:1;
  }}
  .sub {{
    font-size:11px; letter-spacing:.42em;
    color:rgba(0,200,140,.55); text-transform:uppercase; font-weight:300;
  }}
  .welcome {{
    font-size:13px; letter-spacing:.25em;
    color:rgba(255,255,255,.35); text-transform:uppercase;
    margin-top:4px;
  }}

  .divider {{
    width:1px; height:44px;
    background:linear-gradient(to bottom,transparent,rgba(0,200,120,.35),transparent);
    margin:24px 0 20px;
    opacity:0; animation:rise .7s ease 1.3s forwards;
  }}

  .prog-wrap {{
    display:flex; flex-direction:column; align-items:center; gap:10px;
    opacity:0; animation:rise .7s ease 1.5s forwards;
  }}
  .prog-bar {{
    width:220px; height:1px;
    background:rgba(255,255,255,.07); position:relative; overflow:hidden;
  }}
  .prog-fill {{
    position:absolute; left:0; top:0; height:100%;
    background:linear-gradient(90deg,#00e896,#2d4a8a);
    width:0%; transition:width .12s linear;
    box-shadow:0 0 10px rgba(0,232,150,.9);
  }}
  .prog-lbl {{
    font-size:10px; letter-spacing:.32em;
    color:rgba(255,255,255,.28); text-transform:uppercase;
  }}

  .corner {{ position:fixed; width:36px; height:36px; opacity:0; animation:fade .5s ease 1.9s forwards; }}
  .tl {{ top:22px; left:22px; border-top:1px solid rgba(0,200,120,.25); border-left:1px solid rgba(0,200,120,.25); }}
  .tr {{ top:22px; right:22px; border-top:1px solid rgba(0,200,120,.25); border-right:1px solid rgba(0,200,120,.25); }}
  .bl {{ bottom:22px; left:22px; border-bottom:1px solid rgba(0,200,120,.25); border-left:1px solid rgba(0,200,120,.25); }}
  .br {{ bottom:22px; right:22px; border-bottom:1px solid rgba(0,200,120,.25); border-right:1px solid rgba(0,200,120,.25); }}

  .scan {{
    position:fixed; left:0; right:0; height:1px; top:-1px;
    background:linear-gradient(90deg,transparent,rgba(0,200,120,.12),transparent);
    animation:scan 4s linear 1.4s infinite;
  }}

  #flash {{
    position:fixed; inset:0;
    background:#fff; opacity:0; pointer-events:none; z-index:99;
  }}

  @keyframes rise {{ to {{ opacity:1; transform:translateY(0); }} }}
  @keyframes fade {{ to {{ opacity:1; }} }}
  @keyframes scan {{ 0% {{ top:-1px; }} 100% {{ top:100vh; }} }}
</style>
</head>
<body>
<canvas id="c"></canvas>

<div id="ui">
  <div class="logo-wrap">
    <svg width="82" height="82" viewBox="0 0 100 100" fill="none">
      <defs>
        <radialGradient id="g1" cx="50%" cy="30%" r="70%">
          <stop offset="0%" stop-color="#00e896"/>
          <stop offset="60%" stop-color="#1a9e6e"/>
          <stop offset="100%" stop-color="#0f2240"/>
        </radialGradient>
        <filter id="glow"><feGaussianBlur stdDeviation="1.5" result="b"/>
          <feMerge><feMergeNode in="b"/><feMergeNode in="SourceGraphic"/></feMerge>
        </filter>
      </defs>
      <circle cx="50" cy="50" r="47" stroke="url(#g1)" stroke-width="1" fill="none" opacity=".5"/>
      <circle cx="50" cy="50" r="42" stroke="rgba(0,200,120,.15)" stroke-width=".5"
              stroke-dasharray="4 8" fill="none">
        <animateTransform attributeName="transform" type="rotate"
          from="0 50 50" to="360 50 50" dur="12s" repeatCount="indefinite"/>
      </circle>
      <circle cx="50" cy="50" r="37" fill="url(#g1)" opacity=".9"/>
      <path filter="url(#glow)"
        d="M36 28 C36 28,54 28,56 35 C58 42,44 44,44 44 C44 44,62 46,64 54
           C66 62,48 72,44 72 L36 72 L42 72 C42 72,58 70,56 63
           C54 56,38 54,38 54 C38 54,56 52,54 44 C52 36,36 36,36 36 Z"
        fill="none" stroke="rgba(0,232,150,.9)" stroke-width="2"
        stroke-linecap="round" stroke-linejoin="round"/>
      <circle cx="50" cy="50" r="47" fill="none" stroke="rgba(0,200,120,.12)" stroke-width="2">
        <animate attributeName="r" values="47;53;47" dur="3s" repeatCount="indefinite"/>
        <animate attributeName="opacity" values=".12;0;.12" dur="3s" repeatCount="indefinite"/>
      </circle>
    </svg>
    <div class="brand">STINGA</div>
    <div class="sub">AI Finans Platformu · v15.0</div>
    <div class="welcome">{avatar} Hoş geldin, {user_name}</div>
  </div>

  <div class="divider"></div>

  <div class="prog-wrap">
    <div class="prog-bar"><div class="prog-fill" id="pf"></div></div>
    <div class="prog-lbl" id="pl">Sistem başlatılıyor</div>
  </div>
</div>

<div class="corner tl"></div>
<div class="corner tr"></div>
<div class="corner bl"></div>
<div class="corner br"></div>
<div class="scan"></div>
<div id="flash"></div>

<script>
// ── WEBGL SHADER ─────────────────────────────────
const cv = document.getElementById('c');
const gl = cv.getContext('webgl');
function rsz() {{
  cv.width  = window.innerWidth  * devicePixelRatio;
  cv.height = window.innerHeight * devicePixelRatio;
  cv.style.width  = window.innerWidth  + 'px';
  cv.style.height = window.innerHeight + 'px';
  gl.viewport(0,0,cv.width,cv.height);
}}
window.addEventListener('resize',rsz); rsz();

const vs = `attribute vec2 p; void main(){{gl_Position=vec4(p,0,1);}}`;
const fs = `
precision highp float;
uniform vec2 uR; uniform float uT;
#define TAU 6.2831853
vec2 warp(vec2 p,float t){{
  float a=atan(p.y,p.x),r=length(p);
  a+=.3*sin(r*3.-t*.8)+.15*cos(r*5.+t*.5);
  return vec2(cos(a),sin(a))*r;
}}
float ring(vec2 uv,float r,float w){{return smoothstep(w,.0,abs(length(uv)-r));}}
void main(){{
  vec2 uv=(gl_FragCoord.xy*2.-uR)/min(uR.x,uR.y);
  float t=uT*.4;
  float v=1.-smoothstep(.3,1.4,length(uv));
  vec3 col=mix(vec3(0.),vec3(.02,.05,.12),v);

  vec2 wp=warp(uv,t), wp2=warp(uv*.8,t*1.3+1.);
  float aur=0.;
  for(int i=0;i<5;i++){{
    float fi=float(i);
    float ph=fi*1.2+t*(.6+fi*.1);
    float b=sin(wp.x*(2.5+fi*.8)+ph)*.5+.5;
    b*=sin(wp2.y*(1.75+fi*.56)+ph*1.1)*.5+.5;
    aur+=b*(.18-fi*.025);
  }}
  aur=pow(aur,1.4);
  vec3 g1=vec3(0.,.91,.59),g2=vec3(.1,.62,.43),nav=vec3(.11,.19,.36);
  col+=mix(g2,g1,aur)*aur*1.8;

  float tp=mod(t*.7,1.);
  for(int k=0;k<6;k++){{
    float fk=float(k)/6.,ph=mod(fk+tp,1.);
    float al=(1.-ph)*(1.-ph);
    col+=ring(uv,ph*1.6,mix(.018,.004,ph))*al*mix(g1,nav,smoothstep(0.,1.6,length(uv)))*.9;
  }}
  col+=ring(uv,.35,.004)*g1*.45;

  float ang=atan(uv.y,uv.x),bm=0.;
  for(int b=0;b<8;b++){{
    float fb=float(b),ao=fb*TAU/8.+t*(.15+fb*.02);
    bm+=smoothstep(.06,.0,mod(ang-ao,TAU)/TAU)*.07;
  }}
  col+=bm*smoothstep(1.5,.1,length(uv))*g1*.6;
  col+=clamp(.06/(length(uv*.9)+.01)*.18,0.,.5)*mix(g1,vec3(1.),.3);
  float gr=fract(sin(dot(gl_FragCoord.xy,vec2(127.1+uT*.01,311.7+uT*.01)))*43758.5);
  col+=pow(max(col+(gr-.5)*.022,0.),vec3(.88));
  col=mix(col,col*vec3(.85,1.,.9),.25);
  gl_FragColor=vec4(col,1.);
}}`;

function sh(type,src){{const s=gl.createShader(type);gl.shaderSource(s,src);gl.compileShader(s);return s;}}
const pr=gl.createProgram();
gl.attachShader(pr,sh(gl.VERTEX_SHADER,vs));
gl.attachShader(pr,sh(gl.FRAGMENT_SHADER,fs));
gl.linkProgram(pr); gl.useProgram(pr);
const buf=gl.createBuffer();
gl.bindBuffer(gl.ARRAY_BUFFER,buf);
gl.bufferData(gl.ARRAY_BUFFER,new Float32Array([-1,-1,1,-1,-1,1,1,1]),gl.STATIC_DRAW);
const ap=gl.getAttribLocation(pr,'p');
gl.enableVertexAttribArray(ap);
gl.vertexAttribPointer(ap,2,gl.FLOAT,false,0,0);
const uR=gl.getUniformLocation(pr,'uR'),uT=gl.getUniformLocation(pr,'uT');
let t0=null;
(function loop(ts){{
  if(!t0)t0=ts;
  gl.uniform2f(uR,cv.width,cv.height);
  gl.uniform1f(uT,(ts-t0)/1000);
  gl.drawArrays(gl.TRIANGLE_STRIP,0,4);
  requestAnimationFrame(loop);
}})(0);

// ── PROGRESS ─────────────────────────────────────
const pf=document.getElementById('pf'),pl=document.getElementById('pl');
const steps=[[15,'Güvenlik protokolleri'],[35,'AI motoru yükleniyor'],
             [55,'Gemini bağlantısı kuruldu'],[72,'Finans verileri çekiliyor'],
             [88,'Dashboard hazırlanıyor'],[100,'Giriş yapılıyor...']];
let si=0;
function tick(){{
  if(si>=steps.length)return;
  const[p,m]=steps[si++];
  pf.style.width=p+'%'; pl.textContent=m;
  if(si<steps.length){{setTimeout(tick,260+Math.random()*200);}}
  else{{
    setTimeout(()=>{{
      const fl=document.getElementById('flash');
      fl.style.transition='opacity .2s ease';
      fl.style.opacity='1';
      // Streamlit parent'a sinyal gönder
      setTimeout(()=>{{
        try{{window.parent.postMessage({{type:'splash_done'}},  '*');}}catch(e){{}}
        // Fallback: URL hash değiştir
        try{{window.parent.location.hash='splash_done';}}catch(e){{}}
      }},150);
    }},500);
  }}
}}
setTimeout(tick, 1600);
</script>
</body>
</html>
"""

    # Tam ekran component render et
    components.html(splash_html, height=800, scrolling=False)

    # Streamlit tarafı: belirli süre sonra splash_done=True yapıp rerun
    import time
    
    # Progress süresi ~3.5sn + 0.5sn flash = 4sn
    # st.empty() placeholder ile bekle
    placeholder = st.empty()
    
    with placeholder:
        time.sleep(4.2)
    
    placeholder.empty()
    st.session_state.splash_done = True
    st.rerun()


def logout():
    for key in ['authenticated', 'user_info', 'chat_history']:
        if key == 'authenticated': st.session_state[key] = False
        elif key == 'user_info': st.session_state[key] = None
        else: st.session_state[key] = []
    st.rerun()

# ─── ANA UYGULAMA ─────────────────────────────────────────────
init_db()


if not st.session_state.authenticated:
    login()
elif not st.session_state.splash_done:
    show_splash()
else:
    data_store = load_data()
    model = configure_ai()
    user_info = st.session_state.user_info



    # ── Eski bozuk AI_Audit kayıtlarını client-side onar (tek seferlik) ──
    if not st.session_state.get("audit_fixed"):
        st.session_state["audit_fixed"] = True  # Tekrar denemeyi engelle
        try:
            for _e in data_store.get("expenses", []):
                _raw = str(_e.get("AI_Audit", ""))
                if "<div" in _raw or "</" in _raw or "style=" in _raw:
                    _clean = clean_audit(_raw)
                    _api_post("/update-expense", {"ID": _e.get("ID"), "AI_Audit": _clean})
            st.cache_data.clear()
        except Exception:
            pass
    user_name = user_info["name"]
    role = user_info["role"]

    # ── ROL BAZLI FİLTRELEME ──────────────────────────────────
    # admin (Zeynep Özyaman, Şenol Faik Özyaman, Serkan Güzdemir) → tüm fişleri görür + onaylayabilir
    # user  (Okan İlhan) → sadece kendi fişlerini görür
    df_full = pd.DataFrame(data_store.get("expenses", []))
    if role == "user" and not df_full.empty and "Kullanıcı" in df_full.columns:
        df = df_full[df_full["Kullanıcı"] == user_name].copy()
    else:
        df = df_full.copy()
    


    # ── ROBOT: st.markdown + parent window inject ────────────
    import json as _rj, streamlit.components.v1 as _stc
    _rtips = [
        f"Merhaba {user_name}! Hoş geldin 👋",
        "📑 Yeni fiş yüklemek için Fiş Tarama!",
        "⚖️ Onay bekleyen fişleri kontrol et!",
        "🔬 Anomali Dedektörü şüpheli fişleri bulur!",
        "💰 Kasa bakiyeni Finans ekranında gör!",
        "🤖 AI Asistan her konuda yardımcı olur!",
        "📄 Raporları PDF olarak indirebilirsin!",
    ]
    st.markdown(f"""
<style>
#SGNX{{position:fixed;bottom:28px;right:28px;z-index:2147483647;cursor:pointer;width:92px;animation:SGNXF 3.8s ease-in-out infinite;filter:drop-shadow(0 8px 24px rgba(17,133,91,.45));}}
#SGNX:hover{{filter:drop-shadow(0 12px 36px rgba(17,133,91,.8));transform:scale(1.07);}}
@keyframes SGNXF{{0%,100%{{transform:translateY(0)}}50%{{transform:translateY(-11px)}}}}
#SGNXSH{{width:56px;height:11px;margin:3px auto 0;background:radial-gradient(ellipse,rgba(17,133,91,.38) 0%,transparent 70%);border-radius:50%;animation:SGNXS 3.8s ease-in-out infinite;}}
@keyframes SGNXS{{0%,100%{{transform:scaleX(1);opacity:.38}}50%{{transform:scaleX(.45);opacity:.1}}}}
#SGNXH{{width:74px;height:74px;border-radius:50%;margin:0 auto;background:linear-gradient(145deg,#3d4e8a 0%,#2F3C6E 55%,#1a2540 100%);box-shadow:inset 0 4px 10px rgba(255,255,255,.13),inset 0 -4px 10px rgba(0,0,0,.3),0 6px 20px rgba(47,60,110,.55),0 0 0 2px rgba(0,232,150,.22);position:relative;transition:transform .1s ease-out;}}
#SGNXH::before{{content:'';position:absolute;top:11px;left:15px;width:24px;height:15px;border-radius:50%;background:radial-gradient(ellipse,rgba(255,255,255,.2) 0%,transparent 70%);transform:rotate(-20deg);}}
.SGNXEAR{{position:absolute;top:50%;transform:translateY(-50%);width:8px;height:18px;border-radius:4px;background:linear-gradient(180deg,#2F3C6E,#1a2540);border:1px solid rgba(0,232,150,.15);}}
.SGNXEAR::after{{content:'';position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);width:4px;height:4px;border-radius:50%;background:#00e896;box-shadow:0 0 4px #00e896;animation:SGNXBL 2.2s ease-in-out infinite;}}
.SGNXEARL{{left:-5px;}}.SGNXEARR{{right:-5px;}}
#SGNXANT{{position:absolute;top:-18px;left:50%;transform:translateX(-50%);display:flex;flex-direction:column;align-items:center;}}
#SGNXANTB{{width:10px;height:10px;border-radius:50%;background:#00e896;box-shadow:0 0 10px #00e896,0 0 20px rgba(0,232,150,.4);animation:SGNXANTA 1.6s ease-in-out infinite;}}
@keyframes SGNXANTA{{0%,100%{{box-shadow:0 0 8px #00e896,0 0 16px rgba(0,232,150,.4);transform:scale(1)}}50%{{box-shadow:0 0 16px #00e896,0 0 32px rgba(0,232,150,.6);transform:scale(1.3)}}}}
#SGNXANTS{{width:3px;height:13px;background:linear-gradient(180deg,#00e896,#11855B);border-radius:2px;}}
#SGNXEYES{{position:absolute;top:26px;left:50%;transform:translateX(-50%);width:56px;display:flex;justify-content:space-between;}}
.SGNXEYE{{width:23px;height:23px;border-radius:50%;background:#060d1a;box-shadow:inset 0 2px 6px rgba(0,0,0,.8),0 0 0 2px rgba(0,232,150,.3);position:relative;overflow:hidden;animation:SGNXBLINK 7s ease-in-out infinite;}}
.SGNXEYE:nth-child(2){{animation-delay:.08s;}}
@keyframes SGNXBLINK{{0%,44%,56%,100%{{transform:scaleY(1)}}50%{{transform:scaleY(.05)}}}}
.SGNXIRIS{{width:17px;height:17px;border-radius:50%;background:radial-gradient(circle at 35% 35%,#00e896 0%,#11855B 50%,#063d25 100%);position:absolute;top:3px;left:3px;box-shadow:0 0 8px rgba(0,232,150,.5);transition:transform .08s ease-out;}}
.SGNXPUPIL{{width:8px;height:8px;border-radius:50%;background:#000;position:absolute;top:4.5px;left:4.5px;}}
.SGNXEYE::after{{content:'';position:absolute;top:3px;right:3px;width:5px;height:5px;border-radius:50%;background:rgba(255,255,255,.8);pointer-events:none;z-index:2;}}
#SGNXBODY{{width:68px;height:60px;border-radius:20px 20px 16px 16px;background:linear-gradient(145deg,#17a870 0%,#0c6344 60%,#2F3C6E 100%);margin:2px auto 0;position:relative;box-shadow:inset 0 3px 8px rgba(255,255,255,.14),inset 0 -3px 8px rgba(0,0,0,.2),0 6px 20px rgba(12,99,68,.45);border:1.5px solid rgba(0,232,150,.22);}}
#SGNXBODY::before{{content:'';position:absolute;top:6px;left:9px;right:9px;height:16px;background:linear-gradient(180deg,rgba(255,255,255,.14) 0%,transparent 100%);border-radius:9px;}}
#SGNXCHEST{{position:absolute;bottom:9px;left:50%;transform:translateX(-50%);width:40px;height:18px;background:rgba(0,0,0,.25);border-radius:6px;border:1px solid rgba(0,232,150,.18);display:flex;align-items:center;justify-content:center;gap:4px;}}
.SGNXLED{{width:5px;height:5px;border-radius:50%;background:#00e896;box-shadow:0 0 5px #00e896;animation:SGNXBL 1.8s ease-in-out infinite;}}
.SGNXLED:nth-child(2){{animation-delay:.3s;background:#11855B;}}.SGNXLED:nth-child(3){{animation-delay:.6s;}}.SGNXLED:nth-child(4){{animation-delay:.9s;background:#2F3C6E;box-shadow:0 0 5px #3d4e8a;}}
@keyframes SGNXBL{{0%,100%{{opacity:1}}50%{{opacity:.2}}}}
.SGNXARM{{position:absolute;top:9px;width:12px;height:38px;border-radius:6px;background:linear-gradient(180deg,#17a870,#0c6344);border:1px solid rgba(0,232,150,.15);box-shadow:inset 0 2px 4px rgba(255,255,255,.1);}}
.SGNXARML{{left:-14px;animation:SGNXARMA 3.8s ease-in-out infinite;transform-origin:top center;}}
.SGNXARMR{{right:-14px;animation:SGNXARMB 3.8s ease-in-out infinite;transform-origin:top center;}}
@keyframes SGNXARMA{{0%,100%{{transform:rotate(-6deg)}}50%{{transform:rotate(9deg)}}}}
@keyframes SGNXARMB{{0%,100%{{transform:rotate(6deg)}}50%{{transform:rotate(-9deg)}}}}
.SGNXHAND{{position:absolute;bottom:-6px;left:50%;transform:translateX(-50%);width:12px;height:12px;border-radius:50%;background:linear-gradient(135deg,#17a870,#0c6344);border:1px solid rgba(0,232,150,.2);}}
.SGNXLEG{{position:absolute;bottom:-19px;width:16px;height:20px;border-radius:5px 5px 3px 3px;background:linear-gradient(180deg,#0c6344,#083d28);border:1px solid rgba(0,232,150,.12);}}
.SGNXLEGL{{left:9px;}}.SGNXLEGR{{right:9px;}}
.SGNXFOOT{{position:absolute;bottom:-7px;left:50%;transform:translateX(-50%);width:21px;height:7px;border-radius:3px;background:#083d28;border:1px solid rgba(0,232,150,.1);}}
#SGNXBUB{{position:fixed;bottom:196px;right:28px;background:linear-gradient(135deg,#fff,#f4f9f6);border:1.5px solid rgba(17,133,91,.28);border-radius:16px 16px 4px 16px;padding:12px 16px;min-width:210px;max-width:270px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:13px;color:#0f1923;line-height:1.55;box-shadow:0 8px 28px rgba(17,133,91,.14);opacity:0;transform:translateY(8px) scale(.93);transition:all .3s cubic-bezier(.16,1,.3,1);pointer-events:none;z-index:2147483646;}}
#SGNXBUB.on{{opacity:1;transform:translateY(0) scale(1);}}
#SGNXBUB span.badge{{font-size:10px;font-weight:800;color:#11855B;letter-spacing:2px;display:block;margin-bottom:5px;}}
#SGNXBUB::after{{content:'';position:absolute;bottom:-8px;right:18px;border-left:8px solid transparent;border-right:8px solid transparent;border-top:8px solid #fff;}}
#SGNXTC{{position:fixed;bottom:220px;right:28px;display:flex;flex-direction:column;gap:8px;pointer-events:none;z-index:2147483645;}}
.SGNXTOAST{{background:#fff;border-radius:12px;padding:12px 16px;min-width:260px;box-shadow:0 6px 24px rgba(0,0,0,.1);border-left:4px solid #11855B;display:flex;gap:10px;align-items:flex-start;animation:SGNXTI .35s cubic-bezier(.16,1,.3,1) both,SGNXTO .3s ease 4.5s forwards;cursor:pointer;pointer-events:all;}}
.SGNXTOAST.warning{{border-left-color:#d97706;}}.SGNXTOAST.error{{border-left-color:#dc2626;}}
@keyframes SGNXTI{{from{{transform:translateX(110%);opacity:0}}to{{transform:translateX(0);opacity:1}}}}
@keyframes SGNXTO{{to{{transform:translateX(110%);opacity:0}}}}

/* ── ROBOT DUYGU ANİMASYONLARI ── */
/* SEVİNÇ: kollar yukarı + zıplama */
#SGNX.joy{{animation:SGNXJOY .7s ease-out 7 forwards!important;}}
@keyframes SGNXJOY{{
  0%  {{transform:translateY(0) scale(1);}}
  20% {{transform:translateY(-22px) scale(1.08);}}
  40% {{transform:translateY(-32px) scale(1.1);}}
  60% {{transform:translateY(-18px) scale(1.06);}}
  80% {{transform:translateY(-8px) scale(1.02);}}
  100%{{transform:translateY(0) scale(1);}}
}}
#SGNX.joy .SGNXARML{{animation:SGNXARML_JOY .7s ease-out 4 forwards!important;}}
#SGNX.joy .SGNXARMR{{animation:SGNXARMR_JOY .7s ease-out 4 forwards!important;}}
@keyframes SGNXARML_JOY{{
  0%  {{transform:rotate(-6deg);}}
  30% {{transform:rotate(-150deg);}}
  60% {{transform:rotate(-140deg);}}
  80% {{transform:rotate(-145deg);}}
  100%{{transform:rotate(-6deg);}}
}}
@keyframes SGNXARMR_JOY{{
  0%  {{transform:rotate(6deg);}}
  30% {{transform:rotate(150deg);}}
  60% {{transform:rotate(140deg);}}
  80% {{transform:rotate(145deg);}}
  100%{{transform:rotate(6deg);}}
}}
/* Sevinç gözleri: ^^ */
#SGNX.joy .SGNXIRIS{{background:radial-gradient(circle at 50% 80%,#00e896 0%,#11855B 60%,#063d25 100%)!important;}}
#SGNX.joy .SGNXEYE{{animation:SGNXEYEJOY .7s ease 4 forwards!important;transform-origin:center bottom!important;}}
@keyframes SGNXEYEJOY{{
  0%,100%{{transform:scaleY(1);}}
  30%,70%{{transform:scaleY(.15) translateY(4px);}}
}}

/* ÜZÜNTÜ: sarkmak + sallantı */
#SGNX.sad{{animation:SGNXSAD .8s ease-in-out forwards!important;}}
@keyframes SGNXSAD{{
  0%  {{transform:translateY(0) rotate(0);}}
  20% {{transform:translateY(3px) rotate(-4deg);}}
  40% {{transform:translateY(5px) rotate(4deg);}}
  60% {{transform:translateY(3px) rotate(-3deg);}}
  80% {{transform:translateY(4px) rotate(2deg);}}
  100%{{transform:translateY(0) rotate(0);}}
}}
#SGNX.sad .SGNXARML{{animation:SGNXARML_SAD .8s ease-in-out forwards!important;}}
#SGNX.sad .SGNXARMR{{animation:SGNXARMR_SAD .8s ease-in-out forwards!important;}}
@keyframes SGNXARML_SAD{{
  0%,100%{{transform:rotate(-6deg);}}
  50%{{transform:rotate(40deg);}}
}}
@keyframes SGNXARMR_SAD{{
  0%,100%{{transform:rotate(6deg);}}
  50%{{transform:rotate(-40deg);}}
}}
/* Üzüntü gözleri: T_T  */
#SGNX.sad .SGNXIRIS{{background:radial-gradient(circle at 50% 70%,#5a8fff 0%,#2F3C6E 60%,#1a2540 100%)!important;}}
/* Göz yaşı damlaları */
#SGNX.sad .SGNXEYE::before{{
  content:'';position:absolute;bottom:-4px;left:50%;transform:translateX(-50%);
  width:5px;height:10px;
  background:linear-gradient(180deg,rgba(100,160,255,.9),rgba(100,160,255,0));
  border-radius:0 0 4px 4px;
  animation:SGNXTEAR .5s ease-in .2s infinite;
}}
@keyframes SGNXTEAR{{
  0%{{height:0;opacity:.9;transform:translateX(-50%) translateY(0);}}
  100%{{height:14px;opacity:0;transform:translateX(-50%) translateY(8px);}}
}}
/* Anten üzgün renk */
#SGNX.sad #SGNXANTB{{background:#5a8fff!important;box-shadow:0 0 8px #5a8fff!important;animation:none!important;}}

/* Sevinç konfeti overlay */
#SGNXJOYOVERLAY{{
  position:fixed;inset:0;z-index:2147483640;pointer-events:none;
  display:flex;align-items:center;justify-content:center;
  animation:SGNXOVF .4s ease forwards;
}}
@keyframes SGNXOVF{{from{{background:rgba(0,232,150,.18)}}to{{background:rgba(0,0,0,0)}}}}
/* Ret overlay */
#SGNXSADOVERLAY{{
  position:fixed;inset:0;z-index:2147483640;pointer-events:none;
  display:flex;align-items:center;justify-content:center;
  animation:SGNXOVFS .4s ease forwards;
}}
@keyframes SGNXOVFS{{from{{background:rgba(220,38,38,.12)}}to{{background:rgba(0,0,0,0)}}}}
#SGNXNOTIF-BADGE{{position:absolute;top:-10px;right:-10px;background:#dc2626;color:#fff;font-size:10px;font-weight:800;min-width:20px;height:20px;border-radius:50%;display:flex;align-items:center;justify-content:center;border:2px solid #fff;animation:SGNXBADGEP 1.2s ease-in-out infinite;z-index:9999;pointer-events:none;padding:0 3px;}}
@keyframes SGNXBADGEP{{0%,100%{{transform:scale(1);box-shadow:0 0 0 0 rgba(220,38,38,.6)}}50%{{transform:scale(1.18);box-shadow:0 0 0 6px rgba(220,38,38,0)}}}}
</style>
<div id="SGNXBUB"><span class="badge">⚡ STINGA AI</span><span id="SGNXBT">Merhaba!</span></div>
<div id="SGNXTC"></div>
<div id="SGNX">
  <div id="SGNXH">
    <div id="SGNXANT"><div id="SGNXANTB"></div><div id="SGNXANTS"></div></div>
    <div class="SGNXEAR SGNXEARL"></div><div class="SGNXEAR SGNXEARR"></div>
    <div id="SGNXEYES">
      <div class="SGNXEYE"><div class="SGNXIRIS" id="SGNXIL"><div class="SGNXPUPIL"></div></div></div>
      <div class="SGNXEYE"><div class="SGNXIRIS" id="SGNXIR"><div class="SGNXPUPIL"></div></div></div>
    </div>
  </div>
  <div id="SGNXBODY">
    <div class="SGNXARM SGNXARML"><div class="SGNXHAND"></div></div>
    <div class="SGNXARM SGNXARMR"><div class="SGNXHAND"></div></div>
    <div class="SGNXLEG SGNXLEGL"><div class="SGNXFOOT"></div></div>
    <div class="SGNXLEG SGNXLEGR"><div class="SGNXFOOT"></div></div>
    <div id="SGNXCHEST"><div class="SGNXLED"></div><div class="SGNXLED"></div><div class="SGNXLED"></div><div class="SGNXLED"></div></div>
  </div>
  <div id="SGNXSH"></div>
  <div id="SGNXNOTIF-BADGE" style="display:none;">0</div>
</div>
""", unsafe_allow_html=True)
    _stc.html(f"""<!DOCTYPE html><html><head></head><body><script>
(function(){{
  var tips={_rj.dumps(_rtips)},ti=0,cRX=0,cRY=0,cEX=0,cEY=0,mx=window.innerWidth/2,my=window.innerHeight/2;
  // Bildirim badge - sidebar markdown tarafından kontrol edilir
  var _nb=document.getElementById("SGNXNOTIF-BADGE");
  var head=document.getElementById('SGNXH'),iL=document.getElementById('SGNXIL'),iR=document.getElementById('SGNXIR'),
      bub=document.getElementById('SGNXBUB'),bt=document.getElementById('SGNXBT'),wrap=document.getElementById('SGNX');
  function tick(){{
    var tRX=-(my/window.innerHeight-.5)*26,tRY=(mx/window.innerWidth-.5)*26;
    cRX+=(tRX-cRX)*.1;cRY+=(tRY-cRY)*.1;
    if(head)head.style.transform='rotateX('+cRX.toFixed(1)+'deg) rotateY('+cRY.toFixed(1)+'deg)';
    var tEX=(mx/window.innerWidth-.5)*6,tEY=(my/window.innerHeight-.5)*6;
    cEX+=(tEX-cEX)*.12;cEY+=(tEY-cEY)*.12;
    var tr='translate('+cEX.toFixed(1)+'px,'+cEY.toFixed(1)+'px)';
    if(iL)iL.style.transform=tr;if(iR)iR.style.transform=tr;
    requestAnimationFrame(tick);
  }}
  tick();
  document.addEventListener('mousemove',function(e){{mx=e.clientX;my=e.clientY;}});
  // Parent window fare takibi (Streamlit iframe'i dışına çıkar)
  try{{window.parent.document.addEventListener('mousemove',function(e){{
    var fs=window.parent.document.querySelectorAll('iframe'),ol=0,ot=0;
    fs.forEach(function(f){{try{{if(f.contentWindow===window){{var r=f.getBoundingClientRect();ol=r.left;ot=r.top;}}}}catch(x){{}}}});
    mx=e.clientX-ol;my=e.clientY-ot;
  }});}}catch(x){{}}
  if(wrap)wrap.addEventListener('click',function(){{
    if(_nb)_nb.style.display='none';
    if(bt)bt.textContent=tips[ti%tips.length];ti++;
    if(bub){{bub.classList.add('on');clearTimeout(wrap._t);wrap._t=setTimeout(function(){{bub.classList.remove('on');}},4800);}}
  }});
  setTimeout(function(){{if(bt)bt.textContent=tips[0];ti=1;if(bub){{bub.classList.add('on');setTimeout(function(){{bub.classList.remove('on');}},5000);}};}},1200);
  window._sgShowToast=function(title,msg,type){{
    var c=document.getElementById('SGNXTC');if(!c)return;
    var el=document.createElement('div');el.className='SGNXTOAST '+(type||'');
    var ic={{success:'✅',warning:'⚠️',error:'🚨',info:'🔔'}};
    el.innerHTML='<span style="font-size:1.2rem">'+(ic[type]||'📌')+'</span><div><div style="font-weight:700;font-size:.82rem;margin-bottom:2px">'+title+'</div><div style="font-size:.73rem;color:#5a7a6a">'+msg+'</div></div>';
    el.onclick=function(){{el.remove();}};c.appendChild(el);
    setTimeout(function(){{if(el.parentNode)el.remove();}},5000);
  }};

  // ── ROBOT SEVİNÇ (onay) ───────────────────────────────────
  window._sgRobotJoy=function(){{
    var r=document.getElementById('SGNX');
    var bt=document.getElementById('SGNXBT'),bub=document.getElementById('SGNXBUB');
    if(!r)return;
    r.classList.remove('sad');
    r.classList.add('joy');
    // Konfeti overlay
    var ov=document.createElement('div');ov.id='SGNXJOYOVERLAY';
    document.body.appendChild(ov);
    // Canvas konfeti
    var cv=document.createElement('canvas');
    cv.style.cssText='position:fixed;inset:0;width:100%;height:100%;pointer-events:none;z-index:2147483641;';
    cv.width=window.innerWidth;cv.height=window.innerHeight;
    document.body.appendChild(cv);
    var ctx=cv.getContext('2d');
    var particles=[];
    for(var i=0;i<140;i++){{
      particles.push({{
        x:Math.random()*cv.width,y:-20+Math.random()*-80,
        vx:(Math.random()-.5)*8,vy:Math.random()*5+3,
        r:Math.random()*7+3,angle:Math.random()*360,av:(Math.random()-.5)*10,
        color:['#00e896','#17a870','#2F3C6E','#f0a500','#ffffff','#3d4e8a'][Math.floor(Math.random()*6)]
      }});
    }}
    var fr=0;
    function drawC(){{
      ctx.clearRect(0,0,cv.width,cv.height);
      particles.forEach(function(p){{
        p.x+=p.vx;p.y+=p.vy;p.angle+=p.av;p.vy+=.1;
        ctx.save();ctx.translate(p.x,p.y);ctx.rotate(p.angle*Math.PI/180);
        ctx.fillStyle=p.color;ctx.globalAlpha=Math.max(0,1-fr/100);
        ctx.fillRect(-p.r/2,-p.r/2,p.r,p.r*1.7);ctx.restore();
      }});
      fr++;if(fr<110)requestAnimationFrame(drawC);else cv.remove();
    }}
    drawC();
    // Balon mesajı
    if(bt)bt.textContent='🎉 ONAYLANDI! Harika iş!';
    if(bub){{bub.classList.add('on');}}
    setTimeout(function(){{
      r.classList.remove('joy');
      if(ov.parentNode)ov.remove();
      if(bub)bub.classList.remove('on');
    }},5000);
  }};

  // ── ROBOT ÜZÜNTÜ (ret) ───────────────────────────────────
  window._sgRobotSad=function(){{
    var r=document.getElementById('SGNX');
    var bt=document.getElementById('SGNXBT'),bub=document.getElementById('SGNXBUB');
    if(!r)return;
    r.classList.remove('joy');
    r.classList.add('sad');
    // Kırmızı overlay
    var ov=document.createElement('div');ov.id='SGNXSADOVERLAY';
    document.body.appendChild(ov);
    // Balon mesajı
    if(bt)bt.textContent='😢 Fiş reddedildi...';
    if(bub){{bub.classList.add('on');}}
    setTimeout(function(){{
      r.classList.remove('sad');
      if(ov.parentNode)ov.remove();
      if(bub)bub.classList.remove('on');
    }},2500);
  }};

}})();
</script></body></html>""", height=0, scrolling=False)

    # ── SIDEBAR ──────────────────────────────────────────────
    with st.sidebar:
        logo_b64 = get_logo_b64()
        user_xp = data_store.get("xp", {}).get(user_name, 0)
        level = user_xp // 500 + 1
        xp_progress = (user_xp % 500) / 500
        # Okundu takibi session_state id seti ile
        if "read_notif_ids" not in st.session_state:
            st.session_state.read_notif_ids = set()
        _read_ids = st.session_state.read_notif_ids
        def _notif_key(n):
            return f"{n.get('user','')}|{n.get('msg','')}|{n.get('date','')}|{n.get('time','')}"
        my_notifs = [n for n in data_store.get("notifications", [])
                     if (str(n.get("user","")) == user_name or str(n.get("user","")) == "Hepsi"
                         or str(n.get("user","")).lower() == user_name.lower())
                     and _notif_key(n) not in _read_ids]
        # session_state'teki lokal bildirimleri de ekle
        local_notifs = st.session_state.get("local_notifications", [])
        my_notifs = my_notifs + [n for n in local_notifs if not n.get("read", False) and _notif_key(n) not in _read_ids]
        notif_count = len(my_notifs)

        if not df_full.empty and 'Tutar' in df_full.columns:
            my_total = df_full[df_full['Kullanıcı'] == user_name]['Tutar'].sum() if 'Kullanıcı' in df_full.columns else 0
            monthly_limit = user_info.get('monthly_limit', 15000)
            usage_pct = min(my_total / monthly_limit * 100, 100) if monthly_limit > 0 else 0
        else:
            my_total = 0
            monthly_limit = user_info.get('monthly_limit', 15000)
            usage_pct = 0
        usage_color = "#3ecf8e" if usage_pct < 60 else ("#f0a500" if usage_pct < 85 else "#e03a48")
        role_labels = {"admin": "YÖNETİCİ", "user": "PERSONEL"}
        role_colors = {"admin": "#3ecf8e", "user": "#7eb8d4"}
        role_label = user_info.get("title", role_labels.get(role, role.upper()))
        role_color = role_colors.get(role, "#00d4ff")

        # ── 1. LOGO
        logo_img_html = f'<img src="data:image/png;base64,{logo_b64}">' if logo_b64 else '<span style="font-size:2rem;">⚡</span>'
        st.markdown(f"""
        <div style="text-align:center; padding:28px 10px 12px;">
            <div class="slogo-frame" style="display:inline-block;">
                <div class="slogo-core">{logo_img_html}</div>
            </div>
            <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:0.88rem; font-weight:900;
                        letter-spacing:0.3em; color:#0f1923; margin-top:13px; text-transform:uppercase;">
                STINGA PRO
            </div>
            <div style="font-family:'JetBrains Mono',monospace; font-size:0.51rem; color:#a0b8ae;
                        letter-spacing:0.18em; margin-top:4px; text-transform:uppercase;">
                FINANCE v17.0 · STİNGA ENERJİ A.Ş.
            </div>
        </div>
        <div class="ssep"></div>
        """, unsafe_allow_html=True)

        # ── 2. USER CARD
        st.markdown(f"""
        <div class="suser-card">
            <div class="suser-top">
                <div class="suser-ava">{user_info['avatar']}</div>
                <div>
                    <div class="suser-name">{user_name}</div>
                    <div class="suser-role" style="color:{role_color};">{role_label} · LV.{level}</div>
                </div>
            </div>
            <div class="sxp-row">
                <span>⚡ {user_xp} XP</span>
                <span>→ {level * 500} XP</span>
            </div>
            <div class="sxp-track">
                <div class="sxp-fill" style="width:{xp_progress*100:.1f}%;"></div>
            </div>
            <div class="smeta">
                <span>{user_info.get('title', user_info.get('department','—'))}</span>
                <span style="color:{usage_color}; font-weight:700;">%{usage_pct:.0f} limit</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── 3. NOTIFICATION
        # Robot badge'ini _stc.html ile güncelle (st.markdown script çalıştırmaz)
        _stc.html(f"""<!DOCTYPE html><html><body><script>
        (function(){{
          function setBadge(){{
            var frames=window.parent.document.querySelectorAll('iframe');
            for(var i=0;i<frames.length;i++){{
              try{{
                var doc=frames[i].contentDocument||frames[i].contentWindow.document;
                var nb=doc.getElementById('SGNXNOTIF-BADGE');
                if(nb){{
                  nb.textContent='{notif_count}';
                  nb.style.display={notif_count}>0?'flex':'none';
                  return;
                }}
              }}catch(e){{}}
            }}
          }}
          setBadge();
          setTimeout(setBadge,500);
          setTimeout(setBadge,1500);
        }})();
        </script></body></html>""", height=0, scrolling=False)

        if notif_count > 0:
            st.markdown(f"""
            <div class="snotif">
                <span>🔔</span>
                <span class="snotif-lbl">{notif_count} yeni bildirim</span>
                <span class="snotif-num">{notif_count}</span>
            </div>
            """, unsafe_allow_html=True)
            with st.expander("Bildirimleri Gör"):
                # Tüm görünen bildirimleri okundu olarak işaretle (session_state id seti)
                for n in my_notifs:
                    st.session_state.read_notif_ids.add(_notif_key(n))
                for n in st.session_state.get("local_notifications", []):
                    n["read"] = True
                for n in reversed(my_notifs[-5:]):
                    icon = {"xp": "🏆", "info": "ℹ️", "warning": "⚠️", "success": "✅"}.get(n.get("type", "info"), "📌")
                    st.markdown(f"""
                    <div style="padding:7px 0; border-bottom:1px solid rgba(255,255,255,0.06);
                                display:flex; gap:9px; align-items:flex-start;">
                        <span>{icon}</span>
                        <div>
                            <div style="font-size:0.77rem; color:rgba(255,255,255,0.7);">{n['msg']}</div>
                            <div style="font-size:0.63rem; color:rgba(255,255,255,0.28); font-family:'JetBrains Mono',monospace; margin-top:1px;">{n.get('date','')} {n['time']}</div>
                        </div>
                    </div>""", unsafe_allow_html=True)

        # ── 4. SAAT / TARİH (İstanbul UTC+3) — components.html ile çalışır
        _stc.html(f"""<!DOCTYPE html><html><head>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:transparent;overflow:hidden;}}
#box{{
  text-align:center;padding:8px 12px 6px;
  background:rgba(17,133,91,0.08);border-radius:10px;
  margin:4px 8px 2px;border:1px solid rgba(17,133,91,0.15);
}}
#clk{{font-family:'JetBrains Mono',monospace;font-size:1.3rem;
  font-weight:700;color:#11855B;letter-spacing:2px;}}
#dat{{font-family:'JetBrains Mono',monospace;font-size:0.6rem;
  color:#7a96a4;letter-spacing:1px;margin-top:2px;}}
</style>
</head><body>
<div id="box">
  <div id="clk">--:--:--</div>
  <div id="dat">-- --- ----</div>
</div>
<script>
var DAYS=['Pazar','Pazartesi','Sali','Carsamba','Persembe','Cuma','Cumartesi'];
var MONTHS=['Ocak','Subat','Mart','Nisan','Mayis','Haziran','Temmuz','Agustos','Eylul','Ekim','Kasim','Aralik'];
function pad(n){{return n<10?'0'+n:n;}}
function tick(){{
  var now=new Date(new Date().getTime()+3*3600*1000);
  var h=now.getUTCHours(),m=now.getUTCMinutes(),s=now.getUTCSeconds();
  var d=now.getUTCDate(),mo=now.getUTCMonth(),y=now.getUTCFullYear(),dw=now.getUTCDay();
  document.getElementById('clk').textContent=pad(h)+':'+pad(m)+':'+pad(s);
  document.getElementById('dat').textContent=DAYS[dw]+', '+d+' '+MONTHS[mo]+' '+y;
}}
tick();setInterval(tick,1000);
</script>
</body></html>""", height=72, scrolling=False)

        # ── 5. SEPARATOR + NAV HEADER
        st.markdown('<div class="ssep" style="margin-top:6px;"></div><div class="snav-hdr">Navigasyon</div>', unsafe_allow_html=True)

        # ── 5. NAVIGATION (native Streamlit radio — styled via CSS)
        if role == "admin":
            pages_keys = [
                "🏠 Dashboard", "📑 Fiş Tarama", "💰 Finans & Kasa",
                "⚖️ Onay Merkezi", "🔬 Anomali Dedektörü", "📊 Analitik Merkezi",
                "🤖 AI Asistan", "🏆 Leaderboard",
                "🧠 AI Bütçe Koçu", "🔮 Gider Tahmincisi",
                "📂 Gider Kategorileri", "🗄️ Arşiv & Rapor",
            ]
        else:
            # Personel: Dashboard, Fiş Tarama + 2 yeni AI özellik
            pages_keys = [
                "🏠 Dashboard", "📑 Fiş Tarama",
                "🧠 AI Bütçe Koçu", "🔮 Gider Tahmincisi"
            ]

        selected = st.radio("", pages_keys,
                           index=pages_keys.index(st.session_state.selected_page)
                           if st.session_state.selected_page in pages_keys else 0,
                           label_visibility="collapsed")
        st.session_state.selected_page = selected

        # ── 6. LIMIT BAR
        st.markdown(f"""
        <div class="ssep" style="margin-bottom:6px;"></div>
        <div class="slimit">
            <div class="slimit-lbl">Aylık Limit Kullanımı</div>
            <div class="slimit-row">
                <span class="slimit-pct" style="color:{usage_color};">{usage_pct:.0f}%</span>
                <span class="slimit-val">{my_total:,.0f} / {monthly_limit:,.0f} ₺</span>
            </div>
            <div class="slimit-bar">
                <div style="height:100%; width:{usage_pct:.0f}%; background:{usage_color};
                            box-shadow:0 0 8px {usage_color}99; border-radius:99px;
                            transition:width 1s ease;"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── 7. LOGOUT
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("⏻  Oturumu Kapat", use_container_width=True):
            logout()

    # ══════════════════════════════════════════════════════════
    # SAYFA: DASHBOARD
    # ══════════════════════════════════════════════════════════
    if selected == "🏠 Dashboard":
        st.markdown('<div class="page-header"><div class="page-title">OPERASYON MERKEZİ</div></div>', unsafe_allow_html=True)
        
        # KPI Row
        # Admin → tüm ekip verisi; user → sadece kendi
        _kpi_df = df_full if role == "admin" else df
        total_approved = _kpi_df[_kpi_df['Durum']=='Onaylandı']['Tutar'].sum() if not _kpi_df.empty and 'Durum' in _kpi_df.columns else 0
        total_pending  = _kpi_df[_kpi_df['Durum']=='Onay Bekliyor']['Tutar'].sum() if not _kpi_df.empty and 'Durum' in _kpi_df.columns else 0
        crit_risks     = len(_kpi_df[_kpi_df['Risk_Skoru'] > 70]) if not _kpi_df.empty and 'Risk_Skoru' in _kpi_df.columns else 0
        # Kasa bakiyesi: ledger avansları - onaylı harcırah/nakit harcamaları
        # Negatif = şirkete borç (cebinden ödedi)
        my_wallet = get_user_wallet_balance(user_name, data_store)
        total_tx       = len(_kpi_df) if not _kpi_df.empty else 0
        avg_risk       = _kpi_df['Risk_Skoru'].mean() if not _kpi_df.empty and 'Risk_Skoru' in _kpi_df.columns else 0
        
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        
        metrics = [
            (c1, "Onaylı Harcama", f"₺{total_approved:,.0f}", "✓", "#007850"),
            (c2, "Onay Bekleyen", f"₺{total_pending:,.0f}", "⏳", "#d97706"),
            (c3, "Kritik Risk", str(crit_risks), "⛔", "#dc2626"),
            (c4, "Kasa Bakiye",
             f"{'−' if my_wallet < 0 else ''}₺{abs(my_wallet):,.0f}{'  ⚠️ Borç' if my_wallet < 0 else ''}",
             "💰" if my_wallet >= 0 else "🔴",
             "#007850" if my_wallet >= 0 else "#dc2626"),
            (c5, "Toplam İşlem", str(total_tx), "📊", "#283c64"),
            (c6, "Ort. Risk %", f"{avg_risk:.0f}", "🎯", "#ea6c1e"),
        ]
        
        for col, label, value, icon, color in metrics:
            with col:
                st.markdown(f"""
                <div class="metric-card">
                    <div style="font-size:1.5rem; margin-bottom:4px;">{icon}</div>
                    <div style="font-family:'Bebas Neue'; font-size:2rem; color:{color}; 
                                text-shadow:0 0 20px {color}66; line-height:1.1;">{value}</div>
                    <div style="font-size:0.65rem; color:var(--text-muted); text-transform:uppercase; 
                                letter-spacing:1.5px; margin-top:4px;">{label}</div>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Liquidity warning
        if role == "admin":
            total_wallet = sum(data_store.get('wallets',{}).values())
            if total_wallet < total_pending:
                st.markdown(f"""
                <div class="anomaly-alert">
                    <strong style="color:#dc2626;">⚠️ LİKİDİTE UYARISI</strong>
                    <p style="margin:4px 0 0 0; font-size:0.85rem; color:#7f1d1d;">
                        Toplam kasa bakiyesi ({total_wallet:,.0f} ₺), onay bekleyen tutarın ({total_pending:,.0f} ₺) altında.
                        Acil sermaye transferi gerekiyor.
                    </p>
                </div>
                """, unsafe_allow_html=True)
        
        # Charts Row
        if not df.empty:
            col_l, col_r = st.columns([3, 2])
            
            with col_l:
                if 'Tarih' in df.columns and 'Tutar' in df.columns:
                    df_temp = df.copy()
                    df_temp['dt'] = pd.to_datetime(df_temp['Tarih'], errors='coerce')
                    df_temp = df_temp.sort_values('dt')
                    df_temp['cumulative'] = df_temp['Tutar'].cumsum()
                    
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(
                        x=df_temp['dt'], y=df_temp['cumulative'],
                        fill='tozeroy',
                        fillcolor='rgba(0,212,255,0.08)',
                        line=dict(color='#007850', width=2),
                        name='Kümülatif Harcama',
                        hovertemplate='<b>%{x}</b><br>₺%{y:,.0f}<extra></extra>'
                    ))
                    fig.add_trace(go.Bar(
                        x=df_temp['dt'], y=df_temp['Tutar'],
                        marker_color='rgba(139,92,246,0.4)',
                        name='Günlük İşlem',
                        yaxis='y2'
                    ))
                    fig.update_layout(
                        title="📈 Zaman Serisi: Kümülatif & Günlük Harcama",
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568', family='Space Grotesk'),
                        xaxis=dict(gridcolor='rgba(203,213,224,0.8)', showgrid=True),
                        yaxis=dict(gridcolor='rgba(203,213,224,0.8)', showgrid=True),
                        yaxis2=dict(overlaying='y', side='right', showgrid=False),
                        legend=dict(bgcolor='rgba(0,0,0,0)'),
                        hovermode='x unified',
                        height=320
                    )
                    st.plotly_chart(fig, use_container_width=True)
            
            with col_r:
                if 'Proje' in df.columns:
                    proj_data = df.groupby('Proje')['Tutar'].sum().reset_index()
                    fig2 = go.Figure(go.Pie(
                        labels=proj_data['Proje'],
                        values=proj_data['Tutar'],
                        hole=0.6,
                        marker=dict(colors=['#007850','#283c64','#007850','#ea6c1e']),
                        textfont=dict(family='Space Grotesk')
                    ))
                    fig2.update_layout(
                        title="🥧 Proje Bazlı Dağılım",
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568'),
                        height=320,
                        legend=dict(bgcolor='rgba(0,0,0,0)')
                    )
                    st.plotly_chart(fig2, use_container_width=True)
            
            # Risk bubble + category bar
            col_a, col_b = st.columns(2)
            
            with col_a:
                if 'Risk_Skoru' in df.columns:
                    fig3 = px.scatter(
                        df, x='Tarih', y='Tutar', 
                        size='Risk_Skoru', color='Risk_Skoru',
                        hover_name='Firma',
                        color_continuous_scale='RdYlGn_r',
                        title='🎯 Risk Analizi (Balon Büyüklüğü = Risk)',
                        size_max=35,
                        hover_data={'Risk_Skoru': True, 'Proje': True}
                    )
                    fig3.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568'), height=280
                    )
                    st.plotly_chart(fig3, use_container_width=True)
            
            with col_b:
                if 'Kategori' in df.columns:
                    cat_data = df.groupby('Kategori')['Tutar'].sum().sort_values(ascending=True).reset_index()
                    fig4 = go.Figure(go.Bar(
                        x=cat_data['Tutar'], y=cat_data['Kategori'],
                        orientation='h',
                        marker=dict(
                            color=cat_data['Tutar'],
                            colorscale=[[0,'#1a2d4a'],[0.5,'#007850'],[1,'#283c64']]
                        ),
                        text=[f"₺{v:,.0f}" for v in cat_data['Tutar']],
                        textposition='outside'
                    ))
                    fig4.update_layout(
                        title='📂 Kategori Bazlı Harcama',
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568'), height=280,
                        xaxis=dict(showgrid=False), yaxis=dict(showgrid=False)
                    )
                    st.plotly_chart(fig4, use_container_width=True)
        
        else:
            st.markdown("""
            <div class="empty-state">
                <div class="empty-icon">📊</div>
                <div style="color:var(--text-secondary);">Henüz harcama verisi yok. Fiş tarama ile başla!</div>
            </div>
            """, unsafe_allow_html=True)
        
        # AI Daily Brief
        if model and not df_full.empty:
            with st.expander("🤖 Günlük AI Finansal Brifingi", expanded=True):
                if st.button("⚡ Günlük Analizi Oluştur"):
                    with st.spinner("Gemini AI verilerini işliyor..."):
                        insight = generate_ai_insight(df_full if role == "admin" else df, model)
                        st.session_state.last_ai_insight = insight
                
                if st.session_state.last_ai_insight:
                    st.markdown(f"""
                    <div class="ai-bubble">
                        <p style="margin:0; line-height:1.7; color:var(--text-primary); font-size:0.9rem;">
                            {st.session_state.last_ai_insight}
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: FİŞ TARAMA
    # ══════════════════════════════════════════════════════════
    elif selected == "📑 Fiş Tarama":
        st.markdown('<div class="page-header"><div class="page-title">AKILLI FİŞ TARAMA</div></div>', unsafe_allow_html=True)

        tab_fis, tab_manuel = st.tabs(["📸 Fiş Tara (AI)", "✏️ Manuel Giriş"])

        with tab_fis:
            col_form, col_list = st.columns([1.2, 1])

            with col_form:
                st.markdown('<div class="ultra-card">', unsafe_allow_html=True)
                st.markdown("### 📸 Fiş Yükle & AI Analiz")
            
                with st.form("pro_entry", clear_on_submit=False):
                    f = st.file_uploader("Fiş / Fatura Fotoğrafı", type=['jpg','png','jpeg','webp'],
                                         help="Net, iyi aydınlatılmış fotoğraflar için en iyi sonucu alırsınız")
                
                    col_p, col_o = st.columns(2)
                    with col_p:
                        proje = st.selectbox("Proje", ["Maden Sahası", "Aktif Karbon", "Enerji Hatları", "Genel Merkez"])
                    with col_o:
                        oncelik = st.selectbox("Öncelik", ["Normal", "Acil", "Düşük"])

                    # ── Ödeme Tipi ─────────────────────────────────
                    _odeme_secenekler = [
                        "🏦 Harcırahtan Düş (Nakit / Kişisel Kart)",
                        "💳 Şirket Kredi Kartı"
                    ]
                    odeme_tipi = st.radio(
                        "💳 Bu harcama nasıl ödendi?",
                        _odeme_secenekler,
                        index=st.session_state.get("_fis_odeme_idx", 0),
                        horizontal=True,
                        key="fis_odeme_radio",
                        help="Harcırahtan düşülürse kasa bakiyenizden azalır. Şirket kartı ise şirket borcuna eklenir."
                    )
                    # Seçimi session_state'e kaydet (clear_on_submit sonrası kaybolmasın)
                    st.session_state["_fis_odeme_idx"] = _odeme_secenekler.index(odeme_tipi)
                    odeme_turu_sec = "harcirah" if "Harcırahtan" in odeme_tipi else "sirket_karti"

                    notlar = st.text_area("Ek Not (isteğe bağlı)", height=60, placeholder="Harcamayla ilgili açıklama...")
                
                    submitted = st.form_submit_button("🚀 AI ile Tara ve Gönder", use_container_width=True)
                
                    if submitted and f:
                        # Ödeme türünü rerun'dan önce session_state'e kaydet
                        st.session_state["_submitted_odeme_turu"] = odeme_turu_sec
                        with st.spinner("🤖 Gemini AI fişi analiz ediyor..."):
                            progress = st.progress(0)
                            for i in range(70):
                                time.sleep(0.01)
                                progress.progress(i + 1)
                        
                            img = Image.open(f)
                            res_raw = analyze_receipt_pro(img, model)
                            progress.progress(100)
                        
                            if res_raw == "ANAHTAR_DEGISIMI":
                                st.warning("API kotası doldu, yedek anahtara geçildi. Tekrar deneyin.")
                            elif res_raw.startswith("HATA"):
                                st.error(f"Hata: {res_raw}")
                            else:
                                data_ai = extract_json(res_raw)
                                if data_ai:
                                    # ── Mükerrer Fiş Kontrolü (cache bypass) ──
                                    st.cache_data.clear()
                                    _fresh = load_data()
                                    firma_yeni = str(data_ai.get("firma","")).strip().lower()
                                    tutar_yeni = float(data_ai.get("toplam_tutar",0))
                                    tarih_yeni = str(data_ai.get("tarih",""))
                                    dup_blok = None
                                    dup_uyar = None
                                    for e in _fresh.get("expenses",[]):
                                        f_eski = str(e.get("Firma","")).strip().lower()
                                        t_eski = float(e.get("Tutar",0))
                                        d_eski = str(e.get("Tarih",""))
                                        t_esit = abs(t_eski - tutar_yeni) < 1.0
                                        f_esit = f_eski == firma_yeni or (len(firma_yeni)>3 and (firma_yeni in f_eski or f_eski in firma_yeni))
                                        if f_esit and t_esit:
                                            if d_eski == tarih_yeni:
                                                dup_blok = e; break
                                            else:
                                                dup_uyar = e
                                    if dup_blok:
                                        kim = dup_blok.get("Kullanıcı","?")
                                        mesaj = f"Bu fişi **sen** daha önce yükledin." if kim==user_name else f"Bu fişi **{kim}** yüklemiş."
                                        st.error(f"⛔ MÜKERRER FİŞ! {mesaj}\n\nFirma: {dup_blok.get('Firma')} | ₺{float(dup_blok.get('Tutar',0)):,.0f} | {dup_blok.get('Tarih')}")
                                    elif dup_uyar:
                                        st.warning(f"⚠️ Benzer fiş mevcut: {dup_uyar.get('Firma')} ₺{float(dup_uyar.get('Tutar',0)):,.0f} — {dup_uyar.get('Kullanıcı')} tarafından {dup_uyar.get('Tarih')} tarihinde yüklenmiş.")

                                    if not dup_blok:
                                        # ── Tarih Doğrulama (Python tarafında) ──
                                        tarih_str = data_ai.get("tarih", now_ist().strftime("%Y-%m-%d"))
                                        try:
                                            tarih_dt = datetime.strptime(tarih_str, "%Y-%m-%d")
                                            bugun_dt = now_ist()
                                            if tarih_dt > bugun_dt:
                                                # AI yanlış tarih verdiyse bugünün tarihine çek
                                                tarih_str = bugun_dt.strftime("%Y-%m-%d")
                                                data_ai["tarih"] = tarih_str
                                                st.info("ℹ️ Fiş tarihi gelecek olarak tespit edildi, bugünün tarihi kullanıldı.")
                                        except:
                                            tarih_str = now_ist().strftime("%Y-%m-%d")
                                            data_ai["tarih"] = tarih_str

                                        # ── İş Kuralı Motoru Uygula ──
                                        data_ai = apply_business_rules(data_ai, _fresh, user_name)
                                        uyarilar = data_ai.pop("_uyarilar", [])

                                        # Görseli hem lokal hem base64 olarak kaydet
                                        # base64 → bot'tan gelen WhatsApp fişlerinde de görsel görünsün
                                        yukleme_zamani = now_ist().strftime("%Y-%m-%d %H:%M:%S")
                                        path = f"arsiv/{now_ist().strftime('%Y_%m')}"
                                        os.makedirs(path, exist_ok=True)
                                        f_path = os.path.join(path, f"{now_ist().strftime('%H%M%S')}_{f.name}")
                                        img_bytes = f.getbuffer()
                                        with open(f_path, "wb") as fp:
                                            fp.write(img_bytes)
                                        # base64 — DB'ye gömülü, server bağımsız görüntüleme
                                        gorsel_b64 = base64.b64encode(img_bytes).decode("utf-8")
                                        # mime type
                                        ext = f.name.rsplit(".", 1)[-1].lower()
                                        mime = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","webp":"image/webp"}.get(ext,"image/jpeg")
                                        gorsel_data_uri = f"data:{mime};base64,{gorsel_b64}"

                                        new_e = {
                                            "ID": now_ist().strftime("%Y%m%d%H%M%S"),
                                            "Tarih": data_ai.get("tarih", now_ist().strftime("%Y-%m-%d")),
                                            "Saat": data_ai.get("saat", ""),
                                            "Yukleme_Zamani": yukleme_zamani,
                                            "Kullanıcı": user_name,
                                            "Rol": user_info.get("role", "user"),
                                            "Firma": data_ai.get("firma", "Bilinmiyor"),
                                            "Kategori": data_ai.get("kategori", "Diğer"),
                                            "Tutar": float(data_ai.get("toplam_tutar", 0)),
                                            "KDV": float(data_ai.get("kdv_tutari", 0)),
                                            "Odeme_Turu": st.session_state.get("_submitted_odeme_turu", odeme_turu_sec),
                                            "Kalemler": data_ai.get("kalemler", []),
                                            "Kisisel_Giderler": data_ai.get("kisisel_giderler", []),
                                            "Durum": "Onay Bekliyor",
                                            "Dosya_Yolu": f_path,
                                            "Gorsel_B64": gorsel_data_uri,
                                            "Risk_Skoru": int(data_ai.get("risk_skoru", 0)),
                                            "AI_Audit": clean_audit(str(data_ai.get("audit_ozeti", ""))),
                                            "AI_Anomali": data_ai.get("anomali", False),
                                            "AI_Anomali_Aciklama": data_ai.get("anomali_aciklamasi", ""),
                                            "Proje": proje,
                                            "Oncelik": oncelik,
                                            "Notlar": notlar
                                        }
                                    
                                        # ── Railway API'ye gönder ──────────────────
                                        with st.spinner("Railway'e gönderiliyor..."):
                                            ok = api_add_expense(new_e)
                                    
                                        if not ok:
                                            st.error("❌ Fiş API'ye gönderilemedi. Railway bağlantısını kontrol edin.")
                                        else:
                                            # Robot tarama animasyonu
                                            _stc.html('<html><body style="margin:0"><script>(function(){var pd=window.parent.document,r=pd.getElementById("SGNX"),bt=pd.getElementById("SGNXBT"),bub=pd.getElementById("SGNXBUB");if(r){r.classList.remove("sad");r.classList.add("joy");setTimeout(function(){r.classList.remove("joy");},3500);}if(bt)bt.textContent="Fis islendi!";if(bub){bub.classList.add("on");setTimeout(function(){bub.classList.remove("on");if(bt)bt.textContent="Merhaba!";},3800);}var cv=pd.createElement("canvas");cv.id="SGC";cv.style.cssText="position:fixed;top:0;left:0;width:100vw;height:100vh;pointer-events:none;z-index:2147483641;";cv.width=pd.documentElement.clientWidth;cv.height=pd.documentElement.clientHeight;pd.body.appendChild(cv);var ctx=cv.getContext("2d"),p=[],i;for(i=0;i<80;i++)p.push({x:Math.random()*cv.width,y:-20,vx:(Math.random()-.5)*6,vy:Math.random()*4+2,r:Math.random()*5+2,a:Math.random()*360,av:(Math.random()-.5)*8,c:["#00e896","#17a870","#f0a500","#fff"][Math.floor(Math.random()*4)]});var fr=0;function draw(){ctx.clearRect(0,0,cv.width,cv.height);p.forEach(function(q){q.x+=q.vx;q.y+=q.vy;q.a+=q.av;q.vy+=.12;var al=Math.max(0,1-fr/60);if(al>0){ctx.save();ctx.translate(q.x,q.y);ctx.rotate(q.a*Math.PI/180);ctx.fillStyle=q.c;ctx.globalAlpha=al;ctx.fillRect(-q.r/2,-q.r/2,q.r,q.r*1.6);ctx.restore();}});fr++;if(fr<80)requestAnimationFrame(draw);else{var old=pd.getElementById("SGC");if(old)old.remove();}}draw();})()</script></body></html>', height=1, scrolling=False)
                                            # Show result
                                            risk = int(data_ai.get("risk_skoru", 0))
                                            anomali = data_ai.get("anomali", False)
                                        
                                            if role != "admin":
                                                st.success("✅ Fiş sisteme eklendi! +50 XP · ⏳ Onay için yönetici bildirildi.")
                                            else:
                                                st.success("✅ Fiş başarıyla işlendi! +50 XP kazandın!")
                                    
                                        # İş kuralı uyarıları
                                        for uyari in uyarilar:
                                            st.warning(uyari)
                                    
                                        # Result card
                                        st.markdown(f"""
                                        <div class="ultra-card">
                                            <div style="display:flex; justify-content:space-between; align-items:flex-start;">
                                                <div>
                                                    <div style="font-size:1.3rem; font-weight:700; color:var(--text-primary);">
                                                        {data_ai.get('firma','?')}
                                                    </div>
                                                    <div style="font-size:2rem; font-family:'Bebas Neue'; color:var(--accent-blue); margin:4px 0;">
                                                        ₺{float(data_ai.get('toplam_tutar',0)):,.2f}
                                                    </div>
                                                    <div style="font-size:0.8rem; color:var(--text-secondary);">
                                                        {data_ai.get('kategori','?')} · {data_ai.get('odeme_turu','?')} · {data_ai.get('tarih','?')}
                                                    </div>
                                                </div>
                                                <div style="text-align:right;">
                                                    {get_risk_html(risk)}
                                                    {"<br><span style='color:#dc2626; font-size:0.75rem; margin-top:4px; display:block;'>⚠️ ANOMALİ TESPİT EDİLDİ</span>" if anomali else ""}
                                                </div>
                                            </div>
                                            <div class="ai-bubble" style="margin-top:12px;">
                                                <p style="margin:0; font-size:0.85rem; color:var(--text-secondary);">
                                                    {clean_audit(str(data_ai.get('audit_ozeti','Analiz tamamlandı.')))}
                                                </p>
                                            </div>
                                            {"<div class='anomaly-alert' style='margin-top:8px;'><strong style='color:#dc2626;'>🚨 " + str(data_ai.get('anomali_aciklamasi','')) + "</strong></div>" if anomali and data_ai.get('anomali_aciklamasi') else ""}
                                        </div>
                                        """, unsafe_allow_html=True)
                                    
                                        time.sleep(4.0)
                                        st.rerun()
                                else:
                                    st.error("AI fişi okuyamadı. Daha net bir fotoğraf deneyin.")
                    elif submitted and not f:
                        st.warning("Lütfen bir fiş fotoğrafı yükleyin.")
            
                st.markdown('</div>', unsafe_allow_html=True)
        
            with col_list:
                st.markdown("### 📋 Harcamalarım")
                my_exp = df[df['Kullanıcı'] == user_name] if not df.empty and 'Kullanıcı' in df.columns else pd.DataFrame()
            
                if not my_exp.empty:
                    for _, row in my_exp.sort_values('Tarih', ascending=False).head(8).iterrows():
                        risk = row.get('Risk_Skoru', 0)
                        st.markdown(f"""
                        <div class="ultra-card" style="padding:16px; margin:6px 0;">
                            <div style="display:flex; justify-content:space-between; align-items:flex-start;">
                                <div style="flex:1;">
                                    <div style="font-weight:600; font-size:0.9rem; color:var(--text-primary);">{row.get('Firma','?')}</div>
                                    <div style="font-size:0.75rem; color:var(--text-muted); margin:2px 0;">{row.get('Tarih','?')} · {row.get('Proje','?')}</div>
                                    <div style="font-size:0.7rem; color:var(--text-secondary); margin-top:4px; font-style:italic;">
                                        {clean_audit(row.get('AI_Audit',''))[:80]}{'...' if len(clean_audit(row.get('AI_Audit',''))) > 80 else ''}
                                    </div>
                                </div>
                                <div style="text-align:right; margin-left:12px;">
                                    <div style="font-family:'Bebas Neue'; font-size:1.3rem; color:var(--accent-blue);">₺{float(row.get('Tutar',0)):,.0f}</div>
                                    {get_status_html(row.get('Durum','?'))}
                                    <div style="margin-top:4px;">{get_risk_html(risk)}</div>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class="empty-state">
                        <div class="empty-icon">📭</div>
                        <div>Henüz harcama kaydın yok.</div>
                    </div>
                    """, unsafe_allow_html=True)

        with tab_manuel:
            st.markdown("""
            <div style='padding:12px 0 4px;'>
              <div style='font-size:0.85rem;color:#7a96a4;margin-bottom:16px;'>
                Fiş olmadan manuel harcama girişi. Onaya gönderilir.
              </div>
            </div>""", unsafe_allow_html=True)
            with st.form("manuel_giris", clear_on_submit=True):
                mc1, mc2 = st.columns(2)
                with mc1:
                    m_firma    = st.text_input("Firma / Açıklama", placeholder="Örn: Akaryakıt, Market...")
                    m_tutar    = st.number_input("Tutar (₺)", min_value=0.01, step=0.01, format="%.2f")
                    m_kdv      = st.selectbox("KDV Oranı", ["0","1","8","10","20"], index=4)
                with mc2:
                    m_kategori = st.selectbox("Kategori", [
                        "Akaryakıt","Market / Gıda","Kırtasiye","Elektrik / Donanım",
                        "Ulaşım","Konaklama","Yemek","Diğer"])
                    m_proje    = st.selectbox("Proje", ["Maden Sahası","Aktif Karbon","Enerji Hatları","Genel Merkez"])
                    m_odeme = st.radio("Bu harcama nasil odendi?", ["Harcirahtan Dus", "Sirket Kredi Karti"], horizontal=True)
                    m_odeme_turu = "harcirah" if "Harcirahtan" in m_odeme else "sirket_karti"
                m_tarih = st.date_input("Tarih", value=now_ist().date())
                m_notlar = st.text_area("Açıklama / Not", height=70, placeholder="Harcama hakkında kısa not...")
                m_submit = st.form_submit_button("✅ Kaydet ve Onaya Gönder", use_container_width=True)
                if m_submit:
                    if not m_firma.strip():
                        st.error("⚠️ Firma / Açıklama boş bırakılamaz.")
                    elif m_tutar <= 0:
                        st.error("⚠️ Tutar sıfırdan büyük olmalı.")
                    else:
                        kdv_oran  = int(m_kdv) / 100
                        kdv_tutar = round(m_tutar * kdv_oran / (1 + kdv_oran), 2) if kdv_oran > 0 else 0.0
                        net_tutar = round(m_tutar - kdv_tutar, 2)
                        expense = {
                            "ID":            now_ist().strftime("%Y%m%d%H%M%S") + "_M",
                            "Tarih":         str(m_tarih),
                            "Firma":         m_firma.strip(),
                            "Kategori":      m_kategori,
                            "Tutar":         float(m_tutar),
                            "KDV_Orani":     int(m_kdv),
                            "KDV_Tutari":    kdv_tutar,
                            "Net_Tutar":     net_tutar,
                            "Odeme_Turu": m_odeme_turu,
                            "Proje":         m_proje,
                            "Notlar":        m_notlar.strip(),
                            "Kullanıcı":     user_name,
                            "Durum":         "Onay Bekliyor",
                            "Risk_Skoru":    10,
                            "Manuel":        True,
                            "YuklemeZamani": now_ist().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        with st.spinner("Kaydediliyor..."):
                            if api_add_expense(expense):
                                st.success(f"✅ **{m_firma}** — ₺{m_tutar:,.2f} başarıyla kaydedildi. Onay bekleniyor.")
                                st.cache_data.clear()
                            else:
                                st.error("❌ API hatası, lütfen tekrar deneyin.")

    # ══════════════════════════════════════════════════════════
    # SAYFA: ONAY MERKEZİ (sadece admin)
    # ══════════════════════════════════════════════════════════
    elif selected == "⚖️ Onay Merkezi":
        st.markdown('<div class="page-header"><div class="page-title">ONAY MERKEZİ</div></div>', unsafe_allow_html=True)

        if role != "admin":
            st.warning("Bu sayfaya erişim yetkiniz bulunmamaktadır.")
        else:
            pending = df_full[df_full["Durum"] == "Onay Bekliyor"] if not df_full.empty and 'Durum' in df_full.columns else pd.DataFrame()
            sahte_s = df_full[df_full["Durum"] == "Sahte Şüphesi"] if not df_full.empty and 'Durum' in df_full.columns else pd.DataFrame()

            col_om1, col_om2 = st.columns(2)
            col_om1.metric("⏳ Onay Bekleyen", len(pending), f"₺{pending['Tutar'].sum():,.0f}" if not pending.empty else "₺0")
            col_om2.metric("🚨 Sahte Şüphesi", len(sahte_s))
            st.markdown("<br>", unsafe_allow_html=True)

            if pending.empty and sahte_s.empty:
                st.markdown("""
                <div class="ultra-card" style="text-align:center; padding:48px;">
                    <div style="font-size:3rem;">✅</div>
                    <div style="font-size:1.3rem; font-weight:700; color:#007850; margin:12px 0;">Onay Bekleyen İşlem Yok</div>
                    <div style="color:var(--text-muted);">Tüm fişler işlenmiş.</div>
                </div>
                """, unsafe_allow_html=True)
            else:
                tum_bekleyen = pd.concat([pending, sahte_s]).drop_duplicates("ID") if not pending.empty or not sahte_s.empty else pd.DataFrame()
                for idx_r, row in tum_bekleyen.sort_values("Tutar", ascending=False).iterrows():
                    risk = row.get('Risk_Skoru', 0)
                    kaynak_icon = "📱" if row.get("Kaynak") == "WhatsApp" else "💻"
                    # Ödeme türü ve harcırah bilgisi
                    _row_odeme = str(row.get('Odeme_Turu', row.get('OdemeTipi',''))).lower().strip()
                    _row_kullanici = row.get('Kullanıcı','?')
                    _row_tutar = float(row.get('Tutar', 0))
                    _is_harcirah = _row_odeme in ("harcirah","harcırah","harcirahtan dus","harcırahtan düş","harcırahtan düş (nakit / kişisel kart)","nakit","kisisel")
                    if _is_harcirah:
                        _mevcut_bakiye = get_user_wallet_balance(_row_kullanici, data_store)
                        _kalan_bakiye  = _mevcut_bakiye - _row_tutar
                        _kalan_renk    = "#dc2626" if _kalan_bakiye < 0 else "#007850"
                        _harcirah_html = (
                            f'<div style="margin-top:8px; padding:8px 12px; border-radius:8px; '
                            f'background:rgba(245,158,11,0.08); border:1px solid rgba(245,158,11,0.25); font-size:0.8rem;">'
                            f'💵 <strong>{_row_kullanici}</strong> bu ödemeyi <strong>kendi nakdi / şahsi kartıyla</strong> yapmıştır. '
                            f'Mevcut harcırah bakiyesi: <strong>₺{_mevcut_bakiye:,.0f}</strong> &nbsp;·&nbsp; '
                            f'Onaylanırsa kalan bakiye: <strong style="color:{_kalan_renk}">₺{_kalan_bakiye:,.0f}</strong>'
                            f'</div>'
                        )
                    else:
                        _harcirah_html = ""
                    with st.expander(f"{'🔴' if risk > 70 else '🟡'} {kaynak_icon} {row.get('Kullanıcı','?')} · {row.get('Firma','?')} · ₺{float(row.get('Tutar',0)):,.0f} · {row.get('Proje','?')}"):
                        ca, cb = st.columns([2, 1])
                        with ca:
                            st.markdown(f"""
                            <div class="ultra-card">
                                <div style="display:flex; justify-content:space-between; margin-bottom:12px;">
                                    <div>
                                        <div style="font-size:1.1rem; font-weight:700;">{row.get('Firma','?')}</div>
                                        <div style="font-size:0.8rem; color:var(--text-muted);">{row.get('Tarih','?')} · {row.get('Kategori','?')} · {kaynak_icon} {row.get('Kaynak','Dashboard')}</div>
                                    </div>
                                    <div style="text-align:right;">
                                        <div style="font-family:'Bebas Neue'; font-size:2rem; color:var(--accent-blue);">₺{float(row.get('Tutar',0)):,.0f}</div>
                                        {get_risk_html(risk)}
                                        {get_status_html(row.get('Durum','?'))}
                                    </div>
                                </div>
                                <div class="ai-bubble">🤖 {clean_audit(row.get('AI_Audit',''))}</div>
                                {"<div class='anomaly-alert' style='margin-top:8px;'>⚠️ " + strip_html(row.get('AI_Anomali_Aciklama','')) + "</div>" if row.get('AI_Anomali') else ""}
                                <div style="margin-top:8px; font-size:0.75rem; color:var(--text-muted);">
                                    📁 Proje: <strong>{strip_html(row.get('Proje','?'))}</strong> &nbsp;·&nbsp;
                                    ⚡ Öncelik: <strong>{strip_html(row.get('Oncelik','Normal'))}</strong> &nbsp;·&nbsp;
                                    💳 Ödeme: <strong>{odeme_label(row.get('Odeme_Turu', row.get('OdemeTipi','—')))}</strong> &nbsp;·&nbsp;
                                    📱 Kaynak: <strong>{strip_html(row.get('Kaynak','Dashboard'))}</strong>
                                </div>
                                <div style="margin-top:6px; padding:8px 14px; border-radius:8px; font-size:0.82rem; font-weight:700;
                                    {'background:rgba(217,119,6,0.1); border:1px solid rgba(217,119,6,0.3); color:#d97706;' if str(row.get('Odeme_Turu', row.get('OdemeTipi',''))).lower().strip() in ('harcirah','harcırah','harcirahtan dus','harcırahtan düş','harcırahtan düş (nakit / kişisel kart)','nakit','kisisel') else 'background:rgba(8,145,178,0.1); border:1px solid rgba(8,145,178,0.3); color:#0891b2;'}">
                                    {'💵 HARCIRAHTAN DÜŞÜLECEK — Personel şahsi ödeme yapmıştır' if str(row.get('Odeme_Turu', row.get('OdemeTipi',''))).lower().strip() in ('harcirah','harcırah','harcirahtan dus','harcırahtan düş','harcırahtan düş (nakit / kişisel kart)','nakit','kisisel') else '🏦 ŞİRKET KREDİ KARTI — Genel merkezden düşülecek'}
                                </div>
                                {_harcirah_html}
                                {f"<div style='margin-top:6px; font-size:0.75rem;'>📝 {strip_html(row.get('Notlar',''))}</div>" if row.get('Notlar') else ""}
                            </div>
                            """, unsafe_allow_html=True)

                            btn1, btn2 = st.columns(2)
                            if btn1.button("✅ Onayla", key=f"omcent_on_{row['ID']}", use_container_width=True):
                                _stc.html('''<script>
(function(){
  var pd=window.parent.document;
  var cv=pd.createElement('canvas');
  cv.style.cssText='position:fixed;inset:0;width:100%;height:100%;pointer-events:none;z-index:2147483641;';
  cv.width=window.parent.innerWidth;cv.height=window.parent.innerHeight;
  pd.body.appendChild(cv);
  var ctx=cv.getContext('2d');
  var p=[];for(var i=0;i<140;i++)p.push({x:Math.random()*cv.width,y:-20+Math.random()*-80,vx:(Math.random()-.5)*8,vy:Math.random()*5+3,r:Math.random()*7+3,angle:Math.random()*360,av:(Math.random()-.5)*10,color:['#00e896','#17a870','#2F3C6E','#f0a500','#ffffff','#3d4e8a'][Math.floor(Math.random()*6)]});
  var fr=0;function draw(){ctx.clearRect(0,0,cv.width,cv.height);p.forEach(function(q){q.x+=q.vx;q.y+=q.vy;q.angle+=q.av;q.vy+=.1;ctx.save();ctx.translate(q.x,q.y);ctx.rotate(q.angle*Math.PI/180);ctx.fillStyle=q.color;ctx.globalAlpha=Math.max(0,1-fr/100);ctx.fillRect(-q.r/2,-q.r/2,q.r,q.r*1.7);ctx.restore();});fr++;if(fr<180)requestAnimationFrame(draw);else cv.remove();}
  draw();
  var r=pd.getElementById('SGNX');
  if(r){r.classList.remove('sad');r.classList.add('joy');setTimeout(function(){r.classList.remove('joy');},4500);}
  var bt=pd.getElementById('SGNXBT'),bub=pd.getElementById('SGNXBUB');
  if(bt)bt.textContent='🎉 ONAYLANDI! Harika!';
  if(bub){bub.classList.add('on');setTimeout(function(){bub.classList.remove('on');if(bt)bt.textContent='Merhaba!';},4800);}
})();
</script>''', height=1, scrolling=False)
                                with st.spinner("Onaylanıyor..."):
                                    if api_approve(str(row['ID']), "approve", user_name):
                                        st.success("✅ Onaylandı!")
                                        time.sleep(4.8)
                                        st.rerun()
                                    else:
                                        st.error("API hatası, tekrar deneyin")

                            if btn2.button("❌ Reddet", key=f"omcent_ret_{row['ID']}", use_container_width=True):
                                _stc.html('''<script>
(function(){
  var pd=window.parent.document;
  var ov=pd.createElement('div');
  ov.style.cssText='position:fixed;inset:0;background:rgba(220,38,38,.1);z-index:2147483640;pointer-events:none;';
  pd.body.appendChild(ov);setTimeout(function(){ov.remove();},600);
  var r=pd.getElementById('SGNX');
  if(r){r.classList.remove('joy');r.classList.add('sad');setTimeout(function(){r.classList.remove('sad');},4500);}
  var bt=pd.getElementById('SGNXBT'),bub=pd.getElementById('SGNXBUB');
  if(bt)bt.textContent='😢 Fiş reddedildi...';
  if(bub){bub.classList.add('on');setTimeout(function(){bub.classList.remove('on');if(bt)bt.textContent='Merhaba!';},4800);}
})();
</script>''', height=1, scrolling=False)
                                with st.spinner("Reddediliyor..."):
                                    if api_approve(str(row['ID']), "reject", user_name):
                                        st.warning("❌ Reddedildi!")
                                        time.sleep(4.8)
                                        st.rerun()
                                    else:
                                        st.error("API hatası, tekrar deneyin")

                        with cb:
                            dosya   = row.get('Dosya_Yolu', '')
                            b64_uri = row.get('Gorsel_B64', '')
                            if dosya and os.path.exists(str(dosya)):
                                st.image(str(dosya), caption="Orijinal Fiş", use_container_width=True)
                            elif b64_uri:
                                try:
                                    header, b64_data = b64_uri.split(",", 1) if "," in b64_uri else ("", b64_uri)
                                    img_bytes = base64.b64decode(b64_data)
                                    st.image(img_bytes, caption="📱 WhatsApp'tan gönderildi", use_container_width=True)
                                except Exception:
                                    st.markdown("""<div style="height:150px;background:var(--bg-secondary);border-radius:10px;
                                        display:flex;align-items:center;justify-content:center;color:var(--text-muted);">📷 Görsel Yüklenemedi</div>""",
                                        unsafe_allow_html=True)
                            else:
                                st.markdown("""<div style="height:150px;background:var(--bg-secondary);border-radius:10px;
                                    display:flex;align-items:center;justify-content:center;color:var(--text-muted);">📷 Görsel Yok</div>""",
                                    unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: FİNANS & KASA
    # ══════════════════════════════════════════════════════════
    elif selected == "💰 Finans & Kasa":
        st.markdown('<div class="page-header"><div class="page-title">FİNANS MERKEZİ</div></div>', unsafe_allow_html=True)
        
        # Budget overview
        st.markdown("### 📊 Proje Bütçe Durumu")
        budgets = data_store.get("budgets", {})
        
        b_cols = st.columns(4)
        for i, (proj, bdata) in enumerate(budgets.items()):
            limit = bdata.get("limit", 0)
            spent = bdata.get("spent", 0)
            pct = min(spent / limit * 100, 100) if limit > 0 else 0
            color = "#007850" if pct < 60 else ("#d97706" if pct < 85 else "#dc2626")
            
            with b_cols[i % 4]:
                st.markdown(f"""
                <div class="metric-card">
                    <div style="font-size:0.7rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:1px;">{proj}</div>
                    <div style="font-size:1.8rem; font-family:'Bebas Neue'; color:{color}; margin:8px 0;">{pct:.0f}%</div>
                    <div class="budget-track">
                        <div class="budget-fill" style="width:{pct:.0f}%; background:{color};"></div>
                    </div>
                    <div style="font-size:0.7rem; color:var(--text-muted); margin-top:6px;">
                        ₺{spent:,.0f} / ₺{limit:,.0f}
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        col_c, col_d = st.columns([1.2, 1])
        
        with col_c:
            if role == "admin":
                st.markdown("### 💳 Personel Kasa Durumları")
                
                # wallets dict yerine USERS listesini kullan
                # → Railway'deki "Serkan" / "Serkan Güzdemir" duplikasyonu önlenir
                for _ukey, _uinfo in USERS.items():
                    person = _uinfo["name"]
                    bal = get_user_wallet_balance(person, data_store)
                    person_limit = _uinfo.get("monthly_limit", 15000)
                    bal_pct = min(abs(bal) / person_limit * 100, 100) if person_limit > 0 else 0
                    avatar = _uinfo.get("avatar", "👤")
                    
                    st.markdown(f"""
                    <div class="ultra-card" style="padding:16px; margin:8px 0;">
                        <div style="display:flex; align-items:center; justify-content:space-between;">
                            <div style="display:flex; align-items:center; gap:10px;">
                                <span style="font-size:1.5rem;">{avatar}</span>
                                <div>
                                    <div style="font-weight:600;">{person}</div>
                                    <div style="font-size:0.7rem; color:var(--text-muted);">Kasa Bakiyesi</div>
                                </div>
                            </div>
                            <div style="font-family:'Bebas Neue'; font-size:1.8rem; color:{'#dc2626' if bal < 0 else 'var(--accent-blue)'};">
                                {'−' if bal < 0 else ''}₺{abs(bal):,.0f}{'  🔴 Borç' if bal < 0 else ''}
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("### 💸 Harcırah Transfer")
                _transfer_persons = {info["name"]: key for key, info in USERS.items()}
                with st.form("harcirah_form"):
                    col_t1, col_t2 = st.columns(2)
                    with col_t1:
                        target = st.selectbox("Personel", list(_transfer_persons.keys()))
                    with col_t2:
                        amt = st.number_input("Tutar (₺)", min_value=0, step=500, value=1000)
                    aciklama = st.text_input("Açıklama", value="Aylık harcırah")
                    
                    if st.form_submit_button("⚡ Transfer Et", use_container_width=True):
                        with st.spinner("Transfer yapılıyor..."):
                            if api_transfer(target, amt, aciklama, user_name):
                                st.success(f"✅ {target}'e ₺{amt:,.0f} transfer edildi!")
                                time.sleep(0.5)
                                st.rerun()
                            else:
                                st.error("Transfer başarısız, tekrar deneyin")
            else:
                # Ledger-tabanlı kasa hesabı - avans - onaylı harcırah harcamaları
                _w2 = data_store.get("wallets", {})
                my_bal = get_user_wallet_balance(user_name, data_store)
                st.markdown(f"""
                <div class="metric-card" style="margin-bottom:16px;">
                    <div style="font-size:1rem; color:var(--text-secondary);">Mevcut Bakiyeniz</div>
                    <div style="font-family:'Bebas Neue'; font-size:3.5rem; color:var(--accent-green); 
                                text-shadow:0 0 30px rgba(0,255,136,0.4);">
                        {'−' if my_bal < 0 else ''}₺{abs(my_bal):,.0f}{'  🔴 Şirkete borcunuz var' if my_bal < 0 else ''}
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        with col_d:
            st.markdown("### 📜 Son Hareketler")
            ledger = data_store.get("ledger", [])
            if ledger:
                for entry in reversed(ledger[-8:]):
                    st.markdown(f"""
                    <div style="padding:10px 0; border-bottom:1px solid var(--border);">
                        <div style="display:flex; justify-content:space-between;">
                            <span style="font-size:0.8rem; color:var(--text-secondary);">{entry.get('İşlem','?')}</span>
                            <span style="font-family:'JetBrains Mono'; color:var(--accent-blue); font-size:0.85rem;">
                                +₺{entry.get('Miktar',0):,.0f}
                            </span>
                        </div>
                        <div style="font-size:0.7rem; color:var(--text-muted);">{entry.get('Tarih','?')} → {entry.get('Hedef','?')}</div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("Henüz hareket yok.")
        
        # Admin approval center
        if role == "admin":
            st.markdown("---")
            st.markdown("### ⚖️ Onay Merkezi")
            
            # Admin tüm bekleyen fişleri görür
            pending = df_full[df_full["Durum"] == "Onay Bekliyor"] if not df_full.empty and 'Durum' in df_full.columns else pd.DataFrame()
            
            if pending.empty:
                st.success("✅ Onay bekleyen işlem yok. Muhteşem!")
            else:
                for idx, row in pending.iterrows():
                    risk = row.get('Risk_Skoru', 0)
                    risk_color = "#007850" if risk < 30 else ("#d97706" if risk < 70 else "#dc2626")
                    # Ödeme türü ve harcırah bilgisi
                    _row_odeme2 = str(row.get('Odeme_Turu','')).lower().strip()
                    _row_kullanici2 = row.get('Kullanıcı','?')
                    _row_tutar2 = float(row.get('Tutar', 0))
                    _is_harcirah2 = _row_odeme2 in ("harcirah","harcırah","harcirahtan dus","harcırahtan düş","harcırahtan düş (nakit / kişisel kart)","nakit","kisisel")
                    if _is_harcirah2:
                        _mevcut_bakiye2 = get_user_wallet_balance(_row_kullanici2, data_store)
                        _kalan_bakiye2  = _mevcut_bakiye2 - _row_tutar2
                        _kalan_renk2    = "#dc2626" if _kalan_bakiye2 < 0 else "#007850"
                        _harcirah_html2 = (
                            f'<div style="margin-top:8px; padding:8px 12px; border-radius:8px; '
                            f'background:rgba(245,158,11,0.08); border:1px solid rgba(245,158,11,0.25); font-size:0.8rem;">'
                            f'💵 <strong>{_row_kullanici2}</strong> bu ödemeyi <strong>kendi nakdi / şahsi kartıyla</strong> yapmıştır. '
                            f'Mevcut harcırah bakiyesi: <strong>₺{_mevcut_bakiye2:,.0f}</strong> &nbsp;·&nbsp; '
                            f'Onaylanırsa kalan bakiye: <strong style="color:{_kalan_renk2}">₺{_kalan_bakiye2:,.0f}</strong>'
                            f'</div>'
                        )
                    else:
                        _harcirah_html2 = ""
                    
                    with st.expander(f"{'🔴' if risk > 70 else '🟡'} {row['Kullanıcı']} · {row['Firma']} · ₺{row['Tutar']:,.0f} · {row['Proje']}"):
                        ca, cb = st.columns([2, 1])
                        
                        with ca:
                            st.markdown(f"""
                            <div class="ultra-card">
                                <div style="display:flex; justify-content:space-between; margin-bottom:12px;">
                                    <div>
                                        <div style="font-size:1.1rem; font-weight:700;">{row.get('Firma','?')}</div>
                                        <div style="font-size:0.8rem; color:var(--text-muted);">{row.get('Tarih','?')} · {row.get('Kategori','?')}</div>
                                    </div>
                                    <div style="text-align:right;">
                                        <div style="font-family:'Bebas Neue'; font-size:2rem; color:var(--accent-blue);">₺{float(row.get('Tutar',0)):,.0f}</div>
                                        {get_risk_html(risk)}
                                    </div>
                                </div>
                                <div class="ai-bubble">
                                    🤖 {clean_audit(row.get('AI_Audit',''))}
                                </div>
                                {"<div class='anomaly-alert' style='margin-top:8px;'>⚠️ AI Anomali Tespiti: " + strip_html(row.get('AI_Anomali_Aciklama','')) + "</div>" if row.get('AI_Anomali') else ""}
                                <div style="margin-top:8px; font-size:0.75rem; color:var(--text-muted);">
                                    📁 Proje: <strong>{strip_html(row.get('Proje','?'))}</strong> &nbsp;·&nbsp;
                                    ⚡ Öncelik: <strong>{strip_html(row.get('Oncelik','Normal'))}</strong> &nbsp;·&nbsp;
                                    💳 Ödeme: <strong>{odeme_label(row.get('Odeme_Turu','—'))}</strong>
                                </div>
                                {_harcirah_html2}
                                {f"<div style='margin-top:6px; font-size:0.75rem; color:var(--text-secondary);'>📝 Not: {strip_html(row.get('Notlar',''))}</div>" if row.get('Notlar') else ""}
                            </div>
                            """, unsafe_allow_html=True)
                            
                            btn1, btn2 = st.columns(2)
                            if btn1.button("✅ Onayla", key=f"on_{row['ID']}", use_container_width=True):
                                _stc.html('''<script>
(function(){
  var pd=window.parent.document;
  var cv=pd.createElement('canvas');
  cv.style.cssText='position:fixed;inset:0;width:100%;height:100%;pointer-events:none;z-index:2147483641;';
  cv.width=window.parent.innerWidth;cv.height=window.parent.innerHeight;
  pd.body.appendChild(cv);
  var ctx=cv.getContext('2d');
  var p=[];for(var i=0;i<140;i++)p.push({x:Math.random()*cv.width,y:-20+Math.random()*-80,vx:(Math.random()-.5)*8,vy:Math.random()*5+3,r:Math.random()*7+3,angle:Math.random()*360,av:(Math.random()-.5)*10,color:['#00e896','#17a870','#2F3C6E','#f0a500','#ffffff','#3d4e8a'][Math.floor(Math.random()*6)]});
  var fr=0;function draw(){ctx.clearRect(0,0,cv.width,cv.height);p.forEach(function(q){q.x+=q.vx;q.y+=q.vy;q.angle+=q.av;q.vy+=.1;ctx.save();ctx.translate(q.x,q.y);ctx.rotate(q.angle*Math.PI/180);ctx.fillStyle=q.color;ctx.globalAlpha=Math.max(0,1-fr/100);ctx.fillRect(-q.r/2,-q.r/2,q.r,q.r*1.7);ctx.restore();});fr++;if(fr<180)requestAnimationFrame(draw);else cv.remove();}
  draw();
  var r=pd.getElementById('SGNX');
  if(r){r.classList.remove('sad');r.classList.add('joy');setTimeout(function(){r.classList.remove('joy');},4500);}
  var bt=pd.getElementById('SGNXBT'),bub=pd.getElementById('SGNXBUB');
  if(bt)bt.textContent='🎉 ONAYLANDI! Harika!';
  if(bub){bub.classList.add('on');setTimeout(function(){bub.classList.remove('on');if(bt)bt.textContent='Merhaba!';},4800);}
})();
</script>''', height=1, scrolling=False)
                                with st.spinner("Onaylanıyor..."):
                                    if api_approve(str(row['ID']), "approve", user_name):
                                        st.success("✅ Onaylandı!")
                                        time.sleep(4.8)
                                        st.rerun()
                            
                            if btn2.button("❌ Reddet", key=f"ret_{row['ID']}", use_container_width=True):
                                _stc.html('''<script>
(function(){
  var pd=window.parent.document;
  var ov=pd.createElement('div');
  ov.style.cssText='position:fixed;inset:0;background:rgba(220,38,38,.1);z-index:2147483640;pointer-events:none;';
  pd.body.appendChild(ov);setTimeout(function(){ov.remove();},600);
  var r=pd.getElementById('SGNX');
  if(r){r.classList.remove('joy');r.classList.add('sad');setTimeout(function(){r.classList.remove('sad');},4500);}
  var bt=pd.getElementById('SGNXBT'),bub=pd.getElementById('SGNXBUB');
  if(bt)bt.textContent='😢 Fiş reddedildi...';
  if(bub){bub.classList.add('on');setTimeout(function(){bub.classList.remove('on');if(bt)bt.textContent='Merhaba!';},4800);}
})();
</script>''', height=1, scrolling=False)
                                with st.spinner("Reddediliyor..."):
                                    if api_approve(str(row['ID']), "reject", user_name):
                                        st.warning("❌ Reddedildi!")
                                        time.sleep(4.8)
                                        st.rerun()
                        
                        with cb:
                            # Önce lokal dosyayı dene, sonra base64 (WhatsApp fişleri için)
                            dosya   = row.get('Dosya_Yolu', '')
                            b64_uri = row.get('Gorsel_B64', '')
                            if dosya and os.path.exists(dosya):
                                st.image(dosya, caption="Orijinal Fiş", use_container_width=True)
                            elif b64_uri:
                                try:
                                    header, b64_data = b64_uri.split(",", 1) if "," in b64_uri else ("", b64_uri)
                                    img_bytes = base64.b64decode(b64_data)
                                    st.image(img_bytes, caption="📱 WhatsApp'tan gönderilen fiş", use_container_width=True)
                                except Exception:
                                    st.markdown("""
                                    <div style="height:150px; background:var(--bg-secondary); border-radius:10px; 
                                                display:flex; align-items:center; justify-content:center; color:var(--text-muted);">
                                        📷 Görsel Yüklenemedi
                                    </div>
                                    """, unsafe_allow_html=True)
                            else:
                                st.markdown("""
                                <div style="height:150px; background:var(--bg-secondary); border-radius:10px; 
                                            display:flex; align-items:center; justify-content:center; color:var(--text-muted);">
                                    📷 Görsel Yok
                                </div>
                                """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: ANOMALİ DEDEKTÖRÜ
    # ══════════════════════════════════════════════════════════
    elif selected == "🔬 Anomali Dedektörü":
        st.markdown('<div class="page-header"><div class="page-title">ANOMALİ DEDEKTÖRÜ</div></div>', unsafe_allow_html=True)
        
        st.markdown("""
        <div class="ultra-card">
            <p style="color:var(--text-secondary); font-size:0.9rem; margin:0;">
                🔬 Yapay Zeka destekli anomali tespiti motoru — mükerrer fişler, hafta sonu harcamaları, 
                istatistiksel aykırı değerler ve kritik risk skorlarını otomatik olarak tarar.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if not df_full.empty:
            anomalies = detect_anomalies(df_full if role == "admin" else df, model)
            
            if anomalies:
                st.markdown(f"### 🚨 {len(anomalies)} Anomali Tespit Edildi")
                
                for a in anomalies:
                    sev_color = {"high": "#dc2626", "medium": "#d97706", "low": "#007850"}.get(a['severity'], "#4a5568")
                    sev_icon = {"high": "🔴", "medium": "🟡", "low": "🔵"}.get(a['severity'], "⚪")
                    
                    st.markdown(f"""
                    <div class="ultra-card" style="border-left:4px solid {sev_color};">
                        <div style="display:flex; align-items:center; gap:12px;">
                            <span style="font-size:1.5rem;">{sev_icon}</span>
                            <div>
                                <div style="font-weight:600; color:var(--text-primary);">{a['message']}</div>
                                <div style="font-size:0.75rem; color:{sev_color}; text-transform:uppercase; margin-top:2px;">
                                    {a['severity'].upper()} SEVİYE · {a['count']} işlem etkilendi
                                </div>
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="ultra-card" style="border-left:4px solid #007850; text-align:center; padding:40px;">
                    <div style="font-size:3rem;">✅</div>
                    <div style="font-size:1.2rem; font-weight:600; color:#007850; margin:8px 0;">Anomali Tespit Edilmedi</div>
                    <div style="color:var(--text-muted);">Tüm işlemler normal görünüyor.</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Deep analysis
            st.markdown("### 🤖 AI Derin Analiz")
            
            if st.button("🔍 Kapsamlı AI Denetimi Başlat"):
                with st.spinner("Gemini AI tüm verileri tarıyor..."):
                    # Token taşmasını önlemek için özet istatistikler gönder
                    _df_s = df.copy()
                    _tutar_col = next((c for c in ["Tutar","tutar","Amount","amount"] if c in _df_s.columns), None)
                    _kat_col   = next((c for c in ["Kategori","kategori","Category"] if c in _df_s.columns), None)
                    _kul_col   = next((c for c in ["Kullanici","kullanici","Kullanıcı","User","user_name"] if c in _df_s.columns), None)
                    _tip_col   = next((c for c in ["Odeme_Tipi","odeme_tipi","Ödeme Tipi"] if c in _df_s.columns), None)
                    _stat_col  = next((c for c in ["Durum","durum","Status"] if c in _df_s.columns), None)

                    ozet_satirlar = []
                    ozet_satirlar.append(f"Toplam fiş sayısı: {len(_df_s)}")
                    if _tutar_col:
                        ozet_satirlar.append(f"Toplam tutar: {_df_s[_tutar_col].sum():,.0f} ₺")
                        ozet_satirlar.append(f"Ortalama tutar: {_df_s[_tutar_col].mean():,.0f} ₺")
                        ozet_satirlar.append(f"Maks tutar: {_df_s[_tutar_col].max():,.0f} ₺")
                        # En yüksek 10 işlem
                        top10 = _df_s.nlargest(10, _tutar_col)[[c for c in [_tutar_col, _kat_col, _kul_col, _tip_col, _stat_col] if c]].to_dict(orient="records")
                        ozet_satirlar.append(f"En yüksek 10 işlem: {top10}")
                    if _kat_col and _tutar_col:
                        kat_ozet = _df_s.groupby(_kat_col)[_tutar_col].agg(["sum","count","mean"]).round(0).to_dict()
                        ozet_satirlar.append(f"Kategori bazlı özet: {kat_ozet}")
                    if _kul_col and _tutar_col:
                        kul_ozet = _df_s.groupby(_kul_col)[_tutar_col].agg(["sum","count","mean"]).round(0).to_dict()
                        ozet_satirlar.append(f"Kullanıcı bazlı özet: {kul_ozet}")
                    if _stat_col:
                        ozet_satirlar.append(f"Durum dağılımı: {_df_s[_stat_col].value_counts().to_dict()}")

                    veri_ozeti = "\n".join(ozet_satirlar)

                    prompt = f"""Sen bir adli mali denetçisin. Aşağıdaki harcama istatistiklerini incele ve şüpheli durumları rapor et:

{veri_ozeti}

Türkçe olarak şunu belirt:
1. En riskli 3 işlem veya pattern ve nedeni
2. Tespit ettiğin olağandışı durumlar
3. Önerilen aksiyonlar
Kısa ve net ol (max 300 kelime)."""
                    
                    try:
                        response = model.generate_content(prompt)
                        st.markdown(f"""
                        <div class="ai-bubble" style="margin-top:12px;">
                            <p style="margin:0; line-height:1.8; white-space:pre-wrap; font-size:0.9rem;">
                                {response.text}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # AI insight sadece session'da tut (Railway'e yazmıyoruz)
                        if "ai_insights" not in st.session_state:
                            st.session_state.ai_insights = []
                        st.session_state.ai_insights.append({
                            "date": now_ist().strftime("%Y-%m-%d %H:%M"),
                            "type": "anomaly_scan",
                            "content": response.text
                        })
                    except Exception as e:
                        st.error(f"AI yanıt veremedi: {e}")
            
            # Statistical visualization
            st.markdown("### 📊 İstatistiksel Görselleştirme")
            
            col_s1, col_s2 = st.columns(2)
            
            with col_s1:
                if 'Tutar' in df.columns:
                    fig = px.box(df, x='Proje', y='Tutar', color='Proje',
                                title='Proje Bazlı Tutar Dağılımı (Box Plot)',
                                color_discrete_sequence=['#007850','#283c64','#007850','#ea6c1e'])
                    fig.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568'), showlegend=False
                    )
                    st.plotly_chart(fig, use_container_width=True)
            
            with col_s2:
                if 'Risk_Skoru' in df.columns:
                    fig2 = px.histogram(df, x='Risk_Skoru', nbins=10,
                                       title='Risk Skoru Dağılımı',
                                       color_discrete_sequence=['#dc2626'])
                    fig2.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568')
                    )
                    st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Anomali analizi için veri gerekli. Önce fiş taraması yapın.")

    # ══════════════════════════════════════════════════════════
    # SAYFA: ANALİTİK MERKEZİ
    # ══════════════════════════════════════════════════════════
    elif selected == "📊 Analitik Merkezi":
        st.markdown('<div class="page-header"><div class="page-title">ANALİTİK MERKEZİ</div></div>', unsafe_allow_html=True)
        
        if not df.empty:
            tab1, tab2, tab3 = st.tabs(["📈 Trend Analizi", "🗺️ Isı Haritası", "🔮 Tahmin"])
            
            with tab1:
                df_temp = df.copy()
                df_temp['dt'] = pd.to_datetime(df_temp['Tarih'], errors='coerce')
                
                # Monthly trend
                monthly = df_temp.groupby(df_temp['dt'].dt.strftime('%Y-%m')).agg(
                    Toplam=('Tutar', 'sum'),
                    Adet=('Tutar', 'count'),
                    OrtRisk=('Risk_Skoru', 'mean')
                ).reset_index()
                monthly.columns = ['Ay', 'Toplam', 'Adet', 'OrtRisk']
                
                fig = make_subplots(rows=2, cols=1, 
                                   subplot_titles=('Aylık Toplam Harcama (₺)', 'Aylık İşlem Adedi'),
                                   shared_xaxes=True)
                
                fig.add_trace(go.Bar(x=monthly['Ay'], y=monthly['Toplam'],
                                    marker_color='#007850', name='Harcama'), row=1, col=1)
                fig.add_trace(go.Scatter(x=monthly['Ay'], y=monthly['Adet'],
                                        line=dict(color='#283c64', width=2),
                                        fill='tozeroy', fillcolor='rgba(139,92,246,0.1)',
                                        name='Adet'), row=2, col=1)
                
                fig.update_layout(
                    height=450, plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#4a5568'), showlegend=False
                )
                st.plotly_chart(fig, use_container_width=True)
                
                # User comparison
                if 'Kullanıcı' in df.columns:
                    user_stats = df.groupby('Kullanıcı').agg(
                        Toplam=('Tutar', 'sum'),
                        Adet=('Tutar', 'count'),
                        OrtRisk=('Risk_Skoru', 'mean')
                    ).reset_index()
                    
                    fig2 = px.bar(user_stats, x='Kullanıcı', y='Toplam',
                                  color='OrtRisk',
                                  color_continuous_scale='RdYlGn_r',
                                  title='Personel Bazlı Harcama Karşılaştırması',
                                  text=[f"₺{v:,.0f}" for v in user_stats['Toplam']])
                    fig2.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#4a5568')
                    )
                    st.plotly_chart(fig2, use_container_width=True)
            
            with tab2:
                st.markdown("### 🗓️ Günlük Harcama Yoğunluğu")
                
                df_temp2 = df.copy()
                df_temp2['dt'] = pd.to_datetime(df_temp2['Tarih'], errors='coerce')
                df_temp2['Gun'] = df_temp2['dt'].dt.day_name()
                df_temp2['Hafta'] = df_temp2['dt'].dt.isocalendar().week.astype(str)
                
                pivot = df_temp2.pivot_table(
                    values='Tutar', index='Gun', columns='Hafta', 
                    aggfunc='sum', fill_value=0
                )
                
                gun_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
                gun_tr = ['Pzt','Sal','Çar','Per','Cum','Cmt','Paz']
                pivot = pivot.reindex([g for g in gun_order if g in pivot.index])
                
                fig3 = go.Figure(go.Heatmap(
                    z=pivot.values,
                    x=pivot.columns.tolist(),
                    y=[gun_tr[gun_order.index(g)] for g in pivot.index if g in gun_order],
                    colorscale=[[0,'#0a0f1e'],[0.5,'#007850'],[1,'#283c64']],
                    text=[[f"₺{v:,.0f}" for v in row] for row in pivot.values],
                    texttemplate='%{text}',
                    textfont=dict(size=9)
                ))
                fig3.update_layout(
                    title='Haftalık Harcama Yoğunluk Haritası',
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#4a5568'), height=350
                )
                st.plotly_chart(fig3, use_container_width=True)
            
            with tab3:
                st.markdown("### 🔮 AI Harcama Tahmini")
                
                if st.button("📡 Tahmin Oluştur"):
                    with st.spinner("Gemini AI pattern analizi yapıyor..."):
                        pred = predict_monthly_spend(df, model)
                        
                        if pred:
                            current_month_spend = df[
                                pd.to_datetime(df['Tarih'], errors='coerce').dt.strftime('%Y-%m') == now_ist().strftime('%Y-%m')
                            ]['Tutar'].sum() if 'Tarih' in df.columns else 0
                            
                            st.markdown(f"""
                            <div style="display:grid; grid-template-columns:1fr 1fr; gap:16px; margin:16px 0;">
                                <div class="metric-card">
                                    <div style="color:var(--text-muted); font-size:0.75rem; text-transform:uppercase; letter-spacing:1px;">Bu Ay (Mevcut)</div>
                                    <div style="font-family:'Bebas Neue'; font-size:2.5rem; color:var(--accent-blue);">₺{current_month_spend:,.0f}</div>
                                </div>
                                <div class="metric-card">
                                    <div style="color:var(--text-muted); font-size:0.75rem; text-transform:uppercase; letter-spacing:1px;">Gelecek Ay Tahmini</div>
                                    <div style="font-family:'Bebas Neue'; font-size:2.5rem; color:var(--accent-purple);">₺{pred:,.0f}</div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            change_pct = ((pred - current_month_spend) / current_month_spend * 100) if current_month_spend > 0 else 0
                            direction = "📈 artış" if change_pct > 0 else "📉 azalış"
                            st.info(f"AI, geçmiş trendlere göre gelecek ay **{abs(change_pct):.1f}% {direction}** öngörüyor.")
                        else:
                            st.warning("Tahmin için yeterli geçmiş veri yok (en az 2 aylık veri gerekli).")
        else:
            st.info("Analitik görselleştirme için veri gerekli.")

    # ══════════════════════════════════════════════════════════
    # SAYFA: AI ASISTAN
    # ══════════════════════════════════════════════════════════
    elif selected == "🤖 AI Asistan":
        st.markdown('<div class="page-header"><div class="page-title">AI FİNANS ASISTANI</div></div>', unsafe_allow_html=True)
        
        st.markdown("""
        <div class="ultra-card" style="margin-bottom:20px;">
            <p style="color:var(--text-secondary); margin:0; font-size:0.9rem;">
                💬 Stinga AI'a finansal veriler hakkında her şeyi sorabilirsin. Harcama analizi, bütçe karşılaştırması,
                risk değerlendirmesi, trend analizi ve daha fazlası...
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Suggested questions
        st.markdown("**Hızlı Sorular:**")
        q_cols = st.columns(3)
        quick_qs = [
            "En yüksek harcama hangi proje?",
            "Bu ay risk durumu nasıl?",
            "Hangi personel en fazla harcıyor?",
            "Bütçe aşımı var mı?",
            "En çok hangi kategoride harcama var?",
            "Anomalileri özetle"
        ]
        
        for i, q in enumerate(quick_qs):
            with q_cols[i % 3]:
                if st.button(q, key=f"qq_{i}", use_container_width=True):
                    st.session_state.chat_history.append({"role": "user", "content": q})
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Chat history
        for msg in st.session_state.chat_history:
            if msg["role"] == "user":
                st.markdown(f"""
                <div style="display:flex; justify-content:flex-end; margin:8px 0;">
                    <div style="background:rgba(0,212,255,0.1); border:1px solid rgba(0,212,255,0.2); 
                                border-radius:16px 16px 0 16px; padding:12px 16px; max-width:70%;
                                color:var(--text-primary); font-size:0.9rem;">
                        {msg['content']}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="ai-bubble" style="max-width:85%;">
                    <p style="margin:0; line-height:1.7; font-size:0.9rem; white-space:pre-wrap;">
                        {msg['content']}
                    </p>
                </div>
                """, unsafe_allow_html=True)
        
        # Auto-answer last user message
        if st.session_state.chat_history and st.session_state.chat_history[-1]["role"] == "user":
            last_q = st.session_state.chat_history[-1]["content"]
            if model:
                with st.spinner("🤖 Stinga AI düşünüyor..."):
                    answer = generate_ai_insight(df, model, last_q)
                    st.session_state.chat_history.append({"role": "assistant", "content": answer})
                    st.rerun()
        
        # Input
        with st.form("chat_form", clear_on_submit=True):
            col_i, col_b = st.columns([5, 1])
            with col_i:
                user_q = st.text_input("", placeholder="Finansal veriler hakkında bir şey sor...", 
                                       label_visibility="collapsed")
            with col_b:
                sent = st.form_submit_button("→", use_container_width=True)
            
            if sent and user_q.strip():
                st.session_state.chat_history.append({"role": "user", "content": user_q})
                st.rerun()
        
        if st.session_state.chat_history:
            if st.button("🗑️ Sohbeti Temizle"):
                st.session_state.chat_history = []
                st.rerun()

    # ══════════════════════════════════════════════════════════
    # SAYFA: LEADERBOARD
    # ══════════════════════════════════════════════════════════
    elif selected == "🏆 Leaderboard":
        st.markdown('<div class="page-header"><div class="page-title">PERFORMANS SIRALAMASI</div></div>', unsafe_allow_html=True)
        
        st.markdown("""
        <div class="ultra-card">
            <p style="color:var(--text-secondary); margin:0; font-size:0.85rem;">
                🎮 Stinga Pro, harcama yönetimini oyunlaştırıyor. Fiş tarama, onay alma ve dakiklik gibi 
                kriterlere göre XP kazanarak seviye atla ve lider tablosuna çık!
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        xp_data = data_store.get("xp", {"Zeynep Özyaman": 0, "Şenol Faik Özyaman": 0, "Serkan Güzdemir": 0, "Okan İlhan": 0})
        sorted_users = sorted(xp_data.items(), key=lambda x: x[1], reverse=True)
        
        rank_icons = ["🥇", "🥈", "🥉"]
        rank_classes = ["gold", "silver", "bronze"]
        
        st.markdown("### 🏆 XP Sıralaması")
        
        for i, (uname, xp) in enumerate(sorted_users):
            # USERS dict'inde isim alanına göre eşleştir
            user_data = {}
            for _uk, _ui in USERS.items():
                if _ui.get("name","").lower() == uname.lower():
                    user_data = _ui
                    break
            if not user_data:
                # fallback: username key dene
                _lb_key = uname.lower().replace("ş","s").replace("ı","i").replace("ö","o").replace("ü","u").replace("ğ","g").replace("ç","c")
                user_data = USERS.get(_lb_key, USERS.get(uname.lower(), {}))
            avatar = user_data.get('avatar', '👤')
            level = xp // 500 + 1
            monthly_limit = user_data.get('monthly_limit', 15000)
            
            user_expenses = df[df['Kullanıcı'] == uname] if not df.empty and 'Kullanıcı' in df.columns else pd.DataFrame()
            total_spend = user_expenses['Tutar'].sum() if not user_expenses.empty else 0
            approved_count = len(user_expenses[user_expenses['Durum'] == 'Onaylandı']) if not user_expenses.empty and 'Durum' in user_expenses.columns else 0
            
            max_xp = sorted_users[0][1] if sorted_users else 1
            bar_pct = (xp / max_xp * 100) if max_xp > 0 else 0
            
            st.markdown(f"""
            <div class="lb-item" style="{'background:linear-gradient(135deg,rgba(255,215,0,0.1),rgba(255,215,0,0.05));border-color:rgba(255,215,0,0.3);' if i==0 else ''}">
                <div class="lb-rank {'gold' if i==0 else 'silver' if i==1 else 'bronze'}">{rank_icons[i] if i < 3 else str(i+1)}</div>
                <div style="font-size:1.8rem;">{avatar}</div>
                <div style="flex:1;">
                    <div style="font-weight:700; font-size:1rem;">{uname} <span style="font-size:0.7rem; color:var(--text-muted);">Lv.{level}</span></div>
                    <div style="font-size:0.75rem; color:var(--text-secondary);">
                        ₺{total_spend:,.0f} harcama · {approved_count} onaylı fiş
                    </div>
                    <div class="budget-track" style="margin-top:6px; width:200px;">
                        <div class="budget-fill" style="width:{bar_pct:.0f}%; {'background:linear-gradient(90deg,#ffd700,#d97706);' if i==0 else 'background:linear-gradient(90deg,#007850,#283c64);'}"></div>
                    </div>
                </div>
                <div style="text-align:right;">
                    <div style="font-family:'Bebas Neue'; font-size:1.8rem; color:{'#ffd700' if i==0 else 'var(--accent-blue)'};">{xp:,}</div>
                    <div style="font-size:0.7rem; color:var(--text-muted);">XP</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Badges section
        st.markdown("### 🎖️ Başarı Rozetleri")
        
        badge_defs = [
            ("🚀", "İlk Fiş", "İlk fişini tara", 1),
            ("💯", "Seri Teyitçi", "10 fiş tara", 10),
            ("🎯", "Risk Avcısı", "5 yüksek riskli fiş yakala", 5),
            ("⚡", "Hız Ustası", "Aynı gün 3 fiş tara", 3),
            ("🏆", "Finans Gurusu", "500 XP kazan", 500),
            ("💎", "Elit Operatör", "1000 XP kazan", 1000),
        ]
        
        badge_cols = st.columns(6)
        user_xp_val = xp_data.get(user_name, 0)
        user_exp_count = len(df[df['Kullanıcı'] == user_name]) if not df.empty and 'Kullanıcı' in df.columns else 0
        
        for i, (icon, name, desc, req) in enumerate(badge_defs):
            earned = (user_exp_count >= req) or (user_xp_val >= req)
            with badge_cols[i]:
                st.markdown(f"""
                <div style="text-align:center; padding:16px; background:var(--bg-card); border-radius:12px;
                            border:1px solid {'var(--accent-blue)' if earned else 'var(--border)'};
                            opacity:{'1' if earned else '0.4'}; transition:all 0.3s;">
                    <div style="font-size:2rem;">{icon}</div>
                    <div style="font-size:0.75rem; font-weight:700; margin:4px 0; color:{'var(--accent-blue)' if earned else 'var(--text-muted)'};">{name}</div>
                    <div style="font-size:0.65rem; color:var(--text-muted);">{desc}</div>
                    {"<div style='font-size:0.65rem; color:var(--accent-green); margin-top:4px;'>✓ KAZANILDI</div>" if earned else ""}
                </div>
                """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: ARŞİV & RAPOR
    # ══════════════════════════════════════════════════════════
    elif selected == "🗄️ Arşiv & Rapor":
        st.markdown('<div class="page-header"><div class="page-title">ARŞİV & RAPORLAMA</div></div>', unsafe_allow_html=True)
        
        tab_r, tab_a, tab_h, tab_harc, tab_zk, tab_sk = st.tabs([
            "📑 Raporlama", "🔍 Arşiv", "📜 Geçmiş AI Analizleri",
            "💵 Harcırah Harcamaları", "💳 Zeynep Özyaman K.K.", "🏦 Şirket Kredi Kartı"
        ])
        
        with tab_r:
            if not df.empty:
                df['Tarih_DT'] = pd.to_datetime(df['Tarih'], errors='coerce')
                df['Ay_Yil'] = df['Tarih_DT'].dt.strftime('%Y-%m')
                aylar = ["Tüm Zamanlar"] + sorted(df['Ay_Yil'].dropna().unique().tolist(), reverse=True)
                
                col_f1, col_f2 = st.columns(2)
                with col_f1:
                    secilen_ay = st.selectbox("Dönem", aylar)
                with col_f2:
                    secilen_proje = st.selectbox("Proje", ["Tümü"] + df['Proje'].unique().tolist() if 'Proje' in df.columns else ["Tümü"])
                
                filtered = df.copy()
                if secilen_ay != "Tüm Zamanlar":
                    filtered = filtered[filtered['Ay_Yil'] == secilen_ay]
                if secilen_proje != "Tümü" and 'Proje' in filtered.columns:
                    filtered = filtered[filtered['Proje'] == secilen_proje]
                
                # Summary metrics
                if not filtered.empty:
                    fc1, fc2, fc3, fc4 = st.columns(4)
                    fc1.metric("Toplam İşlem", len(filtered))
                    fc2.metric("Toplam Tutar", f"₺{filtered['Tutar'].sum():,.0f}")
                    fc3.metric("Ort. Risk", f"{filtered['Risk_Skoru'].mean():.0f}%" if 'Risk_Skoru' in filtered.columns else "N/A")
                    fc4.metric("Onay Oranı", f"{len(filtered[filtered['Durum']=='Onaylandı'])/len(filtered)*100:.0f}%" if 'Durum' in filtered.columns else "N/A")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                clean_df = filtered.drop(columns=['Tarih_DT','Ay_Yil'], errors='ignore')
                
                d1, d2, d3 = st.columns(3)
                d1.download_button(
                    "📥 CSV İndir",
                    clean_df.to_csv(index=False).encode('utf-8-sig'),
                    f"Stinga_Rapor_{secilen_ay}_{secilen_proje}.csv",
                    "text/csv", use_container_width=True
                )
                d2.download_button(
                    "📊 Muhasebe Excel (.xlsx)",
                    export_excel_muhasebe(clean_df, secilen_ay, logo_path="logo.png"),
                    f"Stinga_Excel_{secilen_ay}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                d3.download_button(
                    "📄 Muhasebe PDF",
                    export_pdf_muhasebe(clean_df, "Mali Rapor", secilen_ay, logo_path="logo.png"),
                    f"Stinga_PDF_{secilen_ay}.pdf",
                    "application/pdf", use_container_width=True
                )
                
                st.dataframe(clean_df.sort_values('Tarih', ascending=False), 
                            use_container_width=True, hide_index=True)
            else:
                st.info("Raporlanacak veri bulunmuyor.")
        
        with tab_a:
            st.markdown("### 🔍 Orijinal Fiş Arşivi")
            
            if not df.empty:
                search = st.text_input("Ara (firma, proje, personel...)", placeholder="🔍 Arama...")
                
                display_df = df.copy()
                if search:
                    mask = display_df.apply(lambda row: search.lower() in str(row).lower(), axis=1)
                    display_df = display_df[mask]
                
                islem_listesi = ["Seçim yapın..."] + display_df.apply(
                    lambda x: f"{x['ID']} | {x.get('Tarih','')} | {x.get('Firma','')} — ₺{float(x.get('Tutar',0)):,.0f} ({x.get('Durum','')})", 
                    axis=1
                ).tolist()
                
                secilen = st.selectbox("İşlem seç:", islem_listesi)
                
                if secilen != "Seçim yapın...":
                    islem_id = secilen.split(" | ")[0]
                    satir = df[df['ID'] == islem_id].iloc[0]
                    
                    col_i, col_img = st.columns([1, 1.5])
                    
                    with col_i:
                        st.markdown(f"""
                        <div class="ultra-card">
                            <div style="font-size:1.3rem; font-weight:700; margin-bottom:16px;">{satir.get('Firma','?')}</div>
                            <div style="display:grid; gap:8px;">
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">PERSONEL</span><br><strong>{satir.get('Kullanıcı','?')}</strong></div>
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">PROJE</span><br><strong>{satir.get('Proje','?')}</strong></div>
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">TUTAR</span><br><strong style="font-family:'Bebas Neue'; font-size:1.5rem; color:var(--accent-blue);">₺{float(satir.get('Tutar',0)):,.2f}</strong></div>
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">KDV</span><br><strong>₺{float(satir.get('KDV',0)):,.2f}</strong></div>
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">ÖDEME</span><br><strong>{'Kredi Kartı' if str(satir.get('Odeme_Turu','')).lower() in ('kredi_karti','kredi kartı') else 'Nakit' if str(satir.get('Odeme_Turu','')).lower()=='nakit' else str(satir.get('Odeme_Turu','?'))}</strong></div>
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">DURUM</span><br>{get_status_html(satir.get('Durum','?'))}</div>
                                <div><span style="color:var(--text-muted); font-size:0.75rem;">RİSK</span><br>{get_risk_html(satir.get('Risk_Skoru',0))}</div>
                            </div>
                            <div class="ai-bubble" style="margin-top:16px;">
                                {clean_audit(satir.get('AI_Audit',''))}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_img:
                        dosya   = satir.get('Dosya_Yolu', '')
                        b64_uri = satir.get('Gorsel_B64', '')
                        if dosya and os.path.exists(dosya):
                            st.image(dosya, caption=f"Orijinal Fiş — {islem_id}", use_container_width=True)
                        elif b64_uri:
                            try:
                                header, b64_data = b64_uri.split(",", 1) if "," in b64_uri else ("", b64_uri)
                                img_bytes = base64.b64decode(b64_data)
                                st.image(img_bytes, caption=f"📱 WhatsApp Fişi — {islem_id}", use_container_width=True)
                            except Exception:
                                st.markdown(f"""
                                <div style="height:300px; background:var(--bg-secondary); border-radius:16px; 
                                            display:flex; flex-direction:column; align-items:center; justify-content:center;
                                            color:var(--text-muted); border:2px dashed var(--border);">
                                    <div style="font-size:3rem; opacity:0.3;">📷</div>
                                    <div style="font-size:0.85rem; margin-top:8px;">Görsel yüklenemedi</div>
                                </div>
                                """, unsafe_allow_html=True)
                        else:
                            st.markdown(f"""
                            <div style="height:300px; background:var(--bg-secondary); border-radius:16px; 
                                        display:flex; flex-direction:column; align-items:center; justify-content:center;
                                        color:var(--text-muted); border:2px dashed var(--border);">
                                <div style="font-size:3rem; opacity:0.3;">📷</div>
                                <div style="font-size:0.85rem; margin-top:8px;">Görsel bulunamadı</div>
                            </div>
                            """, unsafe_allow_html=True)
            else:
                st.info("Arşivde görüntülenecek veri yok.")
        
        with tab_h:
            st.markdown("### 📜 Geçmiş AI Analizleri")
            
            ai_insights = data_store.get("ai_insights", [])
            if ai_insights:
                for ins in reversed(ai_insights[-10:]):
                    with st.expander(f"🤖 {ins.get('date','?')} — {ins.get('type','Analiz').replace('_',' ').title()}"):
                        st.markdown(f"""
                        <div class="ai-bubble">
                            <p style="margin:0; white-space:pre-wrap; font-size:0.85rem; line-height:1.7;">
                                {ins.get('content','İçerik yok.')}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="empty-state">
                    <div class="empty-icon">🤖</div>
                    <div>Henüz AI analizi kaydedilmemiş.</div>
                    <div style="font-size:0.8rem; margin-top:8px;">Anomali Dedektörü'nden derin analiz başlat!</div>
                </div>
                """, unsafe_allow_html=True)

        # ── TAB: HARCIRAH HARCAMALARI ─────────────────────────
        with tab_harc:
            st.markdown("### 💵 Harcırah Harcamaları")
            st.markdown("""
            <div style="background:rgba(17,133,91,0.07); border:1px solid rgba(17,133,91,0.18);
                        border-radius:12px; padding:14px 18px; margin-bottom:16px; font-size:0.85rem; color:#3d5260;">
                Personel harcırah hesabından (nakit / şahsi kart) yapılan tüm harcamalar.
            </div>
            """, unsafe_allow_html=True)

            _harc_odeme_keys = ("harcirah", "harcırah", "harcirahtan dus", "harcırahtan düş",
                                "harcırahtan düş (nakit / kişisel kart)", "nakit", "kisisel")
            if not df_full.empty and 'Odeme_Turu' in df_full.columns:
                harc_df = df_full[df_full['Odeme_Turu'].str.lower().str.strip().isin(_harc_odeme_keys)].copy()
            else:
                harc_df = pd.DataFrame()

            if not harc_df.empty:
                hc1, hc2, hc3 = st.columns(3)
                with hc1:
                    _harc_personel = st.selectbox("Personel", ["Tümü"] + sorted(harc_df['Kullanıcı'].unique().tolist()), key="harc_per")
                with hc2:
                    harc_df['_Tarih_DT'] = pd.to_datetime(harc_df['Tarih'], errors='coerce')
                    harc_df['_Ay_Yil'] = harc_df['_Tarih_DT'].dt.strftime('%Y-%m')
                    _harc_aylar = ["Tüm Zamanlar"] + sorted(harc_df['_Ay_Yil'].dropna().unique().tolist(), reverse=True)
                    _harc_ay = st.selectbox("Dönem", _harc_aylar, key="harc_ay")
                with hc3:
                    _harc_durum = st.selectbox("Durum", ["Tümü", "Onaylandı", "Onay Bekliyor", "Reddedildi"], key="harc_dur")

                _hf = harc_df.copy()
                if _harc_personel != "Tümü":
                    _hf = _hf[_hf['Kullanıcı'] == _harc_personel]
                if _harc_ay != "Tüm Zamanlar":
                    _hf = _hf[_hf['_Ay_Yil'] == _harc_ay]
                if _harc_durum != "Tümü":
                    _hf = _hf[_hf['Durum'] == _harc_durum]

                if not _hf.empty:
                    hm1, hm2, hm3, hm4 = st.columns(4)
                    hm1.metric("Toplam İşlem", len(_hf))
                    hm2.metric("Toplam Tutar", f"₺{_hf['Tutar'].sum():,.0f}")
                    _hf_onay = _hf[_hf['Durum'] == 'Onaylandı']
                    hm3.metric("Onaylı Tutar", f"₺{_hf_onay['Tutar'].sum():,.0f}")
                    hm4.metric("Bekleyen", f"₺{_hf[_hf['Durum']=='Onay Bekliyor']['Tutar'].sum():,.0f}")

                    # Personel bazlı bakiye özeti
                    if _harc_personel == "Tümü":
                        st.markdown("#### 👥 Personel Harcırah Bakiyeleri")
                        for _ukey, _uinfo in USERS.items():
                            _p_name = _uinfo["name"]
                            _p_bal = get_user_wallet_balance(_p_name, data_store)
                            _p_harc = _hf[_hf['Kullanıcı'] == _p_name]['Tutar'].sum() if not _hf[_hf['Kullanıcı'] == _p_name].empty else 0
                            if _p_harc > 0 or _p_bal != 0:
                                st.markdown(f"""
                                <div style="display:flex; justify-content:space-between; align-items:center;
                                            padding:10px 16px; background:#fff; border:1px solid rgba(0,0,0,0.07);
                                            border-radius:10px; margin:4px 0;">
                                    <div><strong>{_uinfo['avatar']} {_p_name}</strong></div>
                                    <div style="display:flex; gap:20px; font-size:0.85rem;">
                                        <span>Harcama: <strong>₺{_p_harc:,.0f}</strong></span>
                                        <span>Bakiye: <strong style="color:{'#dc2626' if _p_bal < 0 else '#007850'}">
                                            {'−' if _p_bal < 0 else ''}₺{abs(_p_bal):,.0f}</strong></span>
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)

                    st.markdown("<br>", unsafe_allow_html=True)
                    _hf_clean = _hf.drop(columns=['_Tarih_DT', '_Ay_Yil'], errors='ignore')
                    hd1, hd2, hd3 = st.columns(3)
                    hd1.download_button("📥 CSV İndir", _hf_clean.to_csv(index=False).encode('utf-8-sig'),
                        f"Harcirah_{_harc_ay}.csv", "text/csv", use_container_width=True, key="harc_csv")
                    hd2.download_button("📊 Excel İndir",
                        export_excel_muhasebe(_hf_clean, f"Harcırah — {_harc_ay}", logo_path="logo.png"),
                        f"Harcirah_{_harc_ay}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="harc_xlsx")
                    hd3.download_button("📄 PDF İndir",
                        export_pdf_muhasebe(_hf_clean, "Harcırah Raporu", f"Harcırah — {_harc_ay}", logo_path="logo.png"),
                        f"Harcirah_{_harc_ay}.pdf", "application/pdf", use_container_width=True, key="harc_pdf")

                    st.dataframe(_hf_clean.sort_values('Tarih', ascending=False), use_container_width=True, hide_index=True)
                else:
                    st.info("Seçili filtrelere uygun harcırah harcaması bulunamadı.")
            else:
                st.markdown("""
                <div class="empty-state"><div class="empty-icon">💵</div>
                <div>Harcırah kaydı bulunamadı.</div></div>""", unsafe_allow_html=True)

        # ── TAB: ZEYNEP ÖZYAMAN ŞAHSİ KREDİ KARTI ───────────
        with tab_zk:
            st.markdown("### 💳 Zeynep Özyaman — Şahsi Kredi Kartı Harcamaları")
            st.markdown("""
            <div style="background:rgba(47,60,110,0.07); border:1px solid rgba(47,60,110,0.18);
                        border-radius:12px; padding:14px 18px; margin-bottom:16px; font-size:0.85rem; color:#3d5260;">
                Zeynep Özyaman'ın şahsi kredi kartından yapılan şirket harcamaları.
            </div>
            """, unsafe_allow_html=True)

            # Zeynep'in harcırah ile yaptığı ödemeler — şahsi kart
            if not df_full.empty and 'Kullanıcı' in df_full.columns and 'Odeme_Turu' in df_full.columns:
                _zk_mask = (
                    df_full['Kullanıcı'].str.contains('Zeynep', case=False, na=False) &
                    df_full['Odeme_Turu'].str.lower().str.strip().isin(_harc_odeme_keys)
                )
                zk_df = df_full[_zk_mask].copy()
            else:
                zk_df = pd.DataFrame()

            if not zk_df.empty:
                zk_df['_Tarih_DT'] = pd.to_datetime(zk_df['Tarih'], errors='coerce')
                zk_df['_Ay_Yil'] = zk_df['_Tarih_DT'].dt.strftime('%Y-%m')
                zc1, zc2 = st.columns(2)
                with zc1:
                    _zk_aylar = ["Tüm Zamanlar"] + sorted(zk_df['_Ay_Yil'].dropna().unique().tolist(), reverse=True)
                    _zk_ay = st.selectbox("Dönem", _zk_aylar, key="zk_ay")
                with zc2:
                    _zk_durum = st.selectbox("Durum", ["Tümü", "Onaylandı", "Onay Bekliyor", "Reddedildi"], key="zk_dur")

                _zkf = zk_df.copy()
                if _zk_ay != "Tüm Zamanlar":
                    _zkf = _zkf[_zkf['_Ay_Yil'] == _zk_ay]
                if _zk_durum != "Tümü":
                    _zkf = _zkf[_zkf['Durum'] == _zk_durum]

                if not _zkf.empty:
                    zm1, zm2, zm3 = st.columns(3)
                    zm1.metric("Toplam İşlem", len(_zkf))
                    zm2.metric("Toplam Tutar", f"₺{_zkf['Tutar'].sum():,.0f}")
                    zm3.metric("Onaylı Tutar", f"₺{_zkf[_zkf['Durum']=='Onaylandı']['Tutar'].sum():,.0f}")

                    # Kategori kırılımı
                    if 'Kategori' in _zkf.columns:
                        _zk_kat = _zkf.groupby('Kategori')['Tutar'].agg(['sum','count']).sort_values('sum', ascending=False)
                        st.markdown("#### 📂 Kategori Kırılımı")
                        for kat, row in _zk_kat.iterrows():
                            st.markdown(f"- **{kat}**: ₺{row['sum']:,.0f} ({int(row['count'])} fiş)")

                    st.markdown("<br>", unsafe_allow_html=True)
                    _zkf_clean = _zkf.drop(columns=['_Tarih_DT', '_Ay_Yil'], errors='ignore')
                    zd1, zd2, zd3 = st.columns(3)
                    zd1.download_button("📥 CSV", _zkf_clean.to_csv(index=False).encode('utf-8-sig'),
                        f"ZeynepKK_{_zk_ay}.csv", "text/csv", use_container_width=True, key="zk_csv")
                    zd2.download_button("📊 Excel",
                        export_excel_muhasebe(_zkf_clean, f"Zeynep Özyaman K.K. — {_zk_ay}", logo_path="logo.png"),
                        f"ZeynepKK_{_zk_ay}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="zk_xlsx")
                    zd3.download_button("📄 PDF",
                        export_pdf_muhasebe(_zkf_clean, "Zeynep Özyaman Şahsi K.K. Raporu", f"Zeynep K.K. — {_zk_ay}", logo_path="logo.png"),
                        f"ZeynepKK_{_zk_ay}.pdf", "application/pdf", use_container_width=True, key="zk_pdf")

                    st.dataframe(_zkf_clean.sort_values('Tarih', ascending=False), use_container_width=True, hide_index=True)
                else:
                    st.info("Seçili filtrelere uygun kayıt bulunamadı.")
            else:
                st.markdown("""
                <div class="empty-state"><div class="empty-icon">💳</div>
                <div>Zeynep Özyaman şahsi kart kaydı bulunamadı.</div></div>""", unsafe_allow_html=True)

        # ── TAB: ŞİRKET KREDİ KARTI ─────────────────────────
        with tab_sk:
            st.markdown("### 🏦 Şirket Kredi Kartı Harcamaları")
            st.markdown("""
            <div style="background:rgba(8,145,178,0.07); border:1px solid rgba(8,145,178,0.18);
                        border-radius:12px; padding:14px 18px; margin-bottom:16px; font-size:0.85rem; color:#3d5260;">
                Stinga Enerji şirket kredi kartından yapılan tüm harcamalar.
            </div>
            """, unsafe_allow_html=True)

            _sirket_odeme_keys = ("sirket_karti", "sirket karti", "kredi_karti", "kredi kartı",
                                   "şirket", "sirket", "şirket kredi kartı", "sirket kredi karti")
            if not df_full.empty and 'Odeme_Turu' in df_full.columns:
                sk_df = df_full[df_full['Odeme_Turu'].str.lower().str.strip().isin(_sirket_odeme_keys)].copy()
            else:
                sk_df = pd.DataFrame()

            if not sk_df.empty:
                sk_df['_Tarih_DT'] = pd.to_datetime(sk_df['Tarih'], errors='coerce')
                sk_df['_Ay_Yil'] = sk_df['_Tarih_DT'].dt.strftime('%Y-%m')
                sc_c1, sc_c2, sc_c3 = st.columns(3)
                with sc_c1:
                    _sk_aylar = ["Tüm Zamanlar"] + sorted(sk_df['_Ay_Yil'].dropna().unique().tolist(), reverse=True)
                    _sk_ay = st.selectbox("Dönem", _sk_aylar, key="sk_ay")
                with sc_c2:
                    _sk_personel = st.selectbox("Personel", ["Tümü"] + sorted(sk_df['Kullanıcı'].unique().tolist()), key="sk_per")
                with sc_c3:
                    _sk_durum = st.selectbox("Durum", ["Tümü", "Onaylandı", "Onay Bekliyor", "Reddedildi"], key="sk_dur")

                _skf = sk_df.copy()
                if _sk_ay != "Tüm Zamanlar":
                    _skf = _skf[_skf['_Ay_Yil'] == _sk_ay]
                if _sk_personel != "Tümü":
                    _skf = _skf[_skf['Kullanıcı'] == _sk_personel]
                if _sk_durum != "Tümü":
                    _skf = _skf[_skf['Durum'] == _sk_durum]

                if not _skf.empty:
                    sm1, sm2, sm3, sm4 = st.columns(4)
                    sm1.metric("Toplam İşlem", len(_skf))
                    sm2.metric("Toplam Tutar", f"₺{_skf['Tutar'].sum():,.0f}")
                    sm3.metric("Onaylı Tutar", f"₺{_skf[_skf['Durum']=='Onaylandı']['Tutar'].sum():,.0f}")
                    sm4.metric("KDV Toplamı", f"₺{_skf['KDV'].sum():,.0f}" if 'KDV' in _skf.columns else "N/A")

                    # Kategori kırılımı
                    if 'Kategori' in _skf.columns:
                        _sk_kat = _skf.groupby('Kategori')['Tutar'].agg(['sum','count']).sort_values('sum', ascending=False)
                        fig_sk = go.Figure(go.Bar(
                            x=_sk_kat['sum'].values, y=_sk_kat.index,
                            orientation='h',
                            marker=dict(color='#0891b2'),
                            text=[f"₺{v:,.0f}" for v in _sk_kat['sum'].values],
                            textposition='outside'
                        ))
                        fig_sk.update_layout(title='Şirket Kartı — Kategori Bazlı Harcama',
                            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                            font=dict(color='#4a5568'), height=280,
                            xaxis=dict(showgrid=False), yaxis=dict(showgrid=False))
                        st.plotly_chart(fig_sk, use_container_width=True)

                    st.markdown("<br>", unsafe_allow_html=True)
                    _skf_clean = _skf.drop(columns=['_Tarih_DT', '_Ay_Yil'], errors='ignore')
                    sd1, sd2, sd3 = st.columns(3)
                    sd1.download_button("📥 CSV", _skf_clean.to_csv(index=False).encode('utf-8-sig'),
                        f"SirketKK_{_sk_ay}.csv", "text/csv", use_container_width=True, key="sk_csv")
                    sd2.download_button("📊 Excel",
                        export_excel_muhasebe(_skf_clean, f"Şirket K.K. — {_sk_ay}", logo_path="logo.png"),
                        f"SirketKK_{_sk_ay}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="sk_xlsx")
                    sd3.download_button("📄 PDF",
                        export_pdf_muhasebe(_skf_clean, "Şirket Kredi Kartı Raporu", f"Şirket K.K. — {_sk_ay}", logo_path="logo.png"),
                        f"SirketKK_{_sk_ay}.pdf", "application/pdf", use_container_width=True, key="sk_pdf")

                    st.dataframe(_skf_clean.sort_values('Tarih', ascending=False), use_container_width=True, hide_index=True)
                else:
                    st.info("Seçili filtrelere uygun kayıt bulunamadı.")
            else:
                st.markdown("""
                <div class="empty-state"><div class="empty-icon">🏦</div>
                <div>Şirket kredi kartı kaydı bulunamadı.</div></div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: AI BÜTÇE KOÇU
    # ══════════════════════════════════════════════════════════
    elif selected == "🧠 AI Bütçe Koçu":
        st.markdown('<div class="page-header"><div class="page-title">🧠 AI BÜTÇE KOÇU</div></div>', unsafe_allow_html=True)

        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(17,133,91,0.12),rgba(47,60,110,0.10));
                    border:1px solid rgba(17,133,91,0.25); border-radius:16px;
                    padding:20px 24px; margin-bottom:24px;">
            <div style="font-size:1.05rem; font-weight:700; color:#11855B; margin-bottom:6px;">
                💡 Kişisel Finans Koçunuz
            </div>
            <div style="font-size:0.85rem; color:#3d5260; line-height:1.6;">
                Harcama verilerinizi analiz ederek bütçe optimizasyon önerileri, tasarruf fırsatları
                ve kişiselleştirilmiş finansal tavsiyeler sunar.
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Kullanıcıya özel veriler
        my_expenses = df.copy()  # df zaten role'e göre filtrelenmiş

        coach_mode = st.selectbox("🎯 Koçluk Modu Seç", [
            "📊 Harcama Analizi & Optimizasyon",
            "💰 Aylık Bütçe Planı Oluştur",
            "✂️ Tasarruf Fırsatları Bul",
            "🎯 Hedef Bazlı Bütçe",
            "📈 Harcama Alışkanlıkları Raporu"
        ])

        hedef_tutar = None
        hedef_aciklama = ""
        if coach_mode == "🎯 Hedef Bazlı Bütçe":
            col_h1, col_h2 = st.columns(2)
            with col_h1:
                hedef_tutar = st.number_input("Hedef Tasarruf Tutarı (₺)", min_value=0, value=5000, step=500)
            with col_h2:
                hedef_aciklama = st.text_input("Tasarruf Hedefi (ör: Ekipman alımı)", value="")

        extra_context = st.text_area(
            "📝 AI'ye ek bilgi ver (isteğe bağlı)",
            placeholder="Örn: Bu ay fazla seyahat harcamamı azaltmak istiyorum...",
            height=80
        )

        if st.button("🧠 Koç Analizi Başlat", use_container_width=True, type="primary"):
            with st.spinner("AI Bütçe Koçunuz analiz yapıyor..."):
                try:
                    # Veri özetini hazırla
                    if not my_expenses.empty and 'Tutar' in my_expenses.columns:
                        toplam = my_expenses['Tutar'].sum()
                        kategori_ozet = ""
                        if 'Kategori' in my_expenses.columns:
                            kat_grp = my_expenses.groupby('Kategori')['Tutar'].sum().sort_values(ascending=False)
                            kategori_ozet = "\n".join([f"  - {k}: ₺{v:,.0f}" for k, v in kat_grp.items()])
                        limit = user_info.get('monthly_limit', 15000)
                        veri_ozet = f"""
Kullanıcı: {user_name}
Unvan: {user_info.get('title', '-')}
Aylık Limit: ₺{limit:,.0f}
Toplam Harcama: ₺{toplam:,.0f}
Limit Kullanımı: %{min(toplam/limit*100,100):.1f}
Fiş Sayısı: {len(my_expenses)}
Kategori Dağılımı:
{kategori_ozet if kategori_ozet else 'Veri yok'}
"""
                    else:
                        veri_ozet = f"Kullanıcı: {user_name}\nHarcama verisi bulunamadı."

                    hedef_text = f"\nHEDEF: {hedef_tutar}₺ tasarruf — {hedef_aciklama}" if hedef_tutar else ""
                    extra_text = f"\nEK BİLGİ: {extra_context}" if extra_context.strip() else ""

                    prompt = f"""Sen Stinga Pro Finance v17.0 uygulamasının AI Bütçe Koçusun. Türkçe yanıt ver.

KULLANICI VERİSİ:
{veri_ozet}{hedef_text}{extra_text}

MOD: {coach_mode}

Lütfen seçilen moda göre detaylı, uygulanabilir ve kişiselleştirilmiş bütçe koçluğu yap.
Somut rakamlar, yüzdeler ve önceliklendirilmiş öneriler sun.
Varsa risk noktalarını vurgula. Emojilerle destekle ama profesyonel kal.
Yanıtın yapısı:
1. Mevcut Durum Değerlendirmesi
2. Tespit Edilen Fırsatlar / Sorunlar  
3. Kişiselleştirilmiş Öneriler (somut adımlar)
4. {('Hedef Gerçekleşme Planı' if hedef_tutar else 'Motivasyon & Özet')}
"""
                    response = genai.GenerativeModel("models/gemini-2.5-flash").generate_content(prompt)
                    coach_result = response.text

                    st.markdown(f"""
                    <div class="ai-bubble" style="margin-top:16px;">
                        <div style="font-size:0.75rem; color:#11855B; font-weight:700; margin-bottom:10px; letter-spacing:0.1em;">
                            🧠 AI BÜTÇE KOÇU ANALİZİ
                        </div>
                        <div style="white-space:pre-wrap; font-size:0.87rem; line-height:1.75; color:#0f1923;">
{coach_result}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    # XP ödülü
                    _xp_data = data_store.get("xp", {})
                    _xp_data[user_name] = _xp_data.get(user_name, 0) + 15
                    data_store["xp"] = _xp_data
                    save_data(data_store)
                    st.success("✅ +15 XP kazandınız!")

                except Exception as e:
                    st.error(f"AI analizi başarısız: {e}")

        # Hızlı ipuçları
        if not my_expenses.empty and 'Kategori' in my_expenses.columns:
            st.markdown("---")
            st.markdown("#### ⚡ Hızlı Bütçe Özeti")
            kat_cols = st.columns(4)
            kat_top = my_expenses.groupby('Kategori')['Tutar'].sum().sort_values(ascending=False).head(4)
            for i, (kat, tutar) in enumerate(kat_top.items()):
                with kat_cols[i]:
                    pct = tutar / my_expenses['Tutar'].sum() * 100
                    color = "#dc2626" if pct > 40 else ("#d97706" if pct > 25 else "#11855B")
                    st.markdown(f"""
                    <div style="background:#fff; border:1px solid rgba(17,133,91,0.15); border-radius:12px;
                                padding:14px; text-align:center;">
                        <div style="font-size:1.3rem; margin-bottom:4px;">
                            {"🍽️" if "Yemek" in kat else "🚗" if "Ulaşım" in kat else "🏨" if "Konaklama" in kat else "📦"}
                        </div>
                        <div style="font-size:0.7rem; color:#3d5260; font-weight:600;">{kat}</div>
                        <div style="font-size:1rem; font-weight:800; color:{color};">₺{tutar:,.0f}</div>
                        <div style="font-size:0.68rem; color:{color};">%{pct:.0f}</div>
                    </div>
                    """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: GİDER TAHMİNCİSİ
    # ══════════════════════════════════════════════════════════
    elif selected == "🔮 Gider Tahmincisi":
        st.markdown('<div class="page-header"><div class="page-title">🔮 GİDER TAHMİNCİSİ</div></div>', unsafe_allow_html=True)

        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(47,60,110,0.12),rgba(139,92,246,0.08));
                    border:1px solid rgba(47,60,110,0.25); border-radius:16px;
                    padding:20px 24px; margin-bottom:24px;">
            <div style="font-size:1.05rem; font-weight:700; color:#2F3C6E; margin-bottom:6px;">
                🔮 Yapay Zeka Destekli Gider Tahmini
            </div>
            <div style="font-size:0.85rem; color:#3d5260; line-height:1.6;">
                Geçmiş harcama örüntülerinizi analiz ederek gelecek dönemler için akıllı tahminler üretir.
                Senaryo bazlı projeksiyon ve erken uyarı sistemi ile bütçe aşımlarını öngörün.
            </div>
        </div>
        """, unsafe_allow_html=True)

        col_t1, col_t2 = st.columns(2)
        with col_t1:
            tahmin_donemi = st.selectbox("📅 Tahmin Dönemi", [
                "Önümüzdeki 1 Ay",
                "Önümüzdeki 3 Ay",
                "Önümüzdeki 6 Ay",
                "Yıl Sonu Tahmini"
            ])
        with col_t2:
            senaryo = st.selectbox("🎭 Senaryo", [
                "📊 Mevcut Trend Devam",
                "📈 Optimistik (%20 Düşüş)",
                "📉 Kötümser (%30 Artış)",
                "🎯 Bütçe Hedefine Göre"
            ])

        proje_filtre = "Tümü"
        if not df.empty and 'Proje' in df.columns:
            projeler = ["Tümü"] + df['Proje'].dropna().unique().tolist()
            proje_filtre = st.selectbox("🏗️ Proje Filtresi", projeler)

        if st.button("🔮 Tahmin Üret", use_container_width=True, type="primary"):
            with st.spinner("AI gider modeli hesaplıyor..."):
                try:
                    filtered_df = df.copy()
                    if proje_filtre != "Tümü" and 'Proje' in filtered_df.columns:
                        filtered_df = filtered_df[filtered_df['Proje'] == proje_filtre]

                    if not filtered_df.empty and 'Tutar' in filtered_df.columns:
                        # Aylık istatistikler
                        filtered_df['Tarih_DT'] = pd.to_datetime(filtered_df['Tarih'], errors='coerce')
                        filtered_df['Ay'] = filtered_df['Tarih_DT'].dt.strftime('%Y-%m')
                        aylik = filtered_df.groupby('Ay')['Tutar'].sum()
                        ort_aylik = aylik.mean()
                        max_ay = aylik.max()
                        min_ay = aylik.min()
                        son_ay = aylik.iloc[-1] if len(aylik) > 0 else 0

                        kategori_trendleri = ""
                        if 'Kategori' in filtered_df.columns:
                            kat_aylik = filtered_df.groupby(['Ay','Kategori'])['Tutar'].sum().unstack(fill_value=0)
                            for kat in kat_aylik.columns[:5]:
                                seri = kat_aylik[kat]
                                trend = "↗️ Artıyor" if len(seri) > 1 and seri.iloc[-1] > seri.iloc[-2] else "↘️ Azalıyor"
                                kategori_trendleri += f"  - {kat}: ort. ₺{seri.mean():,.0f}/ay ({trend})\n"

                        veri_ozet = f"""
Kullanıcı: {user_name}
Proje: {proje_filtre}
Aylık Ortalama: ₺{ort_aylik:,.0f}
En Yüksek Ay: ₺{max_ay:,.0f}
En Düşük Ay: ₺{min_ay:,.0f}
Son Ay: ₺{son_ay:,.0f}
Toplam Fiş: {len(filtered_df)}
Kategori Trendleri:
{kategori_trendleri if kategori_trendleri else 'Veri yok'}
Aylık Limit: ₺{user_info.get('monthly_limit',15000):,.0f}
"""
                    else:
                        veri_ozet = "Yeterli harcama verisi bulunamadı."

                    prompt = f"""Sen Stinga Pro Finance v17.0'ın AI Gider Tahmincisisin. Türkçe yanıt ver.

MEVCUT VERİ:
{veri_ozet}

TAHMİN DÖNEMİ: {tahmin_donemi}
SENARYO: {senaryo}

Lütfen aşağıdaki formatta detaylı gider tahmini yap:

1. 📊 MEVCUT TREND ANALİZİ
   - Genel eğilim değerlendirmesi
   - Öne çıkan pattern'lar

2. 🔮 {tahmin_donemi.upper()} TAHMİN ({senaryo})
   - Tahmini toplam gider: ₺XX,XXX
   - Aylık bazda tahmin kırılımı
   - Kategori bazlı tahminler

3. ⚠️ RİSK & UYARI SİNYALLERİ
   - Bütçe aşım riski var mı?
   - Dikkat edilmesi gereken kalemler

4. 💡 ERKEN ÖNLEM ÖNERİLERİ
   - Tahmin doğrultusunda alınabilecek aksiyonlar

Somut rakamlar ve yüzdeler kullan. Profesyonel ama anlaşılır ol.
"""
                    response = genai.GenerativeModel("models/gemini-2.5-flash").generate_content(prompt)
                    tahmin_result = response.text

                    st.markdown(f"""
                    <div class="ai-bubble" style="margin-top:16px; border-left:3px solid #2F3C6E;">
                        <div style="font-size:0.75rem; color:#2F3C6E; font-weight:700; margin-bottom:10px; letter-spacing:0.1em;">
                            🔮 AI GİDER TAHMİN RAPORU — {tahmin_donemi.upper()}
                        </div>
                        <div style="white-space:pre-wrap; font-size:0.87rem; line-height:1.75; color:#0f1923;">
{tahmin_result}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    # Görsel projeksiyon grafiği
                    if not df.empty and 'Tutar' in df.columns:
                        try:
                            df_tmp = df.copy()
                            df_tmp['Tarih_DT'] = pd.to_datetime(df_tmp['Tarih'], errors='coerce')
                            df_tmp['Ay'] = df_tmp['Tarih_DT'].dt.strftime('%Y-%m')
                            aylik_gercek = df_tmp.groupby('Ay')['Tutar'].sum().reset_index()
                            aylik_gercek.columns = ['Ay', 'Tutar']
                            aylik_gercek['Tip'] = 'Gerçekleşen'

                            # Basit projeksiyon
                            if len(aylik_gercek) >= 2:
                                ort = aylik_gercek['Tutar'].mean()
                                multiplier = 1.0
                                if "Optimistik" in senaryo:
                                    multiplier = 0.80
                                elif "Kötümser" in senaryo:
                                    multiplier = 1.30
                                ay_sayisi = 1 if "1 Ay" in tahmin_donemi else (3 if "3 Ay" in tahmin_donemi else (6 if "6 Ay" in tahmin_donemi else 5))
                                son_ay_dt = pd.to_datetime(aylik_gercek['Ay'].iloc[-1])
                                proj_rows = []
                                for i in range(1, ay_sayisi + 1):
                                    yeni_ay = (son_ay_dt + pd.DateOffset(months=i)).strftime('%Y-%m')
                                    proj_rows.append({'Ay': yeni_ay, 'Tutar': ort * multiplier, 'Tip': 'Tahmin'})
                                proj_df = pd.DataFrame(proj_rows)
                                combined = pd.concat([aylik_gercek, proj_df], ignore_index=True)

                                fig = go.Figure()
                                gercek = combined[combined['Tip'] == 'Gerçekleşen']
                                tahmin_d = combined[combined['Tip'] == 'Tahmin']
                                fig.add_trace(go.Scatter(x=gercek['Ay'], y=gercek['Tutar'],
                                    mode='lines+markers', name='Gerçekleşen',
                                    line=dict(color='#11855B', width=2.5),
                                    marker=dict(size=7)))
                                fig.add_trace(go.Scatter(x=tahmin_d['Ay'], y=tahmin_d['Tutar'],
                                    mode='lines+markers', name='Tahmin',
                                    line=dict(color='#2F3C6E', width=2.5, dash='dash'),
                                    marker=dict(size=7, symbol='diamond')))
                                # Limit çizgisi
                                fig.add_hline(y=user_info.get('monthly_limit', 15000),
                                    line_dash="dot", line_color="#dc2626",
                                    annotation_text="Aylık Limit",
                                    annotation_position="top right")
                                fig.update_layout(
                                    title=f"📈 Gider Trendi & {tahmin_donemi} Projeksiyonu",
                                    xaxis_title="Ay", yaxis_title="Tutar (₺)",
                                    plot_bgcolor='rgba(0,0,0,0)',
                                    paper_bgcolor='rgba(0,0,0,0)',
                                    height=360,
                                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                                )
                                st.plotly_chart(fig, use_container_width=True)
                        except Exception:
                            pass

                    # XP ödülü
                    _xp_data = data_store.get("xp", {})
                    _xp_data[user_name] = _xp_data.get(user_name, 0) + 20
                    data_store["xp"] = _xp_data
                    save_data(data_store)
                    st.success("✅ +20 XP kazandınız!")

                except Exception as e:
                    st.error(f"AI tahmin hatası: {e}")

        # Örnek senaryo kartları (statik, her zaman gösterilir)
        st.markdown("---")
        st.markdown("#### 📋 Senaryo Karşılaştırması")
        sc1, sc2, sc3 = st.columns(3)
        limit = user_info.get('monthly_limit', 15000)
        ort_harcama = df['Tutar'].mean() * 30 if not df.empty and 'Tutar' in df.columns else limit * 0.7
        with sc1:
            st.markdown(f"""
            <div style="background:#f0faf5; border:1px solid #11855B33; border-radius:12px; padding:16px; text-align:center;">
                <div style="font-size:1.4rem;">📊</div>
                <div style="font-weight:700; color:#11855B; margin:6px 0;">Mevcut Trend</div>
                <div style="font-size:1.1rem; font-weight:800; color:#0f1923;">₺{ort_harcama:,.0f}</div>
                <div style="font-size:0.72rem; color:#7a96a4;">tahmini aylık</div>
            </div>
            """, unsafe_allow_html=True)
        with sc2:
            st.markdown(f"""
            <div style="background:#f0f4ff; border:1px solid #2F3C6E33; border-radius:12px; padding:16px; text-align:center;">
                <div style="font-size:1.4rem;">📈</div>
                <div style="font-weight:700; color:#2F3C6E; margin:6px 0;">Optimistik</div>
                <div style="font-size:1.1rem; font-weight:800; color:#0f1923;">₺{ort_harcama*0.8:,.0f}</div>
                <div style="font-size:0.72rem; color:#7a96a4;">%20 tasarruf</div>
            </div>
            """, unsafe_allow_html=True)
        with sc3:
            color = "#dc2626" if ort_harcama * 1.3 > limit else "#d97706"
            st.markdown(f"""
            <div style="background:#fff5f5; border:1px solid {color}33; border-radius:12px; padding:16px; text-align:center;">
                <div style="font-size:1.4rem;">📉</div>
                <div style="font-weight:700; color:{color}; margin:6px 0;">Kötümser</div>
                <div style="font-size:1.1rem; font-weight:800; color:#0f1923;">₺{ort_harcama*1.3:,.0f}</div>
                <div style="font-size:0.72rem; color:{color};">%30 artış riski</div>
            </div>
            """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # SAYFA: GİDER KATEGORİLERİ (Ödeme Tipi + Kategori Bazlı Raporlar)
    # ══════════════════════════════════════════════════════════
    elif selected == "📂 Gider Kategorileri":
        st.markdown('<div class="page-header"><div class="page-title">📂 GİDER KATEGORİLERİ & ÖDEME TİPİ RAPORLARI</div></div>', unsafe_allow_html=True)

        _harc_odeme_keys_gk = ("harcirah", "harcırah", "harcirahtan dus", "harcırahtan düş",
                            "harcırahtan düş (nakit / kişisel kart)", "nakit", "kisisel")
        _sirket_odeme_keys_gk = ("sirket_karti", "sirket karti", "kredi_karti", "kredi kartı",
                               "şirket", "sirket", "şirket kredi kartı", "sirket kredi karti")

        tab_gk_harc, tab_gk_zeynep, tab_gk_sirket = st.tabs([
            "💵 Harcırah Harcamaları", "💳 Zeynep Özyaman Şahsi K.K.", "🏦 Şirket Kredi Kartı"
        ])

        # ── Kategori tanımları — her bir kategorinin hangi tipe ait olduğu
        _GIDER_KATEGORILERI = {
            "ECZANE GİDERLERİ (ŞAHSİ)":      {"tip": "sahsi",  "keywords": ["eczane", "ilaç", "ilac", "pharmacy", "sağlık", "saglik"]},
            "KIRTASİYE GİDERLERİ (ŞİRKET)":  {"tip": "sirket", "keywords": ["kırtasiye", "kirtasiye", "kalem", "kağıt", "kagit", "toner", "ofis"]},
            "HIRDAVAT GİDERLERİ (ŞİRKET)":   {"tip": "sirket", "keywords": ["hırdavat", "hirdavat", "nalbur", "vida", "çivi", "boya", "donanım"]},
            "YAKIT GİDERLERİ (ŞİRKET)":      {"tip": "sirket", "keywords": ["yakıt", "yakit", "akaryakıt", "benzin", "motorin", "shell", "bp", "total", "opet"]},
            "YEMEK GİDERLERİ (ŞİRKET)":      {"tip": "sirket", "keywords": ["yemek", "restoran", "lokanta", "kafe", "cafe", "kebab", "köfte"]},
            "PATENT GİDERLERİ (ŞİRKET)":     {"tip": "sirket", "keywords": ["patent", "marka", "tescil", "lisans"]},
            "ŞENOL BEY SAĞLIK GİDERLERİ (ŞAHSİ)": {"tip": "sahsi", "keywords": ["sağlık", "saglik", "hastane", "doktor", "eczane", "ilaç", "ilac", "muayene"]},
        }

        def _categorize_expense(row):
            """Fişin firma, kategori ve kalemlerinden gider kategorisini tespit et."""
            _text = " ".join([
                str(row.get("Firma", "")).lower(),
                str(row.get("Kategori", "")).lower(),
                str(row.get("Notlar", "")).lower(),
                " ".join([str(k) for k in row.get("Kalemler", [])]).lower() if isinstance(row.get("Kalemler"), list) else str(row.get("Kalemler","")).lower(),
            ])
            for cat_name, cat_info in _GIDER_KATEGORILERI.items():
                for kw in cat_info["keywords"]:
                    if kw in _text:
                        # Şenol Bey Sağlık — sadece Şenol'un fişleri
                        if "ŞENOL" in cat_name.upper():
                            if "şenol" in str(row.get("Kullanıcı","")).lower() or "senol" in str(row.get("Kullanıcı","")).lower():
                                return cat_name
                        else:
                            return cat_name
            return "DİĞER GİDERLER"

        def _render_gider_kategori_tab(source_df, tab_title):
            """Ödeme tipine ait fişleri kategori bazlı raporla."""
            if source_df.empty:
                st.markdown(f"""
                <div class="empty-state"><div class="empty-icon">📂</div>
                <div>{tab_title} kaydı bulunamadı.</div></div>""", unsafe_allow_html=True)
                return

            source_df = source_df.copy()
            source_df['_Gider_Kategori'] = source_df.apply(_categorize_expense, axis=1)
            source_df['_Tarih_DT'] = pd.to_datetime(source_df['Tarih'], errors='coerce')
            source_df['_Ay_Yil'] = source_df['_Tarih_DT'].dt.strftime('%Y-%m')

            # Filtreler
            gkc1, gkc2 = st.columns(2)
            with gkc1:
                _gk_aylar = ["Tüm Zamanlar"] + sorted(source_df['_Ay_Yil'].dropna().unique().tolist(), reverse=True)
                _gk_ay = st.selectbox("Dönem", _gk_aylar, key=f"gk_{tab_title}_ay")
            with gkc2:
                _gk_cats = ["Tümü"] + sorted(source_df['_Gider_Kategori'].unique().tolist())
                _gk_cat = st.selectbox("Gider Kategorisi", _gk_cats, key=f"gk_{tab_title}_cat")

            _gkf = source_df.copy()
            if _gk_ay != "Tüm Zamanlar":
                _gkf = _gkf[_gkf['_Ay_Yil'] == _gk_ay]
            if _gk_cat != "Tümü":
                _gkf = _gkf[_gkf['_Gider_Kategori'] == _gk_cat]

            if _gkf.empty:
                st.info("Seçili filtrelere uygun kayıt yok.")
                return

            # Genel metrikler
            gm1, gm2, gm3 = st.columns(3)
            gm1.metric("Toplam İşlem", len(_gkf))
            gm2.metric("Toplam Tutar", f"₺{_gkf['Tutar'].sum():,.0f}")
            gm3.metric("Onaylı Tutar", f"₺{_gkf[_gkf['Durum']=='Onaylandı']['Tutar'].sum():,.0f}" if 'Durum' in _gkf.columns else "N/A")

            # Kategori bazlı rakamsal rapor kartları
            st.markdown("#### 📊 Kategori Bazlı Rakamsal Rapor")
            _kat_grp = _gkf.groupby('_Gider_Kategori')['Tutar'].agg(['sum', 'count', 'mean']).sort_values('sum', ascending=False)
            _kat_total = _gkf['Tutar'].sum()

            _kat_colors = {
                "ECZANE": "#dc2626", "KIRTASİYE": "#2F3C6E", "HIRDAVAT": "#d97706",
                "YAKIT": "#0891b2", "YEMEK": "#11855B", "PATENT": "#7c3aed",
                "ŞENOL": "#059669", "DİĞER": "#64748b"
            }
            _kat_icons = {
                "ECZANE": "💊", "KIRTASİYE": "📎", "HIRDAVAT": "🔩",
                "YAKIT": "⛽", "YEMEK": "🍽️", "PATENT": "📜",
                "ŞENOL": "🏥", "DİĞER": "📦"
            }

            cols_per_row = 3
            items = list(_kat_grp.iterrows())
            for row_start in range(0, len(items), cols_per_row):
                row_items = items[row_start:row_start + cols_per_row]
                cols = st.columns(cols_per_row)
                for ci, (kat, row) in enumerate(row_items):
                    pct = (row['sum'] / _kat_total * 100) if _kat_total > 0 else 0
                    _c = "#64748b"
                    _ic = "📦"
                    for _k, _v in _kat_colors.items():
                        if _k in kat.upper():
                            _c = _v; break
                    for _k, _v in _kat_icons.items():
                        if _k in kat.upper():
                            _ic = _v; break
                    with cols[ci]:
                        st.markdown(f"""
                        <div style="background:#fff; border:1px solid {_c}33; border-radius:14px;
                                    padding:18px; text-align:center; border-top:3px solid {_c};">
                            <div style="font-size:1.6rem; margin-bottom:6px;">{_ic}</div>
                            <div style="font-size:0.72rem; font-weight:700; color:{_c};
                                        letter-spacing:0.05em; margin-bottom:8px;">{kat}</div>
                            <div style="font-size:1.4rem; font-weight:900; color:#0f1923;">₺{row['sum']:,.0f}</div>
                            <div style="display:flex; justify-content:center; gap:16px; margin-top:8px; font-size:0.72rem; color:#7a96a4;">
                                <span>{int(row['count'])} fiş</span>
                                <span>ort. ₺{row['mean']:,.0f}</span>
                                <span style="font-weight:700; color:{_c};">%{pct:.1f}</span>
                            </div>
                            <div style="margin-top:8px; height:4px; background:rgba(0,0,0,0.06); border-radius:99px;">
                                <div style="height:100%; width:{min(pct,100):.0f}%; background:{_c}; border-radius:99px;"></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

            # Grafik
            if len(_kat_grp) > 1:
                st.markdown("<br>", unsafe_allow_html=True)
                fig_gk = go.Figure(go.Pie(
                    labels=_kat_grp.index.tolist(),
                    values=_kat_grp['sum'].values.tolist(),
                    hole=0.55,
                    textinfo='label+percent',
                    textfont=dict(size=11)
                ))
                fig_gk.update_layout(
                    title=f'{tab_title} — Kategori Dağılımı',
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#4a5568'), height=350,
                    legend=dict(bgcolor='rgba(0,0,0,0)')
                )
                st.plotly_chart(fig_gk, use_container_width=True)

            # Export + tablo
            st.markdown("<br>", unsafe_allow_html=True)
            _gkf_clean = _gkf.drop(columns=['_Tarih_DT', '_Ay_Yil'], errors='ignore')
            _gkf_clean = _gkf_clean.rename(columns={'_Gider_Kategori': 'Gider Kategorisi'})
            gd1, gd2, gd3 = st.columns(3)
            _safe_title = tab_title.replace(" ","_").replace("/","")
            gd1.download_button("📥 CSV", _gkf_clean.to_csv(index=False).encode('utf-8-sig'),
                f"Gider_{_safe_title}_{_gk_ay}.csv", "text/csv", use_container_width=True, key=f"gk_{tab_title}_csv")
            gd2.download_button("📊 Excel",
                export_excel_muhasebe(_gkf_clean.drop(columns=['Gider Kategorisi'], errors='ignore'), f"{tab_title} — {_gk_ay}", logo_path="logo.png"),
                f"Gider_{_safe_title}_{_gk_ay}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key=f"gk_{tab_title}_xlsx")
            gd3.download_button("📄 PDF",
                export_pdf_muhasebe(_gkf_clean.drop(columns=['Gider Kategorisi'], errors='ignore'), f"{tab_title} Raporu", f"{tab_title} — {_gk_ay}", logo_path="logo.png"),
                f"Gider_{_safe_title}_{_gk_ay}.pdf", "application/pdf", use_container_width=True, key=f"gk_{tab_title}_pdf")

            st.dataframe(_gkf_clean.sort_values('Tarih', ascending=False), use_container_width=True, hide_index=True)

        # ── TAB: Harcırah
        with tab_gk_harc:
            st.markdown("### 💵 Harcırah Harcamaları — Kategori Bazlı")
            if not df_full.empty and 'Odeme_Turu' in df_full.columns:
                _harc_gk = df_full[df_full['Odeme_Turu'].str.lower().str.strip().isin(_harc_odeme_keys_gk)].copy()
            else:
                _harc_gk = pd.DataFrame()
            _render_gider_kategori_tab(_harc_gk, "Harcırah")

        # ── TAB: Zeynep Özyaman Şahsi K.K.
        with tab_gk_zeynep:
            st.markdown("### 💳 Zeynep Özyaman Şahsi K.K. — Kategori Bazlı")
            if not df_full.empty and 'Kullanıcı' in df_full.columns and 'Odeme_Turu' in df_full.columns:
                _zk_gk = df_full[
                    df_full['Kullanıcı'].str.contains('Zeynep', case=False, na=False) &
                    df_full['Odeme_Turu'].str.lower().str.strip().isin(_harc_odeme_keys_gk)
                ].copy()
            else:
                _zk_gk = pd.DataFrame()
            _render_gider_kategori_tab(_zk_gk, "Zeynep Özyaman K.K.")

        # ── TAB: Şirket Kredi Kartı
        with tab_gk_sirket:
            st.markdown("### 🏦 Şirket Kredi Kartı — Kategori Bazlı")
            if not df_full.empty and 'Odeme_Turu' in df_full.columns:
                _sk_gk = df_full[df_full['Odeme_Turu'].str.lower().str.strip().isin(_sirket_odeme_keys_gk)].copy()
            else:
                _sk_gk = pd.DataFrame()
            _render_gider_kategori_tab(_sk_gk, "Şirket Kredi Kartı")
